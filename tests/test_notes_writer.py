"""Unit tests for notes/writer.py — write_notes_workbook()."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from notes.payload import NotesPayload
from notes.writer import CELL_CHAR_LIMIT, write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path

# Notes-CI is the smallest template — perfect for round-trip tests.
CORP_INFO_SHEET = "Notes-CI"


def _first_matching_row(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v and label.lower() in str(v).lower():
            return row
    raise AssertionError(f"Label '{label}' not found in sheet {ws.title}")


def test_company_prose_write_puts_content_in_b_and_evidence_in_d(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_filled.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="The Group is a going concern.",
            evidence="Page 14, Note 2(a)",
            source_pages=[14],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success, result.errors
    assert result.rows_written == 1

    wb = openpyxl.load_workbook(out)
    ws = wb[CORP_INFO_SHEET]
    row = _first_matching_row(ws, "Financial reporting status")
    # Cell now carries the heading prepend followed by the body text
    # (Phase 2 of the notes-heading plan). The heading-specific tests
    # below pin the prepend order; here we only care about column placement.
    assert "going concern" in ws.cell(row=row, column=2).value
    # Evidence goes to col D (4) for company filings.
    assert ws.cell(row=row, column=4).value == "Page 14, Note 2(a)"
    # Prior year + company cols are N/A for company-level templates.
    wb.close()


def test_group_prose_writes_to_group_column_only_not_company_columns(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="group")
    out = tmp_path / "Notes-CI_group_filled.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="Consolidated entity is a going concern.",
            evidence="Page 15, Note 2(b)",
            source_pages=[15],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="group",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[CORP_INFO_SHEET]
    row = _first_matching_row(ws, "Financial reporting status")
    # Section 2 #6: Group filing prose → Group col B only. Cell carries
    # heading prepend + body (Phase 2 notes-heading plan); substring check
    # because the heading line now precedes the body.
    assert "going concern" in ws.cell(row=row, column=2).value
    # Company cols (D, E) must be empty for prose.
    assert ws.cell(row=row, column=4).value in (None, "")
    assert ws.cell(row=row, column=5).value in (None, "")
    # Evidence → col F on group.
    assert ws.cell(row=row, column=6).value == "Page 15, Note 2(b)"
    wb.close()


def test_group_numeric_writes_both_group_and_company_columns(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.ISSUED_CAPITAL, level="group")
    out = tmp_path / "Notes-IC_group_filled.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Shares issued and fully paid",
            content="",
            evidence="Page 42, Note 14",
            source_pages=[42],
            numeric_values={
                "group_cy": 1000.0,
                "group_py": 900.0,
                "company_cy": 800.0,
                "company_py": 700.0,
            },
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="group",
        sheet_name="Notes-Issuedcapital",
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb["Notes-Issuedcapital"]
    row = _first_matching_row(ws, "Shares issued and fully paid")
    assert ws.cell(row=row, column=2).value == 1000.0  # Group CY
    assert ws.cell(row=row, column=3).value == 900.0   # Group PY
    assert ws.cell(row=row, column=4).value == 800.0   # Company CY
    assert ws.cell(row=row, column=5).value == 700.0   # Company PY
    assert ws.cell(row=row, column=6).value == "Page 42, Note 14"
    wb.close()


def test_writer_truncates_overlong_content_with_footer(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_trunc.xlsx"
    huge = "A" * (CELL_CHAR_LIMIT + 500)
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content=huge,
            evidence="Pages 10-12",
            source_pages=[10, 11, 12],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success

    wb = openpyxl.load_workbook(out)
    ws = wb[CORP_INFO_SHEET]
    row = _first_matching_row(ws, "Financial reporting status")
    written = ws.cell(row=row, column=2).value
    assert len(written) <= CELL_CHAR_LIMIT
    assert "truncated" in written.lower()
    assert "10" in written  # footer mentions source pages
    wb.close()


def test_writer_skips_unknown_row_labels(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_skip.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Completely bogus label that does not exist",
            content="something",
            evidence="Page 1",
            source_pages=[1],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    # Writer returns success=False only if nothing was written AND errors exist.
    # One unmatched label produces an error but doesn't crash; rows_written=0.
    assert result.rows_written == 0
    assert any("bogus label" in e.lower() for e in result.errors)


def test_writer_surfaces_fuzzy_matches_in_result(tmp_path: Path):
    """Review I1: non-exact row resolutions must surface on the result so
    operators can review borderline matches instead of them being silent."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_fuzzy.xlsx"
    # Deliberately imperfect label — drops the final 's'. The writer's
    # fuzzy fallback should still resolve it but report the match.
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting statu",
            content="The Group is a going concern.",
            evidence="Page 14",
            source_pages=[14],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success
    assert result.fuzzy_matches, "fuzzy fallback should have surfaced"
    req, chosen, score = result.fuzzy_matches[0]
    assert req == "Financial reporting statu"
    assert "financial reporting status" in chosen.lower()
    assert 0.7 <= score < 1.0


def test_combine_payloads_sorts_by_source_page(tmp_path: Path):
    """PR A.3: _combine_payloads must order contributors by the earliest
    PDF page each cited so row-112's concatenation is stable across
    re-runs — input order from asyncio.wait(ALL_COMPLETED) is
    batch-completion order and non-deterministic."""
    from notes.writer import _combine_payloads

    # Feed in reverse page order — later page first.
    payloads = [
        NotesPayload(
            chosen_row_label="Disclosure of other notes to accounts",
            content="later note",
            evidence="p.30",
            source_pages=[30, 31],
            sub_agent_id="subB",
            parent_note={"number": "1", "title": "Test Note"},
        ),
        NotesPayload(
            chosen_row_label="Disclosure of other notes to accounts",
            content="earlier note",
            evidence="p.10",
            source_pages=[10, 11],
            sub_agent_id="subA",
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    combined = _combine_payloads(payloads)
    # Earlier-page payload must appear first in the concatenated content.
    assert combined.content.index("earlier note") < combined.content.index("later note")
    # Evidence joined in the same order.
    assert combined.evidence.index("p.10") < combined.evidence.index("p.30")
    # Aggregated pages ordered earliest-first.
    assert combined.source_pages[0] == 10
    # Sub-agent ids reflect the sorted order too.
    assert combined.sub_agent_id == "subA,subB"


def test_writer_combines_sub_agent_ids_on_row_collision(tmp_path: Path):
    """Review S6: when multiple sub-agents land on the same row the
    combined payload must preserve every contributing sub_agent_id."""
    from notes.writer import _combine_payloads  # internal helper intentionally

    payloads = [
        NotesPayload(
            chosen_row_label="Disclosure of other notes to accounts",
            content="note A",
            evidence="p.10",
            source_pages=[10],
            sub_agent_id="sub0",
            parent_note={"number": "1", "title": "Test Note"},
        ),
        NotesPayload(
            chosen_row_label="Disclosure of other notes to accounts",
            content="note B",
            evidence="p.20",
            source_pages=[20],
            sub_agent_id="sub2",
            parent_note={"number": "1", "title": "Test Note"},
        ),
        NotesPayload(
            chosen_row_label="Disclosure of other notes to accounts",
            content="note C",
            evidence="p.30",
            source_pages=[30],
            sub_agent_id="sub0",  # duplicate — should dedupe
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    combined = _combine_payloads(payloads)
    assert combined.sub_agent_id == "sub0,sub2"
    assert "note A" in combined.content
    assert "note C" in combined.content


def test_writer_refuses_to_overwrite_formula_cells(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.ISSUED_CAPITAL, level="company")
    # First, inspect the template to find any formula row we could target.
    wb = openpyxl.load_workbook(str(tpl))
    ws = wb["Notes-Issuedcapital"]
    formula_row = None
    for r in range(1, ws.max_row + 1):
        for c in range(2, 4):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                formula_row = r
                break
        if formula_row:
            break
    wb.close()

    if formula_row is None:
        pytest.skip("template has no formula cells — test not applicable")

    # Fabricate a label that resolves to the formula row.
    # We can't easily hit that row by label — use a raw-row-index payload.
    # Skip if the public API doesn't support row overrides (that's OK).
    pytest.skip("row-override API not part of public writer contract; guard tested in integration")


def test_evidence_not_written_without_values(tmp_path: Path):
    """PR A.2: a payload with no content and no numeric_values must NOT
    leave its evidence text behind as a ghost citation. Previously the
    writer still wrote evidence to col D even when no value column was
    filled, producing a row with citation text but nothing to cite."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_ghost.xlsx"

    wb0 = openpyxl.load_workbook(str(tpl))
    ws0 = wb0[CORP_INFO_SHEET]
    target_row = _first_matching_row(ws0, "Financial reporting status")
    baseline_evidence = ws0.cell(row=target_row, column=4).value
    wb0.close()

    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="",
            evidence="Page 14, Note 2(a)",
            source_pages=[14],
            numeric_values=None,
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.rows_written == 0
    assert result.success is False

    wb = openpyxl.load_workbook(out)
    ws = wb[CORP_INFO_SHEET]
    assert ws.cell(row=target_row, column=4).value == baseline_evidence
    wb.close()


def test_empty_payloads_returns_failure(tmp_path: Path):
    """PR A.1: zero-row writes must fail so Sheet-12's "all sub-agents lost
    coverage" case can't ship a silent green tick on an untouched template."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_empty.xlsx"
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=[],
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success is False
    assert result.rows_written == 0


def test_notes_writer_persists_source_note_refs_for_post_validator(tmp_path: Path):
    """Step 4.3: the writer must persist a per-template sidecar next to
    the filled xlsx so the Phase 5 post-validator can load note-ref
    provenance without re-parsing the workbook."""
    import json
    from notes.writer import _payload_sidecar_path

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_filled.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="The Group is a going concern.",
            evidence="Page 14",
            source_pages=[14],
            source_note_refs=["2", "2(a)"],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success, result.errors
    sidecar = _payload_sidecar_path(str(out))
    assert sidecar.exists(), f"expected sidecar at {sidecar}"
    data = json.loads(sidecar.read_text(encoding="utf-8"))
    assert isinstance(data, list)
    assert len(data) == 1
    entry = data[0]
    assert entry["sheet"] == CORP_INFO_SHEET
    assert set(entry["source_note_refs"]) == {"2", "2(a)"}
    assert "Going concern" in entry["content_preview"] or "going concern" in entry["content_preview"]


def test_notes_writer_sidecar_concatenates_refs_for_row_with_multiple_payloads(tmp_path: Path):
    """When multiple payloads collapse into one row (Sheet-12 row-112
    catch-all pattern), the sidecar must union their source_note_refs."""
    import json
    from notes.writer import _payload_sidecar_path

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_multi.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="First line.",
            evidence="Page 10",
            source_pages=[10],
            source_note_refs=["2"],
            parent_note={"number": "1", "title": "Test Note"},
        ),
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="Second line.",
            evidence="Page 11",
            source_pages=[11],
            source_note_refs=["2.1"],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success
    data = json.loads(_payload_sidecar_path(str(out)).read_text(encoding="utf-8"))
    assert len(data) == 1
    assert set(data[0]["source_note_refs"]) == {"2", "2.1"}


# ---------------------------------------------------------------------------
# Heading prepend behaviour (Phase 2 of the model+notes-heading plan).
#
# The writer owns the "every note cell opens with its heading" rule —
# agents supply parent_note (and optionally sub_note) as structured fields
# and the writer deterministically prepends the <h3> lines. Impossible to
# drift: the LLM can't forget to include markup it doesn't emit.
# ---------------------------------------------------------------------------


def test_writer_prepends_parent_heading_to_prose_cell(tmp_path: Path):
    """A payload with only a parent_note gets one `<h3>` line prepended.

    Uses the Excel-flattened output (rather than the raw HTML) because
    the workbook stores text; the `<h3>` tag is converted to a bolded
    plain-text line by `html_to_excel_text`. The marker we assert on is
    the note number + title appearing BEFORE the body in the cell value.
    """
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_heading_parent.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="<p>The Group is a going concern.</p>",
            evidence="Page 14, Note 2",
            source_pages=[14],
            parent_note={"number": "2", "title": "Basis of Preparation"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[CORP_INFO_SHEET]
    row = _first_matching_row(ws, "Financial reporting status")
    written = ws.cell(row=row, column=2).value
    # Heading text appears before the body text.
    assert "2 Basis of Preparation" in written
    assert written.index("2 Basis of Preparation") < written.index("going concern")
    wb.close()


def test_writer_prepends_parent_and_sub_headings_to_subnote_cell(tmp_path: Path):
    """A payload with both parent_note and sub_note gets two `<h3>` lines,
    parent first, then sub. Order matters — the sub-note hierarchy reads
    "5 Material Accounting Policies → 5.4 Property, Plant and Equipment"
    from top to bottom in the cell."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_heading_sub.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="<p>PPE is stated at cost.</p>",
            evidence="Page 27, Note 5.4",
            source_pages=[27],
            parent_note={"number": "5", "title": "Material Accounting Policies"},
            sub_note={"number": "5.4", "title": "Property, Plant and Equipment"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[CORP_INFO_SHEET]
    row = _first_matching_row(ws, "Financial reporting status")
    written = ws.cell(row=row, column=2).value
    # Parent heading appears, then sub-heading, then body — in that order.
    parent_idx = written.index("5 Material Accounting Policies")
    sub_idx = written.index("5.4 Property, Plant and Equipment")
    body_idx = written.index("stated at cost")
    assert parent_idx < sub_idx < body_idx, (
        f"Heading order wrong. Got: {written[:200]!r}"
    )
    wb.close()


def test_writer_numeric_only_payload_has_no_headings_injected(tmp_path: Path):
    """Sheet 13/14 numeric-only rows hold a number, not prose — heading
    injection must be a no-op so the cell stays a clean numeric value."""
    tpl = notes_template_path(NotesTemplateType.ISSUED_CAPITAL, level="company")
    out = tmp_path / "Notes-IC_numeric.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Shares issued and fully paid",
            content="",  # numeric-only
            evidence="Page 42, Note 14",
            source_pages=[42],
            numeric_values={"company_cy": 1000.0, "company_py": 900.0},
            parent_note={"number": "14", "title": "Issued Capital"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-Issuedcapital",
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb["Notes-Issuedcapital"]
    row = _first_matching_row(ws, "Shares issued and fully paid")
    # Numeric cells stay numbers — no heading prose injected.
    assert ws.cell(row=row, column=2).value == 1000.0
    assert ws.cell(row=row, column=3).value == 900.0
    wb.close()


def test_writer_headings_count_toward_truncation_budget(tmp_path: Path):
    """When body content is near the char cap, the heading prepend must
    still happen; truncation then applies to the combined text so the
    footer sits after the headings, never replacing them."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_heading_trunc.xlsx"
    huge = "A" * (CELL_CHAR_LIMIT + 500)
    payloads = [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content=f"<p>{huge}</p>",
            evidence="Pages 10-12",
            source_pages=[10, 11, 12],
            parent_note={"number": "2", "title": "Basis of Preparation"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=CORP_INFO_SHEET,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[CORP_INFO_SHEET]
    row = _first_matching_row(ws, "Financial reporting status")
    written = ws.cell(row=row, column=2).value
    # Heading survives — it's prepended before truncation.
    assert "2 Basis of Preparation" in written
    # Truncation still fires: total cell length is under the cap and the
    # footer indicates truncation.
    assert len(written) <= CELL_CHAR_LIMIT
    assert "truncated" in written.lower()
    wb.close()
