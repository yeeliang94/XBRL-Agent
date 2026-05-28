"""Tests for monolith/tools.py — get_state, write_cells, done."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest

from monolith.tools import (
    MonolithToolContext,
    done,
    get_state,
    write_cells,
)
from statement_types import StatementType


_TPL_ROOT = Path(__file__).resolve().parent.parent / "XBRL-template-MFRS" / "Company"


def _merge_company_templates(tmp_path: Path) -> Path:
    """Same merger as in test_monolith_state.py — five Company face
    templates concatenated into one workbook."""
    wb = openpyxl.load_workbook(
        str(_TPL_ROOT / "01-SOFP-CuNonCu.xlsx"), data_only=False,
    )
    for source_name in (
        "03-SOPL-Function.xlsx",
        "05-SOCI-BeforeTax.xlsx",
        "07-SOCF-Indirect.xlsx",
        "09-SOCIE.xlsx",
    ):
        src = openpyxl.load_workbook(str(_TPL_ROOT / source_name), data_only=False)
        for sheet_name in src.sheetnames:
            if sheet_name in wb.sheetnames:
                continue
            target = wb.create_sheet(title=sheet_name)
            src_ws = src[sheet_name]
            for row in src_ws.iter_rows():
                for cell in row:
                    target.cell(row=cell.row, column=cell.column, value=cell.value)
    out = tmp_path / "monolith_filled.xlsx"
    wb.save(str(out))
    wb.close()
    return out


def _ctx(tmp_path: Path, **kw) -> MonolithToolContext:
    wb_path = _merge_company_templates(tmp_path)
    return MonolithToolContext(
        workbook_path=str(wb_path),
        pdf_page_count=120,
        filing_standard="mfrs",
        filing_level="company",
        variants={
            StatementType.SOFP: "CuNonCu",
            StatementType.SOPL: "Function",
            StatementType.SOCI: "BeforeTax",
            StatementType.SOCF: "Indirect",
            StatementType.SOCIE: "Default",
        },
        **kw,
    )


def _seed_one_write(ctx: MonolithToolContext) -> None:
    """Land one accepted write so `done()` clears the empty-workbook guard.

    The accept_imbalance validation tests must hit the per-entry
    validation branch, not the empty-workbook short-circuit added for
    the 2026-05-28 scanned-PDF incident — seed a minimal write first.
    """
    snap = get_state(ctx)
    leaf = next(
        r for r in snap["sheets"]["SOFP"]["rows"] if r["kind"] == "leaf"
    )
    write_cells(ctx, [
        {"sheet": "SOFP-CuNonCu", "row": leaf["row"], "col": "cy",
         "value": 1, "evidence": "seed for done() validation tests"},
    ])


# -----------------------------------------------------------------------------
# get_state
# -----------------------------------------------------------------------------


def test_get_state_returns_snapshot_dict(tmp_path):
    ctx = _ctx(tmp_path)
    state = get_state(ctx)
    assert state["filing"]["standard"] == "mfrs"
    assert state["filing"]["level"] == "company"
    assert state["turn"] == 1
    assert {"SOFP", "SOPL", "SOCI", "SOCF", "SOCIE"} == set(state["sheets"])
    assert isinstance(state["cross_checks"], list)
    # Second call increments the turn counter.
    state = get_state(ctx)
    assert state["turn"] == 2


# -----------------------------------------------------------------------------
# write_cells
# -----------------------------------------------------------------------------


def test_write_cells_accepts_cross_sheet_batch(tmp_path):
    ctx = _ctx(tmp_path)
    # Look up a leaf row on SOFP and a leaf on SOPL from the snapshot so
    # we don't hardcode magic numbers.
    snap = get_state(ctx)
    sofp_leaf = next(
        r for r in snap["sheets"]["SOFP"]["rows"]
        if r["kind"] == "leaf"
    )
    sopl_leaf = next(
        r for r in snap["sheets"]["SOPL"]["rows"]
        if r["kind"] == "leaf"
    )
    out = write_cells(ctx, [
        {"sheet": "SOFP-CuNonCu", "row": sofp_leaf["row"], "col": "cy",
         "value": 12345, "evidence": "Note 14 (p.42)"},
        {"sheet": "SOPL-Function", "row": sopl_leaf["row"], "col": "cy",
         "value": 6789, "evidence": "Note 5 (p.31)"},
    ])
    assert len(out["written"]) >= 2
    assert not [r for r in out["rejected"] if not r["reason"].startswith("warning:")]


def test_write_cells_rejects_abstract_row(tmp_path):
    ctx = _ctx(tmp_path)
    snap = get_state(ctx)
    abstract_row = next(
        r for r in snap["sheets"]["SOFP"]["rows"]
        if r["kind"] == "abstract"
    )
    out = write_cells(ctx, [
        {"sheet": "SOFP-CuNonCu", "row": abstract_row["row"], "col": "cy", "value": 1},
    ])
    # The writer's gotcha #17 guard refuses; the rejection message names
    # the row and points at the leaf alternative.
    assert any(
        "abstract" in (r["reason"] or "").lower()
        for r in out["rejected"]
    )


def test_write_cells_rejects_matrix_col_on_non_matrix_sheet(tmp_path):
    ctx = _ctx(tmp_path)
    out = write_cells(ctx, [
        {"sheet": "SOFP-CuNonCu", "row": 5, "matrix_col": "RetainedEarnings",
         "value": 1},
    ])
    assert any(
        "matrix_col is not valid" in r["reason"]
        for r in out["rejected"]
    )


def test_write_cells_rejects_col_on_socie(tmp_path):
    ctx = _ctx(tmp_path)
    out = write_cells(ctx, [
        {"sheet": "SOCIE", "row": 8, "col": "cy", "value": 1},
    ])
    assert any(
        "use `matrix_col`" in r["reason"]
        for r in out["rejected"]
    )


def test_write_cells_rejects_unknown_matrix_col(tmp_path):
    ctx = _ctx(tmp_path)
    out = write_cells(ctx, [
        {"sheet": "SOCIE", "row": 8, "matrix_col": "DefinitelyNotAColumn",
         "value": 1},
    ])
    assert any(
        "not present on SOCIE" in r["reason"]
        for r in out["rejected"]
    )


def test_write_cells_rejects_duplicate_in_batch(tmp_path):
    ctx = _ctx(tmp_path)
    snap = get_state(ctx)
    leaf = next(
        r for r in snap["sheets"]["SOFP"]["rows"] if r["kind"] == "leaf"
    )
    out = write_cells(ctx, [
        {"sheet": "SOFP-CuNonCu", "row": leaf["row"], "col": "cy", "value": 1},
        {"sheet": "SOFP-CuNonCu", "row": leaf["row"], "col": "cy", "value": 2},
    ])
    assert any(
        "duplicate" in r["reason"]
        for r in out["rejected"]
    )


def test_write_cells_rejects_non_numeric_value(tmp_path):
    ctx = _ctx(tmp_path)
    out = write_cells(ctx, [
        {"sheet": "SOFP-CuNonCu", "row": 5, "col": "cy", "value": "not a number"},
    ])
    assert any(
        "must be a number" in r["reason"]
        for r in out["rejected"]
    )


# -----------------------------------------------------------------------------
# done
# -----------------------------------------------------------------------------


def test_done_on_blank_workbook_returns_not_done(tmp_path):
    """A blank template fails every balance check — `done({})` is not
    accepted; it returns `not_done` with the failing checks named."""
    ctx = _ctx(tmp_path)
    out = done(ctx)
    assert out["status"] == "not_done"
    assert out["failing_checks"], "expected at least one failing check"


def test_done_clears_guard_when_workbook_has_pre_existing_numerics(tmp_path):
    """Peer-review HIGH #1: a resumed run starts with an empty
    `_writes_by_cell` tracker but an already-written workbook. The
    empty-workbook guard must not refuse `done` in that case — read
    the persisted file, not just memory.
    """
    ctx = _ctx(tmp_path)
    # Plant a numeric directly on disk to simulate a previous run's
    # accepted writes. Don't go through write_cells — that would
    # populate `_writes_by_cell` and bypass the half of the guard we
    # want to exercise.
    wb = openpyxl.load_workbook(ctx.workbook_path, data_only=False)
    try:
        wb["SOFP-CuNonCu"].cell(row=5, column=2, value=12345)
        wb.save(ctx.workbook_path)
    finally:
        wb.close()
    assert not ctx._writes_by_cell, (
        "test setup invariant: tracker must stay empty so we exercise "
        "the persisted-data half of the guard"
    )
    snap = get_state(ctx)
    failing_ids = [c["id"] for c in snap["cross_checks"] if not c["pass"]]
    failing_ids.extend([c["id"] for c in snap["verifier"] if not c["pass"]])
    out = done(ctx, accept_imbalance=[
        {
            "check_id": cid,
            "reason": "resumed-run continuation",
            "pdf_page": 1,
            "evidence_excerpt": "stub",
        }
        for cid in failing_ids
    ])
    assert out["status"] == "done", (
        f"expected done to honour pre-existing writes; got: {out}"
    )


def test_done_refuses_when_workbook_has_no_writes(tmp_path):
    """Empty-workbook guard (2026-05-28 scanned-PDF incident).

    Even when the agent supplies an `accept_imbalance` entry for every
    currently-failing check, `done` must refuse if no `write_cells` has
    ever landed — otherwise a vacuous all-zero workbook gets
    rubber-stamped because most cross-checks pass trivially (0 == 0).
    """
    ctx = _ctx(tmp_path)
    snap = get_state(ctx)
    failing_ids = [c["id"] for c in snap["cross_checks"] if not c["pass"]]
    failing_ids.extend([c["id"] for c in snap["verifier"] if not c["pass"]])
    out = done(ctx, accept_imbalance=[
        {
            "check_id": cid,
            "reason": "would-be rubber-stamp",
            "pdf_page": 1,
            "evidence_excerpt": "stub excerpt",
        }
        for cid in failing_ids
    ])
    assert out["status"] == "not_done"
    assert "no cell writes" in out["message"].lower()
    # accept_imbalance entries are NOT promoted to accepted_residuals
    # when the guard fires — the agent must write first, then accept.
    assert out["accepted_residuals"] == []


def test_done_rejects_accept_with_passing_check_id(tmp_path):
    ctx = _ctx(tmp_path)
    _seed_one_write(ctx)
    out = done(ctx, accept_imbalance=[{
        "check_id": "definitely_not_a_real_check",
        "reason": "made up",
        "pdf_page": 1,
        "evidence_excerpt": "x",
    }])
    assert out["status"] == "not_done"
    # The invalid entry is surfaced with a reason.
    assert "invalid_accepts" in out
    assert any(
        "not currently failing" in inv["reason"]
        for inv in out["invalid_accepts"]
    )


def test_done_rejects_accept_with_out_of_range_page(tmp_path):
    ctx = _ctx(tmp_path)
    _seed_one_write(ctx)
    snap = get_state(ctx)
    failing_id = next(
        c["id"] for c in snap["cross_checks"] if not c["pass"]
    )
    out = done(ctx, accept_imbalance=[{
        "check_id": failing_id,
        "reason": "stub",
        "pdf_page": 999999,
        "evidence_excerpt": "x",
    }])
    assert out["status"] == "not_done"
    assert any(
        "out of range" in inv["reason"]
        for inv in out["invalid_accepts"]
    )


def test_done_rejects_accept_with_empty_excerpt(tmp_path):
    ctx = _ctx(tmp_path)
    _seed_one_write(ctx)
    snap = get_state(ctx)
    failing_id = next(
        c["id"] for c in snap["cross_checks"] if not c["pass"]
    )
    out = done(ctx, accept_imbalance=[{
        "check_id": failing_id,
        "reason": "stub",
        "pdf_page": 1,
        "evidence_excerpt": "  ",
    }])
    assert out["status"] == "not_done"
    assert any(
        "evidence_excerpt is required" in inv["reason"]
        for inv in out["invalid_accepts"]
    )


def test_done_accepts_when_every_failing_check_is_named(tmp_path):
    ctx = _ctx(tmp_path)
    _seed_one_write(ctx)
    snap = get_state(ctx)
    failing_ids = [c["id"] for c in snap["cross_checks"] if not c["pass"]]
    failing_ids.extend([c["id"] for c in snap["verifier"] if not c["pass"]])
    out = done(ctx, accept_imbalance=[
        {
            "check_id": cid,
            "reason": "stub for unit test",
            "pdf_page": 1,
            "evidence_excerpt": "stub excerpt",
        }
        for cid in failing_ids
    ])
    assert out["status"] == "done"
    assert out["failing_checks"] == []
    assert len(out["accepted_residuals"]) == len(failing_ids)
