"""SOCF duplicate "Cash and cash equivalents at end of period" label (2026-05-29).

`SOCF-Indirect` carries that mandatory label on TWO rows:
  * the statement line (a blank leaf the agent fills), and
  * lower down, under "Details of cash flows", a reconciliation total that is a
    formula reading the optional cash-on-hand / overdraft breakdown.

`_verify_socf` used "last assignment wins" when locating `cash_end`, so it
bound to the formula row — which evaluates to 0 until the optional breakdown is
filled. The balance identity (end == beginning + net change after FX) could then
never be satisfied with the values the agent is actually meant to enter, and the
agent looped trying to overwrite a protected formula cell until it failed.

The fix binds `cash_end` to the FIRST match (the statement line), matching what
`fill_workbook._find_row_by_label` and `_collect_unfilled_mandatory` resolve to.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from statement_types import StatementType
from tools.verifier import verify_statement

REPO = Path(__file__).resolve().parent.parent


def _build_socf(path: Path) -> None:
    """Minimal SOCF-Indirect reproducing the duplicate-label structure:
    a balanced statement line followed by an unfilled 'Details' total."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCF-Indirect"
    rows = [
        ("*Net cash flows from (used in) operating activities", 100),
        ("*Net increase (decrease) in cash and cash equivalents "
         "before effect of exchange rate changes", 100),
        ("*Net increase (decrease) in cash and cash equivalents "
         "after effect of exchange rate changes", 100),
        ("*Cash and cash equivalents at beginning of period", 50),
        # Statement line — reconciles: 50 + 100 == 150.
        ("*Cash and cash equivalents at end of period", 150),
        ("Details of cash flows", None),
        # Duplicate label as a 'Details' reconciliation total, left blank
        # (0 / unfilled) — binding cash_end here would falsely fail balance.
        ("*Cash and cash equivalents at end of period", None),
    ]
    for i, (label, val) in enumerate(rows, start=1):
        ws.cell(row=i, column=1).value = label
        if val is not None:
            ws.cell(row=i, column=2).value = val
    wb.save(path)
    wb.close()


def test_verify_socf_binds_cash_end_to_statement_line(tmp_path):
    p = tmp_path / "socf.xlsx"
    _build_socf(p)
    result = verify_statement(str(p), StatementType.SOCF, filing_level="company")
    # First-match binding → the statement line (150) reconciles against
    # beginning (50) + net-after-fx (100). The pre-fix last-match binding to
    # the blank 'Details' total would have reported is_balanced False.
    assert result.is_balanced is True, result.mismatches
    assert "cash_end" in result.computed_totals
    assert result.computed_totals["cash_end"] == 150


def test_live_socf_template_has_the_duplicate_label_trap():
    """Structural guard: if the template ever stops carrying two
    'Cash ... at end of period' rows (blank leaf first, formula second), this
    test should be revisited — the first-match fix targets exactly that shape.
    """
    for level in ("Company", "Group"):
        path = REPO / "XBRL-template-MFRS" / level / "07-SOCF-Indirect.xlsx"
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb["SOCF-Indirect"]
        end_rows = [
            r for r in range(1, ws.max_row + 1)
            if (v := ws.cell(r, 1).value)
            and "cash and cash equivalents at end" in str(v).strip().lstrip("*").lower()
        ]
        wb.close()
        assert len(end_rows) == 2, f"{level}: expected 2 end-of-period rows"
        # First occurrence is the blank statement leaf; second is the formula.
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb["SOCF-Indirect"]
        first_b = ws.cell(end_rows[0], 2).value
        second_b = ws.cell(end_rows[1], 2).value
        wb.close()
        assert first_b is None, f"{level}: first end row should be a blank leaf"
        assert isinstance(second_b, str) and second_b.startswith("="), (
            f"{level}: second end row should be a formula total"
        )
