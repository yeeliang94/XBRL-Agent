"""Tests for the generic extraction agent factory (Step 4.2)."""

from unittest.mock import patch
import pytest
from pathlib import Path

from pydantic_ai.models.test import TestModel

from statement_types import StatementType, variants_for
from tools.verifier import VerificationResult


class TestCreateExtractionAgent:
    """Verify the generic factory creates valid agents for each statement type."""

    def _make_agent(self, stmt_type, variant_name, **kwargs):
        from extraction.agent import create_extraction_agent
        return create_extraction_agent(
            statement_type=stmt_type,
            variant=variant_name,
            pdf_path="/tmp/test.pdf",
            template_path="/tmp/test.xlsx",
            model=TestModel(),
            output_dir="/tmp/output",
            **kwargs,
        )

    def test_creates_agent_for_each_statement(self):
        """Factory should produce an agent+deps tuple for every statement type."""
        for stmt in StatementType:
            variants = variants_for(stmt)
            agent, deps = self._make_agent(stmt, variants[0].name)
            assert agent is not None
            assert deps is not None

    def test_deps_carry_statement_metadata(self):
        """AgentDeps should know which statement type and variant it's serving."""
        from extraction.agent import ExtractionDeps
        agent, deps = self._make_agent(StatementType.SOPL, "Function")
        assert isinstance(deps, ExtractionDeps)
        assert deps.statement_type == StatementType.SOPL
        assert deps.variant == "Function"

    def test_deps_carry_page_hints(self):
        """When page_hints are provided, deps should store them."""
        from extraction.agent import ExtractionDeps
        hints = {"face_page": 14, "note_pages": [30, 31]}
        agent, deps = self._make_agent(
            StatementType.SOFP, "CuNonCu", page_hints=hints
        )
        assert deps.page_hints == hints

    def test_deps_default_no_page_hints(self):
        """Without page_hints, deps.page_hints should be None."""
        from extraction.agent import ExtractionDeps
        agent, deps = self._make_agent(StatementType.SOFP, "CuNonCu")
        assert deps.page_hints is None

    def test_output_path_uses_statement_type(self):
        """Each agent's output path should include the statement type for isolation."""
        from extraction.agent import ExtractionDeps
        agent, deps = self._make_agent(StatementType.SOPL, "Function")
        assert deps.statement_type.value in deps.output_dir or "SOPL" in deps.output_dir or deps.output_dir == "/tmp/output"


# ---------------------------------------------------------------------------
# Phase 1.2: mandatory_unfilled surfaces through verify_totals feedback
# ---------------------------------------------------------------------------

def test_verify_totals_tool_output_includes_mandatory_unfilled():
    """The rendered tool output must echo every unfilled label and an
    'Action required:' directive routing the agent back to fix it."""
    from extraction.agent import _format_verify_result

    result = VerificationResult(
        is_balanced=True,
        matches_pdf=None,
        mandatory_unfilled=["*Revenue", "*Finance income"],
    )
    rendered = _format_verify_result(result)
    assert "*Revenue" in rendered
    assert "*Finance income" in rendered
    assert "Action required:" in rendered


# ---------------------------------------------------------------------------
# Phase 1.3: save_result gating
# ---------------------------------------------------------------------------

def test_save_result_blocks_when_verify_totals_never_called():
    """Without a recent verify_totals, save_result returns an error string."""
    from extraction.agent import _check_save_gate, ExtractionDeps
    deps = ExtractionDeps(
        pdf_path="/tmp/x.pdf",
        template_path="/tmp/t.xlsx",
        model="test",
        output_dir="/tmp/o",
        token_report=None,
        statement_type=StatementType.SOPL,
        variant="Function",
    )
    deps.save_attempts = 1
    msg = _check_save_gate(deps)
    assert msg is not None
    assert "verify_totals has not been called" in msg


def test_save_result_blocks_when_verify_totals_failed():
    """If the last verify flagged an imbalance, save_result is blocked."""
    from extraction.agent import _check_save_gate, ExtractionDeps
    deps = ExtractionDeps(
        pdf_path="/tmp/x.pdf",
        template_path="/tmp/t.xlsx",
        model="test",
        output_dir="/tmp/o",
        token_report=None,
        statement_type=StatementType.SOPL,
        variant="Function",
    )
    deps.save_attempts = 1
    deps.last_verify_result = VerificationResult(
        is_balanced=False,
        matches_pdf=None,
        feedback="IMBALANCE: profit != attribution total",
    )
    msg = _check_save_gate(deps)
    assert msg is not None
    assert "refused" in msg.lower()
    assert "IMBALANCE" in msg


def test_save_result_blocks_when_mandatory_unfilled():
    """Unfilled `*` rows alone must block save even if balance is clean."""
    from extraction.agent import _check_save_gate, ExtractionDeps
    deps = ExtractionDeps(
        pdf_path="/tmp/x.pdf",
        template_path="/tmp/t.xlsx",
        model="test",
        output_dir="/tmp/o",
        token_report=None,
        statement_type=StatementType.SOFP,
        variant="CuNonCu",
    )
    deps.save_attempts = 1
    deps.last_verify_result = VerificationResult(
        is_balanced=True,
        matches_pdf=None,
        mandatory_unfilled=["*Revenue"],
    )
    msg = _check_save_gate(deps)
    assert msg is not None
    assert "*Revenue" in msg


def test_save_result_passes_when_verify_clean():
    """A clean verify with no unfilled rows opens the gate (returns None)."""
    from extraction.agent import _check_save_gate, ExtractionDeps
    deps = ExtractionDeps(
        pdf_path="/tmp/x.pdf",
        template_path="/tmp/t.xlsx",
        model="test",
        output_dir="/tmp/o",
        token_report=None,
        statement_type=StatementType.SOFP,
        variant="CuNonCu",
    )
    deps.save_attempts = 1
    deps.last_verify_result = VerificationResult(
        is_balanced=True,
        matches_pdf=None,
        mandatory_unfilled=[],
    )
    assert _check_save_gate(deps) is None


def test_save_result_forced_near_iteration_cap():
    """Peer-review I1: the escape hatch fires when the agent is within
    `_FORCE_SAVE_ITER_MARGIN` iterations of `MAX_AGENT_ITERATIONS`, not
    after a tool-call counter reaches 3. This ensures un-balanceable PDFs
    exhaust the real budget before being force-saved."""
    from agent_tracing import MAX_AGENT_ITERATIONS
    from extraction.agent import (
        _check_save_gate, ExtractionDeps, _FORCE_SAVE_ITER_MARGIN,
    )
    deps = ExtractionDeps(
        pdf_path="/tmp/x.pdf",
        template_path="/tmp/t.xlsx",
        model="test",
        output_dir="/tmp/o",
        token_report=None,
        statement_type=StatementType.SOFP,
        variant="CuNonCu",
    )
    deps.last_verify_result = VerificationResult(
        is_balanced=False,
        matches_pdf=None,
        feedback="still unbalanced",
    )
    # Early iterations: block regardless of save_attempts.
    deps.turn_counter = 1
    deps.save_attempts = 5
    assert _check_save_gate(deps) is not None
    deps.turn_counter = MAX_AGENT_ITERATIONS - _FORCE_SAVE_ITER_MARGIN - 1
    assert _check_save_gate(deps) is not None
    # At the margin: gate opens.
    deps.turn_counter = MAX_AGENT_ITERATIONS - _FORCE_SAVE_ITER_MARGIN
    assert _check_save_gate(deps) is None
    # Past the margin: still open.
    deps.turn_counter = MAX_AGENT_ITERATIONS
    assert _check_save_gate(deps) is None


def test_save_result_attempts_fallback_only_opens_far_past_normal():
    """When turn_counter is zero (test harness without a real coordinator,
    or first save before any iter node fires) the gate falls back to
    save_attempts. That fallback floor is ≥50 so no realistic run can
    accidentally trip it."""
    from extraction.agent import _check_save_gate, ExtractionDeps
    deps = ExtractionDeps(
        pdf_path="/tmp/x.pdf",
        template_path="/tmp/t.xlsx",
        model="test",
        output_dir="/tmp/o",
        token_report=None,
        statement_type=StatementType.SOFP,
        variant="CuNonCu",
    )
    deps.last_verify_result = VerificationResult(
        is_balanced=False,
        matches_pdf=None,
        feedback="still unbalanced",
    )
    deps.turn_counter = 0
    for n in [1, 3, 10, 25, 49]:
        deps.save_attempts = n
        assert _check_save_gate(deps) is not None, f"save_attempts={n} should block"
    deps.save_attempts = 50
    assert _check_save_gate(deps) is None


def test_fill_workbook_resets_last_verify_result(tmp_path):
    """Peer-review S8: exercise the actual fill_workbook tool body so a
    refactor that drops the `deps.last_verify_result = None` line fails
    CI. The previous incarnation set the field manually and asserted —
    which a vacuous refactor would have passed."""
    import openpyxl
    from extraction.agent import create_extraction_agent
    from pydantic_ai.models.test import TestModel as _TM

    # Minimal template that fill_workbook can write into.
    template_path = tmp_path / "tmpl.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Total assets"
    ws["B1"] = 0
    wb.save(str(template_path))

    agent, deps = create_extraction_agent(
        statement_type=StatementType.SOFP,
        variant="CuNonCu",
        pdf_path="/tmp/test.pdf",
        template_path=str(template_path),
        model=_TM(),
        output_dir=str(tmp_path),
    )

    # Pre-load a fake successful verify so we can assert the tool clears it.
    deps.last_verify_result = VerificationResult(
        is_balanced=True, matches_pdf=None, mandatory_unfilled=[],
    )

    # Find the real tool body via the registered toolset and call it.
    # Using the private-but-stable pydantic-ai structure; this is the
    # only way to drive the tool without a full agent run, and matches
    # the pattern other tests (test_correction_agent) already use.
    fill_tool = None
    for ts in getattr(agent, "toolsets", []) or []:
        tools = getattr(ts, "tools", {}) or {}
        if "fill_workbook" in tools:
            fill_tool = tools["fill_workbook"]
            break
    assert fill_tool is not None, "fill_workbook tool not registered"

    # Build a minimal RunContext. The tool only uses ctx.deps, so we
    # can stub the rest.
    import json as _json
    class _Ctx:
        pass
    ctx = _Ctx()
    ctx.deps = deps
    fields_json = _json.dumps({"fields": [
        {"sheet": "SOFP-CuNonCu", "field_label": "Total assets",
         "col": 2, "value": 100, "evidence": "t"},
    ]})
    # Invoke the tool body. `tool.function` is the underlying callable.
    fn = getattr(fill_tool, "function", None) or getattr(fill_tool, "func", None)
    if fn is None:
        # Fallback: try tool.__call__ or tool.run
        fn = fill_tool
    fn(ctx, fields_json)

    assert deps.last_verify_result is None, (
        "fill_workbook must clear last_verify_result on success so save_result "
        "can't be called on stale verification data"
    )


def test_verify_to_save_loop_end_to_end():
    """Step 1.4: simulate the fill -> verify-fail -> fix -> verify-ok -> save
    loop against the real gate helper. Proves the state machine lets the
    agent iterate toward a clean save instead of finalising on a failure."""
    from extraction.agent import _check_save_gate, ExtractionDeps
    deps = ExtractionDeps(
        pdf_path="/tmp/x.pdf",
        template_path="/tmp/t.xlsx",
        model="test",
        output_dir="/tmp/o",
        token_report=None,
        statement_type=StatementType.SOPL,
        variant="Function",
    )
    # Turn 1 — agent fills workbook, verify flags an imbalance.
    deps.last_verify_result = VerificationResult(
        is_balanced=False, matches_pdf=None, feedback="profit != attribution",
    )
    deps.save_attempts = 1
    first_block = _check_save_gate(deps)
    assert first_block is not None and "refused" in first_block.lower()

    # Turn 2 — agent re-fills; fill_workbook clears last_verify_result.
    deps.last_verify_result = None
    deps.save_attempts = 2
    second_block = _check_save_gate(deps)
    assert second_block is not None and "has not been called" in second_block

    # Turn 3 — agent re-verifies; clean.
    deps.last_verify_result = VerificationResult(
        is_balanced=True, matches_pdf=None, mandatory_unfilled=[],
    )
    # save_attempts counter is independent of verify; gate now opens.
    assert _check_save_gate(deps) is None
