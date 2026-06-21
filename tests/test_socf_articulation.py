"""Regression pins for the SOCF articulation hardening (run-50 / Amway).

Two coupled validation gaps let a non-articulating SOCF pass as "completed":

1. The extraction-time verifier silently SKIPPED its closing-cash check when
   the final net-change row label lacked the word "after" (the MPERS
   SOCF-Indirect template labels it plainly "*Net increase (decrease) in cash
   and cash equivalents"). Now the verifier binds that row on "not before" and,
   if it still can't bind, says so loudly instead of skipping in silence.
2. NO cross-check validated SOCF's own internal footing, so every cross-check
   passed and the reviewer never fired. The new `socf_articulation` cross-check
   closes that at the fact layer.

These tests pin both. The verifier tests reproduce run-50's exact figures
(RM 61,976 gap). The cross-check tests run on the REAL MFRS *and* MPERS
templates so the structural net-change resolution is proven across the
"after" / no-"after" label difference (generalizability, not a one-template fix).
"""
from __future__ import annotations

import json
import shutil
import sqlite3
from collections import deque
from pathlib import Path

import openpyxl
import pytest

from statement_types import StatementType
from tools.verifier import verify_statement

REPO = Path(__file__).resolve().parent.parent


# ---------------------------------------------------------------------------
# Verifier (xlsx) — the silently-skipped articulation check now runs + fails.
# ---------------------------------------------------------------------------

def _build_socf_indirect(path, *, closing_cash: float, include_after_row: bool = True):
    """Minimal MPERS-style SOCF-Indirect sheet reproducing run-50's shape.

    The final net-change row is labelled WITHOUT "after" (as the MPERS template
    does). Subtotals are live formulas so the verifier evaluates them; closing
    cash is a hard-typed leaf. With run-50's numbers it foots to 32,666,248 but
    closing cash is the (correct) PDF figure 32,728,224 → a 61,976 gap.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCF-Indirect"
    ws["A1"] = "*Net cash flows from operating activities"
    ws["B1"] = 728419
    ws["A2"] = "*Net cash flows from investing activities"
    ws["B2"] = -2640040
    ws["A3"] = "*Net cash flows from financing activities"
    ws["B3"] = -30000000
    ws["A4"] = "*Net increase (decrease) in cash and cash equivalents before effect of exchange rate changes"
    ws["B4"] = "=B1+B2+B3"
    ws["A5"] = "Effect of exchange rate changes"
    ws["B5"] = 0
    row = 6
    if include_after_row:
        # NOTE: no "after" in the label — exactly the MPERS template wording.
        ws[f"A{row}"] = "*Net increase (decrease) in cash and cash equivalents"
        ws[f"B{row}"] = "=B4+B5"
        row += 1
    ws[f"A{row}"] = "Cash and cash equivalents at beginning of period"
    ws[f"B{row}"] = 64577869
    row += 1
    ws[f"A{row}"] = "Cash and cash equivalents at end of period"
    ws[f"B{row}"] = closing_cash
    wb.save(str(path))


def test_verifier_fails_non_articulating_socf_without_after_label(tmp_path):
    """Run-50 reproduction: closing cash != opening + net change → Check 2
    must FAIL, even though the final net-change row has no "after"."""
    path = tmp_path / "socf_bad.xlsx"
    _build_socf_indirect(path, closing_cash=32728224)  # correct PDF value; doesn't foot

    result = verify_statement(
        str(path), StatementType.SOCF, variant="Indirect", filing_level="company")

    assert result.is_balanced is False
    blob = " ".join(result.mismatches).lower()
    assert "cash at end" in blob
    # The 61,976 gap is named somewhere in the feedback/mismatches.
    assert "61976" in (result.feedback or "").replace(",", "") or \
        any("32666248" in m.replace(",", "") for m in result.mismatches)


def test_verifier_passes_articulating_socf_without_after_label(tmp_path):
    """The same template wording, but closing cash foots → Check 2 PASSES.
    Proves the new matching didn't just hard-fail every no-"after" template."""
    path = tmp_path / "socf_ok.xlsx"
    _build_socf_indirect(path, closing_cash=64577869 - 31911621)  # = 32,666,248

    result = verify_statement(
        str(path), StatementType.SOCF, variant="Indirect", filing_level="company")

    assert result.is_balanced is True


def test_verifier_fails_closed_when_net_change_row_absent(tmp_path):
    """When opening + closing cash are present but the net-change row can't be
    bound, the articulation check must NOT silently skip — it emits a visible
    diagnostic. (The silent skip is how run-50 shipped.)"""
    path = tmp_path / "socf_no_netchange.xlsx"
    _build_socf_indirect(path, closing_cash=32728224, include_after_row=False)

    result = verify_statement(
        str(path), StatementType.SOCF, variant="Indirect", filing_level="company")

    assert "articulation not verified" in (result.feedback or "").lower()


# ---------------------------------------------------------------------------
# Cross-check (facts + xlsx) — internal footing, both MFRS and MPERS.
# ---------------------------------------------------------------------------

from concept_model.cascade import recompute_after_turn  # noqa: E402
from concept_model.exporter import export_run_to_xlsx  # noqa: E402
from concept_model.importer import import_company_targets, import_template  # noqa: E402
from concept_model.parser import parse_template  # noqa: E402
from cross_checks.framework import FactsContext  # noqa: E402
from cross_checks.socf_articulation import SOCFArticulationCheck  # noqa: E402
from db.schema import init_db  # noqa: E402
from tests.shadow_diff import assert_cross_check_parity  # noqa: E402

SOCF_FIXTURES = {
    "mfrs": REPO / "XBRL-template-MFRS" / "Company" / "07-SOCF-Indirect.xlsx",
    "mpers": REPO / "XBRL-template-MPERS" / "Company" / "07-SOCF-Indirect.xlsx",
}


def _leaf_by_label(conn, tid, substr):
    r = conn.execute(
        "SELECT concept_uuid FROM concept_nodes WHERE template_id = ? "
        "AND kind = 'LEAF' AND lower(canonical_label) LIKE ? ORDER BY render_row",
        (tid, f"%{substr.lower()}%"),
    ).fetchone()
    return r[0] if r else None


def _final_net_change_uuid(conn, tid):
    """The COMPUTED final net-change concept, resolved the structural way the
    cross-check does (bottom-most cash 'net increase' row that isn't 'before')."""
    rows = conn.execute(
        "SELECT concept_uuid, canonical_label FROM concept_nodes "
        "WHERE template_id = ? ORDER BY render_row",
        (tid,),
    ).fetchall()
    found = None
    for cu, clabel in rows:
        norm = str(clabel).strip().lstrip("*").strip().lower()
        if "net increase" in norm and "cash" in norm and "before" not in norm:
            found = cu  # bottom-most wins
    return found


def _descendant_leaf(conn, uuid):
    seen, q = set(), deque([uuid])
    while q:
        n = q.popleft()
        if n in seen:
            continue
        seen.add(n)
        kind = conn.execute(
            "SELECT kind FROM concept_nodes WHERE concept_uuid = ?", (n,)
        ).fetchone()
        if kind and kind[0] == "LEAF":
            return n
        for (child,) in conn.execute(
            "SELECT child_uuid FROM concept_edges WHERE parent_uuid = ?", (n,)
        ):
            q.append(child)
    return None


def _seed(conn, run_id, uuid, value):
    conn.execute(
        "INSERT OR REPLACE INTO run_concept_facts(run_id, concept_uuid, period, "
        "entity_scope, value, value_status, source, updated_at) "
        "VALUES (?, ?, 'CY', 'Company', ?, 'observed', 'pdf', '2026-06-21Z')",
        (run_id, uuid, value),
    )


def _fact_value(conn, run_id, uuid):
    r = conn.execute(
        "SELECT value FROM run_concept_facts WHERE run_id = ? AND concept_uuid = ? "
        "AND period = 'CY' AND entity_scope = 'Company'",
        (run_id, uuid),
    ).fetchone()
    return r[0] if r else None


@pytest.mark.parametrize("standard", ["mfrs", "mpers"])
@pytest.mark.parametrize("articulates", [True, False])
def test_socf_articulation_crosscheck_e2e(tmp_path, standard, articulates):
    """On the REAL template (both standards): a non-articulating SOCF FAILS the
    new cross-check and an articulating one PASSES — and the xlsx + fact paths
    agree (parity). This is the deterministic catch that fires the reviewer."""
    fixture = SOCF_FIXTURES[standard]
    db = tmp_path / "xbrl.db"
    init_db(db)
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")

    tree = parse_template(str(fixture))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db, jp)
    import_company_targets(db, tid)

    cur = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
        "VALUES ('2026-06-21T00:00:00Z', 'x.pdf', 'running', '2026-06-21T00:00:00Z')"
    )
    run_id = int(cur.lastrowid)
    conn.commit()

    beginning = _leaf_by_label(conn, tid, "cash and cash equivalents at beginning of period")
    ending = _leaf_by_label(conn, tid, "cash and cash equivalents at end of period")
    netchg = _final_net_change_uuid(conn, tid)
    assert beginning and ending and netchg, "SOCF concepts not resolved"
    op_leaf = _descendant_leaf(conn, netchg)
    assert op_leaf, "no operating leaf under the net-change subtotal"

    # Seed an opening balance + one operating movement, cascade to compute the
    # net change, then set closing cash to foot (or miss by a clear gap).
    _seed(conn, run_id, beginning, 1000.0)
    _seed(conn, run_id, op_leaf, -100.0)
    conn.commit()
    recompute_after_turn(str(db), run_id)
    nc = _fact_value(conn, run_id, netchg)
    assert nc is not None, "cascade did not compute the net-change subtotal"

    closing = 1000.0 + nc + (0.0 if articulates else 50.0)
    _seed(conn, run_id, ending, closing)
    conn.commit()

    ctx = FactsContext(
        conn=conn, run_id=run_id,
        template_ids={StatementType.SOCF: tid},
        filing_level="company", filing_standard=standard,
    )
    facts = SOCFArticulationCheck().run_facts(ctx, tolerance=1.0)

    work = tmp_path / "SOCF_filled.xlsx"
    shutil.copyfile(fixture, work)
    export_run_to_xlsx(str(db), run_id, str(work), template_id=tid)
    xlsx = SOCFArticulationCheck().run(
        {StatementType.SOCF: str(work)}, tolerance=1.0,
        filing_level="company", filing_standard=standard)
    conn.close()

    expected = "passed" if articulates else "failed"
    assert facts.status == expected, facts.message
    assert xlsx.status == expected, xlsx.message
    assert_cross_check_parity(xlsx, facts)


def test_socf_articulation_registered_in_default_checks():
    """The check must be in the registry the server runs, or it never fires."""
    from cross_checks.framework import build_default_cross_checks
    names = {c.name for c in build_default_cross_checks()}
    assert "socf_articulation" in names
