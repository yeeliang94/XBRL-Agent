"""Phase 5 — SOCIE matrix end-to-end (import → facts → export).

Covers steps 5.3 (importer matrix targets), 5.4 (MPERS Group 4-block
vertical mapping + exporter routing) and 5.7 (E2E across MFRS×MPERS ×
Company×Group SOCIE).
"""
from __future__ import annotations

import json
import shutil
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from concept_model.parser import parse_template
from concept_model.importer import import_template
from concept_model.exporter import export_run_to_xlsx
from db.schema import init_db

_ROOT = Path(__file__).resolve().parent.parent
SOCIE = {
    "mfrs_company": _ROOT / "XBRL-template-MFRS" / "Company" / "09-SOCIE.xlsx",
    "mfrs_group": _ROOT / "XBRL-template-MFRS" / "Group" / "09-SOCIE.xlsx",
    "mpers_company": _ROOT / "XBRL-template-MPERS" / "Company" / "09-SOCIE.xlsx",
    "mpers_group": _ROOT / "XBRL-template-MPERS" / "Group" / "09-SOCIE.xlsx",
}


def _import(db: Path, fixture: Path, tmp_path: Path) -> str:
    tree = parse_template(str(fixture))
    jp = tmp_path / f"{tree.template_id}.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    return import_template(db, jp)


def _new_run(db: Path) -> int:
    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
            "VALUES (?, ?, ?, ?)",
            ("2026-05-22T00:00:00Z", "socie.pdf", "running", "2026-05-22T00:00:00Z"),
        )
        run_id = cur.lastrowid
        conn.commit()
        return run_id
    finally:
        conn.close()


# -- 5.3: importer writes shape + matrix_col + targets ----------------


def test_import_marks_template_matrix(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mfrs_company"], tmp_path)
    conn = sqlite3.connect(str(db))
    try:
        (shape,) = conn.execute(
            "SELECT shape FROM concept_templates WHERE template_id = ?", (tid,)
        ).fetchone()
    finally:
        conn.close()
    assert shape == "matrix"


def test_import_populates_matrix_col(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mfrs_company"], tmp_path)
    conn = sqlite3.connect(str(db))
    try:
        cols = {
            r[0] for r in conn.execute(
                "SELECT DISTINCT matrix_col FROM concept_nodes "
                "WHERE template_id = ? AND kind = 'MATRIX_CELL'", (tid,)
            ).fetchall()
        }
    finally:
        conn.close()
    assert "B" in cols and "X" in cols and len(cols) == 23


def test_import_writes_matrix_targets(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mfrs_company"], tmp_path)
    conn = sqlite3.connect(str(db))
    try:
        # A MATRIX_CELL at row 11 (Profit/loss), col B → 2 targets (CY/PY Company).
        node = conn.execute(
            "SELECT concept_uuid FROM concept_nodes "
            "WHERE template_id = ? AND render_row = 11 AND matrix_col = 'B'",
            (tid,),
        ).fetchone()
        assert node is not None
        rows = conn.execute(
            "SELECT period, entity_scope, target_col, target_row "
            "FROM concept_targets WHERE concept_uuid = ? ORDER BY period",
            (node[0],),
        ).fetchall()
    finally:
        conn.close()
    assert sorted(rows) == sorted([
        ("CY", "Company", "B", 11),
        ("PY", "Company", "B", 35),  # block 2 begins at 30 → 11+(30-6)=35
    ])


def test_reimport_flushes_stale_matrix_targets(tmp_path: Path) -> None:
    """Peer-review: a re-import must not leave orphaned concept_targets
    behind. We plant a bogus stale target then re-import and assert it's
    gone (DELETE-then-insert scoped to the template)."""
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mpers_company"], tmp_path)
    conn = sqlite3.connect(str(db))
    node = conn.execute(
        "SELECT concept_uuid FROM concept_nodes WHERE template_id = ? LIMIT 1",
        (tid,),
    ).fetchone()[0]
    # A stale dimension that the real geometry never emits.
    conn.execute(
        "INSERT INTO concept_targets(concept_uuid, entity_scope, period, "
        "target_sheet, target_row, target_col) VALUES (?, 'Group', 'FY', 'SOCIE', 99, 'Z')",
        (node,),
    )
    conn.commit()
    conn.close()

    _import(db, SOCIE["mpers_company"], tmp_path)  # re-import
    conn = sqlite3.connect(str(db))
    stale = conn.execute(
        "SELECT COUNT(*) FROM concept_targets WHERE period = 'FY'"
    ).fetchone()[0]
    conn.close()
    assert stale == 0, "stale concept_targets survived re-import"


def test_import_idempotent(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    _import(db, SOCIE["mpers_group"], tmp_path)
    conn = sqlite3.connect(str(db))
    n1 = conn.execute("SELECT COUNT(*) FROM concept_nodes").fetchone()[0]
    t1 = conn.execute("SELECT COUNT(*) FROM concept_targets").fetchone()[0]
    conn.close()
    _import(db, SOCIE["mpers_group"], tmp_path)
    conn = sqlite3.connect(str(db))
    n2 = conn.execute("SELECT COUNT(*) FROM concept_nodes").fetchone()[0]
    t2 = conn.execute("SELECT COUNT(*) FROM concept_targets").fetchone()[0]
    conn.close()
    assert (n1, t1) == (n2, t2)


# -- 5.4: MPERS Group 4-block vertical mapping ------------------------


def test_mpers_group_targets_map_blocks(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mpers_group"], tmp_path)
    conn = sqlite3.connect(str(db))
    try:
        node = conn.execute(
            "SELECT concept_uuid FROM concept_nodes "
            "WHERE template_id = ? AND render_row = 11 AND matrix_col = 'B'",
            (tid,),
        ).fetchone()
        rows = conn.execute(
            "SELECT period, entity_scope, target_row, target_col "
            "FROM concept_targets WHERE concept_uuid = ?", (node[0],)
        ).fetchall()
    finally:
        conn.close()
    assert sorted(rows) == sorted([
        ("CY", "Group", 11, "B"),
        ("PY", "Group", 35, "B"),
        ("CY", "Company", 59, "B"),
        ("PY", "Company", 83, "B"),
    ])


# -- 5.4 + 5.7: exporter routes matrix facts back ---------------------


def test_mpers_group_export_routes_to_blocks(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mpers_group"], tmp_path)
    run_id = _new_run(db)
    conn = sqlite3.connect(str(db))
    node = conn.execute(
        "SELECT concept_uuid FROM concept_nodes "
        "WHERE template_id = ? AND render_row = 11 AND matrix_col = 'B'", (tid,)
    ).fetchone()[0]
    seeded = [
        ("CY", "Group", 111.0), ("PY", "Group", 222.0),
        ("CY", "Company", 333.0), ("PY", "Company", 444.0),
    ]
    for period, scope, val in seeded:
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, updated_at) "
            "VALUES (?, ?, ?, ?, ?, 'observed', '2026Z')",
            (run_id, node, period, scope, val),
        )
    conn.commit()
    conn.close()

    work = tmp_path / "filled.xlsx"
    shutil.copyfile(SOCIE["mpers_group"], work)
    export_run_to_xlsx(db, run_id, str(work), filing_level="group")

    wb = openpyxl.load_workbook(str(work), data_only=False)
    ws = wb["SOCIE"]
    assert ws["B11"].value == 111.0   # Group CY
    assert ws["B35"].value == 222.0   # Group PY
    assert ws["B59"].value == 333.0   # Company CY
    assert ws["B83"].value == 444.0   # Company PY


def test_mfrs_company_export_routes_components(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mfrs_company"], tmp_path)
    run_id = _new_run(db)
    conn = sqlite3.connect(str(db))
    # Profit (loss) row 11, two components: B (issued capital) and C (retained).
    nodes = {
        r[0]: r[1] for r in conn.execute(
            "SELECT matrix_col, concept_uuid FROM concept_nodes "
            "WHERE template_id = ? AND render_row = 11 AND matrix_col IN ('B','C')",
            (tid,),
        ).fetchall()
    }
    conn.execute(
        "INSERT INTO run_concept_facts(run_id, concept_uuid, period, entity_scope, "
        "value, value_status, updated_at) VALUES (?,?,?,?,?,'observed','Z')",
        (run_id, nodes["C"], "CY", "Company", 999.0),
    )
    conn.execute(
        "INSERT INTO run_concept_facts(run_id, concept_uuid, period, entity_scope, "
        "value, value_status, updated_at) VALUES (?,?,?,?,?,'observed','Z')",
        (run_id, nodes["C"], "PY", "Company", 888.0),
    )
    conn.commit()
    conn.close()

    work = tmp_path / "filled.xlsx"
    shutil.copyfile(SOCIE["mfrs_company"], work)
    export_run_to_xlsx(db, run_id, str(work), filing_level="company")
    wb = openpyxl.load_workbook(str(work), data_only=False)
    ws = wb["SOCIE"]
    assert ws["C11"].value == 999.0   # retained earnings, CY block
    assert ws["C35"].value == 888.0   # retained earnings, PY block (row 35)


# -- regression: source stamping must not clobber matrix value cells --


def test_matrix_export_does_not_stamp_source_into_value_grid(tmp_path: Path) -> None:
    """Peer-review CRITICAL: on MFRS SOCIE the group source column F (and
    company D) sit INSIDE the equity-component value grid (B..X). A fact
    carrying a `source` must not overwrite those real value cells."""
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mfrs_group"], tmp_path)
    run_id = _new_run(db)
    conn = sqlite3.connect(str(db))
    # Profit row 11, col B fact carrying a source. The group source column
    # is F — a real equity-component value cell on SOCIE. Old code would
    # stamp the source string into F11 (order-independent: only this fact
    # touches row 11), corrupting the grid.
    node_b = conn.execute(
        "SELECT concept_uuid FROM concept_nodes "
        "WHERE template_id = ? AND render_row = 11 AND matrix_col = 'B'", (tid,)
    ).fetchone()[0]
    conn.execute(
        "INSERT INTO run_concept_facts(run_id, concept_uuid, period, entity_scope, "
        "value, value_status, source, updated_at) "
        "VALUES (?,?,?,?,?,'observed',?,'Z')",
        (run_id, node_b, "CY", "Group", 111.0, "pdf p.5"),
    )
    conn.commit()
    conn.close()

    work = tmp_path / "filled.xlsx"
    shutil.copyfile(SOCIE["mfrs_group"], work)
    export_run_to_xlsx(db, run_id, str(work), filing_level="group")
    wb = openpyxl.load_workbook(str(work), data_only=False)
    ws = wb["SOCIE"]
    assert ws["B11"].value == 111.0
    # F (group source col) must NOT receive the provenance string — it's a
    # value column in the SOCIE grid.
    assert ws["F11"].value != "pdf p.5"
    assert not isinstance(ws["F11"].value, str)


# -- 5.7: E2E across MFRS×MPERS × Company×Group ------------------------


@pytest.mark.parametrize("name", sorted(SOCIE.keys()))
def test_socie_e2e_all_geometries(name: str, tmp_path: Path) -> None:
    """Compose every layer (parse → import → seed → export) on each of the
    four SOCIE geometries and assert every mapped (period, entity_scope)
    target receives its fact in the right physical cell."""
    fixture = SOCIE[name]
    if not fixture.exists():
        pytest.skip(f"fixture missing: {fixture}")
    is_group = name.endswith("group")
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, fixture, tmp_path)
    run_id = _new_run(db)

    # Pick the "Profit (loss)" movement row, matrix_col B — present in
    # every SOCIE geometry.
    conn = sqlite3.connect(str(db))
    node = conn.execute(
        "SELECT concept_uuid FROM concept_nodes WHERE template_id = ? "
        "AND matrix_col = 'B' AND LOWER(canonical_label) LIKE '%profit%' "
        "ORDER BY render_row LIMIT 1",
        (tid,),
    ).fetchone()[0]
    targets = conn.execute(
        "SELECT period, entity_scope, target_row, target_col "
        "FROM concept_targets WHERE concept_uuid = ?", (node,)
    ).fetchall()
    # Seed a distinct value per (period, scope) target so a mis-route
    # would land the wrong number in the wrong cell.
    seeded: dict[tuple[int, str], float] = {}
    for i, (period, scope, trow, tcol) in enumerate(targets):
        val = 100.0 + i
        seeded[(trow, tcol)] = val
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, updated_at) "
            "VALUES (?, ?, ?, ?, ?, 'observed', 'Z')",
            (run_id, node, period, scope, val),
        )
    conn.commit()
    conn.close()
    assert seeded, "no targets to seed — geometry mapping is empty"

    work = tmp_path / "filled.xlsx"
    shutil.copyfile(fixture, work)
    export_run_to_xlsx(
        db, run_id, str(work),
        filing_level="group" if is_group else "company",
    )
    wb = openpyxl.load_workbook(str(work), data_only=False)
    ws = wb["SOCIE"]
    for (trow, tcol), val in seeded.items():
        assert ws[f"{tcol}{trow}"].value == val, (
            f"{name}: expected {val} at {tcol}{trow}, "
            f"got {ws[f'{tcol}{trow}'].value!r}"
        )


def test_mpers_company_period_as_column(tmp_path: Path) -> None:
    db = tmp_path / "x.db"
    init_db(db)
    tid = _import(db, SOCIE["mpers_company"], tmp_path)
    run_id = _new_run(db)
    conn = sqlite3.connect(str(db))
    node = conn.execute(
        "SELECT concept_uuid FROM concept_nodes "
        "WHERE template_id = ? AND render_row = 10 AND matrix_col = 'B'", (tid,)
    ).fetchone()[0]
    conn.execute(
        "INSERT INTO run_concept_facts(run_id, concept_uuid, period, entity_scope, "
        "value, value_status, updated_at) VALUES (?,?,?,?,?,'observed','Z')",
        (run_id, node, "CY", "Company", 50.0),
    )
    conn.execute(
        "INSERT INTO run_concept_facts(run_id, concept_uuid, period, entity_scope, "
        "value, value_status, updated_at) VALUES (?,?,?,?,?,'observed','Z')",
        (run_id, node, "PY", "Company", 60.0),
    )
    conn.commit()
    conn.close()

    work = tmp_path / "filled.xlsx"
    shutil.copyfile(SOCIE["mpers_company"], work)
    export_run_to_xlsx(db, run_id, str(work), filing_level="company")
    wb = openpyxl.load_workbook(str(work), data_only=False)
    ws = wb["SOCIE"]
    assert ws["B10"].value == 50.0   # CY → col B
    assert ws["C10"].value == 60.0   # PY → col C
