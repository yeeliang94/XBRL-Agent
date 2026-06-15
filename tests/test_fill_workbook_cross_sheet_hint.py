"""Cross-sheet redirect hint — 2026-06-15 SOFP OrderOfLiquidity incident.

Run 168 (Oriental 1936 Berhad, MFRS/Company) flagged an SOFP imbalance equal
to exactly the financial-asset-at-FVTPL (unit trust) amount: RM991,755 (CY) /
RM1,014,701 (PY). Root cause: the agent tried to write that value to the
SUB-sheet under "Investments other than investments accounted for using equity
method", which only exists as a direct DATA_ENTRY leaf on the FACE sheet (no
sub-sheet equivalent). The "no matching label" error said only "check the
label text", giving the agent nowhere to look — it gave up rather than writing
to the face row.

The writer already holds the whole-workbook label index, so it can see the
label IS writable one sheet over. The error must redirect the agent there.
"""
from __future__ import annotations

import openpyxl

from tools.fill_workbook import fill_workbook


def _make_two_sheet_sofp_like(tmp_path) -> str:
    """Face sheet with a direct-entry row that has no sub-sheet equivalent.

    Mirrors the real SOFP OrderOfLiquidity: face row "Investments other than
    investments accounted for using equity method" is a writable DATA_ENTRY
    cell; the sub-sheet jumps from joint ventures straight to receivables and
    never carries that label.
    """
    path = str(tmp_path / "sofp_like.xlsx")
    wb = openpyxl.Workbook()
    face = wb.active
    face.title = "SOFP-OrdOfLiq"
    face["A6"] = "Assets [abstract]"
    face["A7"] = "Total property, plant and equipment"
    face["B7"] = "='SOFP-Sub-OrdOfLiq'!B38"   # formula roll-up, like the real one
    face["A16"] = "Investments other than investments accounted for using equity method"
    # B16 intentionally empty — direct data-entry cell.

    sub = wb.create_sheet("SOFP-Sub-OrdOfLiq")
    sub["A6"] = "Property, plant and equipment [abstract]"
    sub["A38"] = "Total property, plant and equipment"
    sub["A95"] = "Trade and other receivables [abstract]"
    # No "Investments other than..." row anywhere on the sub-sheet.

    wb.save(path)
    wb.close()
    return path


def test_subsheet_miss_redirects_to_writable_face_row(tmp_path):
    template = _make_two_sheet_sofp_like(tmp_path)
    output = str(tmp_path / "filled.xlsx")
    # The exact mistargeting from run 168, msg 9.
    facts = [{
        "sheet": "SOFP-Sub-OrdOfLiq",
        "field_label": "Investments other than investments accounted for using equity method",
        "section": "assets",
        "col": 2,
        "value": 991_755,
    }]

    result = fill_workbook(template, output, facts)

    assert result.fields_written == 0
    assert len(result.errors) == 1
    err = result.errors[0]
    # Must name the sheet the label actually lives on, and the row.
    assert "SOFP-OrdOfLiq" in err, f"error must redirect to the face sheet. Got: {err}"
    assert "16" in err, f"error must name the writable row. Got: {err}"
    # Must NOT fall back to the unhelpful generic hint when a target exists.
    assert "check the exact label text" not in err.lower(), (
        f"a concrete redirect should replace the generic hint. Got: {err}"
    )


def test_face_row_write_succeeds(tmp_path):
    """The corrected write the agent should have made lands cleanly."""
    template = _make_two_sheet_sofp_like(tmp_path)
    output = str(tmp_path / "filled.xlsx")
    facts = [{
        "sheet": "SOFP-OrdOfLiq",
        "field_label": "Investments other than investments accounted for using equity method",
        "section": "assets",
        "col": 2,
        "value": 991_755,
    }]

    result = fill_workbook(template, output, facts)
    assert result.success, result.errors
    assert result.fields_written == 1

    wb = openpyxl.load_workbook(output)
    assert wb["SOFP-OrdOfLiq"].cell(row=16, column=2).value == 991_755
    wb.close()


def test_genuine_typo_still_gets_generic_hint(tmp_path):
    """No redirect when the label exists nowhere — keep the original guidance."""
    template = _make_two_sheet_sofp_like(tmp_path)
    output = str(tmp_path / "filled.xlsx")
    facts = [{
        "sheet": "SOFP-Sub-OrdOfLiq",
        "field_label": "Totally made up label that exists on no sheet",
        "col": 2,
        "value": 123,
    }]

    result = fill_workbook(template, output, facts)
    assert result.fields_written == 0
    err = result.errors[0]
    assert "check the exact label text" in err.lower(), (
        f"with no redirect target, keep the read_template hint. Got: {err}"
    )
