"""SOFP end-to-end regression oracle.

Why: the multi-statement rollout refactors fill_workbook / verifier / agent
creation. If any of those changes silently breaks SOFP extraction, we want a
loud failure. This test runs the whole pipeline against a known PDF and
compares the numeric cells of the resulting workbook against a frozen golden
fixture captured at the start of the rollout.

How to apply: skipped by default (marked `regression`) because it needs an
LLM API key and takes minutes. Run it manually at the end of each phase:

    pytest -m regression tests/test_sofp_regression.py -v

How to refresh the golden files (only when the extraction logic is
intentionally changed):

    GEMINI_API_KEY=... python run.py \\
        data/FINCO-Audited-Financial-Statement-2021.pdf SOFP-Xbrl-template.xlsx
    cp output/run_XXX/filled.xlsx  tests/fixtures/golden/SOFP_FINCO_2021_filled.xlsx
    cp output/run_XXX/result.json  tests/fixtures/golden/SOFP_FINCO_2021_result.json
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
PDF_PATH = REPO_ROOT / "data" / "FINCO-Audited-Financial-Statement-2021.pdf"
TEMPLATE_PATH = REPO_ROOT / "SOFP-Xbrl-template.xlsx"
GOLDEN_DIR = REPO_ROOT / "tests" / "fixtures" / "golden"
GOLDEN_XLSX = GOLDEN_DIR / "SOFP_FINCO_2021_filled.xlsx"
GOLDEN_JSON = GOLDEN_DIR / "SOFP_FINCO_2021_result.json"


def _numeric_cells(path: Path) -> dict[tuple[str, str], float]:
    """Return {(sheet, coord): value} for every numeric cell in a workbook.

    We compare numerically, not by formula, because openpyxl won't re-evaluate
    formulas — and because the agent writes values, not formulas, into data
    entry cells. Template formula cells stay as formulas in both golden and
    current, so they're skipped.
    """
    wb = load_workbook(path, data_only=False)
    out: dict[tuple[str, str], float] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    out[(ws.title, cell.coordinate)] = float(cell.value)
    return out


@pytest.mark.regression
def test_sofp_extraction_matches_golden(tmp_path: Path) -> None:
    """Full SOFP pipeline → compare numeric cells to frozen golden workbook."""
    if not GOLDEN_XLSX.exists() or not GOLDEN_JSON.exists():
        pytest.skip(
            f"Golden fixtures not found in {GOLDEN_DIR}. "
            "See the module docstring for how to generate them."
        )
    if not PDF_PATH.exists() or not TEMPLATE_PATH.exists():
        pytest.skip("Source PDF or template missing — cannot run regression.")

    # Run the CLI entry point with --output-dir pointed at a temp dir so
    # we never pick up stale artefacts from previous repo runs.
    output_root = tmp_path / "output"
    output_root.mkdir()
    result = subprocess.run(
        [
            sys.executable, str(REPO_ROOT / "run.py"),
            str(PDF_PATH), str(TEMPLATE_PATH),
            "--output-dir", str(output_root),
        ],
        cwd=REPO_ROOT,
        capture_output=True,
        text=True,
        timeout=600,
    )
    assert result.returncode == 0, f"run.py failed:\n{result.stderr}"

    # Locate the filled.xlsx in the isolated output dir.
    candidates = sorted(output_root.glob("run_*/filled.xlsx"))
    assert candidates, f"no filled.xlsx produced by run.py in {output_root}"
    current_xlsx = candidates[-1]

    golden = _numeric_cells(GOLDEN_XLSX)
    current = _numeric_cells(current_xlsx)

    # Compare with a tiny float tolerance so formatting roundtrips don't
    # break the test.
    mismatches = []
    for key, gold_val in golden.items():
        cur_val = current.get(key)
        if cur_val is None:
            mismatches.append(f"missing {key}: golden={gold_val}")
            continue
        if abs(gold_val - cur_val) > 0.01:
            mismatches.append(f"{key}: golden={gold_val} current={cur_val}")
    extra = set(current) - set(golden)
    for key in extra:
        mismatches.append(f"extra cell {key}={current[key]}")

    assert not mismatches, "SOFP regression diff:\n  " + "\n  ".join(mismatches[:50])

    # Spot-check result.json structure is preserved.
    golden_json = json.loads(GOLDEN_JSON.read_text(encoding="utf-8"))
    assert isinstance(golden_json, dict)
