"""Peer-review #6 regression — the download path calls
`overlay_notes_cells_into_workbook` against the workbook produced by
`workbook_merger.merge`, not against a bare notes template. Existing
coverage only exercises single-template xlsx files, which miss features
the merged workbook carries (multiple face sheets, merged styling,
different sheet-name collisions). This test seeds a realistic merged
workbook and confirms the overlay:

  * Updates the targeted notes cell with the DB's HTML.
  * Leaves face-statement sheets (SOFP, etc.) byte-identical in terms
    of cell values.
  * Leaves other notes cells that the DB doesn't claim untouched.

Coverage gap this closes: sheet-name ordering from merge could in
principle drift from the per-template ordering the overlay hit before;
a regression there used to be invisible because the single-template
tests never have >1 sheet.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from db import repository as repo
from db.schema import init_db
from notes.payload import NotesPayload
from notes.persistence import (
    overlay_notes_cells_into_workbook,
    persist_notes_cells,
)
from notes.writer import write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path
from statement_types import StatementType, template_path as face_template_path
from workbook_merger import merge


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    p = tmp_path / "xbrl.db"
    init_db(p)
    return p


def _seed_merged_workbook(tmp_path: Path) -> Path:
    """Build a real merged workbook (face SOFP + notes CORP_INFO) via
    the same `workbook_merger.merge` the production code uses."""
    src = face_template_path(StatementType.SOFP, "CuNonCu", level="company")
    face = tmp_path / "SOFP_filled.xlsx"
    wb = openpyxl.load_workbook(src)
    wb.save(face)
    wb.close()

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    notes = tmp_path / "NOTES_CORP_INFO_filled.xlsx"
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="OLD value from agent",
            evidence="Page 14",
            source_pages=[14],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(notes),
        filing_level="company",
        sheet_name="Notes-CI",
    )
    assert result.success

    merged = tmp_path / "filled.xlsx"
    merged_result = merge(
        workbook_paths={StatementType.SOFP: str(face)},
        output_path=str(merged),
        notes_workbook_paths={NotesTemplateType.CORP_INFO: str(notes)},
    )
    assert merged_result.success
    return merged


def _find_row(ws, needle: str) -> int:
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label and needle.lower() in str(label).lower():
            return row
    raise AssertionError(f"no row matching {needle!r}")


def test_overlay_on_merged_workbook_preserves_face_sheets_verbatim(
    tmp_path: Path, db_path: Path,
) -> None:
    """A face-statement sheet (SOFP-CuNonCu) must round-trip unchanged
    — same cell values in column A and B across every populated row.
    """
    merged = _seed_merged_workbook(tmp_path)

    # Snapshot the SOFP sheet cell-by-cell BEFORE overlay.
    before_vals: dict[tuple[int, int], object] = {}
    wb = openpyxl.load_workbook(merged)
    try:
        ws = wb["SOFP-CuNonCu"]
        for row in range(1, ws.max_row + 1):
            for col in (1, 2, 3, 4):
                before_vals[(row, col)] = ws.cell(row=row, column=col).value
    finally:
        wb.close()

    # Seed a notes_cells row so the overlay has work to do.
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "sample.pdf")
    wb = openpyxl.load_workbook(merged)
    try:
        row_num = _find_row(wb["Notes-CI"], "Financial reporting status")
    finally:
        wb.close()
    persist_notes_cells(
        db_path=str(db_path),
        run_id=run_id,
        sheet_name="Notes-CI",
        cells_written=[{
            "sheet": "Notes-CI",
            "row": row_num,
            "label": "Financial reporting status",
            "html": "<p>NEW value from editor</p>",
            "evidence": "Page 14",
            "source_pages": [14],
        }],
    )

    # Run the overlay against the merged workbook (the production path).
    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
    )

    wb = openpyxl.load_workbook(overlaid)
    try:
        # Notes cell reflects the DB value.
        notes_cell = wb["Notes-CI"].cell(row=row_num, column=2).value
        assert notes_cell == "NEW value from editor"

        # Every SOFP cell that existed before is still present with the
        # same value. The overlay must not touch face sheets.
        ws = wb["SOFP-CuNonCu"]
        after_vals: dict[tuple[int, int], object] = {}
        for row in range(1, ws.max_row + 1):
            for col in (1, 2, 3, 4):
                after_vals[(row, col)] = ws.cell(row=row, column=col).value
        assert after_vals == before_vals, (
            "overlay mutated SOFP cells — face sheets must be verbatim"
        )

        # Sheet ordering preserved (face before notes, from merge wiring).
        names = wb.sheetnames
        assert names.index("SOFP-CuNonCu") < names.index("Notes-CI")
    finally:
        wb.close()


def test_overlay_on_merged_workbook_leaves_unclaimed_notes_cells_alone(
    tmp_path: Path, db_path: Path,
) -> None:
    """If the DB only claims one cell on a sheet, the other cells on
    the same sheet (populated by the agent at run time) stay as-is.
    """
    merged = _seed_merged_workbook(tmp_path)

    # Snapshot every populated Notes-CI cell BEFORE overlay.
    before_col_b: dict[int, object] = {}
    wb = openpyxl.load_workbook(merged)
    try:
        ws = wb["Notes-CI"]
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val is not None:
                before_col_b[row] = val
        target_row = _find_row(ws, "Financial reporting status")
    finally:
        wb.close()

    # Seed the SAME cell with updated HTML.
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "sample.pdf")
    persist_notes_cells(
        db_path=str(db_path),
        run_id=run_id,
        sheet_name="Notes-CI",
        cells_written=[{
            "sheet": "Notes-CI",
            "row": target_row,
            "label": "Financial reporting status",
            "html": "<p>edited</p>",
            "evidence": "Page 14",
            "source_pages": [14],
        }],
    )

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
    )

    wb = openpyxl.load_workbook(overlaid)
    try:
        ws = wb["Notes-CI"]
        # Target cell was updated.
        assert ws.cell(row=target_row, column=2).value == "edited"
        # Every OTHER cell that was populated before must still be
        # populated with the same value.
        for row, val in before_col_b.items():
            if row == target_row:
                continue
            assert ws.cell(row=row, column=2).value == val, (
                f"overlay touched unclaimed cell at row {row}"
            )
    finally:
        wb.close()
