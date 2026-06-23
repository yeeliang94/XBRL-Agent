"""Peer-review HIGH (overlay can't represent deletions / stale evidence).

The notes overlay used to be purely additive: it wrote surviving `notes_cells`
rows to column B and never blanked rows the reviewer cleared/moved, nor
refreshed the evidence column. So a reviewer clear/move reintroduced the
original prose on download (duplicate / stale), and reviewer-authored evidence
never reached the export.

These tests pin the fix:
  * a tombstoned coordinate is BLANKED in the overlay (prose + evidence),
  * a "move" (tombstone source + new destination row) yields NO duplicate,
  * the filing-level evidence column is refreshed from notes_cells,
  * revert reconciles tombstones (clears cleared-row tombstones, re-tombstones
    authored rows) so the workbook matches the restored prose.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from db import repository as repo
from db.schema import init_db
from notes.payload import NotesPayload
from notes.persistence import (
    overlay_notes_cells_into_workbook,
    persist_notes_cells,
)
from notes.versioning import ensure_notes_snapshot, revert_notes_to_original
from notes.writer import evidence_col_for, write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path
from statement_types import StatementType, template_path as face_template_path
from workbook_merger import merge

_SHEET = "Notes-CI"
_EV_COL = evidence_col_for("company")  # column D


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    p = tmp_path / "xbrl.db"
    init_db(p)
    return p


def _seed_merged_workbook(tmp_path: Path) -> Path:
    """A real merged workbook with one filled prose note (prose + evidence)."""
    src = face_template_path(StatementType.SOFP, "CuNonCu", level="company")
    face = tmp_path / "SOFP_filled.xlsx"
    wb = openpyxl.load_workbook(src)
    wb.save(face)
    wb.close()

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    notes = tmp_path / "NOTES_CORP_INFO_filled.xlsx"
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="ORIGINAL prose from the agent",
            evidence="Page 14",
            source_pages=[14],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(notes),
        filing_level="company",
        sheet_name=_SHEET,
    )
    assert result.success

    merged = tmp_path / "filled.xlsx"
    merged_result = merge(
        workbook_paths={StatementType.SOFP: str(face)},
        output_path=str(merged),
        notes_workbook_paths={NotesTemplateType.CORP_INFO: str(notes)},
    )
    assert merged_result.success
    return merged


def _find_row(ws, needle: str) -> int:
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label and needle.lower() in str(label).lower():
            return row
    raise AssertionError(f"no row matching {needle!r}")


def _target_row(merged: Path) -> int:
    wb = openpyxl.load_workbook(merged)
    try:
        return _find_row(wb[_SHEET], "Financial reporting status")
    finally:
        wb.close()


def test_tombstone_blanks_cleared_cell_in_overlay(tmp_path, db_path):
    """A reviewer clear (notes_cells deleted + tombstone) blanks the workbook
    cell that was written at merge time — no stale prose on download."""
    merged = _seed_merged_workbook(tmp_path)
    row = _target_row(merged)
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "x.pdf")
        # The reviewer cleared the cell: no notes_cells row, but a tombstone.
        repo.add_notes_tombstone(conn, run_id=run_id, sheet=_SHEET, row=row)

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
        filing_level="company",
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        ws = wb[_SHEET]
        assert ws.cell(row=row, column=2).value is None, "prose not blanked"
        assert ws.cell(row=row, column=_EV_COL).value is None, "evidence not blanked"
    finally:
        wb.close()


def test_move_produces_no_duplicate(tmp_path, db_path):
    """Moving prose to a new row (tombstone source + destination notes_cells)
    leaves the source blank — the original is NOT reintroduced as a duplicate."""
    merged = _seed_merged_workbook(tmp_path)
    src_row = _target_row(merged)
    dst_row = src_row + 3  # an arbitrary other physical row on the sheet
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "x.pdf")
    persist_notes_cells(
        db_path=str(db_path), run_id=run_id, sheet_name=_SHEET,
        cells_written=[{
            "sheet": _SHEET, "row": dst_row, "label": "Moved",
            "html": "<p>MOVED prose</p>", "evidence": "Page 14",
            "source_pages": [14],
        }],
    )
    with repo.db_session(db_path) as conn:
        repo.add_notes_tombstone(conn, run_id=run_id, sheet=_SHEET, row=src_row)

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
        filing_level="company",
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        ws = wb[_SHEET]
        assert ws.cell(row=src_row, column=2).value is None, "source not blanked"
        assert ws.cell(row=dst_row, column=2).value == "MOVED prose"
    finally:
        wb.close()


def test_rerun_after_review_does_not_blank_repopulated_cell(tmp_path, db_path):
    """Peer-review HIGH: a notes-agent rerun repopulates a row the reviewer had
    cleared. The overlay must write the fresh cell and NOT blank it with the
    stale tombstone — both because persist_notes_cells drops the sheet's
    tombstones and because the overlay refuses to blank a coord with live prose."""
    merged = _seed_merged_workbook(tmp_path)
    row = _target_row(merged)
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "x.pdf")
        # Reviewer previously cleared the row (tombstone present).
        repo.add_notes_tombstone(conn, run_id=run_id, sheet=_SHEET, row=row)
    # A notes-agent rerun now repopulates the same row.
    persist_notes_cells(
        db_path=str(db_path), run_id=run_id, sheet_name=_SHEET,
        cells_written=[{
            "sheet": _SHEET, "row": row, "label": "Financial reporting status",
            "html": "<p>FRESH rerun content</p>", "evidence": "Page 3",
            "source_pages": [3],
        }],
    )
    # persist cleared the sheet's tombstones.
    with repo.db_session(db_path) as conn:
        assert (_SHEET, row) not in set(repo.fetch_notes_tombstones(conn, run_id))

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
        filing_level="company",
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        assert wb[_SHEET].cell(row=row, column=2).value == "FRESH rerun content"
    finally:
        wb.close()


def test_overlay_never_blanks_live_coord_even_with_stale_tombstone(tmp_path, db_path):
    """Belt-and-suspenders: even if a tombstone somehow co-exists with a live
    cell at the same coord, the overlay writes the live prose and skips blanking."""
    merged = _seed_merged_workbook(tmp_path)
    row = _target_row(merged)
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "x.pdf")
        repo.upsert_notes_cell(conn, run_id=run_id, sheet=_SHEET, row=row,
                               label="Financial reporting status",
                               html="<p>LIVE</p>", evidence="Page 1")
        # Inject a stale tombstone directly (bypassing the persist cleanup).
        repo.add_notes_tombstone(conn, run_id=run_id, sheet=_SHEET, row=row)

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
        filing_level="company",
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        assert wb[_SHEET].cell(row=row, column=2).value == "LIVE"
    finally:
        wb.close()


def test_evidence_column_refreshed_from_notes_cells(tmp_path, db_path):
    """Reviewer-updated evidence in notes_cells reaches the export's evidence
    column (D for Company) — not left stale/blank."""
    merged = _seed_merged_workbook(tmp_path)
    row = _target_row(merged)
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "x.pdf")
    persist_notes_cells(
        db_path=str(db_path), run_id=run_id, sheet_name=_SHEET,
        cells_written=[{
            "sheet": _SHEET, "row": row, "label": "Financial reporting status",
            "html": "<p>edited prose</p>", "evidence": "Pages 14, 99 — reviewer",
            "source_pages": [14, 99],
        }],
    )
    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
        filing_level="company",
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        ws = wb[_SHEET]
        assert ws.cell(row=row, column=2).value == "edited prose"
        assert ws.cell(row=row, column=_EV_COL).value == "Pages 14, 99 — reviewer"
    finally:
        wb.close()


def test_revert_reconciles_tombstones(tmp_path, db_path):
    """Revert clears a cleared-row tombstone (the row is restored) and
    re-tombstones a reviewer-authored row (so its xlsx prose is blanked)."""
    merged = _seed_merged_workbook(tmp_path)
    orig_row = _target_row(merged)
    authored_row = orig_row + 5
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(conn, "x.pdf")
    # Original extraction prose (one row).
    persist_notes_cells(
        db_path=str(db_path), run_id=run_id, sheet_name=_SHEET,
        cells_written=[{
            "sheet": _SHEET, "row": orig_row, "label": "Financial reporting status",
            "html": "<p>ORIGINAL prose from the agent</p>", "evidence": "Page 14",
            "source_pages": [14],
        }],
    )
    # Snapshot the original (what revert restores to).
    ensure_notes_snapshot(str(db_path), run_id)
    # Reviewer then: clears the original row + authors a brand-new row.
    with repo.db_session(db_path) as conn:
        conn.execute(
            "DELETE FROM notes_cells WHERE run_id=? AND sheet=? AND row=?",
            (run_id, _SHEET, orig_row),
        )
        repo.add_notes_tombstone(conn, run_id=run_id, sheet=_SHEET, row=orig_row)
        repo.upsert_notes_cell(conn, run_id=run_id, sheet=_SHEET, row=authored_row,
                               label="Authored", html="<p>AUTHORED</p>",
                               evidence="Page 20")

    out = revert_notes_to_original(str(db_path), run_id)
    assert out["reverted"] is True

    with repo.db_session(db_path) as conn:
        # The original row is back; the authored row is gone.
        rows = {(c.sheet, c.row) for c in repo.list_notes_cells_for_run(conn, run_id)}
        assert (_SHEET, orig_row) in rows
        assert (_SHEET, authored_row) not in rows
        # Tombstones: original-row tombstone cleared, authored-row tombstoned.
        tombs = set(repo.fetch_notes_tombstones(conn, run_id))
        assert (_SHEET, orig_row) not in tombs
        assert (_SHEET, authored_row) in tombs

    # And the overlay reflects exactly the original: orig restored, authored blank.
    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
        filing_level="company",
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        ws = wb[_SHEET]
        assert ws.cell(row=orig_row, column=2).value == "ORIGINAL prose from the agent"
        assert ws.cell(row=authored_row, column=2).value is None
    finally:
        wb.close()
