"""Phase 5 MPERS hardening — reproduce the three run-#105 SOCIE
cross-check failures in a test, then drive them green.

Red tests for `docs/PLAN-mpers-notes-hardening.md` Phase 5.

Run #105 produced:
  sopl_to_socie_profit   failed — SOPL (-20678) vs SOCIE (-21329), diff 651
  soci_to_socie_tci      failed — "Could not find TCI values: SOCI=-20678, SOCIE=None"
  socie_to_sofp_equity   failed — "Could not find equity values: SOCIE=None, SOFP=963391"

The SOCIE-side value-lookups pinned col 24 (X = "Total") which is
the MFRS matrix layout. MPERS SOCIE has max_col 4 — col 24 is empty
and the helpers return None. The fix: when the filing standard is
MPERS, read col 2 (B = CY) / col 3 (C = PY) like every other MPERS
statement.

Fixtures here synthesise MPERS-shaped workbooks with openpyxl rather
than shipping real filled files — keeps the tests self-contained and
lets us exercise both "all values match" and "one value missing"
without needing a real run.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from cross_checks.socie_to_sofp_equity import SOCIEToSOFPEquityCheck
from cross_checks.sopl_to_socie_profit import SOPLToSOCIEProfitCheck
from cross_checks.soci_to_socie_tci import SOCIToSOCIETCICheck
from statement_types import StatementType


# ---------------------------------------------------------------------------
# Synthetic MPERS / MFRS SOCIE fixtures
# ---------------------------------------------------------------------------

def _build_mpers_company_socie(path: Path, equity_cy: float, equity_py: float,
                                profit_cy: float, tci_cy: float):
    """Minimal MPERS Company SOCIE — matches the 09-SOCIE.xlsx layout
    observed in run #105 (max_row 44, max_col 2-3, labels carry SSM
    type suffixes, `Equity at end of period` at row 44)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"
    rows = [
        (3, "Disclosure of changes in equity [abstract]"),
        (24, "Statement of changes in equity [line items]"),
        (25, "Equity at beginning of period"),
        (30, "Profit (loss)"),
        (31, "Total other comprehensive income"),
        (32, "*Total comprehensive income"),
        (43, "*Total increase (decrease) in equity"),
        (44, "Equity at end of period"),
    ]
    for r, lbl in rows:
        ws.cell(r, 1).value = lbl
    # Populate values in cols B (CY) / C (PY) like MPERS.
    ws.cell(30, 2).value = profit_cy
    ws.cell(31, 2).value = 0
    ws.cell(32, 2).value = tci_cy  # concrete value (not a formula) to keep
                                    # the cross-check test data_only-agnostic
    ws.cell(44, 2).value = equity_cy
    ws.cell(44, 3).value = equity_py
    wb.save(path)


def _build_mpers_company_sofp(path: Path, total_equity: float):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-OrdOfLiq"
    ws.cell(1, 1).value = "Label"
    ws.cell(1, 2).value = "CY"
    ws.cell(5, 1).value = "Total equity"
    ws.cell(5, 2).value = total_equity
    wb.save(path)


def _build_mpers_company_sopl(path: Path, profit: float):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Nature"
    ws.cell(1, 1).value = "Label"
    ws.cell(1, 2).value = "CY"
    ws.cell(5, 1).value = "Profit (loss)"
    ws.cell(5, 2).value = profit
    wb.save(path)


def _build_mpers_company_soci(path: Path, tci: float):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCI-BeforeOfTax"
    ws.cell(1, 1).value = "Label"
    ws.cell(1, 2).value = "CY"
    ws.cell(5, 1).value = "Total comprehensive income"
    ws.cell(5, 2).value = tci
    wb.save(path)


# ---------------------------------------------------------------------------
# 5.1 MPERS — all three checks must resolve their SOCIE-side values
# ---------------------------------------------------------------------------

@pytest.fixture
def mpers_run_config():
    return {
        "statements_to_run": {
            StatementType.SOFP, StatementType.SOCIE,
            StatementType.SOPL, StatementType.SOCI,
        },
        "variants": {StatementType.SOCIE: "Default"},
        "filing_standard": "mpers",
        "filing_level": "company",
    }


@pytest.fixture
def mpers_workbooks(tmp_path):
    paths = {
        StatementType.SOCIE: tmp_path / "SOCIE.xlsx",
        StatementType.SOFP: tmp_path / "SOFP.xlsx",
        StatementType.SOPL: tmp_path / "SOPL.xlsx",
        StatementType.SOCI: tmp_path / "SOCI.xlsx",
    }
    # Concrete reconciling values — same across all three checks.
    equity_cy = 963391.0
    equity_py = 984069.0
    profit = -20678.0
    tci = -20678.0  # TCI == profit (no OCI)
    _build_mpers_company_socie(
        paths[StatementType.SOCIE], equity_cy, equity_py, profit, tci,
    )
    _build_mpers_company_sofp(paths[StatementType.SOFP], equity_cy)
    _build_mpers_company_sopl(paths[StatementType.SOPL], profit)
    _build_mpers_company_soci(paths[StatementType.SOCI], tci)
    return {k: str(v) for k, v in paths.items()}


def test_mpers_socie_to_sofp_equity_resolves(mpers_workbooks, mpers_run_config):
    """MPERS Company SOCIE stores `Equity at end of period` at col 2,
    not col 24. The check must read col 2 when the run is MPERS."""
    check = SOCIEToSOFPEquityCheck()
    result = check.run(
        mpers_workbooks, tolerance=1.0,
        filing_level="company", filing_standard="mpers",
    )
    assert result.status == "passed", (
        f"MPERS SOCIE → SOFP equity check should pass on matching "
        f"values — got {result.status}: {result.message}"
    )


def test_mpers_sopl_to_socie_profit_resolves(mpers_workbooks, mpers_run_config):
    check = SOPLToSOCIEProfitCheck()
    result = check.run(
        mpers_workbooks, tolerance=1.0,
        filing_level="company", filing_standard="mpers",
    )
    assert result.status == "passed", (
        f"MPERS SOPL → SOCIE profit check failed: {result.message}"
    )


def test_mpers_soci_to_socie_tci_resolves(mpers_workbooks, mpers_run_config):
    check = SOCIToSOCIETCICheck()
    result = check.run(
        mpers_workbooks, tolerance=1.0,
        filing_level="company", filing_standard="mpers",
    )
    assert result.status == "passed", (
        f"MPERS SOCI → SOCIE TCI check failed: {result.message}"
    )


# ---------------------------------------------------------------------------
# 5.1 MFRS regression — existing col 24 path must keep working
# ---------------------------------------------------------------------------

def _build_mfrs_company_socie(path: Path, equity_cy: float, profit: float, tci: float):
    """Minimal MFRS SOCIE in the pre-existing matrix layout — values go
    in col 24 (Total) for each row we care about. A non-zero NCI
    numeric in col 23 triggers `has_nci_data` so `socie_column` picks
    col 24 (Total) instead of falling through to col 3 (Retained)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"
    ws.cell(6, 1).value = "*Equity at beginning of period"
    ws.cell(11, 1).value = "*Profit (loss)"
    ws.cell(11, 24).value = profit
    ws.cell(11, 23).value = 0.0001  # non-zero NCI value to trigger
                                     # the Total-column code path
    ws.cell(13, 1).value = "*Total comprehensive income"
    ws.cell(13, 24).value = tci
    ws.cell(25, 1).value = "*Equity at end of period"
    ws.cell(25, 24).value = equity_cy
    wb.save(path)


@pytest.fixture
def mfrs_workbooks(tmp_path):
    paths = {
        StatementType.SOCIE: tmp_path / "SOCIE_mfrs.xlsx",
        StatementType.SOFP: tmp_path / "SOFP_mfrs.xlsx",
        StatementType.SOPL: tmp_path / "SOPL_mfrs.xlsx",
        StatementType.SOCI: tmp_path / "SOCI_mfrs.xlsx",
    }
    equity_cy = 100.0
    profit = 50.0
    tci = 50.0
    _build_mfrs_company_socie(paths[StatementType.SOCIE], equity_cy, profit, tci)
    _build_mpers_company_sofp(paths[StatementType.SOFP], equity_cy)
    _build_mpers_company_sopl(paths[StatementType.SOPL], profit)
    _build_mpers_company_soci(paths[StatementType.SOCI], tci)
    return {k: str(v) for k, v in paths.items()}


def test_mfrs_socie_to_sofp_equity_still_resolves(mfrs_workbooks):
    """Regression — MFRS runs must continue reading col 24 as before."""
    check = SOCIEToSOFPEquityCheck()
    result = check.run(
        mfrs_workbooks, tolerance=1.0,
        filing_level="company", filing_standard="mfrs",
    )
    assert result.status == "passed", (
        f"MFRS regression: SOCIE → SOFP broke after Phase 5 fix: "
        f"{result.message}"
    )


def test_mfrs_sopl_to_socie_profit_still_resolves(mfrs_workbooks):
    check = SOPLToSOCIEProfitCheck()
    result = check.run(
        mfrs_workbooks, tolerance=1.0,
        filing_level="company", filing_standard="mfrs",
    )
    assert result.status == "passed", result.message


def test_mfrs_soci_to_socie_tci_still_resolves(mfrs_workbooks):
    check = SOCIToSOCIETCICheck()
    result = check.run(
        mfrs_workbooks, tolerance=1.0,
        filing_level="company", filing_standard="mfrs",
    )
    assert result.status == "passed", result.message
