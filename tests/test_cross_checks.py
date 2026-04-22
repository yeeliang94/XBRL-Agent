"""Tests for the cross-check framework (Phase 5, Step 5.1 + 5.3)."""
from __future__ import annotations

import os
import tempfile

import openpyxl
import pytest

from statement_types import StatementType
from cross_checks.framework import (
    CrossCheck,
    CrossCheckResult,
    run_all,
)


# --- Fake checks for framework tests ---

class AlwaysPassCheck:
    """A trivial check that always passes — proves the framework calls it."""
    name = "always_pass"
    required_statements = {StatementType.SOFP}

    def applies_to(self, run_config) -> bool:
        return True

    def run(self, workbook_paths, tolerance, filing_level="company") -> CrossCheckResult:
        return CrossCheckResult(
            name=self.name,
            status="passed",
            expected=100.0,
            actual=100.0,
            diff=0.0,
            tolerance=tolerance,
            message="Values match",
        )


class AlwaysFailCheck:
    """A check that always fails — proves failures are reported correctly."""
    name = "always_fail"
    required_statements = {StatementType.SOFP, StatementType.SOPL}

    def applies_to(self, run_config) -> bool:
        return True

    def run(self, workbook_paths, tolerance, filing_level="company") -> CrossCheckResult:
        return CrossCheckResult(
            name=self.name,
            status="failed",
            expected=100.0,
            actual=200.0,
            diff=100.0,
            tolerance=1.0,
            message="Mismatch",
        )


class VariantSpecificCheck:
    """A check that only applies to SOCF Indirect variant."""
    name = "indirect_only"
    required_statements = {StatementType.SOCF}

    def applies_to(self, run_config) -> bool:
        return run_config.get("variants", {}).get(StatementType.SOCF) == "Indirect"

    def run(self, workbook_paths, tolerance, filing_level="company") -> CrossCheckResult:
        return CrossCheckResult(
            name=self.name,
            status="passed",
            expected=0.0,
            actual=0.0,
            diff=0.0,
            tolerance=tolerance,
            message="OK",
        )


# --- Step 5.1: Framework tests ---

class TestCrossCheckFramework:
    def test_framework_runs_all_registered(self):
        """A single fake check that always passes → run_all returns 1 pass, 0 fail."""
        checks = [AlwaysPassCheck()]
        run_config = {"statements_to_run": {StatementType.SOFP}}
        # Provide a dummy path so the workbook-path guard is satisfied
        paths = {StatementType.SOFP: "/tmp/fake.xlsx"}
        results = run_all(checks, workbook_paths=paths, run_config=run_config)

        assert len(results) == 1
        assert results[0].status == "passed"
        assert results[0].name == "always_pass"

    def test_framework_returns_multiple_results(self):
        """Multiple checks produce multiple results in order."""
        checks = [AlwaysPassCheck(), AlwaysFailCheck()]
        run_config = {
            "statements_to_run": {StatementType.SOFP, StatementType.SOPL},
        }
        paths = {StatementType.SOFP: "/tmp/f1.xlsx", StatementType.SOPL: "/tmp/f2.xlsx"}
        results = run_all(checks, workbook_paths=paths, run_config=run_config)

        assert len(results) == 2
        assert results[0].status == "passed"
        assert results[1].status == "failed"

    def test_result_carries_expected_actual_diff(self):
        """CrossCheckResult fields are set correctly by the check."""
        checks = [AlwaysFailCheck()]
        run_config = {
            "statements_to_run": {StatementType.SOFP, StatementType.SOPL},
        }
        paths = {StatementType.SOFP: "/tmp/f1.xlsx", StatementType.SOPL: "/tmp/f2.xlsx"}
        results = run_all(checks, workbook_paths=paths, run_config=run_config)
        r = results[0]

        assert r.expected == 100.0
        assert r.actual == 200.0
        assert r.diff == 100.0
        assert r.message == "Mismatch"


# --- Step 5.3: Variant-aware + missing-statement-aware tests ---

class TestCrossCheckSelection:
    def test_pending_when_statement_not_run(self):
        """If a required statement wasn't extracted, check returns 'pending'."""
        checks = [AlwaysFailCheck()]  # requires SOFP + SOPL
        # Only SOFP was run — SOPL is missing
        run_config = {"statements_to_run": {StatementType.SOFP}}
        results = run_all(checks, workbook_paths={}, run_config=run_config)

        assert len(results) == 1
        assert results[0].status == "pending"
        assert "SOPL" in results[0].message

    def test_not_applicable_when_variant_mismatch(self):
        """A variant-specific check returns 'not_applicable' when variant doesn't match."""
        checks = [VariantSpecificCheck()]
        run_config = {
            "statements_to_run": {StatementType.SOCF},
            "variants": {StatementType.SOCF: "Direct"},  # check wants Indirect
        }
        paths = {StatementType.SOCF: "/tmp/fake.xlsx"}
        results = run_all(checks, workbook_paths=paths, run_config=run_config)

        assert len(results) == 1
        assert results[0].status == "not_applicable"

    def test_variant_match_runs_check(self):
        """When variant matches, the check runs normally."""
        checks = [VariantSpecificCheck()]
        run_config = {
            "statements_to_run": {StatementType.SOCF},
            "variants": {StatementType.SOCF: "Indirect"},
        }
        paths = {StatementType.SOCF: "/tmp/fake.xlsx"}
        results = run_all(checks, workbook_paths=paths, run_config=run_config)

        assert len(results) == 1
        assert results[0].status == "passed"

    def test_socie_missing_triggers_multiple_pending(self):
        """SOCIE is needed by multiple checks — skipping it creates multiple pending."""
        # Two checks both require SOCIE
        class CheckA:
            name = "check_a"
            required_statements = {StatementType.SOPL, StatementType.SOCIE}
            def applies_to(self, rc): return True
            def run(self, wp, tol, filing_level="company"): return CrossCheckResult(
                name=self.name, status="passed", message="OK")

        class CheckB:
            name = "check_b"
            required_statements = {StatementType.SOCI, StatementType.SOCIE}
            def applies_to(self, rc): return True
            def run(self, wp, tol, filing_level="company"): return CrossCheckResult(
                name=self.name, status="passed", message="OK")

        # Run without SOCIE
        run_config = {
            "statements_to_run": {StatementType.SOPL, StatementType.SOCI},
        }
        results = run_all([CheckA(), CheckB()], workbook_paths={}, run_config=run_config)

        assert all(r.status == "pending" for r in results)
        assert all("SOCIE" in r.message for r in results)

    def test_missing_workbook_returns_failed(self):
        """If statement was selected but agent failed (no workbook), check returns 'failed'."""
        checks = [AlwaysPassCheck()]  # requires SOFP
        run_config = {"statements_to_run": {StatementType.SOFP}}
        # SOFP was run but produced no workbook — not in workbook_paths
        results = run_all(checks, workbook_paths={}, run_config=run_config)

        assert len(results) == 1
        assert results[0].status == "failed"
        assert "Workbook missing" in results[0].message

    def test_check_exception_caught_gracefully(self):
        """If a check's .run() raises, it returns failed instead of crashing."""
        class BrokenCheck:
            name = "broken"
            required_statements = {StatementType.SOFP}
            def applies_to(self, rc): return True
            def run(self, wp, tol, filing_level="company"): raise RuntimeError("boom")

        checks = [BrokenCheck()]
        run_config = {"statements_to_run": {StatementType.SOFP}}
        results = run_all(
            checks,
            workbook_paths={StatementType.SOFP: "/tmp/fake.xlsx"},
            run_config=run_config,
        )

        assert len(results) == 1
        assert results[0].status == "failed"
        assert "boom" in results[0].message


# ---------------------------------------------------------------------------
# Phase 4 — MPERS pipeline wiring (mpers_wiring_crosschecks)
# ---------------------------------------------------------------------------


def _make_workbook(sheets: dict[str, list[list]], path: str) -> None:
    """Build a minimal xlsx workbook in the same shape test_cross_checks_impl uses."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
    wb.close()


@pytest.fixture
def tmp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield d


# --- Step 4.1: framework reads applies_to_standard ---

@pytest.mark.mpers_wiring_crosschecks
def test_framework_skips_check_whose_standard_set_excludes_run():
    """A check whose applies_to_standard excludes the run's filing_standard
    must short-circuit to not_applicable, not crash or run blindly."""
    class MfrsOnlyCheck:
        name = "mfrs_only"
        required_statements = {StatementType.SOFP}
        applies_to_standard = frozenset({"mfrs"})
        def applies_to(self, rc): return True
        def run(self, wp, tol, filing_level="company"):
            raise AssertionError("must not run on MPERS")

    results = run_all(
        [MfrsOnlyCheck()],
        workbook_paths={StatementType.SOFP: "/tmp/fake.xlsx"},
        run_config={
            "statements_to_run": {StatementType.SOFP},
            "filing_standard": "mpers",
        },
    )
    assert len(results) == 1
    assert results[0].status == "not_applicable"
    assert "MPERS" in results[0].message


# --- Step 4.2: SOCIE-consuming checks gate out on SoRE ---

@pytest.mark.mpers_wiring_crosschecks
def test_sopl_to_socie_check_not_applicable_on_sore(tmp_dir):
    """MPERS+SoRE runs skip the three SOCIE-matrix checks — they'd fail on
    the missing TCI / per-component columns."""
    from cross_checks.sopl_to_socie_profit import SOPLToSOCIEProfitCheck
    from cross_checks.soci_to_socie_tci import SOCIToSOCIETCICheck
    from cross_checks.socie_to_sofp_equity import SOCIEToSOFPEquityCheck

    # SoRE has no TCI row / no Total (X) column — workbook fixture models that
    # (minimal, just the SoRE sheet with closing RE).
    sopl_path = os.path.join(tmp_dir, "sopl.xlsx")
    _make_workbook({
        "SOPL-Function": [["*Profit (loss)", 1000.0, 800.0]],
    }, sopl_path)
    socie_path = os.path.join(tmp_dir, "socie.xlsx")
    _make_workbook({
        "SoRE": [["*Retained earnings at end of period", 5000.0, 4200.0]],
    }, socie_path)
    sofp_path = os.path.join(tmp_dir, "sofp.xlsx")
    _make_workbook({
        "SOFP-CuNonCu": [
            ["*Total assets", 10_000, 8_000],
            ["*Total equity and liabilities", 10_000, 8_000],
            ["Retained earnings", 5000.0, 4200.0],
            ["*Total equity", 5000.0, 4200.0],
        ],
    }, sofp_path)
    soci_path = os.path.join(tmp_dir, "soci.xlsx")
    _make_workbook({
        "SOCI-NetOfTax": [["*Total comprehensive income", 1000.0, 800.0]],
    }, soci_path)

    run_config = {
        "statements_to_run": {
            StatementType.SOPL, StatementType.SOCI,
            StatementType.SOCIE, StatementType.SOFP,
        },
        "variants": {StatementType.SOCIE: "SoRE"},
        "filing_standard": "mpers",
    }
    paths = {
        StatementType.SOPL: sopl_path,
        StatementType.SOCI: soci_path,
        StatementType.SOCIE: socie_path,
        StatementType.SOFP: sofp_path,
    }
    results = run_all(
        [SOPLToSOCIEProfitCheck(), SOCIToSOCIETCICheck(), SOCIEToSOFPEquityCheck()],
        workbook_paths=paths, run_config=run_config,
    )
    # All three must gate out — they'd fail outright on a SoRE sheet.
    assert [r.status for r in results] == ["not_applicable"] * 3


@pytest.mark.mpers_wiring_crosschecks
def test_sofp_balance_still_runs_on_sore(tmp_dir):
    """The SOFP balance identity is independent of SOCIE variant — must still fire."""
    from cross_checks.sofp_balance import SOFPBalanceCheck

    sofp_path = os.path.join(tmp_dir, "sofp.xlsx")
    _make_workbook({
        "SOFP-CuNonCu": [
            ["*Total assets", 1000.0, 800.0],
            ["*Total equity and liabilities", 1000.0, 800.0],
        ],
    }, sofp_path)
    run_config = {
        "statements_to_run": {StatementType.SOFP},
        "variants": {StatementType.SOCIE: "SoRE"},
        "filing_standard": "mpers",
    }
    results = run_all(
        [SOFPBalanceCheck()],
        workbook_paths={StatementType.SOFP: sofp_path},
        run_config=run_config,
    )
    assert results[0].status == "passed"


# --- Step 4.3: new SoRE → SOFP retained earnings check ---

@pytest.mark.mpers_wiring_crosschecks
def test_sore_to_sofp_retained_earnings_check_exists():
    """Module + class + metadata shape — catches registration regressions."""
    from cross_checks.sore_to_sofp_retained_earnings import (
        SoREToSOFPRetainedEarningsCheck,
    )
    c = SoREToSOFPRetainedEarningsCheck()
    assert c.name == "sore_to_sofp_retained_earnings"
    assert c.required_statements == {StatementType.SOCIE, StatementType.SOFP}
    assert c.applies_to_standard == frozenset({"mpers"})


@pytest.mark.mpers_wiring_crosschecks
def test_sore_to_sofp_retained_earnings_passes_when_matching(tmp_dir):
    from cross_checks.sore_to_sofp_retained_earnings import (
        SoREToSOFPRetainedEarningsCheck,
    )

    sore_path = os.path.join(tmp_dir, "sore.xlsx")
    _make_workbook({
        "SoRE": [
            *([["filler"]] * 20),
            ["Retained earnings at end of period", 1_234_567.0, 1_000_000.0],
        ],
    }, sore_path)
    sofp_path = os.path.join(tmp_dir, "sofp.xlsx")
    _make_workbook({
        "SOFP-CuNonCu": [
            ["*Total assets", 2_000_000, 1_500_000],
            ["*Total equity and liabilities", 2_000_000, 1_500_000],
            ["Retained earnings", 1_234_567.0, 1_000_000.0],
        ],
    }, sofp_path)

    res = SoREToSOFPRetainedEarningsCheck().run(
        {StatementType.SOCIE: sore_path, StatementType.SOFP: sofp_path},
        tolerance=1.0,
    )
    assert res.status == "passed"
    assert res.diff == pytest.approx(0.0, abs=0.01)


@pytest.mark.mpers_wiring_crosschecks
def test_sore_to_sofp_retained_earnings_fails_on_mismatch(tmp_dir):
    from cross_checks.sore_to_sofp_retained_earnings import (
        SoREToSOFPRetainedEarningsCheck,
    )

    sore_path = os.path.join(tmp_dir, "sore.xlsx")
    _make_workbook({
        "SoRE": [
            ["Retained earnings at end of period", 1_234_567.0, 1_000_000.0],
        ],
    }, sore_path)
    sofp_path = os.path.join(tmp_dir, "sofp.xlsx")
    _make_workbook({
        "SOFP-CuNonCu": [
            ["Retained earnings", 1_000_000.0, 900_000.0],
        ],
    }, sofp_path)

    res = SoREToSOFPRetainedEarningsCheck().run(
        {StatementType.SOCIE: sore_path, StatementType.SOFP: sofp_path},
        tolerance=1.0,
    )
    assert res.status == "failed"
    assert res.diff == pytest.approx(234_567.0, abs=0.01)


@pytest.mark.mpers_wiring_crosschecks
def test_sore_to_sofp_retained_earnings_group_reads_both_columns(tmp_dir):
    """Group filings must reconcile Group CY (col B) and Company CY (col D)."""
    from cross_checks.sore_to_sofp_retained_earnings import (
        SoREToSOFPRetainedEarningsCheck,
    )

    sore_path = os.path.join(tmp_dir, "sore.xlsx")
    # Group CY=col B, Group PY=col C, Company CY=col D, Company PY=col E
    _make_workbook({
        "SoRE": [
            ["Retained earnings at end of period", 1_234_567.0, 1_000_000.0, 100_000.0, 80_000.0],
        ],
    }, sore_path)
    sofp_path = os.path.join(tmp_dir, "sofp.xlsx")
    # Group RE matches SoRE; Company RE is off by more than tolerance.
    _make_workbook({
        "SOFP-CuNonCu": [
            ["Retained earnings", 1_234_567.0, 1_000_000.0, 999_999.0, 80_000.0],
        ],
    }, sofp_path)

    res = SoREToSOFPRetainedEarningsCheck().run(
        {StatementType.SOCIE: sore_path, StatementType.SOFP: sofp_path},
        tolerance=1.0,
        filing_level="group",
    )
    assert res.status == "failed"
    assert "Company" in res.message


@pytest.mark.mpers_wiring_crosschecks
def test_server_registers_sore_cross_check():
    """server._build_default_cross_checks() must include the SoRE check so
    MPERS+SoRE runs get a real reconciliation after the three SOCIE-
    consuming checks are gated out."""
    from cross_checks.sore_to_sofp_retained_earnings import (
        SoREToSOFPRetainedEarningsCheck,
    )
    from server import _build_default_cross_checks

    checks = _build_default_cross_checks()
    assert any(isinstance(c, SoREToSOFPRetainedEarningsCheck) for c in checks)
