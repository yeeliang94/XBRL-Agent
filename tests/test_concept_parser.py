"""Phase 0 — Template Parser tests (canonical concept model).

Each test follows the RED→GREEN methodology from docs/PLAN-canonical-concept-model.md.  The five
fixtures pinned in step 0.1 cover the full schema-shape envelope before
SOCIE matrix support arrives in Phase 5:

  1. MFRS Company SOFP-CuNonCu  — face/sub split, base 4-col template
  2. MFRS Group   SOFP-CuNonCu  — 6-col group columns
  3. MFRS Company SOPL-Function — Analysis sub-sheet, abstract-row guard
  4. MPERS Company SOFP-CuNonCu — MPERS label/suffix convention
  5. MPERS Group   SOCIE        — matrix outlier, kept as XFAIL until P5

The parser deliberately delegates abstract-row detection to
``tools.section_headers`` so the writer-side guard (gotcha #17) and the
reader-side parser stay symmetric.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from concept_model import parser as concept_parser


REPO_ROOT = Path(__file__).resolve().parent.parent
MFRS_COMPANY = REPO_ROOT / "XBRL-template-MFRS" / "Company"
MFRS_GROUP = REPO_ROOT / "XBRL-template-MFRS" / "Group"
MPERS_COMPANY = REPO_ROOT / "XBRL-template-MPERS" / "Company"
MPERS_GROUP = REPO_ROOT / "XBRL-template-MPERS" / "Group"


# Pinned fixtures.  Keep this dict tight — adding more fixtures inflates
# every test in the module without adding signal.
FIXTURES: dict[str, Path] = {
    "mfrs_company_sofp_cunoncu":  MFRS_COMPANY / "01-SOFP-CuNonCu.xlsx",
    "mfrs_group_sofp_cunoncu":    MFRS_GROUP   / "01-SOFP-CuNonCu.xlsx",
    "mfrs_company_sopl_function": MFRS_COMPANY / "03-SOPL-Function.xlsx",
    "mpers_company_sofp_cunoncu": MPERS_COMPANY / "01-SOFP-CuNonCu.xlsx",
    "mpers_group_socie":          MPERS_GROUP  / "09-SOCIE.xlsx",
}


def _existing_fixtures() -> list[str]:
    """Skip names whose template files are missing on this machine.

    The MPERS templates are generated; if a contributor pulls without
    running the generator the parser tests should still meaningfully
    cover the MFRS half.
    """
    return [name for name, p in FIXTURES.items() if p.is_file()]


# -- module-scoped parsed-tree fixtures -------------------------------


@pytest.fixture(scope="module")
def parsed_trees() -> dict[str, "concept_parser.ConceptTree"]:
    """Parse each available fixture once for the whole module.

    SOCIE is allowed to raise ``UnsupportedSchemaShape`` (matrix layout
    is Phase 5 territory).  Any other exception fails the suite.
    """
    out: dict[str, concept_parser.ConceptTree] = {}
    for name in _existing_fixtures():
        path = FIXTURES[name]
        try:
            out[name] = concept_parser.parse_template(str(path))
        except concept_parser.UnsupportedSchemaShape:
            # Deferred to Phase 5.  Surfaced as a marker, not a failure.
            out[name] = None  # type: ignore[assignment]
    return out


# -- Step 0.2: abstract-row classification -----------------------------


@pytest.mark.parametrize("fixture_name", _existing_fixtures())
def test_abstract_rows_match_existing_section_headers_detection(
    fixture_name: str, parsed_trees
) -> None:
    """Parser's ABSTRACT row set must equal section_headers' detection
    on the same workbook.  Delegating keeps the writer guard (gotcha
    #17) and the parser in lockstep — without this, an agent could
    write to a row that the parser thought was a leaf."""
    import openpyxl
    from tools.section_headers import (
        discover_section_headers,
        keyword_fallback_for_sheet,
    )

    tree = parsed_trees[fixture_name]
    if tree is None or tree.shape == "matrix":
        # Matrix (SOCIE) templates only canonicalise the first block, so
        # whole-sheet header equivalence doesn't apply — the matrix branch
        # has its own coverage in tests/test_socie_parser_matrix.py.
        pytest.skip("matrix-shape template; covered by SOCIE matrix tests")

    wb = openpyxl.load_workbook(str(FIXTURES[fixture_name]), data_only=False)
    expected: set[tuple[str, int]] = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fallback = keyword_fallback_for_sheet(sheet_name)
        for h in discover_section_headers(ws, extra_keywords=fallback):
            expected.add((sheet_name, h.row))

    actual: set[tuple[str, int]] = {
        (n.render_key["sheet"], n.render_key["row"])
        for n in tree.concepts
        if n.kind == "ABSTRACT"
    }

    assert actual == expected, (
        f"abstract-row mismatch on {fixture_name}: "
        f"only_in_parser={actual - expected} "
        f"only_in_section_headers={expected - actual}"
    )


# -- helper ------------------------------------------------------------


def _by_render(tree, sheet: str, row: int):
    """Find the first concept whose render_key points at (sheet, row)."""
    for n in tree.concepts:
        if n.render_key["sheet"] == sheet and n.render_key["row"] == row:
            return n
    return None


# -- Step 0.3: LEAF vs COMPUTED distinction ----------------------------


def test_computed_rows_have_formulas_leaves_are_empty(parsed_trees) -> None:
    tree = parsed_trees["mfrs_company_sofp_cunoncu"]
    # Three known LEAF rows on the face sheet — col-A label present,
    # no formula in any value column.
    for row in (10, 17, 41):
        node = _by_render(tree, "SOFP-CuNonCu", row)
        assert node is not None, f"row {row} missing"
        assert node.kind == "LEAF", f"row {row} expected LEAF, got {node.kind}"

    # Three known COMPUTED rows — totals with signed-sum formulas.
    for row in (23, 37, 60):
        node = _by_render(tree, "SOFP-CuNonCu", row)
        assert node is not None, f"row {row} missing"
        assert node.kind == "COMPUTED", (
            f"row {row} expected COMPUTED, got {node.kind}"
        )


# -- Step 0.4: signed-sum grammar --------------------------------------


def test_signed_sum_formula_yields_dependency_edges(parsed_trees) -> None:
    """SOFP face row 44 is ``=1*B40+1*B41+-1*B42+1*B43`` — the +- in
    front of B42 must produce a -1 coefficient."""
    tree = parsed_trees["mfrs_company_sofp_cunoncu"]
    node = _by_render(tree, "SOFP-CuNonCu", 44)
    assert node is not None
    assert node.kind == "COMPUTED"

    coeff_by_row = {
        e["ref"]["row"]: e["coefficient"] for e in node.edges
    }
    assert coeff_by_row[40] == 1
    assert coeff_by_row[41] == 1
    assert coeff_by_row[42] == -1
    assert coeff_by_row[43] == 1


# -- Step 0.5: cross-sheet ref -----------------------------------------


def test_cross_sheet_ref_becomes_render_edge_not_concept(parsed_trees) -> None:
    """Face row 8 is ``='SOFP-Sub-CuNonCu'!B39``: the face cell must
    inherit the sub concept's UUID rather than mint a new one."""
    tree = parsed_trees["mfrs_company_sofp_cunoncu"]
    face = _by_render(tree, "SOFP-CuNonCu", 8)
    sub = _by_render(tree, "SOFP-Sub-CuNonCu", 39)
    assert face is not None and sub is not None
    assert face.concept_uuid == sub.concept_uuid, (
        "face row should reuse sub row's identity for the same concept"
    )
    # The face still carries a render edge so the exporter can follow
    # the link when writing back to the face sheet.
    assert face.edges, "cross-sheet face row should carry an edge"


# -- Step 0.6: SUM range -----------------------------------------------


def test_sum_range_formula_expands_to_edges() -> None:
    """We construct a synthetic worksheet to pin the SUM expansion
    semantics — easier than locating the right row inside a real
    template and equally precise."""
    import openpyxl

    from concept_model.parser import _collect_edges, ConceptNode

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUM-Synthetic"
    ws["A1"] = "Total"
    ws["B1"] = "=SUM(B10:B14)"

    node = ConceptNode(
        concept_uuid="u-1",
        parent_uuid=None,
        kind="COMPUTED",
        canonical_label="Total",
        render_key={"sheet": "SUM-Synthetic", "row": 1, "col": "B"},
    )
    _collect_edges(node, "SUM-Synthetic", "=SUM(B10:B14)", [])

    rows = sorted(e["ref"]["row"] for e in node.edges)
    assert rows == [10, 11, 12, 13, 14]
    assert all(e["coefficient"] == 1 for e in node.edges)


# -- Step 0.7: parent / child tree -------------------------------------


def test_concept_tree_parent_links_match_indentation(parsed_trees) -> None:
    """Walk the SOFP face sheet's tree and verify that LEAF rows under
    "Non-current assets" carry that abstract row's UUID as their
    parent."""
    tree = parsed_trees["mfrs_company_sofp_cunoncu"]
    non_current = _by_render(tree, "SOFP-CuNonCu", 7)  # 'Non-current assets'
    assert non_current is not None and non_current.kind == "ABSTRACT"

    # row 10 'Biological assets' is a LEAF sitting under non-current
    # assets in the template; parent must resolve to row 7.
    biological = _by_render(tree, "SOFP-CuNonCu", 10)
    assert biological is not None
    assert biological.parent_uuid == non_current.concept_uuid


# -- Step 0.8: deterministic JSON serialization ------------------------


def test_serialization_is_deterministic_for_same_template() -> None:
    """Parse twice; assert byte-identical JSON.  Anchors UUID5 stability
    so Phase 1's importer can rely on the same template always
    producing the same identities."""
    import json

    from concept_model import parser as cp

    path = str(FIXTURES["mfrs_company_sofp_cunoncu"])
    first = json.dumps(cp.parse_template(path).to_json(),
                       sort_keys=True, indent=2)
    second = json.dumps(cp.parse_template(path).to_json(),
                        sort_keys=True, indent=2)
    assert first == second


# -- Step 0.9: cross-template smoke ------------------------------------


def test_parses_every_live_template_without_error() -> None:
    """Walk both template roots and parse every .xlsx except backups.

    SOCIE is allowed to raise ``UnsupportedSchemaShape`` (matrix layout
    deferred to Phase 5); anything else must succeed."""
    from concept_model import parser as cp

    errors: list[tuple[Path, Exception]] = []
    parsed = 0
    skipped_matrix = 0
    for root in (REPO_ROOT / "XBRL-template-MFRS",
                 REPO_ROOT / "XBRL-template-MPERS"):
        if not root.is_dir():
            continue
        for xlsx in root.rglob("*.xlsx"):
            if any(p.startswith("backup") for p in xlsx.parts):
                continue
            try:
                cp.parse_template(str(xlsx))
                parsed += 1
            except cp.UnsupportedSchemaShape:
                skipped_matrix += 1
            except Exception as exc:
                errors.append((xlsx, exc))

    assert not errors, (
        "parser failed on live templates:\n"
        + "\n".join(f"  {p}: {e!r}" for p, e in errors)
    )
    assert parsed > 0, "no templates were parsed"
