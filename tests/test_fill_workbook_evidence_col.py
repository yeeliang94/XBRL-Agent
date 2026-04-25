"""Pin SOCIE / SoRE evidence-column placement against live template layouts.

`tools/fill_workbook.py` historically wrote SOCIE evidence to col Y (25)
on the assumption that every SOCIE template was the MFRS 24-col matrix.
That worked for MFRS but silently misrouted evidence on:

  - MPERS Company SOCIE  (max_col=4, Source col D)
  - MPERS Group SOCIE    (max_col=4, Source col D — added in Phase 3
                          of PLAN-mpers-group-socie-formulas.md)
  - MPERS Company SoRE   (max_col=4, Source col D)
  - MPERS Group SoRE     (max_col=6, Source col F)

Evidence landed off-screen and the visible "Source" header column stayed
empty. These tests pin the resolved placement: evidence lands at the
template's declared Source header when it has one, falling back to col Y
for the matrix layouts that don't.
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl
import pytest

from tools.fill_workbook import fill_workbook


_REPO_ROOT = Path(__file__).resolve().parent.parent


def _copy_template(src: Path, dst: Path) -> Path:
    shutil.copy2(src, dst)
    return dst


# (template_relpath, sheet_name, filing_level, label, expected_evidence_col)
#
# Evidence column expectations:
#   col D (4)  — MPERS Company/Group SOCIE + MPERS Company SoRE (Source col D)
#   col F (6)  — MPERS Group SoRE (Source col F per 6-col Group layout)
#   col Y (25) — MFRS Company/Group SOCIE (24-col matrix, no Source header
#                — evidence lives just past Total at col 25 by convention)
EVIDENCE_CASES = [
    (
        "XBRL-template-MPERS/Company/09-SOCIE.xlsx",
        "SOCIE",
        "company",
        "Profit (loss)",
        4,
    ),
    (
        "XBRL-template-MPERS/Group/09-SOCIE.xlsx",
        "SOCIE",
        "group",
        "Profit (loss)",
        4,
    ),
    (
        "XBRL-template-MPERS/Company/10-SoRE.xlsx",
        "SoRE",
        "company",
        "Profit (loss)",
        4,
    ),
    (
        "XBRL-template-MPERS/Group/10-SoRE.xlsx",
        "SoRE",
        "group",
        "Profit (loss)",
        6,
    ),
    (
        "XBRL-template-MFRS/Company/09-SOCIE.xlsx",
        "SOCIE",
        "company",
        "Profit (loss)",
        25,
    ),
    (
        "XBRL-template-MFRS/Group/09-SOCIE.xlsx",
        "SOCIE",
        "group",
        "Profit (loss)",
        25,
    ),
]


@pytest.mark.parametrize(
    "rel_path,sheet,filing_level,label,expected_col", EVIDENCE_CASES,
)
def test_socie_evidence_lands_in_source_column(
    tmp_path, rel_path, sheet, filing_level, label, expected_col,
):
    """Evidence column must match the template's declared Source header.

    Parametrising per-template makes a regression message tell you
    exactly which layout broke rather than reporting one big diff.
    """
    src = _REPO_ROOT / rel_path
    assert src.exists(), f"missing template: {rel_path}"
    dst = _copy_template(src, tmp_path / src.name)

    payload = {
        "fields": [
            {
                "sheet": sheet,
                "field_label": label,
                "col": 2,
                "value": 12345,
                "evidence": "Page 14 statement of changes in equity",
            }
        ]
    }
    result = fill_workbook(
        template_path=str(dst),
        output_path=str(dst),
        fields_json=json.dumps(payload),
        filing_level=filing_level,
    )
    assert result.success, f"fill_workbook failed: {result.errors}"

    wb = openpyxl.load_workbook(dst, data_only=False)
    try:
        ws = wb[sheet]
        # Locate the row that received the value so we can read the same
        # row's evidence column. We do this by scanning col B for the
        # value — the writer's label resolution already validated the
        # field landed somewhere.
        value_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 2).value == 12345:
                value_row = r
                break
        assert value_row is not None, (
            f"{rel_path}: value 12345 was not written to col B"
        )
        evidence = ws.cell(value_row, expected_col).value
        assert evidence == "Page 14 statement of changes in equity", (
            f"{rel_path}: expected evidence at col {expected_col} of row "
            f"{value_row}; got cell value {evidence!r}. Other-column scan: "
            + ", ".join(
                f"col{c}={ws.cell(value_row, c).value!r}"
                for c in range(1, ws.max_column + 1)
            )
        )
    finally:
        wb.close()
