"""N1 — notes↔face numeric tie-out checks.

Builds synthetic workbooks pairing a face SOFP sheet with the issued-capital
note and asserts: matched within tolerance → no warning; a 1000×-off note →
WARN; unmapped topics stay silent (no fabricated pairings); a missing sheet is
skipped.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from cross_checks.notes_face_tieouts import (
    _PAIRS,
    check_notes_face_tieouts,
    TieoutWarning,
)
from cross_checks.util import find_label_row, find_sheet


def _build(path: Path, face_share_capital, note_issued_capital,
           *, face_sheet="SOFP-CuNonCu", note_sheet="Notes-Issuedcapital"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    f = wb.create_sheet(face_sheet)
    f.cell(1, 1, "Total assets")
    f.cell(2, 1, "Share capital")
    if face_share_capital is not None:
        f.cell(2, 2, face_share_capital)
    n = wb.create_sheet(note_sheet)
    n.cell(3, 1, "Notes - Issued capital")
    # Mirror the live sheet shape: "Issued capital" is an abstract section
    # header whose value cell is always empty (gotcha #17); the closing
    # amount lives at the "Balance at the end of period" leaf.
    n.cell(5, 1, "Issued capital")
    n.cell(18, 1, "Balance at the end of period")
    if note_issued_capital is not None:
        n.cell(18, 2, note_issued_capital)
    wb.save(str(path))


def test_matched_pair_within_tolerance_no_warning(tmp_path):
    p = tmp_path / "wb.xlsx"
    _build(p, 5_000_000, 5_000_000)
    assert check_notes_face_tieouts(str(p)) == []


def test_1000x_off_note_warns(tmp_path):
    p = tmp_path / "wb.xlsx"
    # Note recorded in units while the face is in thousands → 1000× gap.
    _build(p, 5_000, 5_000_000)
    warns = check_notes_face_tieouts(str(p))
    assert len(warns) == 1
    assert isinstance(warns[0], TieoutWarning)
    assert warns[0].topic == "Share capital"
    assert "reconcile" in warns[0].message.lower()


def test_blank_note_side_is_silent(tmp_path):
    p = tmp_path / "wb.xlsx"
    _build(p, 5_000_000, None)   # note blank → nothing to reconcile
    assert check_notes_face_tieouts(str(p)) == []


def test_missing_note_sheet_is_silent(tmp_path):
    # Only the face sheet exists (run didn't produce the issued-capital note).
    p = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    f = wb.create_sheet("SOFP-CuNonCu")
    f.cell(2, 1, "Share capital")
    f.cell(2, 2, 5_000_000)
    wb.save(str(p))
    assert check_notes_face_tieouts(str(p)) == []


def test_rounding_within_tolerance_no_warning(tmp_path):
    p = tmp_path / "wb.xlsx"
    # Off by 1 on a large magnitude — absorbed by the scaled tolerance.
    _build(p, 5_000_000, 5_000_001)
    assert check_notes_face_tieouts(str(p)) == []


def test_group_filing_checks_company_cy_column(tmp_path):
    # Group filing: Group CY (col B) reconciles, but Company CY (col D) is
    # 1000× off. The check must flag the Company column, not only Group.
    p = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    f = wb.create_sheet("SOFP-CuNonCu")
    f.cell(2, 1, "Share capital")
    f.cell(2, 2, 5_000_000)   # Group CY
    f.cell(2, 4, 3_000_000)   # Company CY
    n = wb.create_sheet("Notes-Issuedcapital")
    n.cell(5, 1, "Balance at the end of period")
    n.cell(5, 2, 5_000_000)   # Group CY — matches
    n.cell(5, 4, 3_000)       # Company CY — 1000× off
    wb.save(str(p))

    warns = check_notes_face_tieouts(str(p), filing_level="group")
    assert len(warns) == 1
    assert warns[0].message.startswith("Company ")


def test_group_filing_both_columns_match_no_warning(tmp_path):
    p = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    f = wb.create_sheet("SOFP-CuNonCu")
    f.cell(2, 1, "Share capital")
    f.cell(2, 2, 5_000_000)
    f.cell(2, 4, 3_000_000)
    n = wb.create_sheet("Notes-Issuedcapital")
    n.cell(5, 1, "Balance at the end of period")
    n.cell(5, 2, 5_000_000)
    n.cell(5, 4, 3_000_000)
    wb.save(str(p))
    assert check_notes_face_tieouts(str(p), filing_level="group") == []


def test_missing_workbook_returns_empty(tmp_path):
    assert check_notes_face_tieouts(str(tmp_path / "nope.xlsx")) == []


# ---------------------------------------------------------------------------
# Live-template pinning (house style: test_notes_prompt_phase1.py
# ::test_live_templates_subtract_dividends_paid). Every seed pair's notes
# label must resolve to a NON-ABSTRACT leaf row on the live templates.
# The pre-fix label ("issued capital") resolved to the 1F3864 section-header
# row (gotcha #17) whose value cell is always empty — so notes_val was None
# on every real run and the tie-out silently no-op'd.
# ---------------------------------------------------------------------------

_REPO_ROOT = Path(__file__).resolve().parent.parent

try:
    from tools.section_headers import _HEADER_FILL_RGB
except ImportError:  # pragma: no cover — helper moved; pin the fills directly
    _HEADER_FILL_RGB = frozenset({"001F3864", "FF1F3864", "00305496", "FF305496"})

_LIVE_ISSUED_CAPITAL_TEMPLATES = [
    ("mfrs", _REPO_ROOT / "XBRL-template-MFRS" / "Company" / "13-Notes-IssuedCapital.xlsx"),
    ("mfrs", _REPO_ROOT / "XBRL-template-MFRS" / "Group" / "13-Notes-IssuedCapital.xlsx"),
    ("mpers", _REPO_ROOT / "XBRL-template-MPERS" / "Company" / "14-Notes-IssuedCapital.xlsx"),
    ("mpers", _REPO_ROOT / "XBRL-template-MPERS" / "Group" / "14-Notes-IssuedCapital.xlsx"),
]


@pytest.mark.parametrize(
    "standard,template_path",
    _LIVE_ISSUED_CAPITAL_TEMPLATES,
    ids=[f"{s}-{p.parent.name}" for s, p in _LIVE_ISSUED_CAPITAL_TEMPLATES],
)
def test_live_templates_seed_pair_resolves_to_non_abstract_leaf(standard, template_path):
    if not template_path.exists():
        pytest.skip(f"live template not present: {template_path}")
    wb = openpyxl.load_workbook(str(template_path))
    try:
        for pair in _PAIRS:
            if pair.applies_to_standard not in ("all", standard):
                continue
            ws = find_sheet(wb, *pair.notes_sheets)
            assert ws is not None, (
                f"{template_path.name}: none of {pair.notes_sheets} found"
            )
            row = find_label_row(ws, list(pair.notes_labels))
            assert row is not None, (
                f"{template_path.name}: notes label {pair.notes_labels} "
                f"resolves to no row — the tie-out would silently no-op"
            )
            cell = ws.cell(row=row, column=1)
            rgb = None
            if cell.fill is not None and cell.fill.start_color is not None:
                rgb = cell.fill.start_color.rgb
            assert rgb not in _HEADER_FILL_RGB, (
                f"{template_path.name} row {row} ({cell.value!r}): notes label "
                f"{pair.notes_labels} resolves to an ABSTRACT section-header "
                f"row (fill {rgb}) — its value cell is never written, so the "
                f"tie-out check would silently no-op (gotcha #17)"
            )
    finally:
        wb.close()
