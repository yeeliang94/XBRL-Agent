"""Tokenizer-based formula evaluator regression tests.

Peer-review finding C3: the pre-existing regex evaluator in
`tools/verifier.py`:
  * only matched a cross-sheet reference when it was the ENTIRE formula body,
    so `='Sub'!B39+'Sub'!B40` fell through and was evaluated as same-sheet
    refs against the current sheet — silently wrong numbers;
  * fell back to summing every `[A-Z]+\\d+` match with weight +1 when the
    term regex didn't consume the body, so formulas containing `IFERROR(...)`
    or `#REF!` returned plausible-looking garbage instead of failing loudly.

The replacement must:
  1. correctly evaluate multi-term cross-sheet formulas;
  2. correctly evaluate SUM(range) across sheets and within sheet;
  3. return a sentinel (0.0 with a warning) when an unsupported construct
     appears — never guess.
"""
from __future__ import annotations

import os
import tempfile

import openpyxl
import pytest

from tools.verifier import _evaluate_formula


@pytest.fixture
def wb():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "t.xlsx")
        book = openpyxl.Workbook()
        main = book.active
        main.title = "Main"
        main["B139"] = 10
        main["B140"] = 20
        main["B141"] = 5
        main["E6"] = 1
        main["F6"] = 2
        main["G6"] = 3
        main["H6"] = 4
        main["I6"] = 5
        main["J6"] = 6
        main["K6"] = 7
        main["L6"] = 8

        sub = book.create_sheet("Sub")
        sub["B39"] = 100
        sub["B40"] = 200

        book.save(path)
        book.close()

        book2 = openpyxl.load_workbook(path)
        try:
            yield book2
        finally:
            book2.close()


def test_multi_term_cross_sheet_formula(wb):
    """The pre-fix regex returned 0 for this formula (matched the first
    ref incorrectly, then fell through). Correct value is 100+200 = 300."""
    warnings: list[str] = []
    result = _evaluate_formula(wb, "Main", "='Sub'!B39+'Sub'!B40", warnings=warnings)
    assert result == 300.0
    assert warnings == []


def test_weighted_sum_formula(wb):
    """=1*B139+1*B140-1*B141 on Main sheet = 10+20-5 = 25."""
    warnings: list[str] = []
    result = _evaluate_formula(wb, "Main", "=1*B139+1*B140-1*B141", warnings=warnings)
    assert result == 25.0
    assert warnings == []


def test_sum_range_formula(wb):
    """=SUM(E6:L6) = 1+2+3+4+5+6+7+8 = 36."""
    warnings: list[str] = []
    result = _evaluate_formula(wb, "Main", "=SUM(E6:L6)", warnings=warnings)
    assert result == 36.0
    assert warnings == []


def test_unsupported_function_returns_sentinel_with_warning(wb):
    """=IFERROR(A1,0) is not supported — pre-fix code summed `A1` as +1,
    returning a wrong-but-plausible number. Must now warn + return 0."""
    warnings: list[str] = []
    result = _evaluate_formula(wb, "Main", "=IFERROR(B139,0)", warnings=warnings)
    assert result == 0.0
    assert warnings, "expected a formula warning for unsupported IFERROR"
    assert "IFERROR" in "; ".join(warnings).upper() or "unsupported" in "; ".join(warnings).lower()


def test_unparseable_formula_emits_warning(wb):
    """#REF! or a malformed token sequence must not be summed as +1."""
    warnings: list[str] = []
    result = _evaluate_formula(wb, "Main", "=#REF!+B139", warnings=warnings)
    # Sentinel behaviour — must not silently return 10 (B139) as if valid.
    assert result == 0.0
    assert warnings, "expected a formula warning for #REF!"


def test_mixed_same_sheet_and_cross_sheet_refs(wb):
    """Real templates mix: ='Sub'!B39+B140 = 100 + 20 = 120."""
    warnings: list[str] = []
    result = _evaluate_formula(wb, "Main", "='Sub'!B39+B140", warnings=warnings)
    assert result == 120.0
    assert warnings == []


def test_empty_or_nonformula_returns_zero(wb):
    """Defensive: non-formula input returns 0 without crashing."""
    assert _evaluate_formula(wb, "Main", "") == 0.0
    assert _evaluate_formula(wb, "Main", "123") == 0.0


def test_sum_with_whitespace_between_args(wb):
    """Peer-review MEDIUM: Excel accepts spaces after commas in SUM, e.g.
    `=SUM(A1, B2)`. The Tokenizer emits WHITE-SPACE tokens that the arg
    walker previously rejected as 'not a plain range', returning 0."""
    warnings: list[str] = []
    # B139=10, B140=20, B141=5 from the fixture; expect 10+20+5 = 35
    result = _evaluate_formula(
        wb, "Main", "=SUM(B139, B140, B141)", warnings=warnings
    )
    assert result == 35.0
    assert warnings == []


def test_signed_sum_with_whitespace_around_operators(wb):
    """Formulas like `=B139 + B140` (spaces around operator) must also work."""
    warnings: list[str] = []
    result = _evaluate_formula(wb, "Main", "=B139 + B140 - B141", warnings=warnings)
    assert result == 25.0  # 10 + 20 - 5
    assert warnings == []


def test_sum_with_whitespace_and_range(wb):
    """Mixed case: SUM with a range and a comma-separated cell, both spaced."""
    warnings: list[str] = []
    result = _evaluate_formula(
        wb, "Main", "=SUM(E6:L6, B139)", warnings=warnings
    )
    # E6..L6 = 1+2+3+4+5+6+7+8 = 36; + B139 (10) = 46
    assert result == 46.0
    assert warnings == []
