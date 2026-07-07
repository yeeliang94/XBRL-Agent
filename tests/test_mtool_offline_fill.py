"""Tests for the offline mTool filler spike (mtool/offline_fill.py).

The tool itself is stdlib-only by contract (it travels to the Windows box as
a single file); openpyxl is used here only to build fixtures and to
independently re-read patched output.
"""
import json
import zipfile

import pytest
from openpyxl import Workbook, load_workbook

from mtool.offline_fill import (
    PrefixedSheetError,
    col_to_idx,
    format_value,
    get_sheet_paths,
    load_workbook_entries,
    main,
    normalize_label,
    patch_cell_in_sheet,
    set_full_calc_on_load,
    split_ref,
    validate_input,
    verify_values,
)

SHEET = "SOFP-Sub-CuNonCu"


def test_xml_escape_strips_illegal_chars_and_stays_valid_xml():
    # A PDF-extracted note payload carrying XML-illegal control chars (NUL,
    # vertical tab, form feed, unit separator) would make xl/sharedStrings.xml
    # unreadable → Excel's "String properties … repair or remove the unreadable
    # content" on open. _xml_escape must drop them while keeping legal
    # whitespace and escaping markup.
    import xml.etree.ElementTree as ET

    from mtool.offline_fill import _xml_escape

    raw = "Lease\x00 note\x0b<b>&</b>\x0c end\x1f\tOK\nline"
    esc = _xml_escape(raw)
    for bad in ("\x00", "\x0b", "\x0c", "\x1f"):
        assert bad not in esc, f"illegal {bad!r} survived"
    assert "\t" in esc and "\n" in esc            # legal whitespace kept
    assert "&amp;" in esc and "&lt;b&gt;" in esc  # markup escaped
    # The exact property that was failing: the text round-trips as valid XML.
    ET.fromstring(f'<t xml:space="preserve">{esc}</t>')


def test_offline_fill_imports_with_no_third_party_deps():
    """The tool travels to Windows as one stdlib-only file (CLAUDE.md #28).

    Import it in a subprocess with site-packages stripped from sys.path so a
    stray `import openpyxl` (or any pip dep) fails loudly here, not on the box.
    """
    import subprocess
    import sys
    from pathlib import Path

    tool = Path(__file__).resolve().parent.parent / "mtool" / "offline_fill.py"
    code = (
        "import sys; "
        "sys.path = [p for p in sys.path if 'site-packages' not in p]; "
        "import importlib.util; "
        f"spec = importlib.util.spec_from_file_location('offline_fill', {str(tool)!r}); "
        "m = importlib.util.module_from_spec(spec); spec.loader.exec_module(m); "
        "print('OK')"
    )
    result = subprocess.run(
        [sys.executable, "-c", code], capture_output=True, text=True,
    )
    assert result.returncode == 0, result.stderr
    assert "OK" in result.stdout


@pytest.fixture
def template(tmp_path):
    """A small workbook mimicking the mTool sub-sheet shape."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["A3"] = "Property, plant and equipment"
    ws["A4"] = "Freehold land"
    ws["B4"].number_format = "#,##0"  # styled empty -> self-closing <c/>
    ws["C4"].number_format = "#,##0"
    ws["A5"] = "Long term leasehold land"  # B5 entirely absent
    ws["A6"] = "Motor vehicles"
    ws["B6"] = 999  # existing numeric value
    ws["A7"] = "*Land"
    ws["B7"] = "=SUM(B4:B6)"  # formula cell — must never be written
    ws["A8"] = "Notes"
    ws["B8"] = "some text"  # shared-string cell
    ws["A9"] = "Duplicate row"
    ws["A10"] = "Duplicate row"
    wb.create_sheet("Other")["A1"] = "unrelated"
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return str(path)


def make_input(tmp_path, writes, columns=None):
    doc = {
        "sheets": {SHEET: {"label_column": "A",
                           "columns": columns or {"current_year": "B",
                                                  "prior_year": "C"}}},
        "writes": writes,
    }
    path = tmp_path / "fill.json"
    path.write_text(json.dumps(doc), encoding="utf-8")
    return str(path)


def run_fill(tmp_path, template, writes, extra_args=(), columns=None):
    out = tmp_path / "filled.xlsx"
    report_path = tmp_path / "report.json"
    code = main([
        "fill", "--workbook", template,
        "--input", make_input(tmp_path, writes, columns),
        "--output", str(out), "--report", str(report_path), *extra_args,
    ])
    report = json.loads(report_path.read_text(encoding="utf-8"))
    return code, report, str(out)


# ------------------------------------------------------------ happy paths

def test_replace_existing_numeric_value(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Motor vehicles",
         "column_role": "current_year", "value": 123},
    ])
    assert code == 0
    assert report["status"] == "ok"
    assert report["written"][0]["action"] == "replaced"
    assert load_workbook(out)[SHEET]["B6"].value == 123


def test_expand_styled_empty_cell(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Freehold land",
         "column_role": "current_year", "value": 1500000},
        {"sheet": SHEET, "label": "Freehold land",
         "column_role": "prior_year", "value": -200.5},
    ])
    assert code == 0
    # openpyxl writes styled empties as paired <c></c> (rebuilt); Excel writes
    # self-closing <c/> (expanded) — the raw-XML test below pins that shape.
    assert {w["action"] for w in report["written"]} <= {"expanded", "rebuilt"}
    ws = load_workbook(out)[SHEET]
    assert ws["B4"].value == 1500000
    assert ws["C4"].value == -200.5
    assert ws["B4"].number_format == "#,##0"  # style survived


def test_insert_missing_cell_in_existing_row(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Long term leasehold land",
         "column_role": "current_year", "value": 820000},
    ])
    assert code == 0
    assert report["written"][0]["action"] == "inserted_cell"
    assert load_workbook(out)[SHEET]["B5"].value == 820000


def test_insert_missing_row_via_cell_override(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "cell": "B100", "value": 42},
    ])
    assert code == 0
    assert report["written"][0]["action"] == "inserted_row"
    ws = load_workbook(out)[SHEET]
    assert ws["B100"].value == 42
    assert ws["B6"].value == 999  # neighbours untouched


def test_untouched_entries_are_byte_identical(tmp_path, template):
    code, _, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Motor vehicles",
         "column_role": "current_year", "value": 123},
    ])
    assert code == 0
    _, data, _ = load_workbook_entries(template)
    patched_entry = get_sheet_paths(data)[SHEET]
    with zipfile.ZipFile(template) as zin, zipfile.ZipFile(out) as zout:
        assert zin.namelist() == zout.namelist()  # order preserved
        for name in zin.namelist():
            if name == patched_entry:
                assert zin.read(name) != zout.read(name)
            else:
                assert zin.read(name) == zout.read(name), name


# ------------------------------------------------------------ guards

def test_formula_cell_is_refused(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "*Land",
         "column_role": "current_year", "value": 777},
    ])
    assert code == 1
    assert report["status"] == "degraded"
    assert report["skipped_formula"][0]["cell"] == "B7"
    assert not report["written"]
    assert load_workbook(out)[SHEET]["B7"].value == "=SUM(B4:B6)"


def test_text_cell_write_is_type_changed(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Notes",
         "column_role": "current_year", "value": 55},
    ])
    assert code == 1  # surfaced for operator review, still applied
    assert report["type_changed"][0]["cell"] == "B8"
    assert load_workbook(out)[SHEET]["B8"].value == 55


def test_duplicate_resolved_target_is_error(tmp_path, template):
    code, report, _ = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Motor vehicles",
         "column_role": "current_year", "value": 1},
        {"sheet": SHEET, "label": "Motor vehicles",
         "column_role": "current_year", "value": 2},
    ])
    assert code == 1
    assert len(report["written"]) == 1
    assert "duplicate write" in report["errors"][0]["error"]


def test_unknown_sheet_is_error(tmp_path, template):
    code, report, _ = run_fill(tmp_path, template, [
        {"sheet": "Nope", "cell": "B2", "value": 1},
    ])
    assert code == 1
    assert "not found" in report["errors"][0]["error"]


# ------------------------------------------------------------ resolution

def test_fuzzy_label_resolves_with_ratio(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Freehod land",  # typo
         "column_role": "current_year", "value": 10},
    ])
    assert code == 0
    entry = report["written"][0]
    assert entry["matched_label"] == "Freehold land"
    assert entry["ratio"] < 1.0
    assert report["fuzzy_matched"] == [entry]  # surfaced for operator review
    assert load_workbook(out)[SHEET]["B4"].value == 10


def test_strict_mode_refuses_fuzzy_match(tmp_path, template):
    out = tmp_path / "filled.xlsx"
    code = main(["fill", "--workbook", template,
                 "--input", make_input(tmp_path, [
                     {"sheet": SHEET, "label": "Freehod land",  # typo
                      "column_role": "current_year", "value": 10}]),
                 "--output", str(out), "--strict",
                 "--report", str(tmp_path / "r.json")])
    report = json.loads((tmp_path / "r.json").read_text())
    assert code == 1
    assert report["strict"] is True
    assert not report["written"]
    assert "strict mode" in report["unresolved"][0]["detail"]


def test_strict_mode_still_allows_exact_match(tmp_path, template):
    _, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "freehold land",  # exact after normalise
         "column_role": "current_year", "value": 10},
    ], extra_args=["--strict"])
    assert report["written"] and report["fuzzy_matched"] == []
    assert load_workbook(out)[SHEET]["B4"].value == 10


def test_doc_level_strict_flag_is_honoured(tmp_path, template):
    out = tmp_path / "filled.xlsx"
    doc = {"strict": True,
           "sheets": {SHEET: {"label_column": "A",
                              "columns": {"current_year": "B"}}},
           "writes": [{"sheet": SHEET, "label": "Freehod land",
                       "column_role": "current_year", "value": 10}]}
    inp = tmp_path / "fill.json"
    inp.write_text(json.dumps(doc), encoding="utf-8")
    code = main(["fill", "--workbook", template, "--input", str(inp),
                 "--output", str(out), "--report", str(tmp_path / "r.json")])
    report = json.loads((tmp_path / "r.json").read_text())
    assert code == 1  # no --strict flag, but doc says strict
    assert report["strict"] is True
    assert not report["written"]


def test_exact_match_is_not_flagged_fuzzy(tmp_path, template):
    _, report, _ = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "freehold land",  # case-only difference
         "column_role": "current_year", "value": 10},
    ])
    assert report["written"] and report["fuzzy_matched"] == []


def test_unresolvable_label_is_reported_not_guessed(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Completely different thing",
         "column_role": "current_year", "value": 10},
    ])
    assert code == 1
    assert report["unresolved"][0]["label"] == "Completely different thing"
    assert not report["written"]


def test_duplicate_label_is_ambiguous(tmp_path, template):
    code, report, _ = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Duplicate row",
         "column_role": "current_year", "value": 10},
    ])
    assert code == 1
    assert "rows [9, 10]" in report["ambiguous"][0]["detail"]


def test_label_normalization():
    assert normalize_label("  Freehold   land : ") == "freehold land"
    assert normalize_label("FREEHOLD LAND") == "freehold land"


# ------------------------------------------------------------ input contract

def test_input_validation_rejects_bad_values():
    doc = {"sheets": {SHEET: {"columns": {"cy": "B"}}}, "writes": [
        {"sheet": SHEET, "label": "x", "column_role": "cy", "value": "(200)"},
        {"sheet": SHEET, "label": "x", "column_role": "cy", "value": True},
        {"sheet": SHEET, "label": "x", "cell": "B2", "value": 1},
        {"sheet": SHEET, "label": "x", "column_role": "missing", "value": 1},
        {"sheet": SHEET, "cell": "2B", "value": 1},
    ]}
    errors = validate_input(doc)
    assert len(errors) == 5
    assert any("'(200)'" in e for e in errors)
    assert any("exactly one of" in e for e in errors)
    assert any("not configured" in e for e in errors)


def test_bom_prefixed_input_file_is_accepted(tmp_path, template):
    """PowerShell writes JSON with a UTF-8 BOM; the loader must tolerate it."""
    out = tmp_path / "filled.xlsx"
    doc = {"sheets": {SHEET: {"label_column": "A",
                              "columns": {"current_year": "B"}}},
           "writes": [{"sheet": SHEET, "label": "Motor vehicles",
                       "column_role": "current_year", "value": 7}]}
    bom_input = tmp_path / "fill_bom.json"
    bom_input.write_bytes(b"\xef\xbb\xbf" + json.dumps(doc).encode("utf-8"))
    code = main(["fill", "--workbook", template, "--input", str(bom_input),
                 "--output", str(out)])
    assert code == 0
    assert load_workbook(out)[SHEET]["B6"].value == 7


def test_invalid_input_exits_2(tmp_path, template):
    bad = tmp_path / "bad.json"
    bad.write_text(json.dumps({"writes": [
        {"sheet": SHEET, "cell": "B2", "value": "nope"}]}), encoding="utf-8")
    code = main(["fill", "--workbook", template, "--input", str(bad),
                 "--output", str(tmp_path / "o.xlsx")])
    assert code == 2


def test_format_value():
    assert format_value(1000) == "1000"
    assert format_value(-125000) == "-125000"
    assert format_value(1500000.0) == "1500000"
    assert format_value(43975.5) == "43975.5"
    with pytest.raises(ValueError):
        format_value(True)
    with pytest.raises(ValueError):
        format_value(float("nan"))


def test_ref_helpers():
    assert split_ref("B12") == ("B", 12)
    assert col_to_idx("A") == 1
    assert col_to_idx("AA") == 27
    with pytest.raises(ValueError):
        split_ref("12B")


# ------------------------------------------------------------ verification

def test_verify_catches_missing_value(template):
    _, data, _ = load_workbook_entries(template)
    entry = get_sheet_paths(data)[SHEET]
    mismatches = verify_values(template, [(entry, "B4", "123")])
    assert mismatches == [{"entry": entry, "cell": "B4",
                           "expected": "123", "found": None}]


def test_dry_run_writes_nothing(tmp_path, template):
    out = tmp_path / "never.xlsx"
    code = main(["fill", "--workbook", template,
                 "--input", make_input(tmp_path, [
                     {"sheet": SHEET, "label": "Motor vehicles",
                      "column_role": "current_year", "value": 5}]),
                 "--output", str(out), "--dry-run"])
    assert code == 0
    assert not out.exists()


def test_force_recalc_sets_flag(tmp_path, template):
    code, report, out = run_fill(tmp_path, template, [
        {"sheet": SHEET, "label": "Motor vehicles",
         "column_role": "current_year", "value": 5},
    ], extra_args=["--force-recalc"])
    assert code == 0
    assert report["force_recalc"]["calcPr_found"] is True
    _, data, _ = load_workbook_entries(out)
    assert b'fullCalcOnLoad="1"' in data["xl/workbook.xml"]


# ------------------------------------------------------------ raw XML quirks
# Excel-authored shapes openpyxl won't reproduce.

WRAP = ('<worksheet xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main"><sheetData>%s</sheetData></worksheet>')


def test_patch_replaces_v_with_attribute_order_quirks():
    xml = WRAP % '<row r="2" spans="1:3"><c s="7" r="B2"><v>1</v></c></row>'
    out, action = patch_cell_in_sheet(xml, "B2", "42")
    assert action == "replaced"
    assert "<v>42</v>" in out


def test_patch_expands_self_closing_styled_empty_cell():
    xml = WRAP % '<row r="2"><c r="B2" s="5"/></row>'
    out, action = patch_cell_in_sheet(xml, "B2", "42")
    assert action == "expanded"
    assert '<c r="B2" s="5"><v>42</v></c>' in out


def test_patch_rebuilds_shared_string_cell_dropping_t():
    xml = WRAP % '<row r="2"><c r="B2" s="3" t="s"><v>17</v></c></row>'
    out, action = patch_cell_in_sheet(xml, "B2", "42")
    assert action == "type_changed"
    assert '<c r="B2" s="3"><v>42</v></c>' in out
    assert 't="s"' not in out


def test_patch_inserts_cell_in_column_order():
    xml = WRAP % '<row r="2"><c r="A2"><v>1</v></c><c r="E2"><v>5</v></c></row>'
    out, action = patch_cell_in_sheet(xml, "C2", "3")
    assert action == "inserted_cell"
    assert out.index('r="A2"') < out.index('r="C2"') < out.index('r="E2"')


def test_patch_inserts_row_in_row_order():
    xml = WRAP % '<row r="1"><c r="A1"><v>1</v></c></row><row r="9"><c r="A9"><v>9</v></c></row>'
    out, action = patch_cell_in_sheet(xml, "B5", "5")
    assert action == "inserted_row"
    assert out.index('<row r="1">') < out.index('<row r="5">') < out.index('<row r="9">')


def test_patch_expands_self_closing_row():
    xml = WRAP % '<row r="2" ht="15"/>'
    out, action = patch_cell_in_sheet(xml, "B2", "7")
    assert action == "inserted_cell"
    assert '<row r="2" ht="15"><c r="B2"><v>7</v></c></row>' in out


def test_patch_handles_empty_self_closing_sheetdata():
    xml = ('<worksheet xmlns="http://schemas.openxmlformats.org/'
           'spreadsheetml/2006/main"><sheetData/></worksheet>')
    out, action = patch_cell_in_sheet(xml, "B2", "7")
    assert action == "inserted_row"
    assert "<sheetData><row" in out


def test_patch_does_not_match_address_inside_formula():
    xml = WRAP % ('<row r="2"><c r="C2"><f>SUM(B9:B10)</f><v>0</v></c></row>'
                  '<row r="9"><c r="A9"/></row>')
    out, action = patch_cell_in_sheet(xml, "B9", "42")
    assert action == "inserted_cell"
    assert "SUM(B9:B10)" in out  # formula text untouched
    assert '<c r="B9"><v>42</v></c>' in out


def test_prefixed_sheet_xml_aborts():
    xml = ('<x:worksheet xmlns:x="http://schemas.openxmlformats.org/'
           'spreadsheetml/2006/main"><x:sheetData/></x:worksheet>')
    with pytest.raises(PrefixedSheetError):
        patch_cell_in_sheet(xml, "B2", "7")


def test_set_full_calc_on_load_variants():
    xml = '<workbook><calcPr calcId="1"/></workbook>'
    out, found = set_full_calc_on_load(xml)
    assert found and 'calcId="1" fullCalcOnLoad="1"/>' in out
    out2, found2 = set_full_calc_on_load(
        '<workbook><calcPr calcId="1" fullCalcOnLoad="0"/></workbook>')
    assert found2 and 'fullCalcOnLoad="1"' in out2
    out3, found3 = set_full_calc_on_load(out)  # idempotent
    assert found3 and out3 == out
    out4, found4 = set_full_calc_on_load("<workbook/>")
    assert not found4 and out4 == "<workbook/>"


# ------------------------------------------------------------ inspect

def test_inspect_lists_sheets(template, capsys):
    assert main(["inspect", "--workbook", template]) == 0
    out = capsys.readouterr().out
    assert SHEET in out and "Other" in out


def test_inspect_dumps_labels_and_cell_kinds(template, capsys):
    assert main(["inspect", "--workbook", template, "--sheet", SHEET]) == 0
    out = capsys.readouterr().out
    assert "Freehold land" in out
    assert "B:F" in out  # formula cell on the *Land row
    assert "B:N" in out  # numeric cell on the Motor vehicles row
    assert "B:E" in out  # styled empty on the Freehold land row


def test_inspect_unknown_sheet_exits_2(template):
    assert main(["inspect", "--workbook", template, "--sheet", "Nope"]) == 2


# ------------------------------------------------------------ footnotes (notes)

from mtool.offline_fill import (  # noqa: E402
    get_defined_names,
    inspect_footnotes,
    read_footnote_rows,
    _parse_defined_ref,
)


def _esc(text):
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# Shared-string table, matching real mTool (openpyxl writes inline strings +
# omits sharedStrings.xml, which the fn_* write path can't exercise).
_FN_STRINGS = [
    "Property, plant and equipment",                                 # 0
    "[Text block added]",                                            # 1
    "Inventories",                                                   # 2
    "Corporate information",                                         # 3
    "fn_14",                                                         # 4
    "Notes-Listofnotes",                                            # 5
    "<html><body><p>Existing PPE note with a <table></table>"        # 6
    "</p></body></html>",
    "fn_20",                                                         # 7
    "fn_99",                                                         # 8
    "<html><body><p>orphan</p></body></html>",                       # 9
    "Deferred taxation",                                             # 10
]

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def _ws(rows, dimension=None):
    """A worksheet XML from {row: [(col, sst_index_or_None), ...]}."""
    body = []
    for r in sorted(rows):
        cells = "".join(
            f'<c r="{col}{r}"/>' if idx is None
            else f'<c r="{col}{r}" t="s"><v>{idx}</v></c>'
            for col, idx in rows[r])
        body.append(f'<row r="{r}">{cells}</row>')
    dim = f'<dimension ref="{dimension}"/>' if dimension else ""
    return (f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<worksheet xmlns="{_NS_MAIN}">{dim}<sheetData>'
            + "".join(body) + "</sheetData></worksheet>")


@pytest.fixture
def footnote_template(tmp_path):
    """A workbook faithfully mimicking mTool's prose-note text-block storage:
    visible trigger cells backed by ``fn_*`` defined names pointing at hidden
    ``+FootnoteTexts`` shared-string payload rows (Windows recon 2026-07-05).
    Hand-built (not openpyxl) so it uses a real shared-strings table."""
    sst = (f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           f'<sst xmlns="{_NS_MAIN}" count="{len(_FN_STRINGS)}" '
           f'uniqueCount="{len(_FN_STRINGS)}">'
           + "".join(f"<si><t>{_esc(s)}</t></si>" for s in _FN_STRINGS)
           + "</sst>")
    sheet1 = _ws({132: [("D", 0), ("E", 1)],      # Notes-Listofnotes: PPE
                  140: [("D", 2), ("E", 1)]})      #   + Inventories (empty pay)
    sheet2 = _ws({14: [("D", 3)],                  # Notes-CI: no fn_* trigger
                  20: [("D", 10)]})                #   a 2nd un-backed label row
    sheet3 = _ws({14: [("A", 4), ("B", 5), ("C", 6)],   # fn_14 populated
                  15: [("A", 7), ("B", 5), ("C", None)],  # fn_20 empty payload
                  16: [("A", 8), ("B", 5), ("C", 9)]},    # fn_99 orphan row
                 dimension="A1:G16")
    workbook = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="{_NS_MAIN}" xmlns:r="{_NS_R}"><sheets>'
        '<sheet name="Notes-Listofnotes" sheetId="1" r:id="rId1"/>'
        '<sheet name="Notes-CI" sheetId="2" r:id="rId2"/>'
        '<sheet name="+FootnoteTexts" sheetId="3" r:id="rId3"/></sheets>'
        '<definedNames>'
        "<definedName name=\"fn_14\">'Notes-Listofnotes'!$E$132</definedName>"
        "<definedName name=\"fn_20\">'Notes-Listofnotes'!$E$140</definedName>"
        '</definedNames></workbook>')
    wb_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        f'2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet2.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet3.xml"/>'
        '<Relationship Id="rId4" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/sharedStrings" '
        'Target="sharedStrings.xml"/></Relationships>')
    content_types = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Types xmlns="http://schemas.openxmlformats.org/package/2006/'
        f'content-types">'
        '<Default Extension="rels" ContentType="application/vnd.'
        'openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.'
        'openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        + "".join(
            f'<Override PartName="/xl/worksheets/sheet{i}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.worksheet+xml"/>' for i in (1, 2, 3))
        + '<Override PartName="/xl/sharedStrings.xml" ContentType="application/'
        'vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '</Types>')
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/'
        '2006/relationships"><Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        'relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>')

    path = tmp_path / "mtool_notes.xlsx"
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", root_rels)
        z.writestr("xl/workbook.xml", workbook)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/sharedStrings.xml", sst)
        z.writestr("xl/worksheets/sheet1.xml", sheet1)
        z.writestr("xl/worksheets/sheet2.xml", sheet2)
        z.writestr("xl/worksheets/sheet3.xml", sheet3)
    return str(path)


def test_parse_defined_ref():
    assert _parse_defined_ref("'Notes-Listofnotes'!$E$132") == \
        ("Notes-Listofnotes", "E132")
    assert _parse_defined_ref("Notes-CI!E14") == ("Notes-CI", "E14")
    assert _parse_defined_ref("'It''s a note'!$A$1") == ("It's a note", "A1")
    assert _parse_defined_ref("'Sheet'!$A$1:$B$2") is None  # ranges rejected
    assert _parse_defined_ref("no-bang") is None


def test_get_defined_names_finds_fn_targets(footnote_template):
    _, data, _ = load_workbook_entries(footnote_template)
    names = get_defined_names(data, "fn_")
    assert set(names) == {"fn_14", "fn_20"}
    assert names["fn_14"] == {"sheet": "Notes-Listofnotes", "cell": "E132",
                              "local_sheet_id": None}


def test_read_footnote_rows_maps_payloads(footnote_template):
    _, data, _ = load_workbook_entries(footnote_template)
    sheet_paths = get_sheet_paths(data)
    from mtool.offline_fill import get_shared_strings
    fn_rows = read_footnote_rows(data, sheet_paths, get_shared_strings(data))
    assert fn_rows["fn_14"]["payload_col"] == "C"
    assert fn_rows["fn_14"]["payload_populated"] is True
    assert fn_rows["fn_20"]["payload_populated"] is False  # present but empty
    assert "fn_99" in fn_rows  # every A-keyed row, defined-name or not


def test_inspect_footnotes_joins_labels_and_payloads(footnote_template):
    _, data, _ = load_workbook_entries(footnote_template)
    info = inspect_footnotes(data)
    assert info["footnote_sheet"] == "+FootnoteTexts"
    by_key = {t["key"]: t for t in info["targets"]}
    # fn_14: backed, populated, and its visible-row label is discoverable.
    assert by_key["fn_14"]["payload_populated"] is True
    assert by_key["fn_14"]["row_text"]["D"] == "Property, plant and equipment"
    # fn_20: backed row but empty payload (a target to fill).
    assert by_key["fn_20"]["has_payload_row"] is True
    assert by_key["fn_20"]["payload_populated"] is False
    # fn_99 is an orphan payload row (no defined name) — surfaced, not a target.
    assert "fn_99" not in by_key
    assert info["orphan_payload_keys"] == ["fn_99"]


def test_footnotes_command_reports_coverage(footnote_template, capsys):
    assert main(["footnotes", "--workbook", footnote_template]) == 0
    out = capsys.readouterr().out
    assert "2 fn_* note target(s)" in out
    assert "1 payload-populated" in out and "1 payload-empty" in out
    assert "Property, plant and equipment" in out
    assert "fn_99" in out  # orphan surfaced


def test_footnotes_command_json(footnote_template, capsys):
    assert main(["footnotes", "--workbook", footnote_template, "--json"]) == 0
    info = json.loads(capsys.readouterr().out)
    assert {t["key"] for t in info["targets"]} == {"fn_14", "fn_20"}


def test_footnotes_command_empty_when_no_fn_targets(template, capsys):
    """A plain sub-sheet workbook has no fn_* — reported, not crashed."""
    assert main(["footnotes", "--workbook", template]) == 0
    assert "no fn_* note targets" in capsys.readouterr().out


# --------------------------------------------------- footnote (note) WRITING

from mtool.offline_fill import (  # noqa: E402
    append_shared_string,
    fill_footnotes,
    get_shared_strings,
    replace_shared_string,
    validate_notes_input,
    wrap_footnote_html,
)

PPE_HTML = ('<h3>Property, plant and equipment</h3><p>Depreciated on a '
            'straight-line basis.</p><table><tr><td>Total &amp; net</td>'
            '<td>9,340</td></tr></table>')


def _payload_of(path, key):
    """Read back the +FootnoteTexts payload text for an fn_ key."""
    from mtool.offline_fill import (get_sheet_paths, load_workbook_entries,
                                    read_footnote_rows)
    _, data, _ = load_workbook_entries(path)
    sheet_paths = get_sheet_paths(data)
    rows = read_footnote_rows(data, sheet_paths, get_shared_strings(data))
    return rows[key]["payload_text"]


def make_notes_input(tmp_path, footnotes):
    path = tmp_path / "notes.json"
    path.write_text(json.dumps({"footnotes": footnotes}), encoding="utf-8")
    return str(path)


def test_wrap_footnote_html_has_tx27_shell():
    out = wrap_footnote_html("<p>x</p>")
    assert "TX27_HTM 27.0.700.500" in out
    assert 'xmlns="http://www.w3.org/1999/xhtml"' in out
    assert "<p>x</p>" in out
    assert "_x000D_" in out


def test_fill_notes_replaces_populated_payload_in_place(tmp_path,
                                                        footnote_template):
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"key": "fn_14", "html": PPE_HTML}]},
        output_path=str(out))
    assert report["status"] == "ok", report
    assert len(report["footnotes_written"]) == 1
    assert report["footnotes_written"][0]["action"] == "shared_string_replaced"
    payload = _payload_of(str(out), "fn_14")
    assert "Property, plant and equipment" in payload
    assert "TX27_HTM" in payload           # wrapped
    assert "Total &amp; net" in payload    # entity round-trips (not &amp;amp;)
    assert "Existing PPE note" not in payload  # old content gone


def test_fill_notes_skips_payload_over_excel_char_limit(tmp_path,
                                                        footnote_template):
    """A note whose wrapped payload exceeds Excel's 32,767-char cell limit is
    SKIPPED (never truncated) so the workbook opens without Excel's repair —
    and flagged with reason 'oversize' + payload_chars. (Windows 2026-07-06
    incident: heavy inline table styling inflated a ~1k-rendered note past 32k.)
    """
    from mtool.offline_fill import EXCEL_CELL_CHAR_LIMIT, wrap_footnote_html
    huge = "<p>" + ("x" * (EXCEL_CELL_CHAR_LIMIT + 5000)) + "</p>"
    assert len(wrap_footnote_html(huge)) > EXCEL_CELL_CHAR_LIMIT
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"key": "fn_14", "html": huge}]},
        output_path=str(out))
    assert report["status"] == "degraded"
    assert report["footnotes_written"] == []
    over = [u for u in report["unresolved"] if u.get("reason") == "oversize"]
    assert len(over) == 1
    assert over[0]["payload_chars"] > EXCEL_CELL_CHAR_LIMIT
    # Every report entry now carries the exporter's size tier (traceability by
    # label — Step 2); None here since this hand-built doc set no tier.
    assert "format_tier" in over[0]
    # workbook is still a readable zip and the old payload is untouched.
    assert "x" * 100 not in _payload_of(str(out), "fn_14")


def test_fill_notes_writes_note_at_exactly_the_limit(tmp_path,
                                                     footnote_template):
    """Boundary: a payload exactly AT the limit still writes (guard is strict >)."""
    from mtool.offline_fill import EXCEL_CELL_CHAR_LIMIT, wrap_footnote_html
    shell = len(wrap_footnote_html(""))
    body = "y" * (EXCEL_CELL_CHAR_LIMIT - shell)
    assert len(wrap_footnote_html(body)) == EXCEL_CELL_CHAR_LIMIT
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template, {"footnotes": [{"key": "fn_14", "html": body}]},
        output_path=str(out))
    assert report["status"] == "ok", report
    assert len(report["footnotes_written"]) == 1


def test_fill_notes_appends_for_empty_payload(tmp_path, footnote_template):
    """fn_20's payload cell (C15) is empty: append a shared string + repoint."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"key": "fn_20", "html": "<p>Inventories note</p>"}]},
        output_path=str(out))
    assert report["status"] == "ok", report
    assert report["footnotes_written"][0]["action"] == "shared_string_appended"
    assert "Inventories note" in _payload_of(str(out), "fn_20")


def test_fill_notes_resolves_by_visible_cell(tmp_path, footnote_template):
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"sheet": "Notes-Listofnotes", "cell": "E132",
                        "html": PPE_HTML}]},
        output_path=str(out))
    assert report["status"] == "ok", report
    assert report["footnotes_written"][0]["key"] == "fn_14"


def test_fill_notes_unresolved_when_no_fn_backs_cell(tmp_path,
                                                     footnote_template):
    """Notes-CI!E14 has no fn_* — V1 refuses, reports unresolved, no crash."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"sheet": "Notes-CI", "cell": "E14",
                        "html": "<p>x</p>"}]},
        output_path=str(out))
    assert report["status"] == "degraded"
    assert len(report["unresolved"]) == 1
    assert not report["footnotes_written"]


def test_fill_notes_other_zip_entries_byte_identical(tmp_path,
                                                     footnote_template):
    """An in-place payload replace touches ONLY sharedStrings.xml; every other
    zip entry is copied verbatim (gotcha #28 discipline)."""
    out = tmp_path / "filled.xlsx"
    fill_footnotes(footnote_template,
                   {"footnotes": [{"key": "fn_14", "html": PPE_HTML}]},
                   output_path=str(out))
    with zipfile.ZipFile(footnote_template) as zin, zipfile.ZipFile(out) as zo:
        changed = {name for name in zin.namelist()
                   if zin.read(name) != zo.read(name)}
    assert changed == {"xl/sharedStrings.xml"}


def test_fill_notes_dry_run_writes_nothing(tmp_path, footnote_template):
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"key": "fn_14", "html": PPE_HTML}]},
        output_path=str(out), dry_run=True)
    assert report["footnotes_written"] and not out.exists()


def test_validate_notes_input():
    assert validate_notes_input({"footnotes": []})
    assert validate_notes_input({"footnotes": [{"key": "fn_1"}]})  # no html
    assert validate_notes_input({"footnotes": [{"html": "x"}]})    # no target
    assert not validate_notes_input(
        {"footnotes": [{"key": "fn_1", "html": "<p>x</p>"}]})


def test_shared_string_ops_roundtrip():
    sst = ('<sst count="2" uniqueCount="2"><si><t>a</t></si>'
           '<si><t>b</t></si></sst>')
    out = replace_shared_string(sst, 0, "<p>x & y</p>")
    assert "&lt;p&gt;x &amp; y&lt;/p&gt;" in out
    out2, idx = append_shared_string(sst, "z")
    # Appending an <si> adds a UNIQUE string but NO cell reference: uniqueCount
    # rises (2 -> 3), `count` (the reference total) stays 2. Conflating them was
    # the sharedStrings "String properties" corruption (Windows 2026-07-06).
    assert idx == 2 and 'count="2"' in out2 and 'uniqueCount="3"' in out2


def _assert_sst_consistent(path):
    """The written sharedStrings header must match reality: uniqueCount = the
    <si> count, count = every t="s" reference across all worksheets. A mismatch
    is what makes Excel 'repair' (corrupt) the file on open."""
    import re
    import zipfile
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        sst = z.read("xl/sharedStrings.xml").decode("utf-8")
        refs = sum(len(re.findall(r'<c\b[^>]*?\bt="s"',
                                  z.read(n).decode("utf-8")))
                   for n in names if n.startswith("xl/worksheets/")
                   and n.endswith(".xml"))
    si_count = len(re.findall(r"<si\b", sst))
    hdr = re.search(r"<sst\b[^>]*>", sst).group(0)
    declared_count = int(re.search(r'\bcount="(\d+)"', hdr).group(1))
    declared_unique = int(re.search(r'\buniqueCount="(\d+)"', hdr).group(1))
    assert declared_unique == si_count, (
        f"uniqueCount={declared_unique} but {si_count} <si> present")
    assert declared_count == refs, (
        f"count={declared_count} but {refs} t=\"s\" references present")


def test_fill_notes_keeps_sst_counts_consistent_on_append(tmp_path,
                                                          footnote_template):
    """Filling an empty payload cell (append <si> + repoint) leaves the header
    exactly consistent — the repoint reuses the cell's reference, so `count`
    must NOT rise for it."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"key": "fn_20", "html": "<p>Inventories note</p>"}]},
        output_path=str(out))
    assert report["status"] == "ok", report
    _assert_sst_consistent(str(out))


def test_create_missing_keeps_sst_counts_consistent(tmp_path,
                                                    orphan_pool_template):
    """The create path adds a new row (3 t=\"s\" refs) + a trigger + appended
    <si> payloads. The header must still reconcile exactly."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        orphan_pool_template,
        {"footnotes": [
            {"label": "Accrued expenses and other liabilities",
             "html": "<p>accruals</p>"},
            {"label": "Capital management", "html": "<p>capman</p>"},
            {"label": "Deferred taxation", "html": "<p>dt</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    _assert_sst_consistent(str(out))


def test_fill_notes_command_end_to_end(tmp_path, footnote_template, capsys):
    out = tmp_path / "filled.xlsx"
    report_path = tmp_path / "report.json"
    code = main(["fill-notes", "--workbook", footnote_template,
                 "--input", make_notes_input(
                     tmp_path, [{"key": "fn_14", "html": PPE_HTML}]),
                 "--output", str(out), "--report", str(report_path)])
    assert code == 0
    report = json.loads(report_path.read_text())
    assert report["status"] == "ok"
    assert "Property, plant and equipment" in _payload_of(str(out), "fn_14")


def test_fill_notes_command_refuses_in_place(footnote_template):
    with pytest.raises(SystemExit):
        main(["fill-notes", "--workbook", footnote_template,
              "--input", footnote_template, "--output", footnote_template])


def test_resolve_footnote_by_label_is_decoration_tolerant():
    from mtool.offline_fill import resolve_footnote_by_label
    # A lightweight target list shaped like inspect_footnotes output.
    targets = [
        {"key": "fn_25", "row_text": {
            "D": "*Disclosure of property, plant and equipment [text block]"}},
        {"key": "fn_11", "row_text": {
            "D": "*Disclosure of corporate information [text block]"}},
    ]
    res = resolve_footnote_by_label("Property, plant and equipment", targets)
    assert res["status"] == "resolved" and res["key"] == "fn_25"
    assert resolve_footnote_by_label("nonexistent note", targets)["status"] \
        == "unresolved"


def test_fill_notes_by_label_resolves_and_writes(tmp_path, footnote_template):
    """Label targeting: no hand-picked key/cell — fill_footnotes matches the
    visible-row label to the fn_* itself."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Property, plant and equipment",
                        "html": PPE_HTML}]},
        output_path=str(out))
    assert report["status"] == "ok", report
    assert report["footnotes_written"][0]["key"] == "fn_14"
    assert "Property, plant and equipment" in _payload_of(str(out), "fn_14")


def test_fill_notes_strict_refuses_non_exact_label(tmp_path, footnote_template):
    """A containment/fuzzy label hit is written lenient but REFUSED under strict
    (doc-level), so a machine doc can't land prose in a near-miss text-block."""
    item = {"label": "plant and equipment", "html": PPE_HTML}  # ~contains fn_14
    # Lenient: resolves via containment to fn_14.
    r1 = fill_footnotes(footnote_template, {"footnotes": [dict(item)]},
                        output_path=str(tmp_path / "a.xlsx"))
    assert r1["footnotes_written"] and r1["footnotes_written"][0]["key"] == "fn_14"
    # Strict (doc-level, as the notes exporter sets): refused, not written.
    r2 = fill_footnotes(footnote_template,
                        {"footnotes": [dict(item)], "strict": True},
                        output_path=str(tmp_path / "b.xlsx"))
    assert r2["strict"] is True
    assert not r2["footnotes_written"] and r2["unresolved"]
    assert "strict" in r2["unresolved"][0]["detail"]


def test_fill_notes_strict_still_allows_exact_label(tmp_path, footnote_template):
    """Exact normalized-core match (decoration aside) survives strict."""
    r = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Property, plant and equipment",
                        "html": PPE_HTML}], "strict": True},
        output_path=str(tmp_path / "c.xlsx"))
    assert r["status"] == "ok" and r["footnotes_written"][0]["key"] == "fn_14"


def test_fill_notes_by_label_unresolved_is_reported(tmp_path, footnote_template):
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Deferred tax liabilities", "html": "<p>x</p>"}]},
        output_path=str(out))
    assert report["status"] == "degraded"
    assert report["unresolved"] and not report["footnotes_written"]


# --------------------------------------------------- create-missing slots

from mtool.offline_fill import (  # noqa: E402
    get_defined_names,
    get_sheet_paths,
    read_sheet_cells,
)


def _visible_cell(path, sheet, addr):
    from mtool.offline_fill import load_workbook_entries as _lwe
    _, data, _ = _lwe(path)
    sp = get_sheet_paths(data)
    rows = read_sheet_cells(data[sp[sheet]], get_shared_strings(data))
    col, row = addr[0], int(addr[1:])
    return rows.get(row, {}).get(col)


def test_create_missing_slot_end_to_end(tmp_path, footnote_template):
    """Notes-CI!E14 has no fn_*; create_missing builds a native-shaped slot."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"sheet": "Notes-CI", "cell": "E14",
                        "html": "<p>Deferred tax note</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    assert len(report["footnotes_created"]) == 1
    created = report["footnotes_created"][0]
    assert created["action"] == "slot_created"
    key = created["key"]

    from mtool.offline_fill import load_workbook_entries as _lwe
    _, data, _ = _lwe(str(out))
    # defined name added, pointing at the visible cell, sheet-scoped.
    defined = get_defined_names(data, "fn_")
    assert defined[key] == {"sheet": "Notes-CI", "cell": "E14",
                            "local_sheet_id": "1"}  # 0-based sheet index
    # visible trigger set to the marker; hidden payload holds the note.
    assert _visible_cell(str(out), "Notes-CI", "E14") == ("S", "[Text block added]")
    assert "Deferred tax note" in _payload_of(str(out), key)
    # +FootnoteTexts dimension extended past the new row.
    sp = get_sheet_paths(data)
    fn_xml = data[sp["+FootnoteTexts"]].decode()
    assert 'ref="A1:G17"' in fn_xml  # was A1:G16, new row 17


def test_create_missing_batch_allocates_distinct_slots(tmp_path,
                                                       footnote_template):
    """Two creates in one pass get distinct fn_ AND distinct rows — the batch
    allocation is drawn from one evolving state (no collision/race)."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [
            {"sheet": "Notes-CI", "cell": "E14", "html": "<p>AAA</p>"},
            {"sheet": "Notes-CI", "cell": "E20", "html": "<p>BBB</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    created = report["footnotes_created"]
    assert len(created) == 2
    keys = [c["key"] for c in created]
    rows = [c["hidden_cell"] for c in created]
    assert len(set(keys)) == 2, keys        # distinct fn_ numbers
    assert len(set(rows)) == 2, rows        # distinct +FootnoteTexts rows
    assert "AAA" in _payload_of(str(out), keys[0])
    assert "BBB" in _payload_of(str(out), keys[1])
    # both defined names present and read-back-consistent
    from mtool.offline_fill import load_workbook_entries as _lwe
    _, data, _ = _lwe(str(out))
    defined = get_defined_names(data, "fn_")
    assert set(keys) <= set(defined)


def test_create_missing_off_by_default_reports_unresolved(tmp_path,
                                                          footnote_template):
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"sheet": "Notes-CI", "cell": "E14", "html": "<p>x</p>"}]},
        output_path=str(out))  # create_missing defaults False
    assert report["status"] == "degraded"
    assert report["unresolved"] and not report["footnotes_created"]
    assert "create_missing" in report["unresolved"][0]["detail"]


def test_create_missing_label_with_no_visible_match_never_creates(
        tmp_path, footnote_template):
    """A label that matches NO existing fn_* AND no visible note-sheet label
    can't locate a cell → stays unresolved, never guessed, even with
    create_missing on."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Totally unknown note", "html": "<p>x</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "degraded"
    assert report["unresolved"] and not report["footnotes_created"]


def test_create_missing_by_label_finds_visible_cell_and_creates(
        tmp_path, footnote_template):
    """The automatic path: a note whose concept has no fn_* (Notes-CI
    'Corporate information' at D14) is CREATED at the trigger cell one column
    right of the visible label (E14) — no hand-picked cell."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Corporate information",
                        "html": "<p>CI note body</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    assert len(report["footnotes_created"]) == 1
    created = report["footnotes_created"][0]
    assert created["action"] == "slot_created"
    assert created["resolved_via"] == "label->cell"
    assert created["visible_cell"] == "Notes-CI!E14"   # D14 label -> E14 trigger
    assert created["label_cell"] == "D14"
    key = created["key"]
    # trigger set + payload landed + defined name points at the discovered cell
    assert _visible_cell(str(out), "Notes-CI", "E14") == ("S", "[Text block added]")
    assert "CI note body" in _payload_of(str(out), key)
    from mtool.offline_fill import load_workbook_entries as _lwe
    _, data, _ = _lwe(str(out))
    assert get_defined_names(data, "fn_")[key]["cell"] == "E14"


def test_create_missing_by_label_off_by_default_stays_unresolved(
        tmp_path, footnote_template):
    """Without create_missing, a label with no existing fn_* is unresolved
    (the pre-existing behaviour) — creation is strictly opt-in."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Corporate information", "html": "<p>x</p>"}]},
        output_path=str(out))  # create_missing defaults False
    assert report["status"] == "degraded"
    assert report["unresolved"] and not report["footnotes_created"]


def test_create_missing_by_label_strict_refuses_non_exact_visible_label(
        tmp_path, footnote_template):
    """Under strict, a near-miss visible label (containment, ratio<1.0) is
    refused rather than creating a slot at a merely-similar row."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Corporate info",  # ~contains 'Corporate information'
                        "html": "<p>x</p>"}], "strict": True},
        output_path=str(out), create_missing=True)
    assert report["strict"] is True
    assert not report["footnotes_created"]
    assert report["unresolved"] and "strict" in report["unresolved"][0]["detail"]


def test_resolve_label_to_note_cell_targets_column_right_of_label():
    from mtool.offline_fill import resolve_label_to_note_cell
    note_sheets = {
        "Notes-CI": {"label_col": "D",
                     "cells": {14: {"D": ("S", "Corporate information")}}},
        "Notes-Listofnotes": {"label_col": "D",
                              "cells": {25: {"D": ("S", "Inventories")}}}}
    res = resolve_label_to_note_cell("Corporate information", note_sheets)
    assert res["status"] == "resolved"
    assert res["sheet"] == "Notes-CI" and res["cell"] == "E14"
    assert res["label_cell"] == "D14"
    assert resolve_label_to_note_cell("no such note", note_sheets)["status"] \
        == "unresolved"


def test_resolve_label_ignores_same_text_outside_the_label_column():
    """Matching is confined to the sheet's label column. A duplicate of the
    label text in another column (a section heading in col A, a stray col F
    cell) must NOT win or force a false ambiguity — only the col-D label
    counts, and the trigger is col E."""
    from mtool.offline_fill import resolve_label_to_note_cell
    note_sheets = {"Notes-CI": {"label_col": "D", "cells": {
        3: {"A": ("S", "Corporate information")},    # a heading, NOT a label
        14: {"D": ("S", "Corporate information"),    # the real label
             "F": ("S", "Corporate information")},   # a stray duplicate
    }}}
    res = resolve_label_to_note_cell("Corporate information", note_sheets)
    assert res["status"] == "resolved"
    assert res["cell"] == "E14" and res["label_cell"] == "D14"


def test_resolve_label_ambiguous_across_two_label_rows():
    """The same label in the label column on two rows is genuinely ambiguous —
    creating at one guessed row is worse than not creating. The tie is
    returned as STRUCTURED candidates (label cell + trigger cell) so a UI can
    offer it as a pick-one decision instead of a dead-end."""
    from mtool.offline_fill import resolve_label_to_note_cell
    note_sheets = {"Notes-CI": {"label_col": "D", "cells": {
        10: {"D": ("S", "Corporate information")},
        20: {"D": ("S", "Corporate information")}}}}
    res = resolve_label_to_note_cell("Corporate information", note_sheets)
    assert res["status"] == "ambiguous"
    assert res["candidates"] == [
        {"sheet": "Notes-CI", "label_cell": "D10", "cell": "E10",
         "matched_label": "Corporate information"},
        {"sheet": "Notes-CI", "label_cell": "D20", "cell": "E20",
         "matched_label": "Corporate information"}]


def test_unresolved_entries_carry_reason_codes(tmp_path, footnote_template):
    """Every unresolved note carries a machine-readable ``reason`` (and, for
    an ambiguous tie, structured candidates) — the decision-UI contract."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [
            # No visible label anywhere -> no_match (create on).
            {"label": "Totally unknown note", "html": "<p>x</p>"},
            # Explicit un-backed cell with create OFF is exercised separately
            # below; here the strict near-miss: 'Corporate info' ~ contains.
            {"label": "Corporate info", "html": "<p>y</p>"}],
         "strict": True},
        output_path=str(out), create_missing=True)
    by_label = {u["label"]: u for u in report["unresolved"]}
    assert by_label["Totally unknown note"]["reason"] == "no_match"
    near = by_label["Corporate info"]
    assert near["reason"] == "strict_near_miss"
    # The refused suggestion is structured (target cell + matched label) so
    # the operator can accept it explicitly.
    assert near["sheet"] == "Notes-CI" and near["cell"] == "E14"
    assert near["matched_label"] == "Corporate information"

    out2 = tmp_path / "filled2.xlsx"
    report2 = fill_footnotes(
        footnote_template,
        {"footnotes": [{"sheet": "Notes-CI", "cell": "E14",
                        "html": "<p>x</p>"}]},
        output_path=str(out2))  # create_missing OFF -> no_slot
    assert report2["unresolved"][0]["reason"] == "no_slot"


def test_note_label_column_derived_from_existing_trigger_and_density():
    from mtool.offline_fill import _note_label_column
    rows = {14: {"D": ("S", "Corporate information")},
            15: {"D": ("S", "Deferred tax")}}
    # With a known trigger col E, the label column is one to its left (D).
    assert _note_label_column(rows, "E") == "D"
    # Without a trigger, fall back to the densest shared-string column (D).
    assert _note_label_column(rows, None) == "D"


def test_create_missing_by_label_does_not_target_heading_in_wrong_column(
        tmp_path, footnote_template):
    """End-to-end: a label whose text also appears as a heading in column A of
    a note sheet creates at the col-E trigger of the col-D label, never col B
    of the heading. (Notes-CI D14 = 'Corporate information'.)"""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"label": "Corporate information",
                        "html": "<p>x</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    assert report["footnotes_created"][0]["visible_cell"] == "Notes-CI!E14"


def test_create_missing_duplicate_labels_do_not_double_create(
        tmp_path, footnote_template):
    """Two notes whose labels resolve to the SAME visible cell create ONE slot;
    the second is reported as a duplicate, never a second fn_* on that cell."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [
            {"label": "Corporate information", "html": "<p>first</p>"},
            {"label": "Corporate information", "html": "<p>second</p>"}]},
        output_path=str(out), create_missing=True)
    assert len(report["footnotes_created"]) == 1
    assert report["errors"] and "duplicate" in report["errors"][0]["error"]
    # Exactly one fn_* now anchors Notes-CI!E14.
    from mtool.offline_fill import load_workbook_entries as _lwe
    _, data, _ = _lwe(str(out))
    anchored = [k for k, d in get_defined_names(data, "fn_").items()
                if d["sheet"] == "Notes-CI" and d["cell"] == "E14"]
    assert len(anchored) == 1, anchored


def test_idx_to_col_roundtrips():
    from mtool.offline_fill import col_to_idx, idx_to_col
    for col in ("A", "E", "Z", "AA", "AZ", "BA"):
        assert idx_to_col(col_to_idx(col)) == col
    assert idx_to_col(col_to_idx("D") + 1) == "E"


def test_cell_pattern_does_not_swallow_past_self_closing_cell():
    """The recon guide's self-closing-cell hazard: patching a `<c/>` must not
    consume following cells/rows up to the next `</c>`."""
    from mtool.offline_fill import _cell_pattern, patch_shared_cell
    xml = ('<worksheet><sheetData>'
           '<row r="15"><c r="A15" t="s"><v>7</v></c><c r="C15"/></row>'
           '<row r="16"><c r="A16" t="s"><v>8</v></c>'
           '<c r="C16" t="s"><v>9</v></c></row>'
           '</sheetData></worksheet>')
    assert _cell_pattern("C15").search(xml).group(0) == '<c r="C15"/>'
    out = patch_shared_cell(xml, "C15", 42)
    assert '<c r="C15" t="s"><v>42</v></c>' in out
    assert '<c r="A16" t="s"><v>8</v></c>' in out   # row 16 intact
    assert '<c r="C16" t="s"><v>9</v></c>' in out
    assert out.count("<row") == 2                    # rows not merged


# ------------------------------------- orphan fn_ pool (Amgen popup incident)

def _write_fn_workbook(tmp_path, name, strings, sheets, defined_names):
    """Minimal hand-built workbook builder for footnote tests: ``sheets`` is an
    ordered list of (sheet_name, worksheet_xml); ``defined_names`` a list of
    raw <definedName> XML strings. Mirrors the footnote_template fixture."""
    sst = (f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           f'<sst xmlns="{_NS_MAIN}" count="{len(strings)}" '
           f'uniqueCount="{len(strings)}">'
           + "".join(f'<si><t xml:space="preserve">{_esc(s)}</t></si>'
                     for s in strings) + "</sst>")
    workbook = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="{_NS_MAIN}" xmlns:r="{_NS_R}"><sheets>'
        + "".join(f'<sheet name="{n}" sheetId="{i+1}" r:id="rId{i+1}"/>'
                  for i, (n, _) in enumerate(sheets))
        + "</sheets><definedNames>" + "".join(defined_names)
        + "</definedNames></workbook>")
    ns_pkg = "http://schemas.openxmlformats.org/package"
    ns_doc = "http://schemas.openxmlformats.org/officeDocument/2006"
    wb_rels = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{ns_pkg}/2006/relationships">'
        + "".join(f'<Relationship Id="rId{i+1}" Type="{ns_doc}/relationships/'
                  f'worksheet" Target="worksheets/sheet{i+1}.xml"/>'
                  for i in range(len(sheets)))
        + f'<Relationship Id="rId{len(sheets)+1}" Type="{ns_doc}/'
          f'relationships/sharedStrings" Target="sharedStrings.xml"/>'
          '</Relationships>')
    content_types = (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Types xmlns="{ns_pkg}/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.'
        'openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.'
        'openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        + "".join(
            f'<Override PartName="/xl/worksheets/sheet{i+1}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.worksheet+xml"/>' for i in range(len(sheets)))
        + '<Override PartName="/xl/sharedStrings.xml" ContentType="application/'
        'vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
        '</Types>')
    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{ns_pkg}/2006/relationships">'
        f'<Relationship Id="rId1" Type="{ns_doc}/relationships/'
        'officeDocument" Target="xl/workbook.xml"/></Relationships>')
    path = tmp_path / name
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", root_rels)
        z.writestr("xl/workbook.xml", workbook)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/sharedStrings.xml", sst)
        for i, (_, xml) in enumerate(sheets):
            z.writestr(f"xl/worksheets/sheet{i+1}.xml", xml)
    return str(path)


_ORPHAN_STRINGS = [
    "Corporate information",                              # 0 (fn_1's label)
    "[Text block added]",                                 # 1
    "Accrued expenses and other liabilities",             # 2
    "Capital management",                                 # 3
    "fn_1",                                               # 4
    "Notes-Listofnotes",                                  # 5
    "<html><body><p>existing CI note</p></body></html>",  # 6
    "fn_5",                                               # 7
    "fn_6",                                               # 8
    "",                                                   # 9 shared empty <si>
    "Deferred taxation",                                  # 10
]


@pytest.fixture
def orphan_pool_template(tmp_path):
    """Real mTool template shape (2026-07-05 Amgen recon): a pool of EMPTY
    pre-provisioned ``+FootnoteTexts`` rows keyed ``fn_N`` in column A with NO
    defined name yet — the free slots mTool assigns when an operator adds a
    text block. fn_5's payload cell shares the "" <si> with a visible cell
    (sharedStrings dedup) to pin the no-replace-on-empty invariant."""
    visible = _ws({9: [("D", 0), ("E", 1)],       # fn_1's trigger row
                   10: [("D", 2)],                 # label, no trigger yet
                   11: [("D", 3)],                 # label, no trigger yet
                   12: [("D", 10), ("F", 9)]})     # F12 shares the "" <si>
    fn_sheet = _ws({1: [("A", 4), ("B", 5), ("C", 6)],   # fn_1 populated
                    5: [("A", 7), ("B", 5), ("C", 9)],   # orphan: t="s" -> ""
                    6: [("A", 8), ("B", 5), ("C", None)]},  # orphan: styled-empty
                   dimension="A1:G6")
    return _write_fn_workbook(
        tmp_path, "orphan_pool.xlsx", _ORPHAN_STRINGS,
        [("Notes-Listofnotes", visible), ("+FootnoteTexts", fn_sheet)],
        ["<definedName name=\"fn_1\">'Notes-Listofnotes'!$E$9</definedName>"])


def test_create_missing_reuses_pre_provisioned_orphan_rows(
        tmp_path, orphan_pool_template):
    """The Amgen fix: creates draw from the template's orphan fn_ pool (no new
    rows, no duplicate column-A keys), and only append past exhaustion."""
    from mtool.offline_fill import _detect_duplicate_fn_keys
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        orphan_pool_template,
        {"footnotes": [
            {"label": "Accrued expenses and other liabilities",
             "html": "<p>accruals body</p>"},
            {"label": "Capital management", "html": "<p>capman body</p>"},
            {"label": "Deferred taxation", "html": "<p>dt body</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    created = report["footnotes_created"]
    assert [c["key"] for c in created] == ["fn_5", "fn_6", "fn_7"]
    # The two orphans are REUSED in place; the third (pool exhausted) appends.
    assert [c["slot_source"] for c in created] == [
        "orphan_reused", "orphan_reused", "row_appended"]
    assert [c["hidden_cell"] for c in created] == ["C5", "C6", "C7"]
    assert report["fn_allocation"] == {
        "orphan_pool_initial": 2, "orphan_reused": 2, "row_appended": 1}
    # The invariant itself: no duplicate column-A join keys in the output.
    assert _detect_duplicate_fn_keys(str(out), "+FootnoteTexts") == []
    # Defined names point at the trigger cells (col E, right of the labels).
    _, data, _ = load_workbook_entries(str(out))
    defined = get_defined_names(data, "fn_")
    assert defined["fn_5"]["cell"] == "E10"
    assert defined["fn_6"]["cell"] == "E11"
    assert defined["fn_7"]["cell"] == "E12"
    # Payloads read back from the reused rows.
    assert "accruals body" in _payload_of(str(out), "fn_5")
    assert "capman body" in _payload_of(str(out), "fn_6")
    assert "dt body" in _payload_of(str(out), "fn_7")


def test_orphan_reuse_never_rewrites_shared_empty_string(
        tmp_path, orphan_pool_template):
    """fn_5's empty payload cell points at the shared "" <si> that visible
    F12 also references. Writing into the orphan must APPEND a fresh <si>,
    never replace the shared one — replacing would rewrite every empty-string
    cell in the workbook (same silent-corruption family as the incident)."""
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        orphan_pool_template,
        {"footnotes": [{"label": "Accrued expenses and other liabilities",
                        "html": "<p>accruals body</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    assert report["footnotes_created"][0]["key"] == "fn_5"
    assert "accruals body" in _payload_of(str(out), "fn_5")
    # The witness cell still reads "" — the shared <si> was left untouched.
    assert _visible_cell(str(out), "Notes-Listofnotes", "F12") == ("S", "")


def test_fallback_append_key_skips_orphan_a_keys(tmp_path, footnote_template):
    """fn_used is seeded from EVERY +FootnoteTexts column-A key, not just
    defined names: with fn_99 present as a populated orphan row, an appended
    slot mints fn_100 — never a key that duplicates an existing row (the
    pre-fix allocator minted from max(defined)+1 = fn_21...)."""
    from mtool.offline_fill import _detect_duplicate_fn_keys
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        footnote_template,
        {"footnotes": [{"sheet": "Notes-CI", "cell": "E14",
                        "html": "<p>Deferred tax note</p>"}]},
        output_path=str(out), create_missing=True)
    assert report["status"] == "ok", report
    created = report["footnotes_created"][0]
    assert created["key"] == "fn_100"
    assert created["slot_source"] == "row_appended"
    assert _detect_duplicate_fn_keys(str(out), "+FootnoteTexts") == []


def test_duplicate_fn_keys_detected_and_degrade_report(tmp_path):
    """Defence-in-depth: a workbook that already carries duplicate column-A
    fn_ keys (the exact broken shape the incident shipped) is flagged loudly
    in the report — read_footnote_rows alone masks it (last row wins, while
    mTool reads the first)."""
    from mtool.offline_fill import _detect_duplicate_fn_keys
    strings = ["Corporate information", "[Text block added]", "fn_1",
               "Notes-Listofnotes",
               "<html><body><p>old</p></body></html>"]
    visible = _ws({9: [("D", 0), ("E", 1)]})
    # Two rows share A="fn_1": the empty pre-provisioned row 1 (what mTool
    # reads) and the populated appended row 7 (where the payload landed).
    fn_sheet = _ws({1: [("A", 2), ("B", 3), ("C", None)],
                    7: [("A", 2), ("B", 3), ("C", 4)]}, dimension="A1:G7")
    path = _write_fn_workbook(
        tmp_path, "dup_keys.xlsx", strings,
        [("Notes-Listofnotes", visible), ("+FootnoteTexts", fn_sheet)],
        ["<definedName name=\"fn_1\">'Notes-Listofnotes'!$E$9</definedName>"])
    assert _detect_duplicate_fn_keys(path, "+FootnoteTexts") == [
        {"key": "fn_1", "rows": [1, 7]}]
    out = tmp_path / "filled.xlsx"
    report = fill_footnotes(
        path, {"footnotes": [{"sheet": "Notes-Listofnotes", "cell": "E9",
                              "html": "<p>new</p>"}]},
        output_path=str(out))
    assert report["status"] == "degraded"
    assert any("duplicate +FootnoteTexts column-A key fn_1" in e["error"]
               for e in report["errors"])


def test_refill_of_patched_template_is_idempotent(
        tmp_path, orphan_pool_template):
    """Re-running the same doc onto the already-patched output takes the
    existing-slot fill path (in-place replace) — no second create, no new
    rows, still zero duplicate keys."""
    from mtool.offline_fill import _detect_duplicate_fn_keys
    doc = {"footnotes": [
        {"label": "Accrued expenses and other liabilities",
         "html": "<p>accruals body v2</p>"},
        {"label": "Capital management", "html": "<p>capman body v2</p>"}]}
    out1 = tmp_path / "filled1.xlsx"
    r1 = fill_footnotes(orphan_pool_template, doc, output_path=str(out1),
                        create_missing=True)
    assert len(r1["footnotes_created"]) == 2
    out2 = tmp_path / "filled2.xlsx"
    r2 = fill_footnotes(str(out1), doc, output_path=str(out2),
                        create_missing=True)
    assert r2["status"] == "ok", r2
    assert r2["footnotes_created"] == []          # slots now exist — filled
    assert len(r2["footnotes_written"]) == 2
    assert {w["action"] for w in r2["footnotes_written"]} == {
        "shared_string_replaced"}
    assert "accruals body v2" in _payload_of(str(out2), "fn_5")
    assert "capman body v2" in _payload_of(str(out2), "fn_6")
    assert _detect_duplicate_fn_keys(str(out2), "+FootnoteTexts") == []

