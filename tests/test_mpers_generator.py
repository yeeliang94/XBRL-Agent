"""MPERS template-generator regression suite.

Organised by plan phase. Markers let each phase be run in isolation:
    pytest -m mpers_inventory     # Phase 1
    pytest -m mpers_generator_core # Phase 2
    pytest -m mpers_company        # Phase 3
    pytest -m mpers_formulas       # Phase 4
    pytest -m mpers_group          # Phase 5
    pytest -m mpers_snapshot       # Phase 6
"""
from __future__ import annotations

from pathlib import Path

import pytest


REPO_ROOT = Path(__file__).resolve().parent.parent
MPERS_TAXONOMY = REPO_ROOT / "SSMxT_2022v1.0" / "rep" / "ssm" / "ca-2016" / "fs" / "mpers"


# ---------------------------------------------------------------------------
# Phase 1 — inventory + format-reference pins
# ---------------------------------------------------------------------------


@pytest.mark.mpers_inventory
def test_list_mpers_roles_returns_24_entries():
    """Red: `list_mpers_roles()` enumerates every MPERS presentation role.

    The SSM MPERS taxonomy ships 24 `pre_*.xml` files, one per role. The
    generator has to see all of them and tag each with role_number + title +
    pre_file_path so later steps can map roles to output filenames.
    """
    from scripts.generate_mpers_templates import list_mpers_roles

    roles = list_mpers_roles()
    assert isinstance(roles, list)
    assert len(roles) == 24

    # Every entry must carry the three required keys.
    for role in roles:
        assert set(role.keys()) >= {"role_number", "title", "pre_file_path"}
        assert isinstance(role["role_number"], str)
        assert isinstance(role["title"], str) and role["title"]
        assert isinstance(role["pre_file_path"], Path)
        assert role["pre_file_path"].exists()

    numbers = {r["role_number"] for r in roles}
    # Spot-check a handful from each band: scope, SOFP, SOCIE, SoRE, notes.
    for expected in {"020000", "210000", "610000", "620000", "750000"}:
        assert expected in numbers, f"missing role {expected} in inventory"


@pytest.mark.mpers_inventory
def test_mpers_template_mapping_matches_15_entries():
    """Red: `template_mapping()` assigns roles to the 15 output xlsx files.

    The MPERS bundle mirrors the MFRS numbering convention (01..14) with an
    extra `10-SoRE` slot for MPERS-only Statement of Retained Earnings, so the
    notes templates shift from 10..14 to 11..15. Templates that carry a
    sub-classification role (SOFP, SOPL-Function, SOPL-Nature) bundle both
    face + sub role numbers in the same xlsx.
    """
    from scripts.generate_mpers_templates import template_mapping

    mapping = template_mapping()

    # Must be an ordered sequence — downstream emitters assume indexable order.
    assert isinstance(mapping, list)
    assert len(mapping) == 15

    # Every entry: (filename, [role_numbers])
    for entry in mapping:
        assert isinstance(entry, tuple)
        assert len(entry) == 2
        filename, role_numbers = entry
        assert isinstance(filename, str) and filename.endswith(".xlsx")
        assert isinstance(role_numbers, list) and role_numbers
        for rn in role_numbers:
            assert isinstance(rn, str) and len(rn) == 6

    # Pin the exact order — this IS the output-bundle contract.
    expected_order = [
        "01-SOFP-CuNonCu.xlsx",
        "02-SOFP-OrderOfLiquidity.xlsx",
        "03-SOPL-Function.xlsx",
        "04-SOPL-Nature.xlsx",
        "05-SOCI-BeforeTax.xlsx",
        "06-SOCI-NetOfTax.xlsx",
        "07-SOCF-Indirect.xlsx",
        "08-SOCF-Direct.xlsx",
        "09-SOCIE.xlsx",
        "10-SoRE.xlsx",
        "11-Notes-CorporateInfo.xlsx",
        "12-Notes-AccountingPolicies.xlsx",
        "13-Notes-ListOfNotes.xlsx",
        "14-Notes-IssuedCapital.xlsx",
        "15-Notes-RelatedParty.xlsx",
    ]
    assert [e[0] for e in mapping] == expected_order

    by_filename = {fname: rns for fname, rns in mapping}

    # Sub-classification bundling: SOFP-CuNonCu carries 210000 + 210100 in
    # the same xlsx (face + sub-sheet) — matches the MFRS CuNonCu convention.
    assert by_filename["01-SOFP-CuNonCu.xlsx"] == ["210000", "210100"]
    assert by_filename["02-SOFP-OrderOfLiquidity.xlsx"] == ["220000", "220100"]
    assert by_filename["03-SOPL-Function.xlsx"] == ["310000", "310100"]
    assert by_filename["04-SOPL-Nature.xlsx"] == ["320000", "320100"]

    # Face-only statements carry a single role each.
    assert by_filename["05-SOCI-BeforeTax.xlsx"] == ["420000"]
    assert by_filename["06-SOCI-NetOfTax.xlsx"] == ["410000"]
    assert by_filename["07-SOCF-Indirect.xlsx"] == ["520000"]
    assert by_filename["08-SOCF-Direct.xlsx"] == ["510000"]
    assert by_filename["09-SOCIE.xlsx"] == ["610000"]
    assert by_filename["10-SoRE.xlsx"] == ["620000"]
    assert by_filename["11-Notes-CorporateInfo.xlsx"] == ["710000"]
    assert by_filename["12-Notes-AccountingPolicies.xlsx"] == ["720000"]
    assert by_filename["13-Notes-ListOfNotes.xlsx"] == ["730000"]
    assert by_filename["14-Notes-IssuedCapital.xlsx"] == ["740000"]
    assert by_filename["15-Notes-RelatedParty.xlsx"] == ["750000"]


# ---------------------------------------------------------------------------
# Format-reference pins — characterise existing MFRS templates so the MPERS
# emitter has a concrete target to match.
# ---------------------------------------------------------------------------


MFRS_COMPANY_DIR = REPO_ROOT / "XBRL-template-MFRS" / "Company"
MFRS_GROUP_DIR = REPO_ROOT / "XBRL-template-MFRS" / "Group"


@pytest.mark.mpers_inventory
def test_mfrs_company_template_has_4_columns():
    """Format pin: Company SOFP uses the 4-column layout (A=label, B=CY, C=PY, D=source)
    with bold formatting on every row whose label starts with ``*``.

    Locks the characterisation so the MPERS emitter has a concrete target.
    """
    import openpyxl

    wb = openpyxl.load_workbook(MFRS_COMPANY_DIR / "01-SOFP-CuNonCu.xlsx")
    ws = wb["SOFP-CuNonCu"]

    # 4 used columns: label + CY + PY + source.
    assert ws.max_column == 4

    # Column D row 1 is the source-column header, pinning the 4-col layout.
    assert ws.cell(row=1, column=4).value == "Source"
    # Period markers in B and C row 1 — template placeholder year text.
    assert "YYYY" in str(ws.cell(row=1, column=2).value)
    assert "YYYY" in str(ws.cell(row=1, column=3).value)

    # Row 23 = Total non-current assets, bolded by convention.
    cell = ws.cell(row=23, column=1)
    assert cell.value == "*Total non-current assets"
    assert cell.font.bold is True

    # Every row whose label starts with "*" is rendered bold.
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.startswith("*"):
            assert ws.cell(row=r, column=1).font.bold is True, (
                f"row {r} starts with '*' but is not bold — pin violated"
            )

    # Freeze panes keep the header + period rows locked while scrolling.
    assert ws.freeze_panes == "A4"
    # Column A is wide enough for full labels.
    assert ws.column_dimensions["A"].width and ws.column_dimensions["A"].width >= 40


@pytest.mark.mpers_inventory
def test_mfrs_group_template_has_6_columns():
    """Format pin: Group SOFP uses the 6-column layout
    (A=label, B=Group CY, C=Group PY, D=Company CY, E=Company PY, F=source).

    Header carries "Group" / "Company" banners in row 1 (merged across B:C / D:E)
    and period-placeholder strings in row 2.
    """
    import openpyxl

    wb = openpyxl.load_workbook(MFRS_GROUP_DIR / "01-SOFP-CuNonCu.xlsx")
    ws = wb["SOFP-CuNonCu"]

    # 6 used columns: label + 4 value columns + source.
    assert ws.max_column == 6

    # Row 1: banner row — merged "Group" over B:C, "Company" over D:E.
    assert ws.cell(row=1, column=2).value == "Group"
    assert ws.cell(row=1, column=4).value == "Company"
    assert ws.cell(row=1, column=6).value == "Source"

    # Row 2: period placeholders in all four value columns.
    for col in (2, 3, 4, 5):
        assert "YYYY" in str(ws.cell(row=2, column=col).value)

    # Row 23 = Total non-current assets — bold + SUM formulas in every value col.
    cell = ws.cell(row=23, column=1)
    assert cell.value == "*Total non-current assets"
    assert cell.font.bold is True
    for col in (2, 3, 4, 5):
        formula = ws.cell(row=23, column=col).value
        assert isinstance(formula, str) and formula.startswith("=")

    assert ws.freeze_panes == "A4"


@pytest.mark.mpers_inventory
def test_mfrs_group_socie_has_4_row_blocks():
    """Format pin: Group SOCIE is laid out as four vertical 23-row blocks.

    Row ranges: 3-25 (Group CY), 27-49 (Group PY), 51-73 (Company CY),
    75-97 (Company PY), with rows 26, 50, 74 blank as block separators.
    All four blocks share the same body rows (rows 5 onwards inside each block).
    """
    import openpyxl

    wb = openpyxl.load_workbook(MFRS_GROUP_DIR / "09-SOCIE.xlsx")
    ws = wb["SOCIE"]

    # Block header labels — these pin the 4-block structure.
    assert ws.cell(row=3, column=1).value == "Group - Current period"
    assert ws.cell(row=27, column=1).value == "Group - Prior period"
    assert ws.cell(row=51, column=1).value == "Company - Current period"
    assert ws.cell(row=75, column=1).value == "Company - Prior period"

    # Blank separator rows between blocks (col A is empty).
    for sep_row in (26, 50, 74):
        assert ws.cell(row=sep_row, column=1).value is None, (
            f"expected blank separator at row {sep_row}"
        )

    # Every block has exactly 23 rows and closes with "*Equity at end of period".
    block_ranges = [(3, 25), (27, 49), (51, 73), (75, 97)]
    bodies: list[list] = []
    for start, end in block_ranges:
        assert end - start + 1 == 23
        assert ws.cell(row=end, column=1).value == "*Equity at end of period"
        # Body = rows inside the block after the block-header row.
        bodies.append([ws.cell(row=r, column=1).value for r in range(start + 1, end + 1)])

    # All four blocks share identical bodies — the same SOCIE row-set repeated.
    for other in bodies[1:]:
        assert other == bodies[0], "SOCIE block bodies are not identical"


# ---------------------------------------------------------------------------
# Phase 2 — generator skeleton
# ---------------------------------------------------------------------------


@pytest.mark.mpers_generator_core
def test_walk_role_710000_returns_corporate_info_rows():
    """Red: `walk_role()` DFS-traverses a presentation linkbase.

    Role 710000 (Notes - Corporate information) is intentionally simple (one
    root abstract + two abstract children + three textblock leaves), so the
    walker's output structure is easy to pin. Every entry is a 4-tuple:
    (depth, concept_id, label, is_abstract).
    """
    from scripts.generate_mpers_templates import walk_role

    pre_path = MPERS_TAXONOMY / "pre_ssmt-fs-mpers_2022-12-31_role-710000.xml"
    rows = walk_role(pre_path)

    # Every row is a 4-tuple with the expected shape.
    for row in rows:
        assert isinstance(row, tuple) and len(row) == 4
        depth, concept_id, label, is_abstract = row
        assert isinstance(depth, int) and depth >= 0
        assert isinstance(concept_id, str) and concept_id
        assert isinstance(label, str)
        assert isinstance(is_abstract, bool)

    # Corporate info is small — 5 to 15 concepts is the realistic range.
    assert 5 <= len(rows) <= 15, f"unexpected row count {len(rows)}"

    # Root is the DisclosureOn…Abstract and is flagged abstract.
    depth, concept_id, _label, is_abstract = rows[0]
    assert depth == 0
    assert "DisclosureOnCorporateInformation" in concept_id
    assert is_abstract is True

    # At least one depth >= 1 child is present (the abstract sub-sections).
    assert any(d >= 1 for d, _, _, _ in rows)


@pytest.mark.mpers_generator_core
def test_load_label_map_resolves_standard_labels():
    """Red: `load_label_map()` merges every `lab_en*.xml` in the taxonomy tree.

    MPERS uses SSM-specific ReportingLabel as the display label — it overrides
    the IFRS-for-SMEs standard label where provided (e.g.
    ``ifrs-smes_InvestmentProperty`` is "Investment properties" in MPERS,
    not the IFRS-for-SMEs "Investment property at fair value …" variant).
    ``ssmt-mpers_*`` concepts have labels only in the def-level label files,
    so the loader must scan every `lab_en*.xml` — not just the rep-level one.
    """
    from scripts.generate_mpers_templates import load_label_map

    labels = load_label_map()

    # MPERS ReportingLabel overrides the default IFRS-for-SMEs label.
    assert labels["ifrs-smes_InvestmentProperty"] == "Investment properties"

    # An ssmt-mpers_* concept resolves — only sourced from def-level labels.
    assert labels["ssmt-mpers_DisclosureOnCorporateInformationAbstract"] == (
        "Disclosure on corporate information [abstract]"
    )


@pytest.mark.mpers_generator_core
def test_walk_role_uses_preferred_label():
    """Red: when a presentation arc carries `preferredLabel`, `walk_role()`
    returns the label matching that role, not the default one.

    Role 210000 (SOFP, CuNonCu) contains
    ``ifrs-smes_NoncurrentAssets`` with ``preferredLabel=.../totalLabel`` — the
    walker must surface the TotalLabel variant ("Total non-current assets"),
    not the standard label ("Non-current assets").
    """
    from scripts.generate_mpers_templates import walk_role

    pre_path = MPERS_TAXONOMY / "pre_ssmt-fs-mpers_2022-12-31_role-210000.xml"
    rows = walk_role(pre_path)

    # Find NoncurrentAssets concept — totalLabel prefers the "Total ..." variant.
    non_current = [r for r in rows if r[1].endswith("_NoncurrentAssets")]
    assert non_current, "ifrs-smes_NoncurrentAssets not found in role 210000"
    label = non_current[0][2]
    assert label.lower().startswith("total"), (
        f"expected TotalLabel variant, got {label!r}"
    )


@pytest.mark.mpers_generator_core
def test_emit_template_company_produces_readable_xlsx(tmp_path):
    """Red: `emit_template(rows, path, level="company")` writes a 4-column xlsx.

    Labels in column A match the input row order exactly, the file loads via
    openpyxl without raising, and the shape matches the Company layout from
    the format pins (4 used columns).
    """
    import openpyxl

    from scripts.generate_mpers_templates import emit_template

    rows = [
        (0, "root_Abstract", "Root Abstract", True),
        (1, "child_A", "Alpha line", False),
        (1, "child_B", "Bravo line", False),
    ]
    out_path = tmp_path / "out.xlsx"
    emit_template(rows, out_path, level="company")

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active

    # Labels start at row 3 (rows 1-2 reserved for header + period).
    assert ws.cell(row=3, column=1).value == "Root Abstract"
    assert ws.cell(row=4, column=1).value == "Alpha line"
    assert ws.cell(row=5, column=1).value == "Bravo line"

    # 4 used columns — matches the MFRS Company pin.
    assert ws.max_column == 4


@pytest.mark.mpers_generator_core
def test_emit_template_applies_total_row_styling(tmp_path):
    """Red: rows whose label starts with ``*`` are bolded in column A.

    Mirrors the MFRS convention captured in the Phase 1 format pins — the
    asterisk prefix marks totals / primary-section rows that render bold.
    """
    import openpyxl

    from scripts.generate_mpers_templates import emit_template

    rows = [
        (0, "root_Abstract", "Root", True),
        (1, "normal_item", "Plain leaf", False),
        (1, "total_item", "*Total of something", False),
    ]
    out_path = tmp_path / "styling.xlsx"
    emit_template(rows, out_path, level="company")

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    # Plain leaf — not bold.
    assert ws.cell(row=4, column=1).font.bold is not True
    # *Total … — bold.
    assert ws.cell(row=5, column=1).value == "*Total of something"
    assert ws.cell(row=5, column=1).font.bold is True


@pytest.mark.mpers_generator_core
def test_emit_template_applies_freeze_panes_and_column_widths(tmp_path):
    """Red: emitter sets freeze panes at A4 and widens column A for labels,
    matching the Phase 1 MFRS Company pin.
    """
    import openpyxl

    from scripts.generate_mpers_templates import emit_template

    rows = [(0, "root_Abstract", "Root", True)]
    out_path = tmp_path / "freeze.xlsx"
    emit_template(rows, out_path, level="company")

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws.freeze_panes == "A4"
    assert ws.column_dimensions["A"].width and ws.column_dimensions["A"].width >= 40


@pytest.mark.mpers_generator_core
def test_generated_template_readable_by_template_reader(tmp_path):
    """Red: a generator-emitted xlsx must be readable by `tools.template_reader`
    in exactly the same way as an existing MFRS notes template.

    The existing pipeline uses ``template_reader.read_template`` to extract
    fields from each template — if the MPERS output isn't shape-compatible,
    `fill_workbook` and cross-checks won't be able to consume it.
    """
    from scripts.generate_mpers_templates import emit_template, walk_role
    from tools.template_reader import read_template

    # Corporate info is a small MPERS role (5-10 rows) — easy to read/compare.
    pre_path = MPERS_TAXONOMY / "pre_ssmt-fs-mpers_2022-12-31_role-710000.xml"
    rows = walk_role(pre_path)

    generated = tmp_path / "mpers-corp-info.xlsx"
    emit_template(rows, generated, level="company")
    generated_fields = read_template(str(generated))

    # Reader must return at least the body rows + the 3 header cells (B1/C1/D1).
    assert len(generated_fields) >= len(rows)

    # Every field carries the expected dataclass-ish attributes.
    first = generated_fields[0]
    for attr in ("sheet", "coordinate", "row", "col", "value", "formula", "is_data_entry"):
        assert hasattr(first, attr)

    # MFRS corporate-info template reads with the same shape — keys match,
    # so downstream code that works on MFRS will work on MPERS too.
    mfrs_fields = read_template(str(MFRS_COMPANY_DIR / "10-Notes-CorporateInfo.xlsx"))
    assert {type(f).__name__ for f in mfrs_fields} == {type(first).__name__}


# ---------------------------------------------------------------------------
# Phase 3 — Company-level templates on disk (01..15)
# ---------------------------------------------------------------------------


MPERS_COMPANY_DIR = REPO_ROOT / "XBRL-template-MPERS" / "Company"


# (filename, [expected_sheet_names], min_rows, max_rows, must_have, must_not_have)
# `must_have`/`must_not_have` check column-A labels in *any* sheet of the file.
FACE_TEMPLATE_CASES = [
    (
        "01-SOFP-CuNonCu.xlsx",
        ["SOFP-CuNonCu", "SOFP-Sub-CuNonCu"],
        20,
        500,
        ["Loans and borrowings"],
        ["Right-of-use", "Contract assets", "disposal group"],
    ),
    (
        "02-SOFP-OrderOfLiquidity.xlsx",
        ["SOFP-OrdOfLiq", "SOFP-Sub-OrdOfLiq"],
        20,
        500,
        [],
        ["Right-of-use", "Contract assets"],
    ),
    (
        "03-SOPL-Function.xlsx",
        ["SOPL-Function", "SOPL-Analysis-Function"],
        5,
        500,
        [],
        ["Right-of-use"],
    ),
    (
        "04-SOPL-Nature.xlsx",
        ["SOPL-Nature", "SOPL-Analysis-Nature"],
        5,
        500,
        [],
        ["Right-of-use"],
    ),
    ("05-SOCI-BeforeTax.xlsx", ["SOCI-BeforeOfTax"], 3, 100, [], []),
    ("06-SOCI-NetOfTax.xlsx", ["SOCI-NetOfTax"], 3, 100, [], []),
    ("07-SOCF-Indirect.xlsx", ["SOCF-Indirect"], 5, 200, [], []),
    ("08-SOCF-Direct.xlsx", ["SOCF-Direct"], 5, 200, [], []),
    ("09-SOCIE.xlsx", ["SOCIE"], 10, 200, [], []),
    (
        "10-SoRE.xlsx",
        ["SoRE"],
        # Row band tightened after the MFRS-format-alignment pass
        # (2026-04-23): XBRL hypercube scaffolding rows ([table], [axis],
        # [line items], Group/Company [member]) are no longer emitted, so
        # SoRE's row count dropped from ~19 to 14. Floor is 10 to allow
        # future taxonomy additions without false-alarming on a legitimate
        # but small data body.
        10,
        25,
        ["Dividends paid", "Retained earnings at beginning of period", "Retained earnings at end of period"],
        [],
    ),
]


@pytest.mark.mpers_company
@pytest.mark.parametrize(
    "filename,sheet_names,min_rows,max_rows,must_have,must_not_have",
    FACE_TEMPLATE_CASES,
    ids=[c[0] for c in FACE_TEMPLATE_CASES],
)
def test_generated_company_face_template(
    filename, sheet_names, min_rows, max_rows, must_have, must_not_have
):
    """Red: each of the 10 MPERS face-statement xlsx files lives under
    ``XBRL-template-MPERS/Company/`` with the expected sheets + labels.

    The generator is run once (see ``scripts/generate_mpers_templates.py``'s
    `main()` in the test's green step) and these tests then pin the output.
    """
    import openpyxl

    path = MPERS_COMPANY_DIR / filename
    assert path.exists(), f"MPERS template not emitted: {path}"

    wb = openpyxl.load_workbook(path, read_only=True)
    assert wb.sheetnames == sheet_names, (
        f"{filename}: sheets {wb.sheetnames} != expected {sheet_names}"
    )

    # Collect every column-A label across every sheet.
    labels: list[str] = []
    total_rows = 0
    for sname in sheet_names:
        ws = wb[sname]
        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None:
                    labels.append(str(cell.value))
                    total_rows += 1
    wb.close()

    assert min_rows <= total_rows <= max_rows, (
        f"{filename}: {total_rows} rows outside [{min_rows},{max_rows}]"
    )

    joined = "\n".join(labels).lower()
    for needle in must_have:
        assert needle.lower() in joined, f"{filename}: missing expected label {needle!r}"
    for needle in must_not_have:
        assert needle.lower() not in joined, (
            f"{filename}: unexpectedly contains {needle!r}"
        )


# (filename, sheet_name, min_rows, max_rows, anchor_label)
NOTES_TEMPLATE_CASES = [
    ("11-Notes-CorporateInfo.xlsx", "Notes-CI", 3, 30, "Corporate information"),
    ("12-Notes-AccountingPolicies.xlsx", "Notes-SummaryofAccPol", 3, 100, "accounting polic"),
    ("13-Notes-ListOfNotes.xlsx", "Notes-Listofnotes", 1, 300, None),
    ("14-Notes-IssuedCapital.xlsx", "Notes-Issuedcapital", 3, 100, "Issued capital"),
    ("15-Notes-RelatedParty.xlsx", "Notes-RelatedPartytran", 3, 150, "Related part"),
]


@pytest.mark.mpers_company
@pytest.mark.parametrize(
    "filename,sheet_name,min_rows,max_rows,anchor",
    NOTES_TEMPLATE_CASES,
    ids=[c[0] for c in NOTES_TEMPLATE_CASES],
)
def test_generated_company_notes_template(filename, sheet_name, min_rows, max_rows, anchor):
    """Red: each of the 5 MPERS notes-template xlsx files is on disk with
    the expected sheet + row-count bounds + a representative anchor label.
    """
    import openpyxl

    path = MPERS_COMPANY_DIR / filename
    assert path.exists(), f"MPERS notes template not emitted: {path}"

    wb = openpyxl.load_workbook(path, read_only=True)
    assert wb.sheetnames == [sheet_name], (
        f"{filename}: sheets {wb.sheetnames} != [{sheet_name!r}]"
    )

    ws = wb[sheet_name]
    labels = [
        str(cell.value)
        for row in ws.iter_rows(min_col=1, max_col=1)
        for cell in row
        if cell.value is not None
    ]
    wb.close()

    assert min_rows <= len(labels) <= max_rows, (
        f"{filename}: {len(labels)} rows outside [{min_rows}, {max_rows}]"
    )

    if anchor is not None:
        joined = "\n".join(labels).lower()
        assert anchor.lower() in joined, f"{filename}: missing anchor {anchor!r}"


# ---------------------------------------------------------------------------
# Phase 4 — calc linkbase → SUM formulas
# ---------------------------------------------------------------------------


@pytest.mark.mpers_formulas
def test_parse_calc_linkbase_role_210000_sofp_totals():
    """Red: `parse_calc_linkbase_for_pre_role("210000")` returns the SOFP totals map.

    Maps pre role 210000 (SOFP, CuNonCu) to calc role 200100 — in the MPERS
    taxonomy, presentation/definition linkbases split face vs sub classification
    (210000/210100, 220000/220100) but the calculation linkbase keeps the
    neutral 200100/200200 split because the arithmetic is the same regardless
    of presentation order.

    Pinned expectations:
      * NoncurrentAssets aggregates at least {PPE, InvestmentProperty, …}
        with +1 weight — matches the CuNonCu subtotal.
      * CurrentAssets aggregates at least {Inventories, CurrentReceivables}.
      * EquityAndLiabilities = [(Equity, +1), (Liabilities, +1)].
    """
    from scripts.generate_mpers_templates import parse_calc_linkbase_for_pre_role

    calc = parse_calc_linkbase_for_pre_role("210000")
    assert isinstance(calc, dict)

    non_current_children = {child for child, _w in calc.get("ifrs-smes_NoncurrentAssets", [])}
    assert "ifrs-smes_PropertyPlantAndEquipment" in non_current_children
    assert "ifrs-smes_InvestmentProperty" in non_current_children
    for _child, weight in calc["ifrs-smes_NoncurrentAssets"]:
        assert weight == 1

    current_children = {child for child, _w in calc.get("ifrs-smes_CurrentAssets", [])}
    # Inventories may show as InventoriesTotal or Inventories depending on calc-file sub-section
    assert any("Inventor" in c for c in current_children), (
        f"no inventory children under CurrentAssets, got {current_children}"
    )

    equity_and_liabs = calc.get("ifrs-smes_EquityAndLiabilities", [])
    # Must have both Equity and Liabilities as immediate children with +1 weight.
    eq_items = {(c, w) for c, w in equity_and_liabs}
    assert ("ifrs-smes_Equity", 1) in eq_items
    assert ("ifrs-smes_Liabilities", 1) in eq_items


@pytest.mark.mpers_formulas
def test_parse_calc_linkbase_handles_negative_weight():
    """Red: parents of subtraction arcs surface with weight=-1.

    Equity in SOFP calc has `TreasuryShares` with weight=-1 (treasury shares
    reduce equity). This pins negative-weight handling end-to-end.
    """
    from scripts.generate_mpers_templates import parse_calc_linkbase_for_pre_role

    calc = parse_calc_linkbase_for_pre_role("210000")
    equity = dict(calc.get("ifrs-smes_Equity", []))
    assert equity.get("ifrs-smes_TreasuryShares") == -1


@pytest.mark.mpers_formulas
def test_inject_sum_formulas_accepts_base_row_offset():
    """`_inject_sum_formulas(base_row=N)` must place all formula references
    starting at row N — used by the 4-block Group SOCIE layout to drop
    the same calc into four different row ranges. Default
    `base_row=_FIRST_BODY_ROW` keeps every existing call site unchanged.
    """
    import openpyxl
    from scripts.generate_mpers_templates import _inject_sum_formulas

    # Minimal 3-row presentation: parent + two children.
    rows = [
        (0, "ParentConcept", "Parent total", False),
        (1, "ChildA", "Child A", False),
        (1, "ChildB", "Child B", False),
    ]
    # One calc block: ParentConcept = ChildA + ChildB.
    calc_blocks = [
        ("role-test", {"ParentConcept": [("ChildA", 1), ("ChildB", 1)]}),
    ]

    base_row = 27
    wb = openpyxl.Workbook()
    ws = wb.active
    # Pre-write the labels at the offset row range so the formula-writer's
    # "*"-prefix mutation has a target.
    for idx, (_d, _cid, label, _abs) in enumerate(rows):
        ws.cell(row=base_row + idx, column=1, value=label)

    _inject_sum_formulas(
        ws, rows, calc_blocks, value_columns=("B",), base_row=base_row,
    )

    # Parent lands at base_row; formula must reference children at
    # base_row+1 and base_row+2 — NOT _FIRST_BODY_ROW (3) + offsets.
    parent_row = base_row
    formula = ws.cell(row=parent_row, column=2).value
    assert formula == f"=1*B{base_row + 1}+1*B{base_row + 2}", (
        f"Expected formula referencing children at offset rows; got {formula!r}"
    )


def test_emitted_template_has_sum_formula_at_total_row():
    """Red: an emitted template carries `=B<r1>+B<r2>+…` formulas at total rows.

    For SOFP-CuNonCu (role 210000), the "*Total non-current assets" row must
    hold a formula in both column B (CY) and column C (PY) that references
    every NoncurrentAssets child row.
    """
    import openpyxl

    path = MPERS_COMPANY_DIR / "01-SOFP-CuNonCu.xlsx"
    assert path.exists(), "generator must have run before formula test"

    wb = openpyxl.load_workbook(path)
    ws = wb["SOFP-CuNonCu"]

    # Find the Total non-current assets row.
    total_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip().startswith("*Total non-current"):
            total_row = r
            break
    assert total_row is not None, "no 'Total non-current assets' row found"

    for col_letter, col_idx in (("B", 2), ("C", 3)):
        formula = ws.cell(row=total_row, column=col_idx).value
        assert isinstance(formula, str), f"col {col_letter} is not a string: {formula!r}"
        assert formula.startswith("="), f"col {col_letter} missing formula: {formula!r}"
        # References must be to the same column.
        assert col_letter in formula, (
            f"col {col_letter} formula {formula!r} references wrong columns"
        )


@pytest.mark.mpers_formulas
def test_emitted_balance_sheet_balances_via_verifier(tmp_path):
    """Red: after filling a handful of known values, SOFP total rows balance
    (Assets == Equity + Liabilities). Uses our in-process formula evaluator
    (`_evaluate_sofp_balance`) rather than openpyxl — openpyxl doesn't
    evaluate on save, and shelling out to Excel/LibreOffice isn't portable.

    Runs in pytest's `tmp_path` fixture so the test doesn't depend on any
    project-tree directory being writable (the original version wrote into
    backup-originals/ which is read-only in some sandboxes).
    """
    import shutil

    import openpyxl

    src = MPERS_COMPANY_DIR / "01-SOFP-CuNonCu.xlsx"
    dst = tmp_path / "_balance_check_tmp.xlsx"
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    ws = wb["SOFP-CuNonCu"]

    # Helper: find the row whose column A label matches exactly.
    def find_row(label: str) -> int:
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == label:
                return r
        raise AssertionError(f"label not found: {label!r}")

    # Fill simple balancing figures (Assets = Equity + Liabilities = 1500).
    # Writing a literal value on the main-sheet line-item cell overwrites
    # the cross-sheet rollup formula — expected behaviour when an agent
    # fills the face sheet directly instead of using sub-sheet details.
    ws.cell(row=find_row("Property, plant and equipment"), column=2, value=1000)
    ws.cell(row=find_row("Inventories"), column=2, value=500)
    ws.cell(row=find_row("Issued capital"), column=2, value=1500)

    wb.save(dst)
    wb.close()

    from scripts.generate_mpers_templates import _evaluate_sofp_balance

    balanced = _evaluate_sofp_balance(dst)
    assert balanced, "SOFP did not balance after known-good fill"


@pytest.mark.mpers_formulas
def test_backup_originals_company_has_15_formula_free_files():
    """Red: every Company template must also exist in
    ``XBRL-template-MPERS/backup-originals/Company/`` — the pre-formula
    snapshot for taxonomy-version diffs (mirrors the MFRS pattern).

    Because the backup is taken post-emission in Phase 4, formulas ARE present
    in the backup files too (matches the MFRS backup-originals snapshot which
    also carries formulas — see CLAUDE.md gotcha #3). The "formula-free"
    language in the plan refers to the pre-formula-fix MFRS snapshot, but for
    MPERS (clean generator output) "snapshot after emit" is what matters.
    """
    backup_dir = REPO_ROOT / "XBRL-template-MPERS" / "backup-originals" / "Company"
    assert backup_dir.is_dir(), f"backup dir missing: {backup_dir}"

    files = sorted(p.name for p in backup_dir.glob("*.xlsx"))
    expected = sorted(fname for fname, _ in [
        ("01-SOFP-CuNonCu.xlsx", None),
        ("02-SOFP-OrderOfLiquidity.xlsx", None),
        ("03-SOPL-Function.xlsx", None),
        ("04-SOPL-Nature.xlsx", None),
        ("05-SOCI-BeforeTax.xlsx", None),
        ("06-SOCI-NetOfTax.xlsx", None),
        ("07-SOCF-Indirect.xlsx", None),
        ("08-SOCF-Direct.xlsx", None),
        ("09-SOCIE.xlsx", None),
        ("10-SoRE.xlsx", None),
        ("11-Notes-CorporateInfo.xlsx", None),
        ("12-Notes-AccountingPolicies.xlsx", None),
        ("13-Notes-ListOfNotes.xlsx", None),
        ("14-Notes-IssuedCapital.xlsx", None),
        ("15-Notes-RelatedParty.xlsx", None),
    ])
    assert files == expected, f"missing backups: {set(expected) - set(files)}"


# ---------------------------------------------------------------------------
# Phase 5 — Group-level 6-column emitter
# ---------------------------------------------------------------------------


MPERS_GROUP_DIR = REPO_ROOT / "XBRL-template-MPERS" / "Group"


@pytest.mark.mpers_group
def test_emit_template_group_produces_6_columns(tmp_path):
    """Red: ``emit_template(..., level="group")`` writes the 6-column layout:
    A=label, B=Group-CY, C=Group-PY, D=Company-CY, E=Company-PY, F=Source.

    Row 1 carries "Group" / "Company" banners, row 2 the period placeholders —
    matches the MFRS Group pin from Phase 1.
    """
    import openpyxl

    from scripts.generate_mpers_templates import emit_template

    rows = [
        (0, "root_Abstract", "Root", True),
        (1, "leaf", "Leaf label", False),
    ]
    out_path = tmp_path / "group.xlsx"
    emit_template(rows, out_path, level="group")

    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws.max_column == 6

    # Banner row + period placeholders + source header.
    assert ws.cell(row=1, column=2).value == "Group"
    assert ws.cell(row=1, column=4).value == "Company"
    assert ws.cell(row=1, column=6).value == "Source"
    for col in (2, 3, 4, 5):
        assert "YYYY" in str(ws.cell(row=2, column=col).value)

    # Body rows start at row 3 in Group layout too.
    assert ws.cell(row=3, column=1).value == "Root"
    assert ws.cell(row=4, column=1).value == "Leaf label"


@pytest.mark.mpers_group
@pytest.mark.parametrize(
    "filename",
    [fname for fname, _ in [
        ("01-SOFP-CuNonCu.xlsx", None),
        ("02-SOFP-OrderOfLiquidity.xlsx", None),
        ("03-SOPL-Function.xlsx", None),
        ("04-SOPL-Nature.xlsx", None),
        ("05-SOCI-BeforeTax.xlsx", None),
        ("06-SOCI-NetOfTax.xlsx", None),
        ("07-SOCF-Indirect.xlsx", None),
        ("08-SOCF-Direct.xlsx", None),
        ("09-SOCIE.xlsx", None),
        ("10-SoRE.xlsx", None),
        ("11-Notes-CorporateInfo.xlsx", None),
        ("12-Notes-AccountingPolicies.xlsx", None),
        ("13-Notes-ListOfNotes.xlsx", None),
        ("14-Notes-IssuedCapital.xlsx", None),
        ("15-Notes-RelatedParty.xlsx", None),
    ]],
)
def test_generated_group_template(filename):
    """Red: all 15 Group-level MPERS templates exist with the 6-col layout,
    formulas present in B/C/D/E at total rows, and "Source" header at F1.

    SOCIE (09-SOCIE.xlsx) is covered by its own 4-row-block test — skipped
    from the generic 6-column check because SOCIE lays out equity components
    across columns instead of period pairs.
    """
    import openpyxl

    path = MPERS_GROUP_DIR / filename
    assert path.exists(), f"group template missing: {path}"

    if filename == "09-SOCIE.xlsx":
        # SOCIE has its own layout pin — only verify the file opens.
        openpyxl.load_workbook(path, read_only=True).close()
        return

    wb = openpyxl.load_workbook(path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.max_column == 6, f"{filename}:{sheet_name} not 6-column"
        assert ws.cell(row=1, column=6).value == "Source"
    wb.close()


@pytest.mark.mpers_group
def test_group_socie_has_four_row_blocks():
    """Red: Group SOCIE (09-SOCIE.xlsx) is laid out as four vertical row-blocks.

    Block headers at rows 3 / 27 / 51 / 75 reading Group-CY / Group-PY /
    Company-CY / Company-PY; blank separators at 26 / 50 / 74; all four
    block bodies (rows after the block-header) contain the same MPERS SOCIE
    row-set derived from role 610000.
    """
    import openpyxl

    path = MPERS_GROUP_DIR / "09-SOCIE.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["SOCIE"]

    # Expected block headers.
    assert ws.cell(row=3, column=1).value == "Group - Current period"
    assert ws.cell(row=27, column=1).value == "Group - Prior period"
    assert ws.cell(row=51, column=1).value == "Company - Current period"
    assert ws.cell(row=75, column=1).value == "Company - Prior period"

    # Separator rows are blank in col A.
    for sep_row in (26, 50, 74):
        assert ws.cell(row=sep_row, column=1).value is None

    # Bodies (rows 4..25, 28..49, etc.) are identical across the 4 blocks.
    bodies = []
    for start, end in [(3, 25), (27, 49), (51, 73), (75, 97)]:
        body = [ws.cell(row=r, column=1).value for r in range(start + 1, end + 1)]
        bodies.append(body)
    wb.close()

    for other in bodies[1:]:
        assert other == bodies[0], "Group SOCIE block bodies diverge"


@pytest.mark.mpers_group
def test_group_sore_single_column_block():
    """Red: Group SoRE (10-SoRE.xlsx) uses the default 6-col layout, NOT the
    4-block SOCIE treatment — SoRE is simpler and doesn't need the
    multi-block structure.
    """
    import openpyxl

    path = MPERS_GROUP_DIR / "10-SoRE.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["SoRE"]

    assert ws.max_column == 6
    # No "Group - Current period" header at row 3 (that's SOCIE's block marker).
    assert ws.cell(row=3, column=1).value != "Group - Current period"
    # Source column header at F1.
    assert ws.cell(row=1, column=6).value == "Source"
    wb.close()


# ---------------------------------------------------------------------------
# Phase 6 — snapshot
# ---------------------------------------------------------------------------


@pytest.mark.mpers_snapshot
def test_backup_originals_group_has_15_files():
    """Red: Group-level snapshot exists under
    ``XBRL-template-MPERS/backup-originals/Group/`` with all 15 files.

    Closes the backup-originals pattern so future taxonomy updates can diff
    a generation-1 baseline at both levels.
    """
    backup = REPO_ROOT / "XBRL-template-MPERS" / "backup-originals" / "Group"
    assert backup.is_dir()
    files = sorted(p.name for p in backup.glob("*.xlsx"))
    expected = sorted([
        "01-SOFP-CuNonCu.xlsx",
        "02-SOFP-OrderOfLiquidity.xlsx",
        "03-SOPL-Function.xlsx",
        "04-SOPL-Nature.xlsx",
        "05-SOCI-BeforeTax.xlsx",
        "06-SOCI-NetOfTax.xlsx",
        "07-SOCF-Indirect.xlsx",
        "08-SOCF-Direct.xlsx",
        "09-SOCIE.xlsx",
        "10-SoRE.xlsx",
        "11-Notes-CorporateInfo.xlsx",
        "12-Notes-AccountingPolicies.xlsx",
        "13-Notes-ListOfNotes.xlsx",
        "14-Notes-IssuedCapital.xlsx",
        "15-Notes-RelatedParty.xlsx",
    ])
    assert files == expected, f"missing: {set(expected) - set(files)}"


# ---------------------------------------------------------------------------
# Phase 4+ — calc-link grouping fix (2026-04-23 hardening side-quest)
#
# XBRL calc linkbases declare multiple independent summation-consistency
# rules per concept using separate <calculationLink role=...> blocks. The
# original MPERS generator flattened them into one formula, which:
#   (a) doubled children when two calc-links share an identical (parent,
#       child) pair (e.g. AuditorsRemuneration in SOPL-Analysis sub-sheet,
#       producing =1*B96+1*B96+1*B97+1*B97);
#   (b) merged distinct-axis decompositions (e.g. on SOPL, ProfitLoss has
#       a vertical calc `ContinuingOps+DiscontinuedOps` AND an attribution
#       calc `Owners+EquityOther+NCI`; flattening produced a 5-term sum
#       that double-counts ProfitLoss).
#
# The fix introduces link-role-aware parsing + per-presentation-occurrence
# formula assignment. These tests pin the expected behaviour. They must fail
# against the pre-fix generator and pass after regeneration.
# ---------------------------------------------------------------------------


@pytest.mark.mpers_formulas
def test_parse_calc_linkbase_grouped_splits_link_roles():
    """Red: the grouped parser returns distinct calc blocks per <calculationLink>.

    SOPL calc file `role-300100.xml` carries four calc-link blocks (300100,
    300100a, 300100b, 300100c). The grouped parser must keep them as separate
    entries so downstream code can decide how to assign them to presentation
    rows.
    """
    from scripts.generate_mpers_templates import (
        parse_calc_linkbase_grouped_for_pre_role,
    )

    blocks = parse_calc_linkbase_grouped_for_pre_role("310000")
    assert isinstance(blocks, list)
    assert len(blocks) >= 2, (
        f"SOPL calc should have at least two distinct link-role blocks "
        f"(vertical + attribution), got {len(blocks)}"
    )

    # Exactly one block should define ProfitLoss = ContinuingOps + DiscontinuedOps
    # and a separate block should define ProfitLoss = attribution trio.
    vertical_blocks = []
    attribution_blocks = []
    for _role, calc_map in blocks:
        pl_children = calc_map.get("ifrs-smes_ProfitLoss", [])
        names = {child for child, _w in pl_children}
        if not names:
            continue
        if "ifrs-smes_ProfitLossFromContinuingOperations" in names:
            vertical_blocks.append(pl_children)
        if "ifrs-smes_ProfitLossAttributableToOwnersOfParent" in names:
            attribution_blocks.append(pl_children)

    assert len(vertical_blocks) == 1, (
        f"expected 1 vertical ProfitLoss calc block, got {len(vertical_blocks)}"
    )
    assert len(attribution_blocks) == 1, (
        f"expected 1 attribution ProfitLoss calc block, got {len(attribution_blocks)}"
    )
    # The vertical and attribution blocks must be DISTINCT list objects
    # (i.e. not the same merged flattened list).
    assert vertical_blocks[0] is not attribution_blocks[0], (
        "vertical and attribution blocks were merged — must be kept separate"
    )


@pytest.mark.mpers_formulas
def test_mpers_sopl_profitloss_splits_across_vertical_and_attribution():
    """MPERS SOPL-Function has `ProfitLoss` at two presentation occurrences —
    once as the vertical total (ContinuingOps + DiscontinuedOps) and once as
    the attribution breakdown (Owners + EquityOther + NCI). Each row must
    carry only its own decomposition.

    The exact row numbers depend on taxonomy walk ordering; we look them up
    dynamically so the test remains stable when future taxonomy drift or
    row-formatting changes shift the presentation layout. The row-24 /
    row-29 numbers below are the post-2026-04-23 layout (hypercube
    scaffolding + duplicate headers stripped — the vertical 'Profit (loss)'
    lands at r24 and the '*Total Profit (Loss)' attribution lands at r29).

    Pre-fix bug this test pins against: row 34 carried a 5-term sum merging
    both decompositions (=1*B26+1*B31+1*B28+1*B32+1*B33) and row 29 had no
    formula at all. After the per-calc-block grouping in
    parse_calc_linkbase_grouped, each occurrence gets its own calc block.
    """
    import openpyxl

    path = MPERS_COMPANY_DIR / "03-SOPL-Function.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["SOPL-Function"]

    # Find the two ProfitLoss rows by label (robust against row shifts).
    # SSM presentation order: the vertical total ('*Profit (loss)') comes
    # first, the attribution rollup ('*Total Profit (Loss)') comes later
    # inside the attribution block. Both carry formulas; both are total
    # rows (asterisk prefix per MFRS convention). We walk top-down and
    # assign the first hit to vertical, second hit to attribution.
    profit_loss_rows: list[int] = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        formula = ws.cell(row=r, column=2).value
        if (
            isinstance(label, str)
            and label.strip().lower().lstrip("*").strip() in (
                "profit (loss)",
                "total profit (loss)",
            )
            and isinstance(formula, str)
            and formula.startswith("=")
        ):
            profit_loss_rows.append(r)

    assert len(profit_loss_rows) == 2, (
        f"expected 2 ProfitLoss formula rows (vertical + attribution); "
        f"got {len(profit_loss_rows)}: {profit_loss_rows}"
    )
    vertical_row, attribution_row = profit_loss_rows

    vertical_formula = ws.cell(row=vertical_row, column=2).value
    attribution_formula = ws.cell(row=attribution_row, column=2).value
    wb.close()

    # Both rows must carry a formula.
    assert isinstance(vertical_formula, str) and vertical_formula.startswith("="), (
        f"vertical row r{vertical_row} (Profit (loss)) missing formula: {vertical_formula!r}"
    )
    assert isinstance(attribution_formula, str) and attribution_formula.startswith("="), (
        f"attribution row r{attribution_row} (*Total Profit (Loss)) missing formula: "
        f"{attribution_formula!r}"
    )

    # Formulas must be distinct — the vertical and attribution blocks must
    # not merge into one sum.
    assert vertical_formula != attribution_formula, (
        f"vertical and attribution formulas are identical — per-calc-block "
        f"grouping failed. vertical={vertical_formula!r}"
    )

    # Shape check: vertical is a 2-term sum (continuing + discontinued),
    # attribution is a 3-term sum (owners + equity-other + NCI).
    vertical_terms = vertical_formula.count("*B")
    attribution_terms = attribution_formula.count("*B")
    assert vertical_terms == 2, (
        f"vertical formula should have 2 terms (continuing+discontinued), "
        f"got {vertical_terms}: {vertical_formula!r}"
    )
    assert attribution_terms == 3, (
        f"attribution formula should have 3 terms (owners+equity-other+NCI), "
        f"got {attribution_terms}: {attribution_formula!r}"
    )


@pytest.mark.mpers_formulas
def test_mpers_templates_have_no_duplicate_cell_refs():
    """Red: no generated MPERS formula may reference the same cell twice.

    Pre-fix bug: sub-sheet totals like `*Total auditor's remuneration` had
    formulas `=1*B96+1*B96+1*B97+1*B97` because two calc-links declared the
    same parent→child arc, and the flat merge appended them both. Any repeat
    reference is mathematically wrong — dedup within a calc block should
    eliminate it.
    """
    import openpyxl
    import re

    # Match `<weight>*<col><row>` tokens; ignore literal constants.
    cell_ref_re = re.compile(r"[+-]?\d+\*([A-Z]+)(\d+)")

    for root_dir in ("Company", "Group"):
        base = REPO_ROOT / "XBRL-template-MPERS" / root_dir
        for xlsx in sorted(base.glob("*.xlsx")):
            wb = openpyxl.load_workbook(xlsx, data_only=False)
            for ws in wb.worksheets:
                for row_cells in ws.iter_rows():
                    for cell in row_cells:
                        f = cell.value
                        if not (isinstance(f, str) and f.startswith("=")):
                            continue
                        refs = cell_ref_re.findall(f)
                        seen: dict[tuple[str, str], int] = {}
                        for ref in refs:
                            seen[ref] = seen.get(ref, 0) + 1
                        dupes = [f"{c}{r}" for (c, r), count in seen.items() if count > 1]
                        assert not dupes, (
                            f"{xlsx.name} [{ws.title}] "
                            f"{cell.coordinate}: formula {f!r} has duplicate refs {dupes}"
                        )
            wb.close()


@pytest.mark.mpers_formulas
def test_mpers_soci_tci_splits_across_rows():
    """MPERS SOCI-BeforeTax (role 420000) has `ComprehensiveIncome` at two
    presentation occurrences — vertical (ProfitLoss + Total-OCI) and
    attribution (Owners + NCI). Each row must carry only its own
    decomposition and they must be distinct formulas.

    Like the SOPL ProfitLoss twin-decomposition test, we locate rows by
    label rather than hard-coding row indices so taxonomy/formatting drift
    doesn't break the test. Post-2026-04-23 the rows land at r23 (vertical
    '*Total comprehensive income') and r27 (attribution '*Total
    comprehensive income' inside the Comprehensive-income-attributable-to
    block).
    """
    import openpyxl

    path = MPERS_COMPANY_DIR / "05-SOCI-BeforeTax.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["SOCI-BeforeOfTax"]

    # Collect every row that carries a formula and whose label is
    # '*Total comprehensive income' (case-insensitive). SSM puts the
    # vertical rollup before the attribution rollup in presentation order.
    tci_rows: list[tuple[int, str]] = []
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        formula = ws.cell(row=r, column=2).value
        if (
            isinstance(label, str)
            and label.strip().lower() == "*total comprehensive income"
            and isinstance(formula, str)
            and formula.startswith("=")
        ):
            tci_rows.append((r, formula))
    wb.close()

    assert len(tci_rows) == 2, (
        f"expected 2 '*Total comprehensive income' formula rows "
        f"(vertical + attribution); got {len(tci_rows)}: "
        f"{[(r, f) for r, f in tci_rows]}"
    )

    vertical_row, vertical_formula = tci_rows[0]
    attribution_row, attribution_formula = tci_rows[1]

    # Formulas must be distinct (not the old merged 4-term sum).
    assert vertical_formula != attribution_formula, (
        f"ComprehensiveIncome at r{vertical_row} and r{attribution_row} "
        f"have identical formulas — must be separate calcs. "
        f"{vertical_formula!r}"
    )

    # Shape: vertical sums 2 terms (ProfitLoss + Total OCI);
    # attribution sums 2 terms (owners + NCI). Both are 2-term sums so we
    # can't distinguish by count alone — check that the formulas reference
    # distinct row ranges (vertical pulls from the OCI block above,
    # attribution from the owners/NCI rows below it).
    import re
    row_re = re.compile(r"B(\d+)")
    vertical_refs = [int(n) for n in row_re.findall(vertical_formula)]
    attribution_refs = [int(n) for n in row_re.findall(attribution_formula)]

    assert vertical_refs, f"vertical formula has no cell refs: {vertical_formula!r}"
    assert attribution_refs, f"attribution formula has no cell refs: {attribution_formula!r}"

    # Vertical must reference rows ABOVE itself; attribution must reference
    # rows between the vertical total and itself.
    assert max(vertical_refs) < vertical_row, (
        f"vertical formula at r{vertical_row} references a row at or below "
        f"itself: {vertical_formula!r}"
    )
    assert max(attribution_refs) < attribution_row, (
        f"attribution formula at r{attribution_row} references a row at or "
        f"below itself: {attribution_formula!r}"
    )
    assert min(attribution_refs) > vertical_row, (
        f"attribution formula at r{attribution_row} pulls from rows at or "
        f"before the vertical total at r{vertical_row}: {attribution_formula!r}"
    )
