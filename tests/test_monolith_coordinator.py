"""Tests for monolith/coordinator.py — covers the bits that don't need
to drive a live LLM. The full mocked-agent run is delegated to a
manual e2e (split-pipeline parallel test pattern)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from monolith.coordinator import (
    MonolithRunConfig,
    _build_initial_user_prompt,
    _default_variant,
    _materialise_workbook,
    _snapshot_workbook,
    _statements_with_writes,
)
from monolith.prompt_renderer import render as render_monolith_prompt
from pydantic_ai.messages import BinaryContent
from statement_types import StatementType


_REPO = Path(__file__).resolve().parent.parent


def _config_company(tmp_path: Path) -> MonolithRunConfig:
    return MonolithRunConfig(
        pdf_path="",
        output_dir=str(tmp_path),
        model="stub",
        statements=set(StatementType),
        variants={},
        filing_level="company",
        filing_standard="mfrs",
    )


def test_materialise_workbook_concatenates_all_five_face_templates(tmp_path):
    config = _config_company(tmp_path)
    out = tmp_path / "monolith_filled.xlsx"
    _materialise_workbook(out, config)
    assert out.exists()
    wb = openpyxl.load_workbook(str(out), data_only=False)
    try:
        # SOCIE is the only sheet not prefix-matched — assert its presence.
        assert "SOCIE" in wb.sheetnames
        # At least one sheet per face statement type.
        for prefix in ("SOFP-", "SOPL-", "SOCI-", "SOCF-"):
            assert any(n.startswith(prefix) for n in wb.sheetnames), (
                f"expected at least one sheet matching prefix {prefix!r}"
            )
    finally:
        wb.close()


def test_materialise_workbook_is_idempotent(tmp_path):
    config = _config_company(tmp_path)
    out = tmp_path / "monolith_filled.xlsx"
    _materialise_workbook(out, config)
    mtime_first = out.stat().st_mtime_ns
    _materialise_workbook(out, config)  # second call should no-op
    assert out.stat().st_mtime_ns == mtime_first


def test_default_variant_picks_a_real_template(tmp_path):
    for stmt in (
        StatementType.SOFP,
        StatementType.SOPL,
        StatementType.SOCI,
        StatementType.SOCF,
        StatementType.SOCIE,
    ):
        variant = _default_variant(stmt, "mfrs")
        assert variant != "NotPrepared"
        assert variant


def test_statements_with_writes_empty_on_fresh_workbook(tmp_path):
    config = _config_company(tmp_path)
    out = tmp_path / "monolith_filled.xlsx"
    _materialise_workbook(out, config)
    assert _statements_with_writes(out, set(StatementType)) == []


def test_statements_with_writes_reports_after_value_lands(tmp_path):
    config = _config_company(tmp_path)
    out = tmp_path / "monolith_filled.xlsx"
    _materialise_workbook(out, config)
    # Plant a numeric value somewhere on SOFP-CuNonCu.
    wb = openpyxl.load_workbook(str(out), data_only=False)
    ws = wb["SOFP-CuNonCu"]
    # Find first data-entry row (col B empty, col A has a label).
    target_row = None
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=1).value and ws.cell(row=r, column=2).value is None:
            target_row = r
            break
    assert target_row is not None
    ws.cell(row=target_row, column=2, value=100.0)
    wb.save(str(out))
    wb.close()
    assert StatementType.SOFP.value in _statements_with_writes(
        out, set(StatementType),
    )


def test_snapshot_workbook_does_not_raise_on_missing(tmp_path):
    # Should be a no-op when the file doesn't exist.
    _snapshot_workbook(tmp_path / "does-not-exist.xlsx")


def test_initial_user_prompt_is_string_for_text_pdf(tmp_path):
    """A PDF with a real text layer gets the plain string prompt — no
    PNG payloads, no extra vision cost on every run."""
    import fitz
    pdf_path = tmp_path / "text.pdf"
    doc = fitz.open()
    try:
        page = doc.new_page()
        page.insert_text((72, 72), "Total assets 12345")
        doc.save(str(pdf_path))
    finally:
        doc.close()
    rendered = render_monolith_prompt(
        str(pdf_path), filing_standard="mfrs", filing_level="company",
    )
    assert rendered.pdf_text_empty is False
    prompt = _build_initial_user_prompt(
        rendered=rendered, pdf_path=str(pdf_path), pdf_page_count=1,
    )
    assert isinstance(prompt, str)
    assert "get_state" in prompt


def test_initial_user_prompt_preloads_vision_for_scanned_pdf(tmp_path):
    """Image-only PDF → opening user message carries one BinaryContent
    per page. Pins the 2026-05-28 scanned-PDF fix; without this, the
    agent saw an all-blank cached text block and bailed."""
    import fitz
    pdf_path = tmp_path / "scanned.pdf"
    doc = fitz.open()
    try:
        for _ in range(2):
            doc.new_page()  # no text — image-only
        doc.save(str(pdf_path))
    finally:
        doc.close()
    rendered = render_monolith_prompt(
        str(pdf_path), filing_standard="mfrs", filing_level="company",
    )
    assert rendered.pdf_text_empty is True
    prompt = _build_initial_user_prompt(
        rendered=rendered, pdf_path=str(pdf_path), pdf_page_count=2,
    )
    assert isinstance(prompt, list)
    pngs = [p for p in prompt if isinstance(p, BinaryContent)]
    assert len(pngs) == 2, "expected one BinaryContent per page"
    assert all(p.media_type == "image/png" for p in pngs)
    # Banner present so the agent knows to use the images.
    text_parts = [p for p in prompt if isinstance(p, str)]
    assert any("image" in t.lower() for t in text_parts)


def test_initial_user_prompt_preloads_only_blank_pages_on_mixed_pdf(tmp_path):
    """Peer-review MEDIUM #3: a mixed text+scan PDF must still take the
    vision-preload path for the scanned pages, not silently swallow
    the empty markers in the cached text block."""
    import fitz
    pdf_path = tmp_path / "mixed.pdf"
    doc = fitz.open()
    try:
        # Page 1: text cover. Pages 2-3: image-only.
        page = doc.new_page()
        page.insert_text((72, 72), "Cover page: ABC Sdn Bhd Annual Report")
        doc.new_page()
        doc.new_page()
        doc.save(str(pdf_path))
    finally:
        doc.close()
    rendered = render_monolith_prompt(
        str(pdf_path), filing_standard="mfrs", filing_level="company",
    )
    assert rendered.pdf_text_empty is False
    assert rendered.blank_pages == [2, 3]
    prompt = _build_initial_user_prompt(
        rendered=rendered, pdf_path=str(pdf_path), pdf_page_count=3,
    )
    assert isinstance(prompt, list)
    pngs = [p for p in prompt if isinstance(p, BinaryContent)]
    assert len(pngs) == 2, "expected one PNG per blank page only"
    text_parts = [p for p in prompt if isinstance(p, str)]
    # Page-number headers identify which pages got preloaded.
    assert any("=== page 2 ===" in t for t in text_parts)
    assert any("=== page 3 ===" in t for t in text_parts)


def test_vision_preload_respects_page_count_cap(tmp_path, monkeypatch):
    """Peer-review HIGH #2: enforce a hard page cap so a 200-page
    scanned annual report can't push tens of MB on the wire."""
    import monolith.coordinator as mc
    monkeypatch.setattr(mc, "MONOLITH_VISION_PRELOAD_MAX_PAGES", 3)
    monkeypatch.setattr(
        mc, "MONOLITH_VISION_PRELOAD_MAX_BYTES", 10 * 1024 * 1024,
    )
    import fitz
    pdf_path = tmp_path / "long_scan.pdf"
    doc = fitz.open()
    try:
        for _ in range(7):
            doc.new_page()
        doc.save(str(pdf_path))
    finally:
        doc.close()
    rendered = render_monolith_prompt(
        str(pdf_path), filing_standard="mfrs", filing_level="company",
    )
    assert rendered.pdf_text_empty is True
    prompt = mc._build_initial_user_prompt(
        rendered=rendered, pdf_path=str(pdf_path), pdf_page_count=7,
    )
    pngs = [p for p in prompt if isinstance(p, BinaryContent)]
    assert len(pngs) == 3, (
        f"expected page-count cap to limit preload to 3 pages, got {len(pngs)}"
    )
    text_parts = " ".join(p for p in prompt if isinstance(p, str))
    # Banner must surface the dropped pages so the agent knows to
    # fetch them mid-run via view_pdf_pages.
    assert "4-7" in text_parts or "4, 5, 6, 7" in text_parts
    assert "view_pdf_pages" in text_parts


def test_vision_preload_respects_byte_cap(tmp_path, monkeypatch):
    """A page-count budget alone isn't enough — a few high-DPI pages
    can still blow the wire budget. Set a tiny byte cap and verify
    pages get dropped once it's exhausted."""
    import monolith.coordinator as mc
    monkeypatch.setattr(mc, "MONOLITH_VISION_PRELOAD_MAX_PAGES", 1000)
    monkeypatch.setattr(mc, "MONOLITH_VISION_PRELOAD_MAX_BYTES", 1024)
    import fitz
    pdf_path = tmp_path / "bytecap.pdf"
    doc = fitz.open()
    try:
        for _ in range(5):
            doc.new_page()
        doc.save(str(pdf_path))
    finally:
        doc.close()
    rendered = render_monolith_prompt(
        str(pdf_path), filing_standard="mfrs", filing_level="company",
    )
    prompt = mc._build_initial_user_prompt(
        rendered=rendered, pdf_path=str(pdf_path), pdf_page_count=5,
    )
    # Either the byte budget yields fewer than 5 PNGs, or the helper
    # falls back to the plain string prompt. Both prove the cap fires;
    # what we forbid is "byte cap silently ignored, 5 PNGs attached".
    if isinstance(prompt, list):
        pngs = [p for p in prompt if isinstance(p, BinaryContent)]
        assert len(pngs) < 5, (
            f"byte cap not enforced; attached {len(pngs)} PNGs"
        )
    else:
        assert isinstance(prompt, str)


def test_vision_preload_uses_lower_dpi_than_view_pdf_pages(tmp_path):
    """Peer-review HIGH #2: opening-turn preload uses 150 DPI; the
    interactive `view_pdf_pages` tool stays at 200 DPI. The cost
    saving comes from the preload, not from degrading on-demand
    vision (where the agent is asking for that specific page)."""
    from monolith.config import MONOLITH_VISION_PRELOAD_DPI
    from tools.pdf_viewer import render_pages_to_png_bytes
    assert MONOLITH_VISION_PRELOAD_DPI < 200, (
        "preload DPI should be lower than view_pdf_pages default to "
        "amortise the front-loaded vision cost"
    )
    # Sanity check the tool's own default hasn't drifted.
    assert (
        render_pages_to_png_bytes.__defaults__[-1] == 200
    ), "view_pdf_pages dpi default expected to stay 200 — update test if intentional"
