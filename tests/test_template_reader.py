from pathlib import Path

from tools.template_reader import read_template

# Anchor on the current MBRS Company SOFP template. The legacy root-level
# `SOFP-Xbrl-template.xlsx` has been removed (see CLAUDE.md §3 / §12).
TEMPLATE = str(
    Path(__file__).resolve().parent.parent
    / "XBRL-template-MFRS"
    / "Company"
    / "01-SOFP-CuNonCu.xlsx"
)


def test_read_template_returns_fields():
    fields = read_template(TEMPLATE)
    assert len(fields) > 0


def test_main_sheet_field_count():
    fields = read_template(TEMPLATE, sheet="SOFP-CuNonCu")
    assert len(fields) >= 70


def test_sub_sheet_field_count():
    fields = read_template(TEMPLATE, sheet="SOFP-Sub-CuNonCu")
    assert len(fields) >= 400


def test_detects_formula_cells():
    fields = read_template(TEMPLATE, sheet="SOFP-CuNonCu")
    formula_fields = [f for f in fields if f.has_formula]
    assert len(formula_fields) > 0


def test_detects_data_entry_cells():
    fields = read_template(TEMPLATE, sheet="SOFP-CuNonCu")
    data_fields = [f for f in fields if not f.has_formula]
    assert len(data_fields) > 0


def test_field_has_coordinate():
    fields = read_template(TEMPLATE, sheet="SOFP-CuNonCu")
    for f in fields:
        assert f.coordinate


def test_field_has_sheet_name():
    fields = read_template(TEMPLATE)
    for f in fields:
        assert f.sheet in ("SOFP-CuNonCu", "SOFP-Sub-CuNonCu")


def test_field_has_value():
    fields = read_template(TEMPLATE, sheet="SOFP-CuNonCu")
    # B1 holds the period header "01/01/YYYY - 31/12/YYYY"
    b1 = [f for f in fields if f.coordinate == "B1"]
    assert len(b1) == 1
    assert b1[0].value is not None


def test_formula_formula_not_value():
    fields = read_template(TEMPLATE, sheet="SOFP-CuNonCu")
    # First data-entry row on the main sheet pulls from the sub-sheet via a
    # cross-sheet formula. Look for *any* such formula rather than a fixed row,
    # since template row numbers have shifted historically.
    cross_sheet = [
        f for f in fields
        if f.has_formula and f.formula and "SOFP-Sub-CuNonCu" in f.formula
    ]
    assert cross_sheet, "expected at least one cross-sheet formula on main sheet"


def test_all_sheets_by_default():
    fields = read_template(TEMPLATE)
    sheets = set(f.sheet for f in fields)
    assert sheets == {"SOFP-CuNonCu", "SOFP-Sub-CuNonCu"}


# ---------------------------------------------------------------------------
# Bug A (2026-04-26): abstract-row marking
#
# XBRL section-header rows ("Interest income", "Other fee and commission
# income", etc.) carry no formula and no value — `is_data_entry` was
# previously True for them, which let agents write residual numbers onto the
# headers instead of the leaves below. The reader now marks them so the
# template summary the agent sees can label them [ABSTRACT].
# ---------------------------------------------------------------------------

# Anchor on the SOPL-Analysis-Function template — it carries the cleanest
# example of header-vs-leaf-vs-total distinctions (see screenshot bug
# 2026-04-26 on FINCO 2023). Row 27 is the dark-navy "Interest income"
# header; row 28-29 are leaves; row 30 is the formula-driven total.
_SOPL_FUNCTION_TEMPLATE = str(
    Path(__file__).resolve().parent.parent
    / "XBRL-template-MFRS"
    / "Company"
    / "03-SOPL-Function.xlsx"
)


def test_abstract_rows_marked_in_sopl_analysis():
    fields = read_template(_SOPL_FUNCTION_TEMPLATE, sheet="SOPL-Analysis-Function")

    # Row 27, col A: the "Interest income" abstract section header.
    a27 = next(
        (f for f in fields if f.coordinate == "A27"), None,
    )
    assert a27 is not None and a27.value == "Interest income"
    assert a27.is_abstract is True, (
        "row 27 ('Interest income') is a dark-navy abstract header — "
        "must be marked is_abstract=True to keep agents from writing to it"
    )

    # Row 28, col A: a leaf line item — must NOT be flagged abstract.
    a28 = next(
        (f for f in fields if f.coordinate == "A28"), None,
    )
    assert a28 is not None
    assert a28.is_abstract is False, (
        "row 28 ('Interest income on loans...') is a leaf, not an abstract "
        "header — agents must be free to write its B/C/D cells"
    )

    # Row 30: the formula-driven 'Total interest income' row. Coloured a
    # different shade (pale blue, in `_TOTAL_FILL_RGB`); not a section
    # header, must not be marked abstract.
    a30 = next(
        (f for f in fields if f.coordinate == "A30"), None,
    )
    assert a30 is not None
    assert a30.is_abstract is False, (
        "row 30 ('Total interest income') is a Total row, not a section "
        "header — the formula-cell guard is what protects it from writes"
    )


_MPERS_GROUP_SOPL_TEMPLATE = str(
    Path(__file__).resolve().parent.parent
    / "XBRL-template-MPERS"
    / "Group"
    / "03-SOPL-Function.xlsx"
)


def test_mpers_templates_carry_header_fills_like_mfrs():
    """Parity pin (2026-04-26): MPERS templates must paint the same dark-
    navy `1F3864` header fill that MFRS uses, so the abstract-row guard
    in fill_workbook works identically across both standards.

    Replaces the prior `test_mpers_templates_lack_header_fills_known_gap`
    which pinned the OBSERVED missing-fill state until the generator was
    fixed. Spot-check on MPERS Group SOPL — if this regresses, the
    generator's `_apply_*_sheet_layout` helpers have stopped painting
    abstract rows.
    """
    import openpyxl
    wb = openpyxl.load_workbook(_MPERS_GROUP_SOPL_TEMPLATE)
    header_rgbs = {"FF1F3864", "001F3864"}
    fills_found = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(1, ws.max_row + 1):
            c = ws.cell(row=r, column=1)
            if c.value is None:
                continue
            if c.fill and c.fill.fgColor and c.fill.fgColor.rgb:
                rgb = c.fill.fgColor.rgb
                if rgb in header_rgbs:
                    fills_found.add(rgb)
    wb.close()
    assert fills_found, (
        "MPERS Group SOPL template carries no dark-navy header fills — "
        "the generator's _apply_*_sheet_layout helpers must paint "
        f"{header_rgbs} on rows where _is_abstract=True. See CLAUDE.md "
        "gotcha #17."
    )


def test_abstract_rows_marked_in_mpers_group_sopl():
    """Positive form (2026-04-26): once the MPERS generator paints
    abstract rows, the reader's is_abstract flag must light up the same
    way it does on MFRS templates."""
    fields = read_template(_MPERS_GROUP_SOPL_TEMPLATE, sheet="SOPL-Analysis-Function")
    abstract_rows = [
        f for f in fields if getattr(f, "is_abstract", False) and f.col == 1
    ]
    assert abstract_rows, (
        "MPERS Group SOPL-Analysis-Function should surface abstract "
        "section headers (e.g. 'Revenue', 'Cost of sales', 'Other "
        "expenses'). If none are detected, either the template generator "
        "has stopped painting fills or the reader's discovery has regressed."
    )
    # Spot-check: at least one common SOPL-Analysis abstract concept must
    # appear in the abstract set. We don't pin the exact list because the
    # SSM linkbase can extend it; the rule is "the section-name labels
    # should be marked".
    abstract_labels = {(f.value or "").strip().lower() for f in abstract_rows}
    expected_any = {"revenue", "cost of sales", "other expenses", "interest income"}
    assert abstract_labels & expected_any, (
        f"none of the expected SOPL-Analysis section-name labels "
        f"({expected_any}) were marked abstract. Got: {abstract_labels}"
    )


def test_abstract_only_set_on_column_a_label_cells():
    """The abstract flag is a row-level property surfaced on the col-A label
    cell. Cells in cols B/C/D never carry data on header rows (they're None
    in the template), so they don't appear in the field list — there's
    nothing to test for them. This keeps the flag's semantics tight.
    """
    fields = read_template(_SOPL_FUNCTION_TEMPLATE, sheet="SOPL-Analysis-Function")
    for f in fields:
        if f.is_abstract:
            assert f.col == 1, (
                f"is_abstract should only be set on col-A label cells; "
                f"got {f.coordinate} (col {f.col})"
            )
