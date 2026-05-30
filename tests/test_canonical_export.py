"""Phase 1 steps 1.11-1.14 — DB-backed Excel export.

The canonical exporter reads facts from ``run_concept_facts`` and
writes them into a fresh copy of the template ``.xlsx``.  It must:

* read exclusively from DB facts (no agent xlsx writes in the loop);
* use the concept's canonical_label as column A (display_label
  overrides are UI-only per PRD §9);
* on aggregate_only parents, replace the live formula with the
  literal value and annotate the source column;
* on not_disclosed leaves, leave the cell blank.
"""
from __future__ import annotations

import json
import shutil
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from concept_model.exporter import export_run_to_xlsx
from concept_model.importer import import_company_targets, import_template
from concept_model.parser import parse_template
from db.schema import init_db


REPO = Path(__file__).resolve().parent.parent
FIXTURE = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"


@pytest.fixture
def seeded(tmp_path: Path) -> tuple[Path, int, str, Path]:
    """Initialise a v4 DB with a parsed SOFP template, a run row, and a
    working copy of the template xlsx that the exporter will fill.

    Returns (db_path, run_id, template_id, xlsx_copy_path).
    """
    db = tmp_path / "xbrl.db"
    init_db(db)

    tree = parse_template(str(FIXTURE))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    template_id = import_template(db, jp)
    import_company_targets(db, template_id)

    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
            "VALUES (?, ?, ?, ?)",
            ("2026-05-21T00:00:00Z", "x.pdf", "running",
             "2026-05-21T00:00:00Z"),
        )
        run_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    work = tmp_path / "filled.xlsx"
    shutil.copyfile(FIXTURE, work)
    return db, run_id, template_id, work


def _uuid_for_row(db: Path, sheet: str, row: int) -> str:
    conn = sqlite3.connect(str(db))
    try:
        r = conn.execute(
            "SELECT concept_uuid FROM concept_nodes "
            "WHERE render_sheet = ? AND render_row = ?",
            (sheet, row),
        ).fetchone()
    finally:
        conn.close()
    assert r is not None, f"no concept at {sheet}!{row}"
    return r[0]


def _seed_fact(db: Path, run_id: int, concept_uuid: str, *,
               value=None, value_status="observed",
               children_status: str | None = None,
               source: str = "pdf p.1") -> None:
    conn = sqlite3.connect(str(db))
    try:
        conn.execute(
            "INSERT OR REPLACE INTO run_concept_facts("
            "run_id, concept_uuid, period, entity_scope, value, "
            "value_status, children_status, source, updated_at) "
            "VALUES (?, ?, 'CY', 'Company', ?, ?, ?, ?, '2026-05-21Z')",
            (run_id, concept_uuid, value, value_status, children_status,
             source),
        )
        conn.commit()
    finally:
        conn.close()


def test_carry_forward_row1_dates_from_scratch(seeded) -> None:
    """Phase 4.1: row-1 reporting-period dates are non-concept cells that don't
    project to facts, so the fact-render keeps the template placeholder
    '01/01/YYYY - 31/12/YYYY'. When the agent's scratch workbook is supplied,
    its real row-1 dates must be carried over — value columns only, and only
    when the scratch holds a real (non-placeholder) date."""
    db, run_id, template_id, work = seeded
    leaf = _uuid_for_row(db, "SOFP-CuNonCu", 10)
    _seed_fact(db, run_id, leaf, value=123.0)

    # Build a scratch workbook from the template, stamping real dates into B1/C1
    # on the face sheet, and leave the sub-sheet's placeholder untouched.
    scratch = work.parent / "SOFP_filled.xlsx"
    shutil.copyfile(FIXTURE, scratch)
    swb = openpyxl.load_workbook(str(scratch), data_only=False)
    swb["SOFP-CuNonCu"]["B1"] = "01/01/2021 - 31/12/2021"
    swb["SOFP-CuNonCu"]["C1"] = "01/01/2020 - 31/12/2020"
    swb["SOFP-CuNonCu"]["A1"] = "SHOULD NOT BE CARRIED"  # col A is never carried
    swb.save(str(scratch))

    export_run_to_xlsx(db, run_id, str(work), carry_forward_row1_from=str(scratch))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    face = wb["SOFP-CuNonCu"]
    # Real dates carried into the value columns.
    assert face["B1"].value == "01/01/2021 - 31/12/2021"
    assert face["C1"].value == "01/01/2020 - 31/12/2020"
    # Col A is not a date placeholder — never carried (stays template/label).
    assert face["A1"].value != "SHOULD NOT BE CARRIED"
    # The fact still landed (carry-forward doesn't disturb values).
    assert face["B10"].value == 123.0
    # The sub-sheet's placeholder stays placeholder (scratch had no real date).
    assert wb["SOFP-Sub-CuNonCu"]["B1"].value == "01/01/YYYY - 31/12/YYYY"


def test_carry_forward_is_placeholder_keyed_not_row1_bound(seeded) -> None:
    """Group templates hold the Group/Company labels in row 1 and the date
    placeholders in row 2 (not row 1). The carry-forward keys on the literal
    'YYYY' placeholder, not a hardcoded row, so it must carry a real date that
    lives below row 1. Simulate by stamping a placeholder into row 2 of the
    template copy and a real date into the scratch's row 2."""
    db, run_id, template_id, work = seeded
    _seed_fact(db, run_id, _uuid_for_row(db, "SOFP-CuNonCu", 10), value=1.0)

    # Canonical copy: put a date placeholder on row 2 (mimics Group layout).
    cwb = openpyxl.load_workbook(str(work), data_only=False)
    cwb["SOFP-CuNonCu"]["B2"] = "01/01/YYYY - 31/12/YYYY"
    cwb.save(str(work))

    # Scratch: real date in the matching row-2 cell.
    scratch = work.parent / "SOFP_filled.xlsx"
    shutil.copyfile(str(work), scratch)
    swb = openpyxl.load_workbook(str(scratch), data_only=False)
    swb["SOFP-CuNonCu"]["B2"] = "01/01/2021 - 31/12/2021"
    swb.save(str(scratch))

    export_run_to_xlsx(db, run_id, str(work), carry_forward_row1_from=str(scratch))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    # Carried even though it's row 2, because it was a "YYYY" placeholder.
    assert wb["SOFP-CuNonCu"]["B2"].value == "01/01/2021 - 31/12/2021"


def test_carry_forward_row1_no_scratch_keeps_placeholder(seeded) -> None:
    """Without a scratch path the export is unchanged — row 1 keeps the
    template placeholder (graceful degradation, no crash)."""
    db, run_id, template_id, work = seeded
    _seed_fact(db, run_id, _uuid_for_row(db, "SOFP-CuNonCu", 10), value=1.0)
    export_run_to_xlsx(db, run_id, str(work))  # no carry_forward_row1_from
    wb = openpyxl.load_workbook(str(work), data_only=False)
    assert wb["SOFP-CuNonCu"]["B1"].value == "01/01/YYYY - 31/12/YYYY"


def test_export_reads_from_db_not_agent_writes(seeded) -> None:
    """Seed two leaf facts only; no other writes happen.  Export must
    still produce non-empty cells at those leaves."""
    db, run_id, template_id, work = seeded
    biological = _uuid_for_row(db, "SOFP-CuNonCu", 10)
    cash_n_equiv_sub = _uuid_for_row(db, "SOFP-Sub-CuNonCu", 193)

    _seed_fact(db, run_id, biological, value=123.0)
    _seed_fact(db, run_id, cash_n_equiv_sub, value=999.0)

    export_run_to_xlsx(db, run_id, str(work))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    assert wb["SOFP-CuNonCu"]["B10"].value == 123.0
    assert wb["SOFP-Sub-CuNonCu"]["B193"].value == 999.0


def test_export_writes_canonical_label_not_display_override(seeded) -> None:
    """Even if the UI override changed the display label, column A
    must keep the canonical SSM label."""
    db, run_id, template_id, work = seeded
    biological = _uuid_for_row(db, "SOFP-CuNonCu", 10)
    conn = sqlite3.connect(str(db))
    try:
        conn.execute(
            "UPDATE concept_nodes SET display_label = ? "
            "WHERE concept_uuid = ?",
            ("Cows and pigs", biological),
        )
        conn.commit()
    finally:
        conn.close()

    _seed_fact(db, run_id, biological, value=10.0)
    export_run_to_xlsx(db, run_id, str(work))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    label = wb["SOFP-CuNonCu"]["A10"].value
    assert "Cows" not in (label or ""), (
        f"display label leaked into export: {label!r}"
    )


def test_aggregate_only_replaces_parent_formula_with_literal(seeded) -> None:
    """Marking a COMPUTED parent aggregate_only must replace the live
    formula with the literal value; child cells stay empty."""
    db, run_id, template_id, work = seeded
    parent = _uuid_for_row(db, "SOFP-Sub-CuNonCu", 39)
    _seed_fact(db, run_id, parent, value=88888.0,
               value_status="user_override",
               children_status="aggregate_only",
               source="pdf p.28; underlying breakdown not disclosed")

    export_run_to_xlsx(db, run_id, str(work))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    cell = wb["SOFP-Sub-CuNonCu"]["B39"]
    assert cell.value == 88888.0
    # The source-column annotation lands somewhere on the same row
    # (col D for Company templates).  Pin the substring rather than
    # the exact column number.
    source_col_text = wb["SOFP-Sub-CuNonCu"]["D39"].value or ""
    assert "aggregate_only" in source_col_text


def test_py_facts_export_to_col_c_on_linear_company(seeded) -> None:
    """PY/Company facts on a linear Company filing must render in
    col C, mirroring cell_resolver.resolve_cell's CY=B / PY=C
    convention. Regression: pre-fix the exporter dropped every
    non-(CY, Company) fact, so the downloaded canonical xlsx had
    empty PY columns even when run_concept_facts carried the data.
    """
    db, run_id, _template_id, work = seeded
    leaf = _uuid_for_row(db, "SOFP-CuNonCu", 10)

    conn = sqlite3.connect(str(db))
    try:
        conn.executemany(
            "INSERT OR REPLACE INTO run_concept_facts("
            "run_id, concept_uuid, period, entity_scope, value, "
            "value_status, source, updated_at) "
            "VALUES (?, ?, ?, 'Company', ?, 'observed', 'pdf p.1', "
            "'2026-05-21Z')",
            [
                (run_id, leaf, "CY", 111.0),
                (run_id, leaf, "PY", 222.0),
            ],
        )
        conn.commit()
    finally:
        conn.close()

    export_run_to_xlsx(db, run_id, str(work))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    assert wb["SOFP-CuNonCu"]["B10"].value == 111.0
    assert wb["SOFP-CuNonCu"]["C10"].value == 222.0


def test_group_facts_dropped_on_company_filing(seeded) -> None:
    """A Group-scope fact on a Company filing has no place to land
    (Group columns D/E only exist on Group templates), so the linear-
    Company exporter drops it. Pinning so a future widening of the
    routing rule can't silently leak Group values into Company B/C.
    """
    db, run_id, _template_id, work = seeded
    leaf = _uuid_for_row(db, "SOFP-CuNonCu", 10)

    conn = sqlite3.connect(str(db))
    try:
        conn.execute(
            "INSERT OR REPLACE INTO run_concept_facts("
            "run_id, concept_uuid, period, entity_scope, value, "
            "value_status, source, updated_at) "
            "VALUES (?, ?, 'CY', 'Group', 999.0, 'observed', "
            "'pdf p.1', '2026-05-21Z')",
            (run_id, leaf),
        )
        conn.commit()
    finally:
        conn.close()

    export_run_to_xlsx(db, run_id, str(work))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    # Col B (CY) and C (PY) must be untouched on a Company filing
    # when the only fact is Group-scope.
    assert wb["SOFP-CuNonCu"]["B10"].value != 999.0
    assert wb["SOFP-CuNonCu"]["C10"].value != 999.0


def test_not_disclosed_leaves_remain_blank_in_excel(seeded) -> None:
    """not_disclosed leaves render as blank cells; a side-channel JSON
    documents which leaves were intentionally blank."""
    db, run_id, template_id, work = seeded
    leaf = _uuid_for_row(db, "SOFP-CuNonCu", 10)
    _seed_fact(db, run_id, leaf, value=None, value_status="not_disclosed")

    export_run_to_xlsx(db, run_id, str(work))

    wb = openpyxl.load_workbook(str(work), data_only=False)
    assert wb["SOFP-CuNonCu"]["B10"].value is None

    # Side-channel JSON next to the xlsx.
    side = Path(str(work) + ".not_disclosed.json")
    assert side.is_file()
    payload = json.loads(side.read_text())
    assert any(
        e.get("concept_uuid") == leaf for e in payload.get("entries", [])
    )
