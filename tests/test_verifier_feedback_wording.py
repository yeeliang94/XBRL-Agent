"""Bug B (2026-04-26) — verifier imbalance feedback must not push the agent
into a catch-all plug.

Old feedback wording was directive: "Action: equity+liabilities section is
too low. Re-examine liabilities or equity sub-items." Combined with the
hard save-gate that refused finalisation while imbalanced, this told the
agent to "fix one side" — and the path of least resistance is to plug a
residual into a catch-all row.

The new wording keeps the diagnostic detail (which side is high/low, which
sub-items to re-examine) but explicitly tells the agent that finishing
with an imbalance is acceptable and that plugging a catch-all is not.

Peer-review #2 + #3 (2026-04-26): the no-plug guidance must apply to ALL
imbalance branches — CY, PY, and (on Group filings) the Company-CY /
Company-PY columns. The original implementation only updated CY, leaving
the agent with no anti-plug guard on the other three branches. These
parametrised tests pin all four.
"""
from __future__ import annotations

import openpyxl
import pytest

from statement_types import StatementType
from tools.verifier import verify_statement, verify_totals


def _no_plug_clause_present(feedback: str) -> bool:
    f = feedback.lower()
    return any(s in f for s in (
        "do not plug",
        "never plug",
        "do not use a catch-all",
        "do not invent a residual",
    ))


def _leave_gap_clause_present(feedback: str) -> bool:
    f = feedback.lower()
    return any(s in f for s in ("leave", "honest", "finish"))


def _make_sofp_with_imbalance(
    tmp_path,
    *,
    cy_imbalanced: bool,
    py_imbalanced: bool,
    company_cy_imbalanced: bool = False,
    company_py_imbalanced: bool = False,
) -> str:
    """Build a SOFP workbook with imbalances on selected period branches.

    Group columns (D=Company CY, E=Company PY) are populated only when the
    matching `company_*_imbalanced` flag is set so company-only tests don't
    accidentally trip the standalone CY/PY checks.
    """
    path = tmp_path / "imbalanced_sofp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Total assets"
    ws["A5"] = "Total equity and liabilities"

    # Group + Company CY (cols B / D), prior year (cols C / E).
    ws["B1"] = 450 if cy_imbalanced else 400
    ws["B5"] = 400
    ws["C1"] = 700 if py_imbalanced else 600
    ws["C5"] = 600

    if company_cy_imbalanced or company_py_imbalanced:
        ws["D1"] = 350 if company_cy_imbalanced else 300
        ws["D5"] = 300
        ws["E1"] = 525 if company_py_imbalanced else 500
        ws["E5"] = 500

    wb.save(str(path))
    return str(path)


def test_imbalance_direction_wording_for_negative_diff(tmp_path):
    """Peer-review 2026-04-26: when assets < equity+liabilities (diff < 0),
    the diagnostic must say 'assets section is lower'. The legacy phrasing
    said 'assets is higher', which inverts which side the agent should
    re-examine and effectively points it at the wrong notes.
    """
    path = tmp_path / "assets_lower.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Total assets"
    ws["B1"] = 400  # lower
    ws["A5"] = "Total equity and liabilities"
    ws["B5"] = 450  # higher → diff < 0
    wb.save(str(path))

    result = verify_totals(str(path))
    feedback = (result.feedback or "").lower()
    assert "assets section is lower" in feedback, (
        f"diff < 0 must say 'assets section is lower'. Got: {result.feedback!r}"
    )
    # Sanity: the wrong phrasing must not appear.
    assert "assets section is higher" not in feedback


def test_imbalance_direction_wording_for_positive_diff(tmp_path):
    """Mirror of the negative-diff test — when diff > 0, the diagnostic
    must point at equity+liabilities being lower (the side the agent
    should re-examine for a missing component)."""
    path = tmp_path / "equity_lower.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Total assets"
    ws["B1"] = 500  # higher
    ws["A5"] = "Total equity and liabilities"
    ws["B5"] = 400  # lower → diff > 0
    wb.save(str(path))

    result = verify_totals(str(path))
    feedback = (result.feedback or "").lower()
    assert "equity+liabilities section is lower" in feedback, (
        f"diff > 0 must say 'equity+liabilities section is lower'. "
        f"Got: {result.feedback!r}"
    )
    assert "equity+liabilities section is higher" not in feedback


@pytest.mark.parametrize("scenario", [
    pytest.param(
        {"cy_imbalanced": True, "py_imbalanced": False, "filing_level": "company"},
        id="cy_only",
    ),
    pytest.param(
        {"cy_imbalanced": False, "py_imbalanced": True, "filing_level": "company"},
        id="py_only",
    ),
    pytest.param(
        {"cy_imbalanced": True, "py_imbalanced": True, "filing_level": "company"},
        id="cy_and_py",
    ),
    pytest.param(
        {
            "cy_imbalanced": False, "py_imbalanced": False,
            "company_cy_imbalanced": True, "company_py_imbalanced": False,
            "filing_level": "group",
        },
        id="group_company_cy_only",
    ),
    pytest.param(
        {
            "cy_imbalanced": False, "py_imbalanced": False,
            "company_cy_imbalanced": False, "company_py_imbalanced": True,
            "filing_level": "group",
        },
        id="group_company_py_only",
    ),
])
def test_imbalance_feedback_is_non_directive_on_every_branch(scenario, tmp_path):
    """Pin: any imbalance — CY, PY, Company CY, or Company PY — must produce
    feedback that (a) names the imbalance, (b) forbids plugging, and
    (c) tells the agent it's okay to finish with the gap."""
    filing_level = scenario.pop("filing_level")
    path = _make_sofp_with_imbalance(tmp_path, **scenario)

    result = verify_totals(path, filing_level=filing_level)

    assert result.is_balanced is False, "test setup error — should be unbalanced"
    feedback = result.feedback or ""

    assert "imbalance" in feedback.lower(), (
        f"feedback must report the imbalance diagnostically. Got: {feedback!r}"
    )
    assert _no_plug_clause_present(feedback), (
        f"every imbalance branch must carry the no-plug clause. "
        f"Got: {feedback!r}"
    )
    assert _leave_gap_clause_present(feedback), (
        f"every imbalance branch must invite the agent to finish honestly. "
        f"Got: {feedback!r}"
    )


# ---------------------------------------------------------------------------
# Phase 7.1 (2026-06-02) — directional sign-error diagnostics for the
# non-SOFP verifiers (SOPL / SOCI / SOCF / SOCIE). When an arithmetic
# identity fails and the gap is ~2x a single component magnitude, that
# component was almost certainly entered with the wrong sign; the feedback
# must name that row WITHOUT weakening the no-plug guard.
# ---------------------------------------------------------------------------


def test_soci_oci_sign_error_is_diagnosed_directionally(tmp_path):
    """An OCI value keyed with the wrong sign makes Total CI miss P&L+OCI by
    exactly 2x the OCI magnitude. The feedback must flag a sign error and
    name the OCI row — and still forbid plugging."""
    path = tmp_path / "soci_oci_sign.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCI-NetOfTax"
    # P&L 1000, OCI should be -200 (a loss) but was keyed +200; correct TCI 800.
    ws["A1"] = "Profit (loss)"
    ws["B1"] = 1000
    ws["A2"] = "Total other comprehensive income"
    ws["B2"] = 200
    ws["A3"] = "Total comprehensive income"
    ws["B3"] = 800
    wb.save(str(path))

    result = verify_statement(
        str(path), StatementType.SOCI, variant="NetOfTax", filing_level="company",
    )
    assert result.is_balanced is False, "test setup error — should be imbalanced"
    feedback = (result.feedback or "").lower()
    assert "diagnostic" in feedback, f"no directional diagnostic. Got: {result.feedback!r}"
    assert "sign error" in feedback, f"diagnostic must name a sign error. Got: {result.feedback!r}"
    assert "other comprehensive income" in feedback, (
        f"diagnostic must name the suspect OCI row. Got: {result.feedback!r}"
    )
    # The no-plug guard must survive the new directional wording (gotcha #17).
    assert _no_plug_clause_present(feedback), (
        f"directional feedback must keep the no-plug clause. Got: {result.feedback!r}"
    )


def test_socf_financing_sign_error_is_diagnosed_directionally(tmp_path):
    """A financing outflow keyed positive shifts Net-increase-before-FX by 2x
    its magnitude; the feedback must name the financing row as the suspect."""
    path = tmp_path / "socf_fin_sign.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCF-Indirect"
    ws["A1"] = "Net cash flows from operating activities"
    ws["B1"] = 500
    ws["A2"] = "Net cash flows from investing activities"
    ws["B2"] = -200
    ws["A3"] = "Net cash flows from financing activities"
    ws["B3"] = 300  # should be -300; net before FX is therefore 0
    ws["A4"] = "Net increase (decrease) in cash before foreign exchange"
    ws["B4"] = 0
    wb.save(str(path))

    result = verify_statement(
        str(path), StatementType.SOCF, variant="Indirect", filing_level="company",
    )
    assert result.is_balanced is False, "test setup error — should be imbalanced"
    feedback = (result.feedback or "").lower()
    assert "diagnostic" in feedback, f"no directional diagnostic. Got: {result.feedback!r}"
    assert "sign error" in feedback
    assert "financing" in feedback, (
        f"diagnostic must name the suspect financing row. Got: {result.feedback!r}"
    )
    assert _no_plug_clause_present(feedback)


def test_non_sofp_diagnostic_silent_when_gap_matches_no_component(tmp_path):
    """The 2x heuristic must not fire spuriously: an imbalance whose gap does
    NOT match 2x any single component carries the raw mismatch + no-plug
    footer but no 'sign error' diagnostic (avoids misleading the agent)."""
    path = tmp_path / "soci_no_match.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCI-NetOfTax"
    # gap = 800 - (1000 + 50) = -250, which is not 2x of 1000 or 50.
    ws["A1"] = "Profit (loss)"
    ws["B1"] = 1000
    ws["A2"] = "Total other comprehensive income"
    ws["B2"] = 50
    ws["A3"] = "Total comprehensive income"
    ws["B3"] = 800
    wb.save(str(path))

    result = verify_statement(
        str(path), StatementType.SOCI, variant="NetOfTax", filing_level="company",
    )
    assert result.is_balanced is False
    feedback = (result.feedback or "").lower()
    assert "sign error" not in feedback, (
        f"2x heuristic fired on a non-matching gap. Got: {result.feedback!r}"
    )
    # But the no-plug guard is still present on every imbalance.
    assert _no_plug_clause_present(feedback)
