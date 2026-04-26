"""RUN-REVIEW P0-3 (2026-04-26): merge-time formula recalc.

Pre-P0-3 the merged workbook shipped with `*Total …` formulas that
openpyxl cannot evaluate without Excel/LibreOffice. With
``data_only=True`` they returned None, breaking downstream readers.
This test pins the recalc helper end-to-end:

1. The MFRS Company SOFP fixture has 18 leaves filled but every
   `*Total …` row reads None pre-recalc.
2. After ``recalc_workbook``, every Total row reads its computed sum.
3. Sheet name casing is preserved (the formulas package uppercases
   sheet names internally; the helper writes back to the original).
4. The MPERS Group fixture exercises the 6-column / cross-sheet
   rollup path so MPERS Group runs aren't silently regressed.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from tools.recalc import recalc_workbook

_FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "run_review"


def test_recalc_populates_total_rows_mfrs_company(tmp_path: Path) -> None:
    """The 18 leaves we pre-fill in `sofp_company_mfrs.xlsx` cascade up
    through *Total inventories, *Total cash, *Total issued capital,
    *Total PPE, etc. Pre-recalc all of these read None; post-recalc
    they read the expected sums."""
    src = _FIXTURE_DIR / "sofp_company_mfrs.xlsx"
    dst = tmp_path / "merged.xlsx"
    shutil.copy(src, dst)

    # Pre-recalc sanity check: Total rows are None.
    pre = load_workbook(dst, data_only=True)
    pre_ws = pre["SOFP-Sub-CuNonCu"]
    for row in (39, 136, 193, 203):
        assert pre_ws.cell(row, 2).value is None, (
            f"Test prerequisite broken: row {row} CY should be None pre-recalc"
        )

    recalc_workbook(dst)

    post = load_workbook(dst, data_only=True)
    post_ws = post["SOFP-Sub-CuNonCu"]
    # *Total inventories — only "Other inventories" was filled at 159,389
    assert post_ws.cell(136, 2).value == 159_389
    # *Total cash and cash equivalents — 61,668 + 63,097 = 124,765
    assert post_ws.cell(193, 2).value == 124_765
    # *Total issued capital — only Capital from ordinary shares 81,804
    assert post_ws.cell(203, 2).value == 81_804
    # *Total PPE — 12420 + 23345 + 56 + 16958 + 2275 + 9525 = 64,579
    assert post_ws.cell(39, 2).value == 64_579


def test_recalc_preserves_sheet_name_casing(tmp_path: Path) -> None:
    """The `formulas` package uppercases sheet names internally
    (SOFP-CuNonCu → SOFP-CUNONCU). The helper must map back to the
    original casing — otherwise every downstream reader that does
    ``wb['SOFP-Sub-CuNonCu']`` breaks loudly."""
    src = _FIXTURE_DIR / "sofp_company_mfrs.xlsx"
    dst = tmp_path / "merged.xlsx"
    shutil.copy(src, dst)
    recalc_workbook(dst)

    wb = load_workbook(dst)
    assert "SOFP-CuNonCu" in wb.sheetnames
    assert "SOFP-Sub-CuNonCu" in wb.sheetnames
    # Negative — neither uppercase variant should appear
    assert "SOFP-CUNONCU" not in wb.sheetnames
    assert "SOFP-SUB-CUNONCU" not in wb.sheetnames


def test_recalc_handles_mpers_group_layout(tmp_path: Path) -> None:
    """RUN-REVIEW MPERS coverage: the MPERS Group fixture has 6 columns
    (Group CY/PY + Company CY/PY + Source). Recalc must compute Total
    rows for every populated column-pair, not just column B."""
    src = _FIXTURE_DIR / "sofp_group_mpers.xlsx"
    dst = tmp_path / "merged.xlsx"
    shutil.copy(src, dst)
    recalc_workbook(dst)

    wb = load_workbook(dst, data_only=True)
    ws = wb["SOFP-Sub-CuNonCu"]
    # Group CY (col B) and Company CY (col D) should both have a Total
    # PPE value. The MPERS template's calc structure differs from MFRS
    # but the formula references whatever it references, so we just
    # assert non-None rather than a specific number.
    total_ppe_row = 26  # *Total property, plant and equipment in MPERS
    assert ws.cell(total_ppe_row, 2).value is not None, (
        "Group CY *Total PPE must be computed post-recalc"
    )
    assert ws.cell(total_ppe_row, 4).value is not None, (
        "Company CY *Total PPE must be computed post-recalc"
    )


def test_recalc_is_idempotent(tmp_path: Path) -> None:
    """Running recalc twice produces the same output as once. Important
    because the merge step might be retried; we don't want to fail or
    drift on a re-run."""
    src = _FIXTURE_DIR / "sofp_company_mfrs.xlsx"
    dst = tmp_path / "merged.xlsx"
    shutil.copy(src, dst)
    recalc_workbook(dst)
    first_pass_value = load_workbook(dst, data_only=True)["SOFP-Sub-CuNonCu"].cell(39, 2).value

    recalc_workbook(dst)
    second_pass_value = load_workbook(dst, data_only=True)["SOFP-Sub-CuNonCu"].cell(39, 2).value

    assert first_pass_value == second_pass_value == 64_579


def test_recalc_returns_path_unchanged_on_missing_file(tmp_path: Path) -> None:
    """Defensive: if the path doesn't exist we log a warning and return
    the path. Caller code can keep going (the fullCalcOnLoad flag is
    the fallback for any Excel user who opens the file)."""
    missing = tmp_path / "does_not_exist.xlsx"
    result = recalc_workbook(missing)
    assert result == missing  # not raised


def test_recalc_skipped_gracefully_when_formulas_missing(
    monkeypatch, tmp_path: Path
) -> None:
    """If the `formulas` package is somehow absent (vendored env, pip
    install issue), recalc must not crash — it should warn and return
    the path untouched. The caller's run continues."""
    src = _FIXTURE_DIR / "sofp_company_mfrs.xlsx"
    dst = tmp_path / "merged.xlsx"
    shutil.copy(src, dst)
    pre_size = dst.stat().st_size

    # Force an ImportError when the helper tries to import formulas
    import builtins
    real_import = builtins.__import__

    def _fake_import(name, *args, **kwargs):
        if name == "formulas":
            raise ImportError("simulated missing formulas package")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", _fake_import)

    result = recalc_workbook(dst)
    assert result == dst
    # File must not have been clobbered
    assert dst.stat().st_size == pre_size
