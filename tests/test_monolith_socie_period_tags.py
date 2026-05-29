"""Peer-review F2: the monolith structural dump tags SOCIE period blocks.

SOCIE repeats identical row labels per period block (and per Group/Company
level). Without a period tag the single monolith agent fills only the current
year and leaves PY empty (observed SOCIE PY-missing failure). The dump must
tag prior-period rows so the agent knows to fill them.
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import pytest

from monolith.prompt_renderer import _describe_sheet_rows, _socie_period_tags

REPO = Path(__file__).resolve().parent.parent


def _tags(level: str) -> Counter:
    lines = _describe_sheet_rows(REPO / "XBRL-template-MFRS" / level / "09-SOCIE.xlsx")
    return Counter(l.split("period: ")[1] for l in lines if "period:" in l)


def test_company_socie_has_two_period_blocks():
    tags = _tags("Company")
    assert set(tags) == {"CURRENT period", "PRIOR period"}
    # Both blocks carry a comparable number of rows (not a 1-row stub).
    assert tags["PRIOR period"] > 5


def test_group_socie_has_four_period_blocks():
    tags = _tags("Group")
    assert set(tags) == {
        "Group — CURRENT period", "Group — PRIOR period",
        "Company — CURRENT period", "Company — PRIOR period",
    }


def test_restated_row_does_not_split_blocks():
    """The '*Equity at beginning of period, restated' row sits just under each
    block opener; it must NOT be counted as a new block start (substring trap)."""
    import openpyxl
    path = REPO / "XBRL-template-MFRS" / "Company" / "09-SOCIE.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["SOCIE"]
    rows = {
        r: {"label": str(ws.cell(r, 1).value), "is_abstract": False}
        for r in range(1, ws.max_row + 1)
        if ws.cell(r, 1).value
    }
    wb.close()
    tags = _socie_period_tags(rows, sorted(rows))
    # Exactly two distinct labels → two blocks, not four.
    assert len(set(tags.values())) == 2
