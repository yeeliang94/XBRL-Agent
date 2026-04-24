"""Step 7 — Excel download regenerates notes sheets from `notes_cells`.

Once the rich editor writes edits back to the DB, the on-disk xlsx is
stale. The download endpoint overlays cells from `notes_cells` onto the
original workbook at stream time — keeping the DB as the canonical
store and the xlsx as a disposable render target.

Tests cover the overlay helper directly (cheap, independent of
FastAPI) and one end-to-end round-trip via the writer → overlay →
Excel inspection.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from db import repository as repo
from db.schema import init_db
from notes.payload import NotesPayload
from notes.persistence import persist_notes_cells
from notes.writer import write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    p = tmp_path / "xbrl.db"
    init_db(p)
    return p


def _seed_run(db_path: Path) -> int:
    with repo.db_session(db_path) as conn:
        return repo.create_run(conn, "sample.pdf")


def _find_row(ws, needle: str) -> int:
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label and needle.lower() in str(label).lower():
            return row
    raise AssertionError(f"no row matching {needle!r}")


def test_overlay_uses_notes_cells_html_when_present(
    tmp_path: Path, db_path: Path,
) -> None:
    """Seed `notes_cells` with HTML distinct from whatever the on-disk
    xlsx holds; the overlay must rewrite the cell to the HTML's
    flattened form."""
    from notes.persistence import overlay_notes_cells_into_workbook

    # Start from the real CORP_INFO template and do a trivial write so we
    # have a "stale" xlsx reflecting prior content.
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "stale.xlsx"
    write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="Original plaintext",
            evidence="Page 1, Note 1", source_pages=[1],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )

    run_id = _seed_run(db_path)
    # Row 8 is a stable prose row on Notes-CI (Financial reporting
    # status); look it up rather than hard-code.
    wb_probe = openpyxl.load_workbook(out)
    try:
        row_num = _find_row(wb_probe["Notes-CI"], "Financial reporting status")
    finally:
        wb_probe.close()

    # Put new HTML into notes_cells — the overlay should render it to
    # plaintext and land it in col B, overwriting the stale content.
    persist_notes_cells(
        db_path=str(db_path), run_id=run_id, sheet_name="Notes-CI",
        cells_written=[{
            "sheet": "Notes-CI", "row": row_num, "label": "Financial reporting status",
            "html": "<p><strong>Active</strong> as of year end.</p>",
            "evidence": "Page 3", "source_pages": [3],
        }],
    )

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=out, run_id=run_id, db_path=str(db_path),
    )

    wb = openpyxl.load_workbook(overlaid)
    try:
        cell = wb["Notes-CI"].cell(row=row_num, column=2).value
    finally:
        wb.close()
    # Overlay flattens HTML → plaintext. <strong> contributes only its
    # text content.
    assert cell == "Active as of year end."


def test_overlay_falls_back_when_no_cells(
    tmp_path: Path, db_path: Path,
) -> None:
    """Empty `notes_cells` → the overlay returns the original path
    untouched (the workbook on disk is the authoritative copy)."""
    from notes.persistence import overlay_notes_cells_into_workbook

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "stale.xlsx"
    write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="Keep me.",
            evidence="Page 1, Note 1", source_pages=[1],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )

    run_id = _seed_run(db_path)
    # No notes_cells rows for this run.

    returned = overlay_notes_cells_into_workbook(
        xlsx_path=out, run_id=run_id, db_path=str(db_path),
    )

    # The helper returns *some* path to a workbook — identity against
    # the input is one clean signal, but a copy is also acceptable as
    # long as the content survives intact.
    wb = openpyxl.load_workbook(returned)
    try:
        row_num = _find_row(wb["Notes-CI"], "Financial reporting status")
        # Cell now carries heading prepend + body (Phase 2 of the notes-
        # heading plan). Substring check because the heading line precedes
        # the body in the flattened Excel text.
        assert "Keep me." in wb["Notes-CI"].cell(row=row_num, column=2).value
    finally:
        wb.close()


def test_download_endpoint_cleans_up_overlay_temp_file(
    tmp_path: Path, db_path: Path, monkeypatch,
) -> None:
    """Peer-review finding #3: the overlay temp file must not accumulate
    in the system temp directory after the download completes."""
    import server
    from fastapi.testclient import TestClient

    # Point the server's audit DB at our fixture DB and seed a run with
    # a merged workbook on disk + a notes_cells row so the overlay path
    # fires (rather than the empty-cells fallback that just returns the
    # on-disk file unchanged).
    monkeypatch.setattr(server, "AUDIT_DB_PATH", db_path)

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    merged = tmp_path / "filled.xlsx"
    write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="Original", evidence="Page 1", source_pages=[1],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(merged),
        filing_level="company",
        sheet_name="Notes-CI",
    )
    with repo.db_session(db_path) as conn:
        run_id = repo.create_run(
            conn, "sample.pdf",
            session_id="sess", output_dir=str(tmp_path),
        )
        repo.mark_run_merged(conn, run_id, str(merged))

    wb_probe = openpyxl.load_workbook(merged)
    try:
        row_num = _find_row(wb_probe["Notes-CI"], "Financial reporting status")
    finally:
        wb_probe.close()
    persist_notes_cells(
        db_path=str(db_path), run_id=run_id, sheet_name="Notes-CI",
        cells_written=[{
            "sheet": "Notes-CI", "row": row_num, "label": "L",
            "html": "<p>Edited via editor</p>", "source_pages": [1],
        }],
    )

    import tempfile
    import glob
    temp_dir = tempfile.gettempdir()
    before = set(glob.glob(str(Path(temp_dir) / "notes_overlay_*.xlsx")))

    client = TestClient(server.app)
    resp = client.get(f"/api/runs/{run_id}/download/filled")
    assert resp.status_code == 200

    # Confirm the overlay fired — parse the response body as an xlsx and
    # verify the edited cell value landed. Excel-compressed raw bytes
    # can't be grep-ed directly.
    import io
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    try:
        assert wb["Notes-CI"].cell(row=row_num, column=2).value == "Edited via editor"
    finally:
        wb.close()

    after = set(glob.glob(str(Path(temp_dir) / "notes_overlay_*.xlsx")))
    # Any file created by this request must have been cleaned up.
    leaked = after - before
    assert not leaked, f"overlay temp files leaked: {leaked}"


def test_overlay_applies_rendered_length_cap_defensively(
    tmp_path: Path, db_path: Path,
) -> None:
    """Peer-review finding #2: the overlay must not flatten oversized
    HTML into an Excel cell that exceeds the documented 30k rendered
    limit, even if a direct DB write path (Step 8 PATCH) skipped the
    cap. The resulting cell must include the truncation footer."""
    from notes.persistence import overlay_notes_cells_into_workbook
    from notes.writer import CELL_CHAR_LIMIT

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "stale.xlsx"
    write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="seed", evidence="Page 1", source_pages=[1],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )

    run_id = _seed_run(db_path)
    wb_probe = openpyxl.load_workbook(out)
    try:
        row_num = _find_row(wb_probe["Notes-CI"], "Financial reporting status")
    finally:
        wb_probe.close()

    # Seed an oversized payload by writing straight into `notes_cells`
    # (simulating a PATCH endpoint that failed to validate). 40k chars
    # exceeds the cap by ~10k.
    from db import repository as repo
    with repo.db_session(db_path) as conn:
        repo.upsert_notes_cell(
            conn, run_id=run_id, sheet="Notes-CI", row=row_num,
            label="Financial reporting status",
            html="<p>" + ("Q" * 40_000) + "</p>",
            evidence="Page 3", source_pages=[3],
        )

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=out, run_id=run_id, db_path=str(db_path),
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        cell = wb["Notes-CI"].cell(row=row_num, column=2).value
    finally:
        wb.close()
    assert cell is not None
    assert len(cell) <= CELL_CHAR_LIMIT, (
        f"overlay cell is {len(cell)} chars, over the {CELL_CHAR_LIMIT} cap"
    )
    assert "[truncated -- see PDF pages 3]" in cell


def test_overlay_flattens_table_html_to_pipe_form(
    tmp_path: Path, db_path: Path,
) -> None:
    """An HTML table in the DB lands in the Excel cell as pipe-separated
    rows — the documented flattener contract from Step 1."""
    from notes.persistence import overlay_notes_cells_into_workbook

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "stale.xlsx"
    write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="ignored",
            evidence="Page 1, Note 1", source_pages=[1],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )

    run_id = _seed_run(db_path)
    wb_probe = openpyxl.load_workbook(out)
    try:
        row_num = _find_row(wb_probe["Notes-CI"], "Financial reporting status")
    finally:
        wb_probe.close()

    html = (
        "<table>"
        "<tr><th>H1</th><th>H2</th></tr>"
        "<tr><td>A</td><td>B</td></tr>"
        "</table>"
    )
    persist_notes_cells(
        db_path=str(db_path), run_id=run_id, sheet_name="Notes-CI",
        cells_written=[{
            "sheet": "Notes-CI", "row": row_num, "label": "Financial reporting status",
            "html": html, "evidence": "Page 3", "source_pages": [3],
        }],
    )

    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=out, run_id=run_id, db_path=str(db_path),
    )
    wb = openpyxl.load_workbook(overlaid)
    try:
        cell = wb["Notes-CI"].cell(row=row_num, column=2).value
    finally:
        wb.close()
    assert cell == "H1 | H2\nA | B"


def test_overlay_logs_warning_when_cell_sheet_not_in_workbook(
    tmp_path: Path, db_path: Path, caplog,
) -> None:
    """Peer-review #5: the overlay used to silently skip cells whose
    sheet is missing from the workbook (e.g. template generator
    rename, MPERS template drop). Users saw a stale xlsx while the
    editor showed the edited content; ops had no signal the drift
    had happened. The overlay now logs at WARNING level so the
    mismatch is visible in logs / the server console.
    """
    import logging as _logging
    from notes.persistence import overlay_notes_cells_into_workbook

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "stale.xlsx"
    write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="OLD",
            evidence="Page 1", source_pages=[1],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )
    run_id = _seed_run(db_path)
    # Seed a cell whose `sheet` column references a non-existent sheet —
    # this mimics a post-rename drift scenario.
    persist_notes_cells(
        db_path=str(db_path),
        run_id=run_id,
        sheet_name="Notes-DoesNotExist",
        cells_written=[{
            "sheet": "Notes-DoesNotExist",
            "row": 4,
            "label": "Stranded cell",
            "html": "<p>orphan</p>",
            "evidence": None,
            "source_pages": [],
        }],
    )

    caplog.set_level(_logging.WARNING, logger="notes.persistence")
    overlay_notes_cells_into_workbook(
        xlsx_path=out, run_id=run_id, db_path=str(db_path),
    )

    # The warning must name the sheet so an operator knows what drifted.
    messages = [r.getMessage() for r in caplog.records]
    assert any(
        "Notes-DoesNotExist" in m and "overlay" in m.lower()
        for m in messages
    ), f"expected a sheet-mismatch warning, got: {messages}"
