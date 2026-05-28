"""Pinning tests for SOCIE matrix-aware mandatory verification.

Peer-review (Edge AFS, 2026-05-28): `_collect_unfilled_mandatory` was a
single helper that scanned only col B (company) or B+D (group). MFRS
SOCIE is a 24-column matrix — valid activity can sit in col C (retained
earnings), col G (FCTR), col V (head office account), col X (total)
while col B is legitimately blank. The generic helper systematically
false-positived sparse SOCIE matrices and blocked save_result on
balanced statements.

The fix: when called for SOCIE with filing_standard='mfrs', the scan
widens to cols B..X (2..24). A row is filled if ANY of those columns
carries a value. MPERS SOCIE/SoRE is a flat B/C layout and keeps the
original col-B scan.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from statement_types import StatementType
from tools.verifier import verify_statement


def _make_socie_with_value_only_in(col_letter: str, *, tmp_path) -> Path:
    """Build a minimal MFRS SOCIE workbook where the mandatory profit row
    has its value only in `col_letter` (a single matrix component column).
    Every other cell on that row is blank.
    """
    path = tmp_path / f"socie_{col_letter}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"
    # Balance scaffolding so the verifier doesn't fail on the balance check
    # before it even reaches the mandatory scan.
    ws["A1"] = "Equity at beginning of period, restated"
    ws["X1"] = 100
    ws["A2"] = "Total increase (decrease) in equity"
    ws["X2"] = 50
    ws["A3"] = "Equity at end of period"
    ws["X3"] = 150
    # The mandatory row under test. Value lives in the matrix component
    # column specified by `col_letter` (C, G, V, X, ...) — col B is blank.
    ws["A4"] = "*Profit (loss)"
    ws[f"{col_letter}4"] = 50
    wb.save(str(path))
    return path


def test_mfrs_socie_mandatory_row_in_col_X_is_not_flagged_unfilled(tmp_path):
    """An MFRS SOCIE `*` row whose value is in col X (total) must NOT be
    reported under mandatory_unfilled — this is the most common matrix
    shape (everything aggregated through the total column)."""
    path = _make_socie_with_value_only_in("X", tmp_path=tmp_path)

    r = verify_statement(
        str(path), StatementType.SOCIE, variant="Default",
        filing_level="company", filing_standard="mfrs",
    )
    assert all("Profit (loss)" not in s for s in r.mandatory_unfilled), (
        f"MFRS SOCIE row with value in col X falsely flagged unfilled: "
        f"{r.mandatory_unfilled!r}"
    )


def test_mfrs_socie_mandatory_row_in_col_C_is_not_flagged_unfilled(tmp_path):
    """An MFRS SOCIE row with value only in col C (retained earnings) must
    not be flagged — the matrix sparsity is legitimate."""
    path = _make_socie_with_value_only_in("C", tmp_path=tmp_path)

    r = verify_statement(
        str(path), StatementType.SOCIE, variant="Default",
        filing_level="company", filing_standard="mfrs",
    )
    assert all("Profit (loss)" not in s for s in r.mandatory_unfilled)


def test_mfrs_socie_mandatory_row_in_col_V_is_not_flagged_unfilled(tmp_path):
    """The Edge AFS shape: value in col V (head office account) for branch
    presentations. The matrix scan must accept this as filled."""
    path = _make_socie_with_value_only_in("V", tmp_path=tmp_path)

    r = verify_statement(
        str(path), StatementType.SOCIE, variant="Default",
        filing_level="company", filing_standard="mfrs",
    )
    assert all("Profit (loss)" not in s for s in r.mandatory_unfilled)


def test_mfrs_socie_truly_blank_mandatory_row_is_still_flagged(tmp_path):
    """The fix must not silence legitimate gaps — a `*` row with EVERY
    column blank (B..X) is still unfilled and must be reported."""
    path = tmp_path / "socie_blank.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"
    ws["A1"] = "Equity at beginning of period, restated"
    ws["X1"] = 100
    ws["A2"] = "Total increase (decrease) in equity"
    ws["X2"] = 0
    ws["A3"] = "Equity at end of period"
    ws["X3"] = 100
    ws["A4"] = "*Profit (loss)"
    # B4..X4 all blank — genuinely unfilled.
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOCIE, variant="Default",
        filing_level="company", filing_standard="mfrs",
    )
    assert any("Profit (loss)" in s for s in r.mandatory_unfilled), (
        f"Genuinely blank mandatory row was silenced by the matrix scan: "
        f"{r.mandatory_unfilled!r}"
    )


def test_mpers_socie_mandatory_keeps_col_B_scan(tmp_path):
    """MPERS SOCIE is a flat B/C layout — the matrix widening must NOT
    apply. A `*` row with value only in col X but col B blank must
    still be reported as unfilled on MPERS."""
    path = tmp_path / "socie_mpers.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"
    ws["A1"] = "Equity at beginning of period, restated"
    ws["B1"] = 100
    ws["A2"] = "Total increase (decrease) in equity"
    ws["B2"] = 50
    ws["A3"] = "Equity at end of period"
    ws["B3"] = 150
    ws["A4"] = "*Profit (loss)"
    # B4 blank, X4 has a value — on MFRS this passes; on MPERS it must fail.
    ws["X4"] = 50
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOCIE, variant="Default",
        filing_level="company", filing_standard="mpers",
    )
    assert any("Profit (loss)" in s for s in r.mandatory_unfilled), (
        f"MPERS SOCIE must keep the col-B scan — col-X value should not "
        f"count as filled. Got mandatory_unfilled={r.mandatory_unfilled!r}"
    )
