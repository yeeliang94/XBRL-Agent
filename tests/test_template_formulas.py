from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parent.parent
SOFP_TEMPLATE = REPO_ROOT / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"


def test_sofp_secured_borrowings_total_uses_leaf_rows() -> None:
    """Row 256 must total the secured-borrowing leaf rows, not the grand total."""
    wb = load_workbook(SOFP_TEMPLATE, data_only=False)
    ws = wb["SOFP-Sub-CuNonCu"]

    assert ws["B256"].value == "=1*B251+1*B252+1*B253+1*B254+1*B255"
    assert ws["C256"].value == "=1*C251+1*C252+1*C253+1*C254+1*C255"

    wb.close()


def test_sofp_template_has_no_formula_cycles() -> None:
    """The SOFP template should not contain circular references."""
    wb = load_workbook(SOFP_TEMPLATE, data_only=False)
    cell_ref = re.compile(r"(?:(?:'([^']+)'|([A-Za-z0-9_ -]+))!)?\$?([A-Z]{1,3})\$?(\d+)")

    formula_cells: dict[tuple[str, str], list[tuple[str, str]]] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                deps: list[tuple[str, str]] = []
                for match in cell_ref.finditer(cell.value):
                    sheet = match.group(1) or match.group(2) or ws.title
                    deps.append((sheet, f"{match.group(3)}{match.group(4)}"))
                formula_cells[(ws.title, cell.coordinate)] = deps

    visited: dict[tuple[str, str], int] = {}
    stack: list[tuple[str, str]] = []

    def dfs(node: tuple[str, str]) -> None:
        visited[node] = 1
        stack.append(node)
        for dep in formula_cells.get(node, []):
            if dep not in formula_cells:
                continue
            state = visited.get(dep, 0)
            if state == 0:
                dfs(dep)
                continue
            if state == 1:
                cycle = stack[stack.index(dep):] + [dep]
                cycle_text = " -> ".join(f"{sheet}!{coord}" for sheet, coord in cycle)
                raise AssertionError(f"Formula cycle detected: {cycle_text}")
        stack.pop()
        visited[node] = 2

    for node in formula_cells:
        if visited.get(node, 0) == 0:
            dfs(node)

    wb.close()


# ---------------------------------------------------------------------------
# 2026-07-03 linkbase regeneration of the hand-built MFRS SOCI + SOCF-Direct
# templates (scripts/regenerate_mfrs_sofp_sopl_formulas.py). The originals
# deviated from the SSM calculation linkbase: OCI component subtotals never
# reached Total OCI, reclassification adjustments were added instead of
# subtracted, and SOCF-Direct subtracted payment rows the linkbase weights +1
# (payments are entered as negative cash effects, matching MPERS and
# prompts/socf.md). These pins guard the regenerated formulas.
# ---------------------------------------------------------------------------

def _formula(template: str, sheet: str, coord: str) -> str:
    wb = load_workbook(REPO_ROOT / "XBRL-template-MFRS" / template, data_only=False)
    try:
        return wb[sheet][coord].value
    finally:
        wb.close()


def test_soci_netoftax_reclass_components_roll_into_total_oci() -> None:
    """Row 36 must sum ALL will-be-reclassified component subtotals; the
    hand-built original summed only rows 34+35, so filled hedge/FX/FVOCI
    values silently never reached Total OCI."""
    for level in ("Company", "Group"):
        f = f"{level}/06-SOCI-NetOfTax.xlsx"
        assert _formula(f, "SOCI-NetOfTax", "B36") == (
            "=1*B20+1*B24+1*B28+1*B33+1*B34+1*B35"
        )
        assert _formula(f, "SOCI-NetOfTax", "B37") == "=1*B15+1*B36"
        # Not-reclassified side: row 13 is the SSM parent of rows 10-12.
        assert _formula(f, "SOCI-NetOfTax", "B13") == "=1*B10+1*B11+1*B12"
        assert _formula(f, "SOCI-NetOfTax", "B15") == "=1*B13+1*B14"


def test_soci_reclassification_adjustments_are_subtracted() -> None:
    """SSM weights reclassification adjustments -1 (they move OCI to P&L).
    The hand-built NetOfTax original added them."""
    for level in ("Company", "Group"):
        f = f"{level}/06-SOCI-NetOfTax.xlsx"
        assert _formula(f, "SOCI-NetOfTax", "B20") == "=1*B18+-1*B19"
        assert _formula(f, "SOCI-NetOfTax", "B24") == "=1*B22+-1*B23"
        assert _formula(f, "SOCI-NetOfTax", "B28") == "=1*B26+-1*B27"
        assert _formula(f, "SOCI-NetOfTax", "B33") == "=1*B30+-1*B31+-1*B32"


def test_soci_beforetax_other_items_subtotal_prevents_double_count() -> None:
    """Row 13 is the SSM parent of rows 11-12; row 15 must sum 13+14 (not the
    flattened 11+12+13+14, which double-counted when row 13 was also filled)."""
    for level in ("Company", "Group"):
        f = f"{level}/05-SOCI-BeforeTax.xlsx"
        assert _formula(f, "SOCI-BeforeOfTax", "B13") == "=1*B11+1*B12"
        assert _formula(f, "SOCI-BeforeOfTax", "B15") == "=1*B13+1*B14"


def test_socf_direct_operating_rows_all_added_per_linkbase() -> None:
    """SSM weights every direct-method operating row +1 — payments are entered
    as negative cash effects (matching MPERS and prompts/socf.md). The
    hand-built original subtracted payment rows, so a prompt-following agent's
    negative entry was double-negated into a cash INCREASE."""
    for level in ("Company", "Group"):
        f = f"{level}/08-SOCF-Direct.xlsx"
        b20 = _formula(f, "SOCF-Direct", "B20")
        assert b20 == (
            "=1*B8+1*B9+1*B10+1*B11+1*B12+1*B13+1*B14"
            "+1*B15+1*B16+1*B17+1*B18+1*B19"
        )
        # The lease-payments row (B11) must stay in the formula: its +1 direct
        # arc must not be cancelled by the indirect-method expansion (see
        # ROLE_EXPANSION_EXCLUDE in scripts/regenerate_mfrs_sofp_sopl_formulas.py).
        assert "B11" in b20


def test_socf_direct_period_bridge_survives_regeneration() -> None:
    """Cash at end = beginning + net increase is a cross-period identity the
    calc linkbase cannot express; regeneration must re-apply it
    (SUPPLEMENTAL_FORMULAS), not drop it."""
    for level in ("Company", "Group"):
        f = f"{level}/08-SOCF-Direct.xlsx"
        assert _formula(f, "SOCF-Direct", "B77") == "=1*B76+1*B75"
        assert _formula(f, "SOCF-Direct", "B73") == "=1*B20+1*B52+1*B72"
