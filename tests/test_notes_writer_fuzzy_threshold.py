"""Threshold + regression tests for notes/writer.py fuzzy label resolution.

Threshold 0.85 is the floor: anything below rejects with an error, anything
at or above writes. Lower scores caused real mis-routings in prior runs
(see the regression cases at the bottom of this file) — loosening the
threshold would silently re-enable those bugs.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from notes.payload import NotesPayload
from notes.writer import (
    BORDERLINE_FUZZY_SCORE,
    _FUZZY_THRESHOLD,
    _resolve_row,
    _build_label_index,
    write_notes_workbook,
)
from notes_types import NotesTemplateType, notes_template_path
import openpyxl


LIST_OF_NOTES_SHEET = "Notes-Listofnotes"
CORP_INFO_SHEET = "Notes-CI"


# ---------------------------------------------------------------------------
# Boundary tests for _resolve_row
# ---------------------------------------------------------------------------

def test_fuzzy_threshold_is_0_85() -> None:
    """Hard-pin the threshold. Lowering this re-admits mis-routings that
    bit real production runs — see the regression cases below. If you have
    a good reason to change it, update those cases first and make sure
    they still protect what they're meant to protect."""
    assert _FUZZY_THRESHOLD == 0.85


def test_fuzzy_threshold_matches_borderline_constant() -> None:
    """The writer rejects below `_FUZZY_THRESHOLD`; the coordinator's
    "borderline warning" logic uses `BORDERLINE_FUZZY_SCORE`. Keeping
    them aligned means anything written is, by definition, above the
    borderline — the warning path in the coordinator is dormant but
    harmless and the two constants cannot drift."""
    assert _FUZZY_THRESHOLD == BORDERLINE_FUZZY_SCORE


def test_resolve_row_rejects_score_below_0_85() -> None:
    """A 0.78-score label (the real "taxation" → "bonds" case from the
    mini run) must resolve to None so the writer routes it to errors
    instead of writing."""
    tpl = notes_template_path(NotesTemplateType.LIST_OF_NOTES, level="company")
    wb = openpyxl.load_workbook(tpl)
    ws = wb[LIST_OF_NOTES_SHEET]
    entries = _build_label_index(ws)
    wb.close()

    # "Disclosure of taxation" scores 0.78 against "Disclosure of bonds".
    assert _resolve_row(entries, "Disclosure of taxation") is None


def test_resolve_row_accepts_score_above_0_85() -> None:
    """A ~0.98 near-match ("statu" → "status") must still resolve —
    legitimate minor typos should not be rejected."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    wb = openpyxl.load_workbook(tpl)
    ws = wb[CORP_INFO_SHEET]
    entries = _build_label_index(ws)
    wb.close()

    # "Financial reporting statu" → "Financial reporting status" scores 0.98.
    resolved = _resolve_row(entries, "Financial reporting statu")
    assert resolved is not None
    row, chosen, score = resolved
    assert "financial reporting status" in chosen.lower()
    assert score >= 0.85


def test_resolve_row_accepts_exact_match_with_star_prefix() -> None:
    """Leading `*` markers (mandatory-row indicators in the template)
    must not break exact matching."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    wb = openpyxl.load_workbook(tpl)
    ws = wb[CORP_INFO_SHEET]
    entries = _build_label_index(ws)
    wb.close()

    resolved = _resolve_row(entries, "Disclosure of corporate information")
    assert resolved is not None
    _, _, score = resolved
    assert score == 1.0


# ---------------------------------------------------------------------------
# End-to-end rejection tests via write_notes_workbook
# ---------------------------------------------------------------------------

def test_writer_rejects_below_threshold_label_as_error(tmp_path: Path) -> None:
    """Writer must surface "no matching row" in result.errors for any
    label below 0.85 — the sub-coordinator and single-sheet agents rely
    on this signal to retry or skip."""
    tpl = notes_template_path(NotesTemplateType.LIST_OF_NOTES, level="company")
    out = tmp_path / "rejected.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Disclosure of taxation",
            content="Income tax expense reconciliation ...",
            evidence="Page 23, Note 10",
            source_pages=[23],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=LIST_OF_NOTES_SHEET,
    )
    assert result.rows_written == 0
    assert any("no matching row" in e.lower() for e in result.errors)
    assert not result.fuzzy_matches, (
        "below-threshold labels must not land on fuzzy_matches — they are "
        "rejected, not borderline"
    )


# ---------------------------------------------------------------------------
# Regression pins for the two real mini-run mis-inserts.
#
# These exact payloads were silently force-inserted into wrong rows on a
# production run against the FYE 2022 audited statements. Keeping them
# here as regression tests guarantees anyone touching the fuzzy logic
# sees the failure mode they'd be re-enabling.
# ---------------------------------------------------------------------------

def test_regression_taxation_must_not_force_insert_into_bonds(
    tmp_path: Path,
) -> None:
    """Mini-run bug: `"Disclosure of taxation"` (fuzzy 0.78) landed the
    entire income-tax-policy paragraph into row 12 `"Disclosure of
    bonds"`. The correct target is row 71 `"Disclosure of income tax
    expense"` — but that scores 0.67, below any sensible threshold.
    The right outcome is REJECTION so the agent must pick a real label
    or skip the note."""
    tpl = notes_template_path(NotesTemplateType.LIST_OF_NOTES, level="company")
    out = tmp_path / "regression_taxation.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Disclosure of taxation",
            content=(
                "Income tax on profit or loss for the financial year "
                "comprises current tax and deferred tax. Current tax is "
                "the expected amount of income taxes payable in respect "
                "of the taxable profit for the financial year."
            ),
            evidence="Pages 19-20, Note 2(g), Note 7 and Note 10",
            source_pages=[19, 20],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=LIST_OF_NOTES_SHEET,
    )
    assert result.rows_written == 0, (
        "Writer silently force-inserted into a wrong row — the fuzzy "
        "threshold is too loose. This regressed the mini-run "
        "taxation->bonds bug."
    )


def test_regression_corporate_info_must_not_force_insert_into_fair_value(
    tmp_path: Path,
) -> None:
    """Mini-run bug: `"Disclosure of corporate information"` (fuzzy
    0.82) landed Sheet-10 content into Sheet-12 row 49 `"Disclosure of
    fair value information"`. Corporate info belongs on a different
    sheet entirely; the right outcome is REJECTION here so the agent
    uses the skip/coverage path instead of force-inserting."""
    tpl = notes_template_path(NotesTemplateType.LIST_OF_NOTES, level="company")
    out = tmp_path / "regression_corpinfo.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Disclosure of corporate information",
            content=(
                "The Company remained dormant during the financial year. "
                "The Company is a public limited company, incorporated "
                "and domiciled in Malaysia."
            ),
            evidence="Page 16, Note 1",
            source_pages=[16],
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=LIST_OF_NOTES_SHEET,
    )
    assert result.rows_written == 0, (
        "Writer silently force-inserted Sheet-10 content into a Sheet-12 "
        "row — the fuzzy threshold is too loose. This regressed the "
        "mini-run corporate-info->fair-value bug."
    )
