import openpyxl
import pytest

from tools.verifier import (
    verify_totals,
    VerificationResult,
    _evaluate_formula,
    _resolve_cell_value,
    verify_statement,
)
from statement_types import StatementType


def _make_template(tmp_path):
    """Create a minimal SOFP-like template with pre-calculated values."""
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP"

    ws["A1"] = "Total assets"
    ws["B1"] = 400
    ws["C1"] = 600
    ws["A2"] = "Current assets"
    ws["B2"] = 100
    ws["C2"] = 200
    ws["A3"] = "Non-current assets"
    ws["B3"] = 300
    ws["C3"] = 400

    ws["A5"] = "Total equity and liabilities"
    ws["B5"] = 400
    ws["C5"] = 600
    ws["A6"] = "Equity"
    ws["B6"] = 250
    ws["C6"] = 350
    ws["A7"] = "Liabilities"
    ws["B7"] = 150
    ws["C7"] = 250

    wb.save(str(path))
    return path


def test_verify_balanced(tmp_path):
    path = _make_template(tmp_path)
    result = verify_totals(str(path))
    assert result.is_balanced


def test_verify_totals_match(tmp_path):
    path = _make_template(tmp_path)
    result = verify_totals(str(path))
    assert result.computed_totals["total_assets_cy"] == 400
    assert result.computed_totals["total_assets_py"] == 600
    assert result.computed_totals["total_equity_liabilities_cy"] == 400
    assert result.computed_totals["total_equity_liabilities_py"] == 600


def test_verify_pdf_values_match(tmp_path):
    path = _make_template(tmp_path)
    pdf_values = {
        "total_assets_cy": 400,
        "total_assets_py": 600,
        "total_equity_liabilities_cy": 400,
        "total_equity_liabilities_py": 600,
    }
    result = verify_totals(str(path), pdf_values=pdf_values)
    assert result.matches_pdf


def test_verify_pdf_values_mismatch(tmp_path):
    path = _make_template(tmp_path)
    pdf_values = {
        "total_assets_cy": 999,
        "total_assets_py": 600,
        "total_equity_liabilities_cy": 400,
        "total_equity_liabilities_py": 600,
    }
    result = verify_totals(str(path), pdf_values=pdf_values)
    assert not result.matches_pdf
    assert len(result.mismatches) > 0


def test_verify_file_not_found():
    with pytest.raises(FileNotFoundError):
        verify_totals("/nonexistent/file.xlsx")


def test_verify_unbalanced(tmp_path):
    path = tmp_path / "unbalanced.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Total assets"
    ws["B1"] = 500
    ws["C1"] = 600
    ws["A2"] = "Total equity and liabilities"
    ws["B2"] = 400
    ws["C2"] = 600
    wb.save(str(path))

    result = verify_totals(str(path))
    assert not result.is_balanced
    assert len(result.mismatches) > 0
    assert "IMBALANCE" in result.feedback
    assert "100" in result.feedback


def test_verify_unbalanced_feedback_direction(tmp_path):
    """When assets > equity+liabilities, feedback should point at the liabilities side."""
    path = tmp_path / "unbalanced.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Total assets"
    ws["B1"] = 1000
    ws["A2"] = "Total equity and liabilities"
    ws["B2"] = 800
    wb.save(str(path))

    result = verify_totals(str(path))
    assert not result.is_balanced
    assert "equity+liabilities section is too low" in result.feedback


def test_evaluate_formula_weighted_sum(tmp_path):
    """Formula parser handles =1*B2+1*B3 style sums."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["B2"] = 100
    ws["B3"] = 200

    result = _evaluate_formula(wb, "Sheet", "=1*B2+1*B3")
    assert result == 300.0
    wb.close()


def test_evaluate_formula_cross_sheet_ref(tmp_path):
    """Formula parser handles ='OtherSheet'!B5 cross-sheet references."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main"
    ws2 = wb.create_sheet("OtherSheet")
    ws2["B5"] = 42

    result = _evaluate_formula(wb, "Main", "='OtherSheet'!B5")
    assert result == 42.0
    wb.close()


def test_evaluate_formula_recursive(tmp_path):
    """Formula evaluation recurses through chained formula cells."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["B1"] = 100
    ws["B2"] = 200
    ws["B3"] = "=1*B1+1*B2"  # = 300
    ws["B4"] = "=1*B3"  # = 300 (references another formula)

    result = _evaluate_formula(wb, "Sheet", "=1*B4")
    assert result == 300.0
    wb.close()


def test_evaluate_formula_recursive_cross_sheet(tmp_path):
    """Recursion works across sheets: main formula -> cross-sheet ref -> formula."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main"
    ws2 = wb.create_sheet("Sub")
    ws2["B1"] = 50
    ws2["B2"] = 70
    ws2["B3"] = "=1*B1+1*B2"  # = 120

    # Main sheet references Sub's formula cell
    result = _evaluate_formula(wb, "Main", "='Sub'!B3")
    assert result == 120.0
    wb.close()


def test_evaluate_formula_cycle_detection(tmp_path):
    """Circular references don't cause infinite recursion."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["B1"] = "=1*B2"
    ws["B2"] = "=1*B1"  # circular

    # Should not hang — returns 0.0 for the cycle
    result = _evaluate_formula(wb, "Sheet", "=1*B1")
    assert result == 0.0
    wb.close()


def test_verify_with_chained_formulas(tmp_path):
    """verify_totals correctly evaluates total rows that reference formula cells."""
    path = tmp_path / "chained.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP"

    ws["A1"] = "Current assets"
    ws["B1"] = 100
    ws["A2"] = "Non-current assets"
    ws["B2"] = 200
    # Subtotal formula
    ws["A3"] = "Sub-total"
    ws["B3"] = "=1*B1+1*B2"  # = 300
    # Total assets references the subtotal formula
    ws["A4"] = "Total assets"
    ws["B4"] = "=1*B3"  # = 300

    ws["A6"] = "Equity"
    ws["B6"] = 150
    ws["A7"] = "Liabilities"
    ws["B7"] = 150
    ws["A8"] = "Total equity and liabilities"
    ws["B8"] = "=1*B6+1*B7"  # = 300

    wb.save(str(path))

    result = verify_totals(str(path))
    assert result.is_balanced
    assert result.computed_totals["total_assets_cy"] == 300.0
    assert result.computed_totals["total_equity_liabilities_cy"] == 300.0


def test_verify_targets_sofp_sheet(tmp_path):
    """Verifier scans SOFP-CuNonCu sheet specifically, not just wb.active."""
    path = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    # Active sheet is NOT the SOFP sheet
    ws_other = wb.active
    ws_other.title = "SomeOtherSheet"
    ws_other["A1"] = "Irrelevant data"

    # Create the SOFP sheet with totals
    ws_sofp = wb.create_sheet("SOFP-CuNonCu")
    ws_sofp["A1"] = "*Total assets"
    ws_sofp["B1"] = 500
    ws_sofp["A2"] = "*Total equity and liabilities"
    ws_sofp["B2"] = 500

    wb.save(str(path))

    result = verify_totals(str(path))
    assert result.is_balanced
    assert result.computed_totals["total_assets_cy"] == 500


# ---------------------------------------------------------------------------
# Phase 1.1: mandatory-field detection on face sheets
# ---------------------------------------------------------------------------

def test_verification_result_has_mandatory_unfilled_field():
    """VerificationResult must carry a `mandatory_unfilled` list so
    verify_totals feedback can surface unfilled `*` rows."""
    r = VerificationResult(is_balanced=True, matches_pdf=None)
    assert r.mandatory_unfilled == []


def test_verify_sopl_reports_unfilled_asterisks(tmp_path):
    """An SOPL sheet with a blank `*Revenue` row should surface 'Revenue'
    under `mandatory_unfilled`, without suppressing the attribution check."""
    path = tmp_path / "sopl.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Function"
    ws["A1"] = "*Revenue"
    # B1 deliberately left blank
    ws["A2"] = "*Cost of sales"
    ws["B2"] = -40
    ws["A3"] = "*Profit (loss)"
    ws["B3"] = 60
    ws["A4"] = "*Total profit (loss)"
    ws["B4"] = 60
    wb.save(str(path))

    r = verify_statement(str(path), StatementType.SOPL, variant="Function")
    assert "*Revenue" in r.mandatory_unfilled


def test_verify_sofp_reports_unfilled_asterisks(tmp_path):
    """Mandatory-unfilled helper must also run on SOFP via verify_totals."""
    path = tmp_path / "sofp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "*Total assets"
    ws["B1"] = 100
    ws["A2"] = "*Cash and cash equivalents"
    # B2 blank — unfilled
    ws["A3"] = "*Total equity and liabilities"
    ws["B3"] = 100
    wb.save(str(path))

    r = verify_totals(str(path))
    assert any("Cash and cash equivalents" in s for s in r.mandatory_unfilled)


def test_verify_socie_reports_unfilled_asterisks(tmp_path):
    """SOCIE verifier must also populate mandatory_unfilled."""
    path = tmp_path / "socie.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"
    ws["A1"] = "Equity at beginning of period, restated"
    ws["X1"] = 100
    ws["A2"] = "Total increase (decrease) in equity"
    ws["X2"] = 50
    ws["A3"] = "Equity at end of period"
    ws["X3"] = 150
    # A mandatory row that is unfilled anywhere on the sheet
    ws["A4"] = "*Profit (loss)"
    # B4 blank
    wb.save(str(path))

    r = verify_statement(str(path), StatementType.SOCIE, variant="Default")
    assert any("Profit (loss)" in s for s in r.mandatory_unfilled)


def test_verify_sopl_group_checks_attribution(tmp_path):
    """SOPL Group filing: owners + NCI attribution rows must sum to Profit
    (loss). A mismatch between (owners, NCI) pair and the profit row must
    drive is_balanced=False."""
    path = tmp_path / "sopl_group_attr.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Function"
    ws["A1"] = "Profit (loss)"
    ws["B1"] = 120  # group profit
    ws["D1"] = 100  # company profit
    ws["A2"] = "Profit (loss), attributable to, owners of parent"
    ws["B2"] = 100  # owners
    ws["D2"] = 100
    ws["A3"] = "Profit (loss), attributable to, non-controlling interests"
    ws["B3"] = 10  # NCI — sum is 110 but profit is 120 → mismatch
    ws["D3"] = 0
    ws["A4"] = "Total profit (loss)"
    ws["B4"] = 120
    ws["D4"] = 100
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOPL, variant="Function", filing_level="group",
    )
    assert r.is_balanced is False
    assert any("attribution" in m.lower() or "owners" in m.lower()
               for m in r.mismatches)


def test_verify_sopl_group_attribution_passes_when_sum_matches(tmp_path):
    """Owners + NCI that actually sums to profit must not fail the check."""
    path = tmp_path / "sopl_group_ok.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Function"
    ws["A1"] = "Profit (loss)"
    ws["B1"] = 110
    ws["D1"] = 100
    ws["A2"] = "Profit (loss), attributable to, owners of parent"
    ws["B2"] = 100
    ws["D2"] = 100
    ws["A3"] = "Profit (loss), attributable to, non-controlling interests"
    ws["B3"] = 10
    ws["D3"] = 0
    ws["A4"] = "Total profit (loss)"
    ws["B4"] = 110
    ws["D4"] = 100
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOPL, variant="Function", filing_level="group",
    )
    # The owners+NCI check should pass; attribution total still matches profit.
    attribution_fails = [m for m in r.mismatches
                         if "owners" in m.lower() and "non-controlling" in m.lower()]
    assert not attribution_fails


def test_verify_sofp_group_checks_equity_attribution(tmp_path):
    """SOFP Group filing: owners + NCI must sum to Total equity."""
    path = tmp_path / "sofp_group_eq.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Equity attributable to owners of parent"
    ws["B1"] = 500
    ws["D1"] = 500
    ws["A2"] = "Non-controlling interests"
    ws["B2"] = 50  # owners+NCI = 550 but total equity is 600 → mismatch
    ws["D2"] = 0
    ws["A3"] = "Total equity"
    ws["B3"] = 600
    ws["D3"] = 500
    ws["A4"] = "Total assets"
    ws["B4"] = 700
    ws["D4"] = 600
    ws["A5"] = "Total equity and liabilities"
    ws["B5"] = 700
    ws["D5"] = 600
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOFP, variant="CuNonCu", filing_level="group",
    )
    assert r.is_balanced is False
    assert any("equity" in m.lower() and ("owners" in m.lower() or "attribut" in m.lower())
               for m in r.mismatches)


def test_verify_sofp_group_equity_attribution_passes_when_sum_matches(tmp_path):
    path = tmp_path / "sofp_group_eq_ok.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Equity attributable to owners of parent"
    ws["B1"] = 500
    ws["D1"] = 500
    ws["A2"] = "Non-controlling interests"
    ws["B2"] = 100
    ws["D2"] = 0
    ws["A3"] = "Total equity"
    ws["B3"] = 600
    ws["D3"] = 500
    ws["A4"] = "Total assets"
    ws["B4"] = 700
    ws["D4"] = 600
    ws["A5"] = "Total equity and liabilities"
    ws["B5"] = 700
    ws["D5"] = 600
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOFP, variant="CuNonCu", filing_level="group",
    )
    equity_attr_fails = [m for m in r.mismatches
                         if "non-controlling" in m.lower() and "equity" in m.lower()]
    assert not equity_attr_fails


def test_verify_sofp_company_does_not_check_equity_attribution(tmp_path):
    """The equity-attribution check is Group-only — running it on a Company
    filing would fail needlessly because standalone SOFPs don't carry NCI."""
    path = tmp_path / "sofp_company.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Total equity"
    ws["B1"] = 600
    ws["A2"] = "Total assets"
    ws["B2"] = 700
    ws["A3"] = "Total equity and liabilities"
    ws["B3"] = 700
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOFP, variant="CuNonCu", filing_level="company",
    )
    attr_fails = [m for m in r.mismatches if "attribut" in m.lower()]
    assert not attr_fails


def test_verify_sopl_group_flags_company_column_unfilled(tmp_path):
    """For group filings, mandatory-unfilled must scan both Group CY (col B)
    and Company CY (col D) — leaving a company column blank still counts."""
    path = tmp_path / "sopl_group.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Function"
    ws["A1"] = "*Revenue"
    ws["B1"] = 100  # Group CY filled
    # D1 (Company CY) blank — unfilled
    ws["A2"] = "*Profit (loss)"
    ws["B2"] = 30
    ws["D2"] = 30
    ws["A3"] = "*Total profit (loss)"
    ws["B3"] = 30
    ws["D3"] = 30
    wb.save(str(path))

    r = verify_statement(
        str(path), StatementType.SOPL, variant="Function", filing_level="group"
    )
    assert any("Revenue" in s for s in r.mandatory_unfilled)
