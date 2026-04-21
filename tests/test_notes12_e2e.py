"""End-to-end tests for Sheet 12 (List of Notes) — Phase C.2.

These tests wire the sub-coordinator into `notes.coordinator` and exercise:
  - `notes.coordinator.run_notes_extraction` dispatches LIST_OF_NOTES
    through the sub-coordinator instead of the single-agent runner.
  - The final xlsx for Notes-12 reflects aggregated payloads from all
    sub-agents (via the writer's row-concatenation logic).
  - `notes12_unmatched.json` is written when any payload lands on row 112.
  - The Sheet-12 prompt file exists and is non-empty.
  - `run.py --notes list_of_notes` is now a valid CLI choice.

All LLM interactions are mocked — `_invoke_sub_agent_once` is patched so
tests never touch the network.
"""
from __future__ import annotations

import asyncio
import json
import subprocess
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from notes.payload import NotesPayload
from notes_types import NotesTemplateType, notes_template_path
from scout.notes_discoverer import NoteInventoryEntry
from scout.infopack import Infopack


# ---------------------------------------------------------------------------
# Prompt file
# ---------------------------------------------------------------------------

def test_listofnotes_prompt_exists_and_references_template_row_count():
    prompt_path = Path(__file__).resolve().parent.parent / "prompts" / "notes_listofnotes.md"
    assert prompt_path.exists(), f"Missing prompt: {prompt_path}"
    text = prompt_path.read_text(encoding="utf-8").lower()
    # Must mention the sheet and the unmatched-row sink.
    assert "notes-listofnotes" in text or "list of notes" in text or "notes 12" in text
    assert "disclosure of other notes to accounts" in text


# ---------------------------------------------------------------------------
# Coordinator dispatch
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_coordinator_dispatches_list_of_notes_through_subcoordinator(tmp_path: Path):
    """run_notes_extraction must use the sub-coordinator (not the single
    agent runner) for LIST_OF_NOTES. The hook point is _run_single_notes_agent
    — when LIST_OF_NOTES is requested, the coordinator should NOT call it."""
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    pdf_path = tmp_path / "dummy.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run={NotesTemplateType.LIST_OF_NOTES},
        filing_level="company",
    )
    infopack = Infopack(
        toc_page=1, page_offset=0,
        notes_inventory=[
            NoteInventoryEntry(4, "Revenue", (20, 21)),
            NoteInventoryEntry(5, "Cost of sales", (22, 23)),
        ],
    )

    captured_invocations: list[dict] = []

    async def fake_invoke(**kwargs):
        captured_invocations.append(kwargs)
        # Return one matched + one unmatched payload per sub-agent.
        batch = kwargs["batch"]
        # Phase 5: invoke returns (payloads, prompt_tokens, completion_tokens).
        payloads = [
            NotesPayload(
                chosen_row_label="Disclosure of borrowings",
                content=f"Revenue for notes {[e.note_num for e in batch]}",
                evidence=f"p.{batch[0].page_range[0]}",
                source_pages=[batch[0].page_range[0]],
            ),
        ]
        return payloads, 0, 0, None

    async def fake_single(*_args, **_kwargs):
        raise AssertionError(
            "Single-agent runner must NOT be called for LIST_OF_NOTES"
        )

    with patch(
        "notes.listofnotes_subcoordinator._invoke_sub_agent_once",
        side_effect=fake_invoke,
    ), patch(
        "notes.coordinator._run_single_notes_agent",
        side_effect=fake_single,
    ):
        result = await run_notes_extraction(config, infopack=infopack)

    # One NotesAgentResult for LIST_OF_NOTES, succeeded, with a workbook.
    assert len(result.agent_results) == 1
    r = result.agent_results[0]
    assert r.template_type == NotesTemplateType.LIST_OF_NOTES
    assert r.status == "succeeded", r.error
    # Sub-invocations happened (at least one).
    assert len(captured_invocations) >= 1


@pytest.mark.asyncio
async def test_list_of_notes_writes_merged_workbook_and_populates_rows(tmp_path: Path):
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    pdf_path = tmp_path / "dummy.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run={NotesTemplateType.LIST_OF_NOTES},
        filing_level="company",
    )
    infopack = Infopack(
        toc_page=1, page_offset=0,
        notes_inventory=[NoteInventoryEntry(i, f"Note {i}", (20 + i, 20 + i)) for i in range(1, 6)],
    )

    async def fake_invoke(**kwargs):
        batch = kwargs["batch"]
        # Emit one canonical label per note in the batch.
        payloads = [
            NotesPayload(
                chosen_row_label="Disclosure of borrowings",
                content=f"note {e.note_num} borrowings",
                evidence=f"p.{e.page_range[0]}",
                source_pages=[e.page_range[0]],
            )
            for e in batch
        ]
        return payloads, 0, 0, None

    with patch(
        "notes.listofnotes_subcoordinator._invoke_sub_agent_once",
        side_effect=fake_invoke,
    ):
        result = await run_notes_extraction(config, infopack=infopack)

    r = result.agent_results[0]
    assert r.status == "succeeded", r.error
    assert r.workbook_path and Path(r.workbook_path).exists(), (
        f"Expected workbook at {r.workbook_path}"
    )

    wb = openpyxl.load_workbook(r.workbook_path)
    ws = wb["Notes-Listofnotes"]

    def _find_row(needle: str) -> int:
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=1).value
            if v and needle.lower() in str(v).lower():
                return row
        raise AssertionError(f"Label containing '{needle}' not in sheet")

    rev_row = _find_row("Disclosure of borrowings")
    # Writer concatenated 5 sub-agent payloads (one per note) into one cell.
    content = ws.cell(row=rev_row, column=2).value
    assert isinstance(content, str)
    for expected in ["note 1 borrowings", "note 2 borrowings", "note 5 borrowings"]:
        assert expected in content
    wb.close()


@pytest.mark.asyncio
async def test_list_of_notes_writes_unmatched_side_log(tmp_path: Path):
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    pdf_path = tmp_path / "dummy.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run={NotesTemplateType.LIST_OF_NOTES},
        filing_level="company",
    )
    infopack = Infopack(
        toc_page=1, page_offset=0,
        notes_inventory=[NoteInventoryEntry(i, f"Note {i}", (20 + i, 20 + i)) for i in range(1, 6)],
    )

    async def fake_invoke(**kwargs):
        batch = kwargs["batch"]
        # Every payload lands in row 112 — all unmatched.
        payloads = [
            NotesPayload(
                chosen_row_label="Disclosure of other notes to accounts",
                content=f"note {e.note_num} weird",
                evidence=f"p.{e.page_range[0]}",
                source_pages=[e.page_range[0]],
            )
            for e in batch
        ]
        return payloads, 0, 0, None

    with patch(
        "notes.listofnotes_subcoordinator._invoke_sub_agent_once",
        side_effect=fake_invoke,
    ):
        await run_notes_extraction(config, infopack=infopack)

    side = tmp_path / "notes12_unmatched.json"
    assert side.exists(), "Expected notes12_unmatched.json side-log"
    data = json.loads(side.read_text(encoding="utf-8"))
    assert data["count"] == 5
    assert len(data["entries"]) == 5


# ---------------------------------------------------------------------------
# CLI exposure
# ---------------------------------------------------------------------------

def test_cli_accepts_list_of_notes_choice():
    """`run.py --notes list_of_notes` must be a valid CLI choice after C.2."""
    script = Path(__file__).resolve().parent.parent / "run.py"
    result = subprocess.run(
        [sys.executable, str(script), "--help"],
        capture_output=True, text=True, timeout=30,
    )
    assert result.returncode == 0
    assert "list_of_notes" in result.stdout, (
        "list_of_notes must be listed in --notes choices"
    )


# ---------------------------------------------------------------------------
# Server-side allowlist
# ---------------------------------------------------------------------------

def test_server_public_notes_allowlist_includes_list_of_notes():
    from server import _PUBLIC_NOTES_TEMPLATES
    assert NotesTemplateType.LIST_OF_NOTES in _PUBLIC_NOTES_TEMPLATES, (
        "Phase C wiring: server must expose LIST_OF_NOTES to the public API"
    )


@pytest.mark.asyncio
async def test_list_of_notes_surfaces_writer_warnings_on_success(tmp_path: Path):
    """Peer review finding 1: when the final writer succeeds overall but
    skipped some rows (e.g. unresolvable labels), those errors used to be
    dropped on the floor. They must now appear on the NotesAgentResult's
    warnings list and in the SSE complete event."""
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    pdf_path = tmp_path / "dummy.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run={NotesTemplateType.LIST_OF_NOTES},
        filing_level="company",
    )
    infopack = Infopack(
        toc_page=1, page_offset=0,
        notes_inventory=[NoteInventoryEntry(i, f"Note {i}", (20 + i, 20 + i)) for i in range(1, 4)],
    )

    async def fake_invoke(**kwargs):
        batch = kwargs["batch"]
        # One valid label + one that will fail to resolve.
        payloads = []
        for e in batch:
            payloads.append(NotesPayload(
                chosen_row_label="Disclosure of borrowings",
                content=f"note {e.note_num}",
                evidence=f"p.{e.page_range[0]}",
                source_pages=[e.page_range[0]],
            ))
            payloads.append(NotesPayload(
                chosen_row_label="zzzzzzz completely bogus row zzzzzzz",
                content=f"bogus {e.note_num}",
                evidence=f"p.{e.page_range[0]}",
                source_pages=[e.page_range[0]],
            ))
        return payloads, 0, 0, None

    emitted_events: list[dict] = []
    queue: asyncio.Queue = asyncio.Queue()

    async def drain():
        while True:
            evt = await queue.get()
            if evt is None:
                break
            emitted_events.append(evt)

    drain_task = asyncio.create_task(drain())

    with patch(
        "notes.listofnotes_subcoordinator._invoke_sub_agent_once",
        side_effect=fake_invoke,
    ):
        result = await run_notes_extraction(
            config, infopack=infopack, event_queue=queue,
        )

    await queue.put(None)
    await drain_task

    r = result.agent_results[0]
    assert r.status == "succeeded"  # partial coverage is still success
    assert r.warnings, "expected writer warnings on NotesAgentResult"
    assert any("bogus" in w.lower() for w in r.warnings), (
        f"bogus-label skip error should be in warnings, got {r.warnings}"
    )

    complete_events = [e for e in emitted_events if e.get("event") == "complete"]
    assert complete_events
    final = complete_events[-1]
    assert final["data"].get("success") is True
    assert final["data"].get("write_errors"), "write_errors must be in complete SSE"
    assert any(
        "bogus" in err.lower() for err in final["data"]["write_errors"]
    )


@pytest.mark.asyncio
async def test_total_failure_surfaces_as_failed_not_succeeded(tmp_path: Path):
    """Peer review #1: when inventory was non-empty but every sub-agent
    returned zero payloads, the fan-out runner must report failure, not
    take the empty-write success path that leaves an empty workbook."""
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    pdf_path = tmp_path / "dummy.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run={NotesTemplateType.LIST_OF_NOTES},
        filing_level="company",
    )
    infopack = Infopack(
        toc_page=1, page_offset=0,
        notes_inventory=[NoteInventoryEntry(i, f"Note {i}", (20, 21)) for i in range(1, 6)],
    )

    async def always_fails(**_kwargs):
        raise RuntimeError("every attempt blows up")

    with patch(
        "notes.listofnotes_subcoordinator._invoke_sub_agent_once",
        side_effect=always_fails,
    ):
        result = await run_notes_extraction(config, infopack=infopack)

    assert len(result.agent_results) == 1
    r = result.agent_results[0]
    assert r.status == "failed", (
        f"Expected failed when every sub-agent lost coverage, got {r.status}"
    )
    assert r.error and "sub-agent" in r.error.lower()


@pytest.mark.asyncio
async def test_partial_coverage_still_succeeds(tmp_path: Path):
    """Counterpart to the total-failure test: when SOME sub-agents produced
    payloads, the fan-out runner must still return succeeded (PLAN §4
    Checkpoint C: partial coverage is success)."""
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    pdf_path = tmp_path / "dummy.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run={NotesTemplateType.LIST_OF_NOTES},
        filing_level="company",
    )
    infopack = Infopack(
        toc_page=1, page_offset=0,
        notes_inventory=[NoteInventoryEntry(i, f"Note {i}", (20, 21)) for i in range(1, 6)],
    )

    call_count = {"n": 0}

    async def one_fails_rest_succeed(**kwargs):
        call_count["n"] += 1
        batch = kwargs["batch"]
        # First invocation fails on both attempts; rest succeed.
        if call_count["n"] <= 2:  # 1 initial + 1 retry from first sub-agent
            raise RuntimeError("one sub-agent is flaky")
        payloads = [
            NotesPayload(
                chosen_row_label="Disclosure of borrowings",
                content=f"note {e.note_num}",
                evidence=f"p.{e.page_range[0]}",
                source_pages=[e.page_range[0]],
            )
            for e in batch
        ]
        return payloads, 0, 0, None

    with patch(
        "notes.listofnotes_subcoordinator._invoke_sub_agent_once",
        side_effect=one_fails_rest_succeed,
    ):
        result = await run_notes_extraction(config, infopack=infopack)

    r = result.agent_results[0]
    assert r.status == "succeeded", r.error
