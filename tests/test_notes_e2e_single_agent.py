"""Phase B integration smoke tests.

These run the notes agent factory with a test model (no real LLM) and
confirm prompts load cleanly and the workbook write path exercises the
correct template for each single-agent sheet (11, 13, 14).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from notes.agent import create_notes_agent, render_notes_prompt
from notes.payload import NotesPayload
from notes.writer import write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path


@pytest.mark.parametrize("template_type,expected_sheet,expected_label_match", [
    (NotesTemplateType.ACC_POLICIES, "Notes-SummaryofAccPol",
     "Description of accounting policy for property, plant and equipment"),
    (NotesTemplateType.ISSUED_CAPITAL, "Notes-Issuedcapital", "Issued capital"),
    (NotesTemplateType.RELATED_PARTY, "Notes-RelatedPartytran", "Related party transactions"),
])
def test_prompt_and_template_alignment(template_type, expected_sheet, expected_label_match):
    prompt = render_notes_prompt(
        template_type=template_type,
        filing_level="company",
        inventory=[],
    )
    assert expected_sheet in prompt

    # Template must contain the expected label so the prompt's instructions
    # are actually achievable.
    tpl = notes_template_path(template_type, level="company")
    wb = openpyxl.load_workbook(tpl)
    ws = wb[expected_sheet]
    labels = [str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)]
    wb.close()
    assert any(expected_label_match.lower() in l.lower() for l in labels), (
        f"template {tpl.name} missing label matching '{expected_label_match}'"
    )


def test_accounting_policies_write_roundtrip(tmp_path: Path):
    out = tmp_path / "NOTES_ACC_POLICIES_filled.xlsx"
    tpl = notes_template_path(NotesTemplateType.ACC_POLICIES, level="company")
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=[
            NotesPayload(
                chosen_row_label="Description of accounting policy for property, plant and equipment",
                content="Items are stated at cost less accumulated depreciation and any impairment losses.",
                evidence="Page 32, Note 2.7",
                source_pages=[32],
            ),
            NotesPayload(
                chosen_row_label="Description of accounting policy for leases",
                content="The Group recognises a right-of-use asset and a lease liability at the commencement date.",
                evidence="Page 35, Note 2.12",
                source_pages=[35],
            ),
        ],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-SummaryofAccPol",
    )
    assert result.success, result.errors
    assert result.rows_written == 2

    wb = openpyxl.load_workbook(out)
    ws = wb["Notes-SummaryofAccPol"]

    def _find_row(needle: str) -> int:
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and needle.lower() in str(v).lower():
                return r
        raise AssertionError(f"{needle} not found")

    ppe_row = _find_row("property, plant and equipment")
    assert ws.cell(row=ppe_row, column=2).value.startswith("Items are stated at cost")
    assert ws.cell(row=ppe_row, column=4).value == "Page 32, Note 2.7"

    leases_row = _find_row("Description of accounting policy for leases")
    assert ws.cell(row=leases_row, column=2).value.startswith("The Group recognises")
    assert ws.cell(row=leases_row, column=4).value == "Page 35, Note 2.12"
    wb.close()
