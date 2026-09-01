"""Bug 5b/5c — fill_workbook safety nets.

5b: if the agent submits {row: N, col: M} WITHOUT a field_label, the write
currently lands wherever row N is — even if col A at row N is blank. That's
exactly how the MPERS SOCIE bug silently wrote values to rows 30/35/49 (no
labels in col A on the MPERS Company template). The writer must reject such
writes with a clear error, preserving the row-1 date-cell carve-out that
CLAUDE.md documents in `_base.md`.

5c: MPERS Group SOCIE has 4 vertical blocks divided by uncoloured text
headers ("Group - Current period", etc.). Without registering those as
section keywords, field_label + section hint cannot pick the right block
and label-based writes on Group filings default to block 1 silently. The
section-header keyword registry must recognise them.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from tools.fill_workbook import fill_workbook, _build_label_index


def _make_company_socie_like(tmp_path) -> str:
    """A minimal MPERS-Company-SOCIE-shaped workbook.

    Mirrors the real MPERS Company SOCIE: labels at rows 5, 10, 24; rows
    25-40 are empty (no col A text). Good enough to exercise the guard
    against the exact "write to row 30" bug the user hit.
    """
    path = str(tmp_path / "socie_like.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"
    ws["A5"] = "Equity at beginning of period"
    ws["A10"] = "Profit (loss)"
    ws["A24"] = "Equity at end of period"
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# 5b — writer guard
# ---------------------------------------------------------------------------

class TestWriterRejectsBlankRowWrites:
    def test_rejects_row_coord_write_when_col_a_is_blank(self, tmp_path):
        template = _make_company_socie_like(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        # Mirrors the MPERS bug — agent submitted {row: 30, col: 2}.
        facts = [
            {"sheet": "SOCIE", "row": 30, "col": 2, "value": 500_000},
        ]

        result = fill_workbook(template, output, facts)

        assert result.fields_written == 0, (
            "writer wrote to a row with no col-A label — this is the exact bug"
        )
        assert any("row 30" in e.lower() for e in result.errors), (
            f"error must name the offending row. Got: {result.errors}"
        )
        # Error message should be actionable — point at the absent label.
        assert any(
            ("blank" in e.lower() or "empty" in e.lower() or "no label" in e.lower())
            for e in result.errors
        ), f"error must explain why. Got: {result.errors}"
        # S-5: the error text must not falsely claim the row doesn't exist
        # in the template — the row is there, it just has no col-A label.
        assert not any(
            "row does not exist" in e.lower() for e in result.errors
        ), (
            "error message should describe the missing LABEL, not falsely "
            f"claim the row is absent. Got: {result.errors}"
        )

    def test_allows_row_1_write_for_date_cells(self, tmp_path):
        """Carve-out: row 1 date cells have no label by design (_base.md)."""
        template = _make_company_socie_like(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        facts = [
            {"sheet": "SOCIE", "row": 1, "col": 2, "value": "01/01/2024 - 31/12/2024"},
        ]

        result = fill_workbook(template, output, facts)

        assert result.success, f"row 1 write should succeed. Errors: {result.errors}"
        assert result.fields_written == 1

    def test_field_label_writes_still_work(self, tmp_path):
        """Regression guard — normal label-based writes unchanged."""
        template = _make_company_socie_like(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        facts = [
            {"sheet": "SOCIE", "field_label": "Profit (loss)", "col": 2, "value": 322_066},
        ]

        result = fill_workbook(template, output, facts)
        assert result.success
        assert result.fields_written == 1

        wb = openpyxl.load_workbook(output)
        assert wb["SOCIE"].cell(row=10, column=2).value == 322_066
        wb.close()

    def test_explicit_row_write_still_works_when_col_a_has_label(self, tmp_path):
        """Explicit row writes are still allowed — just not on blank rows.

        Some agent patterns legitimately use row coordinates (SOCIE MFRS
        matrix is the canonical one). The guard only kicks in when col A
        at that row is genuinely empty.
        """
        template = _make_company_socie_like(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        facts = [
            {"sheet": "SOCIE", "row": 10, "col": 2, "value": 1000},
        ]

        result = fill_workbook(template, output, facts)
        assert result.success
        assert result.fields_written == 1


# ---------------------------------------------------------------------------
# 5c — MPERS Group SOCIE block-header keywords
# ---------------------------------------------------------------------------

class TestMpersGroupSocieBlockHeaders:
    def test_build_label_index_recognises_mpers_group_block_headers(self, tmp_path):
        """Build a minimal MPERS-Group-SOCIE-shaped sheet with 2 blocks.

        Each block has an identical `Profit (loss)` label; the block header
        row (e.g. "Group - Current period") must register as a section so
        the duplicates can be disambiguated.
        """
        path = str(tmp_path / "group_socie.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SOCIE"
        # Block 1
        ws["A3"] = "Group - Current period"
        ws["A11"] = "Profit (loss)"
        # Block 2
        ws["A27"] = "Group - Prior period"
        ws["A35"] = "Profit (loss)"
        wb.save(path)
        wb.close()

        wb = openpyxl.load_workbook(path)
        idx = _build_label_index(wb)
        wb.close()

        # Both profit rows should register, each tagged with a different
        # section picked up from the block header above it.
        socie_entries = idx["SOCIE"]
        profit_entries = [e for e in socie_entries if e.normalized_label == "profit (loss)"]
        assert len(profit_entries) == 2, (
            f"expected two profit entries, one per block. Got {profit_entries}"
        )
        sections = {e.section for e in profit_entries}
        assert "group - current period" in sections, (
            f"block 1 header not registered as section. Sections: {sections}"
        )
        assert "group - prior period" in sections, (
            f"block 2 header not registered as section. Sections: {sections}"
        )

    def test_fill_workbook_disambiguates_mpers_group_blocks_by_section(self, tmp_path):
        """The happy path — same label, different block, section picks the row."""
        path = str(tmp_path / "group_socie.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SOCIE"
        ws["A3"] = "Group - Current period"
        ws["A11"] = "Profit (loss)"
        ws["A27"] = "Group - Prior period"
        ws["A35"] = "Profit (loss)"
        wb.save(path)
        wb.close()

        output = str(tmp_path / "filled.xlsx")
        facts = [
            {"sheet": "SOCIE", "field_label": "Profit (loss)",
             "section": "Group - Current period", "col": 2, "value": 100},
            {"sheet": "SOCIE", "field_label": "Profit (loss)",
             "section": "Group - Prior period", "col": 2, "value": 200},
        ]

        result = fill_workbook(path, output, facts)
        assert result.success, result.errors
        assert result.fields_written == 2

        wb = openpyxl.load_workbook(output)
        assert wb["SOCIE"].cell(row=11, column=2).value == 100
        assert wb["SOCIE"].cell(row=35, column=2).value == 200
        wb.close()

    def test_nested_header_does_not_erase_group_period_block_context(self, tmp_path):
        """The real MPERS layout nests Comprehensive income in each block."""
        path = str(tmp_path / "nested_group_socie.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SOCIE"
        ws["A3"] = "Group - Current period"
        ws["A9"] = "Comprehensive income"
        ws["A9"].fill = openpyxl.styles.PatternFill(
            fill_type="solid", fgColor="1F3864",
        )
        ws["A11"] = "Profit (loss)"
        ws["A27"] = "Group - Prior period"
        ws["A33"] = "Comprehensive income"
        ws["A33"].fill = openpyxl.styles.PatternFill(
            fill_type="solid", fgColor="1F3864",
        )
        ws["A35"] = "Profit (loss)"
        wb.save(path)
        wb.close()

        result = fill_workbook(path, str(tmp_path / "filled.xlsx"), [{
            "sheet": "SOCIE",
            "field_label": "Profit (loss)",
            "section": "Group - Prior period",
            "col": 2,
            "value": 200,
        }])

        assert result.success, result.errors
        wb = openpyxl.load_workbook(tmp_path / "filled.xlsx")
        assert wb["SOCIE"]["B11"].value is None
        assert wb["SOCIE"]["B35"].value == 200
        wb.close()

    def test_duplicate_without_context_fails_closed(self, tmp_path):
        path = str(tmp_path / "ambiguous.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SOCIE"
        ws["A3"] = "Group - Current period"
        ws["A11"] = "Profit (loss)"
        ws["A27"] = "Group - Prior period"
        ws["A35"] = "Profit (loss)"
        wb.save(path)
        wb.close()

        result = fill_workbook(path, str(tmp_path / "filled.xlsx"), [{
            "sheet": "SOCIE",
            "field_label": "Profit (loss)",
            "col": 2,
            "value": 200,
        }])

        assert result.success is False
        assert result.fields_written == 0
        assert "ambiguous" in " ".join(result.errors).lower()


def test_managed_mfrs_socie_allows_exact_later_block_coordinate(tmp_path):
    template = (
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MFRS/Company/09-SOCIE.xlsx"
    )
    output = tmp_path / "filled.xlsx"

    result = fill_workbook(str(template), str(output), [{
        "sheet": "SOCIE", "row": 35, "col": 2, "value": 116_035,
    }])

    assert result.success, result.errors
    wb = openpyxl.load_workbook(output, data_only=False)
    assert wb["SOCIE"]["B35"].value == 116_035
    wb.close()


def test_managed_socie_writability_is_coordinate_not_row_only(tmp_path):
    template = (
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MFRS/Company/09-SOCIE.xlsx"
    )
    result = fill_workbook(str(template), str(tmp_path / "filled.xlsx"), [{
        "sheet": "SOCIE", "row": 35, "col": 25, "value": 116_035,
    }])

    assert result.success is False
    assert result.fields_written == 0
    assert "non-entry" in " ".join(result.errors)


def test_valid_row_with_non_value_column_blames_the_column(tmp_path):
    template = (
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MFRS/Company/01-SOFP-CuNonCu.xlsx"
    )
    result = fill_workbook(str(template), str(tmp_path / "filled.xlsx"), [{
        "sheet": "SOFP-CuNonCu", "row": 10, "col": 4, "value": "source text",
    }])

    message = " ".join(result.errors)
    assert result.success is False
    assert result.fields_written == 0
    assert "row 10" in message
    assert "column D" in message
    assert "row 10 ('Biological assets') is a heading" not in message


def test_live_company_template_allows_current_and_prior_year_slots(tmp_path):
    """Linear manifests are row-complete but not yet column-complete."""
    template = (
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MFRS/Company/01-SOFP-CuNonCu.xlsx"
    )
    output = tmp_path / "filled.xlsx"

    result = fill_workbook(str(template), str(output), [
        {"sheet": "SOFP-CuNonCu", "row": 10, "col": 2, "value": 100},
        {"sheet": "SOFP-CuNonCu", "row": 10, "col": 3, "value": 90},
    ])

    assert result.success, result.errors
    assert result.fields_written == 2
    wb = openpyxl.load_workbook(output, data_only=False)
    assert wb["SOFP-CuNonCu"]["B10"].value == 100
    assert wb["SOFP-CuNonCu"]["C10"].value == 90
    wb.close()


def test_live_group_template_allows_all_four_period_scope_slots(tmp_path):
    template = (
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MFRS/Group/01-SOFP-CuNonCu.xlsx"
    )
    output = tmp_path / "filled.xlsx"

    result = fill_workbook(str(template), str(output), [
        {"sheet": "SOFP-CuNonCu", "row": 10, "col": 2, "value": 400},
        {"sheet": "SOFP-CuNonCu", "row": 10, "col": 3, "value": 300},
        {"sheet": "SOFP-CuNonCu", "row": 10, "col": 4, "value": 200},
        {"sheet": "SOFP-CuNonCu", "row": 10, "col": 5, "value": 100},
    ], filing_level="group")

    assert result.success, result.errors
    assert result.fields_written == 4


def test_live_mpers_group_socie_routes_nested_prior_block_by_section(tmp_path):
    template = (
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MPERS/Group/09-SOCIE.xlsx"
    )
    output = tmp_path / "filled.xlsx"

    result = fill_workbook(str(template), str(output), [{
        "sheet": "SOCIE",
        "field_label": "Profit (loss)",
        "section": "Group - Prior period",
        "col": 2,
        "value": 116_035,
    }])

    assert result.success, result.errors
    wb = openpyxl.load_workbook(output, data_only=False)
    assert wb["SOCIE"]["B11"].value is None
    assert wb["SOCIE"]["B35"].value == 116_035
    wb.close()


def test_incremental_write_keeps_managed_template_writability_contract(tmp_path):
    """A scratch workbook is mutable state, not the filing-target contract.

    Run 103 reproduced this on SOCF: the first write loaded the managed
    template manifest, while the second write loaded the scratch workbook and
    lost that manifest.  The duplicate closing-cash label then became
    ambiguous between the writable statement row and the protected formula
    reconciliation row.
    """
    canonical_template = (
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MFRS/Company/07-SOCF-Indirect.xlsx"
    )
    scratch = tmp_path / "SOCF_filled.xlsx"

    first = fill_workbook(
        str(canonical_template),
        str(scratch),
        [{
            "sheet": "SOCF-Indirect",
            "field_label": "Cash and cash equivalents at beginning of period",
            "col": 2,
            "value": 100,
        }],
    )
    assert first.success, first.errors

    second = fill_workbook(
        str(scratch),
        str(scratch),
        [{
            "sheet": "SOCF-Indirect",
            "field_label": "Cash and cash equivalents at end of period",
            "col": 2,
            "value": 200,
        }],
        canonical_template_path=str(canonical_template),
    )

    assert second.success, second.errors
    assert second.fields_written == 1
    wb = openpyxl.load_workbook(scratch, data_only=False)
    assert wb["SOCF-Indirect"]["B132"].value == 200
    assert str(wb["SOCF-Indirect"]["B137"].value).startswith("=")
    wb.close()
