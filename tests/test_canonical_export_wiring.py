"""Phase C — DB-backed workbook export wiring.

In canonical mode the downloaded workbook must be exported from
run_concept_facts (the authoritative store), not the agent-written scratch
xlsx. `server._export_canonical_workbooks` does this per succeeded statement:
copy the master template, fill it from the DB via export_run_to_xlsx, and point
the merge input at it. These tests pin that the helper repoints
all_workbook_paths and that the exported cell carries the DB fact.
"""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from coordinator import AgentResult
from statement_types import StatementType


REPO = Path(__file__).resolve().parent.parent
CO_SOFP = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"


@pytest.fixture
def seeded(tmp_path: Path):
    from db.schema import init_db
    from concept_model.parser import parse_template, _derive_template_id
    from concept_model.importer import import_template, import_company_targets
    from concept_model.facts_api import write_fact, FactWrite

    db_path = tmp_path / "xbrl.db"
    init_db(db_path)
    tree = parse_template(str(CO_SOFP))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    import_template(db_path, jp)
    template_id = _derive_template_id(CO_SOFP)
    # Phase 6.1: the exporter routes every fact via a single concept_targets
    # lookup and RAISES on an applicable fact with no target (CLAUDE.md gotcha
    # #21). A hand-rolled Company DB must precompute its targets.
    import_company_targets(db_path, template_id)

    conn = sqlite3.connect(str(db_path))
    run_id = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
        "VALUES (?,?,?,?)",
        ("2026-05-25T00:00:00Z", "x.pdf", "running", "2026-05-25T00:00:00Z"),
    ).lastrowid
    leaf = conn.execute(
        "SELECT concept_uuid, render_sheet, render_row FROM concept_nodes "
        "WHERE template_id=? AND render_sheet='SOFP-CuNonCu' AND kind='LEAF' "
        "ORDER BY render_row LIMIT 1",
        (template_id,),
    ).fetchone()
    conn.commit()
    conn.close()

    write_fact(db_path, run_id, FactWrite(
        concept_uuid=leaf[0], value=8888.0, value_status="observed"))
    return db_path, run_id, leaf[1], int(leaf[2])


def test_export_repoints_and_fills_from_db(seeded, tmp_path):
    import server

    db_path, run_id, sheet, row = seeded
    session_dir = tmp_path / "session"
    session_dir.mkdir()

    # Simulate the agent-written scratch workbook the merge would otherwise use.
    scratch = session_dir / "SOFP_filled.xlsx"
    scratch.write_bytes(CO_SOFP.read_bytes())
    all_workbook_paths = {StatementType.SOFP: str(scratch)}

    agent_results = [AgentResult(
        statement_type=StatementType.SOFP, variant="CuNonCu",
        status="succeeded", workbook_path=str(scratch))]

    exported = server._export_canonical_workbooks(
        run_id=run_id,
        agent_results=agent_results,
        all_workbook_paths=all_workbook_paths,
        session_dir=session_dir,
        filing_level="company",
        filing_standard="mfrs",
        db_path=db_path,
    )

    assert StatementType.SOFP in exported
    # all_workbook_paths now points at the canonical export, not the scratch.
    new_path = all_workbook_paths[StatementType.SOFP]
    assert new_path != str(scratch)
    assert Path(new_path).exists()

    wb = openpyxl.load_workbook(new_path, data_only=False)
    assert wb[sheet][f"B{row}"].value == 8888.0


def test_export_preserves_live_formula_for_itemised_computed(seeded, tmp_path):
    """A cascade-computed (itemised) COMPUTED parent must NOT clobber the
    template's live formula with a literal — only aggregate_only does that
    (PRD design decision #19 / gotcha)."""
    from concept_model.exporter import export_run_to_xlsx
    import shutil

    db_path, run_id, _sheet, _row = seeded

    conn = sqlite3.connect(str(db_path))
    try:
        # Find a COMPUTED concept on the face sheet whose render cell holds a
        # live formula in the master template.
        wb_master = openpyxl.load_workbook(str(CO_SOFP), data_only=False)
        computed = None
        for cu, sheet, row, col in conn.execute(
            "SELECT concept_uuid, render_sheet, render_row, render_col "
            "FROM concept_nodes WHERE kind='COMPUTED'"
        ).fetchall():
            if sheet not in wb_master.sheetnames:
                continue
            cell = wb_master[sheet][f"{col or 'B'}{row}"]
            if isinstance(cell.value, str) and cell.value.startswith("="):
                computed = (cu, sheet, int(row), col or "B")
                break
        assert computed, "no COMPUTED concept with a live formula cell found"
        cu, sheet, row, col = computed
        # Insert an itemised observed fact exactly as the cascade would.
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, children_status, source, "
            "updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
            (run_id, cu, "CY", "Company", 12345.0, "observed", "itemised",
             "cascade", "2026-05-25T00:00:00Z"),
        )
        conn.commit()
    finally:
        conn.close()

    work = tmp_path / "filled.xlsx"
    shutil.copyfile(CO_SOFP, work)
    export_run_to_xlsx(db_path, run_id, str(work), filing_level="company")

    wb = openpyxl.load_workbook(str(work), data_only=False)
    cell = wb[sheet][f"{col}{row}"]
    assert isinstance(cell.value, str) and cell.value.startswith("="), (
        f"itemised COMPUTED formula was clobbered: {cell.value!r}"
    )


def test_export_returns_applied_count_and_scopes_by_template(seeded, tmp_path):
    """export_run_to_xlsx returns how many facts it applied, and a
    template_id filter scopes which facts (and which unmapped check) apply."""
    from concept_model.exporter import export_run_to_xlsx
    from concept_model.parser import _derive_template_id
    import shutil

    db_path, run_id, _sheet, _row = seeded
    tid = _derive_template_id(CO_SOFP)
    work = tmp_path / "filled.xlsx"
    shutil.copyfile(CO_SOFP, work)

    applied = export_run_to_xlsx(
        db_path, run_id, str(work), filing_level="company", template_id=tid)
    assert applied == 1  # the single seeded leaf fact


def test_export_zero_facts_returns_zero(seeded, tmp_path):
    """A template with no facts for the run applies nothing (so the caller
    knows not to repoint to a blank workbook)."""
    from concept_model.exporter import export_run_to_xlsx
    from concept_model.parser import _derive_template_id
    import shutil

    db_path, run_id, _sheet, _row = seeded
    # SOPL template id — the run only has SOFP facts.
    sopl = REPO / "XBRL-template-MFRS" / "Company" / "03-SOPL-Function.xlsx"
    sopl_tid = _derive_template_id(sopl)
    work = tmp_path / "sopl.xlsx"
    shutil.copyfile(sopl, work)
    applied = export_run_to_xlsx(
        db_path, run_id, str(work), filing_level="company", template_id=sopl_tid)
    assert applied == 0


def test_export_keeps_scratch_when_no_facts(seeded, tmp_path):
    """_export_canonical_workbooks must NOT repoint a succeeded statement to a
    blank template when the DB has no facts for it (finding 1)."""
    import server

    db_path, run_id, _sheet, _row = seeded
    session_dir = tmp_path / "session"
    session_dir.mkdir()
    scratch = session_dir / "SOPL_filled.xlsx"
    scratch.write_bytes((REPO / "XBRL-template-MFRS" / "Company" / "03-SOPL-Function.xlsx").read_bytes())
    all_workbook_paths = {StatementType.SOPL: str(scratch)}

    agent_results = [AgentResult(
        statement_type=StatementType.SOPL, variant="Function",
        status="succeeded", workbook_path=str(scratch))]

    exported = server._export_canonical_workbooks(
        run_id=run_id, agent_results=agent_results,
        all_workbook_paths=all_workbook_paths, session_dir=session_dir,
        filing_level="company", filing_standard="mfrs", db_path=db_path)

    assert StatementType.SOPL not in exported
    # Scratch workbook is preserved — no blank-template clobber.
    assert all_workbook_paths[StatementType.SOPL] == str(scratch)


def test_group_export_wiring_fills_both_scope_columns(tmp_path):
    """A Group statement exported via the wiring lands Group CY in col B and
    Company CY in col D (the 6-col layout), proving Group routing flows
    through _export_canonical_workbooks (Phase E)."""
    import server
    from db.schema import init_db
    from concept_model.parser import parse_template, _derive_template_id
    from concept_model.importer import (
    import_company_targets,
    import_group_targets,
    import_template,
)
    from concept_model.facts_api import write_fact, FactWrite

    gr_sofp = REPO / "XBRL-template-MFRS" / "Group" / "01-SOFP-CuNonCu.xlsx"
    db_path = tmp_path / "xbrl.db"
    init_db(db_path)
    tree = parse_template(str(gr_sofp))
    jp = tmp_path / "g.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db_path, jp)
    import_group_targets(db_path, tid)

    conn = sqlite3.connect(str(db_path))
    run_id = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
        "VALUES (?,?,?,?)",
        ("2026-05-25T00:00:00Z", "x.pdf", "running", "2026-05-25T00:00:00Z"),
    ).lastrowid
    # A leaf with a Group target at col B.
    tgt = conn.execute(
        "SELECT ct.concept_uuid, ct.target_sheet, ct.target_row FROM concept_targets ct "
        "JOIN concept_nodes n ON n.concept_uuid=ct.concept_uuid "
        "WHERE n.template_id=? AND ct.target_col='B' AND n.kind='LEAF' LIMIT 1",
        (tid,)).fetchone()
    uid, sheet, row = tgt[0], tgt[1], int(tgt[2])
    conn.commit()
    conn.close()

    write_fact(db_path, run_id, FactWrite(
        concept_uuid=uid, period="CY", entity_scope="Group", value=111.0,
        value_status="observed"))
    write_fact(db_path, run_id, FactWrite(
        concept_uuid=uid, period="CY", entity_scope="Company", value=222.0,
        value_status="observed"))

    session_dir = tmp_path / "session"
    session_dir.mkdir()
    all_workbook_paths = {StatementType.SOFP: str(session_dir / "SOFP_filled.xlsx")}
    server._export_canonical_workbooks(
        run_id=run_id,
        agent_results=[AgentResult(statement_type=StatementType.SOFP,
                                   variant="CuNonCu", status="succeeded")],
        all_workbook_paths=all_workbook_paths, session_dir=session_dir,
        filing_level="group", filing_standard="mfrs", db_path=db_path)

    wb = openpyxl.load_workbook(all_workbook_paths[StatementType.SOFP], data_only=False)
    assert wb[sheet][f"B{row}"].value == 111.0  # Group CY
    assert wb[sheet][f"D{row}"].value == 222.0  # Company CY


def test_open_conflict_count_excludes_sentinel(seeded):
    """_open_conflict_count counts open reconciliation conflicts but ignores
    the correction_exhausted sentinel (that has its own run status)."""
    import server

    db_path, run_id, _sheet, _row = seeded
    conn = sqlite3.connect(str(db_path))
    try:
        # One real open conflict + one resolved + one exhaustion sentinel.
        conn.execute(
            "INSERT INTO run_concept_conflicts(run_id, concept_uuid, period, "
            "entity_scope, kind, status, created_at) VALUES "
            "(?,?,?,?,?,?,?)",
            (run_id, "c1", "CY", "Company", "partial_state", "open", "t"))
        conn.execute(
            "INSERT INTO run_concept_conflicts(run_id, concept_uuid, period, "
            "entity_scope, kind, status, created_at) VALUES "
            "(?,?,?,?,?,?,?)",
            (run_id, "c2", "CY", "Company", "partial_state", "resolved", "t"))
        conn.execute(
            "INSERT INTO run_concept_conflicts(run_id, concept_uuid, period, "
            "entity_scope, kind, status, created_at) VALUES "
            "(?,?,?,?,?,?,?)",
            (run_id, "", "", "", "correction_exhausted", "open", "t"))
        conn.commit()
    finally:
        conn.close()

    assert server._open_conflict_count(db_path, run_id) == 1


def test_export_skips_failed_and_skipped_statements(seeded, tmp_path):
    import server

    db_path, run_id, _sheet, _row = seeded
    session_dir = tmp_path / "session"
    session_dir.mkdir()
    all_workbook_paths: dict = {}

    agent_results = [
        AgentResult(statement_type=StatementType.SOPL, variant="Function",
                    status="failed", error="boom"),
        AgentResult(statement_type=StatementType.SOCI, variant="NotPrepared",
                    status="skipped"),
    ]
    exported = server._export_canonical_workbooks(
        run_id=run_id,
        agent_results=agent_results,
        all_workbook_paths=all_workbook_paths,
        session_dir=session_dir,
        filing_level="company",
        filing_standard="mfrs",
        db_path=db_path,
    )
    assert exported == []
    assert all_workbook_paths == {}


def test_event_sink_fires_when_a_statement_export_fails(seeded, tmp_path, monkeypatch):
    """Option C (2026-06-03): when a per-statement export RAISES and the helper
    silently keeps the agent's scratch workbook, the download no longer reflects
    the DB facts. A caller-supplied event_sink must be told, so the live SSE
    stream surfaces the degradation loudly instead of only logging it."""
    import server

    db_path, run_id, _sheet, _row = seeded
    session_dir = tmp_path / "session"
    session_dir.mkdir()
    scratch = session_dir / "SOFP_filled.xlsx"
    scratch.write_bytes(CO_SOFP.read_bytes())
    all_workbook_paths = {StatementType.SOFP: str(scratch)}
    agent_results = [AgentResult(
        statement_type=StatementType.SOFP, variant="CuNonCu",
        status="succeeded", workbook_path=str(scratch))]

    # Force the per-statement export to raise.
    def _boom(*a, **k):
        raise RuntimeError("export blew up")
    monkeypatch.setattr(server, "export_run_to_xlsx", _boom, raising=False)
    # The helper imports export_run_to_xlsx locally from the module, so patch
    # there too.
    import concept_model.exporter as _exp
    monkeypatch.setattr(_exp, "export_run_to_xlsx", _boom)

    events: list = []
    exported = server._export_canonical_workbooks(
        run_id=run_id,
        agent_results=agent_results,
        all_workbook_paths=all_workbook_paths,
        session_dir=session_dir,
        filing_level="company",
        filing_standard="mfrs",
        db_path=db_path,
        event_sink=events.append,
    )

    # Export failed → scratch kept, and the degradation was surfaced.
    assert exported == []
    assert all_workbook_paths[StatementType.SOFP] == str(scratch)
    assert len(events) == 1
    assert events[0]["type"] == "canonical_export_degraded"
    assert "SOFP" in events[0]["message"]
