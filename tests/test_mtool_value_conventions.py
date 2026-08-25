"""Step 5 — pin what our stored values MEAN before translating them.

A translation layer is only as good as its understanding of the input. Before
any scale or sign rule is written, these tests pin the two claims the mTool
exporter's docstring makes about ``run_concept_facts``:

1. **Sign** — the stored sign is already the sign the SSM template's own
   formulas expect (positive expenses in SOPL, per-row formula-ready signs in
   SOCF, positive dividends in SOCIE because the subtotal subtracts the row).
2. **Nothing rescales at extraction time** — a statement printed in RM'000
   stores the RM'000 figure, and the run's ``denomination`` is what says so.

Pure reading + pinning; no behaviour of its own. If a future change flips a
sign at extraction time, these fail here — where it's cheap — rather than in a
filed workbook.
"""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from concept_model.importer import import_company_targets, import_template
from concept_model.parser import parse_template
from db.schema import init_db
from mtool.exporter import build_fill_doc
from mtool.translation import IDENTITY
from mtool.units import NUMERIC_CLASSES

REPO = Path(__file__).resolve().parent.parent
SOCF_INDIRECT = REPO / "XBRL-template-MFRS" / "Company" / "07-SOCF-Indirect.xlsx"


# --------------------------------------------------------------- sign claims

@pytest.mark.parametrize("workbook", [
    "XBRL-template-MFRS/Company/09-SOCIE.xlsx",
    "XBRL-template-MFRS/Group/09-SOCIE.xlsx",
    "XBRL-template-MPERS/Company/10-SoRE.xlsx",
])
def test_socie_subtotal_subtracts_dividends_so_positive_storage_is_right(
        workbook: str):
    """ADR-002, restated as an mTool precondition.

    We store dividends POSITIVE. That is only correct while the template's
    "Total increase (decrease) in equity" formula SUBTRACTS the row — so the
    exporter's decision to emit them unchanged rests on this formula, and this
    test is the tripwire if a regenerated template ever adds instead.
    """
    wb = openpyxl.load_workbook(REPO / workbook)
    try:
        ws = wb[wb.sheetnames[0]]
        div_rows = [
            r for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str)
            and "dividend" in ws.cell(r, 1).value.lower()
            and "paid" in ws.cell(r, 1).value.lower()
        ]
        assert div_rows, f"{workbook}: no 'dividends paid' row found"
        formulas = [
            str(ws.cell(r, c).value)
            for r in range(1, ws.max_row + 1)
            for c in range(2, min(ws.max_column, 30) + 1)
            if isinstance(ws.cell(r, c).value, str)
            and ws.cell(r, c).value.startswith("=")
        ]
        for div_row in div_rows:
            refs = [f"-1*{chr(64 + c)}{div_row}" for c in range(2, 30)]
            refs += [f"-{chr(64 + c)}{div_row}" for c in range(2, 30)]
            assert any(ref in f for f in formulas for ref in refs), (
                f"{workbook} row {div_row}: no subtotal subtracts it — the "
                "positive-dividend storage convention no longer holds")
    finally:
        wb.close()


@pytest.mark.parametrize(
    "label_fragment,stored_value,formula_coefficient",
    [
        # This row is SUBTRACTED by the investing total, so the template-ready
        # fact is a positive magnitude even though it is a cash outflow.
        ("purchase of property, plant and equipment", 250.0, -1),
        # This row is ADDED by the financing total, so it carries the negative
        # cash-direction sign directly.
        ("cash payments for the principal portion", -250.0, 1),
    ],
)
def test_socf_mixed_formula_ready_signs_are_emitted_unchanged(
    tmp_path,
    label_fragment: str,
    stored_value: float,
    formula_coefficient: int,
):
    """mTool receives the final SSM-template sign without another sign flip."""
    db = tmp_path / "xbrl.db"
    init_db(db)
    tree = parse_template(str(SOCF_INDIRECT))
    matching_nodes = [
        node
        for node in tree.concepts
        if node.kind == "LEAF" and label_fragment in node.canonical_label.lower()
    ]
    assert len(matching_nodes) == 1
    target_row = int(matching_nodes[0].render_key["row"])
    parent_edges = [
        edge
        for node in tree.concepts
        for edge in node.edges
        if int(edge["ref"].get("row", -1)) == target_row
    ]
    assert any(
        int(edge["coefficient"]) == formula_coefficient for edge in parent_edges
    ), (
        f"{label_fragment!r} no longer has coefficient {formula_coefficient}; "
        "the pinned stored sign must be reviewed"
    )
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db, jp)
    import_company_targets(db, tid)

    conn = sqlite3.connect(str(db))
    try:
        run_id = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
            "VALUES (?,?,?,?)",
            ("2026-07-27T00:00:00Z", "x.pdf", "completed",
             "2026-07-27T00:00:00Z")).lastrowid
        leaf = conn.execute(
            "SELECT concept_uuid, canonical_label FROM concept_nodes "
            "WHERE kind='LEAF' AND lower(canonical_label) LIKE ? LIMIT 1",
            (f"%{label_fragment}%",),
        ).fetchone()
        assert leaf is not None, f"no SOCF row matching {label_fragment!r}"
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (
                run_id,
                leaf[0],
                "CY",
                "Company",
                stored_value,
                "observed",
                "2026-07-27T00:00:00Z",
            ),
        )
        conn.commit()
    finally:
        conn.close()

    doc = build_fill_doc(db, run_id, filing_standard="mfrs",
                         filing_level="company")
    assert doc["writes"] == [{
        "sheet": doc["writes"][0]["sheet"], "label": leaf[1],
        "column_role": "current_year", "value": int(stored_value)}]


def test_identity_manifest_flips_no_signs():
    """The shipped default asserts sign=+1 for EVERY unit class.

    That is a claim about mTool honouring SSM's own formula conventions, and it
    is exactly what the Windows acceptance run (Step 7) has to confirm before
    any sign rule may be added.
    """
    for unit_class in NUMERIC_CLASSES:
        rule = IDENTITY.units[unit_class]
        assert rule.sign == 1, unit_class
        assert rule.scale == 1.0, unit_class


# ---------------------------------------------------------------- unit claim

def test_denomination_is_reported_not_applied(tmp_path):
    """A 'thousands' run emits the thousands figure verbatim and SAYS so.

    Nothing rescales at extraction time, so the denomination is metadata about
    the stored value, not an instruction the exporter has already carried out.
    """
    db = tmp_path / "xbrl.db"
    init_db(db)
    sofp = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"
    tree = parse_template(str(sofp))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db, jp)
    import_company_targets(db, tid)

    conn = sqlite3.connect(str(db))
    try:
        run_id = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
            "VALUES (?,?,?,?)",
            ("2026-07-27T00:00:00Z", "x.pdf", "completed",
             "2026-07-27T00:00:00Z")).lastrowid
        leaf = conn.execute(
            "SELECT concept_uuid FROM concept_nodes WHERE kind='LEAF' "
            "AND render_sheet LIKE '%Sub%' LIMIT 1").fetchone()
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, updated_at) "
            "VALUES (?,?,?,?,?,?,?)",
            (run_id, leaf[0], "CY", "Company", 1595.0, "observed",
             "2026-07-27T00:00:00Z"))
        conn.commit()
    finally:
        conn.close()

    doc = build_fill_doc(db, run_id, filing_standard="mfrs",
                         filing_level="company", denomination="thousands")
    assert doc["writes"][0]["value"] == 1595
    assert doc["meta"]["denomination"] == "thousands"
    assert doc["meta"]["translation_version"] == "identity-1"
