"""Phase 4.3 — reporting-period date headers are populated deterministically
from the scout-captured period, NOT by the extraction agents.

The period is run-level metadata: the same CY / PY string on every face
statement, repeated across Group's column pairs. So the exporter stamps it onto
the placeholder date-header cells by column parity (CY -> B/D, PY -> C/E),
keyed on the "YYYY" placeholder so it works on every layout:

  * Company  -> dates in ROW 1   (B=CY, C=PY)
  * Group    -> dates in ROW 2   (B=GrpCY, C=GrpPY, D=CoCY, E=CoPY); row 1
                holds the Group/Company column-group labels
  * SOCIE    -> B1 only

The Group case is the bug this fixes: Group dates live in row 2, a labelless
row the writer guard won't let an agent fill, so an agent could never populate
them. Scout metadata sidesteps the agent/writer path entirely.

Scout-metadata takes priority; the agent's scratch workbook remains a fallback
for no-scout runs (covered in test_canonical_export.py).
"""
from __future__ import annotations

import json
import shutil
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from concept_model.exporter import export_run_to_xlsx
from concept_model.importer import (
    import_company_targets,
    import_group_targets,
    import_template,
)
from concept_model.parser import parse_template, _derive_template_id
from db.schema import init_db


REPO = Path(__file__).resolve().parent.parent
CO_SOFP = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"
GRP_SOFP = REPO / "XBRL-template-MFRS" / "Group" / "01-SOFP-CuNonCu.xlsx"

CY = "01/01/2021 - 31/12/2021"
PY = "01/01/2020 - 31/12/2020"


def _seed(tmp_path: Path, fixture: Path, level: str):
    db = tmp_path / "xbrl.db"
    init_db(db)
    tree = parse_template(str(fixture))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db, jp)
    if level == "group":
        import_group_targets(db, tid)
    else:
        import_company_targets(db, tid)
    conn = sqlite3.connect(str(db))
    run_id = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
        "VALUES ('2026-05-30Z','x.pdf','running','2026-05-30Z')"
    ).lastrowid
    conn.commit()
    conn.close()
    work = tmp_path / "filled.xlsx"
    shutil.copyfile(fixture, work)
    return db, run_id, work


def _date_cells(ws, row):
    return {c: ws[f"{c}{row}"].value for c in "BCDE"}


def test_company_dates_synthesized_into_row1(tmp_path: Path):
    db, run_id, work = _seed(tmp_path, CO_SOFP, "company")
    export_run_to_xlsx(
        db, run_id, str(work), filing_level="company",
        reporting_period_cy=CY, reporting_period_py=PY,
    )
    ws = openpyxl.load_workbook(str(work))["SOFP-CuNonCu"]
    # Company dates live in row 1: B=CY, C=PY.
    assert ws["B1"].value == CY
    assert ws["C1"].value == PY


@pytest.mark.skipif(not GRP_SOFP.exists(), reason="group template missing")
def test_group_dates_synthesized_into_row2_by_parity(tmp_path: Path):
    """The headline fix: Group dates are in ROW 2 and the agent can't write
    them. Scout metadata must land them — CY in B & D, PY in C & E."""
    db, run_id, work = _seed(tmp_path, GRP_SOFP, "group")
    # Precondition: row 2 starts as placeholders, row 1 holds the labels.
    pre = openpyxl.load_workbook(str(work))["SOFP-CuNonCu"]
    assert "YYYY" in str(pre["B2"].value)
    assert pre["B1"].value == "Group" and pre["D1"].value == "Company"

    export_run_to_xlsx(
        db, run_id, str(work), filing_level="group",
        reporting_period_cy=CY, reporting_period_py=PY,
    )
    ws = openpyxl.load_workbook(str(work))["SOFP-CuNonCu"]
    got = _date_cells(ws, 2)
    assert got["B"] == CY, got   # Group CY
    assert got["C"] == PY, got   # Group PY
    assert got["D"] == CY, got   # Company CY
    assert got["E"] == PY, got   # Company PY
    # Row 1 column-group labels are untouched (not date placeholders).
    assert ws["B1"].value == "Group"
    assert ws["D1"].value == "Company"


def test_synthesis_only_overwrites_placeholders(tmp_path: Path):
    """A non-placeholder cell in the scan band is never clobbered."""
    db, run_id, work = _seed(tmp_path, CO_SOFP, "company")
    wb = openpyxl.load_workbook(str(work))
    ws = wb["SOFP-CuNonCu"]
    ws["A3"] = "Statement of financial position"  # real label, no YYYY
    wb.save(str(work))
    export_run_to_xlsx(
        db, run_id, str(work), filing_level="company",
        reporting_period_cy=CY, reporting_period_py=PY,
    )
    ws2 = openpyxl.load_workbook(str(work))["SOFP-CuNonCu"]
    assert ws2["A3"].value == "Statement of financial position"


def test_no_metadata_leaves_placeholder_when_no_scratch(tmp_path: Path):
    """No scout dates and no scratch fallback -> placeholder stays (graceful)."""
    db, run_id, work = _seed(tmp_path, CO_SOFP, "company")
    export_run_to_xlsx(db, run_id, str(work), filing_level="company")
    ws = openpyxl.load_workbook(str(work))["SOFP-CuNonCu"]
    assert "YYYY" in str(ws["B1"].value)


def test_reporting_periods_from_infopack_helper():
    """server._reporting_periods_from_infopack accepts the serialised dict
    (download/recheck) and the Infopack object (live pipeline), and coerces
    blanks/None to None so the exporter falls back cleanly."""
    import server
    f = server._reporting_periods_from_infopack
    assert f({"reporting_period_cy": CY, "reporting_period_py": PY}) == (CY, PY)
    assert f(None) == (None, None)
    assert f({"reporting_period_cy": "  ", "reporting_period_py": PY}) == (None, PY)

    class _IP:
        reporting_period_cy = CY
        reporting_period_py = None
    assert f(_IP()) == (CY, None)
