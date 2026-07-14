"""API tests for the benchmark + gold + eval endpoints (api/eval.py).

Spins up the FastAPI app against a temp DB with the live MFRS Company SOFP
template imported, then exercises the full benchmark lifecycle: create via
upload → list → detail → gold grid → spot-edit → run scorecard → delete.
"""
from __future__ import annotations

import json
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl.utils import column_index_from_string

from statement_types import StatementType, template_path


def _import_company_sofp(db_path) -> str:
    from concept_model.importer import import_template, import_company_targets
    from concept_model.parser import parse_template

    tpath = template_path(StatementType.SOFP, "CuNonCu", "company", "mfrs")
    tree = parse_template(str(tpath))
    payload = tree.to_json()
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as fh:
        json.dump(payload, fh, sort_keys=True)
        json_path = fh.name
    try:
        template_id = import_template(db_path, json_path)
    finally:
        Path(json_path).unlink(missing_ok=True)
    import_company_targets(db_path, template_id)
    return template_id


def _fixture_workbook_bytes(db_path, template_id, n=5):
    """Blank a copy of the live template, fill n known leaves, return bytes +
    the written (uuid, value) list."""
    conn = sqlite3.connect(str(db_path))
    leaves = conn.execute(
        "SELECT concept_uuid, render_sheet, render_row, render_col "
        "FROM concept_nodes WHERE template_id = ? AND kind = 'LEAF' "
        "ORDER BY render_sheet, render_row LIMIT ?",
        (template_id, n),
    ).fetchall()
    conn.close()
    tpath = template_path(StatementType.SOFP, "CuNonCu", "company", "mfrs")
    wb = openpyxl.load_workbook(str(tpath), data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=2, max_col=30):
            for cell in row:
                cell.value = None
    written = []
    for i, (uuid, sheet, row, render_col) in enumerate(leaves):
        wb[sheet].cell(row=row, column=column_index_from_string(render_col or "B")).value = 1000.0 + i
        written.append((uuid, 1000.0 + i))
    out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(out.name)
    wb.close()
    data = Path(out.name).read_bytes()
    Path(out.name).unlink(missing_ok=True)
    return data, written


@pytest.fixture
def client(tmp_path, monkeypatch):
    db = tmp_path / "xbrl.db"
    monkeypatch.setenv("XBRL_OUTPUT_DIR", str(tmp_path))
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    monkeypatch.setenv("LLM_PROXY_URL", "")
    import importlib
    import server as srv
    importlib.reload(srv)
    srv.AUDIT_DB_PATH = db

    from db.schema import init_db
    init_db(db)
    template_id = _import_company_sofp(db)
    return TestClient(srv.app), db, template_id, srv


def test_benchmark_lifecycle(client):
    tc, db, template_id, srv = client
    data, written = _fixture_workbook_bytes(db, template_id)

    # Create via upload.
    resp = tc.post(
        "/api/benchmarks",
        data={"name": "FINCO 2021", "filing_standard": "mfrs",
              "filing_level": "company", "document": "FINCO.pdf"},
        files={"file": ("filled.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["ok"] is True
    assert body["ingested"] == len(written)
    bench_id = body["id"]
    assert "SOFP" in body["statements"]

    # List.
    lst = tc.get("/api/benchmarks").json()["benchmarks"]
    assert any(b["id"] == bench_id and b["gold_cell_count"] == len(written)
               for b in lst)

    # Detail.
    detail = tc.get(f"/api/benchmarks/{bench_id}").json()
    assert detail["filing_standard"] == "mfrs"
    assert any(t["template_id"] == template_id for t in detail["templates"])

    # Gold grid (ConceptsPage shape).
    grid = tc.get(f"/api/benchmarks/{bench_id}/concepts").json()
    assert grid["benchmark_id"] == bench_id
    filled = {c["concept_uuid"]: c for c in grid["concepts"] if c["value"] is not None}
    uuid0, val0 = written[0]
    assert uuid0 in filled
    assert abs(filled[uuid0]["value"] - val0) < 1e-9
    assert filled[uuid0]["editable"] is True

    # Spot-edit a gold value.
    patch = tc.patch(
        f"/api/benchmarks/{bench_id}/facts",
        json={"concept_uuid": uuid0, "period": "CY",
              "entity_scope": "Company", "value": 42.0},
    )
    assert patch.status_code == 200, patch.text
    assert patch.json()["value"] == 42.0
    grid2 = tc.get(f"/api/benchmarks/{bench_id}/concepts").json()
    edited = next(c for c in grid2["concepts"] if c["concept_uuid"] == uuid0)
    assert abs(edited["value"] - 42.0) < 1e-9

    # Delete = ARCHIVE (Step 9, PLAN-evals-hardening): historical scores and
    # the benchmark row survive; it just leaves the default picker list.
    dele = tc.delete(f"/api/benchmarks/{bench_id}")
    assert dele.status_code == 200
    body = dele.json()
    assert body["archived"] is True and "scores_kept" in body
    detail = tc.get(f"/api/benchmarks/{bench_id}")
    assert detail.status_code == 200
    assert detail.json()["is_archived"] is True
    default_list = tc.get("/api/benchmarks").json()["benchmarks"]
    assert all(b["id"] != bench_id for b in default_list)
    with_archived = tc.get(
        "/api/benchmarks", params={"include_archived": "true"}
    ).json()["benchmarks"]
    assert any(b["id"] == bench_id for b in with_archived)

    # Restore, then hard-delete (admin-only true-mistake path) → really gone.
    assert tc.post(f"/api/benchmarks/{bench_id}/unarchive").status_code == 200
    hard = tc.delete(f"/api/benchmarks/{bench_id}", params={"hard": "true"})
    assert hard.status_code == 200
    assert hard.json()["hard_deleted"] is True
    assert tc.get(f"/api/benchmarks/{bench_id}").status_code == 404


def test_create_rejects_unmatched_workbook(client):
    tc, db, template_id, srv = client
    # A workbook whose sheets match no template.
    wb = openpyxl.Workbook()
    wb.active.title = "Nope"
    wb.active["B5"] = 1
    out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(out.name)
    wb.close()
    data = Path(out.name).read_bytes()
    Path(out.name).unlink(missing_ok=True)

    resp = tc.post(
        "/api/benchmarks",
        data={"name": "bad", "filing_standard": "mfrs", "filing_level": "company"},
        files={"file": ("bad.xlsx", data, "application/octet-stream")},
    )
    assert resp.status_code == 422
    assert "No" in resp.json()["detail"]


def test_create_rejects_zero_gold_cells(client):
    """Sheets match the template but every value cell is blank → 422, no
    useless 0-cell benchmark is created (P2 / peer-review)."""
    tc, db, template_id, srv = client
    # n=0 → the live template's sheets are present but all value cells blanked.
    data, written = _fixture_workbook_bytes(db, template_id, n=0)
    assert written == []

    resp = tc.post(
        "/api/benchmarks",
        data={"name": "empty", "filing_standard": "mfrs", "filing_level": "company"},
        files={"file": ("empty.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 422
    assert "gold cells" in resp.json()["detail"]
    # Nothing persisted (the route rolled back).
    assert tc.get("/api/benchmarks").json()["benchmarks"] == []


def test_create_rejects_non_xlsx(client):
    tc, db, template_id, srv = client
    resp = tc.post(
        "/api/benchmarks",
        data={"name": "bad", "filing_standard": "mfrs", "filing_level": "company"},
        files={"file": ("notes.txt", b"hello", "text/plain")},
    )
    assert resp.status_code == 400


def test_run_eval_endpoint(client):
    tc, db, template_id, srv = client
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.execute(
        "INSERT INTO eval_benchmarks(name, filing_standard, filing_level) "
        "VALUES ('B', 'mfrs', 'company')"
    )
    bench_id = int(cur.lastrowid)
    cur = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status, benchmark_id) "
        "VALUES ('2026-06-04T00:00:00Z', 'x.pdf', 'completed', ?)",
        (bench_id,),
    )
    run_id = int(cur.lastrowid)
    conn.execute(
        "INSERT INTO eval_scores(run_id, benchmark_id, gold_cells, "
        "matched_cells, missing_cells, mismatch_cells, extra_cells, "
        "scale_mismatch, created_at) VALUES (?, ?, 10, 7, 2, 1, 3, 1, '2026-06-04Z')",
        (run_id, bench_id),
    )
    conn.commit()
    conn.close()

    resp = tc.get(f"/api/runs/{run_id}/eval")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["gold_cells"] == 10
    assert body["matched_cells"] == 7
    assert abs(body["score"] - 0.7) < 1e-9

    # A run with no score → 404.
    cur = sqlite3.connect(str(db)).execute(
        "INSERT INTO runs(created_at, pdf_filename, status) "
        "VALUES ('2026-06-04Z', 'y.pdf', 'completed')"
    )
    assert tc.get("/api/runs/99999/eval").status_code == 404


def test_create_benchmark_from_run_endpoint(client):
    """POST /api/benchmarks/from-run seeds gold straight from a run's facts —
    capturing all leaves with no workbook formula-cache loss (gotcha #23)."""
    tc, db, template_id, srv = client

    # Seed a finished run with a couple of LEAF facts + a COMPUTED total.
    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at, "
            "run_config_json) VALUES (?, ?, ?, ?, ?)",
            ("2026-06-05T00:00:00Z", "x.pdf", "completed_with_errors",
             "2026-06-05T00:00:00Z",
             json.dumps({"filing_standard": "mfrs", "filing_level": "company"})),
        )
        run_id = int(cur.lastrowid)
        leaves = [r[0] for r in conn.execute(
            "SELECT concept_uuid FROM concept_nodes WHERE template_id = ? "
            "AND kind = 'LEAF' ORDER BY render_sheet, render_row LIMIT 2",
            (template_id,),
        ).fetchall()]
        computed = conn.execute(
            "SELECT concept_uuid FROM concept_nodes WHERE template_id = ? "
            "AND kind = 'COMPUTED' LIMIT 1", (template_id,),
        ).fetchone()[0]
        for u, v in ((leaves[0], 111.0), (leaves[1], 222.0)):
            conn.execute(
                "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
                "entity_scope, value, value_status, source, updated_at) "
                "VALUES (?, ?, 'CY', 'Company', ?, 'observed', 'pdf', 'z')",
                (run_id, u, v),
            )
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, source, updated_at) "
            "VALUES (?, ?, 'CY', 'Company', 333.0, 'observed', 'pdf', 'z')",
            (run_id, computed),
        )
        conn.commit()
    finally:
        conn.close()

    resp = tc.post("/api/benchmarks/from-run",
                   json={"run_id": run_id, "name": "Seeded from run"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["ok"] is True
    assert body["ingested"] == 2  # two leaves; COMPUTED excluded
    assert body["source_run_id"] == run_id
    assert "SOFP" in body["statements"]

    # It shows up in the library with the right gold count.
    lst = tc.get("/api/benchmarks").json()["benchmarks"]
    assert any(b["id"] == body["id"] and b["gold_cell_count"] == 2 for b in lst)

    # A not-yet-finished run is rejected with 422.
    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at, "
            "run_config_json) VALUES (?, ?, 'failed', ?, ?)",
            ("2026-06-05T01:00:00Z", "y.pdf", "2026-06-05T01:00:00Z",
             json.dumps({"filing_standard": "mfrs", "filing_level": "company"})),
        )
        bad_run = int(cur.lastrowid)
        conn.commit()
    finally:
        conn.close()
    bad = tc.post("/api/benchmarks/from-run",
                  json={"run_id": bad_run, "name": "bad"})
    assert bad.status_code == 422, bad.text
