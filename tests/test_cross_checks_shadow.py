"""Shadow-diff parity proof for the fact-based cross-checks (item 32, 32a).

The migration gate: for each check, feed the SAME logical values to the xlsx
path (``check.run(workbook_paths, …)``) and the fact path
(``check.run_facts(ctx, …)``) and assert the two ``CrossCheckResult`` objects
are shadow-equal (``tests/shadow_diff.assert_cross_check_parity``). Only once a
check is proven equal here is its xlsx path retired (Phase 4).

Fixtures are hand-built (no import/cascade/export pipeline): a minimal workbook
and a matching DB, with the totals placed on the SAME sheet/row in both so the
target/comparand coordinates line up. This pins the two code paths agree on
identical data — exactly what the gate needs.
"""
from __future__ import annotations

import sqlite3

import openpyxl
import pytest

from cross_checks.framework import FactsContext
from cross_checks.soci_to_socie_tci import SOCIToSOCIETCICheck
from cross_checks.socf_to_sofp_cash import SOCFToSOFPCashCheck
from cross_checks.socie_to_sofp_equity import SOCIEToSOFPEquityCheck
from cross_checks.sofp_balance import SOFPBalanceCheck
from cross_checks.sopl_to_socie_profit import SOPLToSOCIEProfitCheck
from db.schema import init_db
from statement_types import StatementType
from tests.shadow_diff import assert_cross_check_parity


_SHEET = "SOFP-CuNonCu"
_ASSETS_ROW = 20
_EQLIAB_ROW = 40


def _make_workbook(path, rows):
    """rows: {row_index: [colA, colB, colC, colD, colE]} on the SOFP sheet."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(_SHEET)
    for r, cols in rows.items():
        for c_idx, val in enumerate(cols, start=1):
            if val is not None:
                ws.cell(row=r, column=c_idx, value=val)
    wb.save(path)
    wb.close()


def _seed_db(tmp_path, template_id):
    db = tmp_path / "shadow.db"
    init_db(db)
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute(
        "INSERT INTO concept_templates(template_id, source_path) VALUES (?, ?)",
        (template_id, "/tmp/t.xlsx"),
    )
    for uuid, row in (("c_assets", _ASSETS_ROW), ("c_eqliab", _EQLIAB_ROW)):
        label = "Total assets" if uuid == "c_assets" else "Total equity and liabilities"
        conn.execute(
            "INSERT INTO concept_nodes(concept_uuid, template_id, kind, "
            "canonical_label, render_sheet, render_row, render_col) "
            "VALUES (?, ?, 'COMPUTED', ?, ?, ?, 'B')",
            (uuid, template_id, label, _SHEET, row),
        )
    cur = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status) "
        "VALUES ('2026-06-14T00:00:00Z', 'x.pdf', 'completed')"
    )
    run_id = int(cur.lastrowid)
    conn.commit()
    return conn, run_id


def _fact(conn, run_id, uuid, value, *, period="CY", scope="Company"):
    conn.execute(
        "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
        "entity_scope, value, value_status, source) "
        "VALUES (?, ?, ?, ?, ?, 'observed', 'cascade')",
        (run_id, uuid, period, scope, value),
    )


@pytest.mark.parametrize(
    "assets,eqliab,expect_status",
    [
        (1000.0, 1000.0, "passed"),       # balanced
        (1000.0, 940.0, "failed"),        # imbalanced
        (1234.56, 1234.56, "passed"),     # non-round (float-repr parity)
    ],
)
def test_sofp_balance_company_shadow(tmp_path, assets, eqliab, expect_status):
    template_id = "mfrs-company-sofp-cunoncu-v1"

    wb_path = tmp_path / "sofp.xlsx"
    _make_workbook(wb_path, {
        _ASSETS_ROW: ["*Total assets", assets],
        _EQLIAB_ROW: ["*Total equity and liabilities", eqliab],
    })
    xlsx_result = SOFPBalanceCheck().run(
        {StatementType.SOFP: str(wb_path)}, tolerance=1.0, filing_level="company")

    conn, run_id = _seed_db(tmp_path, template_id)
    _fact(conn, run_id, "c_assets", assets)
    _fact(conn, run_id, "c_eqliab", eqliab)
    conn.commit()
    ctx = FactsContext(
        conn=conn, run_id=run_id,
        template_ids={StatementType.SOFP: template_id},
        filing_level="company", filing_standard="mfrs",
    )
    fact_result = SOFPBalanceCheck().run_facts(ctx, tolerance=1.0)
    conn.close()

    assert xlsx_result.status == expect_status
    assert_cross_check_parity(xlsx_result, fact_result)


def test_sofp_balance_group_shadow(tmp_path):
    template_id = "mfrs-group-sofp-cunoncu-v1"
    # Group: col B/C = Group CY/PY, col D/E = Company CY/PY.
    wb_path = tmp_path / "sofp.xlsx"
    _make_workbook(wb_path, {
        _ASSETS_ROW: ["*Total assets", 2000.0, 1900.0, 1200.0, 1100.0],
        _EQLIAB_ROW: ["*Total equity and liabilities", 2000.0, 1900.0, 1200.0, 1100.0],
    })
    xlsx_result = SOFPBalanceCheck().run(
        {StatementType.SOFP: str(wb_path)}, tolerance=1.0, filing_level="group")

    conn, run_id = _seed_db(tmp_path, template_id)
    # Group scope = primary (col B); Company scope = the dual (col D).
    _fact(conn, run_id, "c_assets", 2000.0, scope="Group")
    _fact(conn, run_id, "c_eqliab", 2000.0, scope="Group")
    _fact(conn, run_id, "c_assets", 1200.0, scope="Company")
    _fact(conn, run_id, "c_eqliab", 1200.0, scope="Company")
    conn.commit()
    ctx = FactsContext(
        conn=conn, run_id=run_id,
        template_ids={StatementType.SOFP: template_id},
        filing_level="group", filing_standard="mfrs",
    )
    fact_result = SOFPBalanceCheck().run_facts(ctx, tolerance=1.0)
    conn.close()

    assert xlsx_result.status == "passed"
    assert "Group CY:" in xlsx_result.message and "Company CY:" in xlsx_result.message
    assert_cross_check_parity(xlsx_result, fact_result)


# --------------------------------------------------------------------------
# Cross-statement + SOCIE-matrix checks. Generic builders keep each test a
# minimal (workbook, DB) pair with totals on matching sheet/row/col so the two
# paths' coordinates line up.
# --------------------------------------------------------------------------

_COL = {"B": 2, "C": 3, "W": 23, "X": 24}  # SOCIE matrix column letters → index


def _new_db(tmp_path):
    db = tmp_path / "x.db"
    init_db(db)
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status) "
        "VALUES ('2026-06-14T00:00:00Z', 'x.pdf', 'completed')"
    )
    return conn, int(cur.lastrowid)


def _tpl(conn, template_id):
    conn.execute(
        "INSERT INTO concept_templates(template_id, source_path) VALUES (?, ?)",
        (template_id, "/tmp/t.xlsx"),
    )


def _concept(conn, tid, uuid, kind, label, sheet, row, col, matrix_col=None):
    conn.execute(
        "INSERT INTO concept_nodes(concept_uuid, template_id, kind, "
        "canonical_label, render_sheet, render_row, render_col, matrix_col) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (uuid, tid, kind, label, sheet, row, col, matrix_col),
    )


def _sheet(wb, name, cells):
    """cells: list of (row, col_idx, value)."""
    ws = wb.create_sheet(name)
    for r, c, v in cells:
        ws.cell(row=r, column=c, value=v)


def test_socie_to_sofp_equity_company_shadow(tmp_path):
    sofp_tid, socie_tid = "mfrs-company-sofp-cunoncu-v1", "mfrs-company-socie-v1"
    wb_path = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "SOFP-CuNonCu", [(30, 1, "*Total equity"), (30, 2, 5000.0)])
    # SOCIE equity-at-end in Total column X (24).
    _sheet(wb, "SOCIE", [(25, 1, "*Equity at end of period"), (25, _COL["X"], 5000.0)])
    wb.save(wb_path); wb.close()

    conn, rid = _new_db(tmp_path)
    _tpl(conn, sofp_tid); _tpl(conn, socie_tid)
    _concept(conn, sofp_tid, "sofp_eq", "COMPUTED", "Total equity", "SOFP-CuNonCu", 30, "B")
    _concept(conn, socie_tid, "socie_eq_X", "MATRIX_CELL", "*Equity at end of period", "SOCIE", 25, "X", "X")
    _fact(conn, rid, "sofp_eq", 5000.0)
    _fact(conn, rid, "socie_eq_X", 5000.0)
    conn.commit()
    ctx = FactsContext(conn=conn, run_id=rid,
                       template_ids={StatementType.SOFP: sofp_tid,
                                     StatementType.SOCIE: socie_tid},
                       filing_level="company", filing_standard="mfrs")

    paths = {StatementType.SOFP: str(wb_path), StatementType.SOCIE: str(wb_path)}
    xlsx = SOCIEToSOFPEquityCheck().run(paths, 1.0, filing_level="company", filing_standard="mfrs")
    facts = SOCIEToSOFPEquityCheck().run_facts(ctx, 1.0)
    conn.close()
    assert xlsx.status == "passed"
    assert_cross_check_parity(xlsx, facts)


def test_soci_to_socie_tci_company_shadow(tmp_path):
    soci_tid, socie_tid = "mfrs-company-soci-netoftax-v1", "mfrs-company-socie-v1"
    wb_path = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "SOCI-NetOfTax", [(40, 1, "*Total comprehensive income"), (40, 2, 320.0)])
    _sheet(wb, "SOCIE", [(13, 1, "*Total comprehensive income"), (13, _COL["X"], 320.0)])
    wb.save(wb_path); wb.close()

    conn, rid = _new_db(tmp_path)
    _tpl(conn, soci_tid); _tpl(conn, socie_tid)
    _concept(conn, soci_tid, "soci_tci", "COMPUTED", "Total comprehensive income", "SOCI-NetOfTax", 40, "B")
    _concept(conn, socie_tid, "socie_tci_X", "MATRIX_CELL", "*Total comprehensive income", "SOCIE", 13, "X", "X")
    _fact(conn, rid, "soci_tci", 320.0)
    _fact(conn, rid, "socie_tci_X", 320.0)
    conn.commit()
    ctx = FactsContext(conn=conn, run_id=rid,
                       template_ids={StatementType.SOCI: soci_tid,
                                     StatementType.SOCIE: socie_tid},
                       filing_level="company", filing_standard="mfrs")
    paths = {StatementType.SOCI: str(wb_path), StatementType.SOCIE: str(wb_path)}
    xlsx = SOCIToSOCIETCICheck().run(paths, 1.0, filing_level="company", filing_standard="mfrs")
    facts = SOCIToSOCIETCICheck().run_facts(ctx, 1.0)
    conn.close()
    assert xlsx.status == "passed"
    assert_cross_check_parity(xlsx, facts)


def _profit_fixture(tmp_path, *, nci):
    """Build the SOPL+SOCIE fixture; nci=True puts profit in col X + an NCI
    value in col W, nci=False puts profit in col C (no NCI)."""
    sopl_tid, socie_tid = "mfrs-company-sopl-function-v1", "mfrs-company-socie-v1"
    profit_col = "X" if nci else "C"
    wb_path = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "SOPL-Function", [(50, 1, "*Profit (loss)"), (50, 2, 250.0)])
    socie_cells = [(11, 1, "*Profit (loss)"), (11, _COL[profit_col], 250.0)]
    if nci:
        socie_cells.append((11, _COL["W"], 7.0))  # NCI present → check reads col X
    _sheet(wb, "SOCIE", socie_cells)
    wb.save(wb_path); wb.close()

    conn, rid = _new_db(tmp_path)
    _tpl(conn, sopl_tid); _tpl(conn, socie_tid)
    _concept(conn, sopl_tid, "sopl_profit", "COMPUTED", "Profit (loss)", "SOPL-Function", 50, "B")
    _concept(conn, socie_tid, f"socie_profit_{profit_col}", "MATRIX_CELL", "*Profit (loss)", "SOCIE", 11, profit_col, profit_col)
    _fact(conn, rid, "sopl_profit", 250.0)
    _fact(conn, rid, f"socie_profit_{profit_col}", 250.0)
    if nci:
        _concept(conn, socie_tid, "socie_profit_W", "MATRIX_CELL", "*Profit (loss)", "SOCIE", 11, "W", "W")
        _fact(conn, rid, "socie_profit_W", 7.0)
    conn.commit()
    ctx = FactsContext(conn=conn, run_id=rid,
                       template_ids={StatementType.SOPL: sopl_tid,
                                     StatementType.SOCIE: socie_tid},
                       filing_level="company", filing_standard="mfrs")
    paths = {StatementType.SOPL: str(wb_path), StatementType.SOCIE: str(wb_path)}
    return conn, ctx, paths


def test_sopl_to_socie_profit_no_nci_reads_retained_col_shadow(tmp_path):
    conn, ctx, paths = _profit_fixture(tmp_path, nci=False)
    xlsx = SOPLToSOCIEProfitCheck().run(paths, 1.0, filing_level="company", filing_standard="mfrs")
    facts = SOPLToSOCIEProfitCheck().run_facts(ctx, 1.0)
    conn.close()
    assert xlsx.status == "passed"
    assert_cross_check_parity(xlsx, facts)


def test_sopl_to_socie_profit_with_nci_reads_total_col_shadow(tmp_path):
    conn, ctx, paths = _profit_fixture(tmp_path, nci=True)
    xlsx = SOPLToSOCIEProfitCheck().run(paths, 1.0, filing_level="company", filing_standard="mfrs")
    facts = SOPLToSOCIEProfitCheck().run_facts(ctx, 1.0)
    conn.close()
    assert xlsx.status == "passed"
    assert_cross_check_parity(xlsx, facts)


def test_socf_to_sofp_cash_lineage_message_shadow(tmp_path):
    """Imbalanced with SOFP cash 0 → the 'fill SOFP cash' lineage line must
    appear identically in both paths."""
    socf_tid, sofp_tid = "mfrs-company-socf-indirect-v1", "mfrs-company-sofp-cunoncu-v1"
    wb_path = tmp_path / "wb.xlsx"
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _sheet(wb, "SOCF-Indirect", [(60, 1, "*Cash and cash equivalents at end of period"), (60, 2, 900.0)])
    _sheet(wb, "SOFP-CuNonCu", [(15, 1, "Cash and cash equivalents"), (15, 2, 0.0)])
    wb.save(wb_path); wb.close()

    conn, rid = _new_db(tmp_path)
    _tpl(conn, socf_tid); _tpl(conn, sofp_tid)
    _concept(conn, socf_tid, "socf_cash", "LEAF", "Cash and cash equivalents at end of period", "SOCF-Indirect", 60, "B")
    _concept(conn, sofp_tid, "sofp_cash", "LEAF", "Cash and cash equivalents", "SOFP-CuNonCu", 15, "B")
    _fact(conn, rid, "socf_cash", 900.0)
    _fact(conn, rid, "sofp_cash", 0.0)
    conn.commit()
    ctx = FactsContext(conn=conn, run_id=rid,
                       template_ids={StatementType.SOCF: socf_tid,
                                     StatementType.SOFP: sofp_tid},
                       filing_level="company", filing_standard="mfrs")
    paths = {StatementType.SOCF: str(wb_path), StatementType.SOFP: str(wb_path)}
    xlsx = SOCFToSOFPCashCheck().run(paths, 1.0, filing_level="company")
    facts = SOCFToSOFPCashCheck().run_facts(ctx, 1.0)
    conn.close()
    assert xlsx.status == "failed"
    assert "Fill SOFP cash" in xlsx.message
    assert_cross_check_parity(xlsx, facts)
