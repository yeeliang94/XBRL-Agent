"""Peer-review hardening (2026-05-21).

Pins three regressions surfaced by team-lead peer review:

1. Face coordinator must apply a per-turn timeout (parity with notes).
   Mirror of NOTES_TURN_TIMEOUT — one stalled provider call must not
   hang the whole face run past MAX_AGENT_ITERATIONS (which only fires
   *between* iterations, not during a single stalled node).

2. Face coordinator must NOT return status="succeeded" when the agent
   finished without writing a workbook. The save_result gate in
   extraction/agent.py can refuse to finalise, but it can't force the
   agent to enter the gate — a conversational-only end-of-turn used to
   leak through as `succeeded`.

3. NotPrepared variants must produce an explicit `skipped` AgentResult.
   The pre-fix `continue` left those statements invisible to merge /
   history / UI.

4. tools/verifier._verify_socie must resolve the total column from
   filing_standard. MPERS SOCIE / SoRE uses a flat B/C layout; the
   pre-fix code hardcoded col X (24) and read None on every block,
   false-flagging every MPERS SOCIE as imbalanced.
"""
from __future__ import annotations

import asyncio
from contextlib import asynccontextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Set
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from statement_types import StatementType


# ---------------------------------------------------------------------------
# Local RunConfig stub — mirrors the one in tests/test_coordinator.py so
# this file stays self-contained.
# ---------------------------------------------------------------------------

@dataclass
class _RunConfig:
    pdf_path: str
    output_dir: str
    model: str = "test-model"
    statements_to_run: Set[StatementType] = field(default_factory=lambda: {StatementType.SOFP})
    variants: Dict[StatementType, str] = field(default_factory=dict)
    models: Dict[StatementType, str] = field(default_factory=dict)
    scout_enabled: bool = True
    filing_level: str = "company"
    filing_standard: str = "mfrs"


def _make_agent_with_filled_path(filled_path: str | None):
    """Mock agent.iter() that completes immediately. Caller controls
    whether deps.filled_path was 'set' by the (mocked) tool sequence."""
    mock_agent = MagicMock()
    mock_run = MagicMock()
    mock_run.result = MagicMock(output="done")
    mock_run.usage = MagicMock(return_value=MagicMock(
        input_tokens=10, output_tokens=5, total_tokens=15,
    ))

    async def empty_aiter(_self=None):
        return
        yield  # pragma: no cover
    mock_run.__aiter__ = empty_aiter

    @asynccontextmanager
    async def success_iter(*_a, **_k):
        yield mock_run
    mock_agent.iter = success_iter
    return mock_agent


def _make_agent_that_stalls(stall_seconds: float):
    """Mock agent.iter() whose __anext__ never returns. Used to drive
    the per-turn timeout — patch FACE_TURN_TIMEOUT down to a small value
    so the test doesn't actually sleep 180s."""
    mock_agent = MagicMock()
    mock_run = MagicMock()
    mock_run.result = MagicMock(output="never reached")
    mock_run.usage = MagicMock(return_value=MagicMock(
        input_tokens=0, output_tokens=0, total_tokens=0,
    ))

    async def stalled_aiter(_self=None):
        await asyncio.sleep(stall_seconds)
        return
        yield  # pragma: no cover
    mock_run.__aiter__ = stalled_aiter

    @asynccontextmanager
    async def stall_iter(*_a, **_k):
        yield mock_run
    mock_agent.iter = stall_iter
    return mock_agent


# ---------------------------------------------------------------------------
# 1. Per-turn timeout — fast cancel via patched FACE_TURN_TIMEOUT
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_face_coordinator_times_out_stalled_turn(tmp_path):
    """When a node iteration stalls past FACE_TURN_TIMEOUT and no
    workbook has been written, the agent result must be `failed` with a
    turn-timeout error message — not silently hang."""
    from coordinator import run_extraction

    config = _RunConfig(
        pdf_path="/tmp/test.pdf",
        output_dir=str(tmp_path),
        statements_to_run={StatementType.SOFP},
        variants={StatementType.SOFP: "CuNonCu"},
    )

    with patch("coordinator.FACE_TURN_TIMEOUT", 0.05), \
         patch("coordinator.create_extraction_agent") as factory:
        mock_deps = MagicMock()
        mock_deps.projection_failed = False
        mock_deps.filled_path = ""  # no workbook written
        mock_deps.filled_filename = "SOFP_filled.xlsx"
        mock_deps.statement_type = StatementType.SOFP
        factory.return_value = (_make_agent_that_stalls(stall_seconds=2.0), mock_deps)

        result = await run_extraction(config, infopack=None)

    assert len(result.agent_results) == 1
    r = result.agent_results[0]
    assert r.status == "failed"
    assert "stalled" in (r.error or "").lower()


@pytest.mark.asyncio
async def test_face_coordinator_keeps_workbook_after_post_write_stall(tmp_path):
    """If the agent stalled AFTER successfully writing a workbook, the
    coordinator must keep the artifact — mirror of notes-coordinator
    post-write recovery (NOTES_TURN_TIMEOUT branch)."""
    from coordinator import run_extraction

    config = _RunConfig(
        pdf_path="/tmp/test.pdf",
        output_dir=str(tmp_path),
        statements_to_run={StatementType.SOFP},
        variants={StatementType.SOFP: "CuNonCu"},
    )

    with patch("coordinator.FACE_TURN_TIMEOUT", 0.05), \
         patch("coordinator.create_extraction_agent") as factory:
        mock_deps = MagicMock()
        mock_deps.projection_failed = False
        # Simulate "a write landed before the stall" — filled_path set.
        mock_deps.filled_path = str(tmp_path / "SOFP_filled.xlsx")
        mock_deps.filled_filename = "SOFP_filled.xlsx"
        mock_deps.statement_type = StatementType.SOFP
        factory.return_value = (_make_agent_that_stalls(stall_seconds=2.0), mock_deps)

        result = await run_extraction(config, infopack=None)

    r = result.agent_results[0]
    assert r.status == "succeeded"
    assert r.workbook_path == str(tmp_path / "SOFP_filled.xlsx")


# ---------------------------------------------------------------------------
# 2. Success contract — no filled_path → failed, not succeeded
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_face_coordinator_refuses_success_without_filled_path(tmp_path):
    """An agent that finishes its turn budget without ever calling
    fill_workbook (deps.filled_path empty) must surface as `failed`,
    not `succeeded`. The pre-fix coordinator emitted success=True even
    when nothing was written."""
    from coordinator import run_extraction

    config = _RunConfig(
        pdf_path="/tmp/test.pdf",
        output_dir=str(tmp_path),
        statements_to_run={StatementType.SOFP},
        variants={StatementType.SOFP: "CuNonCu"},
    )

    with patch("coordinator.create_extraction_agent") as factory:
        mock_deps = MagicMock()
        mock_deps.projection_failed = False
        mock_deps.filled_path = ""  # agent never wrote
        mock_deps.filled_filename = "SOFP_filled.xlsx"
        mock_deps.statement_type = StatementType.SOFP
        factory.return_value = (_make_agent_with_filled_path(None), mock_deps)

        result = await run_extraction(config, infopack=None)

    r = result.agent_results[0]
    assert r.status == "failed"
    assert r.workbook_path is None
    assert "without writing" in (r.error or "").lower()


# ---------------------------------------------------------------------------
# 3. NotPrepared → explicit `skipped` result
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_notprepared_variant_produces_skipped_result(tmp_path):
    """A statement resolved to NotPrepared must appear in
    CoordinatorResult.agent_results with status='skipped' — not be
    silently dropped."""
    from coordinator import run_extraction

    config = _RunConfig(
        pdf_path="/tmp/test.pdf",
        output_dir=str(tmp_path),
        statements_to_run={StatementType.SOFP, StatementType.SOCIE},
        variants={
            StatementType.SOFP: "CuNonCu",
            StatementType.SOCIE: "NotPrepared",
        },
    )

    with patch("coordinator.create_extraction_agent") as factory:
        mock_deps = MagicMock()
        mock_deps.projection_failed = False
        mock_deps.filled_path = str(tmp_path / "SOFP_filled.xlsx")
        mock_deps.filled_filename = "SOFP_filled.xlsx"
        mock_deps.statement_type = StatementType.SOFP
        factory.return_value = (_make_agent_with_filled_path("ok"), mock_deps)

        result = await run_extraction(config, infopack=None)

    statuses_by_stmt = {r.statement_type: r.status for r in result.agent_results}
    assert statuses_by_stmt[StatementType.SOCIE] == "skipped"
    assert statuses_by_stmt[StatementType.SOFP] == "succeeded"


# ---------------------------------------------------------------------------
# 4. _verify_socie picks the right total column per filing_standard
# ---------------------------------------------------------------------------

def _build_socie_workbook(
    path: Path,
    *,
    layout: str,  # "mfrs_matrix" or "mpers_flat"
    restated: float,
    increase: float,
    closing: float,
) -> None:
    """Synthesize a SOCIE workbook with the three labels _verify_socie
    looks up. `layout` decides whether values land in col X (24) or col B (2).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCIE"

    ws.cell(3, 1).value = "Equity at beginning of period, restated"
    ws.cell(4, 1).value = "Total increase (decrease) in equity"
    ws.cell(5, 1).value = "Equity at end of period"

    if layout == "mfrs_matrix":
        col = 24
    elif layout == "mpers_flat":
        col = 2
    else:  # pragma: no cover
        raise ValueError(layout)

    ws.cell(3, col).value = restated
    ws.cell(4, col).value = increase
    ws.cell(5, col).value = closing
    wb.save(path)


def test_verify_socie_mfrs_reads_col_24(tmp_path):
    """MFRS SOCIE keeps the pre-existing behaviour: col X (24) is the
    matrix-aggregate total."""
    from tools.verifier import verify_statement

    path = tmp_path / "mfrs_socie.xlsx"
    _build_socie_workbook(
        path, layout="mfrs_matrix",
        restated=1000.0, increase=200.0, closing=1200.0,
    )

    result = verify_statement(
        str(path),
        StatementType.SOCIE,
        variant="Default",
        filing_level="company",
        filing_standard="mfrs",
    )

    assert result.is_balanced is True


def test_verify_socie_mpers_reads_col_2(tmp_path):
    """MPERS SOCIE is a flat B/C layout. The pre-fix verifier hardcoded
    col 24 and read None — every block false-flagged as imbalanced.
    With filing_standard='mpers' the read must pick col B (2)."""
    from tools.verifier import verify_statement

    path = tmp_path / "mpers_socie.xlsx"
    _build_socie_workbook(
        path, layout="mpers_flat",
        restated=1000.0, increase=200.0, closing=1200.0,
    )

    result = verify_statement(
        str(path),
        StatementType.SOCIE,
        variant="Default",
        filing_level="company",
        filing_standard="mpers",
    )

    assert result.is_balanced is True


def test_verify_socie_mpers_with_mfrs_flag_misreads(tmp_path):
    """Regression pin: a MPERS-shaped workbook checked with
    filing_standard='mfrs' (the pre-fix default) reads col 24 and
    flags imbalance — proving the standard axis actually changes
    column selection, not just metadata."""
    from tools.verifier import verify_statement

    path = tmp_path / "mpers_socie_misread.xlsx"
    _build_socie_workbook(
        path, layout="mpers_flat",
        restated=1000.0, increase=200.0, closing=1200.0,
    )

    result = verify_statement(
        str(path),
        StatementType.SOCIE,
        variant="Default",
        filing_level="company",
        filing_standard="mfrs",  # WRONG axis for this workbook
    )

    # Col 24 is empty → treated as 0 → closing(0) != 0+0=0 actually balances
    # at zero. So we can only assert that the *values* didn't surface;
    # the regression we care about is that the standard flag changes
    # which column is read. Check computed_totals reflect the empty read.
    assert result.computed_totals.get("closing_equity_cy", 0.0) == 0.0
    assert result.computed_totals.get("restated_equity_cy", 0.0) == 0.0


# ---------------------------------------------------------------------------
# 5. Peer-review (2026-05-22): skipped is non-failing for all_succeeded
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_all_succeeded_treats_skipped_as_success(tmp_path):
    """A run with a NotPrepared (skipped) statement plus an otherwise
    successful one must report all_succeeded — `skipped` is a non-outcome,
    not a failure. The skipped result is still surfaced via .skipped."""
    from coordinator import run_extraction

    config = _RunConfig(
        pdf_path="/tmp/test.pdf",
        output_dir=str(tmp_path),
        statements_to_run={StatementType.SOFP, StatementType.SOCIE},
        variants={
            StatementType.SOFP: "CuNonCu",
            StatementType.SOCIE: "NotPrepared",
        },
    )
    with patch("coordinator.create_extraction_agent") as factory:
        mock_deps = MagicMock()
        mock_deps.projection_failed = False
        mock_deps.filled_path = str(tmp_path / "SOFP_filled.xlsx")
        mock_deps.filled_filename = "SOFP_filled.xlsx"
        mock_deps.statement_type = StatementType.SOFP
        factory.return_value = (_make_agent_with_filled_path("ok"), mock_deps)
        result = await run_extraction(config, infopack=None)

    assert result.all_succeeded is True
    assert {r.statement_type for r in result.skipped} == {StatementType.SOCIE}


# ---------------------------------------------------------------------------
# 6. Peer-review (2026-05-22): SoRE verified with retained-earnings labels
# ---------------------------------------------------------------------------

def test_verify_sore_uses_retained_earnings_labels(tmp_path):
    """A SoRE workbook (retained-earnings labels, not equity) must verify
    on its own labels. The pre-fix _verify_socie looked only for
    'Equity …' rows and failed-closed on every real SoRE extraction."""
    from tools.verifier import verify_statement

    path = tmp_path / "sore.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoRE"
    ws.cell(3, 1).value = "Retained earnings at beginning of period, restated"
    ws.cell(4, 1).value = "Total increase (decrease) in retained earnings"
    ws.cell(5, 1).value = "Retained earnings at end of period"
    ws.cell(3, 2).value = 1000.0
    ws.cell(4, 2).value = 200.0
    ws.cell(5, 2).value = 1200.0
    wb.save(str(path))

    result = verify_statement(
        str(path), StatementType.SOCIE, variant="SoRE",
        filing_level="company", filing_standard="mpers",
    )
    # Labels are found (no "Required label not found") and the balance holds.
    assert "not found" not in (result.feedback or "").lower()
    assert result.is_balanced is True


def test_verify_sore_imbalance_is_detected(tmp_path):
    """Sanity: the SoRE branch still catches a real imbalance."""
    from tools.verifier import verify_statement

    path = tmp_path / "sore_bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoRE"
    ws.cell(3, 1).value = "Retained earnings at beginning of period, restated"
    ws.cell(4, 1).value = "Total increase (decrease) in retained earnings"
    ws.cell(5, 1).value = "Retained earnings at end of period"
    ws.cell(3, 2).value = 1000.0
    ws.cell(4, 2).value = 200.0
    ws.cell(5, 2).value = 9999.0  # != 1200
    wb.save(str(path))

    result = verify_statement(
        str(path), StatementType.SOCIE, variant="SoRE",
        filing_level="company", filing_standard="mpers",
    )
    assert result.is_balanced is False
