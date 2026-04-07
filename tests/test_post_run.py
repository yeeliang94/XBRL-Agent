"""Tests for post-run workbook merge + field persistence (Phase 7, Step 7.4)."""
from __future__ import annotations

import json
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest

from statement_types import StatementType
from db.schema import init_db
from db import repository as repo
from workbook_merger import merge


@pytest.fixture
def tmp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield d


def _make_workbook(sheets: dict[str, list[list]], path: str):
    """Create a minimal workbook from {sheet_name: [[row_data], ...]}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
    wb.close()


class TestMergedWorkbookAndFieldsPersisted:
    """After a full run, merged workbook exists and fields are in DB."""

    def test_merge_and_persist_fields(self, tmp_dir):
        """End-to-end: merge workbooks + persist extracted fields to DB."""
        db_path = Path(tmp_dir) / "test.db"
        init_db(db_path)

        conn = sqlite3.connect(str(db_path))
        conn.execute("PRAGMA foreign_keys = ON")
        conn.row_factory = sqlite3.Row

        # Create a run and agent rows
        run_id = repo.create_run(conn, "test.pdf")
        sofp_agent_id = repo.create_run_agent(conn, run_id, "SOFP", "CuNonCu", "test-model")
        sopl_agent_id = repo.create_run_agent(conn, run_id, "SOPL", "Function", "test-model")
        conn.commit()

        # Create per-statement workbooks
        sofp_path = str(Path(tmp_dir) / "SOFP_filled.xlsx")
        _make_workbook({"SOFP-CuNonCu": [["Total assets", 1000]], "SOFP-Sub-CuNonCu": [["PPE", 500]]}, sofp_path)

        sopl_path = str(Path(tmp_dir) / "SOPL_filled.xlsx")
        _make_workbook({"SOPL-Function": [["Revenue", 2000]]}, sopl_path)

        workbook_paths = {
            StatementType.SOFP: sofp_path,
            StatementType.SOPL: sopl_path,
        }

        # Merge workbooks
        merged_path = str(Path(tmp_dir) / "filled.xlsx")
        result = merge(workbook_paths, merged_path)
        assert result.success
        assert Path(merged_path).exists()

        # Verify merged workbook
        wb = openpyxl.load_workbook(merged_path)
        assert "SOFP-CuNonCu" in wb.sheetnames
        assert "SOPL-Function" in wb.sheetnames
        wb.close()

        # Persist extracted fields (simulating what server.py does with result.json)
        repo.save_extracted_field(conn, sofp_agent_id, "SOFP-CuNonCu", "Total assets", 2, 1000.0)
        repo.save_extracted_field(conn, sofp_agent_id, "SOFP-Sub-CuNonCu", "PPE", 2, 500.0)
        repo.save_extracted_field(conn, sopl_agent_id, "SOPL-Function", "Revenue", 2, 2000.0)
        conn.commit()

        # Verify fields in DB
        fields = repo.fetch_fields(conn, run_id)
        assert len(fields) == 3
        field_labels = {f.field_label for f in fields}
        assert "Total assets" in field_labels
        assert "Revenue" in field_labels

        # Persist cross-check results
        from cross_checks.framework import CrossCheckResult
        check_result = CrossCheckResult(
            name="sofp_balance", status="passed",
            expected=1000.0, actual=1000.0, diff=0.0,
            tolerance=1.0, message="Balance sheet balances",
        )
        repo.save_cross_check(
            conn, run_id, check_result.name, check_result.status,
            check_result.expected, check_result.actual, check_result.diff,
            check_result.tolerance, check_result.message,
        )
        conn.commit()

        # Verify cross-checks in DB
        checks = repo.fetch_cross_checks(conn, run_id)
        assert len(checks) == 1
        assert checks[0].check_name == "sofp_balance"
        assert checks[0].status == "passed"

        conn.close()
