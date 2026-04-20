"""PLAN §4 Phase E.2 — Excel-cell char-limit guard for the notes writer.

Excel caps a single cell at 32,767 characters. The notes writer must:
  1. Truncate any payload content that exceeds ``CELL_CHAR_LIMIT`` (30K).
  2. Append a footer pointing at the source PDF pages so the reviewer can
     go find the rest.
  3. Keep the total under ``CELL_CHAR_LIMIT`` after the truncation
     (there's no "slight overshoot tolerated" — Excel rejects overlong
     cells outright).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from notes.payload import NotesPayload
from notes.writer import CELL_CHAR_LIMIT, _truncate_with_footer, write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path


def test_short_content_is_passed_through_untouched():
    """Content under the cap must be preserved verbatim — no off-by-one
    where the writer drops the last char or appends a spurious footer."""
    content = "A" * (CELL_CHAR_LIMIT - 1)
    out = _truncate_with_footer(content, source_pages=[1, 2])
    assert out == content
    assert "truncated" not in out.lower()


def test_content_at_exactly_limit_is_not_truncated():
    content = "A" * CELL_CHAR_LIMIT
    out = _truncate_with_footer(content, source_pages=[5])
    assert out == content


def test_35k_content_gets_truncated_with_footer_pointing_at_pages():
    """PLAN §4 E.2 verification: a 35K-char payload must land under 30K
    with a footer that names the source pages."""
    content = "Z" * 35_000
    out = _truncate_with_footer(content, source_pages=[28, 29, 30])

    assert len(out) <= CELL_CHAR_LIMIT
    assert out.endswith("[truncated -- see PDF pages 28, 29, 30]")
    # Head content is preserved up to the truncation boundary.
    assert out.startswith("ZZZZ")
    # Cap minus footer length equals head length — sanity check.
    footer = "\n\n[truncated -- see PDF pages 28, 29, 30]"
    assert len(out) == CELL_CHAR_LIMIT
    assert out == "Z" * (CELL_CHAR_LIMIT - len(footer)) + footer


def test_missing_source_pages_produces_fallback_footer():
    """When the agent forgot to attach source pages to a payload, the
    footer still renders with ``n/a`` so operators notice the gap."""
    content = "X" * (CELL_CHAR_LIMIT + 10)
    out = _truncate_with_footer(content, source_pages=[])
    assert "[truncated -- see PDF pages n/a]" in out
    assert len(out) <= CELL_CHAR_LIMIT


def test_writer_truncates_overlong_payload_end_to_end(tmp_path: Path):
    """End-to-end check: feed an oversized payload through the writer and
    inspect the resulting Excel cell. Uses the smallest real template
    (Notes-CI) to keep the fixture simple."""
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_long.xlsx"

    overflow_body = "LOREM " * 6_000  # ~36K chars
    payload = NotesPayload(
        chosen_row_label="Financial reporting status",
        content=overflow_body,
        evidence="Page 12, Note 2",
        source_pages=[12],
    )

    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=[payload],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )
    assert result.success, result.errors
    assert result.rows_written == 1

    wb = openpyxl.load_workbook(out)
    ws = wb["Notes-CI"]
    # Find the row we wrote to.
    written_cell = None
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label and "financial reporting status" in str(label).lower():
            written_cell = ws.cell(row=row, column=2).value
            break
    assert written_cell is not None, "expected the payload to be written"

    assert len(written_cell) <= CELL_CHAR_LIMIT
    assert written_cell.endswith("[truncated -- see PDF pages 12]")
    wb.close()
