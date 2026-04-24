"""Tests for the 5 P0 cross-statement checks (Phase 5, Step 5.2).

Each test builds minimal workbook fixtures with known values and asserts the
check detects matches and mismatches correctly.
"""
from __future__ import annotations

import os
import tempfile

import openpyxl
import pytest

from statement_types import StatementType
from cross_checks.sofp_balance import SOFPBalanceCheck
from cross_checks.sopl_to_socie_profit import SOPLToSOCIEProfitCheck
from cross_checks.soci_to_socie_tci import SOCIToSOCIETCICheck
from cross_checks.socie_to_sofp_equity import SOCIEToSOFPEquityCheck
from cross_checks.socf_to_sofp_cash import SOCFToSOFPCashCheck


def _make_workbook(sheets: dict[str, list[list]], path: str):
    """Create a minimal workbook from a dict of {sheet_name: [[row_data], ...]}.

    Each row is [col_A, col_B, col_C, ...]. None entries are skipped.
    """
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
    wb.close()


@pytest.fixture
def tmp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield d


# ---------------------------------------------------------------------------
# Check 1: SOFP balance — Total assets = Total equity + liabilities
# ---------------------------------------------------------------------------

class TestSOFPBalance:
    def test_balanced(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                # row 1-9: filler
                *([["filler"]] * 9),
                # row 10: total assets
                ["*Total assets", 1000.0, 800.0],
                # row 11-19: filler
                *([["filler"]] * 9),
                # row 20: total equity and liabilities
                ["*Total equity and liabilities", 1000.0, 800.0],
            ],
        }, path)

        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0,
        )
        assert result.status == "passed"

    def test_imbalanced(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets", 1000.0, 800.0],
                ["*Total equity and liabilities", 900.0, 800.0],
            ],
        }, path)

        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0,
        )
        assert result.status == "failed"
        assert result.diff == pytest.approx(100.0, abs=0.01)

    def test_balanced_within_tolerance(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets", 1000.50, 800.0],
                ["*Total equity and liabilities", 1000.0, 800.0],
            ],
        }, path)

        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0,
        )
        assert result.status == "passed"

    def test_order_of_liquidity_variant(self, tmp_dir):
        """SOFP balance check works on OrderOfLiquidity sheet names too."""
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-OrdOfLiq": [
                ["*Total assets", 500.0, 400.0],
                ["*Total equity and liabilities", 500.0, 400.0],
            ],
        }, path)

        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0,
        )
        assert result.status == "passed"


# ---------------------------------------------------------------------------
# Check 2: SOPL profit = SOCIE profit row
# ---------------------------------------------------------------------------

class TestSOPLToSOCIEProfit:
    def test_matching(self, tmp_dir):
        sopl_path = os.path.join(tmp_dir, "sopl.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")

        _make_workbook({
            "SOPL-Function": [
                # Real template label is "*Profit (loss)" — check uses substring match
                ["*Profit (loss)", 250000.0, 200000.0],
            ],
        }, sopl_path)

        _make_workbook({
            "SOCIE": [
                # Row 1: header
                [None, "Issued capital", "Retained earnings"],
                # Row 2: profit line
                ["*Profit (loss) for the period", None, 250000.0],
            ],
        }, socie_path)

        result = SOPLToSOCIEProfitCheck().run(
            {StatementType.SOPL: sopl_path, StatementType.SOCIE: socie_path},
            tolerance=1.0,
        )
        assert result.status == "passed"

    def test_mismatch(self, tmp_dir):
        sopl_path = os.path.join(tmp_dir, "sopl.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")

        _make_workbook({
            "SOPL-Function": [
                ["*Profit (loss)", 250000.0, 200000.0],
            ],
        }, sopl_path)

        _make_workbook({
            "SOCIE": [
                [None, "Issued capital", "Retained earnings"],
                ["*Profit (loss) for the period", None, 300000.0],
            ],
        }, socie_path)

        result = SOPLToSOCIEProfitCheck().run(
            {StatementType.SOPL: sopl_path, StatementType.SOCIE: socie_path},
            tolerance=1.0,
        )
        assert result.status == "failed"


# ---------------------------------------------------------------------------
# Check 3: SOCI TCI = SOCIE TCI row
# ---------------------------------------------------------------------------

class TestSOCIToSOCIETCI:
    def test_matching(self, tmp_dir):
        soci_path = os.path.join(tmp_dir, "soci.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")

        _make_workbook({
            # Real sheet name is "SOCI-BeforeOfTax" (note the "Of")
            "SOCI-BeforeOfTax": [
                ["*Total comprehensive income for the period", 260000.0, 210000.0],
            ],
        }, soci_path)

        # TCI check always reads Total col X (24) — pad to reach it
        pad = [None] * 22
        _make_workbook({
            "SOCIE": [
                ["*Total comprehensive income for the period", *pad, 260000.0],
            ],
        }, socie_path)

        result = SOCIToSOCIETCICheck().run(
            {StatementType.SOCI: soci_path, StatementType.SOCIE: socie_path},
            tolerance=1.0,
        )
        assert result.status == "passed"

    def test_mismatch(self, tmp_dir):
        soci_path = os.path.join(tmp_dir, "soci.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")

        _make_workbook({
            "SOCI-BeforeOfTax": [
                ["*Total comprehensive income for the period", 260000.0],
            ],
        }, soci_path)

        pad = [None] * 22
        _make_workbook({
            "SOCIE": [
                ["*Total comprehensive income for the period", *pad, 999999.0],
            ],
        }, socie_path)

        result = SOCIToSOCIETCICheck().run(
            {StatementType.SOCI: soci_path, StatementType.SOCIE: socie_path},
            tolerance=1.0,
        )
        assert result.status == "failed"


# ---------------------------------------------------------------------------
# Check 4: SOCIE closing equity = SOFP total equity
# ---------------------------------------------------------------------------

class TestSOCIEToSOFPEquity:
    def test_matching(self, tmp_dir):
        """Equity check always reads Total col X (24)."""
        socie_path = os.path.join(tmp_dir, "socie.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        # Pad to col X (24) — 23 values where index 23 = col X
        pad = [None] * 22
        _make_workbook({
            "SOCIE": [
                ["*Equity at end of period", *pad, 750000.0],
            ],
        }, socie_path)

        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total equity", 750000.0, 600000.0],
            ],
        }, sofp_path)

        result = SOCIEToSOFPEquityCheck().run(
            {StatementType.SOCIE: socie_path, StatementType.SOFP: sofp_path},
            tolerance=1.0,
        )
        assert result.status == "passed"

    def test_mismatch(self, tmp_dir):
        socie_path = os.path.join(tmp_dir, "socie.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        pad = [None] * 22
        _make_workbook({
            "SOCIE": [
                ["*Equity at end of period", *pad, 750000.0],
            ],
        }, socie_path)

        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total equity", 800000.0, 600000.0],
            ],
        }, sofp_path)

        result = SOCIEToSOFPEquityCheck().run(
            {StatementType.SOCIE: socie_path, StatementType.SOFP: sofp_path},
            tolerance=1.0,
        )
        assert result.status == "failed"


# ---------------------------------------------------------------------------
# Check 5: SOCF closing cash = SOFP cash
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Step 5.4: Tolerance config
# ---------------------------------------------------------------------------

class TestToleranceApplied:
    def test_passes_within_tolerance(self, tmp_dir):
        """A check that's RM 0.50 off passes with tol=1, fails with tol=0.25."""
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets", 1000.50, 800.0],
                ["*Total equity and liabilities", 1000.0, 800.0],
            ],
        }, path)

        # With tolerance=1.0 → pass (diff=0.50 < 1.0)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0,
        )
        assert result.status == "passed"
        assert result.diff == pytest.approx(0.50, abs=0.01)

        # With tolerance=0.25 → fail (diff=0.50 > 0.25)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=0.25,
        )
        assert result.status == "failed"
        assert result.diff == pytest.approx(0.50, abs=0.01)

    def test_tolerance_plumbed_through_framework(self, tmp_dir):
        """run_all passes tolerance down to checks."""
        from cross_checks.framework import run_all

        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets", 1000.50, 800.0],
                ["*Total equity and liabilities", 1000.0, 800.0],
            ],
        }, path)

        checks = [SOFPBalanceCheck()]
        run_config = {"statements_to_run": {StatementType.SOFP}}

        # tol=1 → pass
        results = run_all(
            checks,
            workbook_paths={StatementType.SOFP: path},
            run_config=run_config,
            tolerance=1.0,
        )
        assert results[0].status == "passed"

        # tol=0.25 → fail
        results = run_all(
            checks,
            workbook_paths={StatementType.SOFP: path},
            run_config=run_config,
            tolerance=0.25,
        )
        assert results[0].status == "failed"


# ---------------------------------------------------------------------------
# Check 5: SOCF closing cash = SOFP cash
# ---------------------------------------------------------------------------

class TestSOCFToSOFPCash:
    def test_matching(self, tmp_dir):
        socf_path = os.path.join(tmp_dir, "socf.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        _make_workbook({
            "SOCF-Indirect": [
                ["*Cash and cash equivalents at end of period", 2551004.0, 1800000.0],
            ],
        }, socf_path)

        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Cash and cash equivalents", 2551004.0, 1800000.0],
            ],
        }, sofp_path)

        result = SOCFToSOFPCashCheck().run(
            {StatementType.SOCF: socf_path, StatementType.SOFP: sofp_path},
            tolerance=1.0,
        )
        assert result.status == "passed"

    def test_mismatch(self, tmp_dir):
        socf_path = os.path.join(tmp_dir, "socf.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        _make_workbook({
            "SOCF-Indirect": [
                ["*Cash and cash equivalents at end of period", 2551004.0],
            ],
        }, socf_path)

        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Cash and cash equivalents", 2000000.0],
            ],
        }, sofp_path)

        result = SOCFToSOFPCashCheck().run(
            {StatementType.SOCF: socf_path, StatementType.SOFP: sofp_path},
            tolerance=1.0,
        )
        assert result.status == "failed"

    def test_matching_order_of_liquidity_variant(self, tmp_dir):
        """OrdOfLiq SOFP labels the cash row 'Total Cash and bank balances',
        not 'Cash and cash equivalents'. The check must still pass."""
        socf_path = os.path.join(tmp_dir, "socf.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        _make_workbook({
            "SOCF-Indirect": [
                ["*Cash and cash equivalents at end of period", 5023.0, 2955.0],
            ],
        }, socf_path)

        _make_workbook({
            "SOFP-OrdOfLiq": [
                ["Total Cash and bank balances", 5023.0, 2955.0],
            ],
        }, sofp_path)

        result = SOCFToSOFPCashCheck().run(
            {StatementType.SOCF: socf_path, StatementType.SOFP: sofp_path},
            tolerance=1.0,
        )
        assert result.status == "passed"
        assert result.expected == 5023.0
        assert result.actual == 5023.0


# ---------------------------------------------------------------------------
# Smoke tests against real MBRS templates (Finding 5 fix)
# Verifies that find_sheet + find_value_by_label locate the correct cells
# in actual template files. Values are 0 (empty template) but the lookup
# path is validated — no None returns from label mismatches.
# ---------------------------------------------------------------------------

from pathlib import Path
from cross_checks.util import open_workbook, find_sheet, find_value_by_label

TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "XBRL-template-MFRS" / "Company"


def _label_row_exists(ws, label_substr: str) -> bool:
    """Check that a label can be found in column A (without requiring a value)."""
    target = label_substr.strip().lower()
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value is None:
            continue
        normalized = str(cell.value).strip().lstrip("*").strip().lower()
        if normalized == target or target in normalized:
            return True
    return False


class TestRealTemplateSmoke:
    """Verify cross-check label/sheet lookups work against actual MBRS templates.

    Templates are empty, so total-row values are either 0.0 (formula evaluates
    over empty cells) or None (data-entry cell, unfilled). These tests verify
    the lookup *finds the row*, not the value.
    """

    def test_sofp_cunoncuhas_total_rows(self):
        path = str(TEMPLATE_DIR / "01-SOFP-CuNonCu.xlsx")
        wb = open_workbook(path)
        ws = find_sheet(wb, "SOFP-CuNonCu")
        assert ws is not None, "Sheet SOFP-CuNonCu not found"
        # Total rows are formulas — evaluate to 0.0 on empty template
        val = find_value_by_label(ws, "total assets", col=2, wb=wb)
        assert val == 0.0, f"Expected 0.0 for formula-backed total, got {val}"
        val2 = find_value_by_label(ws, "total equity and liabilities", col=2, wb=wb)
        assert val2 == 0.0
        wb.close()

    def test_sofp_ordofliq_has_total_rows(self):
        path = str(TEMPLATE_DIR / "02-SOFP-OrderOfLiquidity.xlsx")
        wb = open_workbook(path)
        ws = find_sheet(wb, "SOFP-OrdOfLiq")
        assert ws is not None
        val = find_value_by_label(ws, "total assets", col=2, wb=wb)
        assert val == 0.0
        wb.close()

    def test_sofp_cash_label_resolves_in_both_variants(self):
        """The cash row has different labels per variant. The candidate-list
        form of find_value_by_label must resolve it in both templates."""
        candidates = ["cash and cash equivalents", "total cash and bank balances"]

        cunoncu = open_workbook(str(TEMPLATE_DIR / "01-SOFP-CuNonCu.xlsx"))
        ws = find_sheet(cunoncu, "SOFP-CuNonCu")
        val = find_value_by_label(ws, candidates, col=2, wb=cunoncu)
        assert val == 0.0, f"CuNonCu cash lookup returned {val}"
        cunoncu.close()

        ordofliq = open_workbook(str(TEMPLATE_DIR / "02-SOFP-OrderOfLiquidity.xlsx"))
        ws = find_sheet(ordofliq, "SOFP-OrdOfLiq")
        val = find_value_by_label(ws, candidates, col=2, wb=ordofliq)
        assert val == 0.0, f"OrdOfLiq cash lookup returned {val}"
        ordofliq.close()

    def test_soci_before_of_tax_sheet_found(self):
        """Real sheet name is SOCI-BeforeOfTax (not BeforeTax)."""
        path = str(TEMPLATE_DIR / "05-SOCI-BeforeTax.xlsx")
        wb = open_workbook(path)
        ws = find_sheet(wb, "SOCI-BeforeOfTax", "SOCI-BeforeTax")
        assert ws is not None, f"SOCI sheet not found. Sheets: {wb.sheetnames}"
        assert _label_row_exists(ws, "total comprehensive income")
        wb.close()

    def test_socie_closing_equity_label(self):
        """Real label is '*Equity at end of period' (not 'Balance at end of period')."""
        path = str(TEMPLATE_DIR / "09-SOCIE.xlsx")
        wb = open_workbook(path)
        ws = find_sheet(wb, "SOCIE")
        assert ws is not None
        assert _label_row_exists(ws, "equity at end of period")
        wb.close()

    def test_sopl_function_profit_label(self):
        """Real label is '*Profit (loss)' (not 'Profit (loss) for the period')."""
        path = str(TEMPLATE_DIR / "03-SOPL-Function.xlsx")
        wb = open_workbook(path)
        ws = find_sheet(wb, "SOPL-Function")
        assert ws is not None
        # Profit (loss) is a formula — evaluates to 0 on empty template
        val = find_value_by_label(ws, "profit (loss)", col=2, wb=wb)
        assert val == 0.0
        wb.close()

    def test_socf_indirect_closing_cash_label(self):
        """Closing cash row exists (data-entry, None on empty template)."""
        path = str(TEMPLATE_DIR / "07-SOCF-Indirect.xlsx")
        wb = open_workbook(path)
        ws = find_sheet(wb, "SOCF-Indirect")
        assert ws is not None
        # Row 132 is a data-entry cell (None on empty template) — just verify label exists
        assert _label_row_exists(ws, "cash and cash equivalents at end of period")
        wb.close()


# ---------------------------------------------------------------------------
# Bug 2 — cross-check messages must not call company filings "Group"
#
# Prior to this fix every check hardcoded "Group:" / "Group CY:" as the
# primary-column prefix. On a company filing there is only ONE column of
# figures, so the label is both redundant and wrong — users saw
# "Group CY: assets (…)" on reports they had explicitly marked Company.
# These tests pin the filing_level-aware prefix in place.
# ---------------------------------------------------------------------------


class TestMessagePrefixHonoursFilingLevel:
    def test_sofp_balance_company_does_not_mention_group(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets", 1000.0, 800.0],
                ["*Total equity and liabilities", 1000.0, 800.0],
            ],
        }, path)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0, filing_level="company",
        )
        assert "Group" not in result.message, (
            f"Company filing should not leak 'Group' label: {result.message!r}"
        )
        # S-3/S-4: company-filing prefix should be explicit and symmetric
        # with the group form ("Group CY:") — bare "CY:" leaves users to
        # guess. "Company CY:" pairs visually with "Group CY:" on group.
        assert "Company CY:" in result.message, (
            f"Company filing prefix should be 'Company CY:' "
            f"(symmetric with 'Group CY:' on group path). Got: {result.message!r}"
        )

    def test_sofp_balance_group_still_labels_both(self, tmp_dir):
        """Regression guard — group filings still dual-report Group + Company."""
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                # Company (col D=4) must exist on a group filing for the check
                # to have something to compare; pad col D too.
                ["*Total assets", 1000.0, 800.0, 600.0, 500.0],
                ["*Total equity and liabilities", 1000.0, 800.0, 600.0, 500.0],
            ],
        }, path)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0, filing_level="group",
        )
        assert "Group CY:" in result.message
        assert "Company CY:" in result.message

    def test_sopl_to_socie_profit_company_does_not_mention_group(self, tmp_dir):
        sopl_path = os.path.join(tmp_dir, "sopl.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")
        _make_workbook({"SOPL-Function": [["*Profit (loss)", 250000.0, 200000.0]]}, sopl_path)
        _make_workbook({
            "SOCIE": [
                [None, "Issued capital", "Retained earnings"],
                ["*Profit (loss) for the period", None, 250000.0],
            ],
        }, socie_path)
        result = SOPLToSOCIEProfitCheck().run(
            {StatementType.SOPL: sopl_path, StatementType.SOCIE: socie_path},
            tolerance=1.0, filing_level="company",
        )
        assert "Group" not in result.message, result.message

    def test_soci_to_socie_tci_company_does_not_mention_group(self, tmp_dir):
        soci_path = os.path.join(tmp_dir, "soci.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")
        _make_workbook({
            "SOCI-BeforeOfTax": [
                ["*Total comprehensive income for the period", 260000.0, 210000.0],
            ],
        }, soci_path)
        pad = [None] * 22
        _make_workbook({
            "SOCIE": [
                ["*Total comprehensive income for the period", *pad, 260000.0],
            ],
        }, socie_path)
        result = SOCIToSOCIETCICheck().run(
            {StatementType.SOCI: soci_path, StatementType.SOCIE: socie_path},
            tolerance=1.0, filing_level="company",
        )
        assert "Group" not in result.message, result.message

    def test_socie_to_sofp_equity_company_does_not_mention_group(self, tmp_dir):
        socie_path = os.path.join(tmp_dir, "socie.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")
        pad = [None] * 22
        _make_workbook({
            "SOCIE": [["*Equity at end of period", *pad, 2_000_000.0]],
        }, socie_path)
        _make_workbook({
            "SOFP-CuNonCu": [["*Total equity", 2_000_000.0, 1_800_000.0]],
        }, sofp_path)
        result = SOCIEToSOFPEquityCheck().run(
            {StatementType.SOCIE: socie_path, StatementType.SOFP: sofp_path},
            tolerance=1.0, filing_level="company",
        )
        assert "Group" not in result.message, result.message

    def test_filing_level_prefix_helper_is_single_source_of_truth(self):
        """S-4: extract `_filing_level_prefix` helper into cross_checks.util
        so adding a 6th check doesn't require copying the same 2-line
        branch. The helper's existence and contract are part of the
        public API for cross-check authors."""
        from cross_checks.util import filing_level_prefix
        # Four-col form (SOFP balance uses CY/PY rows → "Group CY" / "Company CY")
        assert filing_level_prefix("group", with_period=True) == "Group CY"
        assert filing_level_prefix("company", with_period=True) == "Company CY"
        # Three-col form (SOCI → SOCIE / SOPL → SOCIE / SOCF → SOFP use
        # a single-period aggregate so the period marker is redundant)
        assert filing_level_prefix("group", with_period=False) == "Group"
        assert filing_level_prefix("company", with_period=False) == "Company"

    def test_socf_to_sofp_cash_company_does_not_mention_group(self, tmp_dir):
        socf_path = os.path.join(tmp_dir, "socf.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOCF-Indirect": [
                ["*Cash and cash equivalents at end of period", 5023.0, 2955.0],
            ],
        }, socf_path)
        _make_workbook({
            "SOFP-OrdOfLiq": [
                ["Total Cash and bank balances", 5023.0, 2955.0],
            ],
        }, sofp_path)
        result = SOCFToSOFPCashCheck().run(
            {StatementType.SOCF: socf_path, StatementType.SOFP: sofp_path},
            tolerance=1.0, filing_level="company",
        )
        assert "Group" not in result.message, result.message
