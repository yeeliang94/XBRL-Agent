"""Phase 6.1 — tests for `cross_checks.notes_consistency`.

Builds synthetic Sheet-11 + Sheet-12 workbooks in tmp_path and asserts:
- Matching citations produce no warnings.
- Disjoint citations for a paired topic produce one warning.
- Unparseable citations are skipped rather than warned on.
- Single-sheet workbooks (only one of the two notes sheets populated)
  produce no warnings — vacuous consistency.
- Unknown label pairs are ignored (we never fabricate pairings).
- Page ranges ("Pages 20-22") are expanded before comparison.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from cross_checks.notes_consistency import (
    ConsistencyWarning,
    _extract_pages,
    check_notes_consistency,
)


def _build_workbook(
    path: Path,
    sheet11_rows: list[tuple[str, str]] | None = None,
    sheet12_rows: list[tuple[str, str]] | None = None,
) -> None:
    """Write a minimal two-sheet workbook with (col-A label, col-D evidence).
    No content column — col B stays empty; the consistency check only
    reads label + evidence."""
    wb = openpyxl.Workbook()
    # Remove the auto-created default sheet.
    default = wb.active
    wb.remove(default)

    s11 = wb.create_sheet("Notes-SummaryofAccPol")
    s11.cell(row=1, column=1, value="")
    s11.cell(row=1, column=4, value="Source")
    for i, (label, evidence) in enumerate(sheet11_rows or [], start=2):
        s11.cell(row=i, column=1, value=label)
        s11.cell(row=i, column=4, value=evidence)

    s12 = wb.create_sheet("Notes-Listofnotes")
    s12.cell(row=1, column=1, value="")
    s12.cell(row=1, column=4, value="Source")
    for i, (label, evidence) in enumerate(sheet12_rows or [], start=2):
        s12.cell(row=i, column=1, value=label)
        s12.cell(row=i, column=4, value=evidence)

    wb.save(str(path))


# ---------------------------------------------------------------------------
# _extract_pages
# ---------------------------------------------------------------------------


def test_extract_pages_single_page():
    assert _extract_pages("Page 27, Note 2.5(g)") == {27}


def test_extract_pages_range_expanded():
    assert _extract_pages("Pages 20-22, Note 2.5(b)") == {20, 21, 22}


def test_extract_pages_multiple_semicolon_separated():
    assert _extract_pages("Page 27, Note 2.5(g); Page 33") == {27, 33}


def test_extract_pages_reversed_range_normalised():
    """'Pages 30-20' (reversed) must still give a sane set, not {}."""
    assert _extract_pages("Pages 30-20") == {20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30}


def test_extract_pages_no_page_token_returns_empty():
    """A citation that mentions a number but no 'Page' keyword returns
    empty — we don't want Note numbers leaking into the page set."""
    assert _extract_pages("Note 2.5(g); at paragraph 4") == set()


def test_extract_pages_empty_string_returns_empty():
    assert _extract_pages("") == set()


# ---------------------------------------------------------------------------
# Peer-review I-2: abbreviation forms — "pp.", "pp", "p.", "p".
#
# The sub-coordinator's own started-status labels print "pp X-Y"; models
# pick up that convention and emit evidence like "pp. 27-28" in response.
# Without these, the cross-sheet warning silently vanishes for a common
# citation shape.
# ---------------------------------------------------------------------------


def test_extract_pages_pp_dot_with_range():
    """The exact form the peer reviewer flagged: "pp. 27-28, Note 2.5(g)"."""
    assert _extract_pages("pp. 27-28, Note 2.5(g)") == {27, 28}


def test_extract_pages_pp_no_dot_with_range():
    """Matches the sub-coordinator's own started-status text "pp 20-22"."""
    assert _extract_pages("pp 20-22") == {20, 21, 22}


def test_extract_pages_p_dot_single():
    """Singular "p. 33" — seen when a note cites a single page."""
    assert _extract_pages("p. 33") == {33}


def test_extract_pages_p_single_no_dot():
    """Singular "p 5" — bare abbreviation without the trailing dot."""
    assert _extract_pages("p 5, Note 3") == {5}


def test_extract_pages_pp_case_insensitive():
    """Models sometimes capitalise ("PP. 20-22"); the regex is
    case-insensitive so both forms land in the same bucket."""
    assert _extract_pages("PP. 20-22") == {20, 21, 22}


def test_extract_pages_does_not_match_bare_p_inside_word():
    """Guard against false positives — "happy 10" must not match on the
    word-internal 'p' followed by a number. ``\\b`` on the left side of
    the token group + the mandatory space anchor the match."""
    assert _extract_pages("happy 10") == set()


def test_extract_pages_mixed_forms_combine():
    """Multi-citation evidence with mixed abbreviations should union
    correctly — common when a policy row cross-references both a policy
    note and a quantified disclosure."""
    assert _extract_pages("Page 27 and pp. 33-34") == {27, 33, 34}


# ---------------------------------------------------------------------------
# check_notes_consistency
# ---------------------------------------------------------------------------


def test_matching_citations_produce_no_warnings(tmp_path: Path):
    """Happy path: paired topic with overlapping page sets on both
    sheets — no warning."""
    wb_path = tmp_path / "filled.xlsx"
    _build_workbook(
        wb_path,
        sheet11_rows=[
            ("Description of accounting policy for employee benefits",
             "Page 27, Note 2.5(g)"),
        ],
        sheet12_rows=[
            ("Disclosure of employee benefits expense",
             "Page 27, Note 2.5(g)"),
        ],
    )
    assert check_notes_consistency(str(wb_path)) == []


def test_disjoint_citations_produce_one_warning(tmp_path: Path):
    """The FINCO failure mode: same policy, different cited pages."""
    wb_path = tmp_path / "filled.xlsx"
    _build_workbook(
        wb_path,
        sheet11_rows=[
            ("Description of accounting policy for employee benefits",
             "Page 27, Note 2.5(g)"),
        ],
        sheet12_rows=[
            ("Disclosure of employee benefits expense",
             "Page 25, Note 2.5(g)"),  # printed folio, not PDF page
        ],
    )
    warnings = check_notes_consistency(str(wb_path))
    assert len(warnings) == 1
    w = warnings[0]
    assert isinstance(w, ConsistencyWarning)
    assert w.status == "warning"
    assert 27 in w.sheet_11_pages
    assert 25 in w.sheet_12_pages
    assert "printed folio" in w.message


def test_overlapping_but_unequal_page_sets_do_not_warn(tmp_path: Path):
    """Shared ONE page is enough. Sheet 11 cites pp 20-22 (policy),
    Sheet 12 cites pp 20, 31 (policy + schedule) — they share p 20 so
    this is consistent enough."""
    wb_path = tmp_path / "filled.xlsx"
    _build_workbook(
        wb_path,
        sheet11_rows=[
            ("Description of accounting policy for financial instruments",
             "Pages 20-22, Note 2.5(b)"),
        ],
        sheet12_rows=[
            ("Disclosure of financial instruments",
             "Page 20, Note 2.5(b); Page 31, Note 5"),
        ],
    )
    assert check_notes_consistency(str(wb_path)) == []


def test_unparseable_evidence_is_skipped(tmp_path: Path):
    """Evidence string with no 'Page' token — the check can't compare,
    and silently skipping beats flagging every prose-only citation."""
    wb_path = tmp_path / "filled.xlsx"
    _build_workbook(
        wb_path,
        sheet11_rows=[
            ("Description of accounting policy for employee benefits",
             "Note 2.5(g) — section on short-term benefits"),
        ],
        sheet12_rows=[
            ("Disclosure of employee benefits expense",
             "Page 27, Note 2.5(g)"),
        ],
    )
    assert check_notes_consistency(str(wb_path)) == []


def test_single_sheet_produces_no_warnings(tmp_path: Path):
    """If only Sheet 11 is populated, there is no Sheet-12 evidence to
    disagree with — vacuous consistency."""
    wb_path = tmp_path / "filled.xlsx"
    _build_workbook(
        wb_path,
        sheet11_rows=[
            ("Description of accounting policy for employee benefits",
             "Page 27, Note 2.5(g)"),
        ],
        sheet12_rows=[],  # empty
    )
    assert check_notes_consistency(str(wb_path)) == []


def test_unknown_label_pair_ignored(tmp_path: Path):
    """Labels not in the hand-curated pair list are ignored — we never
    guess a pairing the taxonomy doesn't clearly establish."""
    wb_path = tmp_path / "filled.xlsx"
    _build_workbook(
        wb_path,
        sheet11_rows=[("Some ad-hoc policy row", "Page 10")],
        sheet12_rows=[("Disclosure of something unrelated", "Page 99")],
    )
    assert check_notes_consistency(str(wb_path)) == []


def test_missing_workbook_returns_empty(tmp_path: Path):
    """Advisory check must never raise — missing file → []."""
    assert check_notes_consistency(str(tmp_path / "nope.xlsx")) == []


def test_reads_group_filing_evidence_column_f(tmp_path: Path):
    """Group filings place evidence in col F (6), not D (4). The check
    must find it there so Group runs aren't silently skipped."""
    wb_path = tmp_path / "filled.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, label, col_f_evidence in [
        ("Notes-SummaryofAccPol",
         "Description of accounting policy for employee benefits",
         "Page 27, Note 2.5(g)"),
        ("Notes-Listofnotes",
         "Disclosure of employee benefits expense",
         "Page 25, Note 2.5(g)"),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=2, column=1, value=label)
        ws.cell(row=2, column=6, value=col_f_evidence)
    wb.save(str(wb_path))

    warnings = check_notes_consistency(str(wb_path))
    assert len(warnings) == 1
    assert "printed folio" in warnings[0].message
