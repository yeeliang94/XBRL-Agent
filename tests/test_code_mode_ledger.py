"""CodeMode host-side call ledger pins (docs/PLAN-codemode-spike.md, Phase 3).

The ledger is the agreed replacement for the per-tool telemetry that a
`run_code` turn collapses: every sandbox tool call lands as one entry —
recorded in a `finally`, so FAILED calls land too — keyed to its script via
the harness's nested call ids, and flushed into the conversation-trace JSON
as `code_mode_calls`. The bar (plan Step 5 Verify): reconstruct the exact
call sequence of a deliberately failed script from the ledger alone.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest
from pydantic_ai.messages import (
    ModelResponse,
    RetryPromptPart,
    TextPart,
    ToolCallPart,
)
from pydantic_ai.models.function import AgentInfo, FunctionModel

from statement_types import StatementType
from tools.verifier import VerificationResult

from extraction.code_mode import (
    group_code_mode_ledger,
    ledger_echo,
)


def _make_template(tmp_path: Path) -> str:
    path = tmp_path / "sofp_template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP"
    ws["A5"] = "Cash and bank balances"
    ws["A6"] = "Trade receivables"
    wb.save(path)
    wb.close()
    return str(path)


def _build(tmp_path: Path, model):
    from extraction.agent import create_extraction_agent

    return create_extraction_agent(
        statement_type=StatementType.SOFP,
        variant="CuNonCu",
        pdf_path=str(tmp_path / "missing.pdf"),
        template_path=_make_template(tmp_path),
        model=model,
        output_dir=str(tmp_path),
    )


def _flag_on(monkeypatch):
    pytest.importorskip("pydantic_ai_harness")
    monkeypatch.setenv("XBRL_CODE_MODE", "1")
    monkeypatch.setenv("XBRL_CODE_MODE_STATEMENTS", "SOFP")


def _scripted_model(scripts):
    state = {"i": 0, "retries": []}

    def fake(messages, info: AgentInfo) -> ModelResponse:
        for msg in messages:
            for part in getattr(msg, "parts", []):
                if isinstance(part, RetryPromptPart):
                    state["retries"].append(str(part.content))
        if state["i"] < len(scripts):
            code = scripts[state["i"]]
            state["i"] += 1
            return ModelResponse(parts=[ToolCallPart("run_code", {"code": code})])
        return ModelResponse(parts=[TextPart("done")])

    return FunctionModel(fake), state


_WRITE = ('write_facts(facts=[{"sheet": "SOFP", "row": 5, "col": 2, '
          '"value": %d, "evidence": "Page 1, \'cash\'"}])')

SCRIPT_CHAIN = f"""
r = {_WRITE % 111}
v = verify_totals()
s = save_result(fields_json="")
print(r, v, s)
"""

SCRIPT_VERIFY_BLOWS_UP = f"""
r = {_WRITE % 111}
print("write:", r)
v = verify_totals()
print("verify:", v)
"""


def _balanced_verify(*args, **kwargs):
    return VerificationResult(is_balanced=True, matches_pdf=None)


def test_flag_off_ledger_is_none(tmp_path, monkeypatch):
    monkeypatch.delenv("XBRL_CODE_MODE", raising=False)

    def fake(messages, info: AgentInfo) -> ModelResponse:
        return ModelResponse(parts=[TextPart("done")])

    agent, deps = _build(tmp_path, FunctionModel(fake))
    agent.run_sync("go", deps=deps)
    assert deps.code_mode_ledger is None


def test_ledger_records_every_call_in_order(tmp_path, monkeypatch):
    _flag_on(monkeypatch)
    monkeypatch.setattr(
        "extraction.agent._verify_statement_impl", _balanced_verify
    )
    model, _ = _scripted_model([SCRIPT_CHAIN])
    agent, deps = _build(tmp_path, model)
    agent.run_sync("go", deps=deps)

    ledger = deps.code_mode_ledger
    assert [e["tool"] for e in ledger] == [
        "write_facts", "verify_totals", "save_result"
    ]
    for e in ledger:
        assert e["outcome"] == "ok"
        assert e["error"] is None
        assert isinstance(e["duration_ms"], int)
        assert e["result_summary"]
    # write_facts args are visible for debugging (truncated repr).
    assert "111" in ledger[0]["args_summary"]

    grouped = group_code_mode_ledger(ledger)
    assert [g["seq"] for g in grouped] == [1, 2, 3]
    assert {g["script_turn_index"] for g in grouped} == {1}
    assert {g["run_code_call_id"] for g in grouped} != {None}


def test_failed_tool_call_lands_in_ledger(tmp_path, monkeypatch):
    _flag_on(monkeypatch)

    def _explode(*args, **kwargs):
        raise RuntimeError("verifier exploded")

    monkeypatch.setattr("extraction.agent._verify_statement_impl", _explode)
    model, state = _scripted_model([SCRIPT_VERIFY_BLOWS_UP])
    agent, deps = _build(tmp_path, model)
    agent.run_sync("go", deps=deps)

    ledger = deps.code_mode_ledger
    by_tool = {e["tool"]: e for e in ledger}
    assert by_tool["write_facts"]["outcome"] == "ok"
    failed = by_tool["verify_totals"]
    assert failed["outcome"] == "error"
    assert "RuntimeError" in failed["error"]
    assert "verifier exploded" in failed["error"]
    assert failed["result_summary"] is None
    # The exact call sequence of the failed script is reconstructable from
    # the ledger alone — the Phase 3 Verify bar.
    assert [e["tool"] for e in ledger][:2] == ["write_facts", "verify_totals"]


def test_retry_scripts_group_separately(tmp_path, monkeypatch):
    _flag_on(monkeypatch)
    monkeypatch.setattr(
        "extraction.agent._verify_statement_impl", _balanced_verify
    )
    die = f'r = {_WRITE % 111}\nraise ValueError("boom")'
    model, _ = _scripted_model([die, SCRIPT_CHAIN])
    agent, deps = _build(tmp_path, model)
    agent.run_sync("go", deps=deps)

    grouped = group_code_mode_ledger(deps.code_mode_ledger)
    indices = {g["script_turn_index"] for g in grouped}
    assert indices == {1, 2}, (
        "the dead script and the retry must group under distinct scripts"
    )
    first = [g for g in grouped if g["script_turn_index"] == 1]
    assert [g["tool"] for g in first] == ["write_facts"]


def test_trace_file_carries_code_mode_calls(tmp_path):
    from agent_tracing import save_agent_trace

    class _FakeResult:
        def all_messages(self):
            return []

    calls = group_code_mode_ledger([
        {"tool": "write_facts", "tool_call_id": "pyd_ai_abc__1",
         "args_summary": "x", "duration_ms": 3, "outcome": "ok",
         "error": None, "result_summary": "wrote"},
    ])
    save_agent_trace(_FakeResult(), str(tmp_path), "SOFP",
                     turns=[], code_mode_calls=calls)
    payload = json.loads(
        (tmp_path / "SOFP_conversation_trace.json").read_text()
    )
    assert payload["code_mode_calls"][0]["tool"] == "write_facts"
    assert payload["code_mode_calls"][0]["run_code_call_id"] == "pyd_ai_abc"
    assert payload["code_mode_calls"][0]["script_turn_index"] == 1


def test_trace_file_omits_key_when_no_ledger(tmp_path):
    from agent_tracing import save_agent_trace

    class _FakeResult:
        def all_messages(self):
            return []

    save_agent_trace(_FakeResult(), str(tmp_path), "SOFP", turns=[])
    payload = json.loads(
        (tmp_path / "SOFP_conversation_trace.json").read_text()
    )
    assert "code_mode_calls" not in payload


def test_recording_failure_never_breaks_the_call():
    """The never-raise contract: a pathological arg/result whose __str__
    raises must cost only the ledger entry, never the tool call."""
    from extraction.code_mode import ledgered

    class _Unprintable:
        def __str__(self):
            raise RuntimeError("unprintable")
        __repr__ = __str__

    class _Deps:
        code_mode_ledger: list = []

    class _Ctx:
        deps = _Deps()
        tool_call_id = "pyd_ai_x__1"

    @ledgered
    def tool(ctx, value):
        return _Unprintable()

    result = tool(_Ctx(), value=_Unprintable())
    assert isinstance(result, _Unprintable), "the call itself must succeed"

    @ledgered
    def failing_tool(ctx, value):
        raise ValueError("the real error")

    with pytest.raises(ValueError, match="the real error"):
        failing_tool(_Ctx(), value=_Unprintable())


def test_ledger_echo_format():
    assert ledger_echo([]) == ""
    line = ledger_echo([
        {"tool": "write_facts", "outcome": "ok"},
        {"tool": "verify_totals", "outcome": "error",
         "error": "RuntimeError: boom"},
    ])
    assert "write_facts:ok" in line
    assert "verify_totals:FAILED(RuntimeError: boom)" in line
