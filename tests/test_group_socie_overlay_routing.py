"""Group SOCIE must not be handed a layout its template does not have.

Peer review, 2026-08-01. `render_prompt` appended `_group_socie_overlay.md`
to EVERY Group SOCIE regardless of filing standard or variant. That overlay
describes the MFRS matrix — "the same 24 equity-component columns B through
X in each block", four blocks at rows 3-25/27-49/51-73/75-97 — and it tells
the agent to "use explicit `row` and `col` coordinates".

Measured against the live templates, that is wrong twice:

    MFRS  Group SOCIE   24 cols x 97 rows   overlay correct
    MPERS Group SOCIE    4 cols x 97 rows   column claim wrong
    MPERS Group SoRE     6 cols x 16 rows   column AND row claim wrong

The failure is silent, which is why this is pinned rather than left to
review. `concept_model.cell_resolver.resolve_cell` returns None for a
coordinate that maps to no concept and, by its own docstring, "the caller
skips it rather than failing the whole projection". A write to column X of a
4-column sheet is therefore dropped before it reaches `run_concept_facts`,
and because the export re-renders from facts (gotcha #21) the equity
statement lands empty while the agent reports success.

These tests read the LIVE templates, so they keep working if a template is
regenerated. The existing `test_socie_prompt_mpers.py` passes throughout the
bug because it asserts selected phrases independently; the contradiction only
shows up once you look at the fully rendered prompt, which is what the last
test here does.
"""
from __future__ import annotations

import openpyxl
import pytest

from prompts import render_prompt
from statement_types import StatementType, template_path


# Phrases unique to each overlay, used to detect which one was appended.
MATRIX_MARKER = "columns B through X"
FOUR_BLOCK_MARKER = "75-97"
SIX_COL_MARKER = "6 data columns"


def _render(standard: str, variant: str, level: str = "group") -> str:
    """Rendered prompt with en/em dashes folded to '-'.

    The prompt files mix `75-97` and `75–97`; a raw substring check silently
    passes on one spelling and fails on the other.
    """
    text = render_prompt(
        StatementType.SOCIE, variant=variant,
        filing_level=level, filing_standard=standard,
    )
    return text.replace("–", "-").replace("—", "-")


def _template_shape(standard: str, variant: str, level: str = "group"):
    """(max_column, max_row) of the live SOCIE template."""
    path = template_path(StatementType.SOCIE, variant, level=level, standard=standard)
    ws = openpyxl.load_workbook(path)[openpyxl.load_workbook(path).sheetnames[0]]
    return ws.max_column, ws.max_row


# ---------------------------------------------------------------------------
# The templates these prompts describe — pinned so a regeneration that changes
# a shape fails here loudly instead of silently invalidating the prompts.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("standard,variant,cols,rows", [
    ("mfrs", "Default", 24, 97),
    ("mpers", "Default", 4, 97),
    ("mpers", "SoRE", 6, 16),
])
def test_live_group_socie_template_shapes(standard, variant, cols, rows):
    assert _template_shape(standard, variant) == (cols, rows)


# ---------------------------------------------------------------------------
# Routing — each shape gets the overlay that matches it, and only that one.
# ---------------------------------------------------------------------------

def test_mfrs_group_socie_keeps_the_matrix_overlay():
    """No regression: the MFRS template really is 24 cols x 97 rows."""
    prompt = _render("mfrs", "Default")
    assert MATRIX_MARKER in prompt
    assert FOUR_BLOCK_MARKER in prompt


def test_mpers_group_socie_never_claims_24_columns():
    """The MPERS Group SOCIE template has 4 columns. Columns E-X do not
    exist, so a write there is silently dropped by resolve_cell."""
    prompt = _render("mpers", "Default")
    assert MATRIX_MARKER not in prompt
    # It may only DENY the component columns (socie_mpers.md does), never
    # assert them.
    assert "per-equity-component columns" in prompt
    # The 4-block structure is real on MPERS and must survive — it just
    # comes from socie_mpers.md rather than the MFRS overlay.
    assert FOUR_BLOCK_MARKER in prompt
    assert "Company - Prior period" in prompt


def test_mpers_group_sore_gets_no_block_layout_at_all():
    """SoRE is 16 rows. The overlay's block map (rows 27-97) addresses rows
    that do not exist in the sheet."""
    prompt = _render("mpers", "SoRE")
    assert MATRIX_MARKER not in prompt
    assert FOUR_BLOCK_MARKER not in prompt
    assert "51-73" not in prompt
    # SoRE is a plain 6-column Group template: B/C Group, D/E Company, F source.
    assert SIX_COL_MARKER in prompt
    assert "D = Company CY" in prompt


def test_company_filings_get_no_group_overlay():
    for standard, variant in [("mfrs", "Default"), ("mpers", "Default"), ("mpers", "SoRE")]:
        prompt = _render(standard, variant, level="company")
        assert MATRIX_MARKER not in prompt, (standard, variant)
        assert SIX_COL_MARKER not in prompt, (standard, variant)


# ---------------------------------------------------------------------------
# The check the old suite was missing: mutually exclusive instructions in ONE
# rendered prompt. Asserting phrases independently cannot catch this.
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("standard,variant", [
    ("mfrs", "Default"),
    ("mpers", "Default"),
    ("mpers", "SoRE"),
])
@pytest.mark.parametrize("level", ["company", "group"])
def test_rendered_prompt_asserts_one_column_layout(standard, variant, level):
    """A prompt must not both deny and assert the 24-column matrix, and its
    column claims must fit the template it will actually be writing into."""
    prompt = _render(standard, variant, level=level)
    denies_matrix = "NOT the MFRS" in prompt or "no per-equity-component" in prompt
    asserts_matrix = MATRIX_MARKER in prompt
    assert not (denies_matrix and asserts_matrix), (
        f"{standard}/{variant}/{level}: prompt both denies and asserts the "
        "24-column matrix layout"
    )

    max_col, _ = _template_shape(standard, variant, level=level)
    if asserts_matrix:
        assert max_col >= 24, (
            f"{standard}/{variant}/{level}: prompt says columns B through X "
            f"but the template has {max_col} columns"
        )
    if SIX_COL_MARKER in prompt:
        assert max_col == 6, (
            f"{standard}/{variant}/{level}: prompt says 6 data columns but "
            f"the template has {max_col}"
        )
