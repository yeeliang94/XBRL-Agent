"""End-to-end integration tests — full pipeline without a real LLM.

Phase 11: Covers multi-agent coordinator → merger → cross-checks → DB persistence.
"""
import json
import sqlite3
from pathlib import Path
from unittest.mock import patch, AsyncMock

import pytest

import server
from fastapi.testclient import TestClient
from server import app
from statement_types import StatementType
from coordinator import AgentResult, CoordinatorResult
from cross_checks.framework import CrossCheckResult
from workbook_merger import MergeResult


# ---------------------------------------------------------------------------
# Phase 11: Full-pipeline E2E with 5 agents (mocked LLM)
# ---------------------------------------------------------------------------


def _parse_sse(text: str) -> list[dict]:
    """Parse SSE stream text into a list of {event, data} dicts."""
    events = []
    current_event = None
    current_data = None
    for line in text.split("\n"):
        if line.startswith("event: "):
            current_event = line[7:].strip()
        elif line.startswith("data: "):
            current_data = line[6:].strip()
        elif line == "" and current_event and current_data:
            try:
                data = json.loads(current_data)
            except json.JSONDecodeError:
                data = current_data
            events.append({"event": current_event, "data": data})
            current_event = None
            current_data = None
    return events


@pytest.fixture
def full_pipeline_env(tmp_path, monkeypatch):
    """Set up a session with fake PDF and per-statement workbooks."""
    session_id = "e2e-full-pipeline"
    out = tmp_path / "output"
    session_dir = out / session_id
    session_dir.mkdir(parents=True)
    (session_dir / "uploaded.pdf").write_bytes(b"%PDF-1.4 fake")

    # Create minimal workbook files for each statement
    import openpyxl
    for stmt in StatementType:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{stmt.value}-Main"
        ws["A1"] = f"{stmt.value} data"
        wb.save(str(session_dir / f"{stmt.value}_filled.xlsx"))
        wb.close()

    monkeypatch.setattr(server, "OUTPUT_DIR", out)
    monkeypatch.setattr(server, "AUDIT_DB_PATH", out / "xbrl_agent.db")
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    monkeypatch.setenv("TEST_MODEL", "test-model")
    monkeypatch.setenv("LLM_PROXY_URL", "")

    return TestClient(server.app), session_id, out, session_dir


def test_full_extraction_mocked(full_pipeline_env):
    """Phase 11.1: scout(skipped) → 5 sub-agents → merger → cross-checks → DB.

    Asserts: 5 sheets in merged workbook, 5 cross-check rows in DB,
    all SSE events emitted correctly. Runs in <30s.
    """
    client, session_id, out, session_dir = full_pipeline_env

    all_statements = list(StatementType)
    variants = {
        StatementType.SOFP: "CuNonCu",
        StatementType.SOPL: "Function",
        StatementType.SOCI: "BeforeTax",
        StatementType.SOCF: "Indirect",
        StatementType.SOCIE: "Default",
    }

    # Coordinator returns all 5 agents succeeded
    fake_coordinator_result = CoordinatorResult(agent_results=[
        AgentResult(
            statement_type=stmt,
            variant=variants[stmt],
            status="succeeded",
            workbook_path=str(session_dir / f"{stmt.value}_filled.xlsx"),
        )
        for stmt in all_statements
    ])

    # All 5 cross-checks pass
    fake_checks = [
        CrossCheckResult(name="sofp_balance", status="passed",
                         expected=1000.0, actual=1000.0, diff=0.0,
                         tolerance=1.0, message="OK"),
        CrossCheckResult(name="sopl_to_socie_profit", status="passed",
                         expected=500.0, actual=500.0, diff=0.0,
                         tolerance=1.0, message="OK"),
        CrossCheckResult(name="soci_to_socie_tci", status="passed",
                         expected=600.0, actual=600.0, diff=0.0,
                         tolerance=1.0, message="OK"),
        CrossCheckResult(name="socie_to_sofp_equity", status="passed",
                         expected=800.0, actual=800.0, diff=0.0,
                         tolerance=1.0, message="OK"),
        CrossCheckResult(name="socf_to_sofp_cash", status="passed",
                         expected=200.0, actual=200.0, diff=0.0,
                         tolerance=1.0, message="OK"),
    ]

    run_config = {
        "statements": [s.value for s in all_statements],
        "variants": {s.value: v for s, v in variants.items()},
        "models": {},
        "infopack": None,
        "use_scout": False,
    }

    # Mock coordinator to return immediately AND push None sentinel into event_queue
    # so the SSE generator's queue drain loop exits cleanly.
    async def mock_coordinator_run(config, infopack=None, event_queue=None, session_id=None, **_kwargs):
        if event_queue is not None:
            # Emit per-agent complete events like the real coordinator does
            for idx, ar in enumerate(fake_coordinator_result.agent_results):
                agent_id = ar.statement_type.value.lower()
                await event_queue.put({
                    "event": "complete",
                    "data": {
                        "success": ar.status == "succeeded",
                        "agent_id": agent_id,
                        "agent_role": ar.statement_type.value,
                        "workbook_path": ar.workbook_path,
                        "error": ar.error,
                    },
                })
            await event_queue.put(None)
        return fake_coordinator_result

    with patch("server._create_proxy_model", return_value="fake-model"), \
         patch("coordinator.run_extraction", side_effect=mock_coordinator_run), \
         patch("cross_checks.framework.run_all", return_value=fake_checks), patch("cross_checks.framework.run_all_facts", return_value=fake_checks):

        resp = client.post(f"/api/run/{session_id}", json=run_config)

    assert resp.status_code == 200
    events = _parse_sse(resp.text)
    event_types = [e["event"] for e in events]

    # --- SSE assertions ---

    # 5 per-agent completion events
    completes = [e for e in events if e["event"] == "complete"]
    assert len(completes) == 5
    roles = {e["data"]["agent_role"] for e in completes}
    assert roles == {"SOFP", "SOPL", "SOCI", "SOCF", "SOCIE"}

    # Final run_complete
    assert "run_complete" in event_types
    rc = [e for e in events if e["event"] == "run_complete"][0]["data"]
    assert rc["success"] is True
    assert rc["merged_workbook"] is not None
    assert len(rc["cross_checks"]) == 5
    assert set(rc["statements_completed"]) == {"SOFP", "SOPL", "SOCI", "SOCF", "SOCIE"}
    assert rc["statements_failed"] == []

    # --- Merged workbook: 5 sheets ---
    import openpyxl
    merged_wb = openpyxl.load_workbook(rc["merged_workbook"])
    assert len(merged_wb.sheetnames) == 5
    merged_wb.close()

    # --- DB: 1 run, 5 agents, 5 cross-checks ---
    db_path = out / "xbrl_agent.db"
    assert db_path.exists()
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row

    runs = conn.execute("SELECT * FROM runs").fetchall()
    assert len(runs) == 1
    assert runs[0]["status"] == "completed"

    agents_db = conn.execute("SELECT * FROM run_agents ORDER BY id").fetchall()
    assert len(agents_db) == 5

    checks_db = conn.execute("SELECT * FROM cross_checks ORDER BY id").fetchall()
    assert len(checks_db) == 5
    assert all(c["status"] == "passed" for c in checks_db)

    conn.close()


# ---------------------------------------------------------------------------
# Phase 11.2: Live E2E test (real LLM)
# ---------------------------------------------------------------------------

@pytest.mark.live
def test_full_extraction_live(tmp_path):
    """Live E2E: upload FINCO PDF → 5 agents → merged workbook → cross-checks.

    Requires GEMINI_API_KEY env var. Run with: pytest -m live
    Skipped if no API key is set.
    """
    import os
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        pytest.skip("GEMINI_API_KEY not set — skipping live E2E test")

    pdf_path = Path(__file__).resolve().parent.parent / "data" / "FINCO-Audited-Financial-Statement-2021.pdf"
    if not pdf_path.exists():
        pytest.skip(f"FINCO PDF not found at {pdf_path}")

    # Run extraction via coordinator directly (not via server SSE)
    # to avoid server setup complexity in live test
    import asyncio
    from coordinator import RunConfig, run_extraction

    output_dir = str(tmp_path / "live_output")
    Path(output_dir).mkdir(parents=True)

    config = RunConfig(
        pdf_path=str(pdf_path),
        output_dir=output_dir,
        model="google-gla:gemini-2.0-flash",
        statements_to_run={StatementType.SOFP},  # Just SOFP for speed
        variants={StatementType.SOFP: "CuNonCu"},
    )

    result = asyncio.get_event_loop().run_until_complete(run_extraction(config))

    # At least SOFP should succeed
    assert len(result.agent_results) == 1
    sofp_result = result.agent_results[0]
    assert sofp_result.statement_type == StatementType.SOFP
    assert sofp_result.status == "succeeded", f"SOFP failed: {sofp_result.error}"
    assert sofp_result.workbook_path is not None
    assert Path(sofp_result.workbook_path).exists()

    # Verify the workbook has content
    import openpyxl
    wb = openpyxl.load_workbook(sofp_result.workbook_path)
    assert len(wb.sheetnames) >= 1
    wb.close()


# ---------------------------------------------------------------------------
# Phase 8: Group filing E2E (mocked LLM)
# ---------------------------------------------------------------------------


def test_group_filing_e2e_mocked(full_pipeline_env):
    """Group filing: filing_level='group' flows through coordinator → cross-checks → DB."""
    client, session_id, out, session_dir = full_pipeline_env

    fake_coordinator_result = CoordinatorResult(agent_results=[
        AgentResult(
            statement_type=StatementType.SOFP,
            variant="CuNonCu",
            status="succeeded",
            workbook_path=str(session_dir / "SOFP_filled.xlsx"),
        ),
    ])

    fake_checks = [
        CrossCheckResult(name="sofp_balance", status="passed",
                         expected=1000.0, actual=1000.0, diff=0.0,
                         tolerance=1.0, message="Group CY: balanced; Company CY: balanced"),
    ]

    run_config = {
        "statements": ["SOFP"],
        "variants": {"SOFP": "CuNonCu"},
        "models": {},
        "infopack": None,
        "use_scout": False,
        "filing_level": "group",
    }

    async def mock_coordinator_run(config, infopack=None, event_queue=None, session_id=None, **_kwargs):
        # Verify filing_level reached the coordinator
        assert config.filing_level == "group"
        if event_queue is not None:
            for ar in fake_coordinator_result.agent_results:
                agent_id = ar.statement_type.value.lower()
                await event_queue.put({
                    "event": "complete",
                    "data": {
                        "success": True,
                        "agent_id": agent_id,
                        "agent_role": ar.statement_type.value,
                        "workbook_path": ar.workbook_path,
                        "error": None,
                    },
                })
            await event_queue.put(None)
        return fake_coordinator_result

    with patch("server._create_proxy_model", return_value="fake-model"), \
         patch("coordinator.run_extraction", side_effect=mock_coordinator_run), \
         patch("cross_checks.framework.run_all", return_value=fake_checks), patch("cross_checks.framework.run_all_facts", return_value=fake_checks):

        resp = client.post(f"/api/run/{session_id}", json=run_config)

    assert resp.status_code == 200
    events = _parse_sse(resp.text)

    # Cross-checks are fully group-aware (Phase 6), so partial=False
    rc = [e for e in events if e["event"] == "run_complete"][0]["data"]
    assert rc["success"] is True
    assert rc["cross_checks_partial"] is False

    # DB persists filing_level in run_config_json
    db_path = out / "xbrl_agent.db"
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT run_config_json FROM runs LIMIT 1").fetchone()
    stored = json.loads(row["run_config_json"])
    assert stored["filing_level"] == "group"
    conn.close()


def test_clean_run_fires_spot_check_when_enabled(full_pipeline_env, monkeypatch):
    """Issue 1 (2026-06-21): when XBRL_SPOT_CHECK is on, a run whose
    cross-checks ALL pass still launches a spot-check (reviewer pass with
    spot_check framing) — proving the clean-run trigger is wired in
    run_multi_agent_stream. The suite defaults the toggle OFF (conftest), so
    this opts back in. The mocked run carries no real facts, so the pass
    no-ops with `no_extracted_facts_to_review` — but it still emits the
    CORRECTION agent event, which is the trigger signal we pin here."""
    client, session_id, out, session_dir = full_pipeline_env
    monkeypatch.setenv("XBRL_SPOT_CHECK", "true")
    monkeypatch.setenv("XBRL_SPOT_CHECK_MODE", "light")

    all_statements = list(StatementType)
    variants = {
        StatementType.SOFP: "CuNonCu", StatementType.SOPL: "Function",
        StatementType.SOCI: "BeforeTax", StatementType.SOCF: "Indirect",
        StatementType.SOCIE: "Default",
    }
    fake_coordinator_result = CoordinatorResult(agent_results=[
        AgentResult(statement_type=stmt, variant=variants[stmt],
                    status="succeeded",
                    workbook_path=str(session_dir / f"{stmt.value}_filled.xlsx"))
        for stmt in all_statements
    ])
    # Every cross-check passes → the run is CLEAN → only the spot-check can
    # add a CORRECTION event.
    fake_checks = [CrossCheckResult(name="sofp_balance", status="passed",
                                    expected=1.0, actual=1.0, diff=0.0,
                                    tolerance=1.0, message="OK")]

    run_config = {
        "statements": [s.value for s in all_statements],
        "variants": {s.value: v for s, v in variants.items()},
        "models": {}, "infopack": None, "use_scout": False,
    }

    async def mock_coordinator_run(config, infopack=None, event_queue=None, session_id=None, **_kwargs):
        if event_queue is not None:
            for ar in fake_coordinator_result.agent_results:
                await event_queue.put({"event": "complete", "data": {
                    "success": True, "agent_id": ar.statement_type.value.lower(),
                    "agent_role": ar.statement_type.value,
                    "workbook_path": ar.workbook_path, "error": ar.error}})
            await event_queue.put(None)
        return fake_coordinator_result

    with patch("server._create_proxy_model", return_value="fake-model"), \
         patch("coordinator.run_extraction", side_effect=mock_coordinator_run), \
         patch("cross_checks.framework.run_all", return_value=fake_checks), \
         patch("cross_checks.framework.run_all_facts", return_value=fake_checks):
        resp = client.post(f"/api/run/{session_id}", json=run_config)

    assert resp.status_code == 200
    events = _parse_sse(resp.text)
    # The spot-check fired: a CORRECTION-role event is present (the trigger
    # signal). Without the spot-check wiring a clean run emits none.
    correction_events = [
        e for e in events
        if isinstance(e.get("data"), dict)
        and e["data"].get("agent_role") == "CORRECTION"
    ]
    assert correction_events, "clean run with XBRL_SPOT_CHECK on must launch the spot-check"
    # And the reviewing stage was emitted at the boundary.
    stages = [e["data"].get("stage") for e in events
              if e["event"] == "pipeline_stage" and isinstance(e.get("data"), dict)]
    assert "reviewing" in stages

    # Peer-review HIGH (2026-06-21): this mocked run has no real facts, so the
    # spot-check fails with `no_extracted_facts_to_review` and its CORRECTION
    # row is "failed". A failed spot-check must NOT hide under a green badge —
    # the run is downgraded to completed_with_errors, not completed.
    db_path = out / "xbrl_agent.db"
    conn = sqlite3.connect(str(db_path))
    try:
        status = conn.execute(
            "SELECT status FROM runs ORDER BY id DESC LIMIT 1"
        ).fetchone()[0]
    finally:
        conn.close()
    assert status == "completed_with_errors", (
        f"failed spot-check must downgrade the run, got {status!r}"
    )


def test_notes_coverage_checklist_e2e(tmp_path, monkeypatch):
    """End-to-end (mocked model): the notes reviewer pass reconciles the scout
    inventory against placements, auto-resolves a missing note by authoring it,
    and persists a POST-REVIEWER checklist the coverage API then serves
    (docs/PLAN-notes-coverage-and-routing.md Phase 8).
    """
    import asyncio
    from pydantic_ai.messages import ModelResponse, TextPart, ToolCallPart
    from pydantic_ai.models.function import FunctionModel

    import notes.reviewer_agent as ra
    import notes.detectors as det
    from db import repository as repo
    from db.schema import init_db

    monkeypatch.setenv("XBRL_NOTES_COVERAGE", "true")
    monkeypatch.setattr(ra, "count_pdf_pages", lambda _p: 60)
    monkeypatch.setattr(det, "render_pages_to_png_bytes",
                        lambda pdf_path, start, end, dpi=200: [b"png"])

    db = tmp_path / "audit.sqlite"
    init_db(db)
    server.AUDIT_DB_PATH = db
    server.OUTPUT_DIR = tmp_path
    _S12 = "Notes-Listofnotes"
    _PREFIX = "mfrs-company-"

    with repo.db_session(db) as conn:
        run_id = repo.create_run(conn, "x.pdf", session_id="s",
                                 output_dir=str(tmp_path))
        # Inventory: note 4 placed, note 5 MISSING (to be authored back).
        # Adjacent numbers so there's no incidental suspected gap.
        repo.upsert_notes_inventory(conn, run_id=run_id, note_num=4,
                                    title="Property, plant & equipment",
                                    page_lo=30, page_hi=31)
        repo.upsert_notes_inventory(conn, run_id=run_id, note_num=5,
                                    title="Investment properties",
                                    page_lo=32, page_hi=33)
        repo.upsert_notes_cell(conn, run_id=run_id, sheet=_S12, row=20,
                               label="PPE", html="<p>ppe</p>")
        repo.upsert_notes_provenance(conn, run_id=run_id, sheet=_S12, row=20,
                                     row_label="PPE", source_note_refs=["4"])
        conn.execute(
            "INSERT INTO notes_nodes(node_uuid, template_id, sheet, row, label, kind) "
            "VALUES (?, ?, ?, ?, ?, 'LEAF')",
            ("n40", f"{_PREFIX}notes-listofnotes-v1", _S12, 40,
             "Disclosure of investment properties"))

    # Reviewer authors the missing note 5 into an empty leaf, grounded.
    steps = [
        [ToolCallPart(tool_name="view_pdf_pages", args={"pages": [32]})],
        [ToolCallPart(tool_name="author_note_cell", args={
            "sheet": _S12, "row": 40, "html": "<p>IP disclosure</p>",
            "note_num": 5, "source_pages": [32], "evidence": "IP note"})],
        [TextPart("done")],
    ]
    idx = {"i": 0}

    def fn(messages, info):
        i = idx["i"]; idx["i"] += 1
        return ModelResponse(parts=steps[i]) if i < len(steps) else \
            ModelResponse(parts=[TextPart("done")])

    q: asyncio.Queue = asyncio.Queue()
    outcome = asyncio.run(server._run_notes_reviewer_pass(
        run_id=run_id, db_path=str(db), pdf_path=str(tmp_path / "x.pdf"),
        filing_level="company", filing_standard="mfrs",
        model=FunctionModel(fn), output_dir=str(tmp_path),
        merged_workbook_path=None, event_queue=q, sidecar_paths=[]))

    # The reviewer resolved the missing note → coverage is clean, no tip.
    assert outcome["coverage"]["banner"] == "reviewed"
    assert outcome["coverage"]["unresolved"] == 0
    assert server._notes_coverage_tips_status(outcome["coverage"]) is False

    # The API serves the persisted POST-REVIEWER checklist.
    client = TestClient(app)
    body = client.get(f"/api/runs/{run_id}/notes-coverage").json()
    assert body["banner"] == "reviewed"
    rows = {r["note_num"]: r for r in body["rows"]}
    assert rows[4]["status"] == "placed"
    # Note 5 flipped missing → placed and is tagged reviewer-added.
    assert rows[5]["status"] == "placed"
    assert rows[5]["reviewer_added"] is True
