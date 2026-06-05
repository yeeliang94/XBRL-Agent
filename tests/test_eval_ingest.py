"""Reverse-ingestion tests (eval/ingest.py).

Builds a SMALL fixture workbook from a CURRENT canonical template (NOT
``SOFP-Xbrl-reference-FINCO-filled.xlsx`` — it carries a +1 row shift, gotcha
#4): import the live MFRS Company SOFP template into a fresh DB, blank every
value cell, fill a handful of known LEAF cells, then assert ``ingest_workbook``
reproduces exactly those gold facts.
"""
from __future__ import annotations

import json
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string

from db.schema import init_db
from eval.ingest import ingest_workbook, parse_accounting_number
from statement_types import StatementType, template_path


def _import_company_sofp(db_path) -> str:
    """Import the live MFRS Company SOFP-CuNonCu template; return template_id."""
    from concept_model.importer import import_template, import_company_targets
    from concept_model.parser import parse_template

    tpath = template_path(StatementType.SOFP, "CuNonCu", "company", "mfrs")
    tree = parse_template(str(tpath))
    payload = tree.to_json()
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as fh:
        json.dump(payload, fh, sort_keys=True)
        json_path = fh.name
    try:
        template_id = import_template(db_path, json_path)
    finally:
        Path(json_path).unlink(missing_ok=True)
    import_company_targets(db_path, template_id)
    return template_id


def _seed_benchmark(conn, template_id) -> int:
    cur = conn.execute(
        "INSERT INTO eval_benchmarks(name, filing_standard, filing_level) "
        "VALUES ('FINCO 2021', 'mfrs', 'company')"
    )
    bench_id = int(cur.lastrowid)
    conn.execute(
        "INSERT INTO eval_benchmark_templates(benchmark_id, template_id, "
        "statement_type) VALUES (?, ?, 'SOFP')",
        (bench_id, template_id),
    )
    conn.commit()
    return bench_id


def _build_fixture_workbook(conn, template_id, dest_path, n_leaves=5):
    """Blank every value cell of the live template, then fill ``n_leaves``
    known LEAF cells at their CY (col B) coordinate. Returns the list of
    ``(concept_uuid, value)`` written so the test can spot-check."""
    tpath = template_path(StatementType.SOFP, "CuNonCu", "company", "mfrs")
    wb = openpyxl.load_workbook(str(tpath), data_only=False)
    # Blank all value columns so the only numbers are the ones we write.
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_col=2, max_col=30):
            for cell in row:
                cell.value = None

    leaves = conn.execute(
        "SELECT concept_uuid, render_sheet, render_row, render_col "
        "FROM concept_nodes WHERE template_id = ? AND kind = 'LEAF' "
        "ORDER BY render_sheet, render_row LIMIT ?",
        (template_id, n_leaves),
    ).fetchall()
    assert len(leaves) == n_leaves, "template should expose >= n_leaves leaves"

    written = []
    for i, (uuid, sheet, row, render_col) in enumerate(leaves):
        ws = wb[sheet]
        col_idx = column_index_from_string(render_col or "B")
        value = 1000.0 + i  # distinct, deterministic
        ws.cell(row=row, column=col_idx).value = value
        written.append((uuid, value))
    wb.save(dest_path)
    wb.close()
    return written


def test_parse_accounting_number_handles_text_conventions():
    assert parse_accounting_number(1595) == 1595.0
    assert parse_accounting_number(1595.0) == 1595.0
    assert parse_accounting_number("1,595") == 1595.0
    assert parse_accounting_number("(95)") == -95.0
    assert parse_accounting_number("-95") == -95.0
    assert parse_accounting_number("1,234.56") == 1234.56
    assert parse_accounting_number("-") is None
    assert parse_accounting_number("") is None
    assert parse_accounting_number("N/A") is None
    assert parse_accounting_number(None) is None
    # Unbalanced parens are malformed text, NOT a coerced positive value:
    # the old `\(?...\)?` form silently turned "(95" into +95 and stored it.
    assert parse_accounting_number("(95") is None
    assert parse_accounting_number("95)") is None


def test_ingest_reproduces_filled_leaf_cells(tmp_path):
    db = tmp_path / "ingest.db"
    init_db(db)
    template_id = _import_company_sofp(db)

    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        bench_id = _seed_benchmark(conn, template_id)
        fixture = tmp_path / "filled.xlsx"
        written = _build_fixture_workbook(conn, template_id, str(fixture))

        result = ingest_workbook(conn, bench_id, str(fixture), [template_id])
        conn.commit()

        # Exactly the cells we filled became gold (only CY/col B was filled).
        assert result.ingested == len(written)
        gold_count = conn.execute(
            "SELECT COUNT(*) FROM gold_concept_facts WHERE benchmark_id = ?",
            (bench_id,),
        ).fetchone()[0]
        assert gold_count == len(written)

        # Spot-check a known concept's gold value + dimensions.
        uuid, value = written[0]
        row = conn.execute(
            "SELECT value, period, entity_scope, value_status "
            "FROM gold_concept_facts WHERE benchmark_id = ? AND concept_uuid = ?",
            (bench_id, uuid),
        ).fetchone()
        assert row is not None
        assert abs(row[0] - value) < 1e-9
        assert row[1] == "CY"
        assert row[2] == "Company"
        assert row[3] == "observed"
    finally:
        conn.close()


def test_benchmark_concepts_derives_computed_totals_from_gold_leaves(tmp_path):
    """The gold editor must SHOW totals: a COMPUTED parent carries no gold row,
    so its value is derived on-read from the gold leaves (display-only cascade).
    """
    from eval import store

    db = tmp_path / "totals.db"
    init_db(db)
    template_id = _import_company_sofp(db)
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        bench_id = _seed_benchmark(conn, template_id)

        # Find a COMPUTED parent whose children are ALL leaves (single-level
        # sum) so the expected total is unambiguous.
        target = None
        for (puid,) in conn.execute(
            "SELECT concept_uuid FROM concept_nodes "
            "WHERE template_id = ? AND kind = 'COMPUTED'",
            (template_id,),
        ).fetchall():
            edges = conn.execute(
                "SELECT child_uuid, coefficient FROM concept_edges "
                "WHERE parent_uuid = ?",
                (puid,),
            ).fetchall()
            if not edges:
                continue
            child_kinds = {
                conn.execute(
                    "SELECT kind FROM concept_nodes WHERE concept_uuid = ?",
                    (cuid,),
                ).fetchone()[0]
                for cuid, _ in edges
            }
            if child_kinds == {"LEAF"}:
                target = (puid, edges)
                break
        assert target is not None, "expected a COMPUTED parent over LEAF children"
        puid, edges = target

        expected = 0.0
        for i, (cuid, coef) in enumerate(edges):
            val = 100.0 + i
            expected += float(coef) * val
            conn.execute(
                "INSERT INTO gold_concept_facts(benchmark_id, concept_uuid, "
                "period, entity_scope, value, value_status, source, updated_at) "
                "VALUES (?, ?, 'CY', 'Company', ?, 'observed', 'test', '')",
                (bench_id, cuid, val),
            )
        conn.commit()

        rows = {r["concept_uuid"]: r for r in store.benchmark_concepts(conn, bench_id)}
        assert puid in rows, "COMPUTED parent must appear in the grid"
        # Derived on-read into both the primary value and the scope_facts map.
        assert rows[puid]["value"] == round(expected, 2)
        assert rows[puid]["scope_facts"]["Company"]["CY"] == round(expected, 2)
        # Display-only: nothing was persisted for the COMPUTED parent.
        persisted = conn.execute(
            "SELECT COUNT(*) FROM gold_concept_facts "
            "WHERE benchmark_id = ? AND concept_uuid = ?",
            (bench_id, puid),
        ).fetchone()[0]
        assert persisted == 0
    finally:
        conn.close()


def test_ingest_rejects_workbook_matching_no_template(tmp_path):
    db = tmp_path / "ingest.db"
    init_db(db)
    template_id = _import_company_sofp(db)
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        bench_id = _seed_benchmark(conn, template_id)
        # A workbook whose sheet names match nothing in the template set.
        wb = openpyxl.Workbook()
        wb.active.title = "TotallyUnrelatedSheet"
        wb.active["B5"] = 123
        alien = tmp_path / "alien.xlsx"
        wb.save(str(alien))
        wb.close()

        with pytest.raises(ValueError, match="No worksheet"):
            ingest_workbook(conn, bench_id, str(alien), [template_id])
    finally:
        conn.close()


def test_ingest_counts_uncached_formula_cells_as_warning(tmp_path):
    """A gradeable LEAF cell holding a live formula with NO cached value reads
    as ``None`` under ``data_only=True`` and is silently dropped. The ingest
    must COUNT those so the caller can warn (the 2026-06-05 sub-sheet-loss
    incident: machine-exported SOCIE/rollup formulas vanished into sparse gold).
    """
    db = tmp_path / "formula.db"
    init_db(db)
    template_id = _import_company_sofp(db)

    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        bench_id = _seed_benchmark(conn, template_id)

        # Two LEAF coords: one filled with a literal, one with a formula string.
        leaves = conn.execute(
            "SELECT concept_uuid, render_sheet, render_row, render_col "
            "FROM concept_nodes WHERE template_id = ? AND kind = 'LEAF' "
            "ORDER BY render_sheet, render_row LIMIT 2",
            (template_id,),
        ).fetchall()
        assert len(leaves) == 2

        tpath = template_path(StatementType.SOFP, "CuNonCu", "company", "mfrs")
        wb = openpyxl.load_workbook(str(tpath), data_only=False)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_col=2, max_col=30):
                for cell in row:
                    cell.value = None

        lit_uuid, lit_sheet, lit_row, lit_col = leaves[0]
        f_uuid, f_sheet, f_row, f_col = leaves[1]
        wb[lit_sheet].cell(
            row=lit_row, column=column_index_from_string(lit_col or "B")
        ).value = 4242.0
        # A formula with NO cached value — exactly what an un-recalculated
        # openpyxl export produces.
        wb[f_sheet].cell(
            row=f_row, column=column_index_from_string(f_col or "B")
        ).value = "=1000+1"
        fixture = tmp_path / "formula.xlsx"
        wb.save(str(fixture))
        wb.close()

        result = ingest_workbook(conn, bench_id, str(fixture), [template_id])
        conn.commit()

        # The literal became gold; the formula leaf was lost but COUNTED.
        assert result.ingested == 1
        assert result.skipped_formula_cells == 1
        assert conn.execute(
            "SELECT COUNT(*) FROM gold_concept_facts WHERE benchmark_id = ? "
            "AND concept_uuid = ?", (bench_id, f_uuid),
        ).fetchone()[0] == 0
    finally:
        conn.close()


def test_create_from_workbook_names_formula_cause_when_all_cells_uncached(tmp_path):
    """An all-formula export (ingested == 0, skipped_formula_cells > 0) must
    raise an error that names the un-cached-formula cause, not the generic
    'value columns appear empty' message (peer-review follow-up)."""
    from eval import store

    db = tmp_path / "allformula.db"
    init_db(db)
    template_id = _import_company_sofp(db)

    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        leaf = conn.execute(
            "SELECT render_sheet, render_row, render_col FROM concept_nodes "
            "WHERE template_id = ? AND kind = 'LEAF' "
            "ORDER BY render_sheet, render_row LIMIT 1",
            (template_id,),
        ).fetchone()
        sheet, row, col = leaf

        tpath = template_path(StatementType.SOFP, "CuNonCu", "company", "mfrs")
        wb = openpyxl.load_workbook(str(tpath), data_only=False)
        for ws in wb.worksheets:
            for r in ws.iter_rows(min_col=2, max_col=30):
                for cell in r:
                    cell.value = None
        # The ONLY populated gradeable cell is a formula with no cached value.
        wb[sheet].cell(
            row=row, column=column_index_from_string(col or "B")
        ).value = "=1000+1"
        fixture = tmp_path / "allformula.xlsx"
        wb.save(str(fixture))
        wb.close()

        with pytest.raises(ValueError, match="live formulas with no cached value"):
            store.create_benchmark_from_workbook(
                conn, name="all-formula", document=None,
                filing_standard="mfrs", filing_level="company",
                xlsx_path=str(fixture),
            )
    finally:
        conn.close()
