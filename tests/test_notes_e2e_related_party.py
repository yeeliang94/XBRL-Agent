"""Notes 14 (Related Party) — structured numeric round-trip test."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from notes.agent import render_notes_prompt
from notes.payload import NotesPayload
from notes.writer import write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path


SHEET = "Notes-RelatedPartytran"


def test_related_party_prompt_covers_transaction_and_balance_sections():
    prompt = render_notes_prompt(
        template_type=NotesTemplateType.RELATED_PARTY,
        filing_level="company",
        inventory=[],
    )
    assert SHEET in prompt
    prompt_lower = prompt.lower()
    assert "related party" in prompt_lower
    assert "outstanding" in prompt_lower
    assert "numeric_values" in prompt


def _find_row(ws, needle: str) -> int:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and needle.lower() in str(v).lower().lstrip("*"):
            return r
    raise AssertionError(f"{needle} not found")


def test_related_party_company_write_roundtrip(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.RELATED_PARTY, level="company")
    out = tmp_path / "NOTES_RELATED_PARTY_filled.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Management fees",
            content="",
            evidence="Page 55, Note 28",
            source_pages=[55],
            numeric_values={"company_cy": 240_000, "company_py": 200_000},
        ),
        NotesPayload(
            chosen_row_label="Amounts receivable",
            content="",
            evidence="Page 55, Note 28",
            source_pages=[55],
            numeric_values={"company_cy": 1_500_000, "company_py": 1_200_000},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=SHEET,
    )
    assert result.success, result.errors
    assert result.rows_written == 2

    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]
    mgmt_row = _find_row(ws, "Management fees")
    ar_row = _find_row(ws, "Amounts receivable")
    assert ws.cell(row=mgmt_row, column=2).value == 240_000
    assert ws.cell(row=mgmt_row, column=3).value == 200_000
    assert ws.cell(row=ar_row, column=2).value == 1_500_000
    assert ws.cell(row=ar_row, column=3).value == 1_200_000
    assert ws.cell(row=mgmt_row, column=4).value == "Page 55, Note 28"
    wb.close()


def test_related_party_group_write_fills_all_four_columns(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.RELATED_PARTY, level="group")
    out = tmp_path / "NOTES_RP_group.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Rental income",
            content="",
            evidence="Page 58, Note 29",
            source_pages=[58],
            numeric_values={
                "group_cy": 500_000,
                "group_py": 450_000,
                "company_cy": 300_000,
                "company_py": 280_000,
            },
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="group",
        sheet_name=SHEET,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]
    row = _find_row(ws, "Rental income")
    assert ws.cell(row=row, column=2).value == 500_000
    assert ws.cell(row=row, column=3).value == 450_000
    assert ws.cell(row=row, column=4).value == 300_000
    assert ws.cell(row=row, column=5).value == 280_000
    assert ws.cell(row=row, column=6).value == "Page 58, Note 29"
    wb.close()
