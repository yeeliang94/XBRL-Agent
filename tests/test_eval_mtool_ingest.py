"""Reverse-map a filled mTool workbook into gold facts (Step C1).

Builds a synthetic mTool-shaped workbook with openpyxl (as the offline_fill
tests do) and exercises the stdlib reader path end-to-end: strict label
matching, unit conversion, unmatched-row surfacing, and low-confidence refusal.
"""
from __future__ import annotations

import sqlite3

import pytest
from openpyxl import Workbook

from db.schema import init_db
from eval.mtool_ingest import (
    ColumnDetectionError,
    ConceptTarget,
    build_catalogue,
    count_deferred_matrix,
    extract_prose_gold,
    ingest_workbook,
)
from mtool.offline_fill import normalize_label

_TEMPLATE_ID = "mfrs-company-sofp-cunoncu-v1"


def _target(label):
    return ConceptTarget(
        concept_uuid=f"u_{normalize_label(label)}",
        template_id=_TEMPLATE_ID,
        canonical_label=label,
        statement_type="SOFP",
    )


def _catalogue(labels):
    return {"SOFP": {normalize_label(l): _target(l) for l in labels}}


def _mtool_file(tmp_path, rows, sheet="SOFP"):
    """rows: list of (label, cy, py). None value cells are left blank."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i, (label, cy, py) in enumerate(rows, start=3):
        ws[f"A{i}"] = label
        if cy is not None:
            ws[f"B{i}"] = cy
        if py is not None:
            ws[f"C{i}"] = py
    path = tmp_path / "human_filled.xlsx"
    wb.save(path)
    return str(path)


# --- happy path ------------------------------------------------------------

def test_matches_concepts_and_emits_facts(tmp_path):
    path = _mtool_file(tmp_path, [
        ("Cash and bank balances", 1595, 1420),
        ("Trade receivables", 800, None),
        ("Property, plant and equipment", 5000, 4800),
        ("Inventories", 300, 280),
        ("Other assets", 40, 30),
    ])
    catalogue = _catalogue([
        "Cash and bank balances", "Trade receivables",
        "Property, plant and equipment", "Inventories", "Other assets",
    ])
    report = ingest_workbook(path, catalogue, filing_level="company")

    # Cash → 2 facts (CY+PY); Trade receivables → 1 (PY blank); etc.
    facts = {(f.concept_uuid, f.period): f.value for f in report.facts}
    assert facts[("u_cash and bank balances", "CY")] == 1595
    assert facts[("u_cash and bank balances", "PY")] == 1420
    assert facts[("u_trade receivables", "CY")] == 800
    assert ("u_trade receivables", "PY") not in facts  # blank not emitted
    assert report.matched_by_statement["SOFP"] == 5
    assert _TEMPLATE_ID in report.template_ids
    assert report.scale_warning is None


def test_thousands_unit_scales_every_value(tmp_path):
    path = _mtool_file(tmp_path, [
        ("Cash and bank balances", 1595, 1420),
        ("Trade receivables", 800, 700),
        ("Property, plant and equipment", 5000, 4800),
        ("Inventories", 300, 280),
        ("Other assets", 40, 30),
    ])
    catalogue = _catalogue([
        "Cash and bank balances", "Trade receivables",
        "Property, plant and equipment", "Inventories", "Other assets",
    ])
    report = ingest_workbook(
        path, catalogue, filing_level="company", unit_scale=1000.0
    )
    cash_cy = next(
        f.value for f in report.facts
        if f.concept_uuid == "u_cash and bank balances" and f.period == "CY"
    )
    assert cash_cy == 1_595_000


def test_off_template_label_is_unmatched_not_fuzzy_guessed(tmp_path):
    """A labelled value row with no exact concept match is surfaced verbatim,
    never coerced into a nearby concept."""
    path = _mtool_file(tmp_path, [
        ("Cash and bank balances", 1595, 1420),
        ("Trade receivables", 800, 700),
        ("Property, plant and equipment", 5000, 4800),
        ("Inventories", 300, 280),
        ("Sundry receivables and prepayments XYZ", 55, 44),  # off-template
    ])
    catalogue = _catalogue([
        "Cash and bank balances", "Trade receivables",
        "Property, plant and equipment", "Inventories",
    ])
    report = ingest_workbook(path, catalogue, filing_level="company")
    labels = [r["label"] for r in report.unmatched_rows]
    assert "Sundry receivables and prepayments XYZ" in labels
    # And it did NOT become a fact.
    assert all("sundry" not in f.concept_uuid for f in report.facts)


def test_missing_sheet_is_reported(tmp_path):
    path = _mtool_file(tmp_path, [
        ("Cash and bank balances", 1595, 1420),
        ("Trade receivables", 800, 700),
        ("Property, plant and equipment", 5000, 4800),
        ("Inventories", 300, 280),
        ("Other assets", 40, 30),
    ], sheet="SOFP")
    catalogue = _catalogue([
        "Cash and bank balances", "Trade receivables",
        "Property, plant and equipment", "Inventories", "Other assets",
    ])
    catalogue["SOPL"] = {normalize_label("Revenue"): _target("Revenue")}
    report = ingest_workbook(path, catalogue, filing_level="company")
    assert "SOPL" in report.sheets_missing


def test_scale_backstop_warns_on_tiny_figures(tmp_path):
    path = _mtool_file(tmp_path, [
        ("Cash and bank balances", 5, 4),
        ("Trade receivables", 8, 7),
        ("Property, plant and equipment", 9, 8),
        ("Inventories", 3, 2),
        ("Other assets", 4, 3),
    ])
    catalogue = _catalogue([
        "Cash and bank balances", "Trade receivables",
        "Property, plant and equipment", "Inventories", "Other assets",
    ])
    report = ingest_workbook(path, catalogue, filing_level="company")
    assert report.scale_warning is not None
    assert "thousands" in report.scale_warning


# --- column detection ------------------------------------------------------

def test_low_confidence_detection_is_refused(tmp_path):
    """A sheet with too few text labels can't be auto-detected → refuse and ask
    for an explicit map (rather than silently mis-mapping)."""
    path = _mtool_file(tmp_path, [("Cash and bank balances", 1595, 1420)])
    catalogue = _catalogue(["Cash and bank balances"])
    with pytest.raises(ColumnDetectionError):
        ingest_workbook(path, catalogue, filing_level="company")


def test_explicit_column_map_bypasses_detection(tmp_path):
    path = _mtool_file(tmp_path, [("Cash and bank balances", 1595, 1420)])
    catalogue = _catalogue(["Cash and bank balances"])
    override = {
        "SOFP": {
            "label_column": "A",
            "columns": {"current_year": "B", "prior_year": "C"},
        }
    }
    report = ingest_workbook(
        path, catalogue, filing_level="company", column_map_override=override
    )
    assert report.fact_count == 2


# --- DB catalogue ----------------------------------------------------------

# --- prose capture (C2) ----------------------------------------------------

def _mtool_file_with_notes(tmp_path, footnotes):
    """footnotes: list of (fn_key, payload_text | None). Builds a workbook with
    a +FootnoteTexts sheet (col A = fn key, col C = payload)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SOFP"
    ws["A3"] = "Cash and bank balances"
    ws["B3"] = 1595
    fn = wb.create_sheet("+FootnoteTexts")
    for i, (key, payload) in enumerate(footnotes, start=1):
        fn[f"A{i}"] = key
        if payload is not None:
            fn[f"C{i}"] = payload
    path = tmp_path / "notes.xlsx"
    wb.save(path)
    return str(path)


def test_extract_prose_gold_captures_populated_footnotes(tmp_path):
    path = _mtool_file_with_notes(tmp_path, [
        ("fn_1", "<p>Revenue is recognised on delivery.</p>"),
        ("fn_2", "<p>Property is stated at cost.</p>"),
        ("fn_3", None),   # empty payload — not captured
        ("fn_4", "   "),  # whitespace-only — not captured
    ])
    notes = extract_prose_gold(path)
    keys = {n.note_key for n in notes}
    assert keys == {"fn_1", "fn_2"}
    text = {n.note_key: n.text for n in notes}
    assert "Revenue is recognised" in text["fn_1"]


def test_extract_prose_gold_empty_when_no_footnote_sheet(tmp_path):
    path = _mtool_file(tmp_path, [
        ("Cash and bank balances", 1595, 1420),
        ("Trade receivables", 800, 700),
        ("Property, plant and equipment", 5000, 4800),
        ("Inventories", 300, 280),
        ("Other assets", 40, 30),
    ])
    assert extract_prose_gold(path) == []


def test_build_catalogue_scopes_to_family_and_leaves(tmp_path):
    db = tmp_path / "cat.db"
    init_db(db)
    conn = sqlite3.connect(str(db))
    conn.execute(
        "INSERT INTO concept_templates(template_id, source_path) VALUES (?, ?)",
        (_TEMPLATE_ID, "/tmp/t.xlsx"),
    )
    conn.execute(
        "INSERT INTO concept_templates(template_id, source_path) VALUES (?, ?)",
        ("mfrs-group-sofp-cunoncu-v1", "/tmp/g.xlsx"),
    )
    # A LEAF in-family, a COMPUTED in-family (excluded), a LEAF out-of-family.
    conn.execute(
        "INSERT INTO concept_nodes(concept_uuid, template_id, kind, "
        "canonical_label, render_sheet, render_row, render_col) "
        "VALUES ('a', ?, 'LEAF', 'Cash and bank balances', 'SOFP', 5, 'B')",
        (_TEMPLATE_ID,),
    )
    conn.execute(
        "INSERT INTO concept_nodes(concept_uuid, template_id, kind, "
        "canonical_label, render_sheet, render_row, render_col) "
        "VALUES ('b', ?, 'COMPUTED', 'Total assets', 'SOFP', 6, 'B')",
        (_TEMPLATE_ID,),
    )
    conn.execute(
        "INSERT INTO concept_nodes(concept_uuid, template_id, kind, "
        "canonical_label, render_sheet, render_row, render_col) "
        "VALUES ('c', 'mfrs-group-sofp-cunoncu-v1', 'LEAF', 'Cash', 'SOFP', 5, 'B')",
    )
    conn.commit()
    cat = build_catalogue(conn, "mfrs", "company")
    assert set(cat["SOFP"]) == {normalize_label("Cash and bank balances")}
    assert cat["SOFP"][normalize_label("Cash and bank balances")].concept_uuid == "a"


def test_matrix_cells_are_deferred_and_counted_not_silently_dropped(tmp_path):
    """A MATRIX_CELL (SOCIE) concept is NOT ingested (mTool matrix reverse-map is
    deferred, gotcha #28), but it IS counted so the coverage gap is visible —
    never a silent denominator shrink (peer-review HIGH)."""
    db = tmp_path / "matrix.db"
    init_db(db)
    conn = sqlite3.connect(str(db))
    conn.execute(
        "INSERT INTO concept_templates(template_id, source_path) VALUES (?, ?)",
        (_TEMPLATE_ID, "/tmp/t.xlsx"),
    )
    # One LEAF + one MATRIX_CELL in the same family/scope.
    conn.execute(
        "INSERT INTO concept_nodes(concept_uuid, template_id, kind, "
        "canonical_label, render_sheet, render_row, render_col) "
        "VALUES ('leaf1', ?, 'LEAF', 'Cash and bank balances', 'SOFP', 5, 'B')",
        (_TEMPLATE_ID,),
    )
    conn.execute(
        "INSERT INTO concept_nodes(concept_uuid, template_id, kind, "
        "canonical_label, render_sheet, render_row, render_col, matrix_col) "
        "VALUES ('m1', ?, 'MATRIX_CELL', 'Balance at 1 January', 'SOCIE', 3, 'B', 'B')",
        (_TEMPLATE_ID,),
    )
    conn.commit()

    # build_catalogue stays LEAF-only (matrix needs column-aware matching mTool
    # doesn't yet confirm) — the matrix concept is absent from the match set.
    cat = build_catalogue(conn, "mfrs", "company", [_TEMPLATE_ID])
    assert "SOCIE" not in cat
    assert set(cat["SOFP"]) == {normalize_label("Cash and bank balances")}

    # …but it is COUNTED as deferred, so the omission is visible.
    assert count_deferred_matrix(conn, "mfrs", "company", [_TEMPLATE_ID]) == 1
