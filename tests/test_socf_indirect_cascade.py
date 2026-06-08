"""SOCF-Indirect cascade fix (2026-06-06).

A downloaded MFRS SOCF-Indirect workbook showed the disposal of
non-current assets never reaching the operating-activities total: the
hand-authored ``*Total adjustments to reconcile profit (loss)`` formula
(row 57) did not trace back to the SSM calc linkbase
(``cal_ssmt-fs-mfrs_2022-12-31_role-500100.xml``).  The diff against the
linkbase child set of ``AdjustmentsForReconcileProfitLoss`` found four
defects in B57:

1. ``-1 * GainsLossesOnDisposalsOfNoncurrentAssets`` (row 40) MISSING
   -> disposal subtotal orphaned (the reported bug).
2. ``+1 * ImpairmentLossRecognisedInProfitOrLoss`` (row 36) MISSING
   -> impairment subtotal orphaned.  Row 36 itself dropped row 27.
3. row 56 (OtherAdjustmentsToReconcileProfitLoss) counted twice.
4. row 60 (a working-capital inventory item, already in B66) wrongly
   added -> inventory double-counted.

The fix rewrites row 57 to exactly the linkbase child set and completes
row 36 to ``sum(B27:B35)``.  Because the canonical parser builds the
cascade edges from these xlsx formulas, the bug corrupted both the
downloaded total AND the canonical DB / Values total — so this guards
both surfaces.  Applies to MFRS Company + Group (MPERS is generated from
the linkbase and was already correct).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from concept_model.parser import parse_template
from prompts._sign_conventions import _parse_total_formula

REPO = Path(__file__).resolve().parent.parent

# (level, value columns) for each MFRS SOCF-Indirect template under test.
TEMPLATES = [
    ("Company", ("B", "C")),
    ("Group", ("B", "C", "D", "E")),
]

# The exact linkbase child set of row 57 (weight, row), in linkbase order.
EXPECTED_57 = {
    11: 1, 12: 1, 13: -1, 14: 1, 15: -1, 16: 1, 17: -1, 18: 1, 19: 1,
    20: 1, 21: -1, 22: 1, 23: -1, 24: -1, 25: 1, 36: 1, 40: -1, 41: -1,
    42: -1, 43: -1, 44: -1, 45: -1, 46: 1, 54: 1, 55: 1, 56: 1,
}


def _path(level: str) -> Path:
    return REPO / "XBRL-template-MFRS" / level / "07-SOCF-Indirect.xlsx"


@pytest.mark.parametrize("level,cols", TEMPLATES)
def test_total_adjustments_matches_calc_linkbase(level: str, cols) -> None:
    ws = openpyxl.load_workbook(_path(level)).active
    for col in cols:
        terms = _parse_total_formula(ws[f"{col}57"].value)
        got = {row: w for (w, c, row) in terms if c == col}
        expected = {row: w for row, w in EXPECTED_57.items()}
        assert got == expected, f"{level} {col}57 drifted from calc linkbase"


@pytest.mark.parametrize("level,cols", TEMPLATES)
def test_disposal_and_impairment_cascade(level: str, cols) -> None:
    """The two orphaned subtotals must now feed row 57, the duplicate and
    the stray inventory term must be gone."""
    ws = openpyxl.load_workbook(_path(level)).active
    for col in cols:
        rows = [row for (_w, c, row) in _parse_total_formula(ws[f"{col}57"].value) if c == col]
        assert rows.count(40) == 1, "disposal of non-current assets must cascade"
        assert rows.count(36) == 1, "impairment total must cascade"
        assert 60 not in rows, "working-capital inventory row must not be in row 57"
        assert rows.count(56) == 1, "row 56 must not be double-counted"


@pytest.mark.parametrize("level,cols", TEMPLATES)
def test_impairment_subtotal_includes_ppe(level: str, cols) -> None:
    """Row 36 (impairment subtotal) must sum the full 27:35 block."""
    ws = openpyxl.load_workbook(_path(level)).active
    for col in cols:
        rows = {row for (_w, c, row) in _parse_total_formula(ws[f"{col}36"].value) if c == col}
        assert rows == set(range(27, 36)), f"{level} {col}36 must sum rows 27..35"


def test_canonical_parser_emits_disposal_edge() -> None:
    """The cascade edges the canonical model builds from the xlsx must now
    include the disposal (-1) and impairment (+1) children — proving the
    DB/Values total is fixed, not just the downloaded workbook."""
    tree = parse_template(str(_path("Company")))
    n57 = next(n for n in tree.concepts if n.render_key["row"] == 57)
    edges = {(e["ref"]["row"], e["coefficient"]) for e in n57.edges}
    assert (40, -1.0) in edges
    assert (36, 1.0) in edges
    assert all(r != 60 for r, _ in edges)
