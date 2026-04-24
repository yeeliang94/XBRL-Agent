"""Notes workbooks must merge alongside face-statement workbooks."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from notes.payload import NotesPayload
from notes.writer import write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path
from statement_types import StatementType, template_path as face_template_path
from workbook_merger import merge


def _seed_face_workbook(tmp_path: Path) -> Path:
    # Copy the SOFP CuNonCu template verbatim as a "filled" workbook.
    src = face_template_path(StatementType.SOFP, "CuNonCu", level="company")
    dst = tmp_path / "SOFP_filled.xlsx"
    wb = openpyxl.load_workbook(src)
    wb.save(dst)
    wb.close()
    return dst


def _seed_notes_workbook(tmp_path: Path) -> Path:
    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "NOTES_CORP_INFO_filled.xlsx"
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=[NotesPayload(
            chosen_row_label="Financial reporting status",
            content="Going concern.",
            evidence="Page 14",
            source_pages=[14],
            parent_note={"number": "1", "title": "Test Note"},
        )],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )
    assert result.success
    return out


def test_merge_includes_notes_workbook(tmp_path: Path):
    face = _seed_face_workbook(tmp_path)
    notes = _seed_notes_workbook(tmp_path)

    merged_path = tmp_path / "filled.xlsx"
    result = merge(
        workbook_paths={StatementType.SOFP: str(face)},
        output_path=str(merged_path),
        notes_workbook_paths={NotesTemplateType.CORP_INFO: str(notes)},
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(merged_path)
    # Face sheets from SOFP template
    assert "SOFP-CuNonCu" in wb.sheetnames
    # Notes sheet from notes template
    assert "Notes-CI" in wb.sheetnames
    # Face sheets must come first, notes after (Plan A.3 wiring)
    face_idx = wb.sheetnames.index("SOFP-CuNonCu")
    notes_idx = wb.sheetnames.index("Notes-CI")
    assert face_idx < notes_idx
    wb.close()


def test_merge_without_notes_argument_is_backward_compatible(tmp_path: Path):
    face = _seed_face_workbook(tmp_path)
    merged_path = tmp_path / "filled_face_only.xlsx"
    # Call with old positional signature — no notes argument.
    result = merge(
        workbook_paths={StatementType.SOFP: str(face)},
        output_path=str(merged_path),
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(merged_path)
    assert "SOFP-CuNonCu" in wb.sheetnames
    wb.close()
