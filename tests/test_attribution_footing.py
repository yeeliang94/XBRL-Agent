"""Pins for the SOPL / SOCI attribution-footing cross-checks.

These close the same class as the SOCF articulation gap (run-50) for the two
OTHER statements where two subtotals are computed from INDEPENDENT leaf sets:

* SOPL: ``*Profit (loss)`` (down the income statement) vs ``*Total profit
  (loss)`` (from the owners/NCI attribution leaves).
* SOCI: ``*Total comprehensive income`` (= profit + OCI) vs its attribution
  twin (= owners + NCI).

Key behaviours pinned: the check FAILS when the attribution split doesn't sum
to the income-side total, PASSES when it does, and is ``not_applicable`` (never
a false fail) when the attribution isn't disclosed at all — the common
company-filing case, where the cascade writes no fact for the all-blank
attribution total.
"""
from __future__ import annotations

import json
import sqlite3
from collections import deque
from pathlib import Path

import pytest

from concept_model.cascade import recompute_after_turn
from concept_model.importer import import_company_targets, import_template
from concept_model.parser import parse_template
from cross_checks.attribution_footing import (
    SOCIAttributionFootingCheck, SOPLAttributionFootingCheck,
)
from cross_checks.framework import FactsContext, build_default_cross_checks
from db.schema import init_db
from statement_types import StatementType

REPO = Path(__file__).resolve().parent.parent

CASES = {
    "sopl": dict(
        fixture=REPO / "XBRL-template-MFRS" / "Company" / "03-SOPL-Function.xlsx",
        stmt=StatementType.SOPL,
        check=SOPLAttributionFootingCheck,
        income_label="profit (loss)",
        owners_label="profit (loss), attributable to owners of parent",
    ),
    "soci": dict(
        fixture=REPO / "XBRL-template-MFRS" / "Company" / "05-SOCI-BeforeTax.xlsx",
        stmt=StatementType.SOCI,
        check=SOCIAttributionFootingCheck,
        income_label="total comprehensive income",
        owners_label="comprehensive income, attributable to owners of parent",
    ),
}


def _computed_exact(conn, tid, label):
    target = label.strip().lstrip("*").strip().lower()
    for cu, clabel in conn.execute(
        "SELECT concept_uuid, canonical_label FROM concept_nodes "
        "WHERE template_id = ? AND kind = 'COMPUTED' ORDER BY render_row", (tid,),
    ):
        if str(clabel).strip().lstrip("*").strip().lower() == target:
            return cu
    return None


def _leaf_like(conn, tid, substr):
    r = conn.execute(
        "SELECT concept_uuid FROM concept_nodes WHERE template_id = ? "
        "AND kind = 'LEAF' AND lower(canonical_label) LIKE ? ORDER BY render_row",
        (tid, f"%{substr.lower()}%"),
    ).fetchone()
    return r[0] if r else None


def _descendant_leaf(conn, uuid):
    seen, q = set(), deque([uuid])
    while q:
        n = q.popleft()
        if n in seen:
            continue
        seen.add(n)
        k = conn.execute("SELECT kind FROM concept_nodes WHERE concept_uuid = ?", (n,)).fetchone()
        if k and k[0] == "LEAF":
            return n
        for (c,) in conn.execute("SELECT child_uuid FROM concept_edges WHERE parent_uuid = ?", (n,)):
            q.append(c)
    return None


def _seed(conn, run_id, uuid, value):
    conn.execute(
        "INSERT OR REPLACE INTO run_concept_facts(run_id, concept_uuid, period, "
        "entity_scope, value, value_status, source, updated_at) "
        "VALUES (?, ?, 'CY', 'Company', ?, 'observed', 'pdf', '2026-06-21Z')",
        (run_id, uuid, value),
    )


def _val(conn, run_id, uuid):
    r = conn.execute(
        "SELECT value FROM run_concept_facts WHERE run_id = ? AND concept_uuid = ? "
        "AND period = 'CY' AND entity_scope = 'Company'", (run_id, uuid),
    ).fetchone()
    return r[0] if r else None


def _setup(tmp_path, case):
    db = tmp_path / "xbrl.db"
    init_db(db)
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    tree = parse_template(str(case["fixture"]))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db, jp)
    import_company_targets(db, tid)
    cur = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
        "VALUES ('2026-06-21T00:00:00Z', 'x.pdf', 'running', '2026-06-21T00:00:00Z')")
    return db, int(cur.lastrowid), tid, conn


@pytest.mark.parametrize("case_key", ["sopl", "soci"])
@pytest.mark.parametrize("articulates", [True, False])
def test_attribution_footing_catches_mismatch(tmp_path, case_key, articulates):
    case = CASES[case_key]
    db, run_id, tid, conn = _setup(tmp_path, case)

    income_uuid = _computed_exact(conn, tid, case["income_label"])
    owners = _leaf_like(conn, tid, case["owners_label"])
    assert income_uuid and owners, "concepts not resolved"
    income_leaf = _descendant_leaf(conn, income_uuid)
    assert income_leaf

    # Seed one income leaf, cascade to compute the income-side total, then set
    # the owners attribution leaf to match it (articulates) or miss by a clear
    # gap (does not).
    _seed(conn, run_id, income_leaf, 1000.0)
    conn.commit()
    recompute_after_turn(str(db), run_id)
    p = _val(conn, run_id, income_uuid)
    assert p is not None

    _seed(conn, run_id, owners, p if articulates else p + 50.0)
    conn.commit()
    recompute_after_turn(str(db), run_id)

    ctx = FactsContext(
        conn=conn, run_id=run_id, template_ids={case["stmt"]: tid},
        filing_level="company", filing_standard="mfrs")
    result = case["check"]().run_facts(ctx, tolerance=1.0)
    conn.close()
    assert result.status == ("passed" if articulates else "failed"), result.message


@pytest.mark.parametrize("case_key", ["sopl", "soci"])
def test_attribution_footing_not_applicable_when_undisclosed(tmp_path, case_key):
    """No attribution leaves filled → not_applicable, NOT a false fail."""
    case = CASES[case_key]
    db, run_id, tid, conn = _setup(tmp_path, case)
    income_uuid = _computed_exact(conn, tid, case["income_label"])
    income_leaf = _descendant_leaf(conn, income_uuid)
    _seed(conn, run_id, income_leaf, 1000.0)
    conn.commit()
    recompute_after_turn(str(db), run_id)

    ctx = FactsContext(
        conn=conn, run_id=run_id, template_ids={case["stmt"]: tid},
        filing_level="company", filing_standard="mfrs")
    result = case["check"]().run_facts(ctx, tolerance=1.0)
    conn.close()
    assert result.status == "not_applicable", result.message


def test_attribution_footing_checks_registered():
    names = {c.name for c in build_default_cross_checks()}
    assert "sopl_attribution_footing" in names
    assert "soci_attribution_footing" in names


def test_group_company_scope_is_checked(tmp_path):
    """Peer-review (Codex P2): on a group filing the check must reconcile the
    COMPANY column too, not just Group. Group attribution foots but Company
    attribution is wrong → overall FAILED (was silently passing)."""
    from concept_model.importer import import_group_targets
    case = CASES["sopl"]
    db = tmp_path / "xbrl.db"
    init_db(db)
    conn = sqlite3.connect(str(db))
    conn.execute("PRAGMA foreign_keys = ON")
    tree = parse_template(str(case["fixture"]))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db, jp)
    import_group_targets(db, tid)
    cur = conn.execute(
        "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
        "VALUES ('2026-06-21T00:00:00Z', 'x.pdf', 'running', '2026-06-21T00:00:00Z')")
    run_id = int(cur.lastrowid)
    conn.commit()

    income_uuid = _computed_exact(conn, tid, case["income_label"])
    owners = _leaf_like(conn, tid, case["owners_label"])
    income_leaf = _descendant_leaf(conn, income_uuid)

    def seed_scope(uuid, value, scope):
        conn.execute(
            "INSERT OR REPLACE INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, source, updated_at) "
            "VALUES (?, ?, 'CY', ?, ?, 'observed', 'pdf', 't')",
            (run_id, uuid, scope, value))

    # Group: income leaf + owners attribution that FOOTS.
    seed_scope(income_leaf, 1000.0, "Group")
    # Company: income leaf + owners attribution that is WRONG (off by 50).
    seed_scope(income_leaf, 800.0, "Company")
    conn.commit()
    recompute_after_turn(str(db), run_id)
    g_profit = conn.execute(
        "SELECT value FROM run_concept_facts WHERE run_id=? AND concept_uuid=? "
        "AND period='CY' AND entity_scope='Group'", (run_id, income_uuid)).fetchone()[0]
    c_profit = conn.execute(
        "SELECT value FROM run_concept_facts WHERE run_id=? AND concept_uuid=? "
        "AND period='CY' AND entity_scope='Company'", (run_id, income_uuid)).fetchone()[0]
    seed_scope(owners, g_profit, "Group")          # Group foots
    seed_scope(owners, c_profit + 50.0, "Company")  # Company is wrong
    conn.commit()
    recompute_after_turn(str(db), run_id)

    ctx = FactsContext(
        conn=conn, run_id=run_id, template_ids={case["stmt"]: tid},
        filing_level="group", filing_standard="mfrs")
    result = SOPLAttributionFootingCheck().run_facts(ctx, tolerance=1.0)
    conn.close()
    assert result.status == "failed", result.message
    assert "company" in result.message.lower()


def _build_sopl_xlsx(path, *, owners, nci, income=1000.0, owners_blank=False):
    """Minimal SOPL-Function sheet: income total + attribution total formula +
    attribution leaf cells. ``owners_blank`` leaves the leaf cells empty."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Function"
    ws["A1"] = "*Profit (loss)"
    ws["B1"] = income
    ws["A2"] = "Profit (loss), attributable to owners of parent"
    ws["A3"] = "Profit (loss), attributable to non-controlling interests"
    if not owners_blank:
        ws["B2"] = owners
        ws["B3"] = nci
    ws["A4"] = "*Total profit (loss)"
    ws["B4"] = "=B2+B3"
    wb.save(str(path))


def test_xlsx_explicit_zero_attribution_is_validated(tmp_path):
    """Peer-review (Codex P2): an EXPLICIT 0 attribution while income is
    non-zero is a disclosed mismatch — the xlsx path must FAIL it, not skip it
    as not_applicable."""
    path = tmp_path / "sopl_zero.xlsx"
    _build_sopl_xlsx(path, owners=0.0, nci=0.0, income=1000.0)
    result = SOPLAttributionFootingCheck().run(
        {StatementType.SOPL: str(path)}, tolerance=1.0, filing_level="company")
    assert result.status == "failed", result.message


def test_xlsx_blank_attribution_is_not_applicable(tmp_path):
    """The flip side: genuinely BLANK attribution leaves stay not_applicable
    (no false-fail on a company filing that doesn't split the attribution)."""
    path = tmp_path / "sopl_blank.xlsx"
    _build_sopl_xlsx(path, owners=0.0, nci=0.0, income=1000.0, owners_blank=True)
    result = SOPLAttributionFootingCheck().run(
        {StatementType.SOPL: str(path)}, tolerance=1.0, filing_level="company")
    assert result.status == "not_applicable", result.message
