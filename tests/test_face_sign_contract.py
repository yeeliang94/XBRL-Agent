"""End-to-end inventory pins for face-statement formula-ready signs."""
from __future__ import annotations

import subprocess
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from prompts._sign_conventions import (
    _expected_cash_input,
    _formula_occurrences,
    _parse_total_formula,
    face_sign_convention_block,
    socf_sign_convention_block,
)
from statement_types import StatementType


REPO = Path(__file__).resolve().parent.parent


def _path_products(
    ws,
    child_predicate,
    ancestor_predicate,
) -> list[int]:
    """Return products of B-column coefficients from child to ancestor."""
    labels: dict[int, str] = {}
    parents: dict[int, list[tuple[int, int]]] = defaultdict(list)
    for row, values in enumerate(
        ws.iter_rows(min_col=1, max_col=2, values_only=True),
        start=1,
    ):
        label, formula = values
        if label:
            labels[row] = str(label).strip()
        if isinstance(formula, str) and formula.startswith("="):
            for sign, column, child_row in _parse_total_formula(formula):
                if column == "B":
                    parents[child_row].append((row, sign))

    products: list[int] = []
    for child_row, label in labels.items():
        if not child_predicate(label.casefold()):
            continue
        stack = [(child_row, 1, frozenset({child_row}))]
        while stack:
            row, product, visited = stack.pop()
            for parent_row, coefficient in parents.get(row, []):
                if parent_row in visited:
                    continue
                next_product = product * coefficient
                parent_label = labels.get(parent_row, "").casefold()
                if ancestor_predicate(parent_label):
                    products.append(next_product)
                    continue
                stack.append((parent_row, next_product, visited | {parent_row}))
    return products


def _statement_for_template(path: Path) -> StatementType:
    slot = int(path.name[:2])
    if slot <= 2:
        return StatementType.SOFP
    if slot <= 4:
        return StatementType.SOPL
    if slot <= 6:
        return StatementType.SOCI
    if slot <= 8:
        return StatementType.SOCF
    return StatementType.SOCIE


def _face_templates() -> list[Path]:
    paths: list[Path] = []
    for standard in ("MFRS", "MPERS"):
        max_slot = 9 if standard == "MFRS" else 10
        for level in ("Company", "Group"):
            folder = REPO / f"XBRL-template-{standard}" / level
            paths.extend(
                path
                for path in sorted(folder.glob("*.xlsx"))
                if path.name[:2].isdigit() and int(path.name[:2]) <= max_slot
            )
    return paths


@pytest.mark.parametrize("template", _face_templates(), ids=lambda p: str(p.relative_to(REPO)))
def test_every_live_face_template_has_authoritative_sign_guidance(template: Path) -> None:
    statement = _statement_for_template(template)
    block = face_sign_convention_block(template, statement)
    assert block is not None, template
    assert "SIGN CONVENTIONS" in block
    assert "mTool" in block or statement in (StatementType.SOCF, StatementType.SOCIE)


@pytest.mark.parametrize("template", _face_templates(), ids=lambda p: str(p.relative_to(REPO)))
def test_fixed_sign_equity_rows_are_subtracted_in_every_live_formula(
    template: Path,
) -> None:
    """Pin layouts skipped by generic auditors, including MPERS Group SOCIE."""
    # This test mixes coordinate lookups with whole-sheet iteration. In
    # openpyxl's read-only mode each coordinate lookup reparses worksheet XML;
    # normal mode materialises the small template once and makes both access
    # patterns constant-time.
    wb = openpyxl.load_workbook(template, data_only=False)
    try:
        for ws in wb.worksheets:
            sheet_key = ws.title.casefold()
            target_rows = {
                row
                for row in range(1, ws.max_row + 1)
                if (
                    str(ws.cell(row, 1).value or "").strip().casefold()
                    == "treasury shares"
                    and "sofp" in sheet_key
                )
                or (
                    str(ws.cell(row, 1).value or "").strip().casefold()
                    == "dividends paid"
                    and ("socie" in sheet_key or "sore" in sheet_key)
                )
            }
            for target_row in target_rows:
                signs: list[int] = []
                for formula_row in ws.iter_rows():
                    for cell in formula_row:
                        formula = cell.value
                        if not isinstance(formula, str) or not formula.startswith("="):
                            continue
                        formula_column = get_column_letter(cell.column)
                        signs.extend(
                            sign
                            for sign, referenced_column, referenced_row in _parse_total_formula(
                                formula
                            )
                            if referenced_row == target_row
                            and referenced_column == formula_column
                        )
                assert signs, f"{template}:{ws.title} row {target_row} has no formula use"
                assert set(signs) == {-1}, (
                    f"{template}:{ws.title} row {target_row} uses coefficients {signs}"
                )
    finally:
        wb.close()


@pytest.mark.parametrize(
    "standard,level",
    [(standard, level) for standard in ("MFRS", "MPERS") for level in ("Company", "Group")],
)
def test_inventory_direction_guidance_is_pinned_to_profit_formula(
    standard: str,
    level: str,
) -> None:
    path = REPO / f"XBRL-template-{standard}" / level / "04-SOPL-Nature.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb.worksheets[0]
        products = _path_products(
            ws,
            lambda label: "inventories of finished goods" in label,
            lambda label: "profit (loss) before tax" in label,
        )
    finally:
        wb.close()
    assert products and set(products) == {-1}


@pytest.mark.parametrize(
    "level,filename",
    [
        (level, filename)
        for level in ("Company", "Group")
        for filename in ("03-SOPL-Function.xlsx", "04-SOPL-Nature.xlsx")
    ],
)
def test_mpers_impairment_reversal_is_pinned_to_other_income_formula(
    level: str,
    filename: str,
) -> None:
    path = REPO / "XBRL-template-MPERS" / level / filename
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = next(sheet for sheet in wb.worksheets if "Analysis" in sheet.title)
        products = _path_products(
            ws,
            lambda label: label == "(reversal of)/impairment loss on inventories",
            lambda label: "total other income" in label,
        )
    finally:
        wb.close()
    assert products and set(products) == {1}


@pytest.mark.parametrize("level", ["Company", "Group"])
def test_sopl_tax_direction_guidance_is_pinned_to_profit_formula(level: str) -> None:
    path = REPO / "XBRL-template-MFRS" / level / "04-SOPL-Nature.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb.worksheets[0]
        products = _path_products(
            ws,
            lambda label: "total tax expense (income)" in label,
            lambda label: "profit (loss) from continuing operations" in label,
        )
    finally:
        wb.close()
    assert products and set(products) == {-1}


@pytest.mark.parametrize(
    "standard,level",
    [(standard, level) for standard in ("MFRS", "MPERS") for level in ("Company", "Group")],
)
def test_soci_tax_and_reclassification_guidance_is_pinned_to_oci_total(
    standard: str,
    level: str,
) -> None:
    path = REPO / f"XBRL-template-{standard}" / level / "05-SOCI-BeforeTax.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb.worksheets[0]
        tax_products = _path_products(
            ws,
            lambda label: "income tax relating" in label,
            lambda label: "total other comprehensive income" in label,
        )
        reclassification_products = _path_products(
            ws,
            lambda label: label.startswith("reclassification adjustments"),
            lambda label: "total other comprehensive income" in label,
        )
    finally:
        wb.close()
    assert tax_products and set(tax_products) == {-1}
    if standard == "MFRS":
        assert reclassification_products and set(reclassification_products) == {-1}
    else:
        assert reclassification_products == []


@pytest.mark.parametrize(
    "standard,level",
    [(standard, level) for standard in ("MFRS", "MPERS") for level in ("Company", "Group")],
)
def test_socie_treasury_guidance_is_pinned_to_equity_movement_formula(
    standard: str,
    level: str,
) -> None:
    path = REPO / f"XBRL-template-{standard}" / level / "09-SOCIE.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb.worksheets[0]
        products = _path_products(
            ws,
            lambda label: label == "treasury shares transactions",
            lambda label: "total increase (decrease) in equity" in label,
        )
    finally:
        wb.close()
    assert products and set(products) == {1}


@pytest.mark.parametrize(
    "standard,level,filename",
    [
        (standard, level, filename)
        for standard in ("MFRS", "MPERS")
        for level in ("Company", "Group")
        for filename in ("01-SOFP-CuNonCu.xlsx", "02-SOFP-OrderOfLiquidity.xlsx")
    ],
)
def test_retained_earnings_guidance_is_pinned_to_total_equity_formula(
    standard: str,
    level: str,
    filename: str,
) -> None:
    path = REPO / f"XBRL-template-{standard}" / level / filename
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb.worksheets[0]
        products = _path_products(
            ws,
            lambda label: "retained earnings" in label,
            lambda label: "total equity" in label,
        )
    finally:
        wb.close()
    assert products and set(products) == {1}


@pytest.mark.parametrize(
    "standard,level,filename",
    [
        (standard, level, filename)
        for standard in ("MFRS", "MPERS")
        for level in ("Company", "Group")
        for filename in ("07-SOCF-Indirect.xlsx", "08-SOCF-Direct.xlsx")
    ],
)
def test_every_unambiguous_socf_cash_row_states_the_final_input_sign(
    standard: str,
    level: str,
    filename: str,
) -> None:
    path = REPO / f"XBRL-template-{standard}" / level / filename
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb.active
        occurrences, labels = _formula_occurrences(ws)
    finally:
        wb.close()

    block = socf_sign_convention_block(path)
    assert block is not None
    row_lines = {
        int(line.split()[2]): line
        for line in block.splitlines()
        if line.startswith("- Row ")
    }

    for row, row_occurrences in occurrences.items():
        coefficients = {sign for sign, _parent in row_occurrences}
        if len(coefficients) > 1:
            # The taxonomy contains one deliberate mixed-use concept, repeated
            # in Company and Group. It must never fall back to first-wins.
            assert standard == "MFRS" and filename.startswith("07-") and row == 24
            assert "DUAL-USE" in row_lines[row]
            assert "Enter NEGATIVE" in row_lines[row]
            continue
        expected_positive, _reason = _expected_cash_input(
            labels[row],
            next(iter(coefficients)),
        )
        if expected_positive is not None:
            assert "Enter a " in row_lines[row], row_lines[row]
            expected_word = "POSITIVE magnitude" if expected_positive else "NEGATIVE value"
            assert expected_word in row_lines[row], row_lines[row]


@pytest.mark.parametrize(
    "module,result_line",
    [
        (
            "scripts.audit_mfrs_formulas",
            "RESULT: ALL MFRS FORMULAS MATCH THE LINKBASE SIGN CONTRACT",
        ),
        ("scripts.audit_mpers_formulas", "RESULT: ALL FORMULAS CORRECT"),
    ],
)
def test_formula_audit_commands_are_clean(module: str, result_line: str) -> None:
    completed = subprocess.run(
        [sys.executable, "-m", module],
        cwd=REPO,
        check=False,
        capture_output=True,
        text=True,
        timeout=60,
    )
    assert completed.returncode == 0, completed.stdout + completed.stderr
    assert result_line in completed.stdout
