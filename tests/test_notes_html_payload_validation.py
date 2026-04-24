"""Step 5 — HTML payload sanitisation at the writer + agent-tool boundary.

Belt-and-braces guard: even though the prompt now requests HTML, the
model can regress (plaintext Markdown, dangerous tags, inline event
handlers). The sanitiser strips disallowed tags / attributes, wraps
bare prose in `<p>`, and surfaces warnings so the writer doesn't
silently swallow a dropped `<script>` tag.
"""
from __future__ import annotations

import pytest

from notes.html_sanitize import ALLOWED_TAGS, sanitize_notes_html


def test_payload_with_script_tag_strips_it() -> None:
    cleaned, warnings = sanitize_notes_html("<p>x</p><script>alert(1)</script>")
    assert "<script" not in cleaned.lower()
    assert "alert" not in cleaned.lower()
    assert "x" in cleaned
    # Sanitiser surfaces what it removed.
    assert any("script" in w.lower() for w in warnings)


def test_payload_with_inline_event_handler_strips_it() -> None:
    cleaned, warnings = sanitize_notes_html('<a onclick="evil()">x</a>')
    # Onclick removed.
    assert "onclick" not in cleaned.lower()
    assert "evil" not in cleaned.lower()
    # There's a warning somewhere that references the removed handler.
    assert any("onclick" in w.lower() or "event" in w.lower() for w in warnings)


def test_payload_with_no_tags_is_wrapped_in_paragraph() -> None:
    cleaned, _ = sanitize_notes_html("Hello world")
    assert cleaned == "<p>Hello world</p>"


def test_payload_with_disallowed_tag_is_logged_as_warning() -> None:
    cleaned, warnings = sanitize_notes_html("<p>x</p><iframe src='e'></iframe>")
    assert "<iframe" not in cleaned.lower()
    assert any("iframe" in w.lower() for w in warnings)


def test_sanitizer_preserves_allowed_tags_verbatim() -> None:
    # Exercise the whole whitelist end-to-end. bs4's html.parser
    # serialises void tags like <br> as `<br/>`; accept either form
    # since both are valid HTML5 and round-trip through the editor.
    for tag in ALLOWED_TAGS:
        cleaned, warnings = sanitize_notes_html(f"<{tag}>x</{tag}>")
        lower = cleaned.lower()
        assert (f"<{tag}>" in lower) or (f"<{tag}/>" in lower) or \
               (f"<{tag} " in lower), (
            f"{tag} did not survive sanitiser: {cleaned!r}"
        )
        # Allowed tags produce no warnings.
        assert warnings == [], f"{tag} triggered warnings: {warnings}"


def test_sanitizer_strips_style_attributes() -> None:
    # Gotcha #7 — frontend uses inline styles, but agent output must not
    # carry them into the clipboard round-trip.
    cleaned, warnings = sanitize_notes_html(
        '<p style="color:red">x</p>'
    )
    assert "style" not in cleaned.lower()
    assert "color:red" not in cleaned.lower()


def test_sanitizer_empty_input_returns_empty() -> None:
    cleaned, warnings = sanitize_notes_html("")
    assert cleaned == ""
    assert warnings == []


def test_sanitizer_none_input_returns_empty() -> None:
    cleaned, warnings = sanitize_notes_html(None)  # type: ignore[arg-type]
    assert cleaned == ""
    assert warnings == []


def test_wrap_does_not_double_wrap_html_content() -> None:
    # A payload already containing a block-level tag stays as-is.
    cleaned, _ = sanitize_notes_html("<p>already wrapped</p>")
    assert cleaned == "<p>already wrapped</p>"


def test_writer_surfaces_sanitiser_warnings(tmp_path) -> None:
    """End-to-end: a dirty HTML payload produces a row + warnings on
    `NotesWriteResult`. The writer itself sanitises before truncation
    so the cell content is clean."""
    from notes.payload import NotesPayload
    from notes.writer import write_notes_workbook
    from notes_types import NotesTemplateType, notes_template_path

    tpl = notes_template_path(NotesTemplateType.CORP_INFO, level="company")
    out = tmp_path / "Notes-CI_dirty.xlsx"

    payload = NotesPayload(
        chosen_row_label="Financial reporting status",
        content="<p>Active</p><script>alert(1)</script>",
        evidence="Page 12, Note 2",
        source_pages=[12],
        parent_note={"number": "1", "title": "Test Note"},
    )

    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=[payload],
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-CI",
    )
    assert result.success
    assert any("script" in (w or "").lower() for w in result.sanitizer_warnings), (
        f"expected sanitiser warnings, got {result.sanitizer_warnings}"
    )

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb["Notes-CI"]
    # Find the row and check the cell doesn't contain a script fragment.
    cell_value = None
    for row in range(1, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if label and "financial reporting status" in str(label).lower():
            cell_value = ws.cell(row=row, column=2).value
            break
    assert cell_value is not None
    assert "<script" not in (cell_value or "").lower()
    wb.close()
