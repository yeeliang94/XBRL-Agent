"""Tests for workbook merger (Phase 6, Step 6.1).

Verifies that per-statement workbooks can be merged into a single file
with all sheets, values, and formulas preserved.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from statement_types import StatementType


def _make_workbook(sheets: dict[str, list[list]], path: str):
    """Create a minimal workbook from {sheet_name: [[row_data], ...]}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
    wb.close()


@pytest.fixture
def tmp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield d


class TestMergedWorkbookHasAllSheets:
    """Given 5 per-statement workbook fixtures, merged file contains all expected sheets."""

    def test_all_sheets_present(self, tmp_dir):
        """Merged workbook should contain every sheet from every input workbook."""
        from workbook_merger import merge

        # Create 5 per-statement workbooks with unique sheets
        paths = {}
        wb_specs = {
            StatementType.SOFP: {"SOFP-CuNonCu": [["Assets", 100]], "SOFP-Sub-CuNonCu": [["PPE", 50]]},
            StatementType.SOPL: {"SOPL-Function": [["Revenue", 200]], "SOPL-Analysis-Function": [["Sales", 150]]},
            StatementType.SOCI: {"SOCI-BeforeOfTax": [["OCI", 30]]},
            StatementType.SOCF: {"SOCF-Indirect": [["Operating", 80]]},
            StatementType.SOCIE: {"SOCIE": [["Equity", 500]]},
        }

        for stmt, sheets in wb_specs.items():
            p = str(Path(tmp_dir) / f"{stmt.value}_filled.xlsx")
            _make_workbook(sheets, p)
            paths[stmt] = p

        output = str(Path(tmp_dir) / "filled.xlsx")
        result = merge(paths, output)

        assert result.success
        assert Path(output).exists()

        wb = openpyxl.load_workbook(output)
        sheet_names = set(wb.sheetnames)
        expected = {"SOFP-CuNonCu", "SOFP-Sub-CuNonCu", "SOPL-Function",
                    "SOPL-Analysis-Function", "SOCI-BeforeOfTax",
                    "SOCF-Indirect", "SOCIE"}
        assert expected == sheet_names
        wb.close()

    def test_values_intact(self, tmp_dir):
        """Cell values in merged workbook should match source workbooks."""
        from workbook_merger import merge

        paths = {}
        p = str(Path(tmp_dir) / "SOFP_filled.xlsx")
        _make_workbook({"SOFP-CuNonCu": [["Assets", 100, 200]]}, p)
        paths[StatementType.SOFP] = p

        p2 = str(Path(tmp_dir) / "SOPL_filled.xlsx")
        _make_workbook({"SOPL-Function": [["Revenue", 999]]}, p2)
        paths[StatementType.SOPL] = p2

        output = str(Path(tmp_dir) / "filled.xlsx")
        merge(paths, output)

        wb = openpyxl.load_workbook(output, data_only=False)
        assert wb["SOFP-CuNonCu"]["A1"].value == "Assets"
        assert wb["SOFP-CuNonCu"]["B1"].value == 100
        assert wb["SOFP-CuNonCu"]["C1"].value == 200
        assert wb["SOPL-Function"]["B1"].value == 999
        wb.close()

    def test_single_statement_merge(self, tmp_dir):
        """Merger works with just one statement (e.g. SOFP-only run)."""
        from workbook_merger import merge

        p = str(Path(tmp_dir) / "SOFP_filled.xlsx")
        _make_workbook({"SOFP-CuNonCu": [["Assets", 100]]}, p)

        output = str(Path(tmp_dir) / "filled.xlsx")
        result = merge({StatementType.SOFP: p}, output)

        assert result.success
        wb = openpyxl.load_workbook(output)
        assert wb.sheetnames == ["SOFP-CuNonCu"]
        wb.close()

    def test_empty_paths_returns_failure(self, tmp_dir):
        """Merger with no workbook paths should return failure."""
        from workbook_merger import merge

        output = str(Path(tmp_dir) / "filled.xlsx")
        result = merge({}, output)
        assert not result.success

    def test_missing_file_returns_failure(self, tmp_dir):
        """Merger should handle missing workbook files gracefully."""
        from workbook_merger import merge

        paths = {StatementType.SOFP: str(Path(tmp_dir) / "nonexistent.xlsx")}
        output = str(Path(tmp_dir) / "filled.xlsx")
        result = merge(paths, output)
        assert not result.success
        assert len(result.errors) > 0


class TestMergerPreservesFormulas:
    """Formulas in per-statement sheets remain formulas in merged output."""

    def test_formulas_preserved(self, tmp_dir):
        """Formulas should NOT be evaluated — they stay as formula strings."""
        from workbook_merger import merge

        # Create a workbook with formulas
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("SOFP-CuNonCu")
        ws["A1"] = "Assets"
        ws["B1"] = 100
        ws["B2"] = 200
        ws["B3"] = "=SUM(B1:B2)"  # formula — must survive merge
        p = str(Path(tmp_dir) / "SOFP_filled.xlsx")
        wb.save(p)
        wb.close()

        output = str(Path(tmp_dir) / "filled.xlsx")
        merge({StatementType.SOFP: p}, output)

        wb2 = openpyxl.load_workbook(output, data_only=False)
        cell = wb2["SOFP-CuNonCu"]["B3"]
        assert cell.value == "=SUM(B1:B2)", f"Formula was lost: got {cell.value}"
        wb2.close()

    def test_cross_sheet_formulas_preserved(self, tmp_dir):
        """Cross-sheet references within the same statement workbook survive merge."""
        from workbook_merger import merge

        # SOFP main sheet references sub-sheet
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws_main = wb.create_sheet("SOFP-CuNonCu")
        ws_sub = wb.create_sheet("SOFP-Sub-CuNonCu")
        ws_sub["B39"] = 12345
        ws_main["B8"] = "='SOFP-Sub-CuNonCu'!B39"
        p = str(Path(tmp_dir) / "SOFP_filled.xlsx")
        wb.save(p)
        wb.close()

        output = str(Path(tmp_dir) / "filled.xlsx")
        merge({StatementType.SOFP: p}, output)

        wb2 = openpyxl.load_workbook(output, data_only=False)
        assert wb2["SOFP-CuNonCu"]["B8"].value == "='SOFP-Sub-CuNonCu'!B39"
        assert wb2["SOFP-Sub-CuNonCu"]["B39"].value == 12345
        wb2.close()

    def test_styles_preserved(self, tmp_dir):
        """Basic cell formatting (number format, font bold) should survive merge."""
        from workbook_merger import merge
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("SOPL-Function")
        ws["A1"] = "Revenue"
        ws["A1"].font = Font(bold=True)
        ws["B1"] = 1234.56
        ws["B1"].number_format = '#,##0.00'
        p = str(Path(tmp_dir) / "SOPL_filled.xlsx")
        wb.save(p)
        wb.close()

        output = str(Path(tmp_dir) / "filled.xlsx")
        merge({StatementType.SOPL: p}, output)

        wb2 = openpyxl.load_workbook(output)
        assert wb2["SOPL-Function"]["A1"].font.bold is True
        assert wb2["SOPL-Function"]["B1"].number_format == '#,##0.00'
        wb2.close()

    def test_merged_workbook_requests_full_recalculation(self, tmp_dir):
        """Merged downloads should tell Excel to recalculate formula cells on open."""
        from workbook_merger import merge

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("SOFP-CuNonCu")
        ws["B1"] = 100
        ws["B2"] = 200
        ws["B3"] = "=1*B1+1*B2"
        p = str(Path(tmp_dir) / "SOFP_filled.xlsx")
        wb.save(p)
        wb.close()

        output = str(Path(tmp_dir) / "filled.xlsx")
        merge({StatementType.SOFP: p}, output)

        wb2 = openpyxl.load_workbook(output, data_only=False)
        calc = wb2.calculation
        assert calc.fullCalcOnLoad is True
        assert calc.forceFullCalc is True
        assert calc.calcOnSave is True
        wb2.close()
