"""Notes 13 (Issued Capital) — structured numeric round-trip test."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from notes.agent import render_notes_prompt
from notes.payload import NotesPayload
from notes.writer import write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path


SHEET = "Notes-Issuedcapital"


def test_issued_capital_prompt_mentions_movement_table():
    prompt = render_notes_prompt(
        template_type=NotesTemplateType.ISSUED_CAPITAL,
        filing_level="company",
        inventory=[],
    )
    assert "share" in prompt.lower()
    # The prompt must explicitly discuss numeric payloads (this is a
    # numeric template, not a prose one).
    assert "numeric_values" in prompt
    assert SHEET in prompt


def _find_row(ws, needle: str) -> int:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and needle.lower() in str(v).lower().lstrip("*"):
            return r
    raise AssertionError(f"{needle} not found")


def test_issued_capital_company_write_fills_cy_and_py(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.ISSUED_CAPITAL, level="company")
    out = tmp_path / "NOTES_ISSUED_CAPITAL_filled.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Number of shares outstanding at beginning of period",
            content="",
            evidence="Page 45, Note 14",
            source_pages=[45],
            numeric_values={"company_cy": 1_000_000, "company_py": 900_000},
            parent_note={"number": "1", "title": "Test Note"},
        ),
        NotesPayload(
            chosen_row_label="Amount of shares outstanding at beginning of period",
            content="",
            evidence="Page 45, Note 14",
            source_pages=[45],
            numeric_values={"company_cy": 5_000_000, "company_py": 4_500_000},
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=SHEET,
    )
    assert result.success, result.errors
    assert result.rows_written == 2

    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]
    num_row = _find_row(ws, "Number of shares outstanding at beginning")
    amt_row = _find_row(ws, "Amount of shares outstanding at beginning")
    assert ws.cell(row=num_row, column=2).value == 1_000_000
    assert ws.cell(row=num_row, column=3).value == 900_000
    assert ws.cell(row=num_row, column=4).value == "Page 45, Note 14"
    assert ws.cell(row=amt_row, column=2).value == 5_000_000
    assert ws.cell(row=amt_row, column=3).value == 4_500_000
    wb.close()


def test_issued_capital_group_write_fills_all_four_columns(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.ISSUED_CAPITAL, level="group")
    out = tmp_path / "NOTES_IC_group.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Amount of shares outstanding at end of period",
            content="",
            evidence="Page 48, Note 16",
            source_pages=[48],
            numeric_values={
                "group_cy": 10_000_000,
                "group_py": 9_000_000,
                "company_cy": 8_000_000,
                "company_py": 7_500_000,
            },
            parent_note={"number": "1", "title": "Test Note"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="group",
        sheet_name=SHEET,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[SHEET]
    row = _find_row(ws, "Amount of shares outstanding at end")
    assert ws.cell(row=row, column=2).value == 10_000_000  # Group CY
    assert ws.cell(row=row, column=3).value == 9_000_000   # Group PY
    assert ws.cell(row=row, column=4).value == 8_000_000   # Company CY
    assert ws.cell(row=row, column=5).value == 7_500_000   # Company PY
    assert ws.cell(row=row, column=6).value == "Page 48, Note 16"
    wb.close()
