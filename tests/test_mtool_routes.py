"""Route tests for the mTool fill pipeline (api/mtool.py, Phase 4).

Uses the reload-server + swap-AUDIT_DB_PATH pattern from test_eval_routes.py.
AUTH_MODE=dev is the suite default (conftest), so requests auto-session.
"""
from __future__ import annotations

import io
import json
import shutil
import sqlite3
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

REPO = Path(__file__).resolve().parent.parent
SOFP = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"


def _import_company_sofp(db_path) -> str:
    from concept_model.filing_targets import persist_template_manifest
    from concept_model.importer import import_company_targets, import_template
    from concept_model.parser import parse_template
    tree = parse_template(str(SOFP))
    jp = Path(db_path).parent / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    tid = import_template(db_path, jp)
    import_company_targets(db_path, tid)
    persist_template_manifest(db_path, SOFP)
    return tid


def _make_run(db, status="completed") -> int:
    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at, "
            "run_config_json) VALUES (?, ?, ?, ?, ?)",
            ("2026-07-05T00:00:00Z", "x.pdf", status, "2026-07-05T00:00:00Z",
             json.dumps({"filing_standard": "mfrs", "filing_level": "company",
                         "denomination": "thousands"})),
        )
        run_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()
    return run_id


def _seed_distinct_leaves(db, run_id, n=4):
    conn = sqlite3.connect(str(db))
    try:
        leaves = conn.execute(
            """
            SELECT concept_uuid FROM concept_nodes
            WHERE kind='LEAF' AND render_sheet LIKE '%Sub%'
              AND canonical_label IN (
                SELECT canonical_label FROM concept_nodes
                GROUP BY canonical_label HAVING COUNT(*) = 1)
            ORDER BY render_row LIMIT ?
            """, (n,)).fetchall()
        for i, (uuid,) in enumerate(leaves):
            conn.execute(
                "INSERT OR REPLACE INTO run_concept_facts("
                "run_id, concept_uuid, period, entity_scope, value, "
                "value_status, updated_at) VALUES (?,?,?,?,?,?,?)",
                (run_id, uuid, "CY", "Company", 1000 + i, "observed",
                 "2026-07-05T00:00:00Z"))
        conn.commit()
    finally:
        conn.close()
    return len(leaves)


def _download(tc, run_id: int, patch_resp, *, acknowledge: str | None = None):
    """Fetch the workbook a patch produced (Step 11A's second step).

    The patch response carries the report and an artifact id; the file itself
    comes from a separate GET, so the operator can read the report before they
    hold the file. A degraded fill needs an acknowledgement to release it.
    """
    body = patch_resp.json()
    params = {}
    if acknowledge is None and body.get("status") != "ok":
        acknowledge = "understood, filling anyway"
    if acknowledge:
        params["acknowledge_degraded"] = acknowledge
    resp = tc.get(body["download_url"], params=params)
    assert resp.status_code == 200, resp.text
    return resp.content


@pytest.fixture
def client(tmp_path, monkeypatch):
    db = tmp_path / "xbrl.db"
    monkeypatch.setenv("XBRL_OUTPUT_DIR", str(tmp_path))
    # No exposure gate exists (2026-08-05 replay decision) — the routes are
    # live unconditionally; tests/test_mtool_preflight.py pins that.
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    monkeypatch.setenv("LLM_PROXY_URL", "")
    import importlib
    import server as srv
    importlib.reload(srv)
    srv.AUDIT_DB_PATH = db
    from db.schema import init_db
    init_db(db)
    _import_company_sofp(db)
    return TestClient(srv.app), db, srv


# ---------------------------------------------------------------- GET doc

def test_get_fill_doc_on_completed_run(client):
    tc, db, _ = client
    run_id = _make_run(db)
    n = _seed_distinct_leaves(db, run_id)
    resp = tc.get(f"/api/runs/{run_id}/mtool-fill")
    assert resp.status_code == 200, resp.text
    doc = resp.json()
    assert len(doc["writes"]) == n
    assert doc["strict"] is True
    assert doc["meta"]["filing_standard"] == "mfrs"
    assert doc["meta"]["columns_unresolved"] is True


def test_get_fill_doc_on_running_run_is_409(client):
    tc, db, _ = client
    run_id = _make_run(db, status="running")
    resp = tc.get(f"/api/runs/{run_id}/mtool-fill")
    assert resp.status_code == 409


def test_get_fill_doc_unknown_run_is_404(client):
    tc, _, _ = client
    assert tc.get("/api/runs/99999/mtool-fill").status_code == 404


def test_filing_semantics_route_reports_versioned_writable_coverage(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)

    resp = tc.get(f"/api/runs/{run_id}/filing-semantics")

    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["readiness"] == "ready"
    assert body["counts"]["writable_fields"] > 0
    assert body["counts"]["unresolved_fields"] == 0
    assert body["manifest_versions"] == ["2022-v1-slot-semantics-2"]
    assert body["taxonomy_versions"] == ["SSMxT_2022v1.0"]


# ---------------------------------------------------------------- POST patch

def _upload_our_template():
    return {"template": ("01-SOFP-CuNonCu.xlsx", SOFP.read_bytes(),
                         "application/vnd.openxmlformats-officedocument."
                         "spreadsheetml.sheet")}


def test_patch_with_explicit_column_map(client, tmp_path):
    tc, db, _ = client
    run_id = _make_run(db)
    n = _seed_distinct_leaves(db, run_id)
    # Our template's sub-sheet is A=label, B=CY.
    doc = tc.get(f"/api/runs/{run_id}/mtool-fill").json()
    sheet = doc["meta"]["sheets_covered"][0]
    cmap = {sheet: {"label_column": "A", "columns": {"current_year": "B"}}}

    resp = tc.post(
        f"/api/runs/{run_id}/mtool-fill/patch",
        files=_upload_our_template(),
        data={"column_map": json.dumps(cmap), "strict": "true"},
    )
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["status"] == "ok"
    assert report["counts"]["written"] == n
    assert report["counts"]["unresolved"] == 0

    # The returned bytes are a valid workbook with the values.
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(_download(tc, run_id, resp)))
    assert sheet in wb.sheetnames


def test_patch_auto_detects_column_map(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(
        f"/api/runs/{run_id}/mtool-fill/patch",
        files=_upload_our_template(),
        data={"strict": "true"},  # no column_map → auto-detect
    )
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["counts"]["written"] >= 1


def test_patch_unverified_semantic_candidate_still_requires_confirmation(
    client, monkeypatch,
):
    """A recognised taxonomy id must not bypass an unverified column map.

    Unknown/group layouts remain confirmation-gated even when some rows can be
    addressed semantically, because other writes may still use legacy columns.
    """
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)

    import api.mtool as mtool_api

    real_detect = mtool_api.detect_column_map
    real_inspect = mtool_api.inspect_template

    def confirmation_required(*args, **kwargs):
        detected = real_detect(*args, **kwargs)
        for layout in detected.values():
            layout["confidence"] = "high"
            layout["requires_confirmation"] = True
            layout.setdefault("notes", []).append(
                "this unverified template still needs confirmation"
            )
        return detected

    def candidate_semantics(*args, **kwargs):
        inspection = real_inspect(*args, **kwargs)
        inspection["semantic_source"] = "taxonomy-identifiers"
        inspection["mtool_compatibility"] = "candidate-2.2"
        return inspection

    monkeypatch.setattr(mtool_api, "detect_column_map", confirmation_required)
    monkeypatch.setattr(mtool_api, "inspect_template", candidate_semantics)

    resp = tc.post(
        f"/api/runs/{run_id}/mtool-fill/patch",
        files=_upload_our_template(),
        data={"strict": "true"},
    )
    assert resp.status_code == 422, resp.text
    assert "confirm" in str(resp.json()["detail"]).lower()


def test_attention_filing_coverage_degrades_and_withholds_artifact(
    client, monkeypatch,
):
    """Legacy/candidate mapping cannot be reported as a clean filing fill."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)

    import api.mtool as mtool_api

    real_resolve = mtool_api.resolve_filing_doc

    def attention_coverage(*args, **kwargs):
        ready, coverage = real_resolve(*args, **kwargs)
        return ready, {
            **coverage,
            "status": "attention",
            "legacy_label_writes": 1,
        }

    monkeypatch.setattr(mtool_api, "resolve_filing_doc", attention_coverage)

    resp = tc.post(
        f"/api/runs/{run_id}/mtool-fill/patch",
        files=_upload_our_template(),
        data={"strict": "true"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["filing_coverage"]["status"] == "attention"
    assert body["status"] == "degraded"
    assert tc.get(body["download_url"]).status_code == 409


def test_detect_columns_returns_map_and_confidence(client):
    # The up-front pre-flight: detect the layout without writing, so the modal
    # can show the column confirmation alongside the notes check.
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(
        f"/api/runs/{run_id}/mtool-fill/detect-columns",
        files=_upload_our_template(),
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert "detected" in body and isinstance(body["detected"], dict)
    assert body["confidence"] in ("high", "low", "medium")
    # Our own template auto-detects cleanly (same layout the patch path uses).
    sheet = next(iter(body["detected"]))
    assert "label_column" in body["detected"][sheet]
    assert "columns" in body["detected"][sheet]


def test_detect_columns_non_xlsx_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(
        f"/api/runs/{run_id}/mtool-fill/detect-columns",
        files={"template": ("junk.xlsx", b"not a zip",
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")},
    )
    assert resp.status_code == 422


def test_detect_columns_running_run_is_409(client):
    tc, db, _ = client
    run_id = _make_run(db, status="running")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/detect-columns",
                   files=_upload_our_template())
    assert resp.status_code == 409


def test_patch_running_run_is_409(client):
    tc, db, _ = client
    run_id = _make_run(db, status="running")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={})
    assert resp.status_code == 409


def test_patch_non_xlsx_upload_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(
        f"/api/runs/{run_id}/mtool-fill/patch",
        files={"template": ("junk.xlsx", b"not a zip",
                            "application/octet-stream")},
        data={},
    )
    assert resp.status_code == 422


def test_patch_no_facts_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)  # no facts seeded
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={})
    assert resp.status_code == 422


def test_patch_malformed_column_map_is_422_not_500(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    # columns must be an object; a string where a dict belongs used to raise
    # AttributeError deep in apply_column_map -> uncaught 500 + temp leak.
    bad = json.dumps({"SOFP-Sub-CuNonCu": "B"})
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"column_map": bad})
    assert resp.status_code == 422


def test_patch_column_map_bad_json_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"column_map": "{not json"})
    assert resp.status_code == 422


def test_report_is_returned_in_full_not_capped(client):
    """Step 11A: the report moved out of an HTTP header and into the body, so
    the 20-row / 6 KB cap — and the `truncated` flag that admitted to it — are
    gone. Row detail is complete."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200
    assert "X-mTool-Report" not in resp.headers
    body = resp.json()
    assert "counts" in body
    assert "truncated" not in body
    for key in ("unresolved", "skipped_formula", "mismatches", "ambiguous",
                "fuzzy_matched", "errors"):
        assert isinstance(body[key], list)


def test_unresolved_rows_are_all_returned_past_the_old_cap(client):
    """The header cap silently elided detail past 20 rows. Prove it doesn't."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    doc = tc.get(f"/api/runs/{run_id}/mtool-fill").json()
    sheet = doc["meta"]["sheets_covered"][0]
    # Point the label column at an empty column: every write goes unresolved.
    cmap = {sheet: {"label_column": "Z", "columns": {"current_year": "B"}}}
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"column_map": json.dumps(cmap), "strict": "true"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["unresolved"]) == body["counts"]["unresolved"]


def test_error_paths_leave_no_temp_dir_behind(client, tmp_path):
    """An error path must clean up immediately — it produced nothing to keep.

    (A SUCCESS keeps its dir until the artifact is fetched or expires — see
    test_successful_patch_keeps_exactly_one_artifact_dir.)
    """
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
            files=_upload_our_template(),
            data={"column_map": json.dumps({"S": 1})})
    tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
            files={"template": ("junk.xlsx", b"not a zip",
                                "application/vnd.openxmlformats-officedocument"
                                ".spreadsheetml.sheet")})
    staging = tmp_path / "_mtool_tmp"
    leftovers = list(staging.glob("*")) if staging.exists() else []
    assert leftovers == [], f"temp dirs leaked: {leftovers}"


def test_oversized_upload_is_413(client, monkeypatch):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    import api.mtool as m
    monkeypatch.setattr(m, "_MAX_TEMPLATE_BYTES", 1000)  # our template > 1 KB
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={})
    assert resp.status_code == 413


def test_zip_bomb_uncompressed_budget_is_413(client, monkeypatch):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    import api.mtool as m
    monkeypatch.setattr(m, "_MAX_UNCOMPRESSED_BYTES", 100)  # tiny budget
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={})
    assert resp.status_code == 413
    assert "decompresses" in resp.json()["detail"]


def test_directory_bomb_member_count_is_413(client, monkeypatch):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    import api.mtool as m
    monkeypatch.setattr(m, "_MAX_ZIP_MEMBERS", 1)  # our template has > 1 part
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={})
    assert resp.status_code == 413
    assert "too many zip members" in resp.json()["detail"]


def _add_note(db, run_id, sheet, row, label, html):
    conn = sqlite3.connect(str(db))
    try:
        conn.execute(
            "INSERT INTO notes_cells(run_id, sheet, row, label, html, "
            "updated_at) VALUES (?,?,?,?,?,?)",
            (run_id, sheet, row, label, html, "2026-07-05T00:00:00Z"))
        conn.commit()
    finally:
        conn.close()


def test_get_notes_fill_doc(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _add_note(db, run_id, "Notes-Listofnotes", 17,
              "Property, plant and equipment", "<h3>PPE</h3>")
    resp = tc.get(f"/api/runs/{run_id}/mtool-notes-fill")
    assert resp.status_code == 200, resp.text
    doc = resp.json()
    assert doc["meta"]["counts"]["notes"] == 1
    assert doc["footnotes"][0]["label"] == "Property, plant and equipment"


def test_get_notes_fill_doc_running_run_is_409(client):
    tc, db, _ = client
    run_id = _make_run(db, status="running")
    assert tc.get(f"/api/runs/{run_id}/mtool-notes-fill").status_code == 409


def test_notes_fill_doc_honours_per_run_theme(client):
    """A per-run notes_table_style override reaches the mTool fill decoration,
    so the payload matches the in-app editor / manual paste (not DEFAULT)."""
    import json as _json
    tc, db, _ = client
    run_id = _make_run(db)
    _add_note(db, run_id, "Notes-Listofnotes", 17,
              "Property, plant and equipment",
              "<table><tbody><tr><td>x</td></tr></tbody></table>")
    conn = sqlite3.connect(str(db))
    try:
        conn.execute("UPDATE runs SET notes_table_style = ? WHERE id = ?",
                     (_json.dumps({"borderStyle": "double",
                                   "borderColor": "#1F3864"}), run_id))
        conn.commit()
    finally:
        conn.close()
    doc = tc.get(f"/api/runs/{run_id}/mtool-notes-fill").json()
    html = doc["footnotes"][0]["html"].lower()
    assert "3px double #1f3864" in html          # themed, not the #999 default
    assert "1px solid #999" not in html


def test_notes_fill_doc_falls_back_to_firm_default_theme(client, monkeypatch):
    """No per-run override → the firm-wide XBRL_NOTES_TABLE_STYLE applies."""
    tc, db, _ = client
    run_id = _make_run(db)
    _add_note(db, run_id, "Notes-Listofnotes", 17,
              "Property, plant and equipment",
              "<table><tbody><tr><td>x</td></tr></tbody></table>")
    monkeypatch.setenv("XBRL_NOTES_TABLE_STYLE",
                       '{"borderColor": "#abcdef"}')
    doc = tc.get(f"/api/runs/{run_id}/mtool-notes-fill").json()
    assert "1px solid #abcdef" in doc["footnotes"][0]["html"].lower()


def test_patch_fill_notes_off_omits_notes_block(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"strict": "true", "fill_notes": "false"})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert "notes" not in report


def test_patch_fill_notes_on_reports_notes_block(client):
    """With notes present, the header carries a notes block. Our SOFP template
    has no +FootnoteTexts sheet, so notes can't land — but the download must
    still succeed (numbers filled) and surface the notes status gracefully."""
    tc, db, _ = client
    run_id = _make_run(db)
    n = _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["counts"]["written"] == n          # numbers still filled
    assert "notes" in report                          # notes block present
    assert report["notes"]["status"] == "degraded"    # no +FootnoteTexts here
    # Combined status must reflect the notes failure, NOT hide behind numeric ok.
    assert report["status"] == "degraded"
    assert report["numeric_status"] == "ok"
    # The returned bytes are still a valid workbook.
    import openpyxl
    openpyxl.load_workbook(io.BytesIO(_download(tc, run_id, resp)))


def test_patch_notes_styling_none_fills_raw_and_reports_it(client, monkeypatch):
    """The diagnostic "no styling" toggle: notes_styling=none reaches the
    exporter as decorate=False, and the report header carries the honest
    styling_disabled flag so a plain fill can't be misread as a bug."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    import api.mtool as m
    seen = {}
    orig = m.build_notes_fill_doc

    def spy(*a, **k):
        seen.update(k)
        return orig(*a, **k)

    monkeypatch.setattr(m, "build_notes_fill_doc", spy)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"strict": "true", "notes_styling": "none"})
    assert resp.status_code == 200, resp.text
    assert seen.get("decorate") is False
    report = resp.json()
    assert report["notes"]["styling_disabled"] is True


def test_patch_notes_styling_defaults_to_styled(client, monkeypatch):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    import api.mtool as m
    seen = {}
    orig = m.build_notes_fill_doc

    def spy(*a, **k):
        seen.update(k)
        return orig(*a, **k)

    monkeypatch.setattr(m, "build_notes_fill_doc", spy)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200, resp.text
    assert seen.get("decorate") is True
    report = resp.json()
    assert report["notes"]["styling_disabled"] is False
    # The size-tier counters ride along (all zero for this tiny note).
    assert report["notes"]["counts"]["formatting_compacted"] == 0
    assert report["notes"]["counts"]["formatting_reduced"] == 0
    assert report["notes"]["counts"]["formatting_dropped"] == 0
    assert report["notes"]["counts"]["source_styling_dropped"] == 0
    assert report["notes"]["counts"]["white_grid_dropped"] == 0


def test_patch_reports_white_grid_dropped(client):
    """A source-styled note that only fits with the run-76 white grid dropped
    must surface that in the response header — code review 2026-07-20
    (round 3): the exporter counted it but the route never copied it, so the
    modal warning could never appear."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    # Borderless verbatim table sized so full/compact/lite fit only WITHOUT
    # the white grid (65 rows × 6 cols — the measured fallback window).
    cell = '<td style="padding: 1px 0px">39,827</td>'
    mid = ('<table data-source-styled="true">'
           + "".join("<tr>" + cell * 6 + "</tr>" for _ in range(65))
           + "</table>")
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", mid)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["notes"]["counts"]["white_grid_dropped"] == 1


def test_patch_reports_source_styling_dropped(client):
    """A verbatim Word note too big to keep its own styling files destyled —
    and the loss must reach the operator. Code review 2026-07-20 (round 2):
    the exporter recorded the count in the fill doc, but the route only
    copied the three older formatting counters, so the response header (and
    the modal reading it) never saw it."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    cell = ('<td style="padding: 1px 5px; text-align: right; '
            'border-bottom: 1px solid #7F7F7F">1,595</td>')
    big = ("<table>"
           + "".join("<tr>" + cell * 6 + "</tr>" for _ in range(100))
           + "</table>")
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", big)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["notes"]["counts"]["source_styling_dropped"] == 1


def test_patch_bad_notes_styling_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"strict": "true", "notes_styling": "fancy"})
    assert resp.status_code == 422
    assert "notes_styling" in resp.json()["detail"]


def test_patch_notes_exception_preserves_numeric_fill(client, monkeypatch):
    """A notes-fill raise must NOT discard the already-good numeric workbook.

    Notes fill is best-effort (gotcha #22): on an unexpected exception the
    numeric-only fill is still returned (200) with the notes side degraded."""
    tc, db, _ = client
    run_id = _make_run(db)
    n = _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    import api.mtool as m

    def boom(*_a, **_k):
        raise RuntimeError("notes patcher exploded")

    monkeypatch.setattr(m, "fill_footnotes", boom)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["counts"]["written"] == n       # numbers survived the raise
    assert report["notes"]["status"] == "degraded"
    assert report["status"] == "degraded"
    assert report["numeric_status"] == "ok"
    import openpyxl
    openpyxl.load_workbook(io.BytesIO(_download(tc, run_id, resp)))  # still a valid workbook


def test_patch_invalid_notes_doc_reports_degraded(client, monkeypatch):
    """An invalid machine-generated notes doc is reported degraded, not
    silently collapsed into "skipped"."""
    tc, db, _ = client
    run_id = _make_run(db)
    n = _seed_distinct_leaves(db, run_id)
    import api.mtool as m
    # footnotes present but each missing 'html' -> validate_notes_input errors.
    monkeypatch.setattr(
        m, "build_notes_fill_doc",
        lambda *_a, **_k: {"footnotes": [{"label": "PPE"}], "meta": {}})
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["counts"]["written"] == n
    assert report["notes"]["status"] == "degraded"
    assert report["status"] == "degraded"


def test_patch_non_letter_column_map_is_422(client):
    """A structurally-valid map with a non-letter column is rejected up front
    (422), not surfaced only as a buried per-write error."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    doc = tc.get(f"/api/runs/{run_id}/mtool-fill").json()
    sheet = doc["meta"]["sheets_covered"][0]
    bad = json.dumps(
        {sheet: {"label_column": "A", "columns": {"current_year": "1"}}})
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"column_map": bad})
    assert resp.status_code == 422


def test_successful_patch_keeps_exactly_one_artifact_dir(client, tmp_path):
    """Step 11A moved the file behind a second request, so the workbook has to
    outlive the patch response — but only by one bounded holding area."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200
    held = list((tmp_path / "_mtool_tmp").glob("*"))
    assert len(held) == 1, held
    assert (held[0] / "filled.xlsx").exists()


def test_expired_artifact_is_swept_and_404s(client, tmp_path, monkeypatch):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    url = resp.json()["download_url"]

    import api.mtool as m
    monkeypatch.setattr(m, "_ARTIFACT_TTL_S", -1.0)  # everything is expired
    gone = tc.get(url)
    assert gone.status_code == 404
    assert "no longer available" in gone.json()["detail"]
    leftovers = list((tmp_path / "_mtool_tmp").glob("*"))
    assert leftovers == [], f"expired artifact not swept: {leftovers}"


# ------------------------------------------------ notes preview (diagnostic)

def test_notes_preview_reports_plan_and_slot_count(client):
    """The dry-run diagnostic returns a per-note plan + the template's existing
    fn_* slot count, and writes nothing. Our SOFP template has no fn_* slots,
    so a seeded note lands in `unresolved` and template_fn_slots is 0 — the
    signal that the concept isn't popup-backed."""
    tc, db, _ = client
    run_id = _make_run(db)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/notes-preview",
                   files=_upload_our_template(),
                   data={"create_missing_notes": "false"})
    assert resp.status_code == 200, resp.text
    plan = resp.json()
    assert plan["notes_in_run"] == 1
    assert plan["template_fn_slots"] == 0
    assert plan["will_fill_existing"] == [] and plan["will_create"] == []
    # No +FootnoteTexts sheet -> the fill degrades with an error, not a per-note
    # unresolved; either way nothing is planned to write.
    assert plan["unresolved"] or plan["errors"]


def test_notes_preview_honours_notes_styling(client, monkeypatch):
    """The dry-run plan's counts must reflect the styling the operator chose —
    a "no styling" preview reaches the exporter as decorate=False so it can't
    report styled size-degradation tiers."""
    tc, db, _ = client
    run_id = _make_run(db)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    import api.mtool as m
    seen = {}
    orig = m.build_notes_fill_doc

    def spy(*a, **k):
        seen.update(k)
        return orig(*a, **k)

    monkeypatch.setattr(m, "build_notes_fill_doc", spy)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/notes-preview",
                   files=_upload_our_template(),
                   data={"notes_styling": "none"})
    assert resp.status_code == 200, resp.text
    assert seen.get("decorate") is False


def test_notes_preview_bad_notes_styling_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/notes-preview",
                   files=_upload_our_template(),
                   data={"notes_styling": "fancy"})
    assert resp.status_code == 422
    assert "notes_styling" in resp.json()["detail"]


def test_notes_preview_running_run_is_409(client):
    tc, db, _ = client
    run_id = _make_run(db, status="running")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/notes-preview",
                   files=_upload_our_template())
    assert resp.status_code == 409


def test_patch_accepts_create_missing_notes_param(client):
    """The create_missing_notes toggle is wired through the patch endpoint
    (no 4xx/5xx from the extra form field); numeric fill is unaffected."""
    tc, db, _ = client
    run_id = _make_run(db)
    n = _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(),
                   data={"strict": "true", "create_missing_notes": "true"})
    assert resp.status_code == 200, resp.text
    report = resp.json()
    assert report["counts"]["written"] == n


# ------------------------------------------ notes_targets (operator guidance)

def test_patch_malformed_notes_targets_is_422(client):
    """A malformed notes_targets is a caller error (422), never a silent
    fall-back to the label guess the operator just overrode."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    for bad in ("not json", json.dumps(["list"]),
                json.dumps({"99": {"sheet": "S", "cell": "E1"}}),   # out of range
                json.dumps({"0": {"cell": "E14"}}),                  # no sheet
                json.dumps({"0": {"key": "not_a_key"}}),
                json.dumps({"0": {"sheet": "S", "cell": "14E"}})):
        resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                       files=_upload_our_template(),
                       data={"strict": "true", "notes_targets": bad})
        assert resp.status_code == 422, (bad, resp.text)


def test_preview_applies_notes_targets_override(client):
    """An explicit cell decision replaces label matching for that note: the
    re-preview reflects it (the item leaves `unresolved`), keyed by index."""
    tc, db, _ = client
    run_id = _make_run(db)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    override = json.dumps({"0": {"sheet": "Notes-CI", "cell": "e14"}})
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/notes-preview",
                   files=_upload_our_template(),
                   data={"create_missing_notes": "true",
                         "notes_targets": override})
    assert resp.status_code == 200, resp.text
    plan = resp.json()
    # Our SOFP template has no Notes-CI sheet / +FootnoteTexts, so the item
    # can't actually land — but it must be routed via the EXPLICIT-cell path
    # (errors/unresolved mention the cell, not the label-match failure), and
    # the malformed-JSON 422 above proves the field is parsed. The key
    # assertion: no label-based 'no visible note-sheet label matched' refusal.
    details = json.dumps(plan["unresolved"] + plan["errors"])
    assert "no visible note-sheet label matched" not in details


def test_preview_unresolved_entries_are_structured(client):
    """The preview passes through the decision-UI contract fields (index +
    reason) rather than flattening to label+detail."""
    tc, db, _ = client
    run_id = _make_run(db)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/notes-preview",
                   files=_upload_our_template(),
                   data={"create_missing_notes": "true"})
    assert resp.status_code == 200, resp.text
    plan = resp.json()
    if plan["unresolved"]:  # template has no matching sheets -> no_match
        entry = plan["unresolved"][0]
        assert entry["index"] == 0
        assert entry["reason"] in {"no_match", "ambiguous",
                                   "strict_near_miss", "no_slot",
                                   "no_payload_row"}


# ------------------------------------------- peer-review fixes (2026-08-05)

def _corrupt_worksheets(src_path, tmp_path) -> bytes:
    """A workbook with every worksheet's XML truncated mid-element.

    Still a valid ZIP with a readable workbook.xml, so it passes the
    zip-level guards — the parse failure only surfaces when a reader touches
    a damaged sheet part."""
    import zipfile

    out = tmp_path / "corrupt.xlsx"
    with zipfile.ZipFile(src_path) as zin, \
            zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for name in zin.namelist():
            blob = zin.read(name)
            if name.startswith("xl/worksheets/"):
                blob = blob[: len(blob) // 2] + b"<unclosed"
            zout.writestr(name, blob)
    return out.read_bytes()


def _upload_bytes(blob: bytes):
    return {"template": ("t.xlsx", blob,
                         "application/vnd.openxmlformats-officedocument."
                         "spreadsheetml.sheet")}


def test_malformed_worksheet_xml_is_422_on_detect(client, tmp_path):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    blob = _corrupt_worksheets(SOFP, tmp_path)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/detect-columns",
                   files=_upload_bytes(blob))
    assert resp.status_code == 422, resp.text
    assert "could not be read" in str(resp.json()["detail"])


def test_malformed_worksheet_xml_is_422_on_patch(client, tmp_path):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    blob = _corrupt_worksheets(SOFP, tmp_path)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_bytes(blob), data={"strict": "true"})
    assert resp.status_code == 422, resp.text


def test_malformed_worksheet_xml_is_422_on_notes_preview(client, tmp_path):
    """The preview only parses worksheet XML on a template that has footnote
    structure (a bare workbook is refused earlier with a structured error),
    so the fixture needs a +FootnoteTexts sheet and a note in the run."""
    from openpyxl import Workbook

    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    _add_note(db, run_id, "Notes-CI", 12, "Corporate information", "<p>x</p>")
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes-CI"
    ws["D12"] = "Corporate information"
    fn = wb.create_sheet("+FootnoteTexts")
    fn["A1"] = "fn_1"
    fn["B1"] = "Notes-CI"
    fn["C1"] = ""
    clean = tmp_path / "notes_template.xlsx"
    wb.save(clean)
    blob = _corrupt_worksheets(clean, tmp_path)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/notes-preview",
                   files=_upload_bytes(blob))
    assert resp.status_code == 422, resp.text


def _explicit_map_resp(tc, run_id, cmap):
    return tc.post(
        f"/api/runs/{run_id}/mtool-fill/patch",
        files=_upload_our_template(),
        data={"column_map": json.dumps(cmap), "strict": "true"})


def test_cmap_value_column_equal_to_label_column_is_422(client):
    """Figures addressed BY the labels must never overwrite them."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    sheet = tc.get(f"/api/runs/{run_id}/mtool-fill").json()[
        "meta"]["sheets_covered"][0]
    resp = _explicit_map_resp(tc, run_id, {
        sheet: {"label_column": "A", "columns": {"current_year": "A"}}})
    assert resp.status_code == 422, resp.text
    assert "label column" in str(resp.json()["detail"])


def test_cmap_duplicate_value_columns_are_422(client):
    """Two periods pointed at one physical column: last write wins, silently."""
    tc, db, _ = client
    run_id = _make_run(db)
    # Seed BOTH periods so the doc genuinely wants two value columns.
    conn = sqlite3.connect(str(db))
    try:
        leaves = conn.execute(
            "SELECT concept_uuid FROM concept_nodes WHERE kind='LEAF' "
            "AND render_sheet LIKE '%Sub%' AND canonical_label IN ("
            "SELECT canonical_label FROM concept_nodes GROUP BY "
            "canonical_label HAVING COUNT(*) = 1) ORDER BY render_row "
            "LIMIT 2").fetchall()
        for uuid, in leaves:
            for period in ("CY", "PY"):
                conn.execute(
                    "INSERT OR REPLACE INTO run_concept_facts(run_id, "
                    "concept_uuid, period, entity_scope, value, value_status, "
                    "updated_at) VALUES (?,?,?,?,?,?,?)",
                    (run_id, uuid, period, "Company", 7, "observed", "t"))
        conn.commit()
    finally:
        conn.close()
    sheet = tc.get(f"/api/runs/{run_id}/mtool-fill").json()[
        "meta"]["sheets_covered"][0]
    resp = _explicit_map_resp(tc, run_id, {
        sheet: {"label_column": "A",
                "columns": {"current_year": "B", "prior_year": "B"}}})
    assert resp.status_code == 422, resp.text
    assert "same physical column" in str(
        resp.json()["detail"]) or "point at column" in str(
        resp.json()["detail"])


def test_cmap_unexpected_role_is_422(client):
    """A typo'd role must be rejected, not silently ignored."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    sheet = tc.get(f"/api/runs/{run_id}/mtool-fill").json()[
        "meta"]["sheets_covered"][0]
    resp = _explicit_map_resp(tc, run_id, {
        sheet: {"label_column": "A",
                "columns": {"current_year": "B", "current_yaer": "C"}}})
    assert resp.status_code == 422, resp.text
    assert "does not write" in str(resp.json()["detail"])


def test_unit_scale_mismatch_degrades_the_fill(client, monkeypatch):
    """A declared-unit disagreement is the 1000x risk — it must not ride out
    under status "ok" with an unguarded download (peer review, 2026-08-05)."""
    import api.mtool as m
    monkeypatch.setattr(
        m, "unit_scale_mismatches",
        lambda detected, denomination: [{
            "sheet": "S", "column": "E",
            "template_declares": "thousands", "run_denomination": "units"}])
    tc, db, _ = client
    run_id = _make_run(db)
    _seed_distinct_leaves(db, run_id)
    resp = tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=_upload_our_template(), data={"strict": "true"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["unit_scale_warnings"], "warning must be in the report"
    assert body["status"] == "degraded"
    assert body["numeric_status"] == "ok"  # the fill itself was clean
    # Degraded => the artifact is withheld until acknowledged.
    bare = tc.get(body["download_url"])
    assert bare.status_code == 409
    acked = tc.get(body["download_url"],
                   params={"acknowledge_degraded": "checked the units"})
    assert acked.status_code == 200
