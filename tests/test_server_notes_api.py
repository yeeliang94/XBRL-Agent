"""End-to-end server test for notes_to_run plumbing.

Mocks both face + notes coordinators so we exercise the request-parsing,
event-multiplexing, and merger-wiring paths without calling real LLMs.
"""
from __future__ import annotations

import asyncio
import json
import uuid
from pathlib import Path
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient

from coordinator import AgentResult, CoordinatorResult
from notes.coordinator import NotesAgentResult, NotesCoordinatorResult
from notes_types import NotesTemplateType
from statement_types import StatementType
from workbook_merger import MergeResult


def _session(tmp_path: Path) -> tuple[TestClient, str]:
    """Spin up a TestClient pointed at a temp OUTPUT_DIR with a seeded upload."""
    import server as server_module

    server_module.OUTPUT_DIR = tmp_path
    server_module.AUDIT_DB_PATH = tmp_path / "audit.sqlite"
    client = TestClient(server_module.app)
    session_id = str(uuid.uuid4())
    upload_dir = tmp_path / session_id
    upload_dir.mkdir(parents=True)
    (upload_dir / "uploaded.pdf").write_bytes(b"%PDF-1.4\n%fake\n")
    return client, session_id


def _parse_sse(text: str) -> list[dict]:
    events = []
    blocks = text.strip().split("\n\n")
    for blk in blocks:
        lines = [ln for ln in blk.splitlines() if ln]
        if not lines:
            continue
        evt = {}
        for ln in lines:
            if ln.startswith("event:"):
                evt["event"] = ln.split(":", 1)[1].strip()
            elif ln.startswith("data:"):
                evt["data"] = json.loads(ln.split(":", 1)[1].strip())
        if "event" in evt:
            events.append(evt)
    return events


@pytest.mark.asyncio
async def test_notes_to_run_is_accepted_and_passed_to_coordinator(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    client, session_id = _session(tmp_path)

    fake_face = CoordinatorResult(agent_results=[])
    fake_notes = NotesCoordinatorResult(agent_results=[
        NotesAgentResult(
            template_type=NotesTemplateType.CORP_INFO,
            status="succeeded",
            workbook_path=str(tmp_path / session_id / "NOTES_CORP_INFO_filled.xlsx"),
        ),
    ])

    received = {}

    async def mock_face(config, infopack=None, event_queue=None, session_id=None, **_kwargs):
        return fake_face

    async def mock_notes(config, infopack=None, event_queue=None, session_id=None, **_kwargs):
        received["notes_to_run"] = set(config.notes_to_run)
        received["filing_level"] = config.filing_level
        if event_queue is not None:
            await event_queue.put({
                "event": "complete",
                "data": {
                    "agent_id": "notes:CORP_INFO",
                    "agent_role": "CORP_INFO",
                    "success": True,
                    "workbook_path": fake_notes.agent_results[0].workbook_path,
                },
            })
        return fake_notes

    # Seed an existing notes workbook so the merger has something to pick up.
    notes_wb = tmp_path / session_id / "NOTES_CORP_INFO_filled.xlsx"
    # The merger opens it with openpyxl — write a minimal valid workbook.
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "Notes-CI"
    wb.save(notes_wb)
    wb.close()

    with patch("server._create_proxy_model", return_value="fake"), \
         patch("coordinator.run_extraction", side_effect=mock_face), \
         patch("notes.coordinator.run_notes_extraction", side_effect=mock_notes), \
         patch("cross_checks.framework.run_all", return_value=[]):
        resp = client.post(f"/api/run/{session_id}", json={
            "statements": [],
            "variants": {},
            "models": {},
            "filing_level": "company",
            "notes_to_run": ["CORP_INFO"],
        })

    assert resp.status_code == 200
    events = _parse_sse(resp.text)
    assert received["notes_to_run"] == {NotesTemplateType.CORP_INFO}
    assert received["filing_level"] == "company"

    # run_complete event should include notes_completed
    rc = next((e for e in events if e["event"] == "run_complete"), None)
    assert rc is not None
    assert rc["data"].get("notes_completed") == ["CORP_INFO"]
    assert rc["data"].get("notes_failed") == []


def test_public_notes_allowlist_covers_every_enum_member():
    """Review S3: explicit allowlist drift detection. If a new
    NotesTemplateType is added but forgotten in _PUBLIC_NOTES_TEMPLATES,
    this test surfaces the omission loudly."""
    from server import _PUBLIC_NOTES_TEMPLATES
    assert set(_PUBLIC_NOTES_TEMPLATES) == set(NotesTemplateType), (
        "Allowlist drift: every NotesTemplateType member must be in "
        "_PUBLIC_NOTES_TEMPLATES unless intentionally hidden. If hidden, "
        "update this test with an explicit exception."
    )


@pytest.mark.asyncio
async def test_notes_models_resolved_per_template(tmp_path: Path, monkeypatch):
    """Per-note model overrides are routed into NotesRunConfig.models and
    returned by model_for(). Templates without an override fall back to the
    run-wide default, same as face statements.
    """
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    client, session_id = _session(tmp_path)

    # Capture the NotesRunConfig the server hands to the coordinator so we
    # can assert the resolved per-template models wired up correctly.
    captured: dict = {}

    async def mock_notes(config, infopack=None, event_queue=None, session_id=None, **_kwargs):
        captured["config"] = config
        if event_queue is not None:
            for nt in sorted(config.notes_to_run, key=lambda n: n.value):
                await event_queue.put({
                    "event": "complete",
                    "data": {
                        "agent_id": f"notes:{nt.value}",
                        "agent_role": nt.value,
                        "success": True,
                        "workbook_path": str(tmp_path / session_id / f"NOTES_{nt.value}_filled.xlsx"),
                    },
                })
        return NotesCoordinatorResult(agent_results=[
            NotesAgentResult(
                template_type=nt,
                status="succeeded",
                workbook_path=str(tmp_path / session_id / f"NOTES_{nt.value}_filled.xlsx"),
            )
            for nt in sorted(config.notes_to_run, key=lambda n: n.value)
        ])

    # Seed minimal workbooks for the merger — two notes requested.
    import openpyxl
    for nt in ("CORP_INFO", "ACC_POLICIES"):
        wb = openpyxl.Workbook()
        wb.active.title = "Notes-CI" if nt == "CORP_INFO" else "Notes-SummaryofAccPol"
        wb.save(tmp_path / session_id / f"NOTES_{nt}_filled.xlsx")
        wb.close()

    # Proxy-model factory returns a sentinel keyed by model name so we can
    # prove the right string reached the right template.
    proxy_calls: list[str] = []

    def fake_proxy_model(name, _url, _key):
        proxy_calls.append(name)
        return f"model:{name}"

    with patch("server._create_proxy_model", side_effect=fake_proxy_model), \
         patch("coordinator.run_extraction", return_value=__import__("coordinator").CoordinatorResult()), \
         patch("notes.coordinator.run_notes_extraction", side_effect=mock_notes), \
         patch("cross_checks.framework.run_all", return_value=[]):
        resp = client.post(f"/api/run/{session_id}", json={
            "statements": [],
            "notes_to_run": ["CORP_INFO", "ACC_POLICIES"],
            "notes_models": {
                "ACC_POLICIES": "claude-opus-4-6",
                # CORP_INFO intentionally omitted — falls back to default.
            },
        })

    assert resp.status_code == 200
    cfg = captured["config"]
    # ACC_POLICIES picked up its override; CORP_INFO falls back to config.model.
    assert cfg.model_for(NotesTemplateType.ACC_POLICIES) == "model:claude-opus-4-6"
    assert cfg.model_for(NotesTemplateType.CORP_INFO) == cfg.model
    # The override factory should have been called at least once for the
    # overriding model name (among potentially other calls for the default).
    assert "claude-opus-4-6" in proxy_calls


@pytest.mark.asyncio
async def test_notes_models_unknown_key_ignored(tmp_path: Path, monkeypatch):
    """A stray key in notes_models (typo, removed enum member, etc.) must
    not fail the whole run — unknown keys are silently dropped, matching
    how we already handle unknown entries in other optional dicts.
    """
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    client, session_id = _session(tmp_path)

    async def mock_notes(config, infopack=None, event_queue=None, session_id=None, **_kwargs):
        if event_queue is not None:
            await event_queue.put({
                "event": "complete",
                "data": {
                    "agent_id": "notes:CORP_INFO",
                    "agent_role": "CORP_INFO",
                    "success": True,
                    "workbook_path": str(tmp_path / session_id / "NOTES_CORP_INFO_filled.xlsx"),
                },
            })
        return NotesCoordinatorResult(agent_results=[
            NotesAgentResult(
                template_type=NotesTemplateType.CORP_INFO,
                status="succeeded",
                workbook_path=str(tmp_path / session_id / "NOTES_CORP_INFO_filled.xlsx"),
            )
        ])

    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = "Notes-CI"
    wb.save(tmp_path / session_id / "NOTES_CORP_INFO_filled.xlsx")
    wb.close()

    with patch("server._create_proxy_model", return_value="fake"), \
         patch("coordinator.run_extraction", return_value=__import__("coordinator").CoordinatorResult()), \
         patch("notes.coordinator.run_notes_extraction", side_effect=mock_notes), \
         patch("cross_checks.framework.run_all", return_value=[]):
        resp = client.post(f"/api/run/{session_id}", json={
            "statements": [],
            "notes_to_run": ["CORP_INFO"],
            "notes_models": {"MADE_UP": "ghost-model"},
        })

    assert resp.status_code == 200
    events = _parse_sse(resp.text)
    # Run must complete normally — no error events about the stray key.
    rc = next(e for e in events if e["event"] == "run_complete")
    assert rc["data"]["success"] is True


def test_rerun_accepts_one_notes_template(tmp_path: Path, monkeypatch):
    """Peer-review finding #1: /api/rerun/ must accept notes-only payloads
    (1 note, 0 statements) so the frontend rerun button on a failed notes
    tab doesn't produce a guaranteed 400.
    """
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    client, session_id = _session(tmp_path)

    # Patch the inner runner so we don't actually invoke LLMs — we're only
    # asserting the validation gate here.
    async def fake_stream(**_kwargs):
        yield {"event": "status", "data": {"phase": "starting", "message": "ok"}}
        yield {"event": "run_complete", "data": {"success": True}}

    with patch("server.run_multi_agent_stream", side_effect=fake_stream):
        resp = client.post(f"/api/rerun/{session_id}", json={
            "statements": [],
            "notes_to_run": ["CORP_INFO"],
        })
    assert resp.status_code == 200


def test_rerun_rejects_zero_and_multi_agent_payloads(tmp_path: Path, monkeypatch):
    """The relaxed gate still enforces exactly one agent — can't rerun
    nothing and can't rerun two at once. Guards against payloads that
    would re-trigger a full multi-agent run under the rerun endpoint.
    """
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    client, session_id = _session(tmp_path)

    # Zero agents.
    resp = client.post(f"/api/rerun/{session_id}", json={"statements": []})
    assert resp.status_code == 400

    # Two statements.
    resp = client.post(f"/api/rerun/{session_id}", json={"statements": ["SOFP", "SOPL"]})
    assert resp.status_code == 400

    # One statement + one notes template (mixed).
    resp = client.post(f"/api/rerun/{session_id}", json={
        "statements": ["SOFP"],
        "notes_to_run": ["CORP_INFO"],
    })
    assert resp.status_code == 400


@pytest.mark.asyncio
async def test_unknown_notes_template_rejected_with_error(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    client, session_id = _session(tmp_path)

    with patch("server._create_proxy_model", return_value="fake"), \
         patch("coordinator.run_extraction", return_value=CoordinatorResult()), \
         patch("notes.coordinator.run_notes_extraction", return_value=NotesCoordinatorResult()):
        resp = client.post(f"/api/run/{session_id}", json={
            "statements": [],
            "notes_to_run": ["MADE_UP"],
        })

    assert resp.status_code == 200
    events = _parse_sse(resp.text)
    errs = [e for e in events if e["event"] == "error"]
    assert any("MADE_UP" in e["data"].get("message", "") for e in errs)
    # Run must terminate with run_complete success=False
    rc = next(e for e in events if e["event"] == "run_complete")
    assert rc["data"]["success"] is False
