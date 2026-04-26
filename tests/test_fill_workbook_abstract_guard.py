"""Bug A (2026-04-26) — fill_workbook must refuse writes to dark-navy XBRL
abstract section-header rows.

Failure mode that prompted these tests: a Windows extraction run on
SOPL-Analysis-Function wrote 6,092 onto row 27 ("Interest income") — the
section-header row — instead of one of the leaf rows below (28 / 29). The
total row (30, formula `=1*B28+1*B29`) then evaluated to 0 because the
leaves were empty, even though the workbook "looked" populated.

Two coupled defences live in `tools/fill_workbook.py`:

  1. Refuse any write whose target row is an abstract section header.
  2. When a label matches both a header row AND a leaf row in the same
     sheet (the "Other fee and commission income" case), `_find_row_by_label`
     prefers the leaf — so the refusal in (1) is only triggered on labels
     that have no leaf at all.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from tools.fill_workbook import fill_workbook

# The dark-navy ARGB used in MBRS templates for section-header rows.
# Mirrors `tools.section_headers._HEADER_FILL_RGB`. Tests construct minimal
# workbooks rather than depend on the full SSM template so they stay fast.
_HEADER_FILL_ARGB = "FF1F3864"


def _paint_header(ws, row: int) -> None:
    ws.cell(row=row, column=1).fill = PatternFill(
        start_color=_HEADER_FILL_ARGB,
        end_color=_HEADER_FILL_ARGB,
        fill_type="solid",
    )


def _make_sopl_like(tmp_path) -> str:
    """Construct a SOPL-Analysis-Function-shaped sheet: one section header
    'Interest income' (dark navy) followed by two leaves and a total row.
    The shape mirrors rows 27–30 of the real template.
    """
    path = str(tmp_path / "sopl_like.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Analysis-Function"
    ws["A5"] = "Interest income"
    _paint_header(ws, 5)
    ws["A6"] = "Interest income on loans, advances and financing"
    ws["A7"] = "Interest income on other financial assets"
    ws["A8"] = "Total interest income"
    ws["B8"] = "=B6+B7"  # formula row — protected by existing guard
    wb.save(path)
    wb.close()
    return path


def _make_dup_label_sheet(tmp_path) -> str:
    """The "Other fee and commission income" case: identical label appears
    once as a header and once as a leaf, in that order. fill_workbook used
    to return the FIRST exact match (the header); the new behaviour is to
    prefer the leaf."""
    path = str(tmp_path / "dup_label.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOPL-Analysis-Function"
    ws["A5"] = "Other fee and commission income"        # header (navy)
    _paint_header(ws, 5)
    ws["A6"] = "Gross brokerage and other charges"
    ws["A7"] = "Underwriting commissions and fund management income"
    ws["A8"] = "Other fee and commission income"        # leaf (no fill)
    ws["A9"] = "Total other fee and commission income"
    ws["B9"] = "=B6+B7+B8"
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# 1.4 — refuse writes to abstract rows
# ---------------------------------------------------------------------------

class TestRefusesWritesToAbstractRows:
    def test_refuses_write_to_abstract_section_header(self, tmp_path):
        template = _make_sopl_like(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        # Mirrors the screenshot bug — agent wrote 6092 to the "Interest income"
        # header row instead of the leaves below.
        payload = json.dumps([
            {
                "sheet": "SOPL-Analysis-Function",
                "field_label": "Interest income",
                "col": 2,
                "value": 6092,
                "evidence": "Page 36 Note 6",
            },
        ])

        result = fill_workbook(template, output, payload)

        assert result.fields_written == 0, (
            "writer must refuse to land a value on a dark-navy section-"
            "header row"
        )
        assert result.errors, "expected an actionable error message"
        joined = " ".join(result.errors).lower()
        assert "interest income" in joined or "abstract" in joined or "section header" in joined, (
            f"error must name the offending row / call it abstract. Got: {result.errors}"
        )

    def test_writes_to_leaf_row_succeed(self, tmp_path):
        """Regression guard — leaves under a header are still writable."""
        template = _make_sopl_like(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        payload = json.dumps([
            {
                "sheet": "SOPL-Analysis-Function",
                "field_label": "Interest income on loans, advances and financing",
                "col": 2,
                "value": 6092,
            },
        ])

        result = fill_workbook(template, output, payload)
        assert result.success, f"leaf write should succeed. Errors: {result.errors}"
        assert result.fields_written == 1

        wb = openpyxl.load_workbook(output)
        assert wb["SOPL-Analysis-Function"].cell(row=6, column=2).value == 6092
        wb.close()

    def test_refusal_does_not_block_other_writes_in_same_payload(self, tmp_path):
        """Mixed payload: one bad write (header), one good write (leaf).
        The good write must still land — fill_workbook already iterates
        per-mapping, and the header refusal must follow the same pattern
        as the existing formula-cell refusal."""
        template = _make_sopl_like(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        payload = json.dumps([
            {  # bad — abstract header
                "sheet": "SOPL-Analysis-Function",
                "field_label": "Interest income",
                "col": 2,
                "value": 6092,
            },
            {  # good — leaf below
                "sheet": "SOPL-Analysis-Function",
                "field_label": "Interest income on other financial assets",
                "col": 2,
                "value": 5307,
            },
        ])

        result = fill_workbook(template, output, payload)

        assert result.fields_written == 1
        assert result.errors, "the abstract-row write should still produce an error"
        wb = openpyxl.load_workbook(output)
        # Leaf write landed
        assert wb["SOPL-Analysis-Function"].cell(row=7, column=2).value == 5307
        # Header row was NOT touched
        assert wb["SOPL-Analysis-Function"].cell(row=5, column=2).value is None
        wb.close()


# ---------------------------------------------------------------------------
# 1.6 — duplicate label: prefer leaf over header
# ---------------------------------------------------------------------------

class TestAbstractGuardOnRealMpersTemplate:
    """End-to-end pin (2026-04-26): the writer's abstract-row guard must
    refuse header writes on a real MPERS template. Originally a no-op on
    MPERS because the generator did not paint fills (peer-review #8); the
    parity fix makes the guard light up just like on MFRS."""

    _MPERS_TEMPLATE = str(
        Path(__file__).resolve().parent.parent
        / "XBRL-template-MPERS"
        / "Group"
        / "03-SOPL-Function.xlsx"
    )

    def test_writer_refuses_abstract_writes_on_mpers_sopl_analysis(self, tmp_path):
        # 'Revenue' and 'Other expenses' are SOPL-Analysis abstract section
        # headers in the SSM linkbase. Either one is an acceptable target
        # here; we exercise 'Other expenses' because that's the same kind
        # of catch-all-flavoured row the screenshot bug hit.
        output = str(tmp_path / "filled.xlsx")
        payload = json.dumps([{
            "sheet": "SOPL-Analysis-Function",
            "field_label": "Other expenses",
            "col": 2,
            "value": 999_999,
        }])
        result = fill_workbook(self._MPERS_TEMPLATE, output, payload)

        assert result.fields_written == 0, (
            "writer should refuse to write to the 'Other expenses' "
            "abstract section header on the MPERS Group SOPL template. "
            f"Errors: {result.errors}"
        )
        assert result.errors, "expected an actionable error"
        joined = " ".join(result.errors).lower()
        assert "other expenses" in joined or "abstract" in joined or "section header" in joined, (
            f"error must name the offending row. Got: {result.errors}"
        )


class TestDuplicateLabelPrefersLeaf:
    def test_duplicate_label_prefers_leaf_over_header(self, tmp_path):
        """When the same label exists as both a header (row 5) and a leaf
        (row 8), the writer must pick the leaf."""
        template = _make_dup_label_sheet(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        payload = json.dumps([
            {
                "sheet": "SOPL-Analysis-Function",
                "field_label": "Other fee and commission income",
                "col": 2,
                "value": 18282,
            },
        ])

        result = fill_workbook(template, output, payload)

        assert result.success, f"leaf preference should let the write land. Errors: {result.errors}"
        assert result.fields_written == 1

        wb = openpyxl.load_workbook(output)
        ws = wb["SOPL-Analysis-Function"]
        assert ws.cell(row=8, column=2).value == 18282, (
            "value must land on the leaf row 8, not the header row 5"
        )
        assert ws.cell(row=5, column=2).value is None, (
            "header row 5 must remain blank"
        )
        wb.close()
