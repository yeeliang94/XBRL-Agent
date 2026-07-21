"""CodeMode spike seam pins (docs/PLAN-codemode-spike.md, Phase 2).

Four invariants:

1. Flag off ⇒ the optional harness package is NEVER imported (fresh
   subprocess — an in-suite `sys.modules` check goes stale the moment any
   flag-on test imports it) and the factory output is structurally identical
   to today's: same tool roster, no `run_code`, `sequential=False`
   everywhere, no prompt addendum.
2. Flag on ⇒ the sandbox split is exactly the planned one: text tools move
   inside `run_code`; `view_pdf_pages` / `read_template` stay ordinary.
3. The save gate survives inside a script: write→save refused without
   verify; write→verify→save lands.
4. Concurrency + partial failure: a gather over the sequential trio is
   rejected BEFORE any execution; a script that dies after a successful
   write leaves exactly that write applied and the retry turn recovers.

Flag-on tests skip when the optional dependency is absent
(`pip install -r requirements-codemode.txt -c constraints.txt`).
"""
from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
from pydantic_ai.messages import (
    ModelResponse,
    RetryPromptPart,
    TextPart,
    ToolCallPart,
)
from pydantic_ai.models.function import AgentInfo, FunctionModel

from statement_types import StatementType
from tools.verifier import VerificationResult

REPO = Path(__file__).resolve().parent.parent

# Today's exact SOFP tool roster (no scout face refs ⇒ no submit_face_coverage).
BASELINE_TOOLS = {
    "calculator",
    "lookup_definitions",
    "read_template",
    "view_pdf_pages",
    "search_pdf_text",
    "write_facts",
    "verify_totals",
    "save_result",
}

SANDBOXED = {"calculator", "lookup_definitions", "search_pdf_text",
             "write_facts", "verify_totals", "save_result"}


def _make_template(tmp_path: Path) -> str:
    """Minimal SOFP-shaped template: two plain data-entry rows (no header
    fill, no formulas) so fill_workbook's guards all pass."""
    path = tmp_path / "sofp_template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP"
    ws["A5"] = "Cash and bank balances"
    ws["A6"] = "Trade receivables"
    wb.save(path)
    wb.close()
    return str(path)


def _build(tmp_path: Path, model, statement=StatementType.SOFP):
    from extraction.agent import create_extraction_agent

    return create_extraction_agent(
        statement_type=statement,
        variant="CuNonCu",
        pdf_path=str(tmp_path / "missing.pdf"),
        template_path=_make_template(tmp_path),
        model=model,
        output_dir=str(tmp_path),
    )


def _system_text(messages) -> str:
    chunks = []
    for msg in messages:
        for part in getattr(msg, "parts", []):
            if type(part).__name__ in ("SystemPromptPart",):
                chunks.append(str(part.content))
    return "\n".join(chunks)


# ---------------------------------------------------------------------------
# 1. Flag off — no import, identical factory
# ---------------------------------------------------------------------------

_SUBPROCESS_PROBE = """
import os, sys, tempfile
sys.path.insert(0, {repo!r})
os.environ.pop("XBRL_CODE_MODE", None)
import openpyxl
from pydantic_ai.models.test import TestModel
from statement_types import StatementType
from extraction.agent import create_extraction_agent

tmp = tempfile.mkdtemp()
path = os.path.join(tmp, "t.xlsx")
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "SOFP"
ws["A5"] = "Cash"; wb.save(path); wb.close()

def build(stmt):
    return create_extraction_agent(
        statement_type=stmt, variant="CuNonCu",
        pdf_path=os.path.join(tmp, "x.pdf"), template_path=path,
        model=TestModel(), output_dir=tmp,
    )

build(StatementType.SOFP)
assert "pydantic_ai_harness" not in sys.modules, "flag off imported harness"

# Flag on but non-allowlisted statement: still no import.
os.environ["XBRL_CODE_MODE"] = "1"
os.environ["XBRL_CODE_MODE_STATEMENTS"] = "SOFP"
build(StatementType.SOPL)
assert "pydantic_ai_harness" not in sys.modules, "non-allowlisted imported harness"
print("PROBE-OK")
"""


def test_flag_off_harness_never_imported():
    """Fresh subprocess: building the agent flag-off (and flag-on for a
    non-allowlisted statement) must not load the optional package."""
    result = subprocess.run(
        [sys.executable, "-c", _SUBPROCESS_PROBE.format(repo=str(REPO))],
        capture_output=True, text=True, timeout=120,
    )
    assert result.returncode == 0, result.stderr
    assert "PROBE-OK" in result.stdout


def test_flag_off_factory_probe(tmp_path, monkeypatch):
    monkeypatch.delenv("XBRL_CODE_MODE", raising=False)
    seen = {}

    def fake(messages, info: AgentInfo) -> ModelResponse:
        seen["tools"] = {t.name: t for t in info.function_tools}
        seen["system"] = _system_text(messages)
        return ModelResponse(parts=[TextPart("done")])

    agent, deps = _build(tmp_path, FunctionModel(fake))
    agent.run_sync("go", deps=deps)

    assert set(seen["tools"]) == BASELINE_TOOLS
    assert "run_code" not in seen["tools"]
    for name, td in seen["tools"].items():
        assert td.sequential is False, f"{name} unexpectedly sequential flag-off"
    from extraction.code_mode import CODE_MODE_PROMPT_MARKER
    assert CODE_MODE_PROMPT_MARKER not in seen["system"]

    # Schema pin: the `ledgered` wrapper must be invisible to pydantic-ai's
    # schema generation — parameter names and descriptions come from the
    # ORIGINAL functions. (This is what would break if e.g.
    # `from __future__ import annotations` were added to extraction/agent.py:
    # get_type_hints would resolve the wrapper's annotations against
    # code_mode.py's namespace and schema generation would drift/fail.)
    expected_params = {
        "calculator": {"expressions"},
        "lookup_definitions": {"queries"},
        "search_pdf_text": {"queries"},
        "view_pdf_pages": {"pages"},
        "write_facts": {"facts"},
        "verify_totals": set(),
        "save_result": {"fields_json", "acknowledge_unresolved",
                        "unresolved_reason"},
        "read_template": set(),
    }
    for name, params in expected_params.items():
        td = seen["tools"][name]
        schema_props = set((td.parameters_json_schema or {}).get(
            "properties", {}))
        assert schema_props == params, (
            f"{name} schema drifted: {schema_props} != {params}"
        )
        assert td.description, f"{name} lost its docstring description"
    assert seen["tools"]["calculator"].description.startswith(
        "Evaluate arithmetic exactly."
    )


# ---------------------------------------------------------------------------
# 2. Flag on — sandbox split
# ---------------------------------------------------------------------------


def _flag_on(monkeypatch):
    pytest.importorskip("pydantic_ai_harness")
    monkeypatch.setenv("XBRL_CODE_MODE", "1")
    monkeypatch.setenv("XBRL_CODE_MODE_STATEMENTS", "SOFP")


def test_flag_on_sandbox_split(tmp_path, monkeypatch):
    _flag_on(monkeypatch)
    seen = {}

    def fake(messages, info: AgentInfo) -> ModelResponse:
        seen["tools"] = {t.name for t in info.function_tools}
        seen["system"] = _system_text(messages)
        return ModelResponse(parts=[TextPart("done")])

    agent, deps = _build(tmp_path, FunctionModel(fake))
    agent.run_sync("go", deps=deps)

    assert "run_code" in seen["tools"]
    assert seen["tools"] & SANDBOXED == set(), (
        f"sandboxed tools leaked as direct tools: {seen['tools'] & SANDBOXED}"
    )
    # Vision + cached-template tools stay ordinary.
    assert "view_pdf_pages" in seen["tools"]
    assert "read_template" in seen["tools"]
    from extraction.code_mode import CODE_MODE_PROMPT_MARKER
    assert CODE_MODE_PROMPT_MARKER in seen["system"]


# ---------------------------------------------------------------------------
# 3 + 4. Script behavior: save gate, concurrency, partial failure
# ---------------------------------------------------------------------------

_WRITE = ('write_facts(facts=[{"sheet": "SOFP", "row": 5, "col": 2, '
          '"value": %d, "evidence": "Page 1, \'cash\'"}])')

SCRIPT_SAVE_WITHOUT_VERIFY = f"""
r = {_WRITE % 111}
print("write:", r)
s = save_result(fields_json="")
print("save:", s)
"""

SCRIPT_FULL_CHAIN = f"""
r = {_WRITE % 111}
print("write:", r)
v = verify_totals()
print("verify:", v)
s = save_result(fields_json="")
print("save:", s)
"""

SCRIPT_GATHER_SEQUENTIAL = f"""
import asyncio
await asyncio.gather({_WRITE % 111}, verify_totals())
"""

SCRIPT_DIE_AFTER_WRITE = f"""
r = {_WRITE % 111}
print("write ok:", r)
raise ValueError("simulated bug after successful write")
"""

SCRIPT_RECOVER = f"""
r = {_WRITE % 222}
v = verify_totals()
s = save_result(fields_json="")
print(r, v, s)
"""


def _balanced_verify(*args, **kwargs):
    return VerificationResult(is_balanced=True, matches_pdf=None)


def _scripted_model(scripts):
    """FunctionModel that plays the given run_code scripts in order, then
    ends. Also records every RetryPromptPart it is shown."""
    state = {"i": 0, "retries": []}

    def fake(messages, info: AgentInfo) -> ModelResponse:
        for msg in messages:
            for part in getattr(msg, "parts", []):
                if isinstance(part, RetryPromptPart):
                    state["retries"].append(str(part.content))
        if state["i"] < len(scripts):
            code = scripts[state["i"]]
            state["i"] += 1
            return ModelResponse(parts=[ToolCallPart("run_code", {"code": code})])
        return ModelResponse(parts=[TextPart("done")])

    return FunctionModel(fake), state


def test_script_save_without_verify_is_refused(tmp_path, monkeypatch):
    _flag_on(monkeypatch)
    monkeypatch.setattr(
        "extraction.agent._verify_statement_impl", _balanced_verify
    )
    model, state = _scripted_model([SCRIPT_SAVE_WITHOUT_VERIFY])
    agent, deps = _build(tmp_path, model)
    agent.run_sync("go", deps=deps)

    assert deps.result_saved is False
    assert deps.filled_path, "the write itself should have landed"
    # The gate refusal came back through the script's stdout, deduplicated
    # from a crash: the run completed with the refusal text visible.
    assert deps.last_save_error and "refused" in deps.last_save_error


def test_script_full_chain_passes_gate(tmp_path, monkeypatch):
    _flag_on(monkeypatch)
    monkeypatch.setattr(
        "extraction.agent._verify_statement_impl", _balanced_verify
    )
    model, state = _scripted_model([SCRIPT_FULL_CHAIN])
    agent, deps = _build(tmp_path, model)
    agent.run_sync("go", deps=deps)

    assert deps.result_saved is True
    wb = openpyxl.load_workbook(deps.filled_path)
    assert wb["SOFP"]["B5"].value == 111
    wb.close()


def test_gather_over_sequential_trio_rejected_before_execution(
    tmp_path, monkeypatch
):
    _flag_on(monkeypatch)
    model, state = _scripted_model([SCRIPT_GATHER_SEQUENTIAL])
    agent, deps = _build(tmp_path, model)
    agent.run_sync("go", deps=deps)

    # Monty type-checks the script first: sequential tools render as SYNC
    # functions, so gathering them is a type error — rejected before ANY
    # tool executed.
    assert not deps.filled_path, "no write may land from a rejected script"
    assert state["retries"], "the model should have been shown a retry"
    joined = "\n".join(state["retries"])
    assert "error" in joined.lower()


def test_partial_failure_write_survives_and_retry_recovers(
    tmp_path, monkeypatch
):
    _flag_on(monkeypatch)
    monkeypatch.setattr(
        "extraction.agent._verify_statement_impl", _balanced_verify
    )
    model, state = _scripted_model([SCRIPT_DIE_AFTER_WRITE, SCRIPT_RECOVER])
    agent, deps = _build(tmp_path, model)
    agent.run_sync("go", deps=deps)

    # The write from the dead script survived (host-side, by design), the
    # retry's re-write superseded it cleanly, and the workbook is not
    # corrupt: one value, the latest one.
    assert deps.result_saved is True
    wb = openpyxl.load_workbook(deps.filled_path)
    assert wb["SOFP"]["B5"].value == 222
    wb.close()
    # The model was told exactly how far the dead script got.
    joined = "\n".join(state["retries"])
    assert "write ok" in joined, "stdout-before-error must reach the model"
