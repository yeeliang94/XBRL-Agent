"""Phase 6 MPERS hardening — golden E2E regression lock.

Locks the run-#105 failure mode so any future regression on the MPERS
notes pipeline (prompt rendering, suffix normalisation, label catalog,
overlay, SOCIE cross-check) trips this test instead of showing up in
a live run.

The test mocks `_invoke_sub_agent_once` to simulate sub-agents
returning bare MFRS-style labels ("Disclosure of other income") and
asserts that:
  1. The MPERS writer accepts them (Phase 2 suffix normalisation).
  2. They land in the correct MPERS rows.
  3. No `notes12_failures.json` side-log is written for what would
     previously have been silent rejections.

All mocks stop at the sub-coordinator boundary — so the test still
exercises Phases 1-4 (prompt rendering flows through
`create_notes_agent` → the real `render_notes_prompt` with catalog +
overlay) and Phase 2 (the real `write_notes_workbook` with its
suffix-aware normaliser).
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from notes.payload import NotesPayload
from notes_types import NotesTemplateType
from scout.notes_discoverer import NoteInventoryEntry
from scout.infopack import Infopack


@pytest.mark.asyncio
async def test_mpers_list_of_notes_lands_bare_form_labels(tmp_path: Path):
    """The run-#105 golden scenario: sub-agents emit MFRS-style bare
    labels ("Disclosure of other income") — Phase 2 suffix
    normalisation matches them to MPERS `[text block]` rows and ≥ 9
    rows land. Pre-fix this run produced 3 rows."""
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    pdf_path = tmp_path / "dummy.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run={NotesTemplateType.LIST_OF_NOTES},
        filing_level="company",
        filing_standard="mpers",
    )

    # Recreate run-#105's inventory: 14 notes on a Malaysian private-
    # entity filing. Note numbers + titles are the ones from the
    # failing run, so the label choices below mirror what GPT-5.4
    # actually emitted.
    inventory = [
        NoteInventoryEntry(1, "Corporate information", (16, 16)),
        NoteInventoryEntry(2, "Summary of significant accounting policies", (16, 20)),
        NoteInventoryEntry(3, "Cash and short-term funds", (21, 21)),
        NoteInventoryEntry(4, "Financial asset at FVTPL", (22, 22)),
        NoteInventoryEntry(5, "Amount due to immediate holding company", (22, 22)),
        NoteInventoryEntry(6, "Share capital", (22, 22)),
        NoteInventoryEntry(7, "Deferred tax", (22, 22)),
        NoteInventoryEntry(8, "Other income", (23, 23)),
        NoteInventoryEntry(9, "Other expenses", (23, 23)),
        NoteInventoryEntry(10, "Taxation", (23, 23)),
        NoteInventoryEntry(11, "Significant related party transactions", (24, 24)),
        NoteInventoryEntry(12, "Financial risk management", (24, 25)),
        NoteInventoryEntry(13, "Fair value measurement", (26, 26)),
        NoteInventoryEntry(14, "Capital management", (26, 26)),
    ]
    infopack = Infopack(toc_page=1, page_offset=0, notes_inventory=inventory)

    # Per-note label choices mirroring what the real run #105 emitted.
    # Notes 1,2 = corp info / policies → skip (owned by other sheets).
    # Notes 6,11 = share capital / related party → skip (other sheets).
    # Notes 13,14 = concepts genuinely absent from MPERS → must land on
    # the catch-all so nothing drops silently.
    label_for_note: dict[int, str] = {
        3: "Disclosure of cash and cash equivalents",
        4: "Disclosure of financial instruments at fair value through profit or loss",
        5: "Disclosure of trade and other payables",
        7: "Disclosure of deferred tax assets/(liabilities)",
        8: "Disclosure of other income",  # BARE — pre-fix silently rejected
        9: "Disclosure of auditors' remuneration",  # BARE — pre-fix rejected
        10: "Disclosure of income tax expense",  # BARE — pre-fix rejected
        12: "Disclosure of credit risk",  # BARE — pre-fix rejected
        13: "Disclosure of other notes to accounts",  # catch-all
        14: "Disclosure of other notes to accounts",  # catch-all
    }
    skip_notes = {1, 2, 6, 11}

    async def fake_invoke(**kwargs):
        batch = kwargs["batch"]
        payloads = []
        for entry in batch:
            if entry.note_num in skip_notes:
                continue
            label = label_for_note.get(entry.note_num)
            if not label:
                continue
            payloads.append(NotesPayload(
                chosen_row_label=label,
                content=f"Content for note {entry.note_num}",
                evidence=f"p.{entry.page_range[0]}",
                source_pages=[entry.page_range[0]],
                note_num=entry.note_num,
            ))
        return payloads, 0, 0, None

    with patch(
        "notes.listofnotes_subcoordinator._invoke_sub_agent_once",
        side_effect=fake_invoke,
    ):
        result = await run_notes_extraction(config, infopack=infopack)

    r = result.agent_results[0]
    assert r.status == "succeeded", r.error
    assert r.workbook_path and Path(r.workbook_path).exists()

    # --- Content assertions ---
    wb = openpyxl.load_workbook(r.workbook_path)
    ws = wb["Notes-Listofnotes"]
    populated_rows = 0
    for row in range(1, ws.max_row + 1):
        b_val = ws.cell(row=row, column=2).value
        # Skip the template header row that always carries a year range.
        if b_val and str(b_val).strip() and not str(b_val).startswith("01/"):
            populated_rows += 1
    # Run #105 produced 3 rows. Post-fix we expect:
    # - 4 bare-label writes (notes 8, 9, 10, 12)
    # - 4 already-working writes (notes 3, 4, 5, 7)
    # - 1 catch-all merged row for notes 13+14
    # Total: 9 rows minimum. We assert >= 8 to allow 1 row of slack
    # if the fuzzy matcher collapses two into one (e.g. if "trade and
    # other payables" fuzzes to a similar row).
    assert populated_rows >= 8, (
        f"Run-#105 regression: only {populated_rows} rows landed. "
        f"Pre-fix baseline was 3; we expect >= 8 post-Phase-2+3."
    )

    # --- Canary rows: the four bare labels that run-#105 lost ---
    def _find_row(needle: str) -> int | None:
        needle_low = needle.lower()
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=1).value
            if v and needle_low in str(v).lower():
                return row
        return None

    canaries = [
        "Disclosure of other income",
        "Disclosure of auditors' remuneration",
        "Disclosure of income tax expense",
        "Disclosure of credit risk",
    ]
    missing = []
    for label in canaries:
        row = _find_row(label)
        if row is None:
            missing.append(f"{label!r}: row not in template")
            continue
        if not ws.cell(row=row, column=2).value:
            missing.append(f"{label!r}: row {row} empty")
    assert not missing, (
        "Post-fix canary labels failed to land:\n" + "\n".join(missing)
    )

    # --- No failures side-log should have been written ---
    failures_log = tmp_path / "notes_LIST_OF_NOTES_failures.json"
    assert not failures_log.exists(), (
        "Failures side-log appeared — bare-label writes are still "
        f"being rejected. Contents: {failures_log.read_text()[:500]}"
    )
