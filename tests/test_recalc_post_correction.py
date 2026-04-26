"""Peer-review #1 (HIGH, RUN-REVIEW follow-up): recalc must run AFTER
the correction pass, not at merge time.

Failure mode the fix prevents:
- Phase 3 ran recalc at merge time, replacing *Total formulas with
  literal cached values BEFORE the CORRECTION agent ran.
- The correction agent writes to leaves via fill_workbook.py, whose
  formula-cell guard refuses cells whose value starts with "=".
- After recalc, those cells no longer start with "=" — guard is
  silently defeated, AND any leaf-only writes don't propagate to
  the (now-literal) totals, so post-correction cross-checks see
  stale values.

The fix moves recalc to AFTER correction in server.py. This test
pins three properties:

1. `merge()` with explicit `skip_recalc=True` keeps formulas intact
   (the new server.py call site uses this).
2. `fill_workbook` still refuses overwrites of formula cells that
   would have been clobbered by recalc.
3. After a deferred `recalc_workbook` call, totals reflect any
   leaf-level writes the corrector landed (the formula evaluates
   the new leaf values, not the old).
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

from tools.fill_workbook import fill_workbook
from tools.recalc import recalc_workbook
from workbook_merger import merge as merge_workbooks
from statement_types import StatementType


def _build_synthetic_per_statement_workbook(path: Path) -> None:
    """Create a tiny SOFP-shaped workbook with a *Total formula."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("SOFP-CuNonCu")
    ws["A5"] = "Cash"
    ws["B5"] = 100
    ws["A6"] = "Receivables"
    ws["B6"] = 200
    ws["A7"] = "*Total assets"
    ws["B7"] = "=B5+B6"
    wb.save(path)
    wb.close()


def test_merge_with_skip_recalc_preserves_formulas(tmp_path: Path) -> None:
    """The new server.py call site passes skip_recalc=True. The merged
    workbook MUST keep formulas verbatim so the formula-cell guard
    stays effective for any downstream writer."""
    src = tmp_path / "SOFP_filled.xlsx"
    _build_synthetic_per_statement_workbook(src)
    merged = tmp_path / "merged.xlsx"

    merge_workbooks({StatementType.SOFP: str(src)}, str(merged), skip_recalc=True)

    wb = load_workbook(merged, data_only=False)
    assert wb["SOFP-CuNonCu"]["B7"].value == "=B5+B6", (
        "skip_recalc=True must leave formulas as formulas — without "
        "this the post-merge formula-cell guard is silently defeated"
    )
    wb.close()


def test_fill_workbook_refuses_total_cell_after_skip_recalc_merge(
    tmp_path: Path,
) -> None:
    """End-to-end: with skip_recalc=True the *Total cell still
    contains a formula string, so fill_workbook's existing guard
    refuses an overwrite — exactly what would have failed silently
    if recalc had run at merge time."""
    src = tmp_path / "SOFP_filled.xlsx"
    _build_synthetic_per_statement_workbook(src)
    merged = tmp_path / "merged.xlsx"
    merge_workbooks({StatementType.SOFP: str(src)}, str(merged), skip_recalc=True)

    fields = json.dumps([
        {
            "sheet": "SOFP-CuNonCu",
            "row": 7,  # *Total assets — must be refused
            "col": 2,
            "value": 99999,
            "evidence": "synthetic test",
        },
    ])
    result = fill_workbook(
        template_path=str(merged),
        output_path=str(merged),
        fields_json=fields,
    )
    # The write must not have landed
    assert result.fields_written == 0
    assert any("formula" in e.lower() for e in result.errors), (
        f"fill_workbook must refuse formula cells; errors: {result.errors}"
    )


def test_post_correction_recalc_reflects_leaf_writes(tmp_path: Path) -> None:
    """Simulate the full server.py flow:
    1. Merge with skip_recalc=True (formulas survive).
    2. Correction agent edits a LEAF (allowed — col B5 isn't a formula).
    3. Run recalc_workbook AFTER step 2.
    4. The *Total formula's cached value reflects the new leaf.

    Pre-fix, recalc ran in step 1, the formula was already a literal
    100+200=300, the leaf write changed B5 to 500, but the total
    stayed 300 (stale). Post-fix, recalc runs after the leaf write so
    the total becomes 500+200=700.
    """
    src = tmp_path / "SOFP_filled.xlsx"
    _build_synthetic_per_statement_workbook(src)
    merged = tmp_path / "merged.xlsx"
    merge_workbooks({StatementType.SOFP: str(src)}, str(merged), skip_recalc=True)

    # Simulated correction agent write to a LEAF cell
    fields = json.dumps([
        {
            "sheet": "SOFP-CuNonCu",
            "row": 5,  # leaf
            "col": 2,
            "value": 500,
            "evidence": "synthetic correction",
        },
    ])
    fill_result = fill_workbook(
        template_path=str(merged),
        output_path=str(merged),
        fields_json=fields,
    )
    assert fill_result.fields_written == 1

    # Pre-recalc the *Total cell still holds the formula
    pre = load_workbook(merged, data_only=False)
    assert pre["SOFP-CuNonCu"]["B7"].value == "=B5+B6"
    pre.close()

    # Now run the deferred recalc (mirrors server.py post-correction)
    recalc_workbook(merged)

    # Total now reflects the corrected leaf: 500 + 200 = 700
    post = load_workbook(merged, data_only=True)
    assert post["SOFP-CuNonCu"]["B7"].value == 700, (
        "After post-correction recalc, *Total must reflect leaf writes "
        "(500 + 200 = 700). Stale total = pre-fix bug returning."
    )
    post.close()
