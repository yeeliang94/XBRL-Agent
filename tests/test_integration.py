"""Integration test: mocks vision output and asserts workbook creation,
verification, and non-zero token reporting."""

import json
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest


@pytest.fixture
def mock_template(tmp_path):
    template = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Total assets"
    ws["B1"] = 400
    ws["C1"] = 600
    ws["A2"] = "Total equity and liabilities"
    ws["B2"] = 400
    ws["C2"] = 600
    ws["A3"] = "Right-of-use assets"
    ws["B3"] = None
    ws["C3"] = None
    wb.save(str(template))
    return str(template)


@pytest.fixture
def mock_pdf(tmp_path):
    pdf = tmp_path / "test.pdf"
    pdf.write_bytes(b"%PDF-1.4 fake pdf")
    return str(pdf)


def test_integration_mocks_vision_fills_workbook(tmp_path, mock_template, mock_pdf):
    from tools.fill_workbook import fill_workbook
    from tools.verifier import verify_totals
    from tools.template_reader import read_template

    fields_json = json.dumps(
        {
            "fields": [
                {
                    "sheet": "SOFP-CuNonCu",
                    "field_label": "Right-of-use assets",
                    "col": 2,
                    "value": 191518,
                },
            ]
        }
    )

    output = str(tmp_path / "filled.xlsx")
    fill_result = fill_workbook(mock_template, output, fields_json)
    assert fill_result.success
    assert fill_result.fields_written == 1
    assert Path(output).exists()

    wb = openpyxl.load_workbook(output)
    assert wb.active["B3"].value == 191518
    wb.close()


def test_integration_token_tracking(tmp_path):
    from token_tracker import TokenReport, TurnRecord

    report = TokenReport(model="vertex_ai.gemini-3-flash-preview")
    report.add_turn(
        TurnRecord(
            turn=1,
            tool_name="read_template",
            prompt_tokens=8000,
            completion_tokens=500,
            total_tokens=8500,
            thinking_tokens=0,
            cumulative_tokens=8500,
            duration_ms=100,
            timestamp=0,
        )
    )
    report.add_turn(
        TurnRecord(
            turn=2,
            tool_name="view_pdf_pages(1-20)",
            prompt_tokens=25000,
            completion_tokens=2000,
            total_tokens=27000,
            thinking_tokens=1000,
            cumulative_tokens=35500,
            duration_ms=30000,
            timestamp=0,
        )
    )

    # grand_total now includes thinking tokens (peer-review I15):
    # 33000 prompt + 2500 completion + 1000 thinking = 36500
    assert report.grand_total == 36500
    assert report.total_prompt_tokens == 33000
    assert report.total_completion_tokens == 2500
    assert report.total_thinking_tokens == 1000
    assert report.estimate_cost() > 0
    assert "view_pdf_pages" in report.format_table()


def test_integration_full_flow(tmp_path, mock_template):
    from tools.fill_workbook import fill_workbook
    from tools.verifier import verify_totals

    fields_json = json.dumps(
        {
            "fields": [
                {"sheet": "SOFP-CuNonCu", "field_label": "Total assets", "col": 2, "value": 500},
            ]
        }
    )

    output = str(tmp_path / "filled.xlsx")
    fill_result = fill_workbook(mock_template, output, fields_json)
    assert fill_result.success

    pdf_values = {
        "total_assets_cy": 400,
        "total_assets_py": 600,
        "total_equity_liabilities_cy": 400,
        "total_equity_liabilities_py": 600,
    }
    verify_result = verify_totals(output, pdf_values=pdf_values)
    assert not verify_result.matches_pdf
    assert any(
        "total_assets_cy" in m and "500" in m for m in verify_result.mismatches
    )
