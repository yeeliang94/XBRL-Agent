"""Phase 4 — Group filings (4-column / 6-column templates).

Each test pins one step.  The composite key on
``run_concept_facts`` is (run_id, concept_uuid, period, entity_scope)
since Phase 1, so the *dimension* is already wired; Phase 4 proves
end-to-end that:

- a Group filing's 4 facts per concept persist cleanly,
- ``concept_targets`` carries the per-(period, entity_scope) cell
  coordinates that the exporter needs,
- the exporter writes the right value to the right column,
- cross-checks (when Phase-4 routes them) run twice — Group then
  Company — per gotcha #12.

SOCIE is deferred to Phase 5; Group SOCIE is excluded from every
matrix here.
"""
from __future__ import annotations

import json
import shutil
import sqlite3
from pathlib import Path

import openpyxl
import pytest

from concept_model.cascade import recompute_after_turn
from concept_model.exporter import export_run_to_xlsx
from concept_model.importer import import_template
from concept_model.parser import parse_template
from db.schema import init_db


REPO = Path(__file__).resolve().parent.parent
GROUP = REPO / "XBRL-template-MFRS" / "Group"


@pytest.fixture
def group_sofp_db(tmp_path: Path) -> dict:
    """Import the Group SOFP-CuNonCu template into a fresh DB + create
    a run row.  Phase 4 work for SOFP goes through this fixture."""
    fixture = GROUP / "01-SOFP-CuNonCu.xlsx"
    db = tmp_path / "xbrl.db"
    init_db(db)
    tree = parse_template(str(fixture))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    template_id = import_template(db, jp)
    assert template_id == "mfrs-group-sofp-cunoncu-v1"

    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
            "VALUES (?, ?, ?, ?)",
            ("2026-05-21T00:00:00Z", "group.pdf", "running",
             "2026-05-21T00:00:00Z"),
        )
        run_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()
    return {
        "db": db, "run_id": run_id, "template_id": template_id,
        "xlsx_template": fixture,
    }


# -- Step 4.1: entity_scope dimension activated -----------------------


def test_group_run_writes_both_company_and_group_facts(
    group_sofp_db: dict,
) -> None:
    """Write 4 facts for the same LEAF concept across (CY, PY) × (Co,
    Gr).  All 4 persist independently; the composite key allows it."""
    db = group_sofp_db["db"]
    run_id = group_sofp_db["run_id"]

    conn = sqlite3.connect(str(db))
    try:
        leaf = conn.execute(
            "SELECT concept_uuid FROM concept_nodes WHERE template_id = ? "
            "AND kind = 'LEAF' ORDER BY render_sheet, render_row LIMIT 1",
            (group_sofp_db["template_id"],),
        ).fetchone()[0]
    finally:
        conn.close()

    payloads = [
        ("CY", "Company", 100.0),
        ("CY", "Group",   200.0),
        ("PY", "Company", 110.0),
        ("PY", "Group",   220.0),
    ]
    conn = sqlite3.connect(str(db))
    try:
        for period, scope, val in payloads:
            conn.execute(
                "INSERT INTO run_concept_facts(run_id, concept_uuid, "
                "period, entity_scope, value, value_status, source, "
                "updated_at) VALUES (?, ?, ?, ?, ?, 'observed', "
                "'p4', '2026-05Z')",
                (run_id, leaf, period, scope, val),
            )
        conn.commit()
        rows = conn.execute(
            "SELECT period, entity_scope, value FROM run_concept_facts "
            "WHERE concept_uuid = ? ORDER BY period, entity_scope",
            (leaf,),
        ).fetchall()
    finally:
        conn.close()
    assert sorted(rows) == sorted(payloads)


# -- Step 4.2: Group SOFP-CuNonCu import populates concept_targets ----


def test_import_sofp_cunoncu_group_has_correct_render_keys(
    group_sofp_db: dict,
) -> None:
    """The importer (or a Group-aware helper) populates
    ``concept_targets`` with one row per concept × (period,
    entity_scope) — 4 rows per concept on Company+Group templates.

    Column layout per gotcha #12:
      B = Group CY,   C = Group PY
      D = Company CY, E = Company PY
    """
    from concept_model.importer import import_group_targets

    import_group_targets(group_sofp_db["db"], group_sofp_db["template_id"])

    db = group_sofp_db["db"]
    conn = sqlite3.connect(str(db))
    try:
        # Pick a known LEAF concept on the SOFP face — row 10
        # 'Biological assets' is present in both Company and Group
        # templates.  All 4 (period, entity_scope) target rows must
        # exist with the right cell coordinates.
        biological = conn.execute(
            "SELECT concept_uuid FROM concept_nodes "
            "WHERE render_sheet = 'SOFP-CuNonCu' AND render_row = 10 "
            "AND template_id = ?",
            (group_sofp_db["template_id"],),
        ).fetchone()
        assert biological is not None
        uid = biological[0]

        rows = conn.execute(
            "SELECT period, entity_scope, target_col FROM concept_targets "
            "WHERE concept_uuid = ? ORDER BY entity_scope, period",
            (uid,),
        ).fetchall()
    finally:
        conn.close()

    expected = [
        ("CY", "Company", "D"),
        ("PY", "Company", "E"),
        ("CY", "Group",   "B"),
        ("PY", "Group",   "C"),
    ]
    assert sorted(rows) == sorted(expected), rows


# -- Step 4.3: 6-column Excel export -----------------------------------


def test_group_export_populates_all_6_columns(
    group_sofp_db: dict, tmp_path: Path,
) -> None:
    """Seed 4 facts for the same concept; the exporter routes each to
    the right column per ``concept_targets``."""
    from concept_model.importer import import_group_targets

    db = group_sofp_db["db"]
    run_id = group_sofp_db["run_id"]
    import_group_targets(db, group_sofp_db["template_id"])

    conn = sqlite3.connect(str(db))
    try:
        leaf = conn.execute(
            "SELECT concept_uuid, render_sheet, render_row "
            "FROM concept_nodes WHERE render_sheet = 'SOFP-CuNonCu' "
            "AND render_row = 10 AND template_id = ?",
            (group_sofp_db["template_id"],),
        ).fetchone()
    finally:
        conn.close()
    uid, sheet, row_num = leaf[0], leaf[1], int(leaf[2])

    seeded = [
        ("CY", "Group",   1111.0),
        ("PY", "Group",   2222.0),
        ("CY", "Company", 3333.0),
        ("PY", "Company", 4444.0),
    ]
    conn = sqlite3.connect(str(db))
    try:
        for period, scope, val in seeded:
            conn.execute(
                "INSERT INTO run_concept_facts(run_id, concept_uuid, "
                "period, entity_scope, value, value_status, source, "
                "updated_at) VALUES (?, ?, ?, ?, ?, 'observed', 'p4', "
                "'2026-05Z')",
                (run_id, uid, period, scope, val),
            )
        conn.commit()
    finally:
        conn.close()

    work = tmp_path / "group-filled.xlsx"
    shutil.copyfile(group_sofp_db["xlsx_template"], work)
    export_run_to_xlsx(db, run_id, str(work), filing_level="group")

    wb = openpyxl.load_workbook(str(work), data_only=False)
    ws = wb[sheet]
    # B = Group CY, C = Group PY, D = Company CY, E = Company PY
    assert ws[f"B{row_num}"].value == 1111.0, "Group CY at col B"
    assert ws[f"C{row_num}"].value == 2222.0, "Group PY at col C"
    assert ws[f"D{row_num}"].value == 3333.0, "Company CY at col D"
    assert ws[f"E{row_num}"].value == 4444.0, "Company PY at col E"


# -- Step 4.4: cross-checks run twice per gotcha #12 -------------------


def test_group_cross_checks_run_once_per_scope() -> None:
    """A Group filing's cross-checks fire twice — once labelled
    ``Group`` and once labelled ``Company`` — so the Validator tab
    surfaces both result sets distinctly (gotcha #12).

    We pin this by driving ``run_cross_checks_per_scope`` against a
    stub check that records every (filing_level) it was called with.
    """
    from cross_checks.framework import CrossCheck, CrossCheckResult
    from concept_model.group_checks import run_cross_checks_per_scope
    from statement_types import StatementType

    calls: list[str] = []

    class _StubCheck:
        name = "stub_check"
        required_statements: set = set()
        applies_to_standard = {"mfrs", "mpers"}

        def applies_to(self, run_config: dict) -> bool:
            return True

        def run(self, workbook_paths, tolerance, filing_level="company",
                filing_standard="mfrs"):
            calls.append(filing_level)
            return CrossCheckResult(
                name=self.name, status="passed",
                message=f"ran with filing_level={filing_level}",
            )

    results = run_cross_checks_per_scope(
        checks=[_StubCheck()],
        workbook_paths={},
        run_config={
            "statements_to_run": set(),
            "filing_level": "group",
            "filing_standard": "mfrs",
        },
        tolerance=1.0,
    )
    # Stub called twice — once per scope.
    assert sorted(calls) == ["company", "group"]
    # And the results carry a scope label so the UI can render two
    # result sets without conflating them.
    scopes_in_results = sorted({r.message.split("=")[-1] for r in results})
    assert scopes_in_results == ["company", "group"]


def test_company_cross_checks_run_once_only() -> None:
    """Company filings retain the legacy single-pass behaviour — the
    helper short-circuits when ``filing_level != 'group'``."""
    from cross_checks.framework import CrossCheckResult
    from concept_model.group_checks import run_cross_checks_per_scope

    calls: list[str] = []

    class _StubCheck:
        name = "stub"
        required_statements: set = set()
        applies_to_standard = {"mfrs", "mpers"}

        def applies_to(self, run_config: dict) -> bool:
            return True

        def run(self, workbook_paths, tolerance, filing_level="company",
                filing_standard="mfrs"):
            calls.append(filing_level)
            return CrossCheckResult(
                name=self.name, status="passed",
                message=f"ran with filing_level={filing_level}",
            )

    run_cross_checks_per_scope(
        checks=[_StubCheck()],
        workbook_paths={},
        run_config={
            "statements_to_run": set(),
            "filing_level": "company",
            "filing_standard": "mfrs",
        },
        tolerance=1.0,
    )
    assert calls == ["company"]


# -- Steps 4.5-4.10: per-Group-template import + E2E ------------------


PHASE4_GROUP_TEMPLATES = [
    {"id": "sopl_function",   "filename": "03-SOPL-Function.xlsx",
     "template_id": "mfrs-group-sopl-function-v1"},
    {"id": "sopl_nature",     "filename": "04-SOPL-Nature.xlsx",
     "template_id": "mfrs-group-sopl-nature-v1"},
    {"id": "soci_beforetax",  "filename": "05-SOCI-BeforeTax.xlsx",
     "template_id": "mfrs-group-soci-beforetax-v1"},
    {"id": "soci_netoftax",   "filename": "06-SOCI-NetOfTax.xlsx",
     "template_id": "mfrs-group-soci-netoftax-v1"},
    {"id": "socf_indirect",   "filename": "07-SOCF-Indirect.xlsx",
     "template_id": "mfrs-group-socf-indirect-v1"},
    {"id": "socf_direct",     "filename": "08-SOCF-Direct.xlsx",
     "template_id": "mfrs-group-socf-direct-v1"},
]


@pytest.mark.parametrize("spec", PHASE4_GROUP_TEMPLATES,
                         ids=lambda s: s["id"])
def test_import_group_template(tmp_path: Path, spec: dict) -> None:
    """Each Group template imports cleanly and populates
    ``concept_targets`` with 4 rows per LEAF + COMPUTED."""
    from concept_model.importer import import_group_targets

    fixture = GROUP / spec["filename"]
    if not fixture.is_file():
        pytest.skip(f"missing fixture: {fixture}")

    db = tmp_path / "xbrl.db"
    init_db(db)
    tree = parse_template(str(fixture))
    assert tree.template_id == spec["template_id"]
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    template_id = import_template(db, jp)
    rows_written = import_group_targets(db, template_id)
    assert rows_written > 0, "import_group_targets wrote nothing"

    # 4 target rows per non-ABSTRACT concept (excluding SOCIE — none of
    # these templates are SOCIE).
    conn = sqlite3.connect(str(db))
    try:
        n_writable = conn.execute(
            "SELECT COUNT(*) FROM concept_nodes WHERE template_id = ? "
            "AND kind != 'ABSTRACT'", (template_id,),
        ).fetchone()[0]
        n_targets = conn.execute(
            "SELECT COUNT(*) FROM concept_targets WHERE concept_uuid IN "
            "(SELECT concept_uuid FROM concept_nodes WHERE template_id = ?)",
            (template_id,),
        ).fetchone()[0]
    finally:
        conn.close()
    assert n_targets == n_writable * 4


@pytest.mark.parametrize("spec", PHASE4_GROUP_TEMPLATES,
                         ids=lambda s: s["id"])
def test_canonical_e2e_group_template(tmp_path: Path, spec: dict) -> None:
    """For each Group template: import → seed a LEAF in all 4 scopes →
    cascade → export → all 4 columns populated correctly."""
    from concept_model.importer import import_group_targets

    fixture = GROUP / spec["filename"]
    if not fixture.is_file():
        pytest.skip(f"missing fixture: {fixture}")

    db = tmp_path / "xbrl.db"
    init_db(db)
    tree = parse_template(str(fixture))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    template_id = import_template(db, jp)
    import_group_targets(db, template_id)

    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
            "VALUES (?, ?, ?, ?)",
            ("2026-05-21T00:00:00Z", "g.pdf", "running",
             "2026-05-21T00:00:00Z"),
        )
        run_id = cur.lastrowid
        leaf = conn.execute(
            "SELECT concept_uuid, render_sheet, render_row FROM concept_nodes "
            "WHERE template_id = ? AND kind = 'LEAF' "
            "ORDER BY render_sheet, render_row LIMIT 1",
            (template_id,),
        ).fetchone()
        conn.commit()
    finally:
        conn.close()
    assert leaf, f"no LEAF in {template_id}"
    uid, sheet, row_num = leaf[0], leaf[1], int(leaf[2])

    seeded = {
        ("CY", "Group"):   1.0, ("PY", "Group"):   2.0,
        ("CY", "Company"): 3.0, ("PY", "Company"): 4.0,
    }
    conn = sqlite3.connect(str(db))
    try:
        for (period, scope), val in seeded.items():
            conn.execute(
                "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
                "entity_scope, value, value_status, source, updated_at) "
                "VALUES (?, ?, ?, ?, ?, 'observed', 'p4', '2026-05Z')",
                (run_id, uid, period, scope, val),
            )
        conn.commit()
    finally:
        conn.close()
    recompute_after_turn(db, run_id)

    work = tmp_path / f"group-{spec['id']}.xlsx"
    shutil.copyfile(fixture, work)
    export_run_to_xlsx(db, run_id, str(work), filing_level="group")

    wb = openpyxl.load_workbook(str(work), data_only=False)
    ws = wb[sheet]
    assert ws[f"B{row_num}"].value == 1.0, f"Group CY at B (got {ws[f'B{row_num}'].value!r})"
    assert ws[f"C{row_num}"].value == 2.0
    assert ws[f"D{row_num}"].value == 3.0
    assert ws[f"E{row_num}"].value == 4.0


# -- Step 4.11: Phase-4 multi-statement Group E2E ----------------------


def test_e2e_canonical_group_filing_4_statements(
    tmp_path: Path, monkeypatch
) -> None:
    """One DB, 4 Group templates (SOFP+SOPL+SOCI+SOCF), 4 facts per
    concept (CY/PY × Co/Gr), 4 xlsx exports — all green."""
    from concept_model.importer import import_group_targets
    from fastapi.testclient import TestClient

    monkeypatch.setenv("XBRL_CANONICAL_MODE", "1")
    monkeypatch.setenv("XBRL_OUTPUT_DIR", str(tmp_path))
    import importlib
    import server as srv
    importlib.reload(srv)
    db = tmp_path / "xbrl.db"
    srv.AUDIT_DB_PATH = db
    init_db(db)

    statements = [
        ("sofp", "01-SOFP-CuNonCu.xlsx", "mfrs-group-sofp-cunoncu-v1"),
        ("sopl", "03-SOPL-Function.xlsx", "mfrs-group-sopl-function-v1"),
        ("soci", "05-SOCI-BeforeTax.xlsx", "mfrs-group-soci-beforetax-v1"),
        ("socf", "07-SOCF-Indirect.xlsx", "mfrs-group-socf-indirect-v1"),
    ]
    for name, filename, expected_tid in statements:
        fixture = GROUP / filename
        tree = parse_template(str(fixture))
        assert tree.template_id == expected_tid
        jp = tmp_path / f"{name}.json"
        jp.write_text(json.dumps(tree.to_json(), sort_keys=True),
                       encoding="utf-8")
        import_template(db, jp)
        import_group_targets(db, expected_tid)

    conn = sqlite3.connect(str(db))
    try:
        cur = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at) "
            "VALUES (?, ?, ?, ?)",
            ("2026-05-21T00:00:00Z", "group-multi.pdf", "running",
             "2026-05-21T00:00:00Z"),
        )
        run_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    client = TestClient(srv.app)

    # One LEAF per statement, all four scopes.
    leaves: dict[str, tuple[str, str, int]] = {}
    for name, _filename, tid in statements:
        conn = sqlite3.connect(str(db))
        try:
            row = conn.execute(
                "SELECT concept_uuid, render_sheet, render_row "
                "FROM concept_nodes WHERE template_id = ? AND kind = 'LEAF' "
                "ORDER BY render_sheet, render_row LIMIT 1",
                (tid,),
            ).fetchone()
        finally:
            conn.close()
        leaves[name] = (row[0], row[1], int(row[2]))
        uid = row[0]
        for (period, scope), magnitude in [
            (("CY", "Group"),   10.0),
            (("PY", "Group"),   20.0),
            (("CY", "Company"), 30.0),
            (("PY", "Company"), 40.0),
        ]:
            r = client.post(
                f"/api/runs/{run_id}/facts",
                json={
                    "concept_uuid": uid,
                    "period": period,
                    "entity_scope": scope,
                    "value": magnitude,
                    "value_status": "observed",
                    "source": "group e2e",
                },
            )
            assert r.status_code == 200, r.text

    recompute_after_turn(db, run_id)

    # Export each statement and verify the 4-column layout.
    for name, filename, _tid in statements:
        work = tmp_path / f"group-{name}.xlsx"
        shutil.copyfile(GROUP / filename, work)
        export_run_to_xlsx(db, run_id, str(work), filing_level="group")

        uid, sheet, row_num = leaves[name]
        wb = openpyxl.load_workbook(str(work), data_only=False)
        ws = wb[sheet]
        assert ws[f"B{row_num}"].value == 10.0, f"{name} Group CY"
        assert ws[f"C{row_num}"].value == 20.0, f"{name} Group PY"
        assert ws[f"D{row_num}"].value == 30.0, f"{name} Company CY"
        assert ws[f"E{row_num}"].value == 40.0, f"{name} Company PY"


# -- Peer-review #4: Group export rejects unmapped targets -------------


def test_group_export_raises_on_unmapped_target(
    group_sofp_db: dict, tmp_path: Path,
) -> None:
    """If a Group run has a fact whose (concept, period, entity_scope)
    has no row in ``concept_targets``, the exporter must raise rather
    than silently fall back to ``concept_nodes.render_col`` (which is
    always col B and would clobber the Group CY column).

    Reproduces the COALESCE-fallback hazard called out in peer-review
    #4: here we deliberately DON'T call import_group_targets, so no
    target rows exist.
    """
    db = group_sofp_db["db"]
    run_id = group_sofp_db["run_id"]
    # NOTE: import_group_targets intentionally NOT called.

    conn = sqlite3.connect(str(db))
    try:
        leaf = conn.execute(
            "SELECT concept_uuid, render_sheet, render_row FROM concept_nodes "
            "WHERE render_sheet = 'SOFP-CuNonCu' AND render_row = 10 "
            "AND template_id = ?",
            (group_sofp_db["template_id"],),
        ).fetchone()
        uid = leaf[0]
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, source, updated_at) "
            "VALUES (?, ?, 'CY', 'Group', 500.0, 'observed', 'p4', "
            "'2026-05Z')",
            (run_id, uid),
        )
        conn.commit()
    finally:
        conn.close()

    work = tmp_path / "unmapped.xlsx"
    shutil.copyfile(group_sofp_db["xlsx_template"], work)

    with pytest.raises(ValueError, match="no concept_targets"):
        export_run_to_xlsx(db, run_id, str(work), filing_level="group")
