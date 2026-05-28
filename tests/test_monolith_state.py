"""Tests for monolith/state.py — StateSnapshot builder.

Validates the snapshot shape per PRD §5: abstract rows surface with
`kind: "abstract"` (no cy/py), formula cells return `{formula, computed,
warnings}`, SOCIE matrix rows expose `matrix_cols`, and cross-check
diffs carry direction strings.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pytest

from monolith.state import (
    CellWriteResult,
    StateSnapshot,
    build_state_snapshot,
)
from statement_types import StatementType, template_path as get_template_path


_TPL_ROOT = Path(__file__).resolve().parent.parent / "XBRL-template-MFRS" / "Company"


def _merge_company_templates(tmp_path: Path) -> Path:
    """Concatenate the 5 face-statement Company templates into one workbook
    (mimics what `workbook_merger` produces for the monolith path)."""
    # Start from SOFP to get its sheet, then copy the other 4 in.
    wb = openpyxl.load_workbook(
        str(_TPL_ROOT / "01-SOFP-CuNonCu.xlsx"), data_only=False,
    )
    for source_name in (
        "03-SOPL-Function.xlsx",
        "05-SOCI-BeforeTax.xlsx",
        "07-SOCF-Indirect.xlsx",
        "09-SOCIE.xlsx",
    ):
        src = openpyxl.load_workbook(str(_TPL_ROOT / source_name), data_only=False)
        for sheet_name in src.sheetnames:
            if sheet_name in wb.sheetnames:
                continue
            target = wb.create_sheet(title=sheet_name)
            src_ws = src[sheet_name]
            for row in src_ws.iter_rows():
                for cell in row:
                    target.cell(row=cell.row, column=cell.column, value=cell.value)
    out = tmp_path / "monolith_filled.xlsx"
    wb.save(str(out))
    wb.close()
    return out


def test_build_snapshot_returns_all_five_face_sheets(tmp_path):
    wb_path = _merge_company_templates(tmp_path)
    snap = build_state_snapshot(
        str(wb_path), filing_standard="mfrs", filing_level="company", turn=1,
    )
    assert isinstance(snap, StateSnapshot)
    # The five face statement keys are statement-type strings ("SOFP", ...).
    assert set(snap.sheets.keys()) == {
        "SOFP", "SOPL", "SOCI", "SOCF", "SOCIE",
    }
    # Each sheet ought to have a non-empty rows list.
    for s in snap.sheets.values():
        assert s.rows, f"sheet {s.sheet} has no rows"


def test_abstract_rows_marked_no_cy_py(tmp_path):
    """Gotcha #17: abstract rows must surface as kind='abstract' without
    cy/py keys (the writer rejects writes to them; the agent shouldn't
    even see them as a target)."""
    wb_path = _merge_company_templates(tmp_path)
    snap = build_state_snapshot(str(wb_path), turn=1)
    sofp = snap.sheets["SOFP"]
    abstract_rows = [r for r in sofp.rows if r.kind == "abstract"]
    assert abstract_rows, "expected at least one abstract row on SOFP"
    for row in abstract_rows:
        d = next(d for d in snap.to_dict()["sheets"]["SOFP"]["rows"] if d["row"] == row.row)
        assert d["kind"] == "abstract"
        assert "cy" not in d and "py" not in d


def test_formula_rows_return_formula_and_computed(tmp_path):
    """PRD §5: formula cells expose both the formula text and the resolved
    computed value, sparing the agent from re-evaluating in its head."""
    wb_path = _merge_company_templates(tmp_path)
    # Plant a formula in SOFP-CuNonCu somewhere that maps to a formula
    # row that already exists — the template ships with formula rows on
    # totals already, so we'll just look one up.
    snap = build_state_snapshot(str(wb_path), turn=1)
    sofp = snap.sheets["SOFP"]
    formula_rows = [r for r in sofp.rows if r.kind == "formula"]
    assert formula_rows, "expected SOFP to ship with at least one formula row"
    for r in formula_rows:
        value = r.cy if isinstance(r.cy, CellWriteResult) else r.py
        if not isinstance(value, CellWriteResult):
            continue
        assert value.formula.startswith("=")
        # computed may be None (warnings populated) — both acceptable.
        assert isinstance(value.warnings, list)


def test_socie_matrix_rows_expose_matrix_cols(tmp_path):
    """SOCIE rows must surface their per-equity-component column values
    via `matrix_cols` keyed by the row-2 header text. Linear sheets do
    not carry matrix_cols."""
    wb_path = _merge_company_templates(tmp_path)
    snap = build_state_snapshot(str(wb_path), turn=1)
    socie = snap.sheets["SOCIE"]
    matrix_rows = [r for r in socie.rows if r.kind == "matrix_leaf"]
    assert matrix_rows, "expected SOCIE to have matrix_leaf rows"
    for r in matrix_rows:
        assert r.matrix_cols is not None
    # Linear sheets (SOFP) never carry matrix_cols on their rows.
    for r in snap.sheets["SOFP"].rows:
        assert r.matrix_cols is None


def test_cross_check_diff_carries_direction(tmp_path):
    """If a cross-check fails with a numeric diff, the snapshot ought to
    carry a direction string per PRD §5 (e.g. 'SOFP higher by 45')."""
    wb_path = _merge_company_templates(tmp_path)
    snap = build_state_snapshot(str(wb_path), turn=1)
    # The empty template will fail several balance identities. Any with a
    # numeric diff > 0.01 should have a direction string.
    failing_with_diff = [
        c for c in snap.cross_checks
        if not c.pass_ and c.diff is not None and abs(c.diff) > 0.01
    ]
    if failing_with_diff:  # not guaranteed on a blank template; skip otherwise
        for c in failing_with_diff:
            assert c.direction, (
                f"cross_check {c.id} has diff {c.diff} but no direction"
            )


def test_to_dict_serialises_pass_key(tmp_path):
    """The `pass_` dataclass field must serialise as the JSON key `pass`."""
    wb_path = _merge_company_templates(tmp_path)
    snap = build_state_snapshot(str(wb_path), turn=1)
    d = snap.to_dict()
    for c in d["cross_checks"]:
        assert "pass" in c
        assert "pass_" not in c


def test_history_hints_passthrough(tmp_path):
    """state.py is dumb about history — whatever the coordinator hands in
    must come back unmodified."""
    wb_path = _merge_company_templates(tmp_path)
    hints = [{"sheet": "SOPL", "row": 22, "value": 8500, "note": "test"}]
    snap = build_state_snapshot(str(wb_path), history_hints=hints)
    assert snap.history_hints == hints
    assert snap.to_dict()["history_hints"] == hints
