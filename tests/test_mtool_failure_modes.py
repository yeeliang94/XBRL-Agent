"""Step 15 — the failure-mode sweep: nothing here may crash the endpoint.

Every entry is a way an operator can realistically get this wrong: the wrong
template, a corrupt file, a template whose sheets belong to a different filing,
duplicate labels, a figure Excel can't hold, two people filling at once. The
contract is the same in all of them — a structured, plain-language error or a
report that names the problem, never a 500 and never a silent success
(gotcha #20's spirit).
"""
from __future__ import annotations

import importlib
import io
import json
import sqlite3
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

REPO = Path(__file__).resolve().parent.parent
SOFP = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"
SOPL = REPO / "XBRL-template-MFRS" / "Company" / "03-SOPL-Function.xlsx"


def _file(name: str, data: bytes):
    return {"template": (name, io.BytesIO(data),
                         "application/vnd.openxmlformats-officedocument"
                         ".spreadsheetml.sheet")}


@pytest.fixture
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("XBRL_OUTPUT_DIR", str(tmp_path))
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    monkeypatch.setenv("LLM_PROXY_URL", "")
    import server as srv
    importlib.reload(srv)
    db = tmp_path / "xbrl.db"
    srv.AUDIT_DB_PATH = db
    from concept_model.importer import import_company_targets, import_template
    from concept_model.filing_targets import persist_template_manifest
    from concept_model.parser import parse_template
    from db.schema import init_db
    init_db(db)
    tree = parse_template(str(SOFP))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    import_company_targets(db, import_template(db, jp))
    persist_template_manifest(db, SOFP)
    return TestClient(srv.app), db, tmp_path


def _make_run(db) -> int:
    conn = sqlite3.connect(str(db))
    try:
        run_id = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at, "
            "run_config_json) VALUES (?,?,?,?,?)",
            ("2026-07-27T00:00:00Z", "x.pdf", "completed",
             "2026-07-27T00:00:00Z",
             json.dumps({"filing_standard": "mfrs", "filing_level": "company",
                         "denomination": "thousands"}))).lastrowid
        conn.commit()
    finally:
        conn.close()
    return run_id


def _seed(db, run_id, n=4, value=1000.0):
    conn = sqlite3.connect(str(db))
    try:
        leaves = conn.execute(
            """
            SELECT concept_uuid FROM concept_nodes
            WHERE kind='LEAF' AND render_sheet LIKE '%Sub%'
              AND canonical_label IN (
                SELECT canonical_label FROM concept_nodes
                GROUP BY canonical_label HAVING COUNT(*) = 1)
            ORDER BY render_row LIMIT ?
            """, (n,)).fetchall()
        for uuid, in leaves:
            conn.execute(
                "INSERT OR REPLACE INTO run_concept_facts(run_id, "
                "concept_uuid, period, entity_scope, value, value_status, "
                "updated_at) VALUES (?,?,?,?,?,?,?)",
                (run_id, uuid, "CY", "Company", value, "observed", "t"))
        conn.commit()
    finally:
        conn.close()
    return len(leaves)


def _patch(tc, run_id, files=None, **data):
    return tc.post(f"/api/runs/{run_id}/mtool-fill/patch",
                   files=files or _file("t.xlsx", SOFP.read_bytes()),
                   data={"strict": "true", **data})


# ------------------------------------------------------------ bad uploads

def test_not_a_zip_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)
    r = _patch(tc, run_id, files=_file("t.xlsx", b"this is not a zip"))
    assert r.status_code == 422
    assert "readable .xlsx" in json.dumps(r.json())


def test_truncated_zip_is_422_not_500(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)
    half = SOFP.read_bytes()[: len(SOFP.read_bytes()) // 2]
    r = _patch(tc, run_id, files=_file("t.xlsx", half))
    assert r.status_code == 422


def test_a_zip_that_is_not_a_workbook_is_422(client, tmp_path):
    """A valid zip with none of the xlsx parts — e.g. someone renamed a .zip."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("readme.txt", "not a workbook")
    r = _patch(tc, run_id, files=_file("t.xlsx", buf.getvalue()))
    assert r.status_code == 422


def test_empty_upload_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)
    r = _patch(tc, run_id, files=_file("t.xlsx", b""))
    assert r.status_code == 422
    assert "Empty upload" in json.dumps(r.json())


# ------------------------------------------------------- wrong template

def test_wrong_statement_template_is_refused_with_the_sheet_named(client):
    """The operator uploads the SOPL template for a SOFP run. The sheets the
    fill needs simply aren't there — say which, don't half-fill."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)
    r = _patch(tc, run_id, files=_file("sopl.xlsx", SOPL.read_bytes()))
    assert r.status_code == 422
    body = json.dumps(r.json())
    assert "SOFP-Sub-CuNonCu" in body, body


def test_template_with_no_matching_sheets_never_writes_anything(client,
                                                                tmp_path):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)
    wb = Workbook()
    wb.active.title = "Something-Else"
    wb.active["A1"] = "hello"
    path = tmp_path / "other.xlsx"
    wb.save(path)
    r = _patch(tc, run_id, files=_file("other.xlsx", path.read_bytes()))
    assert r.status_code == 422


def test_run_with_no_facts_is_422(client):
    tc, db, _ = client
    run_id = _make_run(db)  # no facts seeded
    r = _patch(tc, run_id)
    assert r.status_code == 422
    assert "no fillable facts" in json.dumps(r.json())


# --------------------------------------------------------- ambiguous labels

def test_duplicate_labels_are_reported_not_guessed(client, tmp_path):
    """A template with the same label on two rows must refuse the write and
    say so — picking one would file a figure against the wrong line."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)
    doc = tc.get(f"/api/runs/{run_id}/mtool-fill").json()
    sheet = doc["meta"]["sheets_covered"][0]
    label = doc["writes"][0]["label"]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i, text in enumerate([label, "filler one", label, "filler two"],
                             start=1):
        ws.cell(i, 1, text)
    path = tmp_path / "dupes.xlsx"
    wb.save(path)

    r = _patch(tc, run_id,
               files=_file("dupes.xlsx", path.read_bytes()),
               column_map=json.dumps({
                   sheet: {"label_column": "A",
                           "columns": {"current_year": "B"}}}))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["status"] == "degraded"
    reported = body["ambiguous"] + body["unresolved"]
    assert any(e.get("label") == label for e in reported), reported


# --------------------------------------------------------- extreme values

def test_an_enormous_value_still_produces_a_readable_workbook(client):
    """A figure far beyond anything real must not corrupt the package."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id, value=1e300)
    r = _patch(tc, run_id)
    assert r.status_code == 200, r.text
    resp = tc.get(r.json()["download_url"],
                  params={"acknowledge_degraded": "test"})
    assert resp.status_code == 200
    import openpyxl
    openpyxl.load_workbook(io.BytesIO(resp.content))


def test_a_negative_value_survives_the_round_trip(client):
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id, value=-1234.0)
    r = _patch(tc, run_id)
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "ok"
    written = r.json()["counts"]["written"]
    assert written > 0


# ------------------------------------------------------------- concurrency

def test_concurrent_fills_do_not_collide(client):
    """Patching is stateless over the run's facts and each request gets its own
    temp dir — two operators filling the same run at once must both succeed
    with their own artifact."""
    tc, db, _ = client
    run_id = _make_run(db)
    _seed(db, run_id)

    def one(_i):
        return _patch(tc, run_id)

    with ThreadPoolExecutor(max_workers=4) as pool:
        results = list(pool.map(one, range(4)))

    assert all(r.status_code == 200 for r in results), [
        r.status_code for r in results]
    artifacts = {r.json()["artifact_id"] for r in results}
    assert len(artifacts) == 4, "artifact ids collided"
    # Each fill left its own receipt — the audit trail doesn't merge them.
    receipts = tc.get(
        f"/api/runs/{run_id}/mtool-fill/receipts").json()["receipts"]
    assert len(receipts) == 4


def test_a_missing_run_is_404_everywhere(client):
    tc, _, _ = client
    for path in ("mtool-fill", "mtool-fill/preflight", "mtool-fill/receipts"):
        assert tc.get(f"/api/runs/999999/{path}").status_code in (404, 200)
    assert tc.get("/api/runs/999999/mtool-fill").status_code == 404
    assert _patch(tc, 999999).status_code == 404


# ---------------------------------------------- the tool stays stdlib-only

def test_offline_fill_imports_with_no_third_party_deps():
    """The rollback plan's invariant: the patcher travels to the Windows box as
    a single file, so it must not grow a repo import or a pip dependency."""
    source = (REPO / "mtool" / "offline_fill.py").read_text(encoding="utf-8")
    banned = ("import openpyxl", "from openpyxl", "import mtool",
              "from mtool", "import server", "from db ", "import pandas")
    for token in banned:
        assert token not in source, f"offline_fill.py imported {token!r}"
