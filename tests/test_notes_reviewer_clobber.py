"""Phase 0 / Step 0 — pin the clobber bug that mandates DB-only writes.

The notes reviewer must mutate ``notes_cells`` (the canonical store), NOT the
merged xlsx. This test demonstrates *why*: the download/finalize path overlays
``notes_cells`` onto the workbook (``overlay_notes_cells_into_workbook``), so a
prose edit made directly in the xlsx — exactly what the legacy validator's
``rewrite_cell`` did — is silently overwritten by the (unchanged) DB row on the
next overlay. Any reviewer that wrote only the xlsx would lose every fix on
download.

Pins the direction taken in docs/PLAN.md; remove/repoint once Phase 2 routes
all reviewer writes through ``notes_cells``.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from db import repository as repo
from db.schema import init_db
from notes.persistence import (
    overlay_notes_cells_into_workbook,
    persist_notes_cells,
)
from notes.validator_agent import NotesValidatorAgentDeps, _rewrite_cell_impl

_SHEET = "Notes-Listofnotes"
_ROW = 49


@pytest.fixture()
def db_path(tmp_path: Path) -> Path:
    p = tmp_path / "xbrl.db"
    init_db(p)
    return p


def test_xlsx_edit_is_clobbered_by_notes_cells_overlay(
    db_path: Path, tmp_path: Path,
) -> None:
    run_id = repo_seed_run(db_path)

    # 1. The DB carries the original extraction prose for this cell.
    persist_notes_cells(
        db_path=str(db_path),
        run_id=run_id,
        sheet_name=_SHEET,
        cells_written=[{
            "sheet": _SHEET,
            "row": _ROW,
            "label": "Disclosure of fair value information",
            "html": "<p>DB ORIGINAL — fair value of investment property</p>",
            "evidence": "Page 19",
            "source_pages": [19],
        }],
    )

    # 2. A merged workbook the (legacy) validator edits IN THE XLSX via
    #    rewrite_cell — the exact code path the old notes validator used.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET
    ws.cell(row=_ROW, column=1, value="Disclosure of fair value information")
    ws.cell(row=_ROW, column=2, value="<p>DB ORIGINAL ...</p>")
    merged = tmp_path / "merged.xlsx"
    wb.save(str(merged))

    deps = NotesValidatorAgentDeps(
        merged_workbook_path=str(merged),
        pdf_path="/tmp/x.pdf",
        sidecar_paths=[],
        filing_level="company",
        filing_standard="mfrs",
        output_dir=str(tmp_path),
        model=None,
    )
    msg = _rewrite_cell_impl(
        merged_workbook_path=str(merged),
        filing_level="company",
        sheet=_SHEET,
        row=_ROW,
        col=2,
        content="VALIDATOR EDIT — should win but will not",
        evidence="Page 99",
        deps=deps,
    )
    assert msg.startswith("OK"), msg

    # Sanity: the edit really did land in the xlsx.
    wb2 = openpyxl.load_workbook(str(merged))
    assert wb2[_SHEET].cell(row=_ROW, column=2).value.startswith("VALIDATOR EDIT")
    wb2.close()

    # 3. The download/finalize overlay runs — and clobbers the xlsx edit with
    #    the (unchanged) DB prose. THIS is the bug: the validator's fix is lost.
    overlaid = overlay_notes_cells_into_workbook(
        xlsx_path=merged, run_id=run_id, db_path=str(db_path),
    )
    wb3 = openpyxl.load_workbook(str(overlaid))
    final_value = wb3[_SHEET].cell(row=_ROW, column=2).value
    wb3.close()

    # The DB prose wins; the validator's xlsx-only edit is gone.
    assert "DB ORIGINAL" in final_value
    assert "VALIDATOR EDIT" not in final_value


def repo_seed_run(db_path: Path) -> int:
    with repo.db_session(db_path) as conn:
        return repo.create_run(
            conn, "sample.pdf", session_id="sess", output_dir="/tmp/sess",
        )
