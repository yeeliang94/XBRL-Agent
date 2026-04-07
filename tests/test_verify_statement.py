"""Tests for parametric verifier (Phase 1, Step 1.3)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from statement_types import StatementType
from tools.verifier import verify_statement, verify_totals


def _make_sofp_template(tmp_path: Path) -> Path:
    """Minimal balanced SOFP fixture."""
    path = tmp_path / "sofp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP-CuNonCu"
    ws["A1"] = "Total assets"
    ws["B1"] = 400
    ws["C1"] = 600
    ws["A2"] = "Total equity and liabilities"
    ws["B2"] = 400
    ws["C2"] = 600
    wb.save(path)
    return path


def _make_generic_template(tmp_path: Path, name: str) -> Path:
    """Minimal workbook so the file-exists check passes."""
    path = tmp_path / f"{name}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name
    ws["A1"] = "anything"
    wb.save(path)
    return path


def test_verify_statement_sofp_delegates_to_verify_totals(tmp_path: Path) -> None:
    path = _make_sofp_template(tmp_path)
    result = verify_statement(str(path), StatementType.SOFP, "CuNonCu")
    expected = verify_totals(str(path))
    assert result.is_balanced is True
    assert result.is_balanced == expected.is_balanced
    assert result.computed_totals == expected.computed_totals


@pytest.mark.parametrize("statement,variant", [
    (StatementType.SOPL, "Function"),
    (StatementType.SOCI, "BeforeTax"),
    (StatementType.SOCF, "Indirect"),
    (StatementType.SOCIE, "Default"),
])
def test_verify_statement_non_sofp_fails_on_empty_workbook(
    tmp_path: Path, statement: StatementType, variant: str
) -> None:
    path = _make_generic_template(tmp_path, statement.value)
    result = verify_statement(str(path), statement, variant)
    # All statement types now have real balance checks. With an empty workbook
    # (no matching data rows), checks fail closed — required labels are missing.
    assert result.is_balanced is False
    assert result.matches_pdf is None
    assert len(result.mismatches) > 0


def test_verify_statement_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        verify_statement(str(tmp_path / "nope.xlsx"), StatementType.SOPL, "Function")


def test_verify_totals_backward_compatible(tmp_path: Path) -> None:
    """Legacy verify_totals() still works exactly as before."""
    path = _make_sofp_template(tmp_path)
    result = verify_totals(str(path))
    assert result.is_balanced is True
    assert result.computed_totals["total_assets_cy"] == 400
