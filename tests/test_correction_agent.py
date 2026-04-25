"""Tests for the cross-check correction agent (Phase 3)."""
from __future__ import annotations

import re
from pathlib import Path

import pytest
from pydantic_ai.models.test import TestModel

from cross_checks.framework import CrossCheckResult
from statement_types import StatementType


def _flatten(text: str) -> str:
    return re.sub(r"\s+", " ", text).lower()


class TestCorrectionAgentFactory:
    """Step 3.1: factory must produce an agent with the expected toolset."""

    def test_correction_agent_factory_returns_agent_with_expected_tools(self):
        from correction.agent import create_correction_agent, CorrectionAgentDeps

        failed = [
            CrossCheckResult(
                name="SOFP balance",
                status="failed",
                expected=1000.0,
                actual=950.0,
                diff=-50.0,
                tolerance=1.0,
                message="SOFP Total assets != Total equity + liabilities",
            ),
        ]
        agent, deps = create_correction_agent(
            merged_workbook_path="/tmp/merged.xlsx",
            pdf_path="/tmp/x.pdf",
            failed_checks=failed,
            infopack=None,
            filing_level="company",
            filing_standard="mfrs",
            model=TestModel(),
            output_dir="/tmp/out",
            statements_to_run={StatementType.SOFP},
        )
        assert agent is not None
        assert isinstance(deps, CorrectionAgentDeps)

        # Tool registration: all four tool names must be present on the agent.
        # pydantic_ai >= 1.77 exposes registered tools via `toolsets`; fall
        # back to probing private attributes only if the public path changes.
        tool_names = _agent_tool_names(agent)
        assert "view_pdf_pages" in tool_names
        assert "inspect_workbook" in tool_names
        assert "fill_workbook" in tool_names
        assert "verify_totals" in tool_names
        assert "run_cross_checks" in tool_names

    def test_deps_carry_failure_context(self):
        from correction.agent import create_correction_agent, CorrectionAgentDeps

        failed = [
            CrossCheckResult(
                name="SOPL->SOCIE profit",
                status="failed",
                message="Profit does not match SOCIE profit row",
            ),
        ]
        _, deps = create_correction_agent(
            merged_workbook_path="/tmp/m.xlsx",
            pdf_path="/tmp/x.pdf",
            failed_checks=failed,
            infopack=None,
            filing_level="group",
            filing_standard="mpers",
            model=TestModel(),
            output_dir="/tmp/out",
            statements_to_run={StatementType.SOPL, StatementType.SOCIE},
        )
        assert deps.failed_checks[0].name == "SOPL->SOCIE profit"
        assert deps.filing_level == "group"
        assert deps.filing_standard == "mpers"
        assert deps.statements_to_run == {StatementType.SOPL, StatementType.SOCIE}

    def test_correction_prompt_has_sign_repair_rules(self):
        body = (Path(__file__).resolve().parent.parent / "prompts" / "correction.md").read_text(
            encoding="utf-8"
        )
        flat = _flatten(body)
        assert "inspect_workbook" in flat
        assert "sign-convention repair rules" in flat
        assert "foreign exchange loss" in flat
        assert "dividends as a positive magnitude" in flat
        assert "nearest subtotal formula subtracts a row" in flat

    def test_inspect_workbook_surfaces_formula_sign(self, tmp_path):
        import openpyxl
        from correction.agent import create_correction_agent

        workbook = tmp_path / "merged.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SOCIE"
        ws["A16"] = "Dividends paid"
        ws["B16"] = 100
        ws["A23"] = "*Total increase (decrease) in equity"
        ws["B23"] = "=1*B12+-1*B16"
        wb.save(workbook)

        agent, deps = create_correction_agent(
            merged_workbook_path=str(workbook),
            pdf_path="/tmp/x.pdf",
            failed_checks=[],
            infopack=None,
            filing_level="company",
            filing_standard="mpers",
            model=TestModel(),
            output_dir=str(tmp_path),
            statements_to_run={StatementType.SOCIE},
        )

        tool_fn = _tool_fn(agent, "inspect_workbook")

        class _Ctx:
            def __init__(self, deps):
                self.deps = deps

        result = tool_fn(_Ctx(deps), '{"sheet":"SOCIE","labels":["Dividends paid"],"context_rows":1}')
        assert "row 16" in result
        assert "B16=100" in result
        assert "B23: =1*B12+-1*B16" in result

    def test_inspect_workbook_handles_invalid_scalars(self, tmp_path):
        """Peer-review M1: a model emitting `"context_rows": "nearby"` or
        `"max_col": null` must not crash the tool. Bad scalars fall back
        to documented defaults and the workbook inspection still runs."""
        import openpyxl
        from correction.agent import create_correction_agent

        workbook = tmp_path / "merged.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SOCIE"
        ws["A16"] = "Dividends paid"
        ws["B16"] = 100
        wb.save(workbook)

        agent, deps = create_correction_agent(
            merged_workbook_path=str(workbook),
            pdf_path="/tmp/x.pdf",
            failed_checks=[],
            infopack=None,
            filing_level="company",
            filing_standard="mpers",
            model=TestModel(),
            output_dir=str(tmp_path),
            statements_to_run={StatementType.SOCIE},
        )

        tool_fn = _tool_fn(agent, "inspect_workbook")

        class _Ctx:
            def __init__(self, deps):
                self.deps = deps

        # Garbage scalars that would have raised before the _coerce_int
        # wrapper landed.
        result = tool_fn(
            _Ctx(deps),
            '{"sheet":"SOCIE","labels":["Dividends paid"],'
            '"context_rows":"nearby","max_col":"all"}',
        )
        # The tool returns a string and the row is still found via the
        # default context_rows + max_col.
        assert isinstance(result, str)
        assert "row 16" in result
        assert "B16=100" in result


class TestCorrectionPassHelper:
    """Step 3.2: _run_correction_pass emits SSE events under CORRECTION
    agent id and short-circuits when no checks have failed."""

    @pytest.mark.asyncio
    async def test_correction_pass_noop_when_no_failures(self):
        import asyncio
        from server import _run_correction_pass

        queue: asyncio.Queue = asyncio.Queue()
        outcome = await _run_correction_pass(
            failed_checks=[],
            merged_workbook_path="/tmp/m.xlsx",
            pdf_path="/tmp/x.pdf",
            infopack=None,
            filing_level="company",
            filing_standard="mfrs",
            model=TestModel(),
            output_dir="/tmp/o",
            event_queue=queue,
        )
        assert outcome["invoked"] is False
        # Nothing should be emitted on the short-circuit path.
        assert queue.empty()

    @pytest.mark.asyncio
    async def test_correction_pass_invokes_agent_on_failure(self, tmp_path, monkeypatch):
        """When failed_checks is non-empty, the helper must create the
        correction agent and emit a complete event under CORRECTION."""
        import asyncio
        from server import _run_correction_pass, CORRECTION_AGENT_ID
        from correction import agent as correction_agent_mod

        created: dict = {}

        def _fake_create(*args, **kwargs):
            created["called"] = True
            created["failed_checks"] = kwargs.get("failed_checks")
            created["filing_standard"] = kwargs.get("filing_standard")

            class _Deps:
                writes_performed = 0

            class _AgentRunCtx:
                def __init__(self):
                    self.ctx = object()

                def __aiter__(self):
                    async def _empty():
                        if False:
                            yield None  # pragma: no cover — empty async gen
                    return _empty()

            class _Agent:
                def iter(self, prompt, deps):
                    class _CM:
                        async def __aenter__(self_inner):
                            return _AgentRunCtx()
                        async def __aexit__(self_inner, *a):
                            return False
                    return _CM()

            return _Agent(), _Deps()

        monkeypatch.setattr(correction_agent_mod, "create_correction_agent", _fake_create)

        queue: asyncio.Queue = asyncio.Queue()
        failed = [
            CrossCheckResult(name="SOPL->SOCIE profit", status="failed",
                             message="profit != socie profit"),
        ]
        outcome = await _run_correction_pass(
            failed_checks=failed,
            merged_workbook_path=str(tmp_path / "m.xlsx"),
            pdf_path=str(tmp_path / "x.pdf"),
            infopack=None,
            filing_level="group",
            filing_standard="mpers",
            model=TestModel(),
            output_dir=str(tmp_path),
            event_queue=queue,
        )
        assert outcome["invoked"] is True
        assert created["called"] is True
        # The helper threads filing_standard through to the factory.
        assert created["filing_standard"] == "mpers"
        assert created["failed_checks"] == failed

        # At least one status + one complete event must be emitted under
        # CORRECTION so the frontend routes them to the pseudo-agent tab.
        events = []
        while not queue.empty():
            events.append(queue.get_nowait())
        assert events, "expected correction-pass to emit at least one event"
        for ev in events:
            assert ev["data"]["agent_id"] == CORRECTION_AGENT_ID
        kinds = {ev["event"] for ev in events}
        assert "status" in kinds
        assert "complete" in kinds


class TestSseEventContract:
    """Phase 7.1: the correction + notes-validator pseudo-agents must emit
    under stable agent_id strings that the frontend's appReducer already
    routes. No new event types — just verifies the IDs are exported."""

    def test_sse_events_include_correction_agent_id(self):
        from server import CORRECTION_AGENT_ID, NOTES_VALIDATOR_AGENT_ID
        assert CORRECTION_AGENT_ID == "CORRECTION"
        assert NOTES_VALIDATOR_AGENT_ID == "NOTES_VALIDATOR"


class TestDrainWhileRunningContract:
    """Peer-review C1 regression: pseudo-agent events enqueued by the
    helper must not be stranded. We test the `_drain_while_running`
    pattern directly — it's the linchpin of the C1 fix."""

    @pytest.mark.asyncio
    async def test_correction_pass_events_drained_after_helper_runs(
        self, monkeypatch,
    ):
        """Simulate the `asyncio.create_task(_run_correction_pass) +
        _drain_while_running` pattern that server.py now uses. Events the
        helper pushes while running (or just before completing) must all
        land in the drained list."""
        import asyncio
        from correction import agent as correction_agent_mod
        from server import _run_correction_pass

        # Fake agent that pushes exactly three events via the helper's
        # internal _emit, then exits cleanly. The helper itself pushes
        # a "status" before iter and a "complete" after — for a total of
        # 5+ events the drain loop must deliver.
        def _fake_create(*args, **kwargs):
            class _Deps:
                writes_performed = 0

            class _Run:
                async def __aenter__(self_inner):
                    return _AgentRun()
                async def __aexit__(self_inner, *a):
                    return False

            class _AgentRun:
                ctx = object()
                def __aiter__(self):
                    async def _empty():
                        if False:
                            yield None  # pragma: no cover
                    return _empty()

            class _Agent:
                def iter(self, prompt, deps):
                    return _Run()

            return _Agent(), _Deps()

        monkeypatch.setattr(
            correction_agent_mod, "create_correction_agent", _fake_create,
        )

        queue: asyncio.Queue = asyncio.Queue()
        failed = [CrossCheckResult(name="X", status="failed", message="y")]

        task = asyncio.create_task(_run_correction_pass(
            failed_checks=failed,
            merged_workbook_path="/tmp/m.xlsx",
            pdf_path="/tmp/x.pdf",
            infopack=None,
            filing_level="company",
            filing_standard="mfrs",
            model=TestModel(),
            output_dir="/tmp/o",
            event_queue=queue,
        ))

        # Replicate server.py's drain-while-running pattern against the
        # real queue and task.
        drained: list[dict] = []
        while not task.done():
            try:
                event = await asyncio.wait_for(queue.get(), timeout=0.3)
            except asyncio.TimeoutError:
                continue
            if event is None:
                continue
            drained.append(event)
        while not queue.empty():
            event = queue.get_nowait()
            if event is None:
                continue
            drained.append(event)
        await task

        # At minimum a status + complete pair should have been drained.
        kinds = {e["event"] for e in drained}
        assert "status" in kinds, drained
        assert "complete" in kinds, drained
        # All pseudo-agent events carry agent_id=CORRECTION.
        for e in drained:
            assert e["data"]["agent_id"] == "CORRECTION"


def _agent_tool_names(agent) -> set[str]:
    """Collect registered tool names from a pydantic-ai Agent.

    Prefers the public `toolsets` attribute introduced in 1.77. Falls
    back to `_function_tools` only for defence-in-depth — if the API
    shifts again we still surface a useful failure rather than silently
    reporting an empty set.
    """
    names: set[str] = set()
    toolsets = getattr(agent, "toolsets", None) or []
    for ts in toolsets:
        tools = getattr(ts, "tools", None) or {}
        for tname in tools:
            names.add(tname)
    if names:
        return names
    legacy = getattr(agent, "_function_tools", None)
    if isinstance(legacy, dict):
        return set(legacy.keys())
    return names


def _tool_fn(agent, tool_name: str):
    """Resolve a registered tool's underlying function from a pydantic-ai Agent.

    Mirrors the defence-in-depth pattern in `_agent_tool_names`: try the
    public `toolsets` shape first (1.77+ exposes `ts.tools` as a dict),
    fall back to a list-shape on `ts.tools`, then to the legacy private
    `_function_tools` map. A clean "tool not registered" AssertionError
    is preferable to an opaque AttributeError when the API shifts.
    """
    def _unwrap(tool):
        return getattr(tool, "function", None) or tool

    for ts in getattr(agent, "toolsets", []) or []:
        tools = getattr(ts, "tools", None)
        if isinstance(tools, dict):
            tool = tools.get(tool_name)
            if tool is not None:
                return _unwrap(tool)
        elif isinstance(tools, (list, tuple)):
            for tool in tools:
                name = getattr(tool, "name", None) or getattr(
                    getattr(tool, "function", None), "__name__", None,
                )
                if name == tool_name:
                    return _unwrap(tool)

    legacy = getattr(agent, "_function_tools", None)
    if isinstance(legacy, dict):
        tool = legacy.get(tool_name)
        if tool is not None:
            return _unwrap(tool)

    raise AssertionError(f"{tool_name} tool not registered")
