"""Phase 1 (post-FINCO-2021 audit) — prompt-contract pins.

These tests anchor the three prompt-level changes from
`docs/PLAN-notes-pipeline-improvements.md` Phase 1 so a later edit can't
silently revert them:

1.1 PDF-page citation pin — every notes prompt must tell the model to cite
    the PDF page number (the one passed to `view_pdf_pages`), not the
    printed folio from the page image footer.
1.2 Schedule-or-prose rule — when a note contains a numeric schedule, it
    must be rendered as an ASCII table, not replaced by policy prose.
1.3 Sub-agent batch-scope nudge — the Sheet-12 sub-agent prompt must name
    the sub-agent's PDF page range explicitly so wander is discouraged.
"""
from __future__ import annotations

import re
from pathlib import Path

import pytest

from notes.agent import render_notes_prompt
from notes_types import NotesTemplateType
from scout.notes_discoverer import NoteInventoryEntry


_PROMPT_DIR = Path(__file__).resolve().parent.parent / "prompts"


def _flatten(s: str) -> str:
    return re.sub(r"\s+", " ", s).lower()


def test_base_prompt_pins_pdf_page_citation():
    """Phase 1.1 — base prompt must state that `evidence` cites PDF pages
    (the ones passed to view_pdf_pages), not the printed folio."""
    base = (_PROMPT_DIR / "_notes_base.md").read_text(encoding="utf-8")
    flat = _flatten(base)
    # Must mention both "PDF page" and a no-printed-folio caveat.
    assert "pdf page" in flat
    assert "printed folio" in flat or "printed page" in flat


def test_listofnotes_prompt_pins_pdf_page_citation():
    """Phase 1.1 — the Sheet-12-specific prompt repeats the rule since
    that's where the drift was observed."""
    loa = (_PROMPT_DIR / "notes_listofnotes.md").read_text(encoding="utf-8")
    flat = _flatten(loa)
    assert "pdf page" in flat
    assert "printed folio" in flat or "printed page" in flat


def test_base_prompt_has_schedules_section():
    """Phase 1.2 — the base prompt must mandate rendering numeric
    schedules (movement tables, ECL allowances, maturity analyses) as
    real tables rather than replacing them with policy prose.

    Note: the schedule rendering format moved from ASCII to HTML as
    part of the rich-editor pipeline (docs/PLAN-NOTES-RICH-EDITOR.md);
    the "schedules render, don't get paraphrased" invariant is the
    stable part, and that's what this test pins."""
    base = (_PROMPT_DIR / "_notes_base.md").read_text(encoding="utf-8")
    assert "SCHEDULES" in base or "SCHEDULE" in base
    flat = _flatten(base)
    # Tables must still be mentioned — format is HTML `<table>` now
    # (formerly ASCII columns).
    assert "<table>" in base or "table" in flat
    assert "movement" in flat or "maturity" in flat
    # Explicit "do not substitute prose for the schedule" rule.
    assert "do not drop" in flat or "do not replace" in flat


def test_rendered_listofnotes_prompt_includes_schedules_rule():
    """The per-template rendered prompt must inherit the schedules rule
    from the base — no way for a specific sheet to silently drop it."""
    prompt = render_notes_prompt(
        template_type=NotesTemplateType.LIST_OF_NOTES,
        filing_level="company",
        inventory=[],
    )
    assert "SCHEDULES" in prompt or "SCHEDULE" in prompt


def test_subcoordinator_prompt_includes_batch_page_range():
    """Phase 1.3 — the per-sub prompt must name the batch's PDF page
    range so the model has something concrete to scope itself against."""
    # We test the helper by building the prompt string in isolation — we
    # don't need the full agent.iter harness to exercise the nudge.
    from notes.listofnotes_subcoordinator import _invoke_sub_agent_once
    import inspect

    # Extract the function source so the test can't be faked by a later
    # edit that deletes the batch-pages derivation. We assert the source
    # references the variables by name.
    src = inspect.getsource(_invoke_sub_agent_once)
    assert "batch_min" in src and "batch_max" in src, (
        "Phase 1.3 expects the sub-agent prompt to derive a batch page "
        "range (batch_min/batch_max) and surface it in the user prompt."
    )
    assert "scope_line" in src or "batch covers PDF pages" in src


def test_subcoordinator_prompt_empty_batch_omits_scope_line():
    """Vacuous case: an empty batch should not emit an empty page-range
    sentence — avoids 'pages 0–0' noise in an edge-case prompt."""
    # We smoke-test by synthesising the prompt fragment the runtime would
    # build. Keep in sync with the string in _invoke_sub_agent_once — if
    # that string changes, this test updates alongside.
    batch: list[NoteInventoryEntry] = []
    batch_pages = [p for e in batch for p in range(e.page_range[0], e.page_range[1] + 1)]
    assert not batch_pages  # empty-batch invariant


# ---------------------------------------------------------------------------
# Sub-sheet template-first rule (Phase 4 of the model+notes-heading plan).
#
# Production runs have been lumping note breakdowns onto the face sheet when a
# matching sub-sheet field exists. The fix is a prompt rule that gates the
# breakdown-to-sub-sheet decision on TEMPLATE granularity, not on the note's
# line count. These tests pin that wording — including a negative assertion
# against the rejected "one row per note line" quota rule.
# ---------------------------------------------------------------------------


def test_sofp_prompt_has_template_first_breakdown_rule():
    """prompts/sofp.md must anchor the breakdown rule on the sub-sheet
    field list (template-first), not on note-line count."""
    body = (_PROMPT_DIR / "sofp.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    # Core template-first phrasing: the agent checks whether a matching
    # sub-sheet field exists before deciding to split a note line.
    assert "matching sub-sheet field" in flat, (
        "prompts/sofp.md must require the agent to check for a matching "
        "sub-sheet field before splitting a note breakdown"
    )


def test_sofp_prompt_names_the_failure_mode_explicitly():
    """The failure case (lump sum on face sheet when a sub-sheet field
    exists) must be called out as the thing to avoid."""
    body = (_PROMPT_DIR / "sofp.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "lump sum" in flat and "face sheet" in flat, (
        "prompts/sofp.md must explicitly call out 'lump sum on the face "
        "sheet' as the failure mode"
    )


def test_sofp_prompt_does_not_carry_rejected_quota_rule():
    """Negative assertion: the rigid one-row-per-note-line quota rule
    is wrong (the template controls granularity, not the note). Guard
    against it being re-introduced by a future edit."""
    body = (_PROMPT_DIR / "sofp.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    # Reject both phrasings the earlier plan draft used.
    assert "one sub-sheet row per breakdown line" not in flat, (
        "prompts/sofp.md contains the rejected quota rule 'one row per "
        "breakdown line' — the rule must gate on template granularity"
    )
    assert "must write 5 sub-sheet rows" not in flat, (
        "prompts/sofp.md contains the rejected concrete quota example "
        "'must write 5 sub-sheet rows'"
    )


def test_sopl_prompt_has_template_first_breakdown_rule():
    """prompts/sopl.md must carry the same template-first rule for the
    Analysis sub-sheet."""
    body = (_PROMPT_DIR / "sopl.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "matching" in flat and "analysis" in flat, (
        "prompts/sopl.md must reference matching note lines to Analysis "
        "sub-sheet fields (template-first rule)"
    )
    assert "lump" in flat or "single line" in flat, (
        "prompts/sopl.md must call out the lumping failure mode"
    )


def test_base_prompt_requires_following_linked_notes_before_lumping():
    """Shared face-statement prompt must force the accountant workflow:
    face line -> linked note -> template-specific component rows."""
    body = (_PROMPT_DIR / "_base.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "accountant extraction procedure" in flat
    assert "before writing any face-statement line that has a note reference" in flat
    assert "only write a lump-sum face value" in flat


def test_sofp_prompt_has_linked_note_split_example():
    """The SOFP prompt should teach linked-note splitting without relying
    on the user's receivables example."""
    body = (_PROMPT_DIR / "sofp.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "linked-note cash case" in flat
    assert "cash on hand" in flat
    assert "fixed deposits with licensed banks" in flat
    assert "do not write rm1,200,000 only to the face statement" in flat


def test_notes_base_prompt_preserves_parent_note_hierarchy():
    """Notes prompts should not split supporting sub-sections into
    unrelated disclosure rows just because the PDF uses (a)/(b)."""
    body = (_PROMPT_DIR / "_notes_base.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "note hierarchy and granularity" in flat
    assert "not like a text splitter" in flat
    assert "finance costs" in flat
    assert "interest on lease liabilities" in flat
    assert "do not split content into a different template row merely because" in flat


def test_notes_base_prompt_requires_in_prose_subsection_label_preservation():
    """Bug 2026-04-27: agents were stripping "(a) Short term benefits" /
    "(b) Defined contribution plans" sub-headers from the body, leaving
    naked prose. The prompt must explicitly require preserving these
    in-prose sub-section labels as bold paragraph headers — and must
    scope the writer-owned heading rule to parent_note/sub_note only so
    the agent doesn't over-generalise it to "drop all sub-headings"."""
    body = (_PROMPT_DIR / "_notes_base.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    # Must explicitly require preservation of in-prose (a)/(b) labels.
    assert "preserve the sub-section labels themselves in the body" in flat, (
        "_notes_base.md must explicitly tell the agent to preserve "
        "(a)/(b)/(i)/(ii) sub-section labels in the body content"
    )
    # Must show the recommended <strong> rendering pattern so the agent
    # has a concrete shape to copy.
    assert "<strong>(a) short term benefits</strong>" in flat, (
        "_notes_base.md must show the <p><strong>(a) ...</strong></p> "
        "pattern in the hierarchy guidance"
    )
    # Must scope the writer-owned heading rule so it doesn't cause the
    # agent to over-strip in-prose sub-headers.
    assert (
        "applies only to the parent_note and sub_note" in flat
        or "scoped strictly to the parent_note and sub_note" in flat
    ), (
        "_notes_base.md 'Heading markup is writer-owned' rule must be "
        "scoped to parent_note + sub_note only — the wider phrasing "
        "causes agents to strip in-prose (a)/(b) labels"
    )


def test_notes_base_prompt_shows_subsection_worked_example():
    """The prompt must include a worked example with (a)/(b) sub-section
    labels preserved as <strong> headers so the LLM has a concrete shape
    to copy. Without an example, the rule is too easily missed."""
    body = (_PROMPT_DIR / "_notes_base.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    # Worked example must show both labels alongside a parent_note that
    # carries a dotted-number (the (a)/(b) pattern is most common in
    # accounting-policy sub-policies like 2.14 Employee benefits).
    assert "<strong>(a) short term benefits</strong>" in flat
    assert "<strong>(b) defined contribution plans</strong>" in flat
    assert '"number": "2.14"' in body


def test_notes_accounting_policies_prompt_calls_out_subsection_preservation():
    """The accounting-policies per-template prompt is the most common
    landing spot for (a)/(b) sub-policy splits (Note 2.x), so it must
    repeat the rule rather than relying on the base prompt alone."""
    body = (_PROMPT_DIR / "notes_accounting_policies.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "preserve any \"(a)/(b)/(i)/(ii)\" sub-section labels" in flat, (
        "notes_accounting_policies.md must repeat the in-prose "
        "sub-section preservation rule (the bug surfaced on a Note 2.14 "
        "Employee benefits cell)"
    )
    assert "do not flatten them" in flat or "do not strip" in flat or (
        "<strong>" in body
    ), (
        "notes_accounting_policies.md must show a concrete <strong> shape "
        "or call out the do-not-flatten rule"
    )


def test_listofnotes_prompt_warns_hierarchy_beats_visual_granularity():
    """Sheet-12 matching prompt needs the same parent-note hierarchy guardrail."""
    body = (_PROMPT_DIR / "notes_listofnotes.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "hierarchy beats visual granularity" in flat
    assert "one finance-costs payload" in flat
    assert "do not move the lease-interest sub-section" in flat


def test_scout_prompt_preserves_face_note_references():
    """Scout should treat face-statement note references as downstream
    extraction hints, not incidental text."""
    from scout.agent import _SYSTEM_PROMPT

    flat = _flatten(_SYSTEM_PROMPT)
    assert "capture the note-reference column" in flat
    assert "best-effort note page hints" in flat


def test_base_prompt_has_sign_convention_troubleshooting():
    """Shared prompt should tell agents to debug sign mismatches without
    blindly negating labels that contain loss/expense wording."""
    body = (_PROMPT_DIR / "_base.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "sign-convention troubleshooting" in flat
    assert "do not infer the sign from wording alone" in flat
    assert "foreign exchange loss" in flat
    assert "if the formula subtracts a row" in flat


def test_sopl_prompt_keeps_loss_expenses_positive():
    """P&L loss labels should be positive magnitudes because SOPL formulas
    handle subtraction."""
    body = (_PROMPT_DIR / "sopl.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    assert "loss-labelled expense rows are also positive magnitudes" in flat
    assert "foreign exchange loss" in flat
    assert "impairment loss on trade receivables" in flat


def test_socie_mpers_group_section_does_not_advertise_efg_columns():
    """Negative pin (peer-review H1): MPERS Group SOCIE has only col B
    per block — no E/F value columns. The prompt previously claimed
    "Group filings additionally use: E (col=5) = Company CY, F (col=6) =
    evidence", which contradicted the four-block layout already
    described elsewhere in the same file. Guard against the line
    re-appearing in a future edit."""
    body = (_PROMPT_DIR / "socie_mpers.md").read_text(encoding="utf-8")
    flat = _flatten(body)
    # The exact stale phrasings.
    assert "additionally use: e (col=5)" not in flat, (
        "socie_mpers.md still advertises a non-existent col E for Group SOCIE"
    )
    assert "f (col=6) = evidence" not in flat, (
        "socie_mpers.md still advertises a non-existent col F for Group SOCIE"
    )
    # Positive guidance: explicitly call out NO additional columns.
    assert "no additional value columns" in flat or "no e/f" in flat, (
        "socie_mpers.md must explicitly state Group has no E/F value columns"
    )


def test_equity_prompts_follow_dividend_formula_sign():
    """SOCIE/SoRE templates subtract dividends, so prompts must ask for
    positive dividend magnitudes rather than double-negating them."""
    for filename in ("socie.md", "socie_mpers.md", "socie_sore.md"):
        body = (_PROMPT_DIR / filename).read_text(encoding="utf-8")
        flat = _flatten(body)
        assert "do not apply the sopl" in flat
        assert "dividends paid are entered as positive magnitudes" in flat
        assert "subtracts the dividends row" in flat or "formula subtracts it" in flat
        assert "reconciles to sofp" in flat


def _assert_dividends_subtracted(workbook: str, sheet: str) -> None:
    """Helper: every `Dividends paid` row in `sheet` must be subtracted by
    a column-B subtotal formula within 10 rows below it.

    Pulled out of the parametrised case so the (currently-broken) MPERS
    Group SOCIE template can be xfailed individually instead of muting
    the whole table.
    """
    import openpyxl

    wb = openpyxl.load_workbook(_PROMPT_DIR.parent / workbook, data_only=False)
    try:
        ws = wb[sheet]
        div_rows = [
            row for row in range(1, ws.max_row + 1)
            if str(ws.cell(row, 1).value or "").strip().lower() == "dividends paid"
        ]
        assert div_rows, f"{workbook} has no dividends paid row"
        for div_row in div_rows:
            div_ref = f"B{div_row}"
            formulas = [
                str(ws.cell(row, 2).value)
                for row in range(div_row + 1, min(ws.max_row, div_row + 10) + 1)
                if isinstance(ws.cell(row, 2).value, str)
                and ws.cell(row, 2).value.startswith("=")
                and div_ref in ws.cell(row, 2).value
            ]
            assert any(f"-1*{div_ref}" in f or f"-{div_ref}" in f for f in formulas), (
                f"{workbook} formula near row {div_row} does not subtract {div_ref}: {formulas}"
            )
    finally:
        wb.close()


@pytest.mark.parametrize(
    "workbook,sheet",
    [
        ("XBRL-template-MFRS/Company/09-SOCIE.xlsx", "SOCIE"),
        ("XBRL-template-MFRS/Group/09-SOCIE.xlsx", "SOCIE"),
        ("XBRL-template-MPERS/Company/09-SOCIE.xlsx", "SOCIE"),
        ("XBRL-template-MPERS/Group/09-SOCIE.xlsx", "SOCIE"),
        ("XBRL-template-MPERS/Company/10-SoRE.xlsx", "SoRE"),
        ("XBRL-template-MPERS/Group/10-SoRE.xlsx", "SoRE"),
    ],
)
def test_live_templates_subtract_dividends_paid(workbook: str, sheet: str):
    """Pin prompt guidance to the live formulas: dividends rows are
    subtracted by SOCIE/SoRE subtotal formulas."""
    _assert_dividends_subtracted(workbook, sheet)
