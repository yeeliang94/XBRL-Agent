"""PLAN-orchestration-hardening item 8 — atomic workbook saves everywhere.

Gotcha #22 fixed the notes validator with tempfile + ``os.replace`` and
flagged the same load→save-in-place pattern as latent in
``tools/fill_workbook.py`` and ``concept_model/exporter.py``. These tests
pin that every live workbook saver now routes through the shared
``utils.workbook_io.atomic_save_workbook`` helper, and that a mid-save
failure leaves the previous file intact (old-or-new, never truncated).
"""
from __future__ import annotations

import openpyxl
import pytest

from utils.workbook_io import atomic_save_workbook


def _make_template(tmp_path, label: str = "Revenue") -> str:
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP"
    ws["A5"] = label
    wb.save(path)
    wb.close()
    return path


class TestSharedHelper:
    def test_validator_alias_is_the_shared_helper(self):
        """The notes validator's import/test contract is a re-export of the
        promoted helper — one mechanism, zero forks."""
        from notes.validator_agent import _atomic_save_workbook

        assert _atomic_save_workbook is atomic_save_workbook

    def test_failed_save_leaves_old_file_intact(self, tmp_path):
        """Mid-save interrupt simulation: a workbook whose save() blows up
        must not clobber or truncate the existing file."""
        target = tmp_path / "out.xlsx"
        wb = openpyxl.Workbook()
        wb.active["A1"] = "original"
        atomic_save_workbook(wb, str(target))
        original_bytes = target.read_bytes()

        class _Boom:
            def save(self, path):
                # Simulate a crash after partially writing the temp file.
                with open(path, "wb") as fh:
                    fh.write(b"PK\x03\x04 truncated")
                raise RuntimeError("disk full mid-save")

        with pytest.raises(RuntimeError):
            atomic_save_workbook(_Boom(), str(target))

        assert target.read_bytes() == original_bytes, (
            "a failed save must leave the previous file untouched"
        )
        # And the temp file is cleaned up — no stragglers accumulate.
        leftovers = [p for p in tmp_path.iterdir() if p.name != "out.xlsx"]
        assert leftovers == [], f"temp files leaked: {leftovers}"


class TestLiveSaversUseAtomicReplace:
    """Monkeypatch-spy: each live-path saver must persist via os.replace
    (the atomic step), never a bare in-place wb.save(target)."""

    def _spy_replace(self, monkeypatch):
        import utils.workbook_io as wio

        calls: list = []
        real_replace = wio.os.replace

        def spy(src, dst):
            calls.append(str(dst))
            return real_replace(src, dst)

        monkeypatch.setattr(wio.os, "replace", spy)
        return calls

    def test_fill_workbook_saves_atomically(self, tmp_path, monkeypatch):
        calls = self._spy_replace(monkeypatch)
        from tools.fill_workbook import fill_workbook

        template = _make_template(tmp_path)
        output = str(tmp_path / "filled.xlsx")
        result = fill_workbook(
            template, output,
            [{"sheet": "SOFP", "field_label": "Revenue", "value": 100,
              "col": 2}],
        )
        assert result.success
        assert output in calls, "fill_workbook must save via os.replace"

    def test_workbook_merger_saves_atomically(self, tmp_path, monkeypatch):
        calls = self._spy_replace(monkeypatch)
        from statement_types import StatementType
        from workbook_merger import merge

        src = _make_template(tmp_path)
        output = str(tmp_path / "merged.xlsx")
        result = merge({StatementType.SOFP: src}, output, skip_recalc=True)
        assert result.success
        assert output in calls, "workbook_merger must save via os.replace"

    def test_notes_writer_saves_atomically(self, tmp_path, monkeypatch):
        calls = self._spy_replace(monkeypatch)
        from notes.payload import NotesPayload
        from notes.writer import write_notes_workbook

        template = _make_template(tmp_path, label="Corporate information")
        output = str(tmp_path / "notes_filled.xlsx")
        result = write_notes_workbook(
            template_path=template,
            payloads=[NotesPayload(
                chosen_row_label="Corporate information",
                content="<p>hello</p>", evidence="p1", source_pages=[1],
                parent_note={"number": "1", "title": "Corporate information"},
            )],
            output_path=output,
            filing_level="company",
            sheet_name="SOFP",
        )
        assert result.success, result.errors
        assert output in calls, "notes writer must save via os.replace"

    def test_canonical_exporter_saves_atomically(self, tmp_path, monkeypatch):
        """The exporter's write path routes through the shared helper. We
        call atomic_save_workbook indirectly by asserting the module binds
        the shared symbol — its full export needs a seeded concept DB, which
        tests/test_canonical_export.py already covers end-to-end."""
        import concept_model.exporter as exporter

        assert exporter.atomic_save_workbook is atomic_save_workbook
