from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import pytest
from openpyxl import Workbook

from concept_model.importer import import_template
from concept_model.parser import parse_template
from concept_model.taxonomy_semantics import (
    _ROLES_BY_FILE,
    semantic_addresses_for,
)
from db.schema import init_db
from eval.mtool_ingest import build_catalogue, ingest_workbook
from mtool.exporter import build_fill_doc
from mtool.offline_fill import (
    fill_workbook,
    load_workbook_entries,
    validate_input,
)
from mtool.template_map import (
    index_workbook,
    inspect_template,
    resolve_filing_doc,
)
from notes_types import NOTES_REGISTRY, notes_template_path
from statement_types import VARIANTS, template_path


REPO = Path(__file__).resolve().parent.parent


def _semantic_template_cases():
    cases = []
    for statement, variant_name in VARIANTS:
        for standard in ("mfrs", "mpers"):
            for level in ("company", "group"):
                try:
                    path = template_path(
                        statement, variant_name, level, standard)
                except ValueError:
                    continue
                if path.exists() and path.name in _ROLES_BY_FILE:
                    cases.append((standard, level, path))
    for note_type, entry in NOTES_REGISTRY.items():
        if not entry.is_numeric:
            continue
        for standard in ("mfrs", "mpers"):
            for level in ("company", "group"):
                path = notes_template_path(note_type, level, standard)
                if path.exists() and path.name in _ROLES_BY_FILE:
                    cases.append((standard, level, path))
    return cases


_KNOWN_EMPTY_SEMANTIC_TEMPLATES: set[tuple[str, str]] = set()


@pytest.mark.parametrize(
    ("standard", "level", "template"),
    _semantic_template_cases(),
    ids=lambda value: value.name if isinstance(value, Path) else str(value),
)
def test_registered_semantic_templates_do_not_silently_lose_all_addresses(
    standard: str, level: str, template: Path,
):
    addresses = semantic_addresses_for(str(template))
    expected_empty = (standard, template.name) in _KNOWN_EMPTY_SEMANTIC_TEMPLATES
    assert bool(addresses) is not expected_empty, (
        f"{standard}/{level}/{template.name} semantic coverage changed: "
        f"{len(addresses)} address(es); update the mapping or the reviewed "
        "exception list explicitly"
    )


def _run(db: Path) -> int:
    conn = sqlite3.connect(db)
    try:
        row = conn.execute(
            "INSERT INTO runs(created_at,pdf_filename,status,started_at) "
            "VALUES ('2026-08-25','source.pdf','completed','2026-08-25')"
        )
        conn.commit()
        return int(row.lastrowid)
    finally:
        conn.close()


@pytest.mark.parametrize(
    ("standard", "level"),
    [("mfrs", "company"), ("mfrs", "group"),
     ("mpers", "company"), ("mpers", "group")],
)
def test_socie_resolves_to_explicit_generated_template_cell(
    tmp_path: Path, standard: str, level: str,
):
    template = (
        REPO / f"XBRL-template-{standard.upper()}" / level.capitalize()
        / "09-SOCIE.xlsx"
    )
    db = tmp_path / "audit.db"
    init_db(db)
    tree = parse_template(str(template))
    payload = tmp_path / "tree.json"
    payload.write_text(json.dumps(tree.to_json()), encoding="utf-8")
    import_template(db, payload)
    run_id = _run(db)

    conn = sqlite3.connect(db)
    try:
        fact = conn.execute(
            """
            SELECT n.concept_uuid, t.period, t.entity_scope,
                   t.target_sheet, t.target_row, t.target_col
            FROM concept_nodes n
            JOIN concept_targets t USING(concept_uuid)
            WHERE n.kind = 'MATRIX_CELL'
              AND NOT EXISTS (
                SELECT 1 FROM concept_edges e
                WHERE e.parent_uuid = n.concept_uuid
              )
            ORDER BY n.render_row, n.matrix_col, t.entity_scope, t.period
            LIMIT 1
            """
        ).fetchone()
        assert fact is not None
        conn.execute(
            "INSERT INTO run_concept_facts("
            "run_id,concept_uuid,period,entity_scope,value,value_status,updated_at) "
            "VALUES (?,?,?,?,123,'observed','2026-08-25')",
            (run_id, fact[0], fact[1], fact[2]),
        )
        conn.commit()
    finally:
        conn.close()

    doc = build_fill_doc(
        db, run_id, filing_standard=standard, filing_level=level)
    assert doc["meta"]["counts"]["excluded_matrix_socie"] == 0
    assert doc["meta"]["counts"]["semantic_mapped"] == 1
    assert doc["writes"][0]["kind"] == "MATRIX_CELL"

    ready, coverage = resolve_filing_doc(str(template), doc)
    assert coverage["status"] == "ready"
    assert coverage["mapped"] == coverage["requested"] == 1
    assert {k: ready["writes"][0][k] for k in ("sheet", "cell", "value")} == {
        "sheet": fact[3], "cell": f"{fact[5]}{fact[4]}", "value": 123,
    }
    assert validate_input(ready) == []

    filled = tmp_path / "filled.xlsx"
    report = fill_workbook(str(template), ready, str(filled), strict=True)
    assert report["status"] == "ok"
    conn = sqlite3.connect(db)
    try:
        catalogue = build_catalogue(conn, standard, level, [tree.template_id])
    finally:
        conn.close()
    reverse = ingest_workbook(
        filled, catalogue, filing_level=level, unit_scale=1.0)
    assert [(f.concept_uuid, f.period, f.entity_scope, f.value)
            for f in reverse.facts] == [(fact[0], fact[1], fact[2], 123.0)]
    assert reverse.semantic_deferred == 0
    assert reverse.matrix_deferred == 0


def test_inspection_reports_supported_target_and_semantic_source(tmp_path: Path):
    template = REPO / "XBRL-template-MFRS/Company/09-SOCIE.xlsx"
    doc = {"writes": [], "sheets": {}}
    report = inspect_template(str(template), doc)
    assert report["supported_mtool_version"] == "2.2"
    assert report["semantic_source"] == "generated-targets"
    assert report["mtool_compatibility"] == "verified-generated"


def _save_semantic_marker_workbook(
    path: Path,
    *,
    two_periods: bool = False,
    sheet_name: str = "SOCIE",
) -> None:
    """Create the smallest mTool-shaped dimensional sheet used by resolver tests.

    The taxonomy cells deliberately use the genuine mTool encoding: XSD hrefs,
    optional label roles, and composite table/axis/member references.  This is
    generated in-test so the suite never depends on the ignored Windows sample.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    def add_block(header_row: int, marker_row: int, end_row: int,
                  primary_row: int, end_date: str) -> None:
        ws.cell(header_row, 5).value = (
            "full_ifrs-cor_2022-03-24.xsd#"
            "ifrs-full_StatementOfChangesInEquityTable::"
            "full_ifrs-cor_2022-03-24.xsd#"
            "ifrs-full_ComponentsOfEquityAxis::"
            "full_ifrs-cor_2022-03-24.xsd#"
            "ifrs-full_IssuedCapitalMember"
        )
        ws.cell(marker_row, 3).value = "#DOM#"
        ws.cell(marker_row, 4).value = "#PRIM#"
        ws.cell(end_row, 3).value = "#ENDT#"
        ws.cell(end_row, 5).value = end_date
        ws.cell(primary_row, 1).value = (
            "full_ifrs-cor_2022-03-24.xsd#"
            "ifrs-full_ProfitLoss@http://www.xbrl.org/2003/role/label"
        )

    add_block(2, 3, 4, 5, "31/12/2024")
    if two_periods:
        add_block(10, 11, 12, 13, "31/12/2023")
    wb.save(path)


def test_index_workbook_indexes_strict_taxonomy_fragments(tmp_path: Path):
    template = tmp_path / "taxonomy-identifiers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Taxonomy"
    ws["A1"] = (
        "full_ifrs-cor_2022-03-24.xsd#"
        "ifrs-full_ProfitLoss@http://www.xbrl.org/2003/role/label"
    )
    ws["E2"] = (
        "full_ifrs-cor_2022-03-24.xsd#ifrs-full_Table::"
        "full_ifrs-cor_2022-03-24.xsd#ifrs-full_Axis::"
        "full_ifrs-cor_2022-03-24.xsd#ifrs-full_Member"
    )
    ws["A3"] = "ordinary-text#must-not-be-indexed"
    wb.save(template)

    _, data, _ = load_workbook_entries(str(template))
    occurrences, _ = index_workbook(data)

    assert occurrences["ifrs-full_ProfitLoss"] == [("Taxonomy", 1, "A")]
    assert occurrences["ifrs-full_Table"] == [("Taxonomy", 2, "E")]
    assert occurrences["ifrs-full_Axis"] == [("Taxonomy", 2, "E")]
    assert occurrences["ifrs-full_Member"] == [("Taxonomy", 2, "E")]
    assert "must-not-be-indexed" not in occurrences


def test_socie_period_markers_disambiguate_repeated_taxonomy_rows(
    tmp_path: Path,
):
    template = tmp_path / "socie-two-periods.xlsx"
    _save_semantic_marker_workbook(template, two_periods=True)
    address = {
        "primary_concept": "ifrs-full_ProfitLoss",
        "dimensions": {
            "ifrs-full_ComponentsOfEquityAxis":
                "ifrs-full_IssuedCapitalMember",
        },
    }
    doc = {
        "meta": {"filing_standard": "mfrs", "filing_level": "company"},
        "sheets": {"SOCIE": {"label_column": None, "columns": {}}},
        "writes": [
            {
                "sheet": "SOCIE", "label": "Profit or loss", "value": 100,
                "period": "CY", "entity_scope": "Company",
                "semantic_address": address,
            },
            {
                "sheet": "SOCIE", "label": "Profit or loss", "value": 90,
                "period": "PY", "entity_scope": "Company",
                "semantic_address": address,
            },
        ],
    }

    ready, coverage = resolve_filing_doc(str(template), doc)

    assert coverage["status"] == "attention"
    assert coverage["mapped"] == coverage["requested"] == 2
    assert coverage["unmapped"] == coverage["ambiguous"] == 0
    assert [(w["cell"], w["value"]) for w in ready["writes"]] == [
        ("E5", 100),
        ("E13", 90),
    ]


def test_dimensional_sheet_reports_missing_prior_year_section(
    tmp_path: Path,
):
    template = tmp_path / "socie-current-year-only.xlsx"
    _save_semantic_marker_workbook(template)
    doc = {
        "meta": {"filing_standard": "mfrs", "filing_level": "company"},
        "sheets": {"SOCIE": {"label_column": None, "columns": {}}},
        "writes": [{
            "sheet": "SOCIE",
            "label": "Profit or loss",
            "value": 90,
            "period": "PY",
            "entity_scope": "Company",
            "semantic_address": {
                "primary_concept": "ifrs-full_ProfitLoss",
                "dimensions": {
                    "ifrs-full_ComponentsOfEquityAxis":
                        "ifrs-full_IssuedCapitalMember",
                },
            },
        }],
    }

    ready, coverage = resolve_filing_doc(str(template), doc)

    assert ready["writes"] == []
    assert coverage["status"] == "blocked"
    unresolved = coverage["unresolved_writes"][0]
    assert unresolved["reason_code"] == "template_period_section_missing"
    assert "no prior-year section" in unresolved["detail"].lower()


def test_taxonomy_identifier_on_another_sheet_is_not_a_valid_target(
    tmp_path: Path,
):
    template = tmp_path / "wrong-sheet-only.xlsx"
    _save_semantic_marker_workbook(template, sheet_name="Other-Sheet")
    doc = {
        "meta": {"filing_standard": "mfrs", "filing_level": "company"},
        "sheets": {"SOCIE": {"label_column": None, "columns": {}}},
        "writes": [{
            "sheet": "SOCIE",
            "label": "Profit or loss",
            "value": 100,
            "period": "CY",
            "entity_scope": "Company",
            "semantic_address": {
                "primary_concept": "ifrs-full_ProfitLoss",
                "dimensions": {
                    "ifrs-full_ComponentsOfEquityAxis":
                        "ifrs-full_IssuedCapitalMember",
                },
            },
        }],
    }

    ready, coverage = resolve_filing_doc(str(template), doc)

    assert ready["writes"] == []
    assert coverage["status"] == "blocked"
    unresolved = coverage["unresolved_writes"][0]
    assert unresolved["reason_code"] == "taxonomy_identifier_missing_on_sheet"
    assert "expected template sheet" in unresolved["detail"].lower()


def test_dimensional_sheet_without_fact_dimension_is_actionably_blocked(
    tmp_path: Path,
):
    template = tmp_path / "category-without-fact-dimension.xlsx"
    _save_semantic_marker_workbook(template)
    doc = {
        "meta": {"filing_standard": "mfrs", "filing_level": "company"},
        "sheets": {"SOCIE": {"label_column": None, "columns": {}}},
        "writes": [{
            "sheet": "SOCIE",
            "label": "Profit or loss",
            "column_role": "current_year",
            "period": "CY",
            "entity_scope": "Company",
            "value": 100,
            "semantic_address": {
                "primary_concept": "ifrs-full_ProfitLoss",
                "dimensions": {},
            },
        }],
    }

    ready, coverage = resolve_filing_doc(str(template), doc)

    assert ready["writes"] == []
    assert coverage["status"] == "blocked"
    unresolved = coverage["unresolved_writes"][0]
    assert unresolved["reason_code"] == "missing_category_dimensions"
    assert "category dimension" in unresolved["detail"].lower()


def test_category_sheet_legacy_fallback_is_structurally_blocked(
    tmp_path: Path,
):
    """A category matrix cannot fall through to blank CY/PY columns.

    Some mTool workbooks omit the taxonomy address needed to choose a share
    class or equity-component column. That is an actionable coverage failure,
    not a malformed column-map exception and not a positional guess.
    """
    template = tmp_path / "legacy-category.xlsx"
    _save_semantic_marker_workbook(
        template,
        sheet_name="Notes-Issuedcapital",
    )
    doc = {
        "meta": {
            "filing_standard": "mfrs",
            "filing_level": "company",
            "denomination": "thousands",
        },
        "sheets": {
            "Notes-Issuedcapital": {
                "label_column": None,
                "columns": {"current_year": None, "prior_year": None},
            },
        },
        "writes": [{
            "sheet": "Notes-Issuedcapital",
            "label": "Number of shares issued",
            "column_role": "current_year",
            "value": 100,
        }],
    }

    ready, coverage = resolve_filing_doc(str(template), doc)

    assert ready["writes"] == []
    assert coverage["status"] == "blocked"
    assert coverage["mapped"] == 0
    assert coverage["unmapped"] == 1
    assert coverage["legacy_label_writes"] == 0
    unresolved = coverage["unresolved_writes"][0]
    assert unresolved["sheet"] == "Notes-Issuedcapital"
    assert "taxonomy" in unresolved["detail"].lower()
    assert "category" in unresolved["detail"].lower()


def test_generated_statement_family_match_uses_exact_token():
    """SOCI must not pass merely because ``SOCIE`` contains that substring."""
    template = REPO / "XBRL-template-MFRS/Company/09-SOCIE.xlsx"
    doc = {
        "meta": {"filing_standard": "mfrs", "filing_level": "company"},
        "sheets": {"SOCI": {"columns": {}}},
        "writes": [{"template_id": "mfrs-company-soci-beforetax-v1"}],
    }
    report = inspect_template(str(template), doc)
    assert report["filing_family_match"] is False
