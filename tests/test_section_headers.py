"""Tests for section-header discovery (Phase 1, Step 1.2)."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from tools.section_headers import discover_section_headers, header_set

REPO = Path(__file__).resolve().parent.parent
MFRS = REPO / "XBRL-template-MFRS" / "Company"


# Every template must have at least a handful of headers — empty results
# indicate the detector missed the colour scheme and fill_workbook would
# silently mis-attribute duplicate labels.
_MFRS_TEMPLATES = [
    "01-SOFP-CuNonCu.xlsx",
    "02-SOFP-OrderOfLiquidity.xlsx",
    "03-SOPL-Function.xlsx",
    "04-SOPL-Nature.xlsx",
    "05-SOCI-BeforeTax.xlsx",
    "06-SOCI-NetOfTax.xlsx",
    "07-SOCF-Indirect.xlsx",
    "08-SOCF-Direct.xlsx",
    "09-SOCIE.xlsx",
]


@pytest.mark.parametrize("fname", _MFRS_TEMPLATES)
def test_section_headers_discovered_from_template(fname: str) -> None:
    """Every MBRS template yields a non-empty header set on its primary sheet."""
    path = MFRS / fname
    wb = openpyxl.load_workbook(path, data_only=False)
    try:
        # Primary sheet is always first. Second sheet (if present) is the sub.
        for sh in wb.sheetnames:
            headers = discover_section_headers(wb[sh])
            assert headers, f"{fname}!{sh} produced no section headers"
            # No header row should be labelled "Total ..." — that's a total row.
            for h in headers:
                assert not h.normalized.startswith("total "), (
                    f"total row leaked through as header: {fname}!{sh} row {h.row} {h.label!r}"
                )
    finally:
        wb.close()


def test_sofp_cunoncu_headers_contain_expected() -> None:
    """Spot-check the canonical SOFP/CuNonCu header list."""
    wb = openpyxl.load_workbook(MFRS / "01-SOFP-CuNonCu.xlsx", data_only=False)
    try:
        headers = header_set(wb, "SOFP-CuNonCu")
    finally:
        wb.close()
    expected = {
        "assets",
        "non-current assets",
        "current assets",
        "equity and liabilities",
        "equity",
        "liabilities",
        "non-current liabilities",
        "current liabilities",
    }
    missing = expected - headers
    assert not missing, f"missing expected headers in SOFP-CuNonCu: {missing}"
