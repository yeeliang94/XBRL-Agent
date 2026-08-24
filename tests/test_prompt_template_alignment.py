"""Prompt instructions that must stay aligned with live template coordinates."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from prompts import render_prompt
from statement_types import StatementType


REPO = Path(__file__).resolve().parent.parent


def _date_rows(path: Path) -> set[int]:
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    return {
        cell.row
        for ws in wb.worksheets
        for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row))
        for cell in row
        if isinstance(cell.value, str) and "YYYY" in cell.value
    }


def test_shared_date_instruction_matches_company_and_group_templates():
    company_paths = [
        p for standard in ("MFRS", "MPERS")
        for p in (REPO / f"XBRL-template-{standard}" / "Company").glob("*.xlsx")
        if "SOCIE" not in p.name
    ]
    group_paths = [
        p for standard in ("MFRS", "MPERS")
        for p in (REPO / f"XBRL-template-{standard}" / "Group").glob("*.xlsx")
        if "SOCIE" not in p.name
    ]
    assert company_paths and group_paths
    assert all(_date_rows(p) == {1} for p in company_paths)
    assert all(_date_rows(p) == {2} for p in group_paths)

    prompt = render_prompt(
        StatementType.SOFP, "CuNonCu", filing_level="group",
        filing_standard="mfrs",
    )
    assert "Group non-SOCIE templates normally use B2:E2" in prompt
    assert "dates in row 1 of every sheet" not in prompt


def test_sore_prompt_matches_live_writable_and_formula_rows():
    prompt_path = REPO / "prompts" / "socie_sore.md"
    prompt = prompt_path.read_text(encoding="utf-8")
    writable_labels = {
        "Retained earnings at beginning of period",
        "Impact of changes in accounting policies",
        "Retained earnings at beginning of period, restated",
        "Profit (loss)",
        "Dividends paid",
        "Retained earnings at end of period",
    }

    for level in ("Company", "Group"):
        path = REPO / "XBRL-template-MPERS" / level / "10-SoRE.xlsx"
        ws = openpyxl.load_workbook(path, data_only=False)["SoRE"]
        formula_rows = {
            row for row in range(1, ws.max_row + 1)
            if any(
                isinstance(ws.cell(row, col).value, str)
                and ws.cell(row, col).value.startswith("=")
                for col in range(2, ws.max_column + 1)
            )
        }
        assert formula_rows == {12, 15}
        for label in writable_labels:
            assert label in prompt

    # The former row map targeted formula/nonexistent cells after the MPERS
    # templates changed. The prompt now uses labels and live writability.
    assert not re.search(r"\brow(?:s)?\s+(12|13|14|16|17|19|20|21)\b", prompt, re.I)
    assert "closing retained-earnings\n   row is a DATA_ENTRY row" in prompt
