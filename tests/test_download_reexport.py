"""Phase 1.3 — download rebuilds the merged workbook from DB facts.

The editable-review plan makes the DB the single source of truth: a
review-UI edit lands in run_concept_facts, and the download must reflect
it without any manual "regenerate" step. server._reexport_and_remerge_from_facts
rebuilds each succeeded statement's workbook from the facts and re-merges
into a temp file (the on-disk merged_workbook_path is left as the fallback).

These pin that:
  * an edited (user_override) fact appears in the re-exported workbook;
  * a run with no facts re-exports to None (legacy fallback path).
"""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl
import pytest


REPO = Path(__file__).resolve().parent.parent
CO_SOFP = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"


@pytest.fixture
def seeded_run(tmp_path: Path):
    """A run with a merged workbook on disk, a succeeded SOFP agent row,
    the SOFP template imported, and one edited leaf fact."""
    import server
    from db.schema import init_db
    from concept_model.parser import parse_template
    from concept_model.importer import import_template, import_company_targets
    from concept_model.facts_api import write_fact, FactWrite

    db_path = tmp_path / "xbrl.db"
    init_db(db_path)
    server.AUDIT_DB_PATH = db_path

    tree = parse_template(str(CO_SOFP))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    _ct_tid = import_template(db_path, jp)
    import_company_targets(db_path, _ct_tid)

    session_dir = tmp_path / "session"
    session_dir.mkdir()
    # Agent scratch workbook + the merged workbook the download points at.
    scratch = session_dir / "SOFP_filled.xlsx"
    scratch.write_bytes(CO_SOFP.read_bytes())
    merged = session_dir / "filled.xlsx"
    merged.write_bytes(CO_SOFP.read_bytes())

    conn = sqlite3.connect(str(db_path))
    try:
        run_id = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at, "
            "session_id, merged_workbook_path, run_config_json) "
            "VALUES (?,?,?,?,?,?,?)",
            ("2026-05-25T00:00:00Z", "x.pdf", "completed",
             "2026-05-25T00:00:00Z", "session", str(merged),
             json.dumps({"filing_level": "company",
                         "filing_standard": "mfrs"})),
        ).lastrowid
        conn.execute(
            "INSERT INTO run_agents(run_id, statement_type, variant, model, "
            "status, started_at) VALUES (?,?,?,?,?,?)",
            (run_id, "SOFP", "CuNonCu", "test", "succeeded",
             "2026-05-25T00:00:00Z"),
        )
        leaf = conn.execute(
            "SELECT concept_uuid, render_sheet, render_row FROM concept_nodes "
            "WHERE render_sheet='SOFP-CuNonCu' AND kind='LEAF' "
            "ORDER BY render_row LIMIT 1"
        ).fetchone()
        conn.commit()
    finally:
        conn.close()

    # A user edit: value_status='user_override', exactly what the PATCH writes.
    write_fact(db_path, run_id, FactWrite(
        concept_uuid=leaf[0], value=7777.0, value_status="user_override",
        source="manual edit", actor="user"))

    return server, db_path, run_id, leaf[1], int(leaf[2])


def test_reexport_reflects_edited_fact(seeded_run):
    server, db_path, run_id, sheet, row = seeded_run

    out = server._reexport_and_remerge_from_facts(run_id)
    try:
        assert out is not None and out.exists(), "expected a re-exported temp file"
        wb = openpyxl.load_workbook(str(out), data_only=False)
        assert wb[sheet][f"B{row}"].value == 7777.0
    finally:
        if out is not None:
            out.unlink(missing_ok=True)


def test_run_without_facts_skips_reexport(tmp_path):
    """A legacy run with no facts returns None so the download falls back to
    the on-disk workbook."""
    import server
    from db.schema import init_db

    db_path = tmp_path / "xbrl.db"
    init_db(db_path)
    server.AUDIT_DB_PATH = db_path
    conn = sqlite3.connect(str(db_path))
    try:
        run_id = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at, "
            "merged_workbook_path) VALUES (?,?,?,?,?)",
            ("2026-05-25T00:00:00Z", "x.pdf", "completed",
             "2026-05-25T00:00:00Z", str(tmp_path / "nope.xlsx")),
        ).lastrowid
        conn.commit()
    finally:
        conn.close()

    assert server._run_has_facts(db_path, run_id) is False
    assert server._reexport_and_remerge_from_facts(run_id) is None


def test_durable_reexport_includes_numeric_note_edit(tmp_path):
    """`_reexport_remerge_durable` overlays numeric-note facts onto the durable
    on-disk merged workbook — so non-download consumers see the edit too
    (peer-review follow-up)."""
    import server
    from db.schema import init_db
    from concept_model.parser import parse_template, _derive_template_id
    from concept_model.importer import import_template, import_company_targets
    from concept_model.bootstrap import import_all_notes_templates
    from notes_types import NotesTemplateType, notes_template_path

    db_path = tmp_path / "xbrl.db"
    init_db(db_path)
    server.AUDIT_DB_PATH = db_path

    # Face statement (so re-export isn't gated to None) + numeric notes registry.
    tree = parse_template(str(CO_SOFP))
    jp = tmp_path / "tree.json"
    jp.write_text(json.dumps(tree.to_json(), sort_keys=True), encoding="utf-8")
    sofp_tid = import_template(db_path, jp)
    import_company_targets(db_path, sofp_tid)
    import_all_notes_templates(db_path)

    cap_tid = "mfrs-company-notes-issuedcapital-v1"
    cap_tpl = notes_template_path(
        NotesTemplateType.ISSUED_CAPITAL, level="company", standard="mfrs"
    )

    session_dir = tmp_path / "session"
    session_dir.mkdir()
    (session_dir / "SOFP_filled.xlsx").write_bytes(CO_SOFP.read_bytes())
    merged = session_dir / "filled.xlsx"
    merged.write_bytes(CO_SOFP.read_bytes())
    # The agent's numeric-notes workbook on disk (un-edited template values).
    (session_dir / "NOTES_ISSUED_CAPITAL_filled.xlsx").write_bytes(
        cap_tpl.read_bytes()
    )

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        run_id = conn.execute(
            "INSERT INTO runs(created_at, pdf_filename, status, started_at, "
            "session_id, merged_workbook_path, run_config_json) "
            "VALUES (?,?,?,?,?,?,?)",
            ("2026-06-17T00:00:00Z", "x.pdf", "completed",
             "2026-06-17T00:00:00Z", "session", str(merged),
             json.dumps({"filing_level": "company", "filing_standard": "mfrs",
                         "notes_to_run": ["ISSUED_CAPITAL"]})),
        ).lastrowid
        conn.execute(
            "INSERT INTO run_agents(run_id, statement_type, variant, model, "
            "status, started_at) VALUES (?,?,?,?,?,?)",
            (run_id, "SOFP", "CuNonCu", "test", "succeeded",
             "2026-06-17T00:00:00Z"),
        )
        tgt = conn.execute(
            "SELECT t.concept_uuid, t.target_sheet, t.target_row, t.target_col "
            "FROM concept_targets t JOIN concept_nodes n "
            "ON n.concept_uuid = t.concept_uuid "
            "WHERE n.template_id = ? AND n.kind = 'LEAF' "
            "AND t.entity_scope = 'Company' AND t.period = 'CY' LIMIT 1",
            (cap_tid,),
        ).fetchone()
        conn.execute(
            "INSERT INTO run_concept_facts(run_id, concept_uuid, period, "
            "entity_scope, value, value_status, updated_at) VALUES "
            "(?, ?, 'CY', 'Company', 55555.0, 'user_override', '2026-06-17T01:00:00Z')",
            (run_id, tgt["concept_uuid"]),
        )
        conn.commit()
    finally:
        conn.close()

    assert server._reexport_remerge_durable(run_id) is True

    from openpyxl.utils import column_index_from_string
    wb = openpyxl.load_workbook(str(merged), data_only=False)
    ws = wb[tgt["target_sheet"]]
    cell = ws.cell(
        row=int(tgt["target_row"]),
        column=column_index_from_string(tgt["target_col"]),
    )
    assert cell.value == 55555.0
