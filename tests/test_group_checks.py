"""Tests for Group-level verifier and cross-check support (Phases 5–6).

Group templates use 6 columns: A=labels, B=Group CY, C=Group PY,
D=Company CY, E=Company PY, F=Source.  Group SOCIE uses 4 vertical
blocks instead of 2.  These tests verify that the verifier and each
cross-check correctly handle both column layouts.
"""
from __future__ import annotations

import os
import tempfile

import openpyxl
import pytest

from statement_types import StatementType
from cross_checks.sofp_balance import SOFPBalanceCheck
from cross_checks.sopl_to_socie_profit import SOPLToSOCIEProfitCheck
from cross_checks.soci_to_socie_tci import SOCIToSOCIETCICheck
from cross_checks.socie_to_sofp_equity import SOCIEToSOFPEquityCheck
from cross_checks.socf_to_sofp_cash import SOCFToSOFPCashCheck


def _make_workbook(sheets: dict[str, list[list]], path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                if val is not None:
                    ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(path)
    wb.close()


@pytest.fixture
def tmp_dir():
    with tempfile.TemporaryDirectory() as d:
        yield d


# ---------------------------------------------------------------------------
# Phase 5: Verifier with Group layout
# ---------------------------------------------------------------------------


class TestVerifierGroup:
    """verify_statement with filing_level='group' checks both column pairs."""

    def test_sofp_group_both_balanced(self, tmp_dir):
        from tools.verifier import verify_statement
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                # A=label, B=Group CY, C=Group PY, D=Company CY, E=Company PY
                ["*Total assets",                   1000.0, 800.0, 500.0, 400.0],
                ["*Total equity and liabilities",   1000.0, 800.0, 500.0, 400.0],
            ],
        }, path)
        result = verify_statement(path, StatementType.SOFP, filing_level="group")
        assert result.is_balanced is True

    def test_sofp_group_company_imbalanced(self, tmp_dir):
        from tools.verifier import verify_statement
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets",                   1000.0, 800.0, 500.0, 400.0],
                ["*Total equity and liabilities",   1000.0, 800.0, 300.0, 400.0],
            ],
        }, path)
        result = verify_statement(path, StatementType.SOFP, filing_level="group")
        assert result.is_balanced is False
        assert "company" in result.feedback.lower() or "company" in str(result.mismatches).lower()

    def test_sofp_company_level_unchanged(self, tmp_dir):
        """Company-level filing_level still checks only B/C."""
        from tools.verifier import verify_statement
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets",                   1000.0, 800.0],
                ["*Total equity and liabilities",   1000.0, 800.0],
            ],
        }, path)
        result = verify_statement(path, StatementType.SOFP, filing_level="company")
        assert result.is_balanced is True


# ---------------------------------------------------------------------------
# Phase 6: Cross-checks with Group columns
# ---------------------------------------------------------------------------


class TestSOFPBalanceGroup:
    """SOFP balance check with Group 6-column layout."""

    def test_group_both_balanced(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets",                   1000.0, 800.0, 500.0, 400.0],
                ["*Total equity and liabilities",   1000.0, 800.0, 500.0, 400.0],
            ],
        }, path)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0, filing_level="group",
        )
        assert result.status == "passed"

    def test_group_company_imbalanced(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets",                   1000.0, 800.0, 500.0, 400.0],
                ["*Total equity and liabilities",   1000.0, 800.0, 300.0, 400.0],
            ],
        }, path)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0, filing_level="group",
        )
        assert result.status == "failed"
        assert "company" in result.message.lower()

    def test_group_consolidated_imbalanced(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets",                   1000.0, 800.0, 500.0, 400.0],
                ["*Total equity and liabilities",    900.0, 800.0, 500.0, 400.0],
            ],
        }, path)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0, filing_level="group",
        )
        assert result.status == "failed"
        assert "group" in result.message.lower()

    def test_company_level_backward_compat(self, tmp_dir):
        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets", 1000.0, 800.0],
                ["*Total equity and liabilities", 1000.0, 800.0],
            ],
        }, path)
        result = SOFPBalanceCheck().run(
            {StatementType.SOFP: path}, tolerance=1.0,
        )
        assert result.status == "passed"


class TestSOPLToSOCIEProfitGroup:
    """SOPL-to-SOCIE profit check: Group col B vs SOCIE block 1, Company col D vs block 3."""

    def test_group_both_match(self, tmp_dir):
        sopl_path = os.path.join(tmp_dir, "sopl.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")

        _make_workbook({
            "SOPL-Function": [
                # A=label, B=Group CY, C=Group PY, D=Company CY, E=Company PY
                ["*Profit (loss)", 250000.0, 200000.0, 100000.0, 80000.0],
            ],
        }, sopl_path)

        # SOCIE: block 1 (Group CY) starts row 3, block 3 (Company CY) starts row 51
        socie_rows = [[None] * 24 for _ in range(97)]
        # Row 1: header (col B = "Issued capital", col C = "Retained earnings")
        socie_rows[0] = [None, "Issued capital", "Retained earnings"] + [None] * 21
        # Block 1 (Group CY) — profit row
        socie_rows[10] = ["*Profit (loss)", None, 250000.0] + [None] * 21
        # Block 3 (Company CY) — profit row at row 59 (block 3 starts at row 51)
        socie_rows[58] = ["*Profit (loss)", None, 100000.0] + [None] * 21

        _make_workbook({"SOCIE": socie_rows}, socie_path)

        result = SOPLToSOCIEProfitCheck().run(
            {StatementType.SOPL: sopl_path, StatementType.SOCIE: socie_path},
            tolerance=1.0, filing_level="group",
        )
        assert result.status == "passed"

    def test_company_level_backward_compat(self, tmp_dir):
        sopl_path = os.path.join(tmp_dir, "sopl.xlsx")
        socie_path = os.path.join(tmp_dir, "socie.xlsx")

        _make_workbook({
            "SOPL-Function": [
                ["*Profit (loss)", 250000.0, 200000.0],
            ],
        }, sopl_path)
        _make_workbook({
            "SOCIE": [
                [None, "Issued capital", "Retained earnings"],
                ["*Profit (loss)", None, 250000.0],
            ],
        }, socie_path)

        result = SOPLToSOCIEProfitCheck().run(
            {StatementType.SOPL: sopl_path, StatementType.SOCIE: socie_path},
            tolerance=1.0,
        )
        assert result.status == "passed"


class TestSOCIEToSOFPEquityGroup:
    """SOCIE-to-SOFP equity: Group SOCIE block 1 col X = Group SOFP col B."""

    def test_group_both_match(self, tmp_dir):
        socie_path = os.path.join(tmp_dir, "socie.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        # SOCIE: equity at end of period in block 1 (row 25) and block 3 (row 73)
        socie_rows = [[None] * 24 for _ in range(97)]
        socie_rows[0] = [None, "Issued capital", "Retained earnings"] + [None] * 21
        # Block 1: row 25 (index 24) — col X (24) = Total
        socie_rows[24] = ["*Equity at end of period"] + [None] * 22 + [5000000.0]
        # Block 3: row 73 (index 72) — col X (24) = Total
        socie_rows[72] = ["*Equity at end of period"] + [None] * 22 + [3000000.0]

        _make_workbook({"SOCIE": socie_rows}, socie_path)

        _make_workbook({
            "SOFP-CuNonCu": [
                # A=label, B=Group CY, C=Group PY, D=Company CY, E=Company PY
                ["*Total equity", 5000000.0, 4000000.0, 3000000.0, 2500000.0],
            ],
        }, sofp_path)

        result = SOCIEToSOFPEquityCheck().run(
            {StatementType.SOCIE: socie_path, StatementType.SOFP: sofp_path},
            tolerance=1.0, filing_level="group",
        )
        assert result.status == "passed"

    def test_company_level_backward_compat(self, tmp_dir):
        socie_path = os.path.join(tmp_dir, "socie.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        socie_rows = [[None] * 24 for _ in range(50)]
        socie_rows[0] = [None, "Issued capital", "Retained earnings"] + [None] * 21
        socie_rows[24] = ["*Equity at end of period"] + [None] * 22 + [5000000.0]

        _make_workbook({"SOCIE": socie_rows}, socie_path)
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total equity", 5000000.0, 4000000.0],
            ],
        }, sofp_path)

        result = SOCIEToSOFPEquityCheck().run(
            {StatementType.SOCIE: socie_path, StatementType.SOFP: sofp_path},
            tolerance=1.0,
        )
        assert result.status == "passed"


class TestSOCFToSOFPCashGroup:
    """SOCF-to-SOFP cash: Group SOCF col B = Group SOFP col B, Company col D = Company col D."""

    def test_group_both_match(self, tmp_dir):
        socf_path = os.path.join(tmp_dir, "socf.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        _make_workbook({
            "SOCF-Indirect": [
                # A=label, B=Group CY, C=Group PY, D=Company CY, E=Company PY
                ["*Cash and cash equivalents at end of period", 120000.0, 100000.0, 80000.0, 60000.0],
            ],
        }, socf_path)

        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Cash and cash equivalents", 120000.0, 100000.0, 80000.0, 60000.0],
            ],
        }, sofp_path)

        result = SOCFToSOFPCashCheck().run(
            {StatementType.SOCF: socf_path, StatementType.SOFP: sofp_path},
            tolerance=1.0, filing_level="group",
        )
        assert result.status == "passed"

    def test_company_level_backward_compat(self, tmp_dir):
        socf_path = os.path.join(tmp_dir, "socf.xlsx")
        sofp_path = os.path.join(tmp_dir, "sofp.xlsx")

        _make_workbook({
            "SOCF-Indirect": [
                ["*Cash and cash equivalents at end of period", 120000.0, 100000.0],
            ],
        }, socf_path)
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Cash and cash equivalents", 120000.0, 100000.0],
            ],
        }, sofp_path)

        result = SOCFToSOFPCashCheck().run(
            {StatementType.SOCF: socf_path, StatementType.SOFP: sofp_path},
            tolerance=1.0,
        )
        assert result.status == "passed"


class TestFrameworkFilingLevel:
    """Framework run_all passes filing_level to checks."""

    def test_run_all_passes_filing_level(self, tmp_dir):
        from cross_checks.framework import run_all

        path = os.path.join(tmp_dir, "sofp.xlsx")
        _make_workbook({
            "SOFP-CuNonCu": [
                ["*Total assets",                   1000.0, 800.0, 500.0, 400.0],
                ["*Total equity and liabilities",   1000.0, 800.0, 500.0, 400.0],
            ],
        }, path)

        results = run_all(
            [SOFPBalanceCheck()],
            {StatementType.SOFP: path},
            {"statements_to_run": {StatementType.SOFP}, "filing_level": "group"},
            tolerance=1.0,
        )
        assert len(results) == 1
        assert results[0].status == "passed"
