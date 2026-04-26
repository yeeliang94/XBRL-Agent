"""Reproducibly build the RUN-REVIEW fixtures.

These fixtures encode the failure modes called out in §6 of `RUN-REVIEW.md`
(Amway MFRS Company FY2024). Each fixture is a copy of a live template
pre-populated with the leaves the reviewer flagged, including the two
"intentionally wrong" classification cells and the PY 1,881 restoration
provision double-booking.

Run this script after a template change to refresh the committed xlsx
files. The committed files are the ground truth — tests should NOT call
this builder at runtime, only the .xlsx outputs.

Usage:
    python tests/fixtures/run_review/_build_fixtures.py
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[3]
FIXTURE_DIR = Path(__file__).resolve().parent

# RUN-REVIEW §6.1 — the 18 leaves observed in the Amway MFRS Company run.
# Row numbers refer to SOFP-Sub-CuNonCu in the live MFRS Company template.
# Format: row -> (cy_value, py_value)
# Double-booking on row 287 PY=1,881 is intentional (also on row 318 PY).
# Rows 135 and 152 are the "intentionally wrong" classifications from the
# review (real values belong on 133 Finished goods and 142 Trade-receivables-
# from-subsidiaries respectively, but the reviewer's run mis-routed them).
MFRS_COMPANY_LEAVES = {
    11: (12_420, 12_688),    # Long term leasehold land
    16: (23_345, 23_978),    # Building on long term leasehold land
    24: (56, None),          # Motor vehicles (PY blank)
    29: (16_958, 9_906),     # Office equipment, fixture and fittings
    37: (2_275, None),       # Construction in progress (PY blank)
    38: (9_525, 9_592),      # Other property, plant and equipment (residual-plug suspect)
    64: (2_398, 849),        # Computer software (intangibles)
    71: (7_648, 7_648),      # Unquoted shares (subsidiaries)
    135: (159_389, 120_685), # Other inventories — INTENTIONALLY WRONG ROW
    139: (21_030, 35_674),   # Trade receivables (current)
    152: (95, 276),          # Other receivables due from subsidiaries — INTENTIONALLY WRONG ROW
    155: (259, 241),         # Other receivables due from other related parties
    158: (2_914, 1_110),     # Prepayments
    163: (1_795, 1_737),     # Deposits
    166: (217, 5_732),       # Other current non-trade receivables
    182: (61_668, 81_920),   # Balances with Licensed Banks
    185: (63_097, 156_767),  # Deposits placed with licensed banks
    200: (81_804, 81_804),   # Capital from ordinary shares
    287: (None, 1_881),      # Provision for decommissioning — DOUBLE-BOOKED PY
    318: (2_761, 1_881),     # Other non-current non-trade payables — also holds 1,881 PY
}

# MPERS Group twin. The structural shape differs (flatter calc per memory,
# fewer rows) so the PY 1,881 double-booking is encoded against the MPERS
# row equivalents, NOT the MFRS row numbers. Group has 6 columns
# (B=Group CY, C=Group PY, D=Company CY, E=Company PY, F=Source).
# To exercise the column-pair-aware double-booking guard (Phase 4.1),
# we duplicate the same Company values into the Group columns — that is
# legitimate consolidation pass-through and MUST NOT trip the guard.
# The PY 1,881 still sits on two rows (190 + 214) within the Company-PY
# column-pair: the genuine bug shape that SHOULD trip the guard.
MPERS_GROUP_LEAVES = {
    10: (12_420, 12_688),    # Long term leasehold land
    15: (23_345, 23_978),    # Building on long term leasehold land
    22: (16_958, 9_906),     # Office equipment, fixture and fittings
    24: (2_275, None),       # Construction in progress (PY blank)
    25: (9_525, 9_592),      # Other property, plant and equipment
    42: (7_648, 7_648),      # Unquoted shares (subsidiaries)
    90: (159_389, 120_685),  # Other inventories — INTENTIONALLY WRONG ROW (MPERS twin)
    105: (95, 276),          # Other receivables due from subsidiaries — INTENTIONALLY WRONG ROW
    108: (259, 241),         # Other receivables due from other related parties
    111: (2_914, 1_110),     # Prepayments
    116: (1_795, 1_737),     # Deposits
    119: (217, 5_732),       # Other current non-trade receivables
    126: (61_668, 81_920),   # Balances with banks
    129: (63_097, 156_767),  # Deposits placed with licensed banks
    140: (81_804, 81_804),   # Capital from ordinary shares
    190: (None, 1_881),      # Provision for decommissioning — DOUBLE-BOOKED PY
    214: (2_761, 1_881),     # Other non-current non-trade payables — also holds 1,881 PY
}


def build_mfrs_company() -> Path:
    """Copy MFRS Company SOFP template and pre-fill the 18 leaves.

    Company layout: B=CY, C=PY, D=Source. We write into B and C only.
    """
    src = REPO_ROOT / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"
    dst = FIXTURE_DIR / "sofp_company_mfrs.xlsx"
    shutil.copy(src, dst)

    wb = load_workbook(dst)
    ws = wb["SOFP-Sub-CuNonCu"]
    for row, (cy, py) in MFRS_COMPANY_LEAVES.items():
        if cy is not None:
            ws.cell(row, 2).value = cy
        if py is not None:
            ws.cell(row, 3).value = py
    wb.save(dst)
    return dst


def build_mpers_group() -> Path:
    """Copy MPERS Group SOFP template and pre-fill leaves into Group + Company column-pairs.

    Group layout: B=Group CY, C=Group PY, D=Company CY, E=Company PY, F=Source.
    Same value in B/D and C/E is legitimate (consolidation pass-through);
    the genuine double-book lives within a single column-pair (rows 190+214 PY).
    """
    src = REPO_ROOT / "XBRL-template-MPERS" / "Group" / "01-SOFP-CuNonCu.xlsx"
    dst = FIXTURE_DIR / "sofp_group_mpers.xlsx"
    shutil.copy(src, dst)

    wb = load_workbook(dst)
    ws = wb["SOFP-Sub-CuNonCu"]
    for row, (cy, py) in MPERS_GROUP_LEAVES.items():
        if cy is not None:
            ws.cell(row, 2).value = cy   # Group CY
            ws.cell(row, 4).value = cy   # Company CY (consolidation pass-through)
        if py is not None:
            ws.cell(row, 3).value = py   # Group PY
            ws.cell(row, 5).value = py   # Company PY
    wb.save(dst)
    return dst


def main() -> None:
    paths = [build_mfrs_company(), build_mpers_group()]
    for p in paths:
        print(f"wrote {p.relative_to(REPO_ROOT)} ({p.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
