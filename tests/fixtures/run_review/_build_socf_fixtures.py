"""Build SOCF fixtures for sign-convention regression tests.

Encodes the two specific divergences from RUN-REVIEW.md §3.6:
- (Gain) loss on disposal of PPE: AI emitted -70, filer used 70
- Cash payments for the principal portion of lease liabilities: AI emitted
  3,732, filer used -3,732

Both are valid in isolation; the "right" sign depends on whether the
*Total formula adds the cell (positive convention) or subtracts it
(negative convention). Phase 4.3 will read those formulas and surface
the convention to the prompt; this fixture lets us pin behaviour
across MFRS Co + MPERS Grp.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[3]
FIXTURE_DIR = Path(__file__).resolve().parent

# MFRS Company SOCF-Indirect (rows from the live template, not RUN-REVIEW's
# §6.2 — the reviewer's row guesses were template-version-dependent).
# Format: row -> (cy_value, py_value)
MFRS_COMPANY_SOCF_VALUES = {
    38: (-70, None),     # (Gain) loss on disposal of PPE — sign-ambiguous
    109: (3_732, None),  # Cash payments for principal portion of lease liability — sign-ambiguous
}

# MPERS Group SOCF-Indirect. Group has 6 cols (B/C=Group CY/PY, D/E=Company CY/PY).
# Same signs in both column-pairs to mirror real consolidation.
MPERS_GROUP_SOCF_VALUES = {
    34: (-70, None),    # (Gain) loss on disposal of PPE
    77: (3_732, None),  # Payments of finance lease liabilities (MPERS uses different label)
}


def build_mfrs_company_socf() -> Path:
    src = REPO_ROOT / "XBRL-template-MFRS" / "Company" / "07-SOCF-Indirect.xlsx"
    dst = FIXTURE_DIR / "socf_company_mfrs.xlsx"
    shutil.copy(src, dst)

    wb = load_workbook(dst)
    ws = wb["SOCF-Indirect"]
    for row, (cy, py) in MFRS_COMPANY_SOCF_VALUES.items():
        if cy is not None:
            ws.cell(row, 2).value = cy
        if py is not None:
            ws.cell(row, 3).value = py
    wb.save(dst)
    return dst


def build_mpers_group_socf() -> Path:
    src = REPO_ROOT / "XBRL-template-MPERS" / "Group" / "07-SOCF-Indirect.xlsx"
    dst = FIXTURE_DIR / "socf_group_mpers.xlsx"
    shutil.copy(src, dst)

    wb = load_workbook(dst)
    ws = wb["SOCF-Indirect"]
    for row, (cy, py) in MPERS_GROUP_SOCF_VALUES.items():
        if cy is not None:
            ws.cell(row, 2).value = cy   # Group CY
            ws.cell(row, 4).value = cy   # Company CY
        if py is not None:
            ws.cell(row, 3).value = py
            ws.cell(row, 5).value = py
    wb.save(dst)
    return dst


def main() -> None:
    paths = [build_mfrs_company_socf(), build_mpers_group_socf()]
    for p in paths:
        print(f"wrote {p.relative_to(REPO_ROOT)} ({p.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
