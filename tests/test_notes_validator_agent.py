"""Tests for the notes post-validator agent (Phase 5)."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest
from pydantic_ai.models.test import TestModel


class TestValidatorFactory:
    """Step 5.1: factory must register the 4 required tools."""

    def test_notes_validator_agent_factory_returns_agent(self, tmp_path):
        from notes.validator_agent import create_notes_validator_agent

        # Minimal merged workbook + sidecar so the factory can load inputs.
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Notes-SummaryofAccPol")
        wb.create_sheet("Notes-Listofnotes")
        merged_path = tmp_path / "merged.xlsx"
        wb.save(str(merged_path))

        sidecar = tmp_path / "NOTES_ACC_POLICIES_filled_payloads.json"
        sidecar.write_text(json.dumps([]), encoding="utf-8")

        agent, deps, ctx = create_notes_validator_agent(
            merged_workbook_path=str(merged_path),
            pdf_path=str(tmp_path / "x.pdf"),
            sidecar_paths=[str(sidecar)],
            filing_level="company",
            filing_standard="mfrs",
            model=TestModel(),
            output_dir=str(tmp_path),
        )
        tool_names = _agent_tool_names(agent)
        assert "view_pdf_pages" in tool_names
        assert "read_cell" in tool_names
        assert "rewrite_cell" in tool_names
        assert "flag_duplication" in tool_names


class TestInventoryCoverageGaps:
    """N3 Stage 1: deterministic per-note_num coverage-gap reporting."""

    def test_gaps_are_inventory_notes_with_no_content(self):
        from notes.validator_agent import inventory_coverage_gaps

        entries = [
            {"sheet": "Notes-Listofnotes", "source_note_refs": ["2.5(g)"]},
            {"sheet": "Notes-SummaryofAccPol", "source_note_refs": ["18"]},
        ]
        # Inventory has notes 2, 5, 18, 24 — 2 and 18 are covered, 5 and 24 not.
        gaps = inventory_coverage_gaps([2, 5, 18, 24], entries)
        assert gaps == [5, 24]

    def test_no_gaps_when_all_covered(self):
        from notes.validator_agent import inventory_coverage_gaps

        entries = [{"source_note_refs": ["2"]}, {"source_note_refs": ["5.1"]}]
        assert inventory_coverage_gaps([2, 5], entries) == []

    def test_empty_inventory_no_gaps(self):
        from notes.validator_agent import inventory_coverage_gaps
        assert inventory_coverage_gaps([], [{"source_note_refs": ["2"]}]) == []

    def test_malformed_ref_does_not_mask_a_note(self):
        from notes.validator_agent import inventory_coverage_gaps
        # A non-numeric ref must not be coerced to note 0 / hide a gap.
        entries = [{"source_note_refs": ["see disclosure", None]}]
        assert inventory_coverage_gaps([3], entries) == [3]

    def test_factory_surfaces_coverage_gaps_in_context(self, tmp_path):
        from notes.validator_agent import create_notes_validator_agent

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Notes-SummaryofAccPol")
        wb.create_sheet("Notes-Listofnotes")
        merged = tmp_path / "merged.xlsx"
        wb.save(str(merged))

        sidecar = tmp_path / "NOTES_LIST_OF_NOTES_filled_payloads.json"
        sidecar.write_text(json.dumps([
            {"sheet": "Notes-Listofnotes", "row": 5, "col": 2,
             "source_note_refs": ["18"], "content_preview": "x"},
        ]), encoding="utf-8")

        _agent, _deps, ctx = create_notes_validator_agent(
            merged_workbook_path=str(merged),
            pdf_path=str(tmp_path / "x.pdf"),
            sidecar_paths=[str(sidecar)],
            filing_level="company", filing_standard="mfrs",
            model=TestModel(), output_dir=str(tmp_path),
            inventory_note_nums=[12, 18, 20],
        )
        # 18 is covered; 12 and 20 are gaps.
        assert ctx["coverage_gaps"] == [12, 20]


class TestDetection:
    """Steps 5.3 and 5.4: pure detection helpers."""

    def test_detects_cross_sheet_duplicate_by_note_ref(self):
        from notes.validator_agent import detect_cross_sheet_duplicates_by_ref

        entries = [
            {"sheet": "Notes-SummaryofAccPol", "row": 20, "col": 2,
             "source_note_refs": ["5.1"], "content_preview": "policy on tax"},
            {"sheet": "Notes-Listofnotes", "row": 55, "col": 2,
             "source_note_refs": ["5.1"], "content_preview": "tax reconcil"},
        ]
        dups = detect_cross_sheet_duplicates_by_ref(entries)
        assert len(dups) == 1
        assert dups[0]["note_ref"] == "5.1"

    def test_does_not_flag_same_sheet_repeats(self):
        from notes.validator_agent import detect_cross_sheet_duplicates_by_ref

        entries = [
            {"sheet": "Notes-Listofnotes", "row": 40, "col": 2,
             "source_note_refs": ["5.1"], "content_preview": "a"},
            {"sheet": "Notes-Listofnotes", "row": 112, "col": 2,
             "source_note_refs": ["5.1"], "content_preview": "a2"},
        ]
        assert detect_cross_sheet_duplicates_by_ref(entries) == []

    def test_detects_overlap_when_refs_missing(self):
        from notes.validator_agent import detect_cross_sheet_overlap_candidates

        # Same text content on both sheets, no note-refs populated.
        text = (
            "The company recognises revenue when control of the goods "
            "is transferred to the buyer, in accordance with MFRS 15."
        )
        entries = [
            {"sheet": "Notes-SummaryofAccPol", "row": 30, "col": 2,
             "source_note_refs": [], "content_preview": text},
            {"sheet": "Notes-Listofnotes", "row": 70, "col": 2,
             "source_note_refs": [], "content_preview": text},
        ]
        cands = detect_cross_sheet_overlap_candidates(entries, threshold=0.5)
        assert len(cands) == 1
        assert cands[0]["score"] >= 0.5

    def test_overlap_skips_pairs_with_matching_refs(self):
        """Overlap fallback should not duplicate what the ref-based path
        already flags (guarded by a disjoint-refs gate)."""
        from notes.validator_agent import detect_cross_sheet_overlap_candidates

        entries = [
            {"sheet": "Notes-SummaryofAccPol", "row": 10, "col": 2,
             "source_note_refs": ["7"], "content_preview": "revenue policy"},
            {"sheet": "Notes-Listofnotes", "row": 90, "col": 2,
             "source_note_refs": ["7"], "content_preview": "revenue policy"},
        ]
        assert detect_cross_sheet_overlap_candidates(entries) == []


class TestSameSheetRowCollisions:
    """Detect a Sheet-12 row that received content from >1 top-level note.

    Reproduces the run-53 failure: `Disclosure of fair value information`
    (row 49) collected both Note 4.1 (investment property) and Note 20.7
    (financial instruments) — two unrelated top-level notes piled onto one
    XBRL concept. The catch-all row stays exempt; a single note split across
    payloads stays unflagged.
    """

    def test_flags_fair_value_row_holding_two_top_level_notes(self):
        from notes.validator_agent import detect_same_sheet_row_collisions

        # Sidecar row 49 unions refs across both notes (as the writer does).
        entries = [
            {"sheet": "Notes-Listofnotes", "row": 49, "col": 2,
             "row_label": "Disclosure of fair value information",
             "source_note_refs": ["4.1", "20", "20.7"],
             "content_preview": "fair value of investment property ... "
                                "fair value of financial instruments"},
        ]
        collisions = detect_same_sheet_row_collisions(entries)
        assert len(collisions) == 1
        c = collisions[0]
        assert c["row"] == 49
        assert c["note_nums"] == [4, 20]
        assert "Disclosure of fair value information" in c["row_label"]

    def test_catch_all_row_is_exempt(self):
        """`Disclosure of other notes to accounts` is the designed catch-all
        — aggregating many notes there is intended, not a finding."""
        from notes.validator_agent import detect_same_sheet_row_collisions

        entries = [
            {"sheet": "Notes-Listofnotes", "row": 112, "col": 2,
             "row_label": "Disclosure of other notes to accounts",
             "source_note_refs": ["6", "9", "21"],
             "content_preview": "misc"},
            # leading '*' + case must not bypass the exemption.
            {"sheet": "Notes-Listofnotes", "row": 112, "col": 2,
             "row_label": "*Disclosure of Other Notes to Accounts",
             "source_note_refs": ["7", "8"],
             "content_preview": "misc2"},
        ]
        assert detect_same_sheet_row_collisions(entries) == []

    def test_single_note_split_across_payloads_not_flagged(self):
        """Acceptance #4: Note 13 split into two payloads (both top-level 13)
        is legitimate aggregation, not a collision."""
        from notes.validator_agent import detect_same_sheet_row_collisions

        entries = [
            {"sheet": "Notes-Listofnotes", "row": 60, "col": 2,
             "row_label": "Disclosure of credit risk",
             "source_note_refs": ["13", "13.4"],
             "content_preview": "credit risk"},
        ]
        assert detect_same_sheet_row_collisions(entries) == []

    def test_sheet_11_entries_ignored(self):
        from notes.validator_agent import detect_same_sheet_row_collisions

        entries = [
            {"sheet": "Notes-SummaryofAccPol", "row": 20, "col": 2,
             "row_label": "Description of accounting policy for X",
             "source_note_refs": ["2", "3"],
             "content_preview": "policies"},
        ]
        assert detect_same_sheet_row_collisions(entries) == []


class TestSubnoteCoverageGaps:
    """Detect a note covered only PARTLY at sub-reference granularity.

    Reproduces the run-53 leases failure: scout discovered 3.1/3.2/3.3/(a)/(b)
    under Note 3; the leases policy cited 3.3 + (b) but dropped (a).
    """

    def test_flags_missing_lettered_subnote_when_sibling_cited(self):
        from notes.validator_agent import detect_subnote_coverage_gaps

        inventory_subnotes = {3: ["3.1", "3.2", "3.3", "(a)", "(b)"]}
        # The leases policy row's sidecar provenance, exactly as observed.
        entries = [
            {"sheet": "Notes-SummaryofAccPol", "row": 42, "col": 2,
             "row_label": "Description of accounting policy for leases",
             "source_note_refs": ["3", "3.3", "(b)"],
             "content_preview": "leases policy ... (b) Recognition exemption"},
        ]
        gaps = detect_subnote_coverage_gaps(inventory_subnotes, entries)
        assert len(gaps) == 1
        g = gaps[0]
        assert g["note_num"] == 3
        # (a) is the dropped lettered policy block; (b) was cited so it must
        # NOT appear as missing.
        assert "(a)" in g["missing_subnote_refs"]
        assert "(b)" not in g["missing_subnote_refs"]
        assert "(b)" in g["cited_subnote_refs"]

    def test_fully_covered_note_not_flagged(self):
        from notes.validator_agent import detect_subnote_coverage_gaps

        inventory_subnotes = {3: ["3.3", "(a)", "(b)"]}
        entries = [
            {"sheet": "Notes-SummaryofAccPol", "row": 42, "col": 2,
             "row_label": "Description of accounting policy for leases",
             "source_note_refs": ["3", "3.3", "(a)", "(b)"],
             "content_preview": "complete leases policy"},
        ]
        assert detect_subnote_coverage_gaps(inventory_subnotes, entries) == []

    def test_note_with_no_subnote_cited_not_flagged(self):
        """A note written as one combined cell citing only the bare number
        shows no sub-level engagement — top-level coverage guards it instead,
        so the proper-subset gate must keep it out of the sub-note report."""
        from notes.validator_agent import detect_subnote_coverage_gaps

        inventory_subnotes = {5: ["5.1", "5.2"]}
        entries = [
            {"sheet": "Notes-Listofnotes", "row": 70, "col": 2,
             "row_label": "Disclosure of X",
             "source_note_refs": ["5"],
             "content_preview": "combined"},
        ]
        assert detect_subnote_coverage_gaps(inventory_subnotes, entries) == []

    def test_lettered_ref_attributed_to_its_own_note(self):
        """A bare lettered ref like (b) is bucketed under the top-level note it
        was written alongside — not leaked to a different note that also has a
        (b) sub-section."""
        from notes.validator_agent import detect_subnote_coverage_gaps

        inventory_subnotes = {3: ["(a)", "(b)"], 9: ["(a)", "(b)"]}
        entries = [
            # Note 3 cites (b) only.
            {"sheet": "Notes-SummaryofAccPol", "row": 42, "col": 2,
             "source_note_refs": ["3", "(b)"], "content_preview": "leases"},
            # Note 9 is untouched at sub-level.
        ]
        gaps = detect_subnote_coverage_gaps(inventory_subnotes, entries)
        # Only note 3 partially covered; note 9 cited nothing so it is NOT a
        # sub-note gap (top-level coverage owns that).
        assert [g["note_num"] for g in gaps] == [3]
        assert gaps[0]["missing_subnote_refs"] == ["(a)"]


class TestFactorySurfacesStructuralFindings:
    """Both new candidate families must reach `context` (and thus the prompt)
    through the factory, the same way coverage_gaps does."""

    def test_factory_surfaces_collisions_and_subnote_gaps(self, tmp_path):
        from notes.validator_agent import create_notes_validator_agent

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Notes-SummaryofAccPol")
        wb.create_sheet("Notes-Listofnotes")
        merged = tmp_path / "merged.xlsx"
        wb.save(str(merged))

        sidecar = tmp_path / "NOTES_LIST_OF_NOTES_filled_payloads.json"
        sidecar.write_text(json.dumps([
            {"sheet": "Notes-Listofnotes", "row": 49, "col": 2,
             "row_label": "Disclosure of fair value information",
             "source_note_refs": ["4.1", "20.7"],
             "content_preview": "fv"},
            {"sheet": "Notes-SummaryofAccPol", "row": 42, "col": 2,
             "row_label": "Description of accounting policy for leases",
             "source_note_refs": ["3", "3.3", "(b)"],
             "content_preview": "leases"},
        ]), encoding="utf-8")

        _agent, _deps, ctx = create_notes_validator_agent(
            merged_workbook_path=str(merged),
            pdf_path=str(tmp_path / "x.pdf"),
            sidecar_paths=[str(sidecar)],
            filing_level="company", filing_standard="mfrs",
            model=TestModel(), output_dir=str(tmp_path),
            inventory_note_nums=[3, 4, 20],
            inventory_subnotes={3: ["3.3", "(a)", "(b)"]},
        )
        assert len(ctx["row_collisions"]) == 1
        assert ctx["row_collisions"][0]["note_nums"] == [4, 20]
        assert len(ctx["subnote_gaps"]) == 1
        assert ctx["subnote_gaps"][0]["note_num"] == 3
        assert "(a)" in ctx["subnote_gaps"][0]["missing_subnote_refs"]

    def test_prompt_block_renders_both_families(self):
        """The candidates must render into the prompt the model sees — pinned
        via the pure helper the factory delegates to (version-independent)."""
        from notes.validator_agent import build_structural_findings_block

        block = build_structural_findings_block(
            row_collisions=[{
                "sheet": "Notes-Listofnotes", "row": 49,
                "row_label": "Disclosure of fair value information",
                "note_nums": [4, 20],
                "source_note_refs": ["4.1", "20.7"],
                "content_preview": "fv",
            }],
            subnote_gaps=[{
                "note_num": 3,
                "missing_subnote_refs": ["(a)"],
                "cited_subnote_refs": ["3.3", "(b)"],
                "all_subnote_refs": ["3.3", "(a)", "(b)"],
            }],
        )
        assert "SAME-SHEET ROW COLLISIONS" in block
        assert "SUB-NOTE COVERAGE GAPS" in block
        assert "Disclosure of fair value information" in block
        assert "(a)" in block

    def test_prompt_block_empty_when_no_findings(self):
        from notes.validator_agent import build_structural_findings_block
        assert build_structural_findings_block([], []) == ""


class TestRewriteCellTool:
    """Step 5.2: `rewrite_cell` tool."""

    def test_rewrite_cell_clears_content_and_evidence(self, tmp_path):
        from notes.validator_agent import (
            NotesValidatorAgentDeps, _rewrite_cell_impl,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notes-SummaryofAccPol"
        ws.cell(row=10, column=2, value="Some prose content")
        ws.cell(row=10, column=4, value="Page 20")  # evidence col on company
        path = tmp_path / "merged.xlsx"
        wb.save(str(path))

        deps = NotesValidatorAgentDeps(
            merged_workbook_path=str(path),
            pdf_path="/tmp/x.pdf",
            sidecar_paths=[],
            filing_level="company",
            filing_standard="mfrs",
            output_dir=str(tmp_path),
            model=TestModel(),
        )
        msg = _rewrite_cell_impl(
            merged_workbook_path=str(path),
            filing_level="company",
            sheet="Notes-SummaryofAccPol",
            row=10,
            col=2,
            content="",
            evidence=None,
            deps=deps,
        )
        assert "cleared" in msg

        wb2 = openpyxl.load_workbook(str(path))
        ws2 = wb2["Notes-SummaryofAccPol"]
        assert ws2.cell(row=10, column=2).value is None
        # Evidence cleared when the primary cell was cleared.
        assert ws2.cell(row=10, column=4).value is None

    def test_rewrite_cell_refuses_formula(self, tmp_path):
        from notes.validator_agent import (
            NotesValidatorAgentDeps, _rewrite_cell_impl,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notes-SummaryofAccPol"
        ws.cell(row=5, column=2, value="=SUM(B1:B4)")
        path = tmp_path / "m.xlsx"
        wb.save(str(path))

        deps = NotesValidatorAgentDeps(
            merged_workbook_path=str(path),
            pdf_path="/tmp/x.pdf",
            sidecar_paths=[],
            filing_level="company",
            filing_standard="mfrs",
            output_dir=str(tmp_path),
            model=TestModel(),
        )
        msg = _rewrite_cell_impl(
            merged_workbook_path=str(path),
            filing_level="company",
            sheet="Notes-SummaryofAccPol",
            row=5, col=2, content="new", evidence=None,
            deps=deps,
        )
        assert "Refusing to overwrite formula cell" in msg


class TestWorkbookIoRaceSafety:
    """Windows EOFError race (2026-05-29): pydantic-ai runs batched tool
    calls on parallel worker threads, so read_cell / rewrite_cell can touch
    the same merged workbook simultaneously. openpyxl's in-place, non-atomic
    save left concurrent loaders reading a truncated zip → EOFError. The
    fix is a per-run io_lock + atomic (tempfile + os.replace) save."""

    def test_deps_carry_an_io_lock(self, tmp_path):
        import threading

        from notes.validator_agent import NotesValidatorAgentDeps

        deps = NotesValidatorAgentDeps(
            merged_workbook_path=str(tmp_path / "m.xlsx"),
            pdf_path="/tmp/x.pdf",
            sidecar_paths=[],
            filing_level="company",
            filing_standard="mfrs",
            output_dir=str(tmp_path),
            model=TestModel(),
        )
        # A lock instance, not the class — must be a usable mutex.
        assert isinstance(deps.io_lock, type(threading.Lock()))

    def test_atomic_save_leaves_no_truncated_file(self, tmp_path):
        """os.replace means a crashing save never overwrites the good file."""
        from notes.validator_agent import _atomic_save_workbook

        wb = openpyxl.Workbook()
        wb.active["A1"] = "good"
        path = tmp_path / "wb.xlsx"
        wb.save(str(path))

        # A workbook whose save blows up mid-stream must not clobber the
        # existing file, and must not leave a stray .xlsx tempfile behind.
        class _Boom(openpyxl.Workbook):
            def save(self, *a, **k):
                raise RuntimeError("save exploded")

        boom = _Boom()
        with pytest.raises(RuntimeError):
            _atomic_save_workbook(boom, str(path))

        # Original file untouched and still readable.
        assert openpyxl.load_workbook(str(path)).active["A1"].value == "good"
        leftovers = list(tmp_path.glob("*.xlsx"))
        assert leftovers == [path], f"tempfile leaked: {leftovers}"

    def test_concurrent_reads_and_writes_never_see_truncated_zip(self, tmp_path):
        """Hammer rewrite_cell from several threads while another thread
        loads the same workbook in a loop. Pre-fix this raised EOFError /
        BadZipFile intermittently; with io_lock + atomic save it is clean."""
        import threading
        import zipfile

        from notes.validator_agent import (
            NotesValidatorAgentDeps, _rewrite_cell_impl,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Notes-Listofnotes"
        for r in range(2, 12):
            ws.cell(row=r, column=2, value=f"row {r}")
        path = tmp_path / "merged.xlsx"
        wb.save(str(path))

        deps = NotesValidatorAgentDeps(
            merged_workbook_path=str(path),
            pdf_path="/tmp/x.pdf",
            sidecar_paths=[],
            filing_level="company",
            filing_standard="mfrs",
            output_dir=str(tmp_path),
            model=TestModel(),
        )

        errors: list[BaseException] = []
        stop = threading.Event()

        def writer(row: int):
            try:
                for i in range(8):
                    _rewrite_cell_impl(
                        merged_workbook_path=str(path),
                        filing_level="company",
                        sheet="Notes-Listofnotes",
                        row=row, col=2, content=f"v{i}", evidence=None,
                        deps=deps,
                    )
            except BaseException as e:  # noqa: BLE001
                errors.append(e)
            finally:
                stop.set()

        def reader():
            try:
                while not stop.is_set():
                    with deps.io_lock:
                        openpyxl.load_workbook(str(path)).close()
            except (EOFError, zipfile.BadZipFile, OSError) as e:
                errors.append(e)

        threads = [threading.Thread(target=writer, args=(r,))
                   for r in range(2, 7)]
        threads.append(threading.Thread(target=reader))
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=30)

        assert not errors, f"concurrent workbook IO raced: {errors!r}"
        # File survived and is still a valid workbook.
        assert openpyxl.load_workbook(str(path)) is not None


class TestServerHook:
    """Step 5.5: the server invokes the validator when both sheets ran."""

    @pytest.mark.asyncio
    async def test_notes_validator_hook_short_circuits_when_single_sheet(self, tmp_path):
        """Trigger condition from plan: run only when BOTH sheet 11 AND
        sheet 12 appear in the notes output."""
        import asyncio
        from server import _run_notes_validator_pass

        queue: asyncio.Queue = asyncio.Queue()
        outcome = await _run_notes_validator_pass(
            merged_workbook_path=str(tmp_path / "m.xlsx"),
            pdf_path=str(tmp_path / "x.pdf"),
            notes_template_outputs={"ACC_POLICIES": str(tmp_path / "a.xlsx")},
            filing_level="company",
            filing_standard="mfrs",
            model=TestModel(),
            output_dir=str(tmp_path),
            event_queue=queue,
        )
        assert outcome["invoked"] is False
        # No events should have been emitted on the short-circuit path.
        assert queue.empty()

    @pytest.mark.asyncio
    async def test_notes_validator_emits_skip_when_both_sheets_but_no_candidates(self, tmp_path):
        """Bug 4a — when both sheets ran but the detectors find nothing to
        resolve, the validator still has to surface a terminal event so the
        frontend tab doesn't hang on 'Waiting for the agent to start…'.

        Seeds a minimal but valid pair of sidecar payloads (empty arrays)
        so both the outer and inner gates pass and we hit the
        'no duplicates + no overlap' short-circuit at server.py:494.
        """
        import asyncio
        from server import _run_notes_validator_pass
        from notes.writer import payload_sidecar_path

        # Both sheets present in notes_template_outputs — outer gate passes.
        acc_xlsx = tmp_path / "NOTES_ACC_POLICIES_filled.xlsx"
        lon_xlsx = tmp_path / "NOTES_LIST_OF_NOTES_filled.xlsx"
        # Workbooks themselves only need to be loadable by the factory.
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Notes-SummaryofAccPol")
        wb.create_sheet("Notes-Listofnotes")
        merged_path = tmp_path / "merged.xlsx"
        wb.save(str(merged_path))

        # Empty sidecars — so detectors report zero candidates.
        for xlsx_path in (acc_xlsx, lon_xlsx):
            payload_sidecar_path(str(xlsx_path)).write_text(
                json.dumps([]), encoding="utf-8",
            )

        queue: asyncio.Queue = asyncio.Queue()
        outcome = await _run_notes_validator_pass(
            merged_workbook_path=str(merged_path),
            pdf_path=str(tmp_path / "x.pdf"),
            notes_template_outputs={
                "ACC_POLICIES": str(acc_xlsx),
                "LIST_OF_NOTES": str(lon_xlsx),
            },
            filing_level="company",
            filing_standard="mfrs",
            model=TestModel(),
            output_dir=str(tmp_path),
            event_queue=queue,
        )
        assert outcome["invoked"] is False
        # Item 15: elapsed is stamped even on the skip path.
        assert outcome["elapsed_seconds"] >= 0

        events: list[dict] = []
        while not queue.empty():
            events.append(queue.get_nowait())

        types = [e["event"] for e in events]
        # Must emit both a status (to seed the tab / show the skip reason)
        # and a terminal complete (so the badge flips off "running").
        assert "status" in types, f"expected status event, got: {types}"
        assert "complete" in types, f"expected complete event, got: {types}"

        # Complete must be a success terminal — the tab renders green, not red.
        complete_event = next(e for e in events if e["event"] == "complete")
        assert complete_event["data"]["success"] is True
        # agent_id must be present so the frontend can route to the tab.
        assert complete_event["data"].get("agent_id")

    @pytest.mark.asyncio
    async def test_coverage_gaps_only_invokes_validator(self, tmp_path, monkeypatch):
        """N3 Stage 1 (peer-review HIGH): a run with missed inventory notes but
        NO duplicate/overlap candidates must still invoke the validator so it
        can investigate the gaps — not short-circuit."""
        import asyncio
        import agent_runner
        from server import _run_notes_validator_pass
        from notes.writer import payload_sidecar_path

        acc_xlsx = tmp_path / "NOTES_ACC_POLICIES_filled.xlsx"
        lon_xlsx = tmp_path / "NOTES_LIST_OF_NOTES_filled.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Notes-SummaryofAccPol")
        wb.create_sheet("Notes-Listofnotes")
        merged_path = tmp_path / "merged.xlsx"
        wb.save(str(merged_path))
        # Empty sidecars → zero dup/overlap candidates AND nothing written, so
        # inventory note 99 is an uncovered gap.
        for xlsx_path in (acc_xlsx, lon_xlsx):
            payload_sidecar_path(str(xlsx_path)).write_text(
                json.dumps([]), encoding="utf-8")

        # Stub the heavy agent loop — we're testing the invoke DECISION, not
        # the agent's behaviour.
        async def _noop_loop(*a, **k):
            return None
        monkeypatch.setattr(agent_runner, "run_agent_loop", _noop_loop)

        queue: asyncio.Queue = asyncio.Queue()
        outcome = await _run_notes_validator_pass(
            merged_workbook_path=str(merged_path),
            pdf_path=str(tmp_path / "x.pdf"),
            notes_template_outputs={
                "ACC_POLICIES": str(acc_xlsx),
                "LIST_OF_NOTES": str(lon_xlsx),
            },
            filing_level="company", filing_standard="mfrs",
            model=TestModel(), output_dir=str(tmp_path),
            event_queue=queue, inventory_note_nums=[99],
        )
        assert outcome["context"]["coverage_gaps"] == [99]
        assert outcome["invoked"] is True

    @pytest.mark.asyncio
    async def test_validator_loop_spec_does_not_bound_inner_streams(self, tmp_path, monkeypatch):
        """Code-review pin (2026-06-13): the validator's rewrite_cell does a
        LOCKED workbook load+save — a legitimately long tool call the 180s
        per-turn timeout must NOT cancel mid-execution. The pass's
        AgentLoopSpec must keep the pre-migration semantics
        (bound_inner_streams=False, the notes/coordinator.py notes_spec
        opt-out precedent)."""
        import asyncio
        import agent_runner
        from server import _run_notes_validator_pass
        from notes.writer import payload_sidecar_path

        acc_xlsx = tmp_path / "NOTES_ACC_POLICIES_filled.xlsx"
        lon_xlsx = tmp_path / "NOTES_LIST_OF_NOTES_filled.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Notes-SummaryofAccPol")
        wb.create_sheet("Notes-Listofnotes")
        merged_path = tmp_path / "merged.xlsx"
        wb.save(str(merged_path))
        # Empty sidecars + an uncovered inventory note → the invoke path.
        for xlsx_path in (acc_xlsx, lon_xlsx):
            payload_sidecar_path(str(xlsx_path)).write_text(
                json.dumps([]), encoding="utf-8")

        captured: dict = {}

        async def _capture_loop(agent_run, deps, spec, emit, turn_records):
            captured["spec"] = spec
            return None

        monkeypatch.setattr(agent_runner, "run_agent_loop", _capture_loop)

        queue: asyncio.Queue = asyncio.Queue()
        outcome = await _run_notes_validator_pass(
            merged_workbook_path=str(merged_path),
            pdf_path=str(tmp_path / "x.pdf"),
            notes_template_outputs={
                "ACC_POLICIES": str(acc_xlsx),
                "LIST_OF_NOTES": str(lon_xlsx),
            },
            filing_level="company", filing_standard="mfrs",
            model=TestModel(), output_dir=str(tmp_path),
            event_queue=queue, inventory_note_nums=[99],
        )
        assert outcome["invoked"] is True
        assert "spec" in captured, "validator pass should reach run_agent_loop"
        assert captured["spec"].bound_inner_streams is False

    @pytest.mark.asyncio
    async def test_no_gaps_no_candidates_still_short_circuits(self, tmp_path):
        """The skip path is preserved when there are neither candidates nor
        coverage gaps (every inventory note covered)."""
        import asyncio
        from server import _run_notes_validator_pass
        from notes.writer import payload_sidecar_path

        acc_xlsx = tmp_path / "NOTES_ACC_POLICIES_filled.xlsx"
        lon_xlsx = tmp_path / "NOTES_LIST_OF_NOTES_filled.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Notes-SummaryofAccPol")
        wb.create_sheet("Notes-Listofnotes")
        merged_path = tmp_path / "merged.xlsx"
        wb.save(str(merged_path))
        # Sidecar covers note 5; inventory is only note 5 → no gaps.
        payload_sidecar_path(str(lon_xlsx)).write_text(json.dumps([
            {"sheet": "Notes-Listofnotes", "row": 5, "col": 2,
             "source_note_refs": ["5"], "content_preview": "x"},
        ]), encoding="utf-8")
        payload_sidecar_path(str(acc_xlsx)).write_text(
            json.dumps([]), encoding="utf-8")

        queue: asyncio.Queue = asyncio.Queue()
        outcome = await _run_notes_validator_pass(
            merged_workbook_path=str(merged_path),
            pdf_path=str(tmp_path / "x.pdf"),
            notes_template_outputs={
                "ACC_POLICIES": str(acc_xlsx),
                "LIST_OF_NOTES": str(lon_xlsx),
            },
            filing_level="company", filing_standard="mfrs",
            model=TestModel(), output_dir=str(tmp_path),
            event_queue=queue, inventory_note_nums=[5],
        )
        assert outcome["context"]["coverage_gaps"] == []
        assert outcome["invoked"] is False


def _agent_tool_names(agent) -> set[str]:
    names: set[str] = set()
    toolsets = getattr(agent, "toolsets", None) or []
    for ts in toolsets:
        tools = getattr(ts, "tools", None) or {}
        for tname in tools:
            names.add(tname)
    if names:
        return names
    legacy = getattr(agent, "_function_tools", None)
    if isinstance(legacy, dict):
        return set(legacy.keys())
    return names
