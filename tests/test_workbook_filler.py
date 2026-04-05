from pathlib import Path

import openpyxl
import pytest

from tools.fill_workbook import fill_workbook, FillResult, _normalize_label, _find_row_by_label, _build_label_index


def _make_template(tmp_path):
    path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Right-of-use assets"
    ws["B1"] = None
    ws["C1"] = None
    ws["A2"] = "Total assets"
    ws["B2"] = "=B1+100"
    ws["C2"] = "=C1+200"
    wb.save(str(path))
    return path


def _make_template_with_duplicates(tmp_path):
    """Template with duplicate labels in different sections (like the real MBRS template)."""
    path = tmp_path / "template_dup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOFP"

    # Non-current section
    ws["A1"] = "Non-current liabilities"
    ws["A2"] = "Lease liabilities"
    ws["B2"] = None
    ws["A3"] = "Borrowings"
    ws["B3"] = None

    # Current section
    ws["A5"] = "Current liabilities"
    ws["A6"] = "Lease liabilities"  # duplicate label
    ws["B6"] = None
    ws["A7"] = "Borrowings"  # duplicate label
    ws["B7"] = None

    wb.save(str(path))
    return path


def test_fill_by_label(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "Sheet", "field_label": "Right-of-use assets", "col": 2, "value": 191518}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert result.success
    assert result.fields_written == 1

    wb = openpyxl.load_workbook(output)
    assert wb.active["B1"].value == 191518
    wb.close()


def test_fill_by_label_fuzzy(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "Sheet", "field_label": "Right of use assets", "col": 2, "value": 100}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert result.success
    assert result.fields_written == 1

    wb = openpyxl.load_workbook(output)
    assert wb.active["B1"].value == 100
    wb.close()


def test_fill_by_row_fallback(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "Sheet", "row": 1, "col": 2, "value": 191518}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert result.success
    assert result.fields_written == 1

    wb = openpyxl.load_workbook(output)
    assert wb.active["B1"].value == 191518
    wb.close()


def test_fill_duplicate_label_with_section(tmp_path):
    """Section hint disambiguates duplicate labels."""
    template = _make_template_with_duplicates(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "SOFP", "field_label": "Lease liabilities", "section": "current liabilities", "col": 2, "value": 36148}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert result.success
    assert result.fields_written == 1

    wb = openpyxl.load_workbook(output)
    # Should write to row 6 (current), not row 2 (non-current)
    assert wb.active["B6"].value == 36148
    assert wb.active["B2"].value is None
    wb.close()


def test_fill_duplicate_label_noncurrent_section(tmp_path):
    """Section hint picks the non-current occurrence."""
    template = _make_template_with_duplicates(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "SOFP", "field_label": "Lease liabilities", "section": "non-current liabilities", "col": 2, "value": 160404}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert result.success
    wb = openpyxl.load_workbook(output)
    assert wb.active["B2"].value == 160404
    assert wb.active["B6"].value is None
    wb.close()


def test_fill_duplicate_label_both_sections(tmp_path):
    """Can fill both occurrences of a duplicate label with different section hints."""
    template = _make_template_with_duplicates(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = """{
        "fields": [
            {"sheet": "SOFP", "field_label": "Lease liabilities", "section": "non-current liabilities", "col": 2, "value": 160404},
            {"sheet": "SOFP", "field_label": "Lease liabilities", "section": "current liabilities", "col": 2, "value": 36148}
        ]
    }"""
    result = fill_workbook(str(template), output, fields_json)

    assert result.success
    assert result.fields_written == 2

    wb = openpyxl.load_workbook(output)
    assert wb.active["B2"].value == 160404
    assert wb.active["B6"].value == 36148
    wb.close()


def test_fill_workbook_preserves_formulas(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "Sheet", "field_label": "Right-of-use assets", "col": 2, "value": 50}]}'
    fill_workbook(str(template), output, fields_json)

    wb = openpyxl.load_workbook(output, data_only=False)
    assert str(wb.active["B2"].value).startswith("=")
    wb.close()


def test_fill_workbook_refuses_formula_cells(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "Sheet", "field_label": "Total assets", "col": 2, "value": 999}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert not result.success
    assert "formula" in " ".join(result.errors).lower()


def test_fill_workbook_missing_template():
    result = fill_workbook("/nonexistent.xlsx", "/tmp/out.xlsx", '{"fields": []}')
    assert not result.success
    assert "not found" in result.errors[0].lower()


def test_fill_workbook_invalid_json(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    result = fill_workbook(str(template), output, "not json")
    assert not result.success
    assert "Invalid JSON" in result.errors[0]


def test_fill_workbook_wrong_sheet(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "NonExistent", "field_label": "Right-of-use assets", "col": 2, "value": 10}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert "NonExistent" in " ".join(result.errors)


def test_fill_workbook_multiple_fields(tmp_path):
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = """{
        "fields": [
            {"sheet": "Sheet", "field_label": "Right-of-use assets", "col": 2, "value": 100},
            {"sheet": "Sheet", "field_label": "Right-of-use assets", "col": 3, "value": 200}
        ]
    }"""
    result = fill_workbook(str(template), output, fields_json)

    assert result.success
    assert result.fields_written == 2

    wb = openpyxl.load_workbook(output)
    assert wb.active["B1"].value == 100
    assert wb.active["C1"].value == 200
    wb.close()


def test_fill_label_not_found(tmp_path):
    """Unmatched label produces an actionable error."""
    template = _make_template(tmp_path)
    output = str(tmp_path / "filled.xlsx")

    fields_json = '{"fields": [{"sheet": "Sheet", "field_label": "Nonexistent field XYZ", "col": 2, "value": 1}]}'
    result = fill_workbook(str(template), output, fields_json)

    assert not result.success
    assert "No matching label" in " ".join(result.errors)
    # Error should suggest checking read_template()
    assert "read_template" in " ".join(result.errors).lower()


def test_normalize_label():
    assert _normalize_label("*Property, plant and equipment") == "property, plant and equipment"
    assert _normalize_label("  Right-of-use assets  ") == "right-of-use assets"
    assert _normalize_label("Total assets") == "total assets"
