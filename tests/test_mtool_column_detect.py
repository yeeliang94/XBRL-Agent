"""Column detection for an uploaded mTool template (mtool/column_detect.py).

Two halves. The positional fallback (our own marker-less templates) is the
original behaviour, now required to declare itself and ask for confirmation
unless the workbook's layout is on file. The semantic path reads a real mTool
template's own marker rows — which is what peer-review finding 3 demanded, and
what stops the detector from mapping a period role onto a share-class column.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from mtool.column_detect import (
    describe_template, detect_column_map, fingerprint_workbook,
    needs_confirmation, overall_confidence, parse_unit_scale,
    unit_scale_mismatches)
from mtool.offline_fill import load_workbook_entries

REPO = Path(__file__).resolve().parent.parent
SOFP = REPO / "XBRL-template-MFRS" / "Company" / "01-SOFP-CuNonCu.xlsx"
SOFP_GROUP = REPO / "XBRL-template-MFRS" / "Group" / "01-SOFP-CuNonCu.xlsx"
# The one real SSM-issued mTool workbook in the repo — the only fixture that
# carries mTool's marker rows.
REAL_MTOOL = REPO / "data" / "MBRS_test.xlsx"

_real_mtool = pytest.mark.skipif(
    not REAL_MTOOL.exists(), reason="real mTool fixture not present")


def test_detects_our_template_layout():
    # Our sub-sheet: labels in A, values from B.
    doc = {"sheets": {"SOFP-Sub-CuNonCu": {
        "label_column": None,
        "columns": {"current_year": None, "prior_year": None}}}}
    result = detect_column_map(str(SOFP), doc)
    sheet = result["SOFP-Sub-CuNonCu"]
    assert sheet["label_column"] == "A"
    assert sheet["columns"] == {"current_year": "B", "prior_year": "C"}
    assert sheet["confidence"] == "high"


def test_detects_real_mtool_style_layout(tmp_path):
    # Mimic the real mTool layout the Windows agent found: labels in D,
    # values in E/F. A..C are blank spacer columns.
    wb = Workbook()
    ws = wb.active
    ws.title = "SOFP-Sub-CuNonCu"
    labels = ["Freehold land", "Long term leasehold land", "Buildings",
              "Motor vehicles", "Machinery", "Plant and equipment",
              "Office equipment", "Computer software"]
    for i, label in enumerate(labels, start=3):
        ws[f"D{i}"] = label
    path = tmp_path / "mtool.xlsx"
    wb.save(path)

    doc = {"sheets": {"SOFP-Sub-CuNonCu": {
        "label_column": None,
        "columns": {"current_year": None, "prior_year": None}}}}
    result = detect_column_map(str(path), doc)
    sheet = result["SOFP-Sub-CuNonCu"]
    assert sheet["label_column"] == "D"
    assert sheet["columns"] == {"current_year": "E", "prior_year": "F"}
    assert sheet["confidence"] == "high"


def test_group_roles_are_never_auto_accepted(tmp_path):
    """Finding 3's headline rule: never auto-proceed on a four-column shape.

    The positional guess is still OFFERED (it seeds the confirm dialog), but
    the map must demand confirmation — a swapped Group/Company pair produces a
    plausible, wrong filing that nothing downstream would catch.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for i in range(3, 12):
        ws[f"A{i}"] = f"Line item {i}"
    path = tmp_path / "g.xlsx"
    wb.save(path)
    doc = {"sheets": {"S": {"label_column": None, "columns": {
        "company_prior_year": None, "group_current_year": None,
        "company_current_year": None, "group_prior_year": None}}}}
    sheet = detect_column_map(str(path), doc)["S"]
    # Canonical order: group CY, group PY, company CY, company PY -> B,C,D,E
    assert sheet["columns"] == {
        "group_current_year": "B", "group_prior_year": "C",
        "company_current_year": "D", "company_prior_year": "E"}
    assert sheet["requires_confirmation"] is True
    assert sheet["confidence"] == "low"
    assert any("group" in n for n in sheet["notes"])


def test_group_layout_of_a_known_template_still_needs_confirmation():
    """Even our OWN Group template — being 'known' doesn't make a positional
    four-column split verified."""
    doc = {"sheets": {"SOFP-Sub-CuNonCu": {"label_column": None, "columns": {
        "group_current_year": None, "group_prior_year": None,
        "company_current_year": None, "company_prior_year": None}}}}
    result = detect_column_map(str(SOFP_GROUP), doc)
    assert needs_confirmation(result) is True


def test_unknown_template_needs_confirmation_even_when_confident(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for i in range(3, 40):
        ws[f"A{i}"] = f"Line item {i}"
    path = tmp_path / "unknown.xlsx"
    wb.save(path)
    doc = {"sheets": {"S": {"label_column": None,
                            "columns": {"current_year": None}}}}
    sheet = detect_column_map(str(path), doc)["S"]
    assert sheet["basis"] == "positional"
    assert sheet["requires_confirmation"] is True
    assert any("haven't seen" in n for n in sheet["notes"])


def test_known_repo_template_may_proceed_unattended():
    doc = {"sheets": {"SOFP-Sub-CuNonCu": {
        "label_column": None,
        "columns": {"current_year": None, "prior_year": None}}}}
    result = detect_column_map(str(SOFP), doc)
    assert needs_confirmation(result) is False


# ------------------------------------------------------------- fingerprints

def test_fingerprint_is_stable_and_registered():
    _, data, _ = load_workbook_entries(str(SOFP))
    fp = fingerprint_workbook(data)
    assert fingerprint_workbook(data) == fp  # deterministic
    entry = describe_template(fp)
    assert entry is not None, "repo templates must be on file (Step 18)"
    assert entry["source"] == "generated"
    assert entry["vouched_by"]  # provenance is recorded, not implied


def test_company_and_group_templates_fingerprint_differently():
    _, company, _ = load_workbook_entries(str(SOFP))
    _, group, _ = load_workbook_entries(str(SOFP_GROUP))
    assert fingerprint_workbook(company) != fingerprint_workbook(group)


@_real_mtool
def test_real_mtool_template_is_registered_and_distinct():
    _, data, _ = load_workbook_entries(str(REAL_MTOOL))
    entry = describe_template(fingerprint_workbook(data))
    assert entry is not None
    assert entry["source"] == "ssm-mtool"


# --------------------------------------------------------------- semantics

@_real_mtool
def test_real_mtool_periods_are_matched_by_date_not_position():
    """The core of finding 3. In the real template E is 2024 and F is 2023, so
    the roles must come from the ``#ENDT#`` dates — the same answer position
    would give, but for a reason that survives a reordered template."""
    doc = {"sheets": {"SOFP-CuNonCu": {
        "label_column": None,
        "columns": {"current_year": None, "prior_year": None}}}}
    sheet = detect_column_map(str(REAL_MTOOL), doc)["SOFP-CuNonCu"]
    assert sheet["basis"] == "semantic"
    assert sheet["label_column"] == "D"          # from mTool's own #PRIM#
    assert sheet["columns"] == {"current_year": "E", "prior_year": "F"}
    assert sheet["confidence"] == "high"
    assert sheet["requires_confirmation"] is False
    assert sheet["period_columns"]["E"] == "01/01/2024 - 31/12/2024"


@_real_mtool
def test_dimensional_sheet_is_refused_not_mapped_by_position():
    """``Notes-Issuedcapital`` lays its columns out as share CLASSES, all for
    the same period. Positional assignment would have written the current year
    into "Ordinary shares" and the prior year into "Redeemable preference
    shares" — a plausible, wrong filing. It must be refused instead."""
    doc = {"sheets": {"Notes-Issuedcapital": {
        "label_column": None,
        "columns": {"current_year": None, "prior_year": None}}}}
    sheet = detect_column_map(str(REAL_MTOOL), doc)["Notes-Issuedcapital"]
    assert sheet["dimensional"] is True
    assert sheet["requires_confirmation"] is True
    assert sheet["confidence"] == "low"
    # Every requested role is OFFERED (the confirm dialog renders an input per
    # key) but BLANK — no guessed period column (peer review, 2026-08-05: the
    # old empty dict left the operator a confirmation with nothing to fill).
    assert sheet["columns"] == {"current_year": "", "prior_year": ""}
    assert any("categories" in n for n in sheet["notes"])


@_real_mtool
def test_declared_unit_scale_is_read_and_reported():
    doc = {"sheets": {"SOFP-CuNonCu": {
        "label_column": None,
        "columns": {"current_year": None, "prior_year": None}}}}
    result = detect_column_map(str(REAL_MTOOL), doc)
    scales = result["SOFP-CuNonCu"]["declared_unit_scales"]
    assert scales == {"E": "thousands", "F": "thousands"}
    # Matching denomination -> silence; mismatch -> a visible warning.
    assert unit_scale_mismatches(result, "thousands") == []
    mismatch = unit_scale_mismatches(result, "units")
    assert len(mismatch) == 2
    assert mismatch[0]["template_declares"] == "thousands"


@pytest.mark.parametrize("declared,expected", [
    ("MYR'000", "thousands"),
    ("MYR", "units"),
    ("", None),
    ("MYR'000000", "millions"),
])
def test_parse_unit_scale(declared, expected):
    assert parse_unit_scale(declared) == expected


def test_missing_sheet_is_low_confidence(tmp_path):
    wb = Workbook()
    wb.active.title = "Present"
    wb.active["A3"] = "x"
    path = tmp_path / "w.xlsx"
    wb.save(path)
    doc = {"sheets": {"Absent": {"label_column": None,
                                 "columns": {"current_year": None}}}}
    result = detect_column_map(str(path), doc)
    assert result["Absent"]["label_column"] is None
    assert result["Absent"]["confidence"] == "low"
    assert overall_confidence(result) == "low"


def test_scarce_labels_are_low_confidence(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["A1"] = "only one label"
    path = tmp_path / "w.xlsx"
    wb.save(path)
    doc = {"sheets": {"S": {"label_column": None,
                            "columns": {"current_year": None}}}}
    result = detect_column_map(str(path), doc)
    assert result["S"]["confidence"] == "low"


# ------------------------------------------- peer-review fixes (2026-08-05)

def _marker_workbook(tmp_path, *, dates=("31/12/2024", "31/12/2023")):
    """A synthetic workbook carrying mTool's own marker rows — semantic path,
    but a fingerprint no registry has ever seen."""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["D2"] = "#PRIM#"
    ws["A3"] = "#ENDT#"
    ws["E3"], ws["F3"] = dates
    ws["A4"] = "#STDTENDTDATE#"
    ws["E4"] = "01/01/2024 - 31/12/2024"
    ws["F4"] = "01/01/2023 - 31/12/2023"
    for i in range(6, 16):
        ws[f"D{i}"] = f"Line item {i}"
    path = tmp_path / "unknown_semantic.xlsx"
    wb.save(path)
    return path


def test_unknown_semantic_template_needs_confirmation(tmp_path):
    """An unknown fingerprint always needs a human — markers make the reading
    semantic, not corroborated. Before this pin, a marker-bearing workbook
    nobody had vouched for sailed through unattended (peer review,
    2026-08-05)."""
    path = _marker_workbook(tmp_path)
    doc = {"sheets": {"S": {"label_column": None, "columns": {
        "current_year": None, "prior_year": None}}}}
    sheet = detect_column_map(str(path), doc)["S"]
    assert sheet["basis"] == "semantic"
    # The reading itself is confident AND correct...
    assert sheet["columns"]["current_year"] == "E"
    assert sheet["columns"]["prior_year"] == "F"
    # ...but permission is a separate question (unknown fingerprint).
    assert sheet["requires_confirmation"] is True
    assert needs_confirmation({"S": sheet}) is True
    assert any("haven't seen" in n for n in sheet["notes"])


def test_group_semantic_layout_offers_every_role_blank(tmp_path):
    """A Group filing on a semantic sheet: confirmation required, and all four
    roles must be PRESENT (blank) so the confirm dialog has fields to edit."""
    path = _marker_workbook(tmp_path)
    doc = {"sheets": {"S": {"label_column": None, "columns": {
        "group_current_year": None, "group_prior_year": None,
        "company_current_year": None, "company_prior_year": None}}}}
    sheet = detect_column_map(str(path), doc)["S"]
    assert sheet["basis"] == "semantic"
    assert sheet["requires_confirmation"] is True
    assert set(sheet["columns"]) == {
        "group_current_year", "group_prior_year",
        "company_current_year", "company_prior_year"}
    assert all(v == "" for v in sheet["columns"].values())


def test_positional_no_label_column_still_offers_roles(tmp_path):
    """Even when nothing can be proposed, the requested roles appear (blank)
    so the operator can supply the mapping by hand."""
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws["B3"] = 123  # numbers only — no label column to find
    path = tmp_path / "numbersonly.xlsx"
    wb.save(path)
    doc = {"sheets": {"S": {"label_column": None, "columns": {
        "current_year": None, "prior_year": None}}}}
    sheet = detect_column_map(str(path), doc)["S"]
    assert sheet["basis"] == "positional"
    assert sheet["confidence"] == "low"
    assert sheet["columns"] == {"current_year": "", "prior_year": ""}
