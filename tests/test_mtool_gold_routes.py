"""Endpoint test for POST /api/benchmarks/from-mtool (Step C3).

Imports the live MFRS Company SOFP template into a temp DB, builds a synthetic
mTool-shaped workbook whose sheet name matches a render_sheet and whose col-A
labels are the real canonical labels, then ingests it through the endpoint.
(A real mTool file uses mTool's own sheet names — the render_sheet↔mTool sheet
mapping is the Windows-recon gate per gotcha #28; this pins the reverse-mapping
logic itself.)
"""
from __future__ import annotations

import importlib
import json
import sqlite3
import tempfile
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from statement_types import StatementType, template_path

_TEMPLATE_ID = "mfrs-company-sofp-cunoncu-v1"


def _import_company_sofp(db_path) -> str:
    from concept_model.importer import import_company_targets, import_template
    from concept_model.parser import parse_template

    tpath = template_path(StatementType.SOFP, "CuNonCu", "company", "mfrs")
    tree = parse_template(str(tpath))
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as fh:
        json.dump(tree.to_json(), fh, sort_keys=True)
        json_path = fh.name
    try:
        template_id = import_template(db_path, json_path)
    finally:
        Path(json_path).unlink(missing_ok=True)
    import_company_targets(db_path, template_id)
    return template_id


def _pick_leaves(db_path, template_id, n=6):
    """Return (render_sheet, [(label, uuid), ...]) for the render_sheet with the
    most distinct-labelled LEAF rows — so column detection is high-confidence."""
    conn = sqlite3.connect(str(db_path))
    rows = conn.execute(
        "SELECT render_sheet, canonical_label, concept_uuid FROM concept_nodes "
        "WHERE template_id = ? AND kind = 'LEAF'",
        (template_id,),
    ).fetchall()
    conn.close()
    by_sheet: dict[str, list] = {}
    seen: dict[str, set] = {}
    for sheet, label, uuid in rows:
        if not label or not label.strip():
            continue
        seen.setdefault(sheet, set())
        if label in seen[sheet]:
            continue  # skip duplicate labels within a sheet (ambiguous)
        seen[sheet].add(label)
        by_sheet.setdefault(sheet, []).append((label, uuid))
    best_sheet = max(by_sheet, key=lambda s: len(by_sheet[s]))
    return best_sheet, by_sheet[best_sheet][:n]


def _mtool_workbook_bytes(sheet, leaves, with_notes=True):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i, (label, _uuid) in enumerate(leaves, start=3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = 100 + i   # CY
        ws[f"C{i}"] = 90 + i    # PY
    if with_notes:
        fn = wb.create_sheet("+FootnoteTexts")
        fn["A1"] = "fn_1"
        fn["C1"] = "<p>Significant accounting policies.</p>"
    out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(out.name)
    data = Path(out.name).read_bytes()
    Path(out.name).unlink(missing_ok=True)
    return data


@pytest.fixture
def client(tmp_path, monkeypatch):
    db = tmp_path / "xbrl.db"
    monkeypatch.setenv("XBRL_OUTPUT_DIR", str(tmp_path))
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    monkeypatch.setenv("LLM_PROXY_URL", "")
    import server as srv

    importlib.reload(srv)
    srv.AUDIT_DB_PATH = db
    from db.schema import init_db

    init_db(db)
    template_id = _import_company_sofp(db)
    return TestClient(srv.app), db, template_id


def test_from_mtool_happy_path(client):
    tc, db, template_id = client
    sheet, leaves = _pick_leaves(db, template_id)
    data = _mtool_workbook_bytes(sheet, leaves)

    resp = tc.post(
        "/api/benchmarks/from-mtool",
        data={
            "name": "FINCO human mTool",
            "filing_standard": "mfrs",
            "filing_level": "company",
            "unit": "thousands",
            "template_ids": template_id,
            "document": "FINCO.pdf",
        },
        files={"file": ("human.xlsx", data,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["ok"] is True
    # Each of the N leaves emits 2 facts (CY+PY), scaled by 1000.
    assert body["ingested"] == len(leaves) * 2
    assert body["prose_notes_captured"] == 1

    # Gold is real: the grid shows filled cells scaled to thousands.
    bench_id = body["id"]
    grid = tc.get(f"/api/benchmarks/{bench_id}/concepts").json()
    filled = [c for c in grid["concepts"] if c["value"] is not None]
    assert filled, "expected gold values in the benchmark grid"
    assert any(abs(c["value"]) >= 100_000 for c in filled)  # ×1000 applied


def test_list_eval_templates_for_family(client):
    """Step C4: the mTool-gold form's variant picker enumerates imported face
    templates for a filing family."""
    tc, db, template_id = client
    resp = tc.get("/api/eval/templates?standard=mfrs&level=company")
    assert resp.status_code == 200, resp.text
    templates = resp.json()["templates"]
    ids = {t["template_id"] for t in templates}
    assert template_id in ids
    sofp = next(t for t in templates if t["template_id"] == template_id)
    assert sofp["statement"] == "SOFP"
    assert sofp["variant"]  # non-empty variant token

    # A family with nothing imported returns an empty list, not an error.
    empty = tc.get("/api/eval/templates?standard=mpers&level=group")
    assert empty.status_code == 200
    assert empty.json()["templates"] == []


def test_unit_is_mandatory(client):
    tc, db, template_id = client
    sheet, leaves = _pick_leaves(db, template_id)
    data = _mtool_workbook_bytes(sheet, leaves)
    resp = tc.post(
        "/api/benchmarks/from-mtool",
        data={"name": "x", "filing_standard": "mfrs", "filing_level": "company",
              "template_ids": template_id},  # no unit
        files={"file": ("h.xlsx", data, "application/octet-stream")},
    )
    assert resp.status_code == 422  # missing required Form field


def test_bad_unit_rejected(client):
    tc, db, template_id = client
    sheet, leaves = _pick_leaves(db, template_id)
    data = _mtool_workbook_bytes(sheet, leaves)
    resp = tc.post(
        "/api/benchmarks/from-mtool",
        data={"name": "x", "filing_standard": "mfrs", "filing_level": "company",
              "unit": "millions", "template_ids": template_id},
        files={"file": ("h.xlsx", data, "application/octet-stream")},
    )
    assert resp.status_code == 400
    assert "unit" in resp.json()["detail"].lower()


def test_no_matching_values_is_422(client):
    tc, db, template_id = client
    # A workbook whose labels match nothing → zero gold → 422.
    data = _mtool_workbook_bytes("SOFP", [
        (f"Totally off-template line {i}", None) for i in range(6)
    ])
    resp = tc.post(
        "/api/benchmarks/from-mtool",
        data={"name": "x", "filing_standard": "mfrs", "filing_level": "company",
              "unit": "full", "template_ids": template_id},
        files={"file": ("h.xlsx", data, "application/octet-stream")},
    )
    assert resp.status_code == 422


def test_wrong_shaped_column_map_is_400_not_500(client):
    """Syntactically-valid JSON with the wrong shape (e.g. []) is a user-fixable
    400, never a 500 inside ingest (peer-review LOW)."""
    tc, db, template_id = client
    sheet, leaves = _pick_leaves(db, template_id)
    data = _mtool_workbook_bytes(sheet, leaves)
    for bad in ("[]", "{}", '{"SOFP": 5}', '{"SOFP": {"columns": {}}}'):
        resp = tc.post(
            "/api/benchmarks/from-mtool",
            data={"name": "x", "filing_standard": "mfrs",
                  "filing_level": "company", "unit": "full",
                  "template_ids": template_id, "column_map": bad},
            files={"file": ("h.xlsx", data, "application/octet-stream")},
        )
        assert resp.status_code == 400, f"{bad} -> {resp.status_code}"


def test_bad_json_column_map_is_400(client):
    tc, db, template_id = client
    sheet, leaves = _pick_leaves(db, template_id)
    data = _mtool_workbook_bytes(sheet, leaves)
    resp = tc.post(
        "/api/benchmarks/from-mtool",
        data={"name": "x", "filing_standard": "mfrs", "filing_level": "company",
              "unit": "full", "template_ids": template_id, "column_map": "{not json"},
        files={"file": ("h.xlsx", data, "application/octet-stream")},
    )
    assert resp.status_code == 400
