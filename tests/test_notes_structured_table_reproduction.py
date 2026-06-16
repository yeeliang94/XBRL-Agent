"""Part C: structured notes sheets (Issued Capital / Related Party) also
reproduce the disclosed table as HTML into the top-level text-block row.

The numeric grid still fills as before; the reproduced table is an ADDITION
that lands as a prose cell (col B of the text-block row + a notes_cells row
so it shows in the Review tab). See CLAUDE.md gotcha #14/#16 and the
notes_issued_capital / notes_related_party prompts.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from notes.payload import NotesPayload
from notes.writer import write_notes_workbook
from notes_types import NotesTemplateType, notes_template_path


def _row_for(ws, label: str) -> int:
    target = label.lower().lstrip("*").strip()
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v and str(v).lower().lstrip("*").strip() == target:
            return row
    raise AssertionError(f"Label '{label}' not found in sheet {ws.title}")


_TABLE_HTML = (
    "<table>"
    "<tr><th>Movement</th><th>2024</th><th>2023</th></tr>"
    "<tr><td>Balance at beginning</td><td>1,000</td><td>900</td></tr>"
    "<tr><td>Shares issued</td><td>200</td><td>100</td></tr>"
    "</table>"
)


def test_issued_capital_reproduces_table_into_textblock_row(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.ISSUED_CAPITAL, level="company")
    sheet = "Notes-Issuedcapital"
    out = tmp_path / "issued_filled.xlsx"

    payloads = [
        # The numeric grid — unchanged behaviour.
        NotesPayload(
            chosen_row_label="Shares issued during financial year",
            content="",
            evidence="Page 40, Note 14",
            source_pages=[40],
            numeric_values={"company_cy": 200, "company_py": 100},
            parent_note={"number": "14", "title": "Share capital"},
        ),
        # The reproduced disclosure table — prose payload to the text-block row.
        NotesPayload(
            chosen_row_label="Disclosure of classes of share capital",
            content=_TABLE_HTML,
            evidence="Page 40, Note 14",
            source_pages=[40],
            parent_note={"number": "14", "title": "Share capital"},
        ),
    ]

    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=sheet,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[sheet]
    # Numeric grid still filled.
    num_row = _row_for(ws, "Shares issued during financial year")
    assert ws.cell(row=num_row, column=2).value == 200
    assert ws.cell(row=num_row, column=3).value == 100
    # Text-block row carries the flattened table in col B.
    tb_row = _row_for(ws, "Disclosure of classes of share capital")
    b = ws.cell(row=tb_row, column=2).value
    assert b and "Balance at beginning" in b and "Shares issued" in b
    wb.close()

    # The reproduced table reaches notes_cells (Review tab) as rich HTML —
    # only the prose row, never the numeric rows.
    table_cells = [c for c in result.cells_written if "<table" in c["html"]]
    assert len(table_cells) == 1
    assert table_cells[0]["row"] == tb_row
    # Numeric rows are not persisted as editor cells.
    assert all(c["row"] == tb_row for c in result.cells_written)


def test_related_party_reproduces_table_into_textblock_row(tmp_path: Path):
    tpl = notes_template_path(NotesTemplateType.RELATED_PARTY, level="company")
    sheet = "Notes-RelatedPartytran"
    out = tmp_path / "related_filled.xlsx"

    rp_table = (
        "<table>"
        "<tr><th>Transaction</th><th>2024</th></tr>"
        "<tr><td>Dividend income</td><td>5,000</td></tr>"
        "</table>"
    )
    payloads = [
        NotesPayload(
            chosen_row_label="Dividend income",
            content="",
            evidence="Page 52, Note 28",
            source_pages=[52],
            numeric_values={"company_cy": 5000},
            parent_note={"number": "28", "title": "Related party transactions"},
        ),
        NotesPayload(
            # This label exists verbatim on both MFRS and MPERS.
            chosen_row_label="Disclosure of transactions between related parties",
            content=rp_table,
            evidence="Page 52, Note 28",
            source_pages=[52],
            parent_note={"number": "28", "title": "Related party transactions"},
        ),
    ]

    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name=sheet,
    )
    assert result.success, result.errors

    wb = openpyxl.load_workbook(out)
    ws = wb[sheet]
    assert ws.cell(row=_row_for(ws, "Dividend income"), column=2).value == 5000
    tb_row = _row_for(ws, "Disclosure of transactions between related parties")
    assert "Dividend income" in (ws.cell(row=tb_row, column=2).value or "")
    wb.close()

    table_cells = [c for c in result.cells_written if "<table" in c["html"]]
    assert len(table_cells) == 1
    assert table_cells[0]["row"] == tb_row


@pytest.mark.parametrize("standard", ["mfrs", "mpers"])
def test_textblock_rows_are_writable_on_both_standards(standard: str, tmp_path: Path):
    """The reproduced-table home row must accept a prose write (not a
    formula / abstract-guarded cell) on both filing standards."""
    tpl = notes_template_path(
        NotesTemplateType.RELATED_PARTY, level="company", standard=standard,
    )
    out = tmp_path / f"rp_{standard}.xlsx"
    payloads = [
        NotesPayload(
            chosen_row_label="Disclosure of transactions between related parties",
            content="<table><tr><th>X</th></tr><tr><td>1</td></tr></table>",
            evidence="Page 1",
            source_pages=[1],
            parent_note={"number": "1", "title": "RP"},
        ),
    ]
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=payloads,
        output_path=str(out),
        filing_level="company",
        sheet_name="Notes-RelatedPartytran",
    )
    assert result.success, result.errors
    assert result.rows_written == 1
    # No formula-cell refusal errors leaked through.
    assert not result.errors
