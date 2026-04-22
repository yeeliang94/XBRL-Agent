"""PLAN §4 Phase E.3 — full-pipeline E2E across all 5 notes sheets.

Mocks the per-sheet agent runner and the Sheet-12 sub-agent so no real
LLM is invoked. Exercises:

  * The coordinator's fan-out across all five ``NotesTemplateType`` values.
  * The writer for sheets 10/11/13/14 (prose + numeric payloads).
  * The Sheet-12 sub-coordinator + writer path with row-112 concatenation.
  * The workbook merger at the end, producing a single ``filled.xlsx``
    with one sheet per notes template, preserving face sheets when
    requested alongside.

Every payload goes through the real ``notes.writer.write_notes_workbook``
so template-shape regressions surface immediately.
"""
from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from notes.coordinator import (
    NotesAgentResult,
    NotesRunConfig,
    run_notes_extraction,
)
from notes.listofnotes_subcoordinator import SubAgentRunResult
from notes.payload import NotesPayload
from notes.writer import write_notes_workbook
from notes_types import (
    NOTES_REGISTRY,
    NotesTemplateType,
    notes_template_path,
)
from scout.infopack import Infopack
from scout.notes_discoverer import NoteInventoryEntry


# ---------------------------------------------------------------------------
# Per-template fixture payloads — realistic enough to exercise the writer's
# row resolution, evidence column placement, and Sheet-12 concatenation.
# ---------------------------------------------------------------------------

_FIXTURE_PAYLOADS: dict[NotesTemplateType, list[NotesPayload]] = {
    NotesTemplateType.CORP_INFO: [
        NotesPayload(
            chosen_row_label="Financial reporting status",
            content="The Group is a going concern and has adequate liquidity.",
            evidence="Page 14, Note 1",
            source_pages=[14],
        ),
    ],
    NotesTemplateType.ACC_POLICIES: [
        NotesPayload(
            chosen_row_label="Description of accounting policy for property, plant and equipment",
            content="PPE is stated at cost less accumulated depreciation.",
            evidence="Page 32, Note 2.7",
            source_pages=[32],
        ),
        NotesPayload(
            chosen_row_label="Description of accounting policy for leases",
            content="Right-of-use assets recognised at lease commencement.",
            evidence="Page 35, Note 2.12",
            source_pages=[35],
        ),
    ],
    NotesTemplateType.ISSUED_CAPITAL: [
        NotesPayload(
            chosen_row_label="Issued capital",
            content="",
            evidence="Page 55, Note 18",
            source_pages=[55],
            numeric_values={"company_cy": 10_000_000, "company_py": 9_500_000},
        ),
    ],
    NotesTemplateType.RELATED_PARTY: [
        NotesPayload(
            chosen_row_label="Related party transactions",
            content="",
            evidence="Page 71, Note 24",
            source_pages=[71],
            numeric_values={"company_cy": 450_000, "company_py": 320_000},
        ),
    ],
}

# Sheet-12 — exercised via a mocked sub-coordinator that hands back
# realistic aggregated payloads.
_LIST_OF_NOTES_PAYLOADS: list[NotesPayload] = [
    NotesPayload(
        chosen_row_label="Disclosure of revenue from contract customers",
        content="Revenue is recognised when performance obligations are met.",
        evidence="Page 28-30, Note 4",
        source_pages=[28, 29, 30],
        sub_agent_id="notes:LIST_OF_NOTES:sub0",
    ),
    # A catch-all row-112 entry so the writer's row-concatenation path runs.
    NotesPayload(
        chosen_row_label="Disclosure of other notes to accounts",
        content="Segment disclosure omitted per MFRS 8 exemption.",
        evidence="Page 62, Note 20",
        source_pages=[62],
        sub_agent_id="notes:LIST_OF_NOTES:sub3",
    ),
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_fixture_for(template_type: NotesTemplateType, output_dir: Path) -> str:
    """Call the real writer with the fixture payloads for ``template_type``
    and return the absolute output path."""
    entry = NOTES_REGISTRY[template_type]
    tpl = notes_template_path(template_type, level="company")
    out = output_dir / f"NOTES_{template_type.value}_filled.xlsx"
    result = write_notes_workbook(
        template_path=str(tpl),
        payloads=_FIXTURE_PAYLOADS[template_type],
        output_path=str(out),
        filing_level="company",
        sheet_name=entry.sheet_name,
    )
    assert result.success, (
        f"writer failed for {template_type.value}: {result.errors}"
    )
    assert result.rows_written > 0
    return str(out)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_all_five_notes_sheets_run_end_to_end(tmp_path: Path):
    """Run the coordinator for all 5 notes templates with mocked agents.
    Verify every NotesAgentResult is a success and every workbook file
    lands on disk with the fixture content visible in its cells."""

    pdf_path = tmp_path / "uploaded.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run=set(NotesTemplateType),
        filing_level="company",
    )
    infopack = Infopack(
        toc_page=1,
        page_offset=0,
        notes_inventory=[
            NoteInventoryEntry(note_num=1, title="Corporate information", page_range=(14, 14)),
            NoteInventoryEntry(note_num=2, title="Summary of accounting policies", page_range=(32, 35)),
            NoteInventoryEntry(note_num=4, title="Revenue", page_range=(28, 30)),
            NoteInventoryEntry(note_num=18, title="Issued capital", page_range=(55, 55)),
            NoteInventoryEntry(note_num=24, title="Related party transactions", page_range=(71, 71)),
        ],
    )

    async def fake_single_agent(**kwargs):
        tt = kwargs["template_type"]
        workbook_path = _write_fixture_for(tt, Path(kwargs["output_dir"]))
        return NotesAgentResult(
            template_type=tt,
            status="succeeded",
            workbook_path=workbook_path,
        )

    # Sheet 12: bypass the sub-agent fan-out by mocking the coordinator's
    # sub-coordinator call. Aggregated payloads flow directly to the real
    # writer inside _run_list_of_notes_fanout.
    from notes.listofnotes_subcoordinator import ListOfNotesSubResult

    async def fake_sub_coordinator(**kwargs):
        return ListOfNotesSubResult(
            sub_agent_results=[
                SubAgentRunResult(
                    sub_agent_id=f"notes:LIST_OF_NOTES:sub{i}",
                    batch=[],
                    payloads=[],
                    status="succeeded",
                    retry_count=0,
                )
                for i in range(5)
            ],
            aggregated_payloads=_LIST_OF_NOTES_PAYLOADS,
            unmatched_payloads=[
                p for p in _LIST_OF_NOTES_PAYLOADS
                if "other notes" in p.chosen_row_label.lower()
            ],
        )

    with patch("notes.coordinator._run_single_notes_agent", side_effect=fake_single_agent), \
         patch(
             # Patch at the coordinator's lookup site — B.2 moved the import
             # from function scope to module top, so the coordinator captures
             # its own reference to this function at import time.
             "notes.coordinator.run_listofnotes_subcoordinator",
             side_effect=fake_sub_coordinator,
         ):
        result = await run_notes_extraction(config, infopack=infopack)

    # Every sheet succeeded.
    assert result.all_succeeded, [r.error for r in result.agent_results]
    by_type = {r.template_type: r for r in result.agent_results}
    assert set(by_type) == set(NotesTemplateType)
    for tt, r in by_type.items():
        assert r.status == "succeeded", f"{tt.value} status={r.status} error={r.error}"
        assert r.workbook_path, f"{tt.value} missing workbook_path"
        assert Path(r.workbook_path).exists()

    # Spot-check each sheet — prose, numeric, and Sheet-12 concatenation.
    wb_ci = openpyxl.load_workbook(by_type[NotesTemplateType.CORP_INFO].workbook_path)
    ws_ci = wb_ci["Notes-CI"]
    assert _cell_contains(ws_ci, "Financial reporting status", col=2, needle="going concern")
    wb_ci.close()

    wb_ic = openpyxl.load_workbook(by_type[NotesTemplateType.ISSUED_CAPITAL].workbook_path)
    ws_ic = wb_ic["Notes-Issuedcapital"]
    # Numeric values land in col B (CY) and C (PY). Exact-match on the
    # data-row label so we don't collide with the sheet-header row
    # "Notes - Issued capital" above it.
    ic_row = _first_row_matching(ws_ic, "Issued capital", exact=True)
    assert ws_ic.cell(row=ic_row, column=2).value == 10_000_000
    assert ws_ic.cell(row=ic_row, column=3).value == 9_500_000
    wb_ic.close()

    # Sheet 12 — the sub-coordinator aggregated two payloads. The writer
    # placed the revenue payload on its own row and the row-112 catch-all.
    wb_12 = openpyxl.load_workbook(by_type[NotesTemplateType.LIST_OF_NOTES].workbook_path)
    ws_12 = wb_12["Notes-Listofnotes"]
    row_112 = _first_row_matching(ws_12, "disclosure of other notes to accounts")
    assert "segment disclosure omitted" in str(
        ws_12.cell(row=row_112, column=2).value or ""
    ).lower()
    wb_12.close()


@pytest.mark.asyncio
async def test_partial_run_still_writes_successful_sheets(tmp_path: Path):
    """If the ISSUED_CAPITAL sheet blows up but the other four succeed,
    the run as a whole is not-all-succeeded but every successful sheet
    still has a workbook on disk. This is the PLAN §4 Checkpoint E.1
    ‘partial coverage is success per sheet’ contract at the coordinator
    level."""
    pdf_path = tmp_path / "uploaded.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")

    # Exclude LIST_OF_NOTES from this scenario so we don't need to also
    # mock the sub-coordinator — this test is about single-agent sheets.
    templates = {
        NotesTemplateType.CORP_INFO,
        NotesTemplateType.ACC_POLICIES,
        NotesTemplateType.ISSUED_CAPITAL,
        NotesTemplateType.RELATED_PARTY,
    }
    config = NotesRunConfig(
        pdf_path=str(pdf_path),
        output_dir=str(tmp_path),
        model="test",
        notes_to_run=templates,
        filing_level="company",
    )

    async def flaky_runner(**kwargs):
        tt = kwargs["template_type"]
        if tt == NotesTemplateType.ISSUED_CAPITAL:
            return NotesAgentResult(
                template_type=tt,
                status="failed",
                error="retries exhausted — model kept timing out",
            )
        return NotesAgentResult(
            template_type=tt,
            status="succeeded",
            workbook_path=_write_fixture_for(tt, Path(kwargs["output_dir"])),
        )

    with patch("notes.coordinator._run_single_notes_agent", side_effect=flaky_runner):
        result = await run_notes_extraction(config, infopack=None)

    by_type = {r.template_type: r for r in result.agent_results}
    assert by_type[NotesTemplateType.ISSUED_CAPITAL].status == "failed"
    for tt in templates - {NotesTemplateType.ISSUED_CAPITAL}:
        assert by_type[tt].status == "succeeded"
        assert Path(by_type[tt].workbook_path).exists()
    assert not result.all_succeeded


# ---------------------------------------------------------------------------
# Small helpers to keep assertions readable
# ---------------------------------------------------------------------------


def _first_row_matching(ws, needle: str, *, exact: bool = False) -> int:
    """Find the first row whose column-A label contains (or equals) ``needle``.

    ``exact=True`` matches the normalized label (strip, lstrip '*', lower),
    which is necessary when multiple labels share a common substring — e.g.
    the Notes-Issuedcapital template has both 'Notes - Issued capital'
    (section header) and 'Issued capital' (data row).
    """
    target = needle.strip().lstrip("*").strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not v:
            continue
        normalized = str(v).strip().lstrip("*").strip().lower()
        if exact and normalized == target:
            return r
        if not exact and target in normalized:
            return r
    raise AssertionError(f"No row found matching '{needle}' in sheet {ws.title}")


def _cell_contains(ws, label_needle: str, *, col: int, needle: str) -> bool:
    row = _first_row_matching(ws, label_needle)
    value = ws.cell(row=row, column=col).value
    return value is not None and needle.lower() in str(value).lower()
