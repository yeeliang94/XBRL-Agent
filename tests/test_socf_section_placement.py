"""SOCF lines placed in the wrong section — the error no arithmetic can see.

Run-84 finding (2026-08-05). A dividend of 65,345 was entered under operating
activities. Every cash-flow check still passed, because moving a line between
sections leaves the net change in cash unchanged: ``socf_articulation``
(opening + net change == closing) and ``socf_to_sofp_cash`` both compare totals,
and the totals were right. The operating subtotal was wrong by 65,345 and
nothing reported it.

The check warns and never fails, by decision. MFRS permits dividends paid in
either operating or financing — the live MFRS Company SOCF-Indirect template
carries a "Dividends paid" row in BOTH sections for that reason — so there is no
fixed correct answer, only "does it match the source". That makes the check only
as reliable as the scout's reading of the face page's section headings, which
has never been measured.

These pin both halves: it must catch a real misplacement, and it must stay
quiet on everything it cannot actually see.
"""
from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

from cross_checks.socf_section_placement import (
    check_socf_section_placement,
    classify_section,
)

REPO = Path(__file__).resolve().parent.parent
SOCF_INDIRECT = REPO / "XBRL-template-MFRS" / "Company" / "07-SOCF-Indirect.xlsx"


def _ref(label, section):
    return SimpleNamespace(label=label, section=section)


@pytest.fixture
def socf_workbook(tmp_path: Path):
    """A copy of the live template with values written to named rows."""
    def _build(*, rows: dict[str, float]):
        wb = openpyxl.load_workbook(str(SOCF_INDIRECT))
        ws = wb["SOCF-Indirect"]
        wanted = {k.strip().lstrip("*").strip().lower(): v
                  for k, v in rows.items()}
        placed: dict[str, int] = {}
        for r in range(1, ws.max_row + 1):
            raw = ws.cell(row=r, column=1).value
            if raw is None:
                continue
            key = str(raw).strip().lstrip("*").strip().lower()
            if key in wanted and key not in placed:
                # Only a row with no formula of its own is an input row.
                if not (isinstance(ws.cell(row=r, column=2).value, str)
                        and str(ws.cell(row=r, column=2).value).startswith("=")):
                    ws.cell(row=r, column=2).value = wanted[key]
                    placed[key] = r
        assert len(placed) == len(wanted), (
            f"fixture could not place {set(wanted) - set(placed)}")
        out = tmp_path / "filled.xlsx"
        wb.save(str(out))
        wb.close()
        return str(out), placed
    return _build


# --- section classification ------------------------------------------------

@pytest.mark.parametrize("heading,expected", [
    ("Cash flows from (used in) operating activities", "operating"),
    ("Cash flows from (used in) investing activities", "investing"),
    ("Cash flows from (used in) financing activities", "financing"),
    ("CASH FLOWS FROM FINANCING ACTIVITIES", "financing"),
    (None, None),
    ("", None),
    ("Changes in working capital", None),
    # Ambiguous: names two sections, so it identifies neither.
    ("Reconciliation of operating to investing activities", None),
])
def test_classify_section(heading, expected):
    assert classify_section(heading) == expected


# --- the warning it exists for ---------------------------------------------

def test_dividend_in_operating_is_flagged_when_source_says_financing(
    socf_workbook,
):
    """The run-84 case, on the live template."""
    path, placed = socf_workbook(rows={"Dividends paid": 65345.0})

    warns = check_socf_section_placement(
        path, [_ref("Dividends paid",
                    "Cash flows from (used in) financing activities")])

    assert len(warns) == 1
    w = warns[0]
    assert w.status == "warning"
    assert w.template_section == "operating"
    assert w.source_section == "financing"
    assert w.row == placed["dividends paid"]
    # The message has to admit it might be the wrong one.
    assert "is the one that is wrong" in w.message


def test_agreement_produces_no_warning(socf_workbook):
    path, _ = socf_workbook(rows={"Dividends received": 1000.0})

    warns = check_socf_section_placement(
        path, [_ref("Dividends received",
                    "Cash flows from (used in) operating activities")])

    assert warns == []


def test_label_decoration_does_not_manufacture_a_mismatch(socf_workbook):
    """The two sides come from a template cell and a PDF reading; they agree on
    wording far more often than on punctuation and case."""
    path, _ = socf_workbook(rows={"Dividends received": 1000.0})

    warns = check_socf_section_placement(
        path, [_ref("  dividends received:  ", "operating activities")])

    assert warns == []


# --- silence where it cannot see -------------------------------------------

def test_no_scout_refs_yields_no_warnings(socf_workbook):
    """Nothing observed means not assessed. Silence here must not be read as
    agreement — which is why the check reports nothing rather than a pass."""
    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    assert check_socf_section_placement(path, []) == []


def test_ref_without_a_section_is_skipped(socf_workbook):
    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    assert check_socf_section_placement(
        path, [_ref("Dividends paid", None)]) == []


def test_unclassifiable_source_heading_is_skipped(socf_workbook):
    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    assert check_socf_section_placement(
        path, [_ref("Dividends paid", "Note 14")]) == []


def test_label_the_agent_never_wrote_is_skipped(socf_workbook):
    """A line the scout saw but the agent left empty has no placement to
    disagree about."""
    path, _ = socf_workbook(rows={"Dividends received": 10.0})
    assert check_socf_section_placement(
        path, [_ref("Interest paid", "financing activities")]) == []


def test_dict_refs_are_accepted(socf_workbook):
    """A persisted infopack round-trips face_line_refs as plain dicts."""
    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})

    warns = check_socf_section_placement(
        path, [{"label": "Dividends paid", "section": "financing activities"}])

    assert len(warns) == 1


# --- it must never be able to break a run ----------------------------------

def test_unreadable_workbook_returns_no_warnings(tmp_path: Path):
    bad = tmp_path / "not-a-workbook.xlsx"
    bad.write_text("nope", encoding="utf-8")
    assert check_socf_section_placement(
        str(bad), [_ref("Dividends paid", "financing")]) == []


def test_workbook_without_a_socf_sheet_returns_no_warnings(tmp_path: Path):
    wb = openpyxl.Workbook()
    wb.active.title = "SOFP-CuNonCu"
    out = tmp_path / "no-socf.xlsx"
    wb.save(str(out))
    wb.close()
    assert check_socf_section_placement(
        str(out), [_ref("Dividends paid", "financing")]) == []


def test_malformed_ref_does_not_raise(socf_workbook):
    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    assert check_socf_section_placement(path, [object()]) == []


# --- server wiring ---------------------------------------------------------

def test_server_adapter_folds_warnings_into_cross_check_results(socf_workbook):
    """Rides the same warning channel as the notes advisory checks, so it
    reaches persistence, the SSE stream and the Cross-checks tab without any
    new plumbing."""
    import server

    path, placed = socf_workbook(rows={"Dividends paid": 65345.0})
    infopack = SimpleNamespace(statements={
        "SOCF": SimpleNamespace(face_line_refs=[
            _ref("Dividends paid", "financing activities")]),
    })

    results = server._run_socf_section_placement(path, infopack, run_id=1)

    assert len(results) == 1
    assert results[0].status == "warning", "this check must never fail a run"
    assert results[0].name.startswith("SOCF section:")
    assert results[0].target_sheet == "SOCF"
    assert results[0].target_row == placed["dividends paid"]


def test_server_adapter_reads_an_enum_keyed_infopack(socf_workbook):
    """A live infopack keys `statements` by StatementType, not by string."""
    import server
    from statement_types import StatementType

    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    infopack = SimpleNamespace(statements={
        StatementType.SOCF: SimpleNamespace(face_line_refs=[
            _ref("Dividends paid", "financing activities")]),
    })

    assert len(server._run_socf_section_placement(path, infopack, run_id=1)) == 1


def test_server_adapter_survives_a_missing_infopack(socf_workbook):
    import server
    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    assert server._run_socf_section_placement(path, None, run_id=1) == []


# ------------------------------------------- the advisory must survive review
#
# `cross_check_results` is REPLACED wholesale by the post-reviewer re-run, and
# the final persistence writes whatever that list holds. An advisory computed
# in the initial block only is therefore deleted the moment the reviewer makes
# a fix — the run files with the warning gone (peer review, 2026-08-05).


def test_one_aggregator_owns_the_advisory(socf_workbook):
    """`_run_notes_advisories` includes the SOCF placement check, so both the
    initial pass and the post-reviewer re-run get it from the same place."""
    import asyncio

    import server
    from statement_types import StatementType

    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    infopack = SimpleNamespace(statements={
        StatementType.SOCF: SimpleNamespace(face_line_refs=[
            _ref("Dividends paid", "financing activities")]),
    })

    results = asyncio.run(server._run_notes_advisories(
        path, 1, filing_level="company", filing_standard="mfrs",
        infopack=infopack,
    ))
    assert any("section" in (r.name or "").lower()
               or "section" in (r.message or "").lower()
               for r in results), (
        "the SOCF placement advisory must come from the shared aggregator"
    )


def test_the_advisory_has_no_second_call_site():
    """Structural guard. The check must be reachable ONLY through the shared
    aggregator — a direct call in one of the two pipeline blocks is how it went
    missing in the first place."""
    import inspect

    import server

    pipeline = inspect.getsource(server.run_multi_agent_stream)
    assert "_run_socf_section_placement" not in pipeline, (
        "call it via _run_notes_advisories, not directly in the pipeline — "
        "a direct call lands in only one of the two cross-check passes"
    )
    aggregator = inspect.getsource(server._run_notes_advisories)
    assert "_run_socf_section_placement" in aggregator


def test_both_cross_check_passes_recompute_the_advisories():
    """Both the initial pass and the post-reviewer re-run must call the
    aggregator, or the surviving list is missing every advisory."""
    import inspect

    import server

    src = inspect.getsource(server.run_multi_agent_stream)
    assert src.count("_run_notes_advisories(") == 2, (
        "expected exactly two call sites: the initial pass and the "
        "post-reviewer re-run that replaces the results list"
    )
    # …and the second one is after the reviewer stage.
    first = src.find("_run_notes_advisories(")
    second = src.find("_run_notes_advisories(", first + 1)
    reviewing = src.find('_emit_stage("reviewing")')
    assert first < reviewing < second, (
        "the second recompute must follow the reviewer stage — that is the "
        "pass whose replacement of cross_check_results drops the advisories"
    )


def test_the_aggregator_surfaces_notes_consistency_too(socf_workbook):
    """Regression guard for this refactor's own bug: moving the block to a
    module-level function left `check_notes_consistency` unbound, and the
    aggregator's broad except turned the NameError into an empty list — which
    reads as "nothing to warn about". Every advisory family must reach the
    output, not just the one being worked on."""
    import asyncio
    from types import SimpleNamespace as NS
    from unittest.mock import patch

    import server

    path, _ = socf_workbook(rows={"Dividends paid": 65345.0})
    warning = NS(
        sheet_11_label="Trade receivables", sheet_12_label="Trade receivables",
        message="Sheet 11 cites [21]; Sheet 12 cites [19]. No overlap.",
    )
    with patch("cross_checks.notes_consistency.check_notes_consistency",
               return_value=[warning]):
        results = asyncio.run(server._run_notes_advisories(
            path, 1, filing_level="company", filing_standard="mfrs",
            infopack=None,
        ))
    assert any(r.name.startswith("Notes consistency:") for r in results), (
        "the notes-consistency family vanished from the shared aggregator"
    )
