"""Layout tests for the MPERS Group SOCIE template.

These tests are red until Phase 3 of `docs/PLAN-mpers-group-socie-formulas.md`
lands. They pin the four-block layout (rows 3-25, 27-49, 51-73, 75-97)
emitted by `scripts/generate_mpers_templates.py::_apply_group_socie_layout`
to:

  * carry value cells in column B (no longer label-only),
  * land a `*Total increase (decrease) in equity` subtotal in each block
    that subtracts `Dividends paid` per ADR-002,
  * leave the `Equity at end of period` row formula-free in every block
    (per gotcha #12 the agent enters this directly).

Once Phase 3 is green the prompt fallback in `prompts/socie_mpers.md` and
the `test_mpers_group_socie_subtracts_dividends_paid` seat in
`test_notes_prompt_phase1.py` can be removed (Phase 4).
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest


_REPO_ROOT = Path(__file__).resolve().parent.parent
_TEMPLATE = _REPO_ROOT / "XBRL-template-MPERS" / "Group" / "09-SOCIE.xlsx"

# Block ranges match cross_checks/util.py::SOCIE_GROUP_BLOCKS — keep in
# sync. Each tuple is (header_row, last_body_row); the body of each
# block runs from header_row+1 through last_body_row.
BLOCK_RANGES = [(3, 25), (27, 49), (51, 73), (75, 97)]


def _norm(value) -> str:
    return str(value or "").strip().lstrip("*").strip().lower()


@pytest.fixture(scope="module")
def socie_ws():
    """Open the live MPERS Group SOCIE workbook once for all assertions.

    Module-scoped so we only pay openpyxl's parse cost once. data_only=False
    so formulas come back as strings, not evaluated values.
    """
    wb = openpyxl.load_workbook(_TEMPLATE, data_only=False)
    try:
        yield wb["SOCIE"]
    finally:
        wb.close()


def test_mpers_group_socie_has_value_cells(socie_ws):
    """Template must populate at least column B; the original
    label-only layout had max_col=1."""
    assert socie_ws.max_column >= 2, (
        f"Expected value column(s) in MPERS Group SOCIE; "
        f"max_col={socie_ws.max_column} (label-only layout)"
    )


@pytest.mark.parametrize("start,end", BLOCK_RANGES)
def test_mpers_group_socie_has_per_block_subtotal(socie_ws, start, end):
    """Each block must carry a `*Total increase (decrease) in equity`
    row with a column-B formula that subtracts the block's
    `Dividends paid` row.

    Parametrising per block makes failure messages tell you which
    block is missing the subtotal rather than reporting one big diff.
    """
    total_row = next(
        (
            r for r in range(start, end + 1)
            if _norm(socie_ws.cell(r, 1).value).endswith(
                "total increase (decrease) in equity"
            )
        ),
        None,
    )
    assert total_row is not None, (
        f"Block ({start},{end}): no `*Total increase (decrease) in equity` "
        f"label row found"
    )

    formula = socie_ws.cell(total_row, 2).value
    assert isinstance(formula, str) and formula.startswith("="), (
        f"Block ({start},{end}) row {total_row}: expected a formula in col B, "
        f"got {formula!r}"
    )

    div_row = next(
        (
            r for r in range(start, end + 1)
            if _norm(socie_ws.cell(r, 1).value) == "dividends paid"
        ),
        None,
    )
    assert div_row is not None, (
        f"Block ({start},{end}): no `Dividends paid` row found"
    )
    assert f"-1*B{div_row}" in formula or f"-B{div_row}" in formula, (
        f"Block ({start},{end}) subtotal at row {total_row} does not "
        f"subtract B{div_row}: {formula}"
    )


@pytest.mark.parametrize("start,end", BLOCK_RANGES)
def test_mpers_group_socie_closing_balance_is_blank(socie_ws, start, end):
    """`Equity at end of period` is the closing balance per gotcha #12 —
    the agent enters it directly because the calc linkbase doesn't roll
    it up from movement rows. Pin that the template never injects a
    formula here.
    """
    closing_row = next(
        (
            r for r in range(start, end + 1)
            if _norm(socie_ws.cell(r, 1).value) == "equity at end of period"
        ),
        None,
    )
    assert closing_row is not None, (
        f"Block ({start},{end}): no `Equity at end of period` row found"
    )
    value = socie_ws.cell(closing_row, 2).value
    assert not (isinstance(value, str) and value.startswith("=")), (
        f"Block ({start},{end}) row {closing_row}: closing balance must be "
        f"formula-free, got {value!r}"
    )
