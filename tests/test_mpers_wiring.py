"""MPERS pipeline wiring tests — one marker per plan phase.

Exercises the `filing_standard` axis across the whole pipeline, from the
template registry up through the CLI and server request body. Every test
is paired with the production change that makes it pass in
`docs/PLAN-mpers-pipeline-wiring.md`.
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path
from typing import Optional
from unittest.mock import patch

import pytest

from notes_types import NotesTemplateType, notes_template_path
from statement_types import (
    StatementType,
    VARIANTS,
    template_path,
)


# ---------------------------------------------------------------------------
# Phase 1 — Registry (mpers_wiring_registry)
# ---------------------------------------------------------------------------

@pytest.mark.mpers_wiring_registry
def test_template_path_resolves_mpers_company() -> None:
    """MPERS Company SOFP-CuNonCu template is resolvable and on disk."""
    p = template_path(
        StatementType.SOFP, "CuNonCu", level="company", standard="mpers",
    )
    assert p.as_posix().endswith(
        "XBRL-template-MPERS/Company/01-SOFP-CuNonCu.xlsx"
    ), f"unexpected path: {p}"
    assert p.exists(), f"missing template: {p}"


@pytest.mark.mpers_wiring_registry
def test_template_path_default_still_mfrs() -> None:
    """Backward compat: no `standard` kwarg keeps MFRS routing."""
    p = template_path(StatementType.SOFP, "CuNonCu", "company")
    assert "XBRL-template-MFRS" in p.as_posix()
    assert p.exists()


@pytest.mark.mpers_wiring_registry
def test_variant_has_applies_to_standard_default() -> None:
    """Every pre-existing (non-SoRE) variant carries the default set.

    SoRE is the only registered MPERS-only variant — every other entry must
    keep the wide-open default so existing runs are unaffected.
    """
    from statement_types import Variant

    # Dataclass default is the wide-open set.
    sample = Variant(statement=StatementType.SOFP, name="probe", template_filename="x")
    assert sample.applies_to_standard == frozenset({"mfrs", "mpers"})

    for (_, name), v in VARIANTS.items():
        if name == "SoRE":
            continue  # intentionally MPERS-only (step 1.2)
        assert v.applies_to_standard == frozenset({"mfrs", "mpers"}), (
            f"{v.statement.value}/{v.name} unexpectedly narrowed"
        )


@pytest.mark.mpers_wiring_registry
def test_sore_registered_as_mpers_only_socie_variant() -> None:
    """SoRE exists on the SOCIE row, MPERS-only, with the real template."""
    v = VARIANTS[(StatementType.SOCIE, "SoRE")]
    assert v.template_filename == "10-SoRE.xlsx"
    assert v.applies_to_standard == frozenset({"mpers"})


@pytest.mark.mpers_wiring_registry
def test_template_path_rejects_sore_on_mfrs() -> None:
    """MFRS has no SoRE — asking for it must raise with a clear message."""
    with pytest.raises(ValueError, match="MPERS"):
        template_path(StatementType.SOCIE, "SoRE", "company", "mfrs")


@pytest.mark.mpers_wiring_registry
def test_variants_for_standard_filters_by_applicability() -> None:
    """variants_for_standard(SOCIE, ...) returns Default only on MFRS,
    Default + SoRE on MPERS (Default first, pinning coordinator fallback)."""
    from statement_types import variants_for_standard
    mfrs = variants_for_standard(StatementType.SOCIE, "mfrs")
    assert [v.name for v in mfrs] == ["Default"]

    mpers = variants_for_standard(StatementType.SOCIE, "mpers")
    names = [v.name for v in mpers]
    assert names[0] == "Default", "Default must be first — coordinator falls back to [0]"
    assert "SoRE" in names


@pytest.mark.mpers_wiring_registry
def test_notes_template_path_resolves_mpers_shifted_numbering() -> None:
    """MPERS notes are shifted one slot up (11..15) from MFRS's 10..14."""
    expected_prefixes = {
        NotesTemplateType.CORP_INFO: "11-",
        NotesTemplateType.ACC_POLICIES: "12-",
        NotesTemplateType.LIST_OF_NOTES: "13-",
        NotesTemplateType.ISSUED_CAPITAL: "14-",
        NotesTemplateType.RELATED_PARTY: "15-",
    }
    for nt, prefix in expected_prefixes.items():
        p = notes_template_path(nt, "company", standard="mpers")
        assert p.name.startswith(prefix), (
            f"{nt.value}: expected prefix {prefix}, got {p.name}"
        )
        assert p.exists(), f"missing MPERS notes template: {p}"


@pytest.mark.mpers_wiring_registry
def test_notes_template_path_default_still_mfrs() -> None:
    """No `standard` kwarg keeps the MFRS 10- numbering."""
    p = notes_template_path(NotesTemplateType.CORP_INFO, "company")
    assert p.name == "10-Notes-CorporateInfo.xlsx"
    assert "XBRL-template-MFRS" in p.as_posix()
    assert p.exists()


# ---------------------------------------------------------------------------
# Phase 2 — Coordinator plumbing (mpers_wiring_coordinator)
# ---------------------------------------------------------------------------

@pytest.mark.mpers_wiring_coordinator
def test_coordinator_passes_standard_into_template_path(monkeypatch) -> None:
    """run_extraction threads RunConfig.filing_standard into template_path()."""
    import coordinator as coord
    from coordinator import RunConfig, run_extraction

    captured: dict = {}

    def spy_template_path(stmt, variant, level="company", standard="mfrs"):
        captured["standard"] = standard
        captured["level"] = level
        captured["statement"] = stmt
        captured["variant"] = variant
        # Return a placeholder so _run_single_agent can short-circuit via mock
        return Path("/tmp/spy-template.xlsx")

    async def fake_single_agent(**kwargs):
        # Minimal AgentResult so the coordinator completes cleanly.
        from coordinator import AgentResult
        return AgentResult(
            statement_type=kwargs["statement_type"],
            variant=kwargs["variant"],
            status="succeeded",
        )

    monkeypatch.setattr(coord, "get_template_path", spy_template_path)
    monkeypatch.setattr(coord, "_run_single_agent", fake_single_agent)

    cfg = RunConfig(
        pdf_path="/tmp/fake.pdf",
        output_dir="/tmp",
        statements_to_run={StatementType.SOFP},
        variants={StatementType.SOFP: "CuNonCu"},
        filing_standard="mpers",
    )
    asyncio.run(run_extraction(cfg, push_sentinel=False))
    assert captured["standard"] == "mpers"
    assert captured["level"] == "company"


@pytest.mark.mpers_wiring_coordinator
def test_coordinator_variant_fallback_uses_standard_filter(monkeypatch) -> None:
    """With no variant hint and no infopack, coordinator uses
    variants_for_standard(…) which lists Default first on MPERS."""
    import coordinator as coord
    from coordinator import RunConfig, run_extraction

    captured_variant: dict = {}

    async def fake_single_agent(**kwargs):
        captured_variant["variant"] = kwargs["variant"]
        from coordinator import AgentResult
        return AgentResult(
            statement_type=kwargs["statement_type"],
            variant=kwargs["variant"],
            status="succeeded",
        )

    # Avoid touching the real filesystem; the test is about variant selection.
    monkeypatch.setattr(
        coord, "get_template_path",
        lambda stmt, variant, level="company", standard="mfrs": Path("/tmp/x.xlsx"),
    )
    monkeypatch.setattr(coord, "_run_single_agent", fake_single_agent)

    cfg = RunConfig(
        pdf_path="/tmp/fake.pdf",
        output_dir="/tmp",
        statements_to_run={StatementType.SOCIE},
        variants={},
        filing_standard="mpers",
    )
    asyncio.run(run_extraction(cfg, push_sentinel=False))
    assert captured_variant["variant"] == "Default", (
        "Default must remain the fallback even when SoRE is also registered"
    )


@pytest.mark.mpers_wiring_coordinator
def test_coordinator_drops_scout_suggestion_when_standard_mismatches(
    monkeypatch,
) -> None:
    """Peer-review HIGH: scout suggesting an MPERS-only variant on an MFRS
    run must not crash the coordinator at template_path() time. The
    suggestion is advisory; coordinator falls back to the registry default."""
    import coordinator as coord
    from coordinator import RunConfig, run_extraction
    from scout.infopack import Infopack, StatementPageRef

    captured: dict = {}

    def spy_template_path(stmt, variant, level="company", standard="mfrs"):
        captured.setdefault("calls", []).append({
            "variant": variant, "standard": standard,
        })
        return Path("/tmp/x.xlsx")

    async def fake_single_agent(**kwargs):
        from coordinator import AgentResult
        captured.setdefault("variants", []).append(kwargs["variant"])
        return AgentResult(
            statement_type=kwargs["statement_type"],
            variant=kwargs["variant"],
            status="succeeded",
        )

    monkeypatch.setattr(coord, "get_template_path", spy_template_path)
    monkeypatch.setattr(coord, "_run_single_agent", fake_single_agent)

    # Scout suggests SoRE (MPERS-only) for an MFRS run.
    ip = Infopack(
        toc_page=1, page_offset=0,
        statements={
            StatementType.SOCIE: StatementPageRef(
                variant_suggestion="SoRE", face_page=10,
            ),
        },
    )
    cfg = RunConfig(
        pdf_path="/tmp/fake.pdf", output_dir="/tmp",
        statements_to_run={StatementType.SOCIE},
        variants={},
        filing_standard="mfrs",
    )

    # Before the fix: this raised ValueError synchronously inside the loop.
    # After the fix: coordinator drops the suggestion and resolves to Default.
    asyncio.run(run_extraction(cfg, infopack=ip, push_sentinel=False))
    assert captured["variants"] == ["Default"], (
        f"expected MFRS fallback to Default, got {captured['variants']}"
    )


@pytest.mark.mpers_wiring_coordinator
def test_create_extraction_agent_accepts_filing_standard() -> None:
    """create_extraction_agent passes filing_standard through to deps."""
    from pydantic_ai.models.test import TestModel
    from extraction.agent import create_extraction_agent

    agent, deps = create_extraction_agent(
        statement_type=StatementType.SOFP,
        variant="CuNonCu",
        pdf_path="/tmp/does-not-matter.pdf",
        template_path=str(template_path(StatementType.SOFP, "CuNonCu", "company")),
        model=TestModel(),
        output_dir="/tmp",
        filing_standard="mpers",
    )
    assert deps.filing_standard == "mpers"


@pytest.mark.mpers_wiring_coordinator
def test_notes_coordinator_passes_standard_into_agent_factory(
    monkeypatch, tmp_path,
) -> None:
    """run_notes_extraction threads NotesRunConfig.filing_standard into
    create_notes_agent(). Peer-review MEDIUM: the earlier version of this
    test stubbed out the single-agent runner, which skipped the factory
    entirely — the spy never triggered. This version patches the factory
    itself so we assert on the actual `filing_standard=` kwarg the
    coordinator propagates."""
    from pydantic_ai.models.test import TestModel
    import notes.coordinator as ncoord
    from notes.coordinator import NotesRunConfig, run_notes_extraction

    captured: list = []

    # Real factory so the template path lookup exercises notes_template_path
    # with the real standard kwarg, but wrap it to capture the kwargs.
    from notes.agent import create_notes_agent as real_create_notes_agent

    def spy_create_notes_agent(**kwargs):
        captured.append({
            "template_type": kwargs["template_type"],
            "filing_level": kwargs["filing_level"],
            "filing_standard": kwargs.get("filing_standard", "mfrs"),
        })
        # Swap in a TestModel so no network key is required, then delegate
        # to the real factory to exercise notes_template_path routing.
        kwargs["model"] = TestModel()
        return real_create_notes_agent(**kwargs)

    # Short-circuit the iter() loop by patching _invoke_single_notes_agent_once
    # to return a canned outcome — the factory is what we care about, not
    # the agent run itself.
    from notes.coordinator import _SingleAgentOutcome

    async def fake_invoke(**kwargs):
        # Call the factory through our spy so captured[] is populated, then
        # return a canned outcome. The factory itself is fast (no network)
        # because we swap in a TestModel above.
        spy_create_notes_agent(
            template_type=kwargs["template_type"],
            pdf_path=kwargs["pdf_path"],
            inventory=kwargs["inventory"],
            filing_level=kwargs["filing_level"],
            model=kwargs["model"],
            output_dir=kwargs["output_dir"],
            page_hints=kwargs.get("page_hints"),
            page_offset=kwargs.get("page_offset", 0),
            filing_standard=kwargs.get("filing_standard", "mfrs"),
        )
        return _SingleAgentOutcome(
            filled_path=str(tmp_path / "NOTES_CORP_INFO_filled.xlsx"),
        )

    monkeypatch.setattr(ncoord, "_invoke_single_notes_agent_once", fake_invoke)

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    # Point utils.paths at tmp so the writable-output-dir check accepts us.
    import utils.paths as up
    monkeypatch.setattr(up, "PROJECT_ROOT", tmp_path.resolve())

    cfg = NotesRunConfig(
        pdf_path="/tmp/fake.pdf",
        output_dir=str(out_dir),
        model=TestModel(),
        notes_to_run={NotesTemplateType.CORP_INFO},
        filing_standard="mpers",
    )
    asyncio.run(run_notes_extraction(cfg))

    assert captured, "factory was never invoked"
    assert captured[0]["filing_standard"] == "mpers"
    assert captured[0]["template_type"] == NotesTemplateType.CORP_INFO


@pytest.mark.mpers_wiring_coordinator
def test_notes_agent_deps_carry_filing_standard(tmp_path) -> None:
    """create_notes_agent exposes filing_standard on its deps."""
    from pydantic_ai.models.test import TestModel
    from notes.agent import create_notes_agent

    agent, deps = create_notes_agent(
        template_type=NotesTemplateType.CORP_INFO,
        pdf_path="/tmp/no.pdf",
        inventory=[],
        filing_level="company",
        model=TestModel(),
        output_dir=str(tmp_path),
        filing_standard="mpers",
    )
    assert deps.filing_standard == "mpers"
    # Template path routed through the MPERS tree and the shifted 11- prefix.
    assert "XBRL-template-MPERS" in deps.template_path
    assert "11-Notes-CorporateInfo.xlsx" in deps.template_path


@pytest.mark.mpers_wiring_coordinator
def test_run_cli_parses_standard_flag() -> None:
    """run.py's real parser recognises --standard with choices mfrs/mpers.

    Peer-review LOW: earlier revisions of this test built their own
    argparse.Parser and only grepped run.py for "--standard" — a refactor
    that silently dropped the production flag would have slipped through.
    This version calls run.build_parser() directly."""
    from run import build_parser

    parser = build_parser()
    args = parser.parse_args(
        ["data/foo.pdf", "--standard", "mpers", "--statements", "SOFP"],
    )
    assert args.standard == "mpers"
    # Default stays MFRS — backward-compat contract for every
    # pre-existing CLI one-liner in CLAUDE.md and README.
    default_args = parser.parse_args(["data/foo.pdf", "--statements", "SOFP"])
    assert default_args.standard == "mfrs"


# ---------------------------------------------------------------------------
# Phase 3 — Server (mpers_wiring_server)
# ---------------------------------------------------------------------------

@pytest.mark.mpers_wiring_server
def test_run_config_request_accepts_mpers_filing_standard() -> None:
    from server import RunConfigRequest

    req = RunConfigRequest(statements=["SOFP"], filing_standard="mpers")
    assert req.filing_standard == "mpers"

    # Default is still mfrs so existing callers keep working.
    req2 = RunConfigRequest(statements=["SOFP"])
    assert req2.filing_standard == "mfrs"


@pytest.mark.mpers_wiring_server
def test_api_rejects_sore_on_mfrs_filing(tmp_path, monkeypatch) -> None:
    """Early validation: SoRE variant on an MFRS run must fail fast with a
    clear message naming SoRE and MPERS — before any template resolution."""
    import json
    import server
    from fastapi.testclient import TestClient

    session_id = "test-mpers-reject"
    out = tmp_path / "output"
    (out / session_id).mkdir(parents=True)
    (out / session_id / "uploaded.pdf").write_bytes(b"%PDF-1.4 fake")
    monkeypatch.setattr(server, "OUTPUT_DIR", out)
    monkeypatch.setattr(server, "AUDIT_DB_PATH", out / "xbrl_agent.db")
    fake_env = tmp_path / ".env-test"
    fake_env.write_text("")
    monkeypatch.setattr(server, "ENV_FILE", fake_env)
    monkeypatch.setenv("GOOGLE_API_KEY", "test-key-12345")
    monkeypatch.setenv("TEST_MODEL", "test-model-default")
    monkeypatch.setenv("LLM_PROXY_URL", "")
    client = TestClient(server.app)

    run_config = {
        "statements": ["SOCIE"],
        "variants": {"SOCIE": "SoRE"},
        "models": {},
        "infopack": None,
        "use_scout": False,
        "filing_standard": "mfrs",
    }
    with patch("server._create_proxy_model", return_value="fake-model"):
        resp = client.post(f"/api/run/{session_id}", json=run_config)

    assert resp.status_code == 200
    # Parse SSE payloads. Peer-review S9: three tight assertions so a
    # future wording drift or event-shape change doesn't silently pass:
    # (1) the run ends with success=False
    # (2) the error message names both SoRE and MPERS
    # (3) the run_complete event is present (not just an intermediate error)
    messages: list[str] = []
    run_complete_events: list[dict] = []
    for line in resp.text.splitlines():
        if line.startswith("data:"):
            try:
                payload = json.loads(line[len("data:"):].strip())
            except json.JSONDecodeError:
                continue
            if "message" in payload:
                messages.append(str(payload["message"]))
            # _fail_run emits a run_complete event with success=False.
            if payload.get("success") is False and "message" in payload:
                run_complete_events.append(payload)

    assert run_complete_events, (
        f"expected a run_complete event with success=False, got no matching events"
    )
    assert any("SoRE" in m and "MPERS" in m for m in messages), (
        f"expected SoRE/MPERS error message, got {messages}"
    )


# ---------------------------------------------------------------------------
# Phase 5 — Scout standard detection (mpers_wiring_scout)
# ---------------------------------------------------------------------------

@pytest.mark.mpers_wiring_scout
def test_infopack_has_detected_standard_default_unknown() -> None:
    """Fresh Infopack defaults to detected_standard='unknown'."""
    from scout.infopack import Infopack
    ip = Infopack(toc_page=1, page_offset=0)
    assert ip.detected_standard == "unknown"


@pytest.mark.mpers_wiring_scout
def test_infopack_json_roundtrip_preserves_detected_standard() -> None:
    """to_json/from_json must preserve the field; legacy JSON (no field)
    returns 'unknown' without raising."""
    from scout.infopack import Infopack
    ip = Infopack(toc_page=5, page_offset=3, detected_standard="mpers")
    raw = ip.to_json()
    restored = Infopack.from_json(raw)
    assert restored.detected_standard == "mpers"

    # Legacy JSON without detected_standard still deserialises to "unknown".
    legacy = '{"toc_page": 1, "page_offset": 0}'
    legacy_ip = Infopack.from_json(legacy)
    assert legacy_ip.detected_standard == "unknown"


@pytest.mark.mpers_wiring_scout
@pytest.mark.parametrize(
    "text, expected",
    [
        ("MPERS - Section 3 Statement of Retained Earnings", "mpers"),
        ("prepared in accordance with MFRS 101", "mfrs"),
        ("", "unknown"),
        # Both frameworks mentioned an equal number of times → tie → unknown.
        ("MFRS 101 and MPERS", "unknown"),
    ],
)
def test_detect_filing_standard_cases(text, expected) -> None:
    from scout.standard_detector import detect_filing_standard
    assert detect_filing_standard(text) == expected


@pytest.mark.mpers_wiring_scout
def test_scout_fills_detected_standard_from_toc_text(monkeypatch, tmp_path) -> None:
    """_find_toc_impl populates deps.detected_standard from the TOC text."""
    from scout.agent import ScoutDeps, _find_toc_impl
    import scout.agent as scout_agent_mod

    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"%PDF-1.4 placeholder")

    class _FakePage:
        def __init__(self, text): self._text = text
        def get_text(self): return self._text

    class _FakeDoc:
        def __init__(self, pages): self._pages = pages
        def __getitem__(self, idx): return self._pages[idx]
        def __len__(self): return len(self._pages)
        def close(self): pass

    class _FakeCandidate:
        def __init__(self, pn): self.page_number = pn

    toc_text = (
        "Table of Contents\n"
        "Prepared in accordance with MPERS - Section 3\n"
        "Statement of Retained Earnings .............. 12\n"
    )
    monkeypatch.setattr(
        scout_agent_mod, "find_toc_candidate_pages",
        lambda _path: [_FakeCandidate(1)],
    )
    monkeypatch.setattr(
        scout_agent_mod.fitz, "open",
        lambda _path: _FakeDoc([_FakePage(toc_text)]),
    )

    deps = ScoutDeps(
        pdf_path=pdf, pdf_length=100,
        statements_to_find=None, on_progress=None,
    )
    _find_toc_impl(deps)
    assert deps.detected_standard == "mpers"


@pytest.mark.mpers_wiring_scout
def test_scout_prefers_sore_variant_on_mpers_sore_page() -> None:
    """With MPERS standard active, a SoRE-shaped SOCIE page scores SoRE,
    not Default. MFRS runs stay on Default because SoRE is filtered out."""
    from scout.variant_detector import detect_variant_from_signals

    sore_page_text = (
        "Statement of Retained Earnings\n"
        "Retained earnings at beginning of period ...\n"
        "Dividends paid\n"
        "Retained earnings at end of period ...\n"
    )
    assert detect_variant_from_signals(
        StatementType.SOCIE, sore_page_text, standard="mpers",
    ) == "SoRE"

    # Same text through an MFRS run must not land on SoRE — the MFRS
    # registry only knows Default.
    assert detect_variant_from_signals(
        StatementType.SOCIE, sore_page_text, standard="mfrs",
    ) == "Default"


@pytest.mark.mpers_wiring_scout
def test_variant_detector_unknown_standard_falls_back_to_full_candidate_set() -> None:
    """Peer-review I4: when the caller passes `standard=None` (scout's
    'unknown' detection) the detector walks every registered variant,
    including MPERS-only ones. That's the current shipped behaviour — a
    future "tighten unknown → raise" refactor must confront this test."""
    from scout.variant_detector import detect_variant_from_signals

    sore_page_text = (
        "Statement of Retained Earnings\n"
        "Retained earnings at beginning of period ...\n"
        "Retained earnings at end of period ...\n"
    )
    # standard=None → no filter → SoRE is a candidate and wins the score.
    # If the detector is ever tightened (e.g. raise on unknown, or treat
    # unknown as MFRS), this assertion flips and the test author has to
    # decide whether the change is intended.
    assert detect_variant_from_signals(
        StatementType.SOCIE, sore_page_text, standard=None,
    ) == "SoRE"

    # Sanity: the "unknown" string path through _check_variant_signals_impl
    # (scout.agent:_check_variant_signals_impl) passes None here, so the
    # same input exercises the same code path as the scanned-PDF / ambiguous
    # TOC case where detect_filing_standard returned "unknown".


# ---------------------------------------------------------------------------
# Phase 6 — SoRE prompt (mpers_wiring_prompts)
# ---------------------------------------------------------------------------

@pytest.mark.mpers_wiring_e2e
def test_claude_md_mpers_status_updated() -> None:
    """Regression grep: CLAUDE.md section 15 must reflect the live wiring —
    the placeholder sentence from the pre-wiring draft is gone, and the
    `filing_standard` axis is listed in the Files-That-Must-Stay-in-Sync
    table. Without this pin, a future CLAUDE.md edit could quietly drop
    the MPERS guidance and new contributors would follow the stale note."""
    from pathlib import Path
    doc = Path(__file__).resolve().parent.parent / "CLAUDE.md"
    text = doc.read_text(encoding="utf-8")
    # The old placeholder is gone — case-insensitive so title-case
    # ("Not Yet Pipeline-Wired") and uppercase ("NOT yet pipeline-wired")
    # variants all trip the guard (peer-review C2: the case-sensitive
    # version let both known stale forms through silently).
    assert "not yet pipeline-wired" not in text.lower(), (
        "CLAUDE.md still claims MPERS templates are 'not yet pipeline-wired'"
    )
    # The new axis is listed in the Files-in-Sync table.
    assert "Filing standard" in text or "filing_standard" in text, (
        "CLAUDE.md missing a `filing_standard` Files-That-Must-Stay-in-Sync row"
    )


# Live smoke tests (Step 9.1 / 9.2). Skipped by default; set MPERS_TEST_PDF to
# the path of an MPERS filing and run `pytest -m live` to exercise the full
# pipeline. No mocks — real scout, real extraction, real cross-checks.

@pytest.mark.live
@pytest.mark.mpers_wiring_e2e
def test_e2e_mpers_company_smoke(tmp_path) -> None:
    import os
    pdf = os.environ.get("MPERS_TEST_PDF")
    if not pdf or not Path(pdf).exists():
        pytest.skip("MPERS_TEST_PDF not set or file missing — skipping MPERS live smoke")
    from coordinator import RunConfig, run_extraction
    import openpyxl

    output_dir = tmp_path / "mpers_co"
    output_dir.mkdir(parents=True)
    cfg = RunConfig(
        pdf_path=pdf,
        output_dir=str(output_dir),
        statements_to_run={StatementType.SOFP},
        variants={StatementType.SOFP: "CuNonCu"},
        filing_level="company",
        filing_standard="mpers",
    )
    asyncio.run(run_extraction(cfg))

    # Find any produced workbook; assert SOFP sheet exists and has no
    # MFRS-only rows (ROU / contract-asset labels that MPERS doesn't carry).
    produced = list(output_dir.rglob("*.xlsx"))
    assert produced, "no workbook produced by MPERS company run"
    wb = openpyxl.load_workbook(produced[0], data_only=False)
    assert any("SOFP" in s for s in wb.sheetnames), (
        f"no SOFP sheet in {wb.sheetnames}"
    )
    ws = next(wb[s] for s in wb.sheetnames if "SOFP" in s)
    labels = {
        str(ws.cell(row=r, column=1).value or "").strip().lower()
        for r in range(1, ws.max_row + 1)
    }
    # MPERS has no right-of-use asset or contract asset line items.
    for mfrs_only in ("right-of-use assets", "contract assets"):
        assert not any(mfrs_only in l for l in labels), (
            f"{mfrs_only!r} present in MPERS SOFP — template swap did not take effect"
        )


@pytest.mark.live
@pytest.mark.mpers_wiring_e2e
def test_e2e_mpers_group_sore_smoke(tmp_path) -> None:
    import os
    pdf = os.environ.get("MPERS_TEST_PDF")
    if not pdf or not Path(pdf).exists():
        pytest.skip("MPERS_TEST_PDF not set or file missing — skipping MPERS live smoke")
    from coordinator import RunConfig, run_extraction
    from cross_checks.framework import run_all as run_cross_checks
    from server import _build_default_cross_checks

    output_dir = tmp_path / "mpers_gr"
    output_dir.mkdir(parents=True)
    cfg = RunConfig(
        pdf_path=pdf,
        output_dir=str(output_dir),
        statements_to_run={StatementType.SOFP, StatementType.SOCIE},
        variants={
            StatementType.SOFP: "CuNonCu",
            StatementType.SOCIE: "SoRE",
        },
        filing_level="group",
        filing_standard="mpers",
    )
    result = asyncio.run(run_extraction(cfg))
    # Gather the per-statement workbook paths coordinator produced.
    workbook_paths = {
        r.statement_type: r.workbook_path for r in result
        if getattr(r, "workbook_path", None)
    }
    check_config = {
        "statements_to_run": cfg.statements_to_run,
        "variants": cfg.variants,
        "filing_level": "group",
        "filing_standard": "mpers",
    }
    results = run_cross_checks(
        _build_default_cross_checks(), workbook_paths, check_config,
    )
    by_name = {r.name: r for r in results}
    # SOCIE-consuming checks must short-circuit on SoRE.
    for name in (
        "sopl_to_socie_profit",
        "soci_to_socie_tci",
        "socie_to_sofp_equity",
    ):
        if name in by_name:
            assert by_name[name].status == "not_applicable", (
                f"{name} should have been gated out on SoRE, got {by_name[name].status}"
            )
    # The SoRE-specific check must have actually fired (passed or failed is
    # fine on a real extraction; not_applicable means the gate mis-fired).
    sore = by_name.get("sore_to_sofp_retained_earnings")
    assert sore is not None, "SoRE cross-check not registered"
    assert sore.status != "not_applicable", (
        f"SoRE check did not fire (status={sore.status}); registry gate regressed"
    )


@pytest.mark.mpers_wiring_prompts
def test_render_prompt_socie_sore_uses_variant_file() -> None:
    """prompts/socie_sore.md is picked up for the SoRE variant and includes
    MPERS-specific wording (the retained-earnings movement schedule), not
    the SOCIE matrix layout from socie.md."""
    from prompts import render_prompt

    rendered = render_prompt(
        StatementType.SOCIE, "SoRE", filing_level="company",
        filing_standard="mpers",
    )
    # SoRE-specific phrase — must come from socie_sore.md.
    assert "Retained earnings at end of period" in rendered
    # Sanity: MATRIX-layout phrasing from socie.md must NOT leak in.
    assert "MATRIX layout" not in rendered
