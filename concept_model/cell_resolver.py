"""Reverse cell→fact resolver + write projection (Phase B).

The extraction agent writes resolved ``(sheet, row, col, value)`` cells into a
scratch xlsx. In canonical mode each write is *also* projected into
``run_concept_facts`` so the DB — not the agent's xlsx — becomes the
authoritative fact store the Concepts UI and the DB exporter read.

``resolve_cell`` is the linchpin: it maps a written cell back to its
``(concept_uuid, period, entity_scope)``. Group and matrix (SOCIE) templates
carry per-(scope, period) coordinates in ``concept_targets`` (column letters
B/C/D/E…), so a reverse lookup there handles every dimension uniformly. Linear
Company filings have no targets — they fall back to ``concept_nodes`` keyed on
``(render_sheet, render_row)`` plus the fixed B=CY / C=PY column convention.
"""
from __future__ import annotations

import logging
import sqlite3
from dataclasses import dataclass, field
from typing import Optional

from fastapi import HTTPException
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from concept_model.facts_api import FactWrite, apply_fact

logger = logging.getLogger(__name__)


def resolve_cell(
    conn: sqlite3.Connection,
    template_id: str,
    sheet: str,
    row: int,
    col_num: int,
) -> Optional[tuple[str, str, str]]:
    """Map a written cell to ``(concept_uuid, period, entity_scope)``.

    Returns ``None`` when the cell can't be mapped to a concept (e.g. an
    evidence/source column, or a row with no matching concept_node) — the
    caller skips it rather than failing the whole projection.
    """
    col_letter = get_column_letter(col_num)

    # 1. concept_targets reverse lookup — Group 6-col + SOCIE matrix carry
    # an explicit (scope, period) per target cell.
    row_t = conn.execute(
        "SELECT ct.concept_uuid, ct.entity_scope, ct.period "
        "FROM concept_targets ct "
        "JOIN concept_nodes n ON n.concept_uuid = ct.concept_uuid "
        "WHERE n.template_id = ? AND ct.target_sheet = ? "
        "AND ct.target_row = ? AND ct.target_col = ?",
        (template_id, sheet, row, col_letter),
    ).fetchone()
    if row_t is not None:
        # Contract is (concept_uuid, period, entity_scope); the SELECT
        # returns (concept_uuid, entity_scope, period) — swap the dims.
        return (row_t[0], row_t[2], row_t[1])

    # 2. Company linear fallback — no targets; key on render coordinates
    # and the fixed column convention (B=CY, C=PY, Company scope).
    node = conn.execute(
        "SELECT concept_uuid FROM concept_nodes "
        "WHERE template_id = ? AND render_sheet = ? AND render_row = ?",
        (template_id, sheet, row),
    ).fetchone()
    if node is None:
        # 2b. Alias fallback — a concept that surfaces on more than one
        # physical sheet (the cross-sheet rollup case: face row anchors
        # the same canonical concept as a sub-sheet *Total) stores the
        # demoted face coord in ``concept_render_aliases``. Without
        # this lookup an agent write to the face cell would be silently
        # skipped even though the canonical UUID exists. Join through
        # concept_nodes so the template_id scope is honoured.
        alias = conn.execute(
            "SELECT a.concept_uuid FROM concept_render_aliases a "
            "JOIN concept_nodes n ON n.concept_uuid = a.concept_uuid "
            "WHERE n.template_id = ? "
            "AND a.alias_sheet = ? AND a.alias_row = ?",
            (template_id, sheet, row),
        ).fetchone()
        if alias is None:
            return None
        node = alias
    if col_num == 2:
        period = "CY"
    elif col_num == 3:
        period = "PY"
    else:
        # Any other column on a Company linear template is non-value
        # (evidence/source) — nothing to project.
        return None
    return (node[0], period, "Company")


@dataclass
class ProjectionResult:
    """Outcome of projecting a batch of cell writes into facts.

    ``projected`` landed as facts; ``skipped`` didn't resolve to a concept
    (evidence columns, or a cell whose concept isn't on this render sheet);
    ``rejected`` resolved but the facts API refused (e.g. a value on a
    formula concept). ``skipped``/``rejected`` carry human-readable cell
    descriptors so the caller can surface *which* cells were dropped.
    """
    projected: int = 0
    skipped: list[str] = field(default_factory=list)
    rejected: list[str] = field(default_factory=list)

    @property
    def has_gaps(self) -> bool:
        return bool(self.skipped or self.rejected)


def project_writes(
    db_path,
    run_id: int,
    template_id: str,
    writes,
    *,
    filing_level: str = "company",
) -> ProjectionResult:
    """Project resolved cell writes into ``run_concept_facts``.

    ``writes`` is an iterable of dicts with ``sheet``, ``row``, ``col``,
    ``value`` and optional ``evidence``. Returns a :class:`ProjectionResult`
    so callers can see how many facts landed vs were skipped/rejected — a
    single bad cell never aborts a run, but the gaps are no longer silent.
    """
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.row_factory = sqlite3.Row
    result = ProjectionResult()
    try:
        for w in writes:
            sheet = w["sheet"]
            row = int(w["row"])
            col = int(w["col"])
            cell_desc = f"{sheet}!{get_column_letter(col)}{row}"
            resolved = resolve_cell(conn, template_id, sheet, row, col)
            if resolved is None:
                result.skipped.append(cell_desc)
                logger.debug(
                    "project_writes: unresolved cell %s (template %s)",
                    cell_desc, template_id,
                )
                continue
            concept_uuid, period, entity_scope = resolved
            try:
                # Defer the commit — the whole batch lands in one transaction
                # so a mid-batch crash rolls back cleanly (no half-projected
                # run) instead of leaving committed-up-to-cell-N state.
                apply_fact(
                    conn,
                    run_id,
                    FactWrite(
                        concept_uuid=concept_uuid,
                        period=period,
                        entity_scope=entity_scope,
                        value=w.get("value"),
                        value_status="observed",
                        source=w.get("evidence") or None,
                        evidence=w.get("evidence") or None,
                        actor="extraction",
                    ),
                    commit=False,
                )
                result.projected += 1
            except HTTPException as exc:
                result.rejected.append(f"{cell_desc} ({exc.detail})")
                logger.info(
                    "project_writes: facts API rejected %s — %s",
                    cell_desc, exc.detail,
                )
            except ValidationError as exc:
                # The agent-controlled `value` is the only free field, and the
                # tool layer accepts Union[int, float, str] (for row-1 date
                # cells). A non-numeric value on a *resolvable* concept — e.g. a
                # text-disclosure/title row — makes FactWrite(value=...) raise
                # here because facts_api.FactWrite.value is Optional[float].
                # Before this catch it escaped to the outer rollback and aborted
                # the ENTIRE batch (run 49: one SOCI text row nuked both numeric
                # profit facts -> projection_failed). Reject this one cell and
                # keep projecting the numeric facts — restoring the docstring's
                # "a single bad cell never aborts a run" contract.
                # Truncate the raw value: a text/title row can carry a long
                # sentence and `rejected` surfaces into the run/review summary.
                raw = repr(w.get("value"))
                if len(raw) > 80:
                    raw = raw[:77] + "..."
                result.rejected.append(f"{cell_desc} (non-numeric value {raw})")
                logger.info(
                    "project_writes: non-numeric/invalid value for %s — %s",
                    cell_desc, exc,
                )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return result
