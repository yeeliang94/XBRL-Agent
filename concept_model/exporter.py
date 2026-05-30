"""Phase 1 step 1.11-1.14 — DB-backed Excel exporter.

Reads facts from ``run_concept_facts`` and writes them into a fresh
copy of the template ``.xlsx``.  Three rules from the PRD that this
module owns:

* **Canonical label only in col A** — the UI's ``display_label`` is
  never exported (PRD §9 boundary).  We honour this by ignoring the
  display_label column entirely.
* **aggregate_only branch** — when a COMPUTED parent's
  ``children_status='aggregate_only'``, we replace the live formula
  with the literal value AND annotate the source column so M-Tool
  reviewers know why the breakdown is missing.
* **not_disclosed branch** — leaves marked ``not_disclosed`` stay
  blank in Excel; a side-channel JSON next to the xlsx documents
  which leaves were intentionally blank (so a downstream reviewer can
  distinguish "agent didn't try" from "agent confirmed missing").
"""
from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any

import openpyxl


_DEFAULT_SOURCE_COL = {
    "company": "D",   # MFRS/MPERS Company templates: D = Source
    "group": "F",     # Group templates: F = Source
}

# Which entity scopes a filing level renders. A fact whose scope isn't in
# this set has no place on the workbook (e.g. a Group-scope fact on a
# Company filing — the Group columns only exist on Group templates) and is
# dropped. A fact whose scope IS applicable but has no precomputed
# concept_targets row is an importer bug, surfaced loudly (see below).
_APPLICABLE_SCOPES = {
    "company": ("Company",),
    "group": ("Group", "Company"),
}


def _is_date_placeholder(value) -> bool:
    """True for an unfilled reporting-period date-header cell."""
    return isinstance(value, str) and "YYYY" in value


def _stamp_period_headers(
    wb,
    reporting_period_cy: str | None,
    reporting_period_py: str | None,
    carry_forward_row1_from: str | Path | None,
) -> None:
    """Fill the placeholder reporting-period date headers in-place.

    Scout metadata first (by column parity), else carry-forward from the
    agent's scratch workbook. See the call site for the full rationale. Both
    paths only overwrite cells still holding the "YYYY" placeholder.
    """
    import logging

    if reporting_period_cy or reporting_period_py:
        for d_ws in wb.worksheets:
            for drow in d_ws.iter_rows(min_row=1, max_row=3):
                for d_cell in drow:
                    if not _is_date_placeholder(d_cell.value):
                        continue
                    is_cy = d_cell.column % 2 == 0  # B,D -> CY; C,E -> PY
                    new = reporting_period_cy if is_cy else reporting_period_py
                    if new:
                        d_cell.value = new
        return

    # Fallback: carry the agent-typed dates over from the scratch workbook.
    if not (carry_forward_row1_from and Path(carry_forward_row1_from).exists()):
        return
    src_wb = None
    try:
        src_wb = openpyxl.load_workbook(carry_forward_row1_from, data_only=False)
        for sheet in wb.sheetnames:
            if sheet not in src_wb.sheetnames:
                continue
            s_ws, d_ws = src_wb[sheet], wb[sheet]
            for drow in d_ws.iter_rows(min_row=1, max_row=3):
                for d_cell in drow:
                    if not _is_date_placeholder(d_cell.value):
                        continue
                    s_val = s_ws.cell(row=d_cell.row, column=d_cell.column).value
                    if isinstance(s_val, str) and s_val and "YYYY" not in s_val:
                        d_cell.value = s_val
    except Exception:  # noqa: BLE001 — a carry-forward hiccup must not sink the export
        logging.getLogger(__name__).warning(
            "reporting-period carry-forward from %s failed",
            carry_forward_row1_from, exc_info=True,
        )
    finally:
        if src_wb is not None:
            src_wb.close()


def export_run_to_xlsx(
    db_path: str | Path,
    run_id: int,
    xlsx_path: str | Path,
    *,
    filing_level: str = "company",
    template_id: str | None = None,
    reporting_period_cy: str | None = None,
    reporting_period_py: str | None = None,
    carry_forward_row1_from: str | Path | None = None,
) -> int:
    """Fill ``xlsx_path`` in-place with the canonical-mode facts.

    The template is expected to already exist at ``xlsx_path`` — the
    caller copies the master template into the run's output dir before
    invoking us.  This keeps the formula network intact for any rows we
    don't override.

    ``carry_forward_row1_from`` is the agent's scratch workbook. Row 1 holds
    the reporting-period date headers (e.g. "01/01/2021 - 31/12/2021"), which
    are NOT XBRL concepts so they never project to ``run_concept_facts`` — they
    land in ``proj.skipped``. The fresh template copy we fill carries the
    literal placeholder "01/01/YYYY - 31/12/YYYY" there, so a download rendered
    purely from facts would ship placeholder dates on every face sheet. When a
    scratch path is given we copy its real row-1 date cells over (value columns
    only — never col A or the Source header), keeping the fact-render faithful.

    ``template_id`` scopes the fact query (and the unmapped-targets check)
    to a single template. The per-statement export pass passes it so one
    statement's unmapped Group/matrix fact can't fail every other
    statement's export, and so the returned count reflects only this
    template (peer-review findings 1 + 3). When ``None`` the whole run is
    exported (legacy callers / tests).

    Returns the number of facts that applied to a sheet present in this
    workbook — the caller uses a zero count to avoid repointing a download
    at a blank template.
    """
    db_path = str(db_path)
    xlsx_path = str(xlsx_path)
    source_col = _DEFAULT_SOURCE_COL.get(filing_level, "D")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        # Pull every fact joined to its concept_node and template shape,
        # LEFT JOINing ``concept_targets`` so each (period, entity_scope)
        # dimension routes to its dedicated cell. As of Phase 6.1 routing is
        # a SINGLE keyed lookup for every shape: the importer precomputes a
        # concept_targets row for every dimension a filing renders
        # (import_company_targets B=CY/C=PY, import_group_targets B/C/D/E,
        # matrix targets inline). The exporter no longer falls back to
        # ``concept_nodes.render_col`` (the render_* columns remain in the
        # SELECT but are unused by routing now). Facts whose scope doesn't
        # apply to this filing are dropped; an applicable fact with no target
        # row is an importer bug and raises.
        rows = conn.execute(
            """
            SELECT f.concept_uuid, f.period, f.entity_scope,
                   f.value, f.value_status, f.children_status,
                   f.source, f.evidence,
                   n.canonical_label, n.kind, n.render_sheet,
                   n.render_row, n.render_col,
                   tpl.shape AS shape,
                   t.target_col AS target_col,
                   t.target_sheet AS target_sheet,
                   t.target_row AS target_row
            FROM run_concept_facts f
            JOIN concept_nodes n ON n.concept_uuid = f.concept_uuid
            JOIN concept_templates tpl ON tpl.template_id = n.template_id
            LEFT JOIN concept_targets t
              ON t.concept_uuid = f.concept_uuid
              AND t.period = f.period
              AND t.entity_scope = f.entity_scope
            WHERE f.run_id = ?
              AND (? IS NULL OR n.template_id = ?)
            """,
            (run_id, template_id, template_id),
        ).fetchall()
    finally:
        conn.close()

    # Resolve each fact to a physical (sheet, row, col) target via a single
    # concept_targets lookup. Matrix (SOCIE) always routes via targets for
    # every dimension. Linear facts whose entity_scope doesn't apply to this
    # filing level (e.g. a Group-scope fact on a Company filing) are dropped;
    # an applicable fact with no target row is an importer bug → raise.
    applicable = _APPLICABLE_SCOPES.get(
        (filing_level or "").lower(), ("Company",)
    )
    routed: list[dict[str, Any]] = []
    unmapped: list[tuple] = []
    for r in rows:
        is_matrix = r["shape"] == "matrix"
        if not is_matrix and r["entity_scope"] not in applicable:
            # Out-of-scope fact for this filing level — no cell to land in.
            continue
        if r["target_col"] is None:
            # In-scope (or matrix) fact with no precomputed target: the
            # importer (import_company_targets / import_group_targets / the
            # matrix importer) didn't map this dimension. Surface loudly.
            unmapped.append((r["concept_uuid"], r["period"], r["entity_scope"]))
            continue
        sheet, row, col = r["target_sheet"], int(r["target_row"]), r["target_col"]
        routed.append({
            "sheet": sheet, "row": row, "col": col,
            "concept_uuid": r["concept_uuid"],
            "canonical_label": r["canonical_label"],
            "kind": r["kind"],
            "shape": r["shape"],
            "value": r["value"],
            "value_status": r["value_status"],
            "children_status": r["children_status"],
            "source": r["source"],
        })

    if unmapped:
        raise ValueError(
            "Export found facts with no concept_targets mapping "
            f"(run {run_id}): {unmapped[:5]}"
            f"{'…' if len(unmapped) > 5 else ''}. Run import_company_targets / "
            "import_group_targets (or re-import matrix templates) for every "
            "template in the run."
        )

    not_disclosed: list[dict[str, Any]] = []
    applied = 0

    for r in routed:
        sheet = r["sheet"]
        row = int(r["row"])
        col = r["col"] or "B"
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        # A fact whose sheet is in this workbook counts as applied — the
        # caller uses applied>0 to decide whether the export is worth
        # repointing the download at (vs keeping the agent's scratch xlsx).
        applied += 1

        # Column A: canonical label only.  This is idempotent — the
        # template already carries it — but we re-stamp so a Phase-5
        # template-schema drift wouldn't silently leak the old text.
        ws[f"A{row}"] = r["canonical_label"]

        if r["value_status"] == "not_disclosed":
            # Side-channel only; cell stays blank.  Don't touch
            # whatever formula was there (a downstream user might
            # re-fill it).  Preserve the formula by writing back its
            # original value verbatim.
            not_disclosed.append({
                "concept_uuid": r["concept_uuid"],
                "sheet": sheet,
                "row": row,
                "label": r["canonical_label"],
            })
            continue

        if (
            r["kind"] == "COMPUTED"
            and r["children_status"] == "aggregate_only"
            and r["value"] is not None
        ):
            # Replace the live formula with the literal value and
            # annotate the source column so reviewers know why the
            # underlying breakdown is missing.
            cell = ws[f"{col}{row}"]
            cell.value = float(r["value"])
            annotation = "aggregate_only"
            if r["source"]:
                annotation = f"{annotation} — {r['source']}"
            ws[f"{source_col}{row}"] = annotation
            continue

        if r["value"] is not None:
            target = ws[f"{col}{row}"]
            # Preserve live template formulas for *itemised* totals. The PRD
            # promise is "Excel formulas stay live for itemised concepts; only
            # aggregate_only (handled above) replaces a formula with a
            # literal". The cascade persists COMPUTED / matrix totals computed
            # from their children as observed + children_status='itemised' —
            # for those we leave the live formula so Excel/M-Tool recompute it
            # rather than clobbering it with a stale snapshot literal. A raw
            # observed value (children_status NULL — a direct override) or a
            # leaf still gets its literal written.
            existing = target.value
            is_formula_cell = isinstance(existing, str) and existing.startswith("=")
            if is_formula_cell and r["children_status"] == "itemised":
                pass
            else:
                target.value = float(r["value"])

        # Stamp the source column when the fact carries one — but NEVER on
        # matrix (SOCIE) templates. There the fixed source_col (D/F) lands
        # *inside* the equity-component value grid (MFRS spans B..X), so
        # stamping a provenance string would clobber a real Treasury-shares
        # / reserve value. SOCIE source + evidence stay in the DB
        # (run_concept_facts) — the xlsx is a flattened snapshot, the
        # canonical store is the source of truth (cf. gotcha #16 for notes).
        # Per-cell source on a 23-column matrix row is also semantically
        # ambiguous (which component does it describe?), so skipping is
        # both safe and correct.
        if (
            r["shape"] != "matrix"
            and r["source"]
            and r["children_status"] != "aggregate_only"
        ):
            ws[f"{source_col}{row}"] = r["source"]

    # Reporting-period date headers are NON-concept cells (they never project
    # to ``run_concept_facts``), so the fresh template copy keeps the literal
    # "01/01/YYYY - 31/12/YYYY" placeholder. Populate them deterministically.
    #
    # The period is RUN-LEVEL metadata — the same CY / PY string on every face
    # statement, just repeated across Group's column pairs — so we never ask the
    # five extraction agents to re-type it. Two sources, in priority order:
    #
    #   1. Scout-captured period (``reporting_period_cy`` / ``_py``). Stamped by
    #      column PARITY, which holds across every layout: value columns are
    #      CY, PY, CY, PY from B — Company B(CY)/C(PY); Group B(GrpCY)/C(GrpPY)/
    #      D(CoCY)/E(CoPY); SOCIE B(CY) only. Even cols (B=2,D=4) → CY, odd
    #      (C=3,E=5) → PY. Layout-independent and fixes Group, whose dates live
    #      in row 2 (row 1 holds the Group/Company labels) — a row the writer
    #      guard won't let an agent fill.
    #   2. Fallback for no-scout runs: the agent's scratch workbook, where it
    #      typed the dates into row 1 itself. Copied cell-for-cell.
    #
    # Both only ever overwrite a cell that still holds the "YYYY" placeholder,
    # so they are self-targeting — never a data, label, or Source cell.
    _stamp_period_headers(
        wb, reporting_period_cy, reporting_period_py, carry_forward_row1_from
    )

    wb.save(xlsx_path)

    # Side-channel JSON for not_disclosed leaves.  Always written
    # (even if empty) so downstream tools don't have to special-case
    # absence.
    side = Path(xlsx_path + ".not_disclosed.json")
    side.write_text(
        json.dumps({"run_id": run_id, "entries": not_disclosed},
                   indent=2, sort_keys=True),
        encoding="utf-8",
    )
    return applied
