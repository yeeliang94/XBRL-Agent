"""Taxonomy identities for canonical template cells.

The canonical parser historically retained only the visible Excel label.  That
is insufficient for mTool filing: labels repeat, while a filing fact is
identified by its taxonomy concept plus any dimensions.  This module restores
that identity from the same presentation-linkbase roles that generated the
repository templates.

The public interface is deliberately small: :func:`semantic_addresses_for`
returns a coordinate-keyed mapping for one template.  Callers do not need to
know role numbers, taxonomy paths, SOCIE row blocks, or component-member ids.
Failure is fail-closed: an unrecognised or drifted template receives no
semantic address and remains usable by the extraction pipeline, but mTool
readiness will report the missing filing identity.
"""
from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

import openpyxl


TAXONOMY_VERSION = "SSMxT_2022v1.0"
ADDRESS_VERSION = "2022-v1"

# The presentation roles are the source used by the template generators.
_ROLES_BY_FILE: dict[str, tuple[str, ...]] = {
    "01-SOFP-CuNonCu.xlsx": ("210000", "210100"),
    "02-SOFP-OrderOfLiquidity.xlsx": ("220000", "220100"),
    "03-SOPL-Function.xlsx": ("310000", "310100"),
    "04-SOPL-Nature.xlsx": ("320000", "320100"),
    "05-SOCI-BeforeTax.xlsx": ("420000",),
    "06-SOCI-NetOfTax.xlsx": ("410000",),
    "07-SOCF-Indirect.xlsx": ("520000",),
    "08-SOCF-Direct.xlsx": ("510000",),
    "09-SOCIE.xlsx": ("610000",),
    "10-SoRE.xlsx": ("620000",),
    "10-Notes-CorporateInfo.xlsx": ("710000",),
    "11-Notes-CorporateInfo.xlsx": ("710000",),
    "11-Notes-AccountingPolicies.xlsx": ("720000",),
    "12-Notes-AccountingPolicies.xlsx": ("720000",),
    "12-Notes-ListOfNotes.xlsx": ("730000",),
    "13-Notes-ListOfNotes.xlsx": ("730000",),
    "13-Notes-IssuedCapital.xlsx": ("740000",),
    "14-Notes-RelatedParty.xlsx": ("750000",),
    "14-Notes-IssuedCapital.xlsx": ("740000",),
    "15-Notes-RelatedParty.xlsx": ("750000",),
}

# These presentation wrappers exist in the MFRS roles but have no physical
# row in the reviewed SSM workbooks.  Naming them here makes the exception
# auditable and prevents the old one-title-row tolerance from shifting every
# later field by one row.
_REVIEWED_ROLE_OMISSIONS: dict[tuple[str, str], frozenset[str]] = {
    ("mfrs", "13-Notes-IssuedCapital.xlsx"): frozenset({
        "ifrs-full_DisclosureOfClassesOfShareCapitalAbstract",
    }),
    ("mfrs", "14-Notes-RelatedParty.xlsx"): frozenset({
        "ifrs-full_DisclosureOfTransactionsBetweenRelatedPartiesAbstract",
    }),
}


@dataclass(frozen=True)
class TaxonomyConcept:
    taxonomy_version: str
    namespace_uri: str
    local_name: str
    source_element_id: str
    abstract: bool
    concept_role: str
    data_type: str | None
    period_type: str | None
    balance: str | None
    substitution_group: str | None

    @property
    def reportable(self) -> bool:
        return self.concept_role == "PRIMARY_ITEM" and not self.abstract


def _concept_role(element_id: str, attrs: dict[str, str]) -> str:
    if attrs.get("abstract", "false").lower() == "true":
        return "ABSTRACT"
    substitution = attrs.get("substitutionGroup", "")
    local_name = attrs.get("name", "")
    if substitution.endswith("dimensionItem") or local_name.endswith("Axis"):
        return "DIMENSION"
    if substitution.endswith("hypercubeItem") or local_name.endswith("Table"):
        return "HYPERCUBE"
    if local_name.endswith("Member") or element_id.endswith("Member"):
        return "MEMBER"
    if local_name.endswith("LineItems") or element_id.endswith("LineItems"):
        return "LINE_ITEMS"
    return "PRIMARY_ITEM"


@lru_cache(maxsize=1)
def taxonomy_registry() -> dict[str, TaxonomyConcept]:
    """Read the committed SSM schemas into an element-id registry."""
    root = Path(__file__).resolve().parent.parent / "SSMxT_2022v1.0"
    xs_element = "{http://www.w3.org/2001/XMLSchema}element"
    xbrli = "{http://www.xbrl.org/2003/instance}"
    out: dict[str, TaxonomyConcept] = {}
    for path in root.rglob("*.xsd"):
        try:
            schema = ET.parse(path).getroot()
        except ET.ParseError:
            continue
        namespace_uri = schema.attrib.get("targetNamespace", "")
        for element in schema.findall(xs_element):
            element_id = element.attrib.get("id")
            local_name = element.attrib.get("name")
            if not element_id or not local_name:
                continue
            attrs = dict(element.attrib)
            concept = TaxonomyConcept(
                taxonomy_version=TAXONOMY_VERSION,
                namespace_uri=namespace_uri,
                local_name=local_name,
                source_element_id=element_id,
                abstract=attrs.get("abstract", "false").lower() == "true",
                concept_role=_concept_role(element_id, attrs),
                data_type=attrs.get("type"),
                period_type=attrs.get(f"{xbrli}periodType"),
                balance=attrs.get(f"{xbrli}balance"),
                substitution_group=attrs.get("substitutionGroup"),
            )
            previous = out.get(element_id)
            if previous is not None and previous != concept:
                raise ValueError(
                    f"Taxonomy element id {element_id!r} is defined inconsistently"
                )
            out[element_id] = concept
    return out


def taxonomy_concept(source_element_id: str) -> TaxonomyConcept | None:
    return taxonomy_registry().get(source_element_id)

# MFRS SOCIE's columns are explicit members of ComponentsOfEquityAxis.  The
# presentation linkbase contains the core members; the SSM extension members
# below are defined elsewhere in the same taxonomy package.
_MFRS_SOCIE_COMPONENTS: tuple[str, ...] = (
    "ifrs-full_IssuedCapitalMember",
    "ifrs-full_RetainedEarningsMember",
    "ifrs-full_TreasurySharesMember",
    "ifrs-full_CapitalReserveMember",
    "ifrs-full_ReserveOfGainsAndLossesOnHedgingInstrumentsThatHedgeInvestmentsInEquityInstrumentsMember",
    "ssmt-mfrs_ForeignCurrencyTranslationReserveMember",
    "ifrs-full_ReserveOfSharebasedPaymentsMember",
    "ifrs-full_RevaluationSurplusMember",
    "ifrs-full_StatutoryReserveMember",
    "ifrs-full_WarrantReserveMember",
    "ssmt-mfrs_OtherNondistributableReserveMember",
    "ssmt_NonDistributableReservesMember",
    "ssmt-mfrs_FairValueAdjustmentReserveMember",
    "ssmt_ReserveOfNoncurrentAssetsClassifiedAsHeldForSaleMember",
    "ssmt-mfrs_ConsolidatedReserveMember",
    "ssmt-mfrs_WarrantyReserveMember",
    "ssmt-mfrs_OtherDistributableReserveMember",
    "ssmt_DistributableReservesMember",
    "ssmt_ReservesMember",
    "ifrs-full_EquityAttributableToOwnersOfParentMember",
    "ssmt-mfrs_OtherComponentsOfEquityMember",
    "ifrs-full_NoncontrollingInterestsMember",
    "ifrs-full_EquityMember",
)

def _standard_and_level(path: Path) -> tuple[str, str] | None:
    lowered = [part.lower() for part in path.parts]
    standard = (
        "mpers" if any("xbrl-template-mpers" in part for part in lowered)
        else "mfrs" if any("xbrl-template-mfrs" in part for part in lowered)
        else None
    )
    level = "group" if "group" in lowered else "company" if "company" in lowered else None
    if standard is None or level is None:
        return None
    return standard, level


@lru_cache(maxsize=64)
def _role_rows(standard: str, role: str) -> tuple[tuple[int, str, str, bool], ...]:
    """Read one role through the generator's tested presentation walker.

    The generator's public adapter owns its cache/context isolation, so this
    consumer does not reach into generator-private globals.
    """
    from scripts import generate_mpers_templates as taxonomy

    root = Path(__file__).resolve().parent.parent
    tax_dir = root / "SSMxT_2022v1.0/rep/ssm/ca-2016/fs" / standard
    return tuple(taxonomy.walk_role_for_taxonomy(tax_dir, standard, role))


def _address(primary: str, dimensions: dict[str, str] | None = None) -> dict[str, Any]:
    concept = taxonomy_concept(primary)
    address = {
        "primary_concept": primary,
        "dimensions": dict(sorted((dimensions or {}).items())),
        "taxonomy_version": TAXONOMY_VERSION,
        "address_version": ADDRESS_VERSION,
    }
    if concept is not None:
        address.update({
            "namespace_uri": concept.namespace_uri,
            "local_name": concept.local_name,
            "concept_role": concept.concept_role,
            "abstract": concept.abstract,
            "reportable": concept.reportable,
            "data_type": concept.data_type,
            "period_type": concept.period_type,
            "balance": concept.balance,
        })
    return address


def _linear_addresses(path: Path, standard: str) -> dict[tuple[str, int, str | None], dict[str, Any]]:
    roles = _ROLES_BY_FILE.get(path.name)
    if not roles:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        if len(wb.worksheets) != len(roles):
            return {}
        out: dict[tuple[str, int, str | None], dict[str, Any]] = {}
        for ws, role in zip(wb.worksheets, roles):
            sheet_rows = [
                row for row in range(1, ws.max_row + 1)
                if ws.cell(row, 1).value not in (None, "")
            ]
            taxonomy_rows = list(_role_rows(standard, role))
            if path.name in {"13-Notes-IssuedCapital.xlsx", "14-Notes-RelatedParty.xlsx"}:
                omissions = _REVIEWED_ROLE_OMISSIONS.get((standard, path.name), frozenset())
                taxonomy_rows = [
                    item for item in taxonomy_rows
                    if not item[1].lower().endswith(
                        ("table", "axis", "member", "lineitems")
                    )
                    and item[1] not in omissions
                ]
            # One MFRS SOCI template carries a display-only title above the exact
            # linkbase rows.  A larger discrepancy means the template drifted and
            # must not receive guessed identities.
            if (
                standard == "mfrs"
                and path.name == "05-SOCI-BeforeTax.xlsx"
                and len(sheet_rows) == len(taxonomy_rows) + 1
            ):
                sheet_rows = sheet_rows[1:]
            if len(sheet_rows) != len(taxonomy_rows):
                continue
            for row, (_depth, concept_id, _label, _abstract) in zip(sheet_rows, taxonomy_rows):
                out[(ws.title, row, None)] = _address(concept_id)
        return out
    finally:
        wb.close()


def _matrix_addresses(path: Path, standard: str, level: str) -> dict[tuple[str, int, str | None], dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb["SOCIE"] if "SOCIE" in wb.sheetnames else wb[wb.sheetnames[0]]
        taxonomy_rows = list(_role_rows(standard, "610000"))

        if standard == "mpers":
            # MPERS's single value column is not a component dimension.  The
            # canonical parser starts at Equity and excludes the two title rows.
            line_items = taxonomy_rows[2:]
            # Company begins at row 5; Group has one additional block heading.
            base_row = 6 if level == "group" else 5
            rows = list(range(base_row, base_row + len(line_items)))
            return {
                (ws.title, row, "B"): _address(concept_id)
                for row, (_depth, concept_id, _label, _abstract) in zip(rows, line_items)
            }

        marker = "ifrs-full_StatementOfChangesInEquityLineItems"
        try:
            start = next(i for i, item in enumerate(taxonomy_rows) if item[1] == marker) + 1
        except StopIteration:
            return {}
        line_items = taxonomy_rows[start:]
        # The MFRS canonical matrix uses its first CY block, rows 6..25, as the
        # concept home; later blocks are render targets for period/scope.
        rows = list(range(6, 6 + len(line_items)))
        if len(line_items) != 20 or ws.max_column < 24:
            return {}
        axis = "ifrs-full_ComponentsOfEquityAxis"
        out: dict[tuple[str, int, str | None], dict[str, Any]] = {}
        for row, (_depth, primary, _label, _abstract) in zip(rows, line_items):
            for offset, member in enumerate(_MFRS_SOCIE_COMPONENTS, start=2):
                col = openpyxl.utils.get_column_letter(offset)
                out[(ws.title, row, col)] = _address(primary, {axis: member})
        return out
    finally:
        wb.close()


@lru_cache(maxsize=64)
def semantic_addresses_for(path_value: str) -> dict[tuple[str, int, str | None], dict[str, Any]]:
    """Return semantic addresses keyed by ``(sheet, row, matrix_col)``.

    ``matrix_col`` is ``None`` for linear concepts.  The returned dictionaries
    are safe to serialize into the concept-tree JSON.
    """
    path = Path(path_value).resolve()
    family = _standard_and_level(path)
    if family is None or path.name not in _ROLES_BY_FILE:
        return {}
    standard, level = family
    if path.name == "09-SOCIE.xlsx":
        return _matrix_addresses(path, standard, level)
    return _linear_addresses(path, standard)


__all__ = [
    "ADDRESS_VERSION",
    "TAXONOMY_VERSION",
    "semantic_addresses_for",
    "taxonomy_concept",
    "taxonomy_registry",
]
