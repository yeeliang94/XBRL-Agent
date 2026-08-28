"""Canonical filing-target registry and all-template coverage audit.

This module is the seam shared by extraction, review, canonical persistence,
and mTool filing.  It separates taxonomy capability from workbook slot role;
callers consume the resulting writable-target decision instead of recreating
label, style, or row-position rules.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass
from functools import lru_cache
from hashlib import sha256
from pathlib import Path
import json
import sqlite3
from typing import Any

import openpyxl

from concept_model.notes_parser import parse_notes_template
from concept_model.parser import ConceptNode, _derive_template_id, parse_template
from concept_model.taxonomy_semantics import taxonomy_concept


MANIFEST_VERSION = "2022-v1-slot-semantics-1"
WRITABLE_SLOT_ROLES = frozenset({"INPUT", "MATRIX_INPUT"})
SLOT_ROLES = frozenset({
    "PRESENTATION_ONLY",
    "INPUT",
    "FORMULA",
    "MATRIX_INPUT",
    "MATRIX_FORMULA",
    "PERIOD_METADATA",
    "UNMAPPED",
})

_PROSE_FILENAMES = frozenset({
    "10-Notes-CorporateInfo.xlsx",
    "11-Notes-CorporateInfo.xlsx",
    "11-Notes-AccountingPolicies.xlsx",
    "12-Notes-AccountingPolicies.xlsx",
    "12-Notes-ListOfNotes.xlsx",
    "13-Notes-ListOfNotes.xlsx",
})


@dataclass(frozen=True)
class FilingTarget:
    template_id: str
    sheet: str
    row: int
    col: str | None
    target_id: str
    canonical_target_id: str
    label: str
    slot_role: str
    value_kind: str
    taxonomy_element_id: str | None
    namespace_uri: str | None
    local_name: str | None
    concept_role: str | None
    reportable: bool
    dimensions: dict[str, str]
    mapping_source: str
    exception_code: str | None = None

    @property
    def writable(self) -> bool:
        return self.slot_role in WRITABLE_SLOT_ROLES and self.reportable


def active_template_paths(repository_root: str | Path) -> list[Path]:
    root = Path(repository_root)
    return sorted(
        path
        for standard in ("MFRS", "MPERS")
        for level in ("Company", "Group")
        for path in (root / f"XBRL-template-{standard}" / level).glob("*.xlsx")
    )


def _is_managed_template(path_value: str | Path) -> bool:
    """Whether the file is one of this repository's active SSM templates.

    Synthetic and legacy workbooks intentionally retain the existing
    style/label guards. They have no authoritative taxonomy manifest, so
    trying to parse them as an active variant would turn an unavailable
    contract into a false rejection (or a parser error).
    """
    path = Path(path_value).resolve()
    root = Path(__file__).resolve().parent.parent
    try:
        relative = path.relative_to(root)
    except ValueError:
        return False
    parts = relative.parts
    return (
        len(parts) == 3
        and parts[0] in {"XBRL-template-MFRS", "XBRL-template-MPERS"}
        and parts[1] in {"Company", "Group"}
        and path.suffix.lower() == ".xlsx"
    )


def _workbook_fingerprint(path: Path) -> str:
    return sha256(path.read_bytes()).hexdigest()


def _reviewed_exception(path: Path, node: ConceptNode | None = None) -> str | None:
    is_mfrs = any(part == "XBRL-template-MFRS" for part in path.parts)
    if node is None and path.name == "13-Notes-IssuedCapital.xlsx" and is_mfrs:
        return "MFRS_ISSUED_CAPITAL_WRAPPER_OMITTED"
    if node is None and path.name == "14-Notes-RelatedParty.xlsx" and is_mfrs:
        return "MFRS_RELATED_PARTY_WRAPPER_OMITTED"
    if node is not None and node.render_key.get("semantic_address") is None:
        if path.name == "05-SOCI-BeforeTax.xlsx" and node.render_key.get("row") == 3:
            return "PRESENTATION_TITLE_WITHOUT_TAXONOMY_SLOT"
        if path.name == "09-SOCIE.xlsx" and node.kind == "ABSTRACT":
            return "SOCIE_SECTION_HEADER_WITHOUT_TAXONOMY_SLOT"
    return None


def _numeric_targets(path: Path) -> tuple[str, list[FilingTarget]]:
    tree = parse_template(str(path))
    targets: list[FilingTarget] = []
    for node in tree.concepts:
        rk = node.render_key
        address = rk.get("semantic_address") or {}
        element_id = address.get("primary_concept")
        concept = taxonomy_concept(element_id) if element_id else None
        slot_role = str(rk.get("slot_role") or "UNMAPPED")
        exception = _reviewed_exception(path, node)
        if slot_role == "UNMAPPED" and exception:
            slot_role = "PRESENTATION_ONLY"
        targets.append(FilingTarget(
            template_id=tree.template_id,
            sheet=str(rk.get("sheet") or ""),
            row=int(rk.get("row") or 0),
            col=rk.get("matrix_col") or rk.get("col"),
            target_id=(
                f"{node.concept_uuid}:{rk.get('sheet')}:{rk.get('row')}:"
                f"{rk.get('matrix_col') or rk.get('col')}"
            ),
            canonical_target_id=node.concept_uuid,
            label=node.canonical_label,
            slot_role=slot_role,
            value_kind="numeric",
            taxonomy_element_id=element_id,
            namespace_uri=concept.namespace_uri if concept else None,
            local_name=concept.local_name if concept else None,
            concept_role=concept.concept_role if concept else None,
            reportable=bool(concept and concept.reportable),
            dimensions=dict(address.get("dimensions") or {}),
            mapping_source=(
                "reviewed_exception" if exception and element_id
                else "presentation_linkbase" if element_id
                else "none"
            ),
            exception_code=exception,
        ))
    return tree.template_id, targets


def _prose_targets(path: Path) -> tuple[str, list[FilingTarget]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = wb.sheetnames[0]
    finally:
        wb.close()
    template_id, nodes = parse_notes_template(str(path), sheet)
    targets: list[FilingTarget] = []
    for node in nodes:
        address = node.semantic_address or {}
        concept = (
            taxonomy_concept(node.taxonomy_element_id)
            if node.taxonomy_element_id else None
        )
        targets.append(FilingTarget(
            template_id=template_id,
            sheet=node.sheet,
            row=node.row,
            col="B",
            target_id=node.node_uuid,
            canonical_target_id=node.node_uuid,
            label=node.label,
            slot_role=node.slot_role,
            value_kind="html",
            taxonomy_element_id=node.taxonomy_element_id,
            namespace_uri=concept.namespace_uri if concept else None,
            local_name=concept.local_name if concept else None,
            concept_role=concept.concept_role if concept else None,
            reportable=bool(concept and concept.reportable),
            dimensions=dict(address.get("dimensions") or {}),
            mapping_source=("presentation_linkbase" if concept else "none"),
        ))
    return template_id, targets


@lru_cache(maxsize=128)
def _targets_for_template_cached(
    resolved_path: str, file_size: int, modified_ns: int,
) -> tuple[str, tuple[FilingTarget, ...]]:
    """Parse a static template once per on-disk revision.

    Size and nanosecond mtime participate in the cache key so template
    regeneration invalidates the entry without introducing mutable global
    state or a manual cache-reset protocol.
    """
    del file_size, modified_ns  # cache-key material; parsing only needs the path
    path = Path(resolved_path)
    if path.name in _PROSE_FILENAMES:
        template_id, targets = _prose_targets(path)
    else:
        template_id, targets = _numeric_targets(path)
    return template_id, tuple(targets)


def targets_for_template(path_value: str | Path) -> tuple[str, list[FilingTarget]]:
    path = Path(path_value).resolve()
    stat = path.stat()
    template_id, targets = _targets_for_template_cached(
        str(path), stat.st_size, stat.st_mtime_ns,
    )
    return template_id, list(targets)


def list_writable_targets(path_value: str | Path) -> list[FilingTarget]:
    if not _is_managed_template(path_value):
        return []
    return [target for target in targets_for_template(path_value)[1] if target.writable]


def writable_rows(
    path_value: str | Path, sheet: str,
) -> frozenset[int] | None:
    if not _is_managed_template(path_value):
        return None
    return frozenset(
        target.row for target in list_writable_targets(path_value)
        if target.sheet == sheet
    )


def resolve_writable_html_target(
    conn: sqlite3.Connection,
    *,
    family_prefix: str | None = None,
    template_id: str | None = None,
    sheet: str,
    row: int,
) -> dict[str, Any] | None:
    """Resolve one writable HTML field through the canonical slot manifest.

    Most HTML note fields live in ``notes_nodes``. Numeric notes are different:
    their figures live in ``concept_nodes``, while each sheet also has one
    taxonomy text-block field for the reproduced disclosure table. The v41
    slot manifest is the shared contract that represents both tracks, so HTML
    persistence and editing must resolve through it rather than assuming every
    HTML field belongs to the prose registry.

    Callers resolving a run-scoped write pass ``family_prefix``; callers that
    already selected one template may pass ``template_id``. Exactly one
    selector is required. Ambiguity fails closed: the selected scope must
    contain exactly one matching slot for a physical coordinate before callers
    may write it.
    """
    if (family_prefix is None) == (template_id is None):
        raise ValueError("Pass exactly one of family_prefix or template_id.")
    template_predicate = (
        "ts.template_id = ?" if template_id is not None
        else "ts.template_id LIKE ?"
    )
    if template_id is not None:
        template_selector = template_id
    else:
        assert family_prefix is not None
        template_selector = family_prefix + "%"
    matches = conn.execute(
        f"""
        SELECT ts.template_id,
               ts.canonical_target_id AS concept_uuid,
               ts.label
        FROM template_slots ts
        LEFT JOIN taxonomy_concepts tc
          ON tc.source_element_id = ts.taxonomy_element_id
        WHERE {template_predicate}
          AND ts.sheet = ? AND ts.row = ? AND ts.col = 'B'
          AND ts.validation_status = 'writable'
          AND ts.slot_role = 'INPUT'
          AND (
            ts.value_kind = 'html'
            OR LOWER(COALESCE(tc.data_type, '')) LIKE '%textblockitemtype'
          )
        """,
        (template_selector, sheet, int(row)),
    ).fetchall()
    if len(matches) != 1:
        return None
    match = matches[0]
    return {
        "template_id": match[0],
        "concept_uuid": match[1],
        "label": match[2],
    }


def persist_template_manifest(db_path: str | Path, path_value: str | Path) -> int:
    """Replace one template's v41 taxonomy/slot manifest atomically."""
    path = Path(path_value).resolve()
    fingerprint = _workbook_fingerprint(path)
    template_id = _derive_template_id(path)
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    existing = conn.execute(
        "SELECT COUNT(*), MIN(manifest_version), MAX(manifest_version), "
        "MIN(workbook_fingerprint), MAX(workbook_fingerprint) "
        "FROM template_slots WHERE template_id = ?",
        (template_id,),
    ).fetchone()
    if (
        existing
        and int(existing[0] or 0) > 0
        and existing[1] == existing[2] == MANIFEST_VERSION
        and existing[3] == existing[4] == fingerprint
    ):
        conn.close()
        return int(existing[0])

    try:
        parsed_template_id, targets = targets_for_template(path)
    except Exception:
        conn.close()
        raise
    if parsed_template_id != template_id:
        conn.close()
        raise ValueError(
            f"Template id changed while parsing {path}: "
            f"{template_id!r} != {parsed_template_id!r}"
        )
    linked_ids = {
        element_id
        for target in targets
        for element_id in (
            [target.taxonomy_element_id] if target.taxonomy_element_id else []
        ) + list(target.dimensions.keys()) + list(target.dimensions.values())
    }
    conn.execute("BEGIN")
    try:
        for element_id in sorted(linked_ids):
            concept = taxonomy_concept(element_id)
            if concept is None:
                raise ValueError(
                    f"Template {template_id} references unknown taxonomy element "
                    f"{element_id!r}"
                )
            conn.execute(
                """
                INSERT INTO taxonomy_concepts(
                    source_element_id, taxonomy_version, namespace_uri,
                    local_name, abstract, concept_role, data_type, period_type,
                    balance, substitution_group
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(source_element_id) DO UPDATE SET
                    taxonomy_version = excluded.taxonomy_version,
                    namespace_uri = excluded.namespace_uri,
                    local_name = excluded.local_name,
                    abstract = excluded.abstract,
                    concept_role = excluded.concept_role,
                    data_type = excluded.data_type,
                    period_type = excluded.period_type,
                    balance = excluded.balance,
                    substitution_group = excluded.substitution_group
                """,
                (
                    concept.source_element_id, concept.taxonomy_version,
                    concept.namespace_uri, concept.local_name,
                    int(concept.abstract), concept.concept_role,
                    concept.data_type, concept.period_type, concept.balance,
                    concept.substitution_group,
                ),
            )

        conn.execute("DELETE FROM template_slots WHERE template_id = ?", (template_id,))
        conn.execute(
            "DELETE FROM template_manifest_exceptions WHERE template_id = ?",
            (template_id,),
        )
        for target in targets:
            if target.writable:
                validation = "writable"
            elif target.taxonomy_element_id or target.exception_code:
                validation = "non_writable"
            else:
                validation = "unresolved"
            conn.execute(
                """
                INSERT INTO template_slots(
                    target_id, canonical_target_id, template_id, sheet, row, col, label, slot_role,
                    value_kind, taxonomy_element_id, dimensions_json,
                    mapping_source, manifest_version, workbook_fingerprint,
                    validation_status, exception_code
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    target.target_id, target.canonical_target_id,
                    target.template_id, target.sheet,
                    target.row, target.col or "", target.label, target.slot_role,
                    target.value_kind, target.taxonomy_element_id,
                    json.dumps(target.dimensions, sort_keys=True, separators=(",", ":")),
                    target.mapping_source, MANIFEST_VERSION, fingerprint,
                    validation, target.exception_code,
                ),
            )

        template_exception = _reviewed_exception(path)
        if template_exception:
            conn.execute(
                """
                INSERT INTO template_manifest_exceptions(
                    template_id, exception_code, manifest_version,
                    workbook_fingerprint
                ) VALUES (?, ?, ?, ?)
                """,
                (template_id, template_exception, MANIFEST_VERSION, fingerprint),
            )

        standard, level = template_id.split("-", 2)[:2]
        if path.name in _PROSE_FILENAMES:
            # Existing databases predate template-scoped prose identities.
            # Upgrade valid rows deterministically by the run's exact family
            # and physical slot; never infer identity from the stored label.
            conn.execute(
                """
                UPDATE notes_cells AS c
                SET concept_uuid = (
                      SELECT nn.node_uuid FROM notes_nodes nn
                      WHERE nn.template_id = ?
                        AND nn.sheet = c.sheet AND nn.row = c.row
                        AND nn.kind = 'LEAF' AND nn.slot_role = 'INPUT'
                    ),
                    invalid_target = 0,
                    invalid_target_reason = NULL
                WHERE EXISTS (
                    SELECT 1 FROM runs r
                    WHERE r.id = c.run_id
                      AND LOWER(COALESCE(
                        json_extract(r.run_config_json, '$.filing_standard'),
                        'mfrs'
                      )) = ?
                      AND LOWER(COALESCE(
                        json_extract(r.run_config_json, '$.filing_level'),
                        'company'
                      )) = ?
                )
                  AND EXISTS (
                    SELECT 1 FROM notes_nodes nn
                    WHERE nn.template_id = ?
                      AND nn.sheet = c.sheet AND nn.row = c.row
                      AND nn.kind = 'LEAF' AND nn.slot_role = 'INPUT'
                  )
                """,
                (template_id, standard, level, template_id),
            )

        # Quarantine historical numeric facts on structural/presentation slots.
        # Formula-owned totals remain valid canonical facts; only slots whose
        # role says presentation/unmapped are invalid targets.
        conn.execute(
            """
            UPDATE run_concept_facts
            SET invalid_target = CASE WHEN concept_uuid IN (
                    SELECT canonical_target_id FROM template_slots
                    WHERE template_id = ?
                      AND slot_role IN ('PRESENTATION_ONLY', 'UNMAPPED')
                ) THEN 1 ELSE 0 END,
                invalid_target_reason = CASE WHEN concept_uuid IN (
                    SELECT canonical_target_id FROM template_slots
                    WHERE template_id = ?
                      AND slot_role IN ('PRESENTATION_ONLY', 'UNMAPPED')
                ) THEN 'The stored value targets a presentation-only template row.'
                  ELSE NULL END
            WHERE concept_uuid IN (
                SELECT concept_uuid FROM concept_nodes WHERE template_id = ?
            )
            """,
            (template_id, template_id, template_id),
        )
        conn.execute(
            """
            UPDATE notes_cells
            SET invalid_target = 1,
                invalid_target_reason =
                  'The stored note targets a heading or other non-entry row.'
            WHERE EXISTS (
                SELECT 1
                FROM notes_nodes nn
                JOIN template_slots ts
                  ON ts.canonical_target_id = nn.node_uuid
                WHERE nn.template_id = ?
                  AND nn.sheet = notes_cells.sheet
                  AND nn.row = notes_cells.row
                  AND ts.slot_role IN ('PRESENTATION_ONLY', 'UNMAPPED')
            )
              AND EXISTS (
                SELECT 1 FROM runs r
                WHERE r.id = notes_cells.run_id
                  AND LOWER(COALESCE(
                    json_extract(r.run_config_json, '$.filing_standard'),
                    'mfrs'
                  )) = ?
                  AND LOWER(COALESCE(
                    json_extract(r.run_config_json, '$.filing_level'),
                    'company'
                  )) = ?
              )
            """,
            (template_id, standard, level),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return len(targets)


def semantic_coverage_for_run(
    db_path: str | Path,
    run_id: int,
    *,
    filing_standard: str,
    filing_level: str,
) -> dict[str, Any]:
    """Return the auditable field-semantics contract for one filing run.

    A run can predate v41 or be only partly bootstrapped. Such a run is
    reported as ``needs_review``; an empty manifest is never described as
    complete. Invalid historical values remain visible as quarantined counts.
    """
    family_prefix = f"{filing_standard.lower()}-{filing_level.lower()}-"
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        selected = {
            row[0]
            for row in conn.execute(
                """
                SELECT DISTINCT n.template_id
                FROM run_concept_facts f
                JOIN concept_nodes n ON n.concept_uuid = f.concept_uuid
                WHERE f.run_id = ? AND n.template_id LIKE ?
                UNION
                SELECT DISTINCT nn.template_id
                FROM notes_cells c
                JOIN notes_nodes nn ON nn.node_uuid = c.concept_uuid
                WHERE c.run_id = ? AND nn.template_id LIKE ?
                """,
                (run_id, family_prefix + "%", run_id, family_prefix + "%"),
            ).fetchall()
        }
        slot_rows = conn.execute(
            """
            SELECT template_id, COUNT(*) AS total,
                   SUM(CASE WHEN validation_status = 'writable' THEN 1 ELSE 0 END)
                     AS writable,
                   SUM(CASE WHEN validation_status = 'unresolved' THEN 1 ELSE 0 END)
                     AS unresolved,
                   MAX(manifest_version) AS manifest_version
            FROM template_slots
            WHERE template_id LIKE ?
            GROUP BY template_id ORDER BY template_id
            """,
            (family_prefix + "%",),
        ).fetchall()
        manifests = {row["template_id"]: row for row in slot_rows}
        taxonomy_versions = [
            str(row[0]) for row in conn.execute(
                """
                SELECT DISTINCT tc.taxonomy_version
                FROM template_slots ts
                JOIN taxonomy_concepts tc
                  ON tc.source_element_id = ts.taxonomy_element_id
                WHERE ts.template_id LIKE ?
                ORDER BY tc.taxonomy_version
                """,
                (family_prefix + "%",),
            ).fetchall()
            if row[0]
        ]
        missing = sorted(template_id for template_id in selected if template_id not in manifests)
        invalid_facts = conn.execute(
            """
            SELECT COUNT(*) FROM run_concept_facts f
            WHERE f.run_id = ? AND (
              f.invalid_target = 1 OR (
                EXISTS (SELECT 1 FROM template_slots ts
                        WHERE ts.canonical_target_id = f.concept_uuid)
                AND NOT EXISTS (
                  SELECT 1 FROM template_slots ts
                  WHERE ts.canonical_target_id = f.concept_uuid
                    AND ts.slot_role IN (
                      'INPUT', 'MATRIX_INPUT', 'FORMULA', 'MATRIX_FORMULA'
                    )
                )
              )
            )
            """,
            (run_id,),
        ).fetchone()[0]
        invalid_notes = conn.execute(
            """
            SELECT COUNT(*) FROM notes_cells c
            WHERE c.run_id = ? AND (
              c.invalid_target = 1 OR (
                EXISTS (
                  SELECT 1 FROM notes_nodes nn
                  JOIN template_slots ts
                    ON ts.canonical_target_id = nn.node_uuid
                  WHERE nn.template_id LIKE ?
                    AND nn.sheet = c.sheet AND nn.row = c.row
                )
                AND NOT EXISTS (
                  SELECT 1 FROM notes_nodes nn
                  JOIN template_slots ts
                    ON ts.canonical_target_id = nn.node_uuid
                  WHERE nn.template_id LIKE ?
                    AND nn.node_uuid = c.concept_uuid
                    AND nn.sheet = c.sheet AND nn.row = c.row
                    AND ts.validation_status = 'writable'
                    AND ts.slot_role = 'INPUT'
                )
              )
            )
            """,
            (run_id, family_prefix + "%", family_prefix + "%"),
        ).fetchone()[0]
        slot_exceptions = [dict(row) for row in conn.execute(
            """
            SELECT exception_code, COUNT(*) AS count
            FROM template_slots
            WHERE template_id LIKE ? AND exception_code IS NOT NULL
            GROUP BY exception_code ORDER BY exception_code
            """,
            (family_prefix + "%",),
        ).fetchall()]
        manifest_exceptions = [dict(row) for row in conn.execute(
            """
            SELECT exception_code, COUNT(*) AS count
            FROM template_manifest_exceptions
            WHERE template_id LIKE ?
            GROUP BY exception_code ORDER BY exception_code
            """,
            (family_prefix + "%",),
        ).fetchall()]
    finally:
        conn.close()

    unresolved = sum(int(row["unresolved"] or 0) for row in slot_rows)
    blockers: list[dict[str, Any]] = []
    if not slot_rows:
        blockers.append({
            "code": "template_catalog_missing",
            "count": 1,
            "message": (
                "The filing-field catalog for this filing family is missing. "
                "Rebuild the template catalog before filing."
            ),
            "examples": [],
        })
    if missing:
        blockers.append({
            "code": "template_manifest_missing",
            "count": len(missing),
            "message": (
                "The filing fields for part of this run have not been matched "
                "to the selected template. Rebuild the run's template catalog "
                "before filing."
            ),
            "examples": missing[:8],
        })
    if unresolved:
        blockers.append({
            "code": "template_fields_unresolved",
            "count": unresolved,
            "message": (
                "Some template rows have no confirmed field identity. They "
                "must be mapped or accepted as a reviewed exception."
            ),
            "examples": [],
        })
    quarantined = int(invalid_facts) + int(invalid_notes)
    if quarantined:
        blockers.append({
            "code": "invalid_targets_quarantined",
            "count": quarantined,
            "message": (
                "Stored values are not linked to writable filing fields. "
                "Move or remove them before filing."
            ),
            "examples": [],
        })

    exception_counts: dict[str, int] = {}
    for item in [*slot_exceptions, *manifest_exceptions]:
        code = str(item["exception_code"])
        exception_counts[code] = exception_counts.get(code, 0) + int(item["count"])

    return {
        "run_id": run_id,
        "filing_standard": filing_standard.lower(),
        "filing_level": filing_level.lower(),
        "readiness": "ready" if not blockers else "needs_review",
        "selected_templates": sorted(selected),
        "counts": {
            "catalog_templates": len(slot_rows),
            "selected_templates": len(selected),
            "template_slots": sum(int(row["total"] or 0) for row in slot_rows),
            "writable_fields": sum(int(row["writable"] or 0) for row in slot_rows),
            "unresolved_fields": unresolved,
            "quarantined_values": quarantined,
        },
        "manifest_versions": sorted({
            str(row["manifest_version"]) for row in slot_rows
            if row["manifest_version"]
        }),
        "taxonomy_versions": taxonomy_versions,
        "reviewed_exceptions": [
            {"exception_code": code, "count": count}
            for code, count in sorted(exception_counts.items())
        ],
        "blockers": blockers,
    }


def audit_active_templates(repository_root: str | Path) -> dict[str, Any]:
    repository_root = Path(repository_root).resolve()
    results: list[dict[str, Any]] = []
    missing_required: list[dict[str, Any]] = []
    structural_writable: list[dict[str, Any]] = []
    reviewed: list[dict[str, Any]] = []
    numeric_slots = 0
    prose_slots = 0
    worksheets = 0
    unclassified = 0

    paths = active_template_paths(repository_root)
    for path in paths:
        relative_path = path.resolve().relative_to(repository_root).as_posix()
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        try:
            sheet_count = len(wb.sheetnames)
        finally:
            wb.close()
        worksheets += sheet_count
        template_id, targets = targets_for_template(path)
        is_prose = path.name in _PROSE_FILENAMES
        if is_prose:
            prose_slots += len(targets)
        else:
            numeric_slots += len(targets)
        writable = [target for target in targets if target.writable]
        mapped_writable = [
            target for target in writable if target.taxonomy_element_id
        ]
        for target in targets:
            item = {"path": relative_path, **asdict(target)}
            if target.slot_role not in SLOT_ROLES:
                unclassified += 1
            if target.slot_role in WRITABLE_SLOT_ROLES and not target.writable:
                missing_required.append(item)
            if target.writable and target.concept_role != "PRIMARY_ITEM":
                structural_writable.append(item)
            if target.exception_code:
                reviewed.append({
                    "code": target.exception_code,
                    "path": relative_path,
                    "sheet": target.sheet,
                    "row": target.row,
                })
        file_exception = _reviewed_exception(path)
        if file_exception:
            reviewed.append({"code": file_exception, "path": relative_path})
        results.append({
            "template_id": template_id,
            "path": relative_path,
            "filename": path.name,
            "fingerprint": _workbook_fingerprint(path),
            "worksheets": sheet_count,
            "slots": len(targets),
            "writable_slots": len(writable),
            "mapped_writable_slots": len(mapped_writable),
            "manifest_version": MANIFEST_VERSION,
            "semantic_coverage": (
                "complete" if len(writable) == len(mapped_writable) else "incomplete"
            ),
        })

    return {
        "manifest_version": MANIFEST_VERSION,
        "templates": len(paths),
        "worksheets": worksheets,
        "numeric_slots": numeric_slots,
        "prose_slots": prose_slots,
        "unclassified_slots": unclassified,
        "missing_required_mappings": missing_required,
        "structural_writable_slots": structural_writable,
        "reviewed_exceptions": reviewed,
        "template_results": results,
    }


__all__ = [
    "FilingTarget",
    "MANIFEST_VERSION",
    "WRITABLE_SLOT_ROLES",
    "active_template_paths",
    "audit_active_templates",
    "list_writable_targets",
    "persist_template_manifest",
    "semantic_coverage_for_run",
    "targets_for_template",
    "writable_rows",
]
