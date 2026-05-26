"""Phase 0 — XBRL template parser.

Walks a template ``.xlsx`` and emits a ``ConceptTree`` of typed nodes
(ABSTRACT section-headers, LEAF data-entry rows, COMPUTED formula rows)
plus the dependency edges between them.  No DB writes; pure read.

The parser is deliberately layered on existing primitives:

* abstract-row detection delegates to ``tools.section_headers`` so the
  writer's header-guard (gotcha #17) and the parser stay symmetric;
* indentation is read off ``cell.alignment.indent`` because that is the
  same signal ``section_headers.py`` relies on for hierarchy hints.

Formula-grammar coverage is intentionally narrow.  Today we recognise:

* signed sums      e.g. ``=1*B10+1*B11-1*B12`` and bare adds ``=B10+B11``;
* cross-sheet refs e.g. ``='SOFP-Sub-CuNonCu'!B39``;
* ``SUM`` ranges   e.g. ``=SUM(B10:B14)`` (single column/row only).

Anything else raises ``UnknownFormulaShape`` so callers can collect the
list of surprises rather than silently skip them.  SOCIE's matrix layout
raises ``UnsupportedSchemaShape`` until Phase 5 lands.
"""
from __future__ import annotations

import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from tools.section_headers import (
    discover_section_headers,
    keyword_fallback_for_sheet,
)


# UUID namespace for concept IDs.  Stable across runs so the same
# template always emits the same UUIDs — Phase 1's importer relies on
# this for idempotent upserts.
_CONCEPT_NS = uuid.UUID("8c2dc94e-1d2a-4d3f-9c1e-b6f0e8a3a8e0")


# ---------------------------------------------------------------------------
# Public exceptions
# ---------------------------------------------------------------------------


class UnknownFormulaShape(Exception):
    """Formula grammar we haven't taught the parser yet.

    Carries the raw formula + cell coordinate so the caller can log a
    catalogue of surprise shapes instead of failing the whole template.
    """

    def __init__(self, sheet: str, coordinate: str, formula: str) -> None:
        super().__init__(f"{sheet}!{coordinate}: {formula!r}")
        self.sheet = sheet
        self.coordinate = coordinate
        self.formula = formula


class UnsupportedSchemaShape(Exception):
    """Whole-template shape the linear concept tree can't represent.

    SOCIE is the only known case in Phase 0 — its (row, col) matrix
    geometry needs a different kind enum (MATRIX_CELL) that arrives in
    Phase 5.
    """


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------


@dataclass
class ConceptNode:
    """One row of the template.

    ``concept_uuid`` is deterministic (UUID5) so re-running the parser
    yields the same identity every time — important for idempotent DB
    upserts in Phase 1 and for stable cross-template references.
    """

    concept_uuid: str
    parent_uuid: str | None
    kind: str              # "ABSTRACT" | "LEAF" | "COMPUTED"
    canonical_label: str   # column-A text, trimmed; leading '*' kept
    render_key: dict[str, Any]
    edges: list[dict[str, Any]] = field(default_factory=list)
    # Indentation level (cell.alignment.indent) — used only during tree
    # assembly, not serialised.  Kept on the dataclass for debugging.
    _indent: int = 0


@dataclass
class ConceptTree:
    template_id: str
    concepts: list[ConceptNode]
    # "linear" for the face statements (LEAF/COMPUTED rows); "matrix" for
    # SOCIE, where concept identity is (row, equity-component-column).
    shape: str = "linear"

    def to_json(self) -> dict[str, Any]:
        return {
            "template_id": self.template_id,
            "shape": self.shape,
            "concepts": [
                {
                    "concept_uuid": n.concept_uuid,
                    "parent_uuid": n.parent_uuid,
                    "kind": n.kind,
                    "canonical_label": n.canonical_label,
                    "render_key": n.render_key,
                    "edges": n.edges,
                }
                for n in self.concepts
            ],
        }


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def parse_template(xlsx_path: str) -> ConceptTree:
    """Parse a template into a ``ConceptTree``.

    SOCIE templates take the matrix branch (Phase 5): concept identity is
    (row, equity-component-column), emitted as MATRIX_CELL concepts.  Every
    other template uses the linear LEAF/COMPUTED model.
    """
    path = Path(xlsx_path)
    template_id = _derive_template_id(path)

    # SOCIE is a (row, col) matrix and doesn't fit the linear LEAF/
    # COMPUTED model.  Dispatch to the matrix parser instead.
    if "socie" in path.stem.lower():
        return _parse_socie_matrix(path, template_id)

    wb = openpyxl.load_workbook(str(path), data_only=False)
    nodes: list[ConceptNode] = []
    # First pass: per-sheet, classify every column-A row and collect any
    # formula edges.  Cross-sheet references are stashed for the second
    # pass once every node has a UUID we can point at.
    coord_to_uuid: dict[tuple[str, int, str], str] = {}
    pending_cross_refs: list[tuple[ConceptNode, str, str]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        _classify_sheet(
            ws,
            sheet_name,
            template_id,
            nodes,
            coord_to_uuid,
            pending_cross_refs,
        )

    # Second pass: resolve cross-sheet references.  These do NOT add new
    # concepts — the face cell shares the sub cell's identity but keeps
    # its own render_key so the exporter can still find the right
    # coordinate when writing back to xlsx.
    _resolve_cross_refs(nodes, coord_to_uuid, pending_cross_refs)

    return ConceptTree(template_id=template_id, concepts=nodes)


# ---------------------------------------------------------------------------
# SOCIE matrix parser (Phase 5)
# ---------------------------------------------------------------------------


# A SOCIE block is a self-contained statement-of-changes-in-equity, opened
# by an "Equity at beginning of period" row and closed by "Equity at end of
# period".  Group filings stack four such blocks vertically (gotcha #12);
# MFRS Company stacks two (CY then PY); MPERS Company has a single block
# where the period is a column (B=CY, C=PY) instead of a stacked block.
_SOCIE_BLOCK_OPEN = "equity at beginning of period"
_SOCIE_BLOCK_CLOSE = "equity at end of period"


def _socie_norm(value: Any) -> str:
    return str(value or "").strip().lstrip("*").strip().lower()


def _find_socie_blocks(ws: Worksheet) -> list[tuple[int, int]]:
    """Return ordered (begin_row, end_row) pairs, one per stacked block."""
    begins = [
        r for r in range(1, ws.max_row + 1)
        if _socie_norm(ws.cell(r, 1).value) == _SOCIE_BLOCK_OPEN
    ]
    ends = [
        r for r in range(1, ws.max_row + 1)
        if _socie_norm(ws.cell(r, 1).value) == _SOCIE_BLOCK_CLOSE
    ]
    return list(zip(begins, ends))


def _socie_component_cols(ws: Worksheet) -> list[str]:
    """Equity-component column letters from the row-2 header.

    MFRS SOCIE carries 23 component headers in B..X; MPERS SOCIE has no
    row-2 headers and uses a single value column (B).
    """
    cols = [
        get_column_letter(c)
        for c in range(2, ws.max_column + 1)
        if ws.cell(2, c).value not in (None, "")
    ]
    return cols or ["B"]


def _socie_block_dims(n_blocks: int, idx: int) -> tuple[str, str] | None:
    """Map a stacked-block index to its (period, entity_scope).

    Returns None for the single-block MPERS-Company case, where the period
    is a column rather than a stacked block.
    """
    if n_blocks == 4:
        return [
            ("CY", "Group"), ("PY", "Group"),
            ("CY", "Company"), ("PY", "Company"),
        ][idx]
    if n_blocks == 2:
        return [("CY", "Company"), ("PY", "Company")][idx]
    return None


def _parse_socie_matrix(path: Path, template_id: str) -> ConceptTree:
    """Parse a SOCIE template into a matrix-shaped ``ConceptTree``.

    Canonical concepts live in the FIRST block; each carries `matrix_col`
    (the equity-component column) and a `targets` list mapping every
    (period, entity_scope) dimension to its physical (sheet, row, col).
    The exporter and importer route facts back through those targets — the
    same separation Phase 4 uses for Group columns, except here the
    period/scope dimension shifts the ROW (stacked blocks) rather than the
    column.
    """
    wb = openpyxl.load_workbook(str(path), data_only=False)
    sheet_name = "SOCIE" if "SOCIE" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    blocks = _find_socie_blocks(ws)
    if not blocks:
        raise UnsupportedSchemaShape(
            f"SOCIE template has no recognisable equity blocks ({path.name})"
        )
    components = _socie_component_cols(ws)
    base_begin, base_end = blocks[0]
    n_blocks = len(blocks)
    single_block = n_blocks == 1

    fallback = keyword_fallback_for_sheet(sheet_name)
    abstract_rows = {
        h.row for h in discover_section_headers(ws, extra_keywords=fallback)
    }

    nodes: list[ConceptNode] = []
    pending_cross_refs: list[tuple[ConceptNode, str, str]] = []

    for row in range(base_begin, base_end + 1):
        raw = ws.cell(row, 1).value
        if raw is None or not str(raw).strip():
            continue
        label = str(raw).strip()

        # Grouping sub-headers inside a block ("Changes in equity",
        # "Comprehensive income", …) carry no values — one ABSTRACT
        # concept, no per-column cells.
        if row in abstract_rows:
            uid = _mint_uuid(template_id, sheet_name, row, label)
            nodes.append(ConceptNode(
                concept_uuid=uid,
                parent_uuid=None,
                kind="ABSTRACT",
                canonical_label=label,
                render_key={"sheet": sheet_name, "row": row, "col": "A",
                            "matrix_col": None, "targets": []},
            ))
            continue

        for col in components:
            targets = _socie_targets_for(
                sheet_name, row, col, blocks, base_begin,
                single_block,
            )
            uid = _mint_uuid(
                template_id, sheet_name, row, f"{col}::{label}"
            )
            node = ConceptNode(
                concept_uuid=uid,
                parent_uuid=None,
                kind="MATRIX_CELL",
                canonical_label=label,
                render_key={
                    "sheet": sheet_name,
                    "row": row,
                    "col": col,
                    "matrix_col": col,
                    "targets": targets,
                },
            )
            # Capture any in-cell formula as dependency edges.  SOCIE
            # formulas are within-column signed sums (e.g. =B6+B7); a
            # shape we don't recognise is skipped rather than fatal so a
            # surprise template doesn't break the whole matrix.
            formula = ws.cell(row, get_col_index(col)).value
            if isinstance(formula, str) and formula.startswith("="):
                try:
                    _collect_edges(node, sheet_name, formula, pending_cross_refs)
                except UnknownFormulaShape:
                    pass
            nodes.append(node)

    return ConceptTree(
        template_id=template_id, concepts=nodes, shape="matrix"
    )


def _socie_targets_for(
    sheet_name: str,
    base_row: int,
    col: str,
    blocks: list[tuple[int, int]],
    base_begin: int,
    single_block: bool,
) -> list[dict[str, Any]]:
    """Build the per-(period, entity_scope) render targets for one cell."""
    if single_block:
        # MPERS Company: period is a column (B=CY, C=PY) at the same row.
        return [
            {"period": "CY", "entity_scope": "Company",
             "sheet": sheet_name, "row": base_row, "col": "B"},
            {"period": "PY", "entity_scope": "Company",
             "sheet": sheet_name, "row": base_row, "col": "C"},
        ]
    targets: list[dict[str, Any]] = []
    n_blocks = len(blocks)
    for idx, (b_begin, _b_end) in enumerate(blocks):
        dims = _socie_block_dims(n_blocks, idx)
        if dims is None:
            continue
        period, scope = dims
        targets.append({
            "period": period,
            "entity_scope": scope,
            "sheet": sheet_name,
            "row": base_row + (b_begin - base_begin),
            "col": col,
        })
    return targets


def get_col_index(letter: str) -> int:
    from openpyxl.utils.cell import column_index_from_string
    return column_index_from_string(letter)


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------


def _derive_template_id(path: Path) -> str:
    """Stable, human-readable id derived from the on-disk path.

    Format: ``{standard}-{level}-{slug}-v1`` so two templates with the
    same filename in different folders (Company vs Group, MFRS vs
    MPERS) yield different ids.
    """
    parts = list(path.parts)
    standard = "mfrs"
    level = "company"
    for part in parts:
        low = part.lower()
        if "mpers" in low:
            standard = "mpers"
        if low == "group":
            level = "group"
        elif low == "company":
            level = "company"
    slug = re.sub(r"^\d+-", "", path.stem).lower()
    return f"{standard}-{level}-{slug}-v1"


def _classify_sheet(
    ws: Worksheet,
    sheet_name: str,
    template_id: str,
    out_nodes: list[ConceptNode],
    coord_to_uuid: dict[tuple[str, int, str], str],
    pending_cross_refs: list[tuple[ConceptNode, str, str]],
) -> None:
    """Walk one worksheet top-to-bottom; emit ConceptNode rows.

    Header rows become ABSTRACT concepts; rows whose first non-A cell
    starts with ``=`` become COMPUTED; everything else with a label
    becomes a LEAF.
    """
    fallback = keyword_fallback_for_sheet(sheet_name)
    abstract_rows = {
        h.row for h in discover_section_headers(ws, extra_keywords=fallback)
    }

    # Parent tracking: each indent level maps to the most-recently-seen
    # ABSTRACT concept at that level.  Leaves and computed rows attach
    # to the deepest ABSTRACT above them.
    parent_by_indent: dict[int, str] = {}

    for row_num in range(1, ws.max_row + 1):
        label_cell = ws.cell(row=row_num, column=1)
        raw = label_cell.value
        if raw is None:
            continue
        label = str(raw).strip()
        if not label:
            continue

        kind = _classify_row(ws, row_num, label_cell, abstract_rows)
        if kind is None:
            # Row has text but no recognised role (e.g. a meta header
            # cell at the top of the sheet) — skip rather than emit a
            # ghost concept.
            continue

        indent = int(getattr(label_cell.alignment, "indent", 0) or 0)
        # Pick parent = the deepest ABSTRACT whose indent is strictly
        # less than ours.  For ABSTRACTs at indent 0 there is no parent.
        parent_uuid: str | None = None
        for lvl in sorted(parent_by_indent.keys(), reverse=True):
            if lvl < indent or kind != "ABSTRACT":
                if lvl <= indent and not (kind == "ABSTRACT" and lvl == indent):
                    parent_uuid = parent_by_indent[lvl]
                    break

        # Default value-column for the render key.  Company templates
        # use col B for CY; Group templates still anchor the concept's
        # identity on the row + first value column.  Phase 4 will add
        # the full per-scope target table.
        render_col = "B"

        node_uuid = _mint_uuid(template_id, sheet_name, row_num, label)
        node = ConceptNode(
            concept_uuid=node_uuid,
            parent_uuid=parent_uuid,
            kind=kind,
            canonical_label=label,
            render_key={
                "sheet": sheet_name,
                "row": row_num,
                "col": render_col,
            },
            _indent=indent,
        )
        out_nodes.append(node)
        coord_to_uuid[(sheet_name, row_num, render_col)] = node_uuid

        # Maintain the indent-stack: when a new ABSTRACT row appears,
        # everything at >= its indent goes out of scope.
        if kind == "ABSTRACT":
            for lvl in list(parent_by_indent.keys()):
                if lvl >= indent:
                    del parent_by_indent[lvl]
            parent_by_indent[indent] = node_uuid

        # Collect formula edges (COMPUTED rows only).  Cross-sheet refs
        # are deferred to the second pass.
        if kind == "COMPUTED":
            formula = _first_formula_in_row(ws, row_num)
            if formula:
                _collect_edges(node, sheet_name, formula, pending_cross_refs)


def _classify_row(
    ws: Worksheet,
    row: int,
    label_cell,
    abstract_rows: set[int],
) -> str | None:
    """Classify a single row as ABSTRACT, LEAF, or COMPUTED.

    ABSTRACT comes from the shared section-headers detector so the
    writer guard and the parser agree.  Everything else inspects the
    first non-A cell: starts with '=' → COMPUTED; non-empty literal →
    (rare) treat as data; empty → LEAF.
    """
    if row in abstract_rows:
        return "ABSTRACT"

    # Scan the value columns (B onward) for a formula.  We stop at the
    # first one found because every template we've seen either has the
    # same formula across all value columns or none at all.
    for col in range(2, max(3, ws.max_column + 1)):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if isinstance(val, str) and val.startswith("="):
            return "COMPUTED"
    return "LEAF"


def _first_formula_in_row(ws: Worksheet, row: int) -> str | None:
    """Return the first formula found in the row (cols B..max), or None."""
    for col in range(2, max(3, ws.max_column + 1)):
        val = ws.cell(row=row, column=col).value
        if isinstance(val, str) and val.startswith("="):
            return val
    return None


# -- formula grammar --------------------------------------------------


_SIGNED_SUM_TERM_RE = re.compile(
    r"""
    (?P<sign>[+-])?              # optional leading sign
    \s*
    (?:(?P<coef>\d+(?:\.\d+)?)\s*\*\s*)?   # optional coefficient
    (?P<col>[A-Z]+)(?P<row>\d+)             # cell ref
    """,
    re.VERBOSE,
)

_CROSS_SHEET_RE = re.compile(
    r"""^=\s*'(?P<sheet>[^']+)'\s*!\s*(?P<col>[A-Z]+)(?P<row>\d+)\s*$""",
    re.VERBOSE,
)

_SUM_RANGE_RE = re.compile(
    r"""^=\s*SUM\s*\(
        \s*(?P<col1>[A-Z]+)(?P<row1>\d+)
        \s*:\s*
        (?P<col2>[A-Z]+)(?P<row2>\d+)\s*
    \)\s*$""",
    re.VERBOSE | re.IGNORECASE,
)


def _collect_edges(
    node: ConceptNode,
    sheet_name: str,
    formula: str,
    pending_cross_refs: list[tuple[ConceptNode, str, str]],
) -> None:
    """Parse one formula and attach the dependency edges to ``node``.

    Side-effect on ``pending_cross_refs`` for the cross-sheet case so
    the caller can resolve them once every node has a UUID.
    """
    body = formula.strip()

    # 1. cross-sheet reference: face cell points to a sub-sheet cell.
    m = _CROSS_SHEET_RE.match(body)
    if m:
        pending_cross_refs.append(
            (node, m.group("sheet"), f"{m.group('col')}{m.group('row')}")
        )
        return

    # 2. SUM range — expand to one +1 edge per cell in the range.
    m = _SUM_RANGE_RE.match(body)
    if m:
        c1, r1 = m.group("col1"), int(m.group("row1"))
        c2, r2 = m.group("col2"), int(m.group("row2"))
        if c1 != c2 and r1 != r2:
            # Multi-cell rectangle — not seen in our templates; refuse
            # rather than silently fan out a 2-D grid.
            raise UnknownFormulaShape(
                sheet_name, node.render_key.get("coord", "?"), formula
            )
        if c1 == c2:
            for r in range(r1, r2 + 1):
                node.edges.append({
                    "ref": {"sheet": sheet_name, "col": c1, "row": r},
                    "coefficient": 1,
                })
        else:
            # row range — rare but symmetric
            from openpyxl.utils.cell import column_index_from_string
            i1 = column_index_from_string(c1)
            i2 = column_index_from_string(c2)
            for ci in range(i1, i2 + 1):
                node.edges.append({
                    "ref": {
                        "sheet": sheet_name,
                        "col": get_column_letter(ci),
                        "row": r1,
                    },
                    "coefficient": 1,
                })
        return

    # 3. signed sum: ``=1*B10+1*B11-1*B12`` or bare ``=B10+B11``.
    if body.startswith("="):
        terms = _parse_signed_sum(body, sheet_name, formula)
        for sign, coef, col, row in terms:
            node.edges.append({
                "ref": {"sheet": sheet_name, "col": col, "row": int(row)},
                "coefficient": sign * coef,
            })
        return

    raise UnknownFormulaShape(sheet_name, "?", formula)


def _parse_signed_sum(
    body: str, sheet_name: str, formula_raw: str
) -> list[tuple[int, float, str, str]]:
    """Tokenise a ``=...`` signed-sum into (sign, coef, col, row) tuples.

    Refuses anything that doesn't fully tokenise — surprise grammar is
    better surfaced than silently truncated.
    """
    expr = body.lstrip("=").replace(" ", "")
    # Collapse sign-on-sign sequences ("+-" / "-+" / "--" / "++") that
    # appear in templates where a negative coefficient is written as a
    # subtraction (``+-1*B42``).  Folding to a single sign keeps the
    # term regex simple.
    while True:
        new = (
            expr.replace("+-", "-")
                .replace("-+", "-")
                .replace("--", "+")
                .replace("++", "+")
        )
        if new == expr:
            break
        expr = new
    # Insert explicit '+' before each unsigned bare-add term so the
    # regex hits every term uniformly.
    if expr and expr[0] not in "+-":
        expr = "+" + expr

    pos = 0
    out: list[tuple[int, float, str, str]] = []
    while pos < len(expr):
        m = _SIGNED_SUM_TERM_RE.match(expr, pos)
        if not m or m.end() == pos:
            raise UnknownFormulaShape(sheet_name, "?", formula_raw)
        sign = -1 if m.group("sign") == "-" else 1
        coef = float(m.group("coef")) if m.group("coef") else 1.0
        out.append((sign, coef, m.group("col"), m.group("row")))
        pos = m.end()
    if not out:
        raise UnknownFormulaShape(sheet_name, "?", formula_raw)
    return out


# -- cross-sheet resolution -------------------------------------------


def _resolve_cross_refs(
    nodes: list[ConceptNode],
    coord_to_uuid: dict[tuple[str, int, str], str],
    pending: list[tuple[ConceptNode, str, str]],
) -> None:
    """Wire face-sheet rows to their sub-sheet sources.

    The face cell DOES NOT mint a new concept — it inherits the sub
    cell's identity but keeps its own render_key so the exporter knows
    where to write the value on the face sheet.  Implemented by
    rewriting the face node's ``concept_uuid`` to point at the sub
    node's UUID, and adding a sentinel edge so cascades can still
    follow the link.
    """
    for face_node, target_sheet, target_coord in pending:
        m = re.match(r"^([A-Z]+)(\d+)$", target_coord)
        if not m:
            continue
        col, row = m.group(1), int(m.group(2))
        target_uuid = coord_to_uuid.get((target_sheet, row, col))
        if target_uuid is None:
            # Sub row uses a different value column (B is the default
            # we indexed); fall back to a sheet+row lookup.
            for (s, r, _c), uid in coord_to_uuid.items():
                if s == target_sheet and r == row:
                    target_uuid = uid
                    break
        if target_uuid is None:
            continue

        # Replace the face node's identity with the sub node's so any
        # downstream lookup by canonical_label converges.  Keep the
        # render_key untouched so the exporter still writes to the
        # face sheet.
        face_node.concept_uuid = target_uuid
        face_node.edges = [{
            "ref": {"sheet": target_sheet, "col": col, "row": row},
            "coefficient": 1,
            "kind": "cross_sheet",
        }]


# -- UUID minting -----------------------------------------------------


def _mint_uuid(template_id: str, sheet: str, row: int, label: str) -> str:
    """Deterministic UUID5 per (template, sheet, row, canonical_label)."""
    name = f"{template_id}::{sheet}::{row}::{label}"
    return str(uuid.uuid5(_CONCEPT_NS, name))


def mint_notes_concept_uuid(sheet: str, row: int, label: str) -> str:
    """Deterministic UUID5 for a notes cell (Phase 7).

    Notes templates aren't parsed into the concept tree, but every notes
    cell still gets a stable concept identity so the unified fact store
    can address it. Keyed on (sheet, row, label) with a ``notes::``
    prefix so it can never collide with a face-statement concept UUID
    (which is keyed on ``template_id::…``). Re-writing the same cell —
    including a Sheet-12 LIST_OF_NOTES fan-out row — yields the same UUID.
    """
    name = f"notes::{sheet}::{row}::{label}"
    return str(uuid.uuid5(_CONCEPT_NS, name))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def _cli(argv: Iterable[str]) -> int:
    import argparse
    import json
    import sys

    parser = argparse.ArgumentParser(
        description="Parse an XBRL template into a concept tree (JSON)."
    )
    parser.add_argument("xlsx", nargs="?", help="Path to .xlsx template")
    parser.add_argument("--pretty", action="store_true",
                        help="Pretty-print JSON output")
    parser.add_argument("--all", action="store_true",
                        help="Batch-dump every live template into "
                             "output/concept_trees/{template_id}.json")
    args = parser.parse_args(list(argv))

    if args.all:
        return _cli_all(pretty=args.pretty)

    if not args.xlsx:
        parser.error("xlsx path required unless --all is passed")

    tree = parse_template(args.xlsx)
    indent = 2 if args.pretty else None
    json.dump(tree.to_json(), sys.stdout, indent=indent, sort_keys=True)
    sys.stdout.write("\n")
    return 0


def _cli_all(pretty: bool) -> int:
    import json
    import sys

    repo = Path(__file__).resolve().parent.parent
    out_dir = repo / "output" / "concept_trees"
    out_dir.mkdir(parents=True, exist_ok=True)

    template_roots = [
        repo / "XBRL-template-MFRS",
        repo / "XBRL-template-MPERS",
    ]
    skipped: list[tuple[Path, str]] = []
    ok = 0
    for root in template_roots:
        if not root.is_dir():
            continue
        for xlsx in sorted(root.rglob("*.xlsx")):
            # Non-authoritative snapshots living under the template root:
            # backup-originals/, archive-*/ (pre-regeneration captures),
            # and snapshot-*/. Parsing these as live templates would
            # overwrite the real template's JSON output (same template_id).
            if any(
                p.startswith(("backup", "archive", "snapshot"))
                for p in xlsx.parts
            ):
                continue
            try:
                tree = parse_template(str(xlsx))
            except UnsupportedSchemaShape as exc:
                skipped.append((xlsx, str(exc)))
                continue
            indent = 2 if pretty else None
            (out_dir / f"{tree.template_id}.json").write_text(
                json.dumps(tree.to_json(), indent=indent, sort_keys=True),
                encoding="utf-8",
            )
            ok += 1
    sys.stderr.write(
        f"wrote {ok} concept trees to {out_dir} "
        f"({len(skipped)} skipped: matrix shape)\n"
    )
    return 0


if __name__ == "__main__":
    import sys

    raise SystemExit(_cli(sys.argv[1:]))
