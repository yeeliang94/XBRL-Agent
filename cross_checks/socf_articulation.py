"""Internal-footing check: SOCF articulates (opening + net change == closing).

The cross-statement check ``socf_to_sofp_cash`` only compares SOCF's
closing-cash LEAF to SOFP's cash leaf — if both carry the same (correct)
closing figure it passes even when SOCF's own subtotals don't foot. A sign
error on an operating-section adjustment row leaves closing cash right (it's a
hard-typed leaf the agent reads straight from the PDF) while the COMPUTED net
change is wrong, so the statement looks fine at the headline and silently
fails to articulate (run-50 / Amway, RM 61,976; the two errors partially
offset, so the gap looked small).

Two independent validation layers each had a gap that let this class through:
the extraction-time verifier silently skipped its articulation check when the
net-change row label didn't match, and NO cross-check validated SOCF's own
internal footing — so every cross-check passed and the reviewer never fired.
This check closes the cross-check-layer gap at the FACT level: a
non-articulating SOCF now FAILS and triggers the reviewer.

Mirrors the verifier's identity ``closing cash == opening cash + net change
after FX`` (``tools/verifier._verify_socf`` Check 2). Implements both the
xlsx ``run`` and fact-based ``run_facts`` paths like every other cross-check;
the fact path is default-on (``XBRL_FACT_BASED_CHECKS``).
"""
from __future__ import annotations

from typing import Dict, Optional

from statement_types import StatementType
from cross_checks.framework import CrossCheckResult, Comparand
from cross_checks.util import (
    open_workbook, find_sheet, find_value_by_label, find_label_row,
    filing_level_prefix,
)
from cross_checks._format import fmt_amount, fmt_diff


# Magnitude-scaled tolerance, mirroring tools.verifier._balance_tolerance so
# the cross-check layer absorbs a large entity's accumulated rounding without
# ever tolerating a real, line-item-sized discrepancy. The passed-in tolerance
# (default RM 1) is the floor.
def _effective_tolerance(tolerance: float, *magnitudes: Optional[float]) -> float:
    scale = max((abs(m) for m in magnitudes if m is not None), default=0.0)
    return max(tolerance, 1e-6 * scale)


def _find_net_change_xlsx(ws, col: int, wb):
    """Bottom-most cash net-change row that is NOT the before-FX subtotal.

    The xlsx twin of ``facts_util.read_socf_net_change`` — resolves the FINAL
    (after-FX) net-change row structurally so it binds whether the template
    says "after effect of exchange rate changes" (MFRS) or labels the row
    plainly (MPERS). Returns ``(value, row)``; value is None when the row
    isn't found or carries no usable number.
    """
    from openpyxl.utils import get_column_letter
    from tools.verifier import _resolve_cell_value

    match_row = None
    for r in range(1, ws.max_row + 1):
        raw_label = ws.cell(row=r, column=1).value
        if raw_label is None:
            continue
        norm = str(raw_label).strip().lstrip("*").strip().lower()
        if (
            ("net increase" in norm or "net change" in norm)
            and "cash" in norm
            and "before" not in norm
        ):
            match_row = r  # keep scanning — bottom-most match wins
    if match_row is None:
        return None, None
    raw = ws.cell(row=match_row, column=col).value
    if raw is None:
        return None, match_row
    if isinstance(raw, str) and raw.startswith("="):
        cell_ref = f"{get_column_letter(col)}{match_row}"
        return _resolve_cell_value(wb, ws.title, cell_ref), match_row
    try:
        return float(raw), match_row
    except (ValueError, TypeError):
        return None, match_row


class SOCFArticulationCheck:
    name = "socf_articulation"
    required_statements = {StatementType.SOCF}

    def applies_to(self, run_config: dict) -> bool:
        # Every SOCF — Indirect or Direct, MFRS or MPERS — has this identity.
        return True

    def _evaluate(
        self,
        beginning: Optional[float],
        net_change: Optional[float],
        ending: Optional[float],
        tolerance: float,
        label: str,
    ) -> tuple[bool, str]:
        """Return (passed, message-part) for one entity scope."""
        if beginning is None or net_change is None or ending is None:
            return False, (
                f"{label}: missing SOCF row(s) "
                f"(opening={beginning}, net change={net_change}, "
                f"closing={ending}) — cannot verify articulation"
            )
        expected = beginning + net_change
        diff = abs(ending - expected)
        eff_tol = _effective_tolerance(tolerance, beginning, net_change, ending)
        passed = diff <= eff_tol
        return passed, (
            f"{label}: closing cash ({fmt_amount(ending)}) "
            f"{'==' if passed else '!='} opening ({fmt_amount(beginning)}) + "
            f"net change ({fmt_amount(net_change)}) = {fmt_amount(expected)}, diff={fmt_diff(diff)}"
        )

    def run(self, workbook_paths: Dict[StatementType, str], tolerance: float,
            filing_level: str = "company", filing_standard: str = "mfrs") -> CrossCheckResult:
        wb = open_workbook(workbook_paths[StatementType.SOCF])
        ws = find_sheet(wb, "SOCF-Indirect", "SOCF-Direct")
        if ws is None:
            wb.close()
            return CrossCheckResult(
                name=self.name, status="failed",
                message="No SOCF sheet found in workbook",
            )
        sheet = ws.title

        def _read(col: int):
            beginning = find_value_by_label(
                ws, "cash and cash equivalents at beginning of period",
                col=col, wb=wb)
            ending = find_value_by_label(
                ws, "cash and cash equivalents at end of period",
                col=col, wb=wb)
            net_change, nc_row = _find_net_change_xlsx(ws, col, wb)
            return beginning, net_change, ending, nc_row

        primary_label = filing_level_prefix(filing_level, with_period=False)
        beginning, net_change, ending, nc_row = _read(2)
        passed, part = self._evaluate(
            beginning, net_change, ending, tolerance, primary_label)
        parts = [part]

        co_passed = True
        co_vals = (None, None, None)
        if filing_level == "group":
            co_beginning, co_net_change, co_ending, _ = _read(4)
            co_vals = (co_beginning, co_net_change, co_ending)
            co_passed, co_part = self._evaluate(
                co_beginning, co_net_change, co_ending, tolerance, "Company")
            parts.append(co_part)

        wb.close()

        # Rows for the cash leaves so the comparand coords match the fact path
        # (which resolves real (sheet,row)) — keeps the e2e shadow-parity exact.
        ending_row = find_label_row(ws, "cash and cash equivalents at end of period")
        beginning_row = find_label_row(ws, "cash and cash equivalents at beginning of period")

        socf = StatementType.SOCF.value
        comparands = [
            Comparand(label="Cash and cash equivalents at end of period",
                      sheet=sheet, value=ending, role="lhs", statement=socf,
                      row=ending_row),
            Comparand(label="Cash and cash equivalents at beginning of period",
                      sheet=sheet, value=beginning, role="leaf", statement=socf,
                      row=beginning_row),
            Comparand(label="Net increase (decrease) in cash and cash equivalents",
                      sheet=sheet, value=net_change, role="leaf", statement=socf,
                      row=nc_row),
        ]
        if filing_level == "group":
            comparands += [
                Comparand(label="Cash and cash equivalents at end of period [company]",
                          sheet=sheet, value=co_vals[2], role="lhs", statement=socf),
                Comparand(label="Cash and cash equivalents at beginning of period [company]",
                          sheet=sheet, value=co_vals[0], role="leaf", statement=socf),
                Comparand(label="Net increase (decrease) in cash and cash equivalents [company]",
                          sheet=sheet, value=co_vals[1], role="leaf", statement=socf),
            ]

        return CrossCheckResult(
            name=self.name,
            status="passed" if passed and co_passed else "failed",
            expected=(beginning + net_change) if (beginning is not None and net_change is not None) else None,
            actual=ending,
            diff=(abs(ending - (beginning + net_change)) if (beginning is not None and net_change is not None and ending is not None) else None),
            tolerance=tolerance,
            message="; ".join(parts),
            target_sheet=sheet,
            target_row=nc_row,
            comparands=comparands,
        )

    def run_facts(self, ctx, tolerance: float) -> CrossCheckResult:
        """Fact-based twin of :meth:`run` (item 32). Reads the opening/closing
        cash leaves and the COMPUTED net-change subtotal from
        ``run_concept_facts`` by uuid (the cascade persists COMPUTED parents,
        so the net change is present as a fact)."""
        from cross_checks.facts_util import (
            primary_scope, read_labelled_value, read_socf_net_change,
        )

        scope = primary_scope(ctx.filing_level)

        def _read(entity_scope: str):
            beginning = read_labelled_value(
                ctx, StatementType.SOCF,
                "cash and cash equivalents at beginning of period",
                "CY", entity_scope)
            ending = read_labelled_value(
                ctx, StatementType.SOCF,
                "cash and cash equivalents at end of period",
                "CY", entity_scope)
            net_change = read_socf_net_change(ctx, "CY", entity_scope)
            return beginning, ending, net_change

        primary_label = filing_level_prefix(ctx.filing_level, with_period=False)
        beginning, ending, net_change = _read(scope)
        sheet = (net_change.sheet or ending.sheet or beginning.sheet or "SOCF")
        passed, part = self._evaluate(
            beginning.value, net_change.value, ending.value, tolerance,
            primary_label)
        parts = [part]

        co_passed = True
        co_b = co_e = co_n = None
        if ctx.filing_level == "group":
            cb, ce, cn = _read("Company")
            co_b, co_e, co_n = cb.value, ce.value, cn.value
            co_passed, co_part = self._evaluate(
                co_b, co_n, co_e, tolerance, "Company")
            parts.append(co_part)

        socf = StatementType.SOCF.value
        comparands = [
            Comparand(label="Cash and cash equivalents at end of period",
                      sheet=sheet, value=ending.value, role="lhs",
                      statement=socf, row=ending.row),
            Comparand(label="Cash and cash equivalents at beginning of period",
                      sheet=sheet, value=beginning.value, role="leaf",
                      statement=socf, row=beginning.row),
            Comparand(label="Net increase (decrease) in cash and cash equivalents",
                      sheet=sheet, value=net_change.value, role="leaf",
                      statement=socf, row=net_change.row),
        ]
        if ctx.filing_level == "group":
            comparands += [
                Comparand(label="Cash and cash equivalents at end of period [company]",
                          sheet=sheet, value=co_e, role="lhs", statement=socf),
                Comparand(label="Cash and cash equivalents at beginning of period [company]",
                          sheet=sheet, value=co_b, role="leaf", statement=socf),
                Comparand(label="Net increase (decrease) in cash and cash equivalents [company]",
                          sheet=sheet, value=co_n, role="leaf", statement=socf),
            ]

        b, n, e = beginning.value, net_change.value, ending.value
        complete = b is not None and n is not None and e is not None
        return CrossCheckResult(
            name=self.name,
            status="passed" if passed and co_passed else "failed",
            expected=(b + n) if (b is not None and n is not None) else None,
            actual=e,
            diff=(abs(e - (b + n)) if complete else None),
            tolerance=tolerance,
            message="; ".join(parts),
            target_sheet=sheet,
            target_row=net_change.row,
            comparands=comparands,
        )
