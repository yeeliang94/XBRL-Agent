"""Advisory check: SOCF lines sit in the section the source printed them in.

Every other cash-flow check is an arithmetic identity, and none of them can see
this class of error. Moving a line between operating, investing and financing
does not change the net change in cash, so ``socf_articulation`` (opening + net
change == closing) still passes, the closing-cash tie-out to SOFP still passes,
and the statement reports a wrong operating cash flow while every check is
green. That is what happened in run 84: a dividend of 65,345 was placed in
operating activities, distorting the operating subtotal by that amount with no
effect on the total.

**Why this warns and never fails.** There is often no single correct answer.
MFRS permits dividends paid in either operating or financing, and the template
carries a "Dividends paid" row in BOTH sections precisely for that reason. The
only test is whether the placement matches where the SOURCE statement prints
it — which means this check is only as good as the scout's reading of the face
page's section headings, and that accuracy has never been measured. A warning
that is sometimes wrong costs an operator a glance; a failure that is sometimes
wrong sends the reviewer chasing a non-problem on every run.

**Deliberately conservative.** Labels are matched EXACTLY after normalisation —
no fuzzy matching (gotcha #13's rule that scout output is advisory cuts both
ways: it must not manufacture warnings either). A line the scout did not record,
recorded without a section, or recorded under a heading we cannot classify is
skipped rather than guessed at. The check reports what it can see and stays
silent about the rest.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Optional

import openpyxl


# The three cash-flow sections, keyed by the word that identifies them in both
# a template section header and a scout-observed heading. Order matters only for
# determinism of the scan.
_SECTIONS = ("operating", "investing", "financing")

_SOCF_SHEETS = ("SOCF-Indirect", "SOCF-Direct")

# Value columns to test for "the agent wrote something on this row". Company
# filings use B/C; group filings add D/E. Reading all four is harmless — an
# empty cell simply doesn't count.
_VALUE_COLUMNS = (2, 3, 4, 5)


@dataclass
class SectionPlacementWarning:
    """One line whose section differs from where the scout saw it printed."""
    status: str  # always "warning"
    label: str
    row: int
    template_section: str
    source_section: str
    message: str


def _normalize(label: str) -> str:
    """Lower-case, strip the template's leading ``*`` and trailing punctuation,
    collapse whitespace. The two sides of the comparison come from different
    places — a template cell and a scout reading of a PDF line — so they agree
    on wording far more often than on decoration."""
    text = str(label or "").strip().lstrip("*").strip()
    text = re.sub(r"\s+", " ", text)
    return text.rstrip(":.").strip().lower()


def classify_section(heading: Optional[str]) -> Optional[str]:
    """Map a heading to ``operating`` / ``investing`` / ``financing``.

    Returns None when the heading names none of them — a subtotal line, a
    blank, or wording we don't recognise. None means "don't compare", never
    "no disagreement".
    """
    if not heading:
        return None
    lowered = str(heading).lower()
    hits = [s for s in _SECTIONS if s in lowered]
    # A heading naming two sections (a reconciliation line quoting both) is
    # ambiguous, so it identifies neither.
    return hits[0] if len(hits) == 1 else None


def _sheet_for(wb) -> Optional[Any]:
    for name in _SOCF_SHEETS:
        if name in wb.sheetnames:
            return wb[name]
    return None


def template_sections_by_label(ws) -> dict[str, tuple[str, int]]:
    """Map each written row's normalised label to ``(section, row)``.

    A row's section is the nearest section heading ABOVE it. Headings are found
    by wording rather than by fill colour: the fill marks every abstract row
    (gotcha #17), of which the three section headers are only a few, and the
    wording is what carries the operating/investing/financing meaning anyway.

    Only rows carrying a value are returned — an untouched template row has no
    placement to disagree about.
    """
    out: dict[str, tuple[str, int]] = {}
    current: Optional[str] = None
    for row in range(1, ws.max_row + 1):
        raw = ws.cell(row=row, column=1).value
        if raw is None:
            continue
        label = str(raw)
        section = classify_section(label)
        # A heading is a row that names a section and carries no value of its
        # own. "Net cash flows from (used in) operating activities" is a
        # subtotal, not a heading — but it also names its own section, so
        # treating it as one changes nothing for the rows beneath it.
        has_value = any(
            ws.cell(row=row, column=c).value not in (None, "")
            for c in _VALUE_COLUMNS
        )
        if section is not None and not has_value:
            current = section
            continue
        if current is None or not has_value:
            continue
        # A formula cell is a computed total, not something the agent placed.
        if all(
            isinstance(ws.cell(row=row, column=c).value, str)
            and str(ws.cell(row=row, column=c).value).startswith("=")
            for c in _VALUE_COLUMNS
            if ws.cell(row=row, column=c).value not in (None, "")
        ):
            continue
        key = _normalize(label)
        # First occurrence wins. A label appearing in two sections (the
        # dividends case) is exactly what we want to catch, and the agent
        # wrote to one of them — whichever carries a value is that one.
        out.setdefault(key, (current, row))
    return out


def check_socf_section_placement(
    workbook_path: str,
    face_line_refs: Any,
) -> list[SectionPlacementWarning]:
    """Compare each written SOCF line's section against the scout's reading.

    ``face_line_refs`` is the SOCF entry's ``face_line_refs`` from the run's
    infopack — objects (or dicts) carrying ``label`` and ``section``. An empty
    list yields no warnings: with nothing observed there is nothing to compare,
    and silence here means "not assessed", not "agrees".

    Never raises: a workbook we can't read, a missing sheet, or a malformed ref
    returns an empty list. This is an advisory pass and must not be able to
    break a run (gotcha #20's sibling rule).
    """
    if not face_line_refs:
        return []
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
    except Exception:  # noqa: BLE001 — advisory only
        return []
    try:
        ws = _sheet_for(wb)
        if ws is None:
            return []
        placed = template_sections_by_label(ws)
    finally:
        wb.close()

    warnings: list[SectionPlacementWarning] = []
    seen: set[str] = set()
    for ref in face_line_refs:
        label = getattr(ref, "label", None)
        section = getattr(ref, "section", None)
        if label is None and isinstance(ref, dict):
            label, section = ref.get("label"), ref.get("section")
        source_section = classify_section(section)
        if source_section is None:
            continue
        key = _normalize(label)
        if not key or key in seen:
            continue
        entry = placed.get(key)
        if entry is None:
            continue
        template_section, row = entry
        if template_section == source_section:
            continue
        seen.add(key)
        warnings.append(SectionPlacementWarning(
            status="warning",
            label=str(label).strip(),
            row=row,
            template_section=template_section,
            source_section=source_section,
            message=(
                f"'{str(label).strip()}' was entered under {template_section} "
                f"activities (row {row}), but the source statement appears to "
                f"print it under {source_section} activities. Moving a line "
                f"between sections does not change total cash, so no other "
                f"check can see this — confirm against the PDF. If the source "
                f"really does print it under {template_section}, this warning "
                f"is the one that is wrong."
            ),
        ))
    return warnings
