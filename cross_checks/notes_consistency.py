"""Phase 6.1 — post-run notes cross-sheet consistency check.

Sheet 11 (`Notes-SummaryofAccPol`) and Sheet 12 (`Notes-Listofnotes`)
often describe the same underlying disclosure from different angles.
When both sheets end up populated for one topic, the evidence strings
should agree on which PDF page the disclosure lives on. On the FINCO
2021 run we observed:

- Sheet 11 row "Description of accounting policy for employee benefits"
  cited "Page 27, Note 2.5(g)".
- Sheet 12 row "Disclosure of employee benefits expense" cited
  "Page 25, Note 2.5(g)".

The prose is the same; the page number drifted because a sub-agent read
the printed folio instead of the PDF page number. A WARN-level check
after the run surfaces these mismatches without blocking the merge —
operators get a signal to open both sheets and re-reconcile.

Kept intentionally small:
- Standalone module (not plugged into the face-statement cross-check
  runner). The server/coordinator can import + call it when ready.
- Topic pairing is a conservative hand-coded dict. Unknown topics
  produce no warnings — we never fabricate a pairing the taxonomy
  doesn't clearly establish.
- Page numbers are parsed out of the human-readable evidence string
  with a permissive regex. The check short-circuits to "could not
  compare" when either side has no parseable page.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import openpyxl


# Hand-curated topic pairings. Both sides of each pair are row labels
# exactly as they appear in col A of the respective templates (leading
# `*` stripped, trailing whitespace stripped). Only include pairings
# where the overlap is unambiguous — where the Sheet-11 policy row
# narrates the same accounting treatment that the Sheet-12 disclosure
# row then quantifies (or vice versa).
_TOPIC_PAIRS: list[tuple[str, str]] = [
    ("Description of accounting policy for cash and cash equivalents",
     "Disclosure of cash and cash equivalents"),
    ("Description of accounting policy for employee benefits",
     "Disclosure of employee benefits expense"),
    ("Description of accounting policy for deferred income tax",
     "Disclosure of deferred tax assets/(liabilities)"),
    ("Description of accounting policy for income tax",
     "Disclosure of income tax expense"),
    ("Description of accounting policy for leases",
     "Disclosure of leases"),
    ("Description of accounting policy for fair value measurement",
     "Disclosure of fair value measurement"),
    ("Description of accounting policy for financial instruments",
     "Disclosure of financial instruments"),
    ("Description of accounting policy for impairment of financial assets",
     "Disclosure of allowance for credit losses"),
    ("Description of accounting policy for recognition of revenue and "
     "other income",
     "Disclosure of revenue from contract customers"),
    ("Description of accounting policy for other provisions",
     "Disclosure of other provisions"),
    ("Description of accounting policy for property, plant and equipment",
     "Disclosure of property, plant and equipment"),
]


# Evidence strings are free-form human prose. We only need to pluck
# integer page numbers out. Examples:
#   "Page 27, Note 2.5(g)"              → {27}
#   "Pages 20-21, Note 2.5(b)(i)"       → {20, 21}
#   "Page 27, Note 2.5(g); Page 33"     → {27, 33}
#   "pp. 27-28, Note 2.5(g)"            → {27, 28}
#   "p. 33"                             → {33}
# The regex matches the "Page"/"Pages"/"p."/"pp." forms followed by an
# integer or integer-range. Bare numbers elsewhere in the string
# (e.g. "Note 2.5(g)" → the "2.5") are NOT counted because they are
# always preceded by "Note " or "(", never by a page marker.
#
# Peer-review I-2: the original regex missed "pp." / "p." — the exact
# abbreviation the Sheet-12 sub-coordinator prints in its own started-
# status labels ("pp X-Y"). Model-emitted evidence using that form was
# silently yielding set(), defeating the whole cross-sheet check for a
# common citation shape. ``\.?`` on each form keeps an unabbreviated
# "pp 20-22" (no dot) working too.
_PAGE_TOKEN = re.compile(
    r"\b(?:pages?|pp?)\.?\s+(\d+)(?:\s*[-–]\s*(\d+))?",
    re.IGNORECASE,
)


def _extract_pages(evidence: str) -> set[int]:
    """Parse page numbers from an evidence string. Returns {} on failure."""
    if not evidence:
        return set()
    pages: set[int] = set()
    for m in _PAGE_TOKEN.finditer(evidence):
        start = int(m.group(1))
        end_raw = m.group(2)
        if end_raw:
            end = int(end_raw)
            # Defensive: guard against reversed ranges ("Pages 30-20").
            if end < start:
                start, end = end, start
            pages.update(range(start, end + 1))
        else:
            pages.add(start)
    return pages


def _norm(label: str) -> str:
    """Normalise a label for matching: strip, drop leading '*', lowercase."""
    if not label:
        return ""
    return label.strip().lstrip("*").strip().lower()


@dataclass
class ConsistencyWarning:
    """One inconsistency between paired Sheet-11 and Sheet-12 rows.

    ``status`` is always "warning" by design — this check never fails
    the merge. The field is kept so a future evolution can introduce
    stricter levels without breaking existing consumers.
    """
    status: str  # always "warning"
    sheet_11_label: str
    sheet_12_label: str
    sheet_11_evidence: str
    sheet_12_evidence: str
    sheet_11_pages: list[int]
    sheet_12_pages: list[int]
    message: str


def _read_label_and_evidence(
    ws, label_col: int = 1, evidence_cols: tuple[int, ...] = (4, 6),
) -> dict[str, str]:
    """Build {normalised label: evidence} for every row with content.

    ``evidence_cols`` covers both Company (col D = 4) and Group (col F = 6)
    layouts. We scan both and take the first non-empty STRING. The string
    guard matters on Group filings: numeric notes rows hold VALUES in col D
    (gotcha #14 — B/C/D/E are value columns, evidence moves to col F), so
    accepting any non-empty cell would capture the number as "evidence" and
    never reach the real citation in col F.
    """
    # Peer-review S5: openpyxl's per-cell `ws.cell(row, col)` accessor
    # is roughly an order of magnitude slower than `iter_rows`, which
    # streams a single pass through the underlying XML. ~138 rows is
    # negligible today; switching pre-emptively makes the cost stay
    # flat as new templates land.
    max_evidence_col = max(evidence_cols) if evidence_cols else label_col
    max_col = max(label_col, max_evidence_col)
    out: dict[str, str] = {}
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True,
    ):
        raw_label = row[label_col - 1] if len(row) >= label_col else None
        if not raw_label or not str(raw_label).strip():
            continue
        label = _norm(str(raw_label))
        evidence = ""
        for col in evidence_cols:
            if len(row) < col:
                continue
            v = row[col - 1]
            # Strings only: on Group filings numeric notes rows carry the
            # VALUE in col D and the citation in col F (gotcha #14) —
            # a numeric cell must fall through to the next evidence col.
            if isinstance(v, str) and v.strip():
                evidence = v.strip()
                break
        if evidence:
            out[label] = evidence
    return out


# ---------------------------------------------------------------------------
# N4 — generalized citation-consistency check.
#
# The curated-pairs pass above only compares ~10 hand-coded topic pairs across
# Sheets 11/12. But the underlying failure — one sub-agent cites the printed
# folio, another the PDF page — applies to ANY two cells citing the same note
# ref. This second pass groups EVERY filled cell's citation by its note ref
# across all notes sheets and WARNs when one ref's cited pages diverge grossly
# (span beyond the note's known page range). It infers nothing about which
# rows pair up — it groups by the agents' OWN citations, so it never fabricates
# a pairing (the module's conservatism invariant).
# ---------------------------------------------------------------------------

# A note ref token: "Note 2.5(g)", "Note 18", "note 2.14(a)". Captures the ref
# body after the "Note " marker, leaving the page parsing to _PAGE_TOKEN.
_NOTE_REF_TOKEN = re.compile(
    r"\bnote\s+(\d+(?:\.\d+)*(?:\([a-z0-9]+\))*)",
    re.IGNORECASE,
)

# Default allowed page span for a note when the inventory doesn't tell us its
# range — generous so the generic pass only flags GROSS drift (folio-vs-PDF),
# leaving tight same-topic drift to the curated pass.
_DEFAULT_NOTE_SPAN = 3


@dataclass
class CitationWarning:
    """One note ref cited with grossly divergent pages across cells."""
    status: str  # always "warning"
    note_ref: str
    citations: list[tuple[str, list[int]]]  # (location, pages)
    message: str


def _extract_note_refs(evidence: str) -> list[str]:
    """Note refs in an evidence string ('2.5(g)', '18'), lowercased."""
    if not evidence:
        return []
    return [m.group(1).lower() for m in _NOTE_REF_TOKEN.finditer(evidence)]


def _top_note_num(ref: str) -> Optional[int]:
    """Top-level integer note number of a ref ('2.5(g)' → 2)."""
    head = ref.split(".")[0].split("(")[0]
    try:
        return int(head)
    except ValueError:
        return None


def _read_rows_label_and_evidence(
    ws, label_col: int = 1, evidence_cols: tuple[int, ...] = (4, 6),
) -> list[tuple[int, str, str]]:
    """Row-level (row, label, evidence) for every populated row.

    Unlike :func:`_read_label_and_evidence` (which keys by normalised label and
    so COLLAPSES duplicate-label rows — fine for the curated pass that looks up
    unique policy labels), the generic citation pass must see EVERY cell:
    notes sheets legitimately repeat labels ("Balance at the end of period",
    "Total"), and two same-label rows citing one note ref with divergent pages
    is exactly the drift N4 hunts (peer-review MEDIUM). Keyed by row, nothing
    is overwritten.
    """
    max_col = max((label_col, *evidence_cols)) if evidence_cols else label_col
    out: list[tuple[int, str, str]] = []
    row_idx = 0
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True,
    ):
        row_idx += 1
        raw_label = row[label_col - 1] if len(row) >= label_col else None
        if not raw_label or not str(raw_label).strip():
            continue
        evidence = ""
        for col in evidence_cols:
            if len(row) < col:
                continue
            v = row[col - 1]
            # Strings only: on Group filings numeric notes rows carry the
            # VALUE in col D and the citation in col F (gotcha #14) —
            # a numeric cell must fall through to the next evidence col.
            if isinstance(v, str) and v.strip():
                evidence = v.strip()
                break
        if evidence:
            out.append((row_idx, _norm(str(raw_label)), evidence))
    return out


def check_notes_citation_consistency(
    workbook_path: str,
    note_spans: Optional[dict[int, int]] = None,
    default_span: int = _DEFAULT_NOTE_SPAN,
) -> list[CitationWarning]:
    """Generic citation-consistency pass across ALL notes sheets.

    Groups every filled cell's citation by its note ref. A ref cited on two or
    more cells whose pages are disjoint AND span more than the note's allowed
    range (``note_spans[top_note_num]`` if known, else ``default_span``) yields
    one WARN. Evidence with no parseable page or no single note ref is skipped
    (never a false positive). Advisory only — never raises.
    """
    path = Path(workbook_path)
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception:  # noqa: BLE001 — advisory, never raise
        return []
    try:
        notes_sheets = [n for n in wb.sheetnames if n.lower().startswith("notes-")]
        # ref -> list of (location, page-set)
        by_ref: dict[str, list[tuple[str, set[int]]]] = {}
        for sheet_name in notes_sheets:
            # Row-level read (not label-keyed) so duplicate-label rows on one
            # sheet each contribute a citation (peer-review MEDIUM).
            for row_idx, label, evidence in _read_rows_label_and_evidence(
                wb[sheet_name]
            ):
                refs = _extract_note_refs(evidence)
                # Only single-ref evidence — multi-ref strings can't be
                # attributed to one note unambiguously, so we skip them.
                if len(set(refs)) != 1:
                    continue
                pages = _extract_pages(evidence)
                if not pages:
                    continue
                by_ref.setdefault(refs[0], []).append(
                    (f"{sheet_name}!row{row_idx} {label}", pages)
                )

        warnings: list[CitationWarning] = []
        for ref, cites in by_ref.items():
            if len(cites) < 2:
                continue
            page_sets = [p for _, p in cites]
            union = set().union(*page_sets)
            span = max(union) - min(union)
            allowed = default_span
            top = _top_note_num(ref)
            if note_spans and top is not None and top in note_spans:
                allowed = max(note_spans[top], 0)
            if span <= allowed:
                continue
            # Require at least one disjoint pair — overlapping citations of a
            # genuinely multi-page note are consistent even if the span is wide.
            disjoint = any(
                page_sets[i].isdisjoint(page_sets[j])
                for i in range(len(page_sets))
                for j in range(i + 1, len(page_sets))
            )
            if not disjoint:
                continue
            warnings.append(CitationWarning(
                status="warning",
                note_ref=ref,
                citations=[(loc, sorted(p)) for loc, p in cites],
                message=(
                    f"Note {ref} is cited with disjoint pages that span "
                    f"{span} (> allowed {allowed}): "
                    + "; ".join(f"{loc} → {sorted(p)}" for loc, p in cites)
                    + ". One citation may use the printed folio instead of the "
                    "PDF page."
                ),
            ))
        return warnings
    finally:
        wb.close()


def check_notes_consistency(
    workbook_path: str,
    sheet_11_name: str = "Notes-SummaryofAccPol",
    sheet_12_name: str = "Notes-Listofnotes",
) -> list[ConsistencyWarning]:
    """Compare paired rows across the two notes sheets.

    Returns a list of WARNINGs — empty when the run is clean. Any
    sheet-missing or read error is swallowed and returns [] (this check
    is advisory and must never break the merge).
    """
    path = Path(workbook_path)
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception:  # noqa: BLE001 — advisory check, never raise
        return []

    try:
        if sheet_11_name not in wb.sheetnames or sheet_12_name not in wb.sheetnames:
            # If either sheet is missing there's nothing to compare. A
            # user who requested only one of the two sheets gets no
            # warnings (correct behaviour — there IS no inconsistency
            # when only one side was populated).
            return []

        s11 = _read_label_and_evidence(wb[sheet_11_name])
        s12 = _read_label_and_evidence(wb[sheet_12_name])

        warnings: list[ConsistencyWarning] = []
        for lbl11_raw, lbl12_raw in _TOPIC_PAIRS:
            lbl11 = _norm(lbl11_raw)
            lbl12 = _norm(lbl12_raw)
            ev11 = s11.get(lbl11)
            ev12 = s12.get(lbl12)
            if not ev11 or not ev12:
                # Only one side populated → no comparison to make.
                continue

            pages11 = _extract_pages(ev11)
            pages12 = _extract_pages(ev12)
            if not pages11 or not pages12:
                # One or both citations had no parseable page number. We
                # choose to skip rather than warn — warning on un-
                # parseable evidence would be noisy and actionable only
                # by a prompt edit, which the writer can't surface.
                continue

            # Disjoint pages → true inconsistency. Non-disjoint is OK
            # even when the sets aren't equal (a Sheet-11 policy may
            # point at pp 20-22 while Sheet-12 pulls the schedule from
            # p 31 — as long as they share ONE page, the citations are
            # consistent enough).
            if pages11.isdisjoint(pages12):
                warnings.append(ConsistencyWarning(
                    status="warning",
                    sheet_11_label=lbl11_raw,
                    sheet_12_label=lbl12_raw,
                    sheet_11_evidence=ev11,
                    sheet_12_evidence=ev12,
                    sheet_11_pages=sorted(pages11),
                    sheet_12_pages=sorted(pages12),
                    message=(
                        f"Sheet 11 '{lbl11_raw}' cites pages "
                        f"{sorted(pages11)}; Sheet 12 '{lbl12_raw}' "
                        f"cites pages {sorted(pages12)}. No overlap — "
                        f"one of the two may have cited the printed "
                        f"folio instead of the PDF page."
                    ),
                ))

        return warnings
    finally:
        wb.close()
