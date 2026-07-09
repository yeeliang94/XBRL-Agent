"""Internal-footing checks: profit / TCI reconcile to their attribution split.

The SOCF articulation gap (run-50) was one instance of a wider class — a
statement whose two SUBTOTALS are computed from INDEPENDENT leaf sets and must
agree, with NO cross-check firing the reviewer when they don't. The other two
genuinely-catchable instances (verified against the live MFRS/MPERS templates):

* **SOPL** — `*Profit (loss)` (r26, computed down the income statement:
  `=continuing + discontinued`) must equal `*Total profit (loss)` (r31,
  computed from the attribution leaves: `=owners + other-components + NCI`).
* **SOCI** — `*Total comprehensive income` (computed `=profit + OCI`) must
  equal the attribution `*Total comprehensive income` (computed `=owners + NCI`).

These are NOT tautological: each side is computed from a different leaf set, so
an agent that mis-enters the attribution split makes them diverge — exactly the
SOCF failure mode (a hard, independently-sourced number disagreeing with a
computed one). The within-statement "computed total == its own leaves" identity
IS tautological in fact-space (the cascade builds the total from the leaves), so
no check is added for it. SOCIE's closing equity is likewise computed-from-cells
and already reconciled to SOFP cross-statement (``socie_to_sofp_equity``), so it
gets no internal check here either.

The verifier already runs these identities at extraction time, but (a) only on
group filings, (b) with guards that silently skip when an attribution row isn't
found, and (c) without firing the reviewer. These cross-checks close all three:
they self-gate to ``not_applicable`` when the attribution side isn't disclosed
(the cascade writes no fact for an all-blank computed total), so they never
false-fail a company filing that simply doesn't split the attribution.

Both ``run`` (xlsx) and ``run_facts`` (facts, default-on) paths, group
dual-pass, magnitude-scaled tolerance — same contract as every other check.
"""
from __future__ import annotations

from typing import Dict, Optional

from statement_types import StatementType
from cross_checks.framework import CrossCheckResult, Comparand
from cross_checks.util import (
    open_workbook, find_sheet, find_value_by_label, find_label_row,
    filing_level_prefix,
)
from cross_checks._format import fmt_amount, fmt_diff


def _effective_tolerance(tolerance: float, *magnitudes: Optional[float]) -> float:
    scale = max((abs(m) for m in magnitudes if m is not None), default=0.0)
    return max(tolerance, 1e-6 * scale)


def _find_value_last_match(ws, label: str, col: int, wb):
    """Value + row of the BOTTOM-MOST row whose col-A label matches ``label``
    exactly (xlsx twin of facts_util.read_labelled_value_last). Used for SOCI's
    duplicate "*Total comprehensive income" rows."""
    from openpyxl.utils import get_column_letter
    from tools.verifier import _resolve_cell_value

    target = label.strip().lstrip("*").strip().lower()
    match_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        if str(v).strip().lstrip("*").strip().lower() == target:
            match_row = r  # bottom-most wins
    if match_row is None:
        return None, None
    raw = ws.cell(match_row, col).value
    if raw is None:
        return None, match_row
    if isinstance(raw, str) and raw.startswith("="):
        ref = f"{get_column_letter(col)}{match_row}"
        return _resolve_cell_value(wb, ws.title, ref), match_row
    try:
        return float(raw), match_row
    except (ValueError, TypeError):
        return None, match_row


class _AttributionFootingBase:
    """Shared logic: income-side total == attribution-side total, self-gating
    to not_applicable when the attribution side isn't disclosed."""

    # Subclasses set these.
    name: str = ""
    required_statements: set = set()
    _stmt: StatementType = None
    _sheets: tuple = ()
    _income_label: str = ""
    _attribution_label: str = ""
    _noun: str = "total"
    # The attribution LEAF rows (owners / NCI / other components). The xlsx
    # path gates on whether any of these CELLS carry a value, because a
    # formula-evaluated attribution total reads 0 (not blank) when its leaves
    # are empty — without this gate a company filing that doesn't split the
    # attribution would false-fail. The facts path needs no equivalent: the
    # cascade writes no row for an all-blank computed total (reads as None).
    _attribution_leaf_labels: tuple = ()

    def applies_to(self, run_config: dict) -> bool:
        return True

    def _eval_scope(self, income, attribution, inc_row, attr_row, sheet,
                    tolerance, label, *, is_company: bool) -> dict:
        """Evaluate ONE entity scope. Returns a dict carrying the verdict
        (passed / failed / not_applicable), a message part, comparands, and the
        numeric anchor fields. ``not_applicable`` when this scope's attribution
        isn't disclosed — so it never false-fails a scope that simply doesn't
        split the attribution (e.g. a group's Company column on a wholly-owned
        parent)."""
        stmt_val = self._stmt.value
        suffix = " [company]" if is_company else ""
        if income is None or attribution is None:
            return {
                "verdict": "not_applicable",
                "msg": (
                    f"{label}: {self._noun} attribution not disclosed "
                    f"(income={income}, attribution={attribution})"
                ),
                "comparands": [],
            }
        diff = abs(income - attribution)
        eff_tol = _effective_tolerance(tolerance, income, attribution)
        passed = diff <= eff_tol
        return {
            "verdict": "passed" if passed else "failed",
            "msg": (
                f"{label}: {self._noun} ({fmt_amount(income)}) "
                f"{'==' if passed else '!='} attribution split ({fmt_amount(attribution)}), "
                f"diff={fmt_diff(diff)}"
            ),
            "comparands": [
                Comparand(label=self._income_label + suffix, sheet=sheet,
                          value=income, role="lhs", statement=stmt_val, row=inc_row),
                Comparand(label=self._attribution_label + suffix, sheet=sheet,
                          value=attribution, role="rhs", statement=stmt_val,
                          row=attr_row),
            ],
            "expected": income, "actual": attribution, "diff": diff,
            "sheet": sheet, "row": attr_row,
        }

    def _combine(self, evals: list, tolerance: float) -> CrossCheckResult:
        """Combine per-scope verdicts: failed if ANY disclosed scope fails;
        else passed if any scope was checked; else not_applicable (all scopes
        undisclosed). Numeric anchor fields come from the first failing scope
        (most actionable), else the first passing one."""
        verdicts = [e["verdict"] for e in evals]
        if "failed" in verdicts:
            status = "failed"
        elif "passed" in verdicts:
            status = "passed"
        else:
            status = "not_applicable"
        comparands = [c for e in evals for c in e["comparands"]]
        anchor = (next((e for e in evals if e["verdict"] == "failed"), None)
                  or next((e for e in evals if e["verdict"] == "passed"), None))
        res = CrossCheckResult(
            name=self.name, status=status, tolerance=tolerance,
            message="; ".join(e["msg"] for e in evals),
            comparands=comparands,
        )
        if anchor:
            res.expected = anchor["expected"]
            res.actual = anchor["actual"]
            res.diff = anchor["diff"]
            res.target_sheet = anchor["sheet"]
            res.target_row = anchor["row"]
        return res

    def _scopes(self, filing_level):
        """The (entity_scope-or-column, label, is_company) passes for this
        filing. Group filings dual-pass Group + Company, mirroring every other
        group-aware check — otherwise a wrong Company-column attribution split
        would pass silently."""
        primary = filing_level_prefix(filing_level, with_period=False)
        if filing_level == "group":
            return [("Group", 2, primary, False), ("Company", 4, "Company", True)]
        return [("Company", 2, primary, False)]

    # --- xlsx path ---------------------------------------------------------
    def run(self, workbook_paths: Dict[StatementType, str], tolerance: float,
            filing_level: str = "company", filing_standard: str = "mfrs") -> CrossCheckResult:
        wb = open_workbook(workbook_paths[self._stmt])
        ws = find_sheet(wb, *self._sheets)
        if ws is None:
            wb.close()
            return CrossCheckResult(
                name=self.name, status="failed",
                message=f"No {self._stmt.value} sheet found in workbook")
        sheet = ws.title
        evals = []
        for _scope, col, label, is_company in self._scopes(filing_level):
            income = find_value_by_label(ws, self._income_label, col=col, wb=wb)
            inc_row = find_label_row(ws, self._income_label)
            attribution, attr_row = self._read_attribution_xlsx(ws, col, wb)
            # Gate: if no attribution leaf cell carries a value, the split isn't
            # disclosed — drop the (formula-zero) total so we report
            # not_applicable instead of false-failing (parity with the facts
            # path's None for an all-blank computed total).
            if attribution is not None and not self._attribution_present_xlsx(ws, col):
                attribution = None
            evals.append(self._eval_scope(
                income, attribution, inc_row, attr_row, sheet, tolerance,
                label, is_company=is_company))
        wb.close()
        return self._combine(evals, tolerance)

    def _attribution_present_xlsx(self, ws, col: int) -> bool:
        """True when at least one attribution leaf cell in ``col`` carries a
        NUMERIC value — including an explicit 0. An explicit zero is a disclosed
        split (the agent said "this component is zero"), so it must be validated
        against the income total, not skipped: a 0 attribution while income is
        non-zero is a real mismatch. Only a truly BLANK cell (None) is "not
        disclosed" — mirrors the facts path, where not_disclosed reads as None
        but an explicit_zero leaf makes the cascade write a real 0 total."""
        targets = {lbl.strip().lstrip("*").strip().lower()
                   for lbl in self._attribution_leaf_labels}
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            if str(v).strip().lstrip("*").strip().lower() in targets:
                raw = ws.cell(r, col).value
                if raw is None:
                    continue
                if isinstance(raw, str) and raw.startswith("="):
                    continue  # a formula here isn't an entered leaf
                try:
                    float(raw)
                    return True  # any numeric leaf, incl 0, is a disclosed split
                except (ValueError, TypeError):
                    continue
        return False

    def _read_attribution_xlsx(self, ws, col, wb):
        """Default: the attribution total has a DISTINCT label (SOPL)."""
        val = find_value_by_label(ws, self._attribution_label, col=col, wb=wb)
        row = find_label_row(ws, self._attribution_label)
        return val, row

    # --- facts path --------------------------------------------------------
    def run_facts(self, ctx, tolerance: float) -> CrossCheckResult:
        from cross_checks.facts_util import read_labelled_value
        evals = []
        for entity_scope, _col, label, is_company in self._scopes(ctx.filing_level):
            income = read_labelled_value(ctx, self._stmt, self._income_label, "CY", entity_scope)
            attribution = self._read_attribution_facts(ctx, "CY", entity_scope)
            sheet = attribution.sheet or income.sheet or self._stmt.value
            evals.append(self._eval_scope(
                income.value, attribution.value, income.row, attribution.row,
                sheet, tolerance, label, is_company=is_company))
        return self._combine(evals, tolerance)

    def _read_attribution_facts(self, ctx, period, scope):
        # EXACT-match only (read_labelled_value_last), never the substring
        # fallthrough of read_labelled_value: the attribution label "total
        # profit (loss)" substring-matches the income row "profit (loss)", so a
        # substring read would fall through to the income total when the real
        # attribution total is absent and spuriously pass. Exact match binds
        # only the attribution row (None when undisclosed → not_applicable).
        from cross_checks.facts_util import read_labelled_value_last
        return read_labelled_value_last(ctx, self._stmt, self._attribution_label, period, scope)


class SOPLAttributionFootingCheck(_AttributionFootingBase):
    name = "sopl_attribution_footing"
    required_statements = {StatementType.SOPL}
    _stmt = StatementType.SOPL
    _sheets = ("SOPL-Function", "SOPL-Nature")
    _income_label = "profit (loss)"
    _attribution_label = "total profit (loss)"
    _noun = "profit"
    _attribution_leaf_labels = (
        "profit (loss), attributable to owners of parent",
        "profit (loss) attributable to equity other components",
        "profit (loss), attributable to non-controlling interests",
    )


class SOCIAttributionFootingCheck(_AttributionFootingBase):
    name = "soci_attribution_footing"
    required_statements = {StatementType.SOCI}
    _stmt = StatementType.SOCI
    _sheets = ("SOCI-BeforeTax", "SOCI-BeforeOfTax", "SOCI-NetOfTax")
    # SOCI carries "*Total comprehensive income" on BOTH the income-side row
    # (= profit + OCI) and the attribution row (= owners + NCI). The income
    # side is the FIRST occurrence; the attribution side is the LAST.
    _income_label = "total comprehensive income"
    _attribution_label = "total comprehensive income"
    _noun = "total comprehensive income"
    _attribution_leaf_labels = (
        "comprehensive income, attributable to owners of parent",
        "comprehensive income, attributable to non-controlling interests",
    )

    def _read_attribution_xlsx(self, ws, col, wb):
        return _find_value_last_match(ws, self._attribution_label, col, wb)

    def _read_attribution_facts(self, ctx, period, scope):
        from cross_checks.facts_util import read_labelled_value_last
        attr = read_labelled_value_last(ctx, self._stmt, self._attribution_label, period, scope)
        # If only ONE "total comprehensive income" row exists (no attribution
        # rows in this template), first == last; treat as not-disclosed so we
        # don't compare the income total to itself.
        from cross_checks.facts_util import read_labelled_value
        income = read_labelled_value(ctx, self._stmt, self._income_label, period, scope)
        if attr.row is not None and income.row is not None and attr.row == income.row:
            from cross_checks.facts_util import LabelledValue
            return LabelledValue(None, attr.sheet, attr.row)
        return attr
