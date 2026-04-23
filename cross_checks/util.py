"""Shared helpers for cross-check implementations.

All checks need to open a workbook, find a sheet by partial name, and look
up a cell value by its column-A label. This module centralises that logic.
"""
from __future__ import annotations

from typing import Optional, Sequence, Union

import openpyxl

# Reuse the formula evaluator from the verifier — it handles cross-sheet
# references and weighted sums found in MBRS templates.
from tools.verifier import _resolve_cell_value

# SOCIE column constants (pre-MBRS layout)
_SOCIE_NCI_COL = 23       # W — Non-controlling interests
_SOCIE_TOTAL_COL = 24     # X — Total
_SOCIE_RETAINED_COL = 3   # C — Retained earnings

# Group SOCIE block row ranges (each block is a complete SOCIE for one entity/period)
SOCIE_GROUP_BLOCKS = {
    "group_cy":   (3, 25),
    "group_py":   (27, 49),
    "company_cy": (51, 73),
    "company_py": (75, 97),
}


def is_sore_run(run_config: dict) -> bool:
    """True when the run's SOCIE slot is filled by the MPERS SoRE variant.

    Lives here so the three SOCIE-consuming cross-checks
    (sopl_to_socie_profit, soci_to_socie_tci, socie_to_sofp_equity) can all
    read the SOCIE→SoRE contract from one place (peer-review S4). SoRE
    replaces the matrix SOCIE on MPERS filings, so those three checks have
    nothing to reconcile against and must gate themselves out.
    """
    # Local import to avoid a circular dependency if util.py is later
    # imported from statement_types (today it only goes the other direction).
    from statement_types import StatementType
    variants = run_config.get("variants", {}) or {}
    return variants.get(StatementType.SOCIE) == "SoRE"


def has_nci_data(ws, start_row: int = 1, end_row: Optional[int] = None) -> bool:
    """Check whether the SOCIE sheet has actual NCI data filled in.

    The template's NCI column (W) contains metadata strings (e.g. the header
    'Non-controlling interests' in row 2) and formula scaffolding.  This
    function only counts **numeric** non-zero values — strings and formulas
    are ignored, so metadata rows don't trigger false positives.

    When end_row is provided, only checks within that range (for Group SOCIE
    where each block may have different NCI characteristics).
    """
    max_row = end_row if end_row is not None else ws.max_row
    for row in range(start_row, max_row + 1):
        val = ws.cell(row=row, column=_SOCIE_NCI_COL).value
        if val is None or val == 0:
            continue
        if isinstance(val, str):
            continue
        return True
    return False


# MPERS SOCIE doesn't use the MFRS matrix layout — values land in col B
# (CY) and col C (PY) like every other MPERS statement. The dimensional
# breakdown lives on separate axes, not across 24 columns.
_MPERS_SOCIE_CY_COL = 2
_MPERS_SOCIE_PY_COL = 3


def socie_column(
    ws,
    start_row: int = 1,
    end_row: Optional[int] = None,
    filing_standard: str = "mfrs",
) -> int:
    """Return the correct SOCIE read column for a CY value.

    On MFRS: Total (X=24) if NCI data exists, Retained earnings (C=3)
    otherwise. The NCI check scans a row range so Group SOCIE blocks
    can be scoped independently.

    On MPERS: always col B (2). MPERS SOCIE is a flat 2-column
    (CY/PY) layout with dimensional members on separate rows — the
    matrix logic doesn't apply. Bypassing the NCI scan also means we
    don't accidentally pick up a stray dimensional string in col W
    as "NCI data".

    Defaults to MFRS so callers that haven't migrated (tests, etc.)
    keep their pre-Phase-5 behaviour.
    """
    if filing_standard == "mpers":
        return _MPERS_SOCIE_CY_COL
    return _SOCIE_TOTAL_COL if has_nci_data(ws, start_row, end_row) else _SOCIE_RETAINED_COL


def socie_total_column(filing_standard: str = "mfrs") -> int:
    """Return the SOCIE column that holds an aggregate Total (CY).

    Used by cross-checks that read equity-at-end or TCI — neither has
    a "per-component" variant, they're always totals across the
    dimensional axes.

    MFRS → col X (24) unconditionally (the pre-existing contract;
    equity/TCI tests rely on this).
    MPERS → col B (2) because MPERS SOCIE is a flat two-column layout
    with dimensional members on axis rows, not across 24 columns.
    """
    return _MPERS_SOCIE_CY_COL if filing_standard == "mpers" else _SOCIE_TOTAL_COL


def socie_py_column(filing_standard: str = "mfrs") -> int:
    """Return the CY→PY offset column for SOCIE reads.

    MPERS SOCIE stores PY in col C (3). MFRS SOCIE PY blocks are in
    separate row ranges (not a PY column on the same row), so this
    helper's MFRS branch is only ever needed for Phase-5 symmetry; the
    existing MFRS checks read PY via block-range lookups and don't
    call here.
    """
    if filing_standard == "mpers":
        return _MPERS_SOCIE_PY_COL
    # MFRS doesn't have a "PY column" on the same row — blocks are
    # row-range-separated. Returning col 25 is technically meaningless
    # but keeps the symmetric helper shape; callers on MFRS should
    # pass block ranges via find_value_in_block instead of reading
    # this column directly.
    return _SOCIE_TOTAL_COL + 1


def find_value_in_block(
    ws,
    label_substr: Union[str, Sequence[str]],
    col: int,
    start_row: int,
    end_row: int,
    wb: openpyxl.Workbook = None,
) -> Optional[float]:
    """Like find_value_by_label but restricted to a row range (for Group SOCIE blocks)."""
    candidates: list[str]
    if isinstance(label_substr, str):
        candidates = [label_substr]
    else:
        candidates = list(label_substr)

    for candidate in candidates:
        target = candidate.strip().lower()
        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=1)
            if cell.value is None:
                continue
            normalized = str(cell.value).strip().lstrip("*").strip().lower()
            if normalized == target or target in normalized or normalized in target:
                val_cell = ws.cell(row=r, column=col)
                raw = val_cell.value
                if raw is None:
                    continue
                if isinstance(raw, str) and raw.startswith("="):
                    if wb is None:
                        continue
                    from openpyxl.utils import get_column_letter
                    cell_ref = f"{get_column_letter(col)}{r}"
                    resolved = _resolve_cell_value(wb, ws.title, cell_ref)
                    if resolved is not None:
                        return resolved
                    continue
                try:
                    return float(raw)
                except (ValueError, TypeError):
                    continue
    return None


def open_workbook(path: str) -> openpyxl.Workbook:
    """Open a workbook in data_only=False mode (preserves formulas for evaluation)."""
    return openpyxl.load_workbook(path, data_only=False)


def find_sheet(wb: openpyxl.Workbook, *candidates: str):
    """Return the first sheet whose name matches one of the candidates.

    Matching is case-insensitive. Returns None if no match found.
    """
    for name in wb.sheetnames:
        for cand in candidates:
            if name.lower() == cand.lower():
                return wb[name]
    return None


def find_value_by_label(
    ws,
    label_substr: Union[str, Sequence[str]],
    col: int = 2,
    wb: openpyxl.Workbook = None,
) -> Optional[float]:
    """Scan column A for a row whose label contains `label_substr` (case-insensitive),
    then return the numeric value from the specified column on that row.

    When `wb` is provided, formula cells are evaluated recursively (required for
    real MBRS templates where total rows are always formulas). Without `wb`,
    formula cells return None.

    Args:
        ws: openpyxl worksheet.
        label_substr: substring to match in column A (case-insensitive, stripped of
            leading *). May also be a sequence of candidate labels — each is tried
            in order until one yields a value. Use this for fields whose wording
            differs across template variants (e.g. SOFP cash row is "cash and cash
            equivalents" in CuNonCu but "total cash and bank balances" in OrdOfLiq).
        col: which column to read the value from (2=B for CY, 3=C for PY, etc.).
        wb: the parent workbook — needed for formula evaluation.

    Returns:
        The float value, or None if no candidate label was found or cells were empty.
    """
    candidates: list[str]
    if isinstance(label_substr, str):
        candidates = [label_substr]
    else:
        candidates = list(label_substr)

    for candidate in candidates:
        target = candidate.strip().lower()

        # Collect all matching rows: exact matches first, then substring matches.
        # Multiple rows can share the same label (e.g. SOCF has a data-entry row
        # and a formula row both labelled "Cash and cash equivalents at end of period").
        # We try each match in priority order, skipping rows with no value.
        exact_rows: list[int] = []
        substr_rows: list[int] = []
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value is None:
                continue
            normalized = str(cell.value).strip().lstrip("*").strip().lower()
            if normalized == target:
                exact_rows.append(cell.row)
            elif target in normalized or normalized in target:
                substr_rows.append(cell.row)

        # Try exact matches first, then substring matches
        for match_row in exact_rows + substr_rows:
            val_cell = ws.cell(row=match_row, column=col)
            raw = val_cell.value
            if raw is None:
                continue

            # Formula cell — evaluate it if we have the workbook
            if isinstance(raw, str) and raw.startswith("="):
                if wb is None:
                    continue
                from openpyxl.utils import get_column_letter
                cell_ref = f"{get_column_letter(col)}{match_row}"
                resolved = _resolve_cell_value(wb, ws.title, cell_ref)
                if resolved is not None:
                    return resolved
                continue

            try:
                return float(raw)
            except (ValueError, TypeError):
                continue

    return None
