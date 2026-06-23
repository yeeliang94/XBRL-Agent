"""Persist per-run notes HTML payloads into the audit DB.

Step 6 of docs/Archive/PLAN-NOTES-RICH-EDITOR.md. The coordinator hands the
list of cells a notes agent wrote through this module after each
successful run; the result is the canonical payload the post-run
editor reads/writes and the Excel download overlays at stream time.

The helper is deliberately self-contained: it opens its own sqlite
connection, runs its work inside a single transaction, and closes
the connection on the way out. No event-loop coupling, no recorder
dependency — reliable to call from inside the coordinator's async
hot path or from a unit test.
"""
from __future__ import annotations

import logging
import shutil
import tempfile
from pathlib import Path
from typing import Iterable, Mapping

from db import repository as repo
from notes.html_to_text import html_to_excel_text
from notes.writer import evidence_col_for, truncate_with_footer

logger = logging.getLogger(__name__)


def persist_notes_cells(
    *,
    db_path: str,
    run_id: int,
    sheet_name: str,
    cells_written: Iterable[Mapping[str, object]],
) -> int:
    """Replace every notes_cells row for (run_id, sheet_name) with a
    fresh set from ``cells_written``.

    Each entry in ``cells_written`` is a dict-like with keys:

        sheet, row, label, html, evidence (optional), source_pages (optional list[int])

    Returns the number of rows upserted (0 if ``cells_written`` is
    empty — which is a legitimate outcome for a "no prose, numeric-only"
    sheet; the clobber still runs so prior content doesn't linger).

    Clobber-then-upsert is done in a single transaction so a crash
    mid-batch cannot leave the DB with a partial replacement.
    """
    cells_list = list(cells_written)

    # repo.db_session handles pragmas (foreign_keys, journal_mode,
    # busy_timeout), commits on success, and rolls back on exception —
    # so we don't repeat that boilerplate here. BEGIN IMMEDIATE is not
    # set explicitly because db_session's default journal_mode=WAL gives
    # us the right reader/writer behaviour, and the context manager's
    # commit is atomic at transaction end.
    with repo.db_session(db_path) as conn:
        repo.delete_notes_cells_for_run_sheet(
            conn, run_id=run_id, sheet=sheet_name,
        )
        # A rerun (regenerate) supersedes any prior reviewer pass over this
        # sheet: its tombstones reference the OLD extraction, so drop them or
        # the overlay would blank these freshly-written cells (peer-review HIGH).
        repo.clear_notes_tombstones_for_sheet(conn, run_id, sheet_name)
        for cell in cells_list:
            # Each cell dict must at minimum carry sheet/row/label/html.
            # `source_pages` is optional (defaults to []), `evidence`
            # may be None for rows the agent chose to leave uncited.
            source_pages = cell.get("source_pages") or []
            repo.upsert_notes_cell(
                conn,
                run_id=run_id,
                sheet=str(cell.get("sheet") or sheet_name),
                row=int(cell["row"]),
                label=str(cell["label"]),
                html=str(cell["html"]),
                evidence=(
                    str(cell["evidence"])
                    if cell.get("evidence") is not None
                    else None
                ),
                source_pages=[int(p) for p in source_pages],
            )
    return len(cells_list)


def persist_notes_review_inputs(
    *,
    db_path: str,
    run_id: int,
    sidecar_entries: Iterable[Mapping[str, object]],
    inventory: Iterable[Mapping[str, object]],
) -> tuple[int, int]:
    """Mirror the notes reviewer's detector inputs into the DB (Step 1).

    The structural detectors need each written cell's source note refs and the
    scout sub-note inventory. Both live only in on-disk run-dir files today
    (``*_payloads.json`` sidecars, ``infopack.json``), which a manual re-review
    on a fresh process can't rely on. We copy them into ``notes_cell_provenance``
    + ``run_notes_inventory`` once at extraction completion so the reviewer
    recomputes findings from the database (docs/PLAN.md Step 1/2).

    Best-effort by contract — the caller wraps this so a provenance-write
    failure never fails the run; the reviewer falls back to the sidecars when a
    row is absent. Returns ``(provenance_rows, inventory_rows)`` written.
    """
    prov = list(sidecar_entries)
    inv = list(inventory)
    with repo.db_session(db_path) as conn:
        for e in prov:
            row = e.get("row")
            sheet = e.get("sheet")
            if row is None or not sheet:
                continue
            refs = e.get("source_note_refs") or []
            repo.upsert_notes_provenance(
                conn,
                run_id=run_id,
                sheet=str(sheet),
                row=int(row),  # type: ignore[arg-type]
                row_label=str(e.get("row_label") or ""),
                source_note_refs=[str(r) for r in refs],  # type: ignore[union-attr]
                content_preview=(
                    str(e.get("content_preview"))
                    if e.get("content_preview") is not None
                    else None
                ),
            )
        for item in inv:
            note_num = item.get("note_num")
            if note_num is None:
                continue
            subs = item.get("subnote_refs") or []
            repo.upsert_notes_inventory(
                conn,
                run_id=run_id,
                note_num=int(note_num),  # type: ignore[arg-type]
                title=str(item.get("title") or ""),
                subnote_refs=[str(s) for s in subs],  # type: ignore[union-attr]
                page_lo=item.get("page_lo"),  # type: ignore[arg-type]
                page_hi=item.get("page_hi"),  # type: ignore[arg-type]
            )
    return len(prov), len(inv)


def overlay_notes_cells_into_workbook(
    *,
    xlsx_path: Path | str,
    run_id: int,
    db_path: str,
    filing_level: str = "company",
) -> Path:
    """Return a path to an xlsx whose notes sheets reflect the DB payload.

    The overlay is AUTHORITATIVE for the notes prose region, not merely
    additive: it writes each surviving ``notes_cells`` row (prose to col B,
    evidence to the filing-level evidence column) AND blanks every coordinate
    the reviewer emptied (``notes_cell_tombstones``, v25). Without the blanking
    pass a reviewer clear / move-out would leave the original prose written at
    merge time in the workbook, so the download reintroduced it as a duplicate.
    ``filing_level`` selects the evidence column (D=Company / F=Group) — the
    reviewer can update grounded evidence, so it must round-trip to the export.

    Returns ``xlsx_path`` unchanged only when there is nothing to apply (no
    notes_cells rows AND no tombstones). Otherwise copies the workbook into a
    temp file and returns that path; the caller cleans the temp file up after
    streaming. The flattened form is intentional: Excel has no HTML rendering
    layer, so tables become pipe-separated rows, lists gain `- ` / `1. `
    markers, and inline styling is dropped. The editor UI surfaces the rich
    HTML from the DB directly.
    """
    xlsx_path = Path(xlsx_path)

    # Same db_session-based path as persist_notes_cells — read-only,
    # but still benefits from the shared pragmas (WAL + busy_timeout).
    with repo.db_session(db_path) as conn:
        cells = repo.list_notes_cells_for_run(conn, run_id)
        tombstones = repo.fetch_notes_tombstones(conn, run_id)

    if not cells and not tombstones:
        return xlsx_path

    ev_col = evidence_col_for(filing_level)

    # Lazy import so the FastAPI layer doesn't force openpyxl into
    # every test harness that only touches the helper.
    import openpyxl

    # Temp file mirrors the original so the final streaming step
    # serves exactly what the caller requested. `delete=False` keeps
    # the file alive after the `with` block — FileResponse closes
    # the response asynchronously and an auto-deleted tempfile would
    # race the stream.
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, prefix="notes_overlay_",
    )
    tmp.close()
    tmp_path = Path(tmp.name)
    shutil.copy(str(xlsx_path), str(tmp_path))

    def _set_cell(ws, row: int, column: int, value) -> None:
        """Write a cell unless it carries a formula (never clobber one)."""
        ws_cell = ws.cell(row=row, column=column)
        if isinstance(ws_cell.value, str) and ws_cell.value.startswith("="):
            logger.warning(
                "overlay skipping formula cell at %s!R%dC%d (run_id=%s)",
                ws.title, row, column, run_id,
            )
            return
        ws_cell.value = value

    # Coords with live prose — a tombstone must NEVER blank one of these (a
    # rerun-after-review can leave a stale tombstone on a now-repopulated row;
    # this is the self-healing backstop to persist_notes_cells clearing them).
    live_coords = {(c.sheet, c.row) for c in cells}

    wb = openpyxl.load_workbook(str(tmp_path))
    try:
        for cell in cells:
            if cell.sheet not in wb.sheetnames:
                # Cell references a sheet the workbook doesn't carry.
                # Expected on partial merges (download for a face-only
                # run that still has stray notes rows in the DB), but
                # also signals real drift — e.g. template generator
                # renamed Notes-SummaryofAccPol. Log at WARNING so
                # operators can spot the mismatch; the download stays
                # stale vs the editor until either the DB rows are
                # cleaned up or the sheet name is reconciled.
                logger.warning(
                    "overlay: skipping cell (run_id=%s, row=%d) — "
                    "sheet %r not in workbook (sheets: %s)",
                    run_id, cell.row, cell.sheet,
                    ", ".join(wb.sheetnames),
                )
                continue
            ws = wb[cell.sheet]
            # Defence-in-depth: the writer already truncates HTML to
            # the 30k rendered cap before persisting, but a future
            # direct-PATCH path (plan Step 8) could write over-limit
            # content. Re-apply the writer's truncation before
            # flattening so Excel never silently clips a cell and
            # reviewers always see the truncation footer.
            truncated = truncate_with_footer(cell.html, cell.source_pages)
            _set_cell(ws, cell.row, 2, html_to_excel_text(truncated))
            # Evidence (audit trail) — reviewer writes update it, so refresh the
            # filing-level evidence column rather than leaving it stale / blank.
            _set_cell(ws, cell.row, ev_col, cell.evidence or None)

        # Blank every coordinate the reviewer emptied (clear / move-out /
        # authored-then-reverted) so a deletion actually reaches the export —
        # but never a coord that now carries live prose (stale tombstone after
        # a rerun-after-review).
        for sheet, row in tombstones:
            if sheet not in wb.sheetnames or (sheet, row) in live_coords:
                continue
            ws = wb[sheet]
            _set_cell(ws, row, 2, None)
            _set_cell(ws, row, ev_col, None)

        wb.save(str(tmp_path))
    finally:
        wb.close()
    return tmp_path


def overlay_numeric_facts_into_workbook(
    *,
    xlsx_path: Path | str,
    run_id: int,
    db_path: str,
) -> Path:
    """Return an xlsx whose NUMERIC notes cells reflect ``run_concept_facts``.

    Numeric notes (sheets 13/14) live in the canonical fact store — not
    ``notes_cells`` — so their post-run edits (``PATCH /facts``) never touch the
    agent-written workbook the merge sources from disk. This is the numeric
    counterpart of :func:`overlay_notes_cells_into_workbook`: it writes each
    numeric-note fact onto its target cell (resolved via ``concept_targets``,
    which carries the per-(scope, period) column) at download time.

    Returns ``xlsx_path`` unchanged when the run has no numeric-note facts.
    Formula cells are never overwritten — face/total formulas stay live so
    Excel recomputes them (export-keeps-live-formulas). Numeric notes carry no
    prose, so this never collides with the HTML overlay above (different
    sheets / different cells).
    """
    import sqlite3

    from notes_types import notes_template_ids

    xlsx_path = Path(xlsx_path)

    # Scope by the exact numeric-notes template_id set rather than a
    # '%-notes-%' LIKE, so a face slug containing "notes" can never be picked
    # up by this overlay (PLAN-notes-template-registry code-review hardening).
    notes_ids = sorted(notes_template_ids(numeric_only=True))
    if not notes_ids:
        return xlsx_path
    placeholders = ",".join("?" * len(notes_ids))

    with repo.db_session(db_path) as conn:
        conn.row_factory = sqlite3.Row
        facts = conn.execute(
            f"""
            SELECT t.target_sheet AS sheet, t.target_row AS row,
                   t.target_col AS col, f.value AS value
            FROM run_concept_facts f
            JOIN concept_nodes n ON n.concept_uuid = f.concept_uuid
            JOIN concept_targets t
              ON t.concept_uuid = f.concept_uuid
             AND t.entity_scope = f.entity_scope
             AND t.period = f.period
            WHERE f.run_id = ?
              AND n.kind = 'LEAF'
              AND n.template_id IN ({placeholders})
            """,
            (run_id, *notes_ids),
        ).fetchall()

    if not facts:
        return xlsx_path

    import openpyxl
    from openpyxl.utils import column_index_from_string

    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, prefix="numeric_overlay_",
    )
    tmp.close()
    tmp_path = Path(tmp.name)
    shutil.copy(str(xlsx_path), str(tmp_path))

    wb = openpyxl.load_workbook(str(tmp_path))
    changed = False
    try:
        for fr in facts:
            if fr["sheet"] not in wb.sheetnames:
                continue
            ws = wb[fr["sheet"]]
            cell = ws.cell(
                row=int(fr["row"]),
                column=column_index_from_string(fr["col"]),
            )
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Never clobber a live total formula.
                continue
            cell.value = fr["value"]
            changed = True
        if changed:
            wb.save(str(tmp_path))
    finally:
        wb.close()

    if not changed:
        tmp_path.unlink(missing_ok=True)
        return xlsx_path
    return tmp_path
