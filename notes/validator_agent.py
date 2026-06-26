"""Notes post-validator agent factory (Phase 5).

Runs once after the notes pass + merge, with access to the MERGED
workbook (so it can see Sheet 11 and Sheet 12 together). Its job is to
detect and resolve cross-sheet duplication between "Material Accounting
Policies" (Sheet 11) and "List of Notes" (Sheet 12).

Tools:
    - view_pdf_pages — same tool the face agents use
    - read_cell       — read current contents of a cell in the merged wb
    - rewrite_cell    — overwrite a data-entry cell (or clear it)
    - flag_duplication — record an agent note for the audit log

The detection of candidate duplicates is done in Python (Step 5.3 / 5.4)
and the list is injected into the agent prompt. The agent applies the
"Material Accounting Policies → Sheet 11, otherwise → Sheet 12" rule
itself (per PLAN D3) using the PDF as the source of truth.
"""
from __future__ import annotations

import json
import logging
import os
import threading
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, List, Optional, Union

import openpyxl
from pydantic_ai import Agent, RunContext
from pydantic_ai.messages import BinaryContent
from pydantic_ai.models import Model
from model_settings import build_model_settings

from notes.writer import payload_sidecar_path
from tools.pdf_viewer import count_pdf_pages, render_pages_to_png_bytes
# Promoted to utils/workbook_io.py (item 8) so every workbook saver shares
# one atomic mechanism. Re-exported under the original private name to keep
# this module's import/test contract (tests/test_notes_validator_agent.py)
# untouched.
from utils.workbook_io import atomic_save_workbook as _atomic_save_workbook

logger = logging.getLogger(__name__)

# The structural detectors + their sidecar / DB-provenance / scout-inventory
# loaders moved to notes/detectors.py (the neutral home) so the live notes
# REVIEWER no longer imports from this dead-but-green module — keeping it
# working after the eventual validator deletion (docs/PLAN.md Step 1).
# Re-exported here so this module's own remaining code + its pinning tests
# (tests/test_notes_validator_agent.py) keep their import surface unchanged.
from notes.detectors import (  # noqa: F401
    _render_single_page,
    detect_cross_sheet_duplicates_by_ref,
    detect_cross_sheet_overlap_candidates,
    detect_same_sheet_row_collisions,
    detect_subnote_coverage_gaps,
    detect_title_format_issues,
    inventory_coverage_gaps,
    load_inventory_from_db,
    load_provenance_entries,
    load_sidecar_entries,
)

# Defense-in-depth: the validator is chartered to reconcile cross-sheet
# duplication between Sheet 11 (policies) and Sheet 12 (list of notes).
# It must never touch face-statement sheets or other notes sheets —
# an agent confusion could otherwise overwrite numeric cells with prose.
# Prompt says "never write prose to Sheet 13 or 14"; this allowlist
# enforces it in code so a mis-targeted rewrite fails loudly.
_REWRITE_ALLOWED_SHEETS: frozenset[str] = frozenset({
    "Notes-SummaryofAccPol",
    "Notes-Listofnotes",
})


class NotesValidatorAgentDeps:
    """Dependencies carried through the post-validator's tool calls."""

    def __init__(
        self,
        merged_workbook_path: str,
        pdf_path: str,
        sidecar_paths: List[str],
        filing_level: str,
        filing_standard: str,
        output_dir: str,
        model: Any,
    ):
        self.merged_workbook_path = merged_workbook_path
        self.pdf_path = pdf_path
        self.sidecar_paths = list(sidecar_paths)
        self.filing_level = filing_level
        self.filing_standard = filing_standard
        self.output_dir = output_dir
        self.model = model
        self.pdf_page_count = 0
        # Pages the agent has actually rendered via view_pdf_pages this run.
        # The notes-reviewer write guard requires every fix to cite a page in
        # this set — so an authored/edited cell can never claim grounding on a
        # page the agent never looked at (docs/PLAN.md Step 4 + Step 6).
        self.viewed_pages: set[int] = set()
        # Agent-recorded corrections + rationales so operators can audit
        # what the validator chose to do. One file lands at
        # `notes_validator_log.json` next to the merged workbook.
        self.correction_log: list[dict] = []
        self.writes_performed = 0
        # Serialises every workbook load/save across the validator's tool
        # calls. pydantic-ai runs batched tool calls in parallel worker
        # threads (default parallel_execution_mode); without this lock a
        # `read_cell` load_workbook could hit `merged_workbook_path` mid-way
        # through a `rewrite_cell` non-atomic `wb.save`, reading a truncated
        # zip → EOFError. The validator only ever touches this one workbook,
        # so a single lock is sufficient. See gotcha #6 / Windows race
        # (2026-05-29). Pinned by tests/test_notes_validator_agent.py.
        self.io_lock = threading.Lock()


_SHEET_11_TAB = "Notes-SummaryofAccPol"
_SHEET_12_TAB = "Notes-Listofnotes"


def build_validator_prompt_body(
    duplicates: list[dict],
    overlap_candidates: list[dict],
    filing_level: str,
    filing_standard: str,
) -> str:
    """Render the dynamic portion of the validator prompt — the list of
    candidates the agent must resolve. Surfaces the literal worksheet tab
    names (`Notes-SummaryofAccPol`, `Notes-Listofnotes`) so the agent's
    first `read_cell` / `rewrite_cell` call uses the real sheet names
    `_rewrite_cell_impl` will accept (peer-review Codex P2)."""
    lines: list[str] = [
        "=== CROSS-SHEET DUPLICATE CANDIDATES ===",
        "",
        f"filing_level: {filing_level}",
        f"filing_standard: {filing_standard}",
        "",
        "Worksheet tab names to pass to read_cell / rewrite_cell:",
        f"  • Sheet 11 → {_SHEET_11_TAB!r}",
        f"  • Sheet 12 → {_SHEET_12_TAB!r}",
        "",
    ]
    if not duplicates and not overlap_candidates:
        lines.append(
            "No candidates — no ref-based duplicates and no overlap fallback "
            "hits. You may exit immediately with a short confirmation."
        )
        return "\n".join(lines)

    def _fmt_pages(entry: dict) -> str:
        pages = entry.get("source_pages") or []
        if not pages:
            return "pages: (none — use scout hints or scan the PDF)"
        return f"pages: {sorted(set(pages))}"

    if duplicates:
        lines.append(
            "REF-BASED DUPLICATES (same source_note_refs value on both sheets):"
        )
        for d in duplicates:
            a = d["sheet_11"]
            b = d["sheet_12"]
            lines.append(
                f"  • Note ref {d['note_ref']!r}: "
                f"Sheet 11 ({_SHEET_11_TAB!r}) row {a.get('row')} vs "
                f"Sheet 12 ({_SHEET_12_TAB!r}) row {b.get('row')}"
            )
            lines.append(f"    Sheet 11 {_fmt_pages(a)}")
            lines.append(
                f"    Sheet 11 preview: {a.get('content_preview', '')!r}"
            )
            lines.append(f"    Sheet 12 {_fmt_pages(b)}")
            lines.append(
                f"    Sheet 12 preview: {b.get('content_preview', '')!r}"
            )
        lines.append("")

    if overlap_candidates:
        lines.append(
            "OVERLAP FALLBACK (content-similarity candidates without matching "
            "refs — treat as probable duplicates and verify against the PDF):"
        )
        for c in overlap_candidates:
            a = c["sheet_11"]
            b = c["sheet_12"]
            lines.append(
                f"  • Score {c['score']}: "
                f"Sheet 11 ({_SHEET_11_TAB!r}) row {a.get('row')} vs "
                f"Sheet 12 ({_SHEET_12_TAB!r}) row {b.get('row')}"
            )
            lines.append(f"    Sheet 11 {_fmt_pages(a)}")
            lines.append(
                f"    Sheet 11 preview: {a.get('content_preview', '')!r}"
            )
            lines.append(f"    Sheet 12 {_fmt_pages(b)}")
            lines.append(
                f"    Sheet 12 preview: {b.get('content_preview', '')!r}"
            )
        lines.append("")

    lines.append(
        "For each candidate, view the relevant PDF pages and apply the "
        "heading rule: headings like 'Material Accounting Policies', "
        "'Significant Accounting Policies', or 'Summary of Material "
        "Accounting Policies' belong on Sheet 11; everything else belongs "
        "on Sheet 12. Rewrite the wrong cell (set content to empty string "
        f"to delete) — pass sheet={_SHEET_11_TAB!r} or {_SHEET_12_TAB!r} "
        "verbatim — then call `flag_duplication` to record your reasoning."
    )
    return "\n".join(lines)


def build_structural_findings_block(
    row_collisions: list[dict],
    subnote_gaps: list[dict],
) -> str:
    """Render the SAME-SHEET ROW COLLISIONS + SUB-NOTE COVERAGE GAPS prompt
    blocks. Returns "" when both are empty (nothing appended).

    Kept as a pure helper (rather than inlined in the factory) so the prompt
    contract is pinnable without standing up an Agent — the same reason the
    duplicate/overlap rendering lives in :func:`build_validator_prompt_body`.
    """
    out = ""
    if row_collisions:
        out += (
            "\n\n=== SAME-SHEET ROW COLLISIONS (one Sheet-12 row holds content "
            "from >1 unrelated top-level note) ===\n"
            "A specific disclosure row received prose from multiple top-level "
            "notes. The writer concatenates payloads that land on the same row; "
            "for the catch-all 'Disclosure of other notes to accounts' that is "
            "intended, but on a SPECIFIC row it usually means two different "
            "notes were force-matched to one XBRL concept. For each, view the "
            "cited pages, decide which note legitimately owns the row, and clear "
            "the mis-placed content (rewrite_cell content=\"\") ONLY if there is "
            "a clearly correct owner. If both genuinely belong or you are "
            "unsure, flag_duplication decision=\"no_action\" with your reasoning "
            "— surfacing it for a human is better than deleting valid content.\n"
        )
        for c in row_collisions:
            out += (
                f"  • Sheet 12 row {c['row']} {c['row_label']!r}: notes "
                f"{c['note_nums']} (refs {c['source_note_refs']}) — preview: "
                f"{c['content_preview']!r}\n"
            )
    if subnote_gaps:
        out += (
            "\n\n=== SUB-NOTE COVERAGE GAPS (scout saw sub-sections a note only "
            "PARTLY covered) ===\n"
            "For each note, scout discovered these sub-sections but the written "
            "content cited only some of them. View the note's pages and judge "
            "whether each MISSING sub-section is a real omission (e.g. a "
            "lettered (a)/(b) policy block dropped) or simply folded into the "
            "prose / non-applicable. Report genuine omissions; never fabricate.\n"
        )
        for g in subnote_gaps:
            out += (
                f"  • Note {g['note_num']}: cited {g['cited_subnote_refs']}, "
                f"MISSING {g['missing_subnote_refs']} (scout saw "
                f"{g['all_subnote_refs']})\n"
            )
    return out


# ---------------------------------------------------------------------------
# Agent factory
# ---------------------------------------------------------------------------


_PROMPT_PATH = Path(__file__).resolve().parent.parent / "prompts" / "notes_validator.md"


def create_notes_validator_agent(
    merged_workbook_path: str,
    pdf_path: str,
    sidecar_paths: List[str],
    filing_level: str,
    filing_standard: str,
    model: Union[str, Model],
    output_dir: str,
    inventory_note_nums: Optional[List[int]] = None,
    inventory_subnotes: Optional[dict] = None,
    run_id: Optional[int] = None,
    db_path: Optional[str] = None,
) -> tuple[Agent[NotesValidatorAgentDeps, str], NotesValidatorAgentDeps, dict]:
    """Build the notes post-validator agent.

    Returns (agent, deps, context) where `context` is a dict describing
    what the detectors found before the run — useful for test assertions
    and for the coordinator to short-circuit when there are no candidates.

    ``inventory_note_nums`` (N3 Stage 1) is the scout's notes inventory by
    integer note_num. When supplied, the deterministic
    :func:`inventory_coverage_gaps` reports which inventory notes have NO
    content on ANY sheet, surfaced in ``context['coverage_gaps']`` and the
    prompt so the validator agent investigates them (gotcha-#14-safe: code
    reports gaps by note_num, the AGENT judges content adequacy).

    ``inventory_subnotes`` maps a top-level note_num → scout's discovered
    sub-reference strings under it. When supplied,
    :func:`detect_subnote_coverage_gaps` reports notes that were only PARTLY
    covered at sub-reference granularity (e.g. a leases policy citing ``3.3``
    + ``(b)`` but dropping ``(a)``), surfaced in ``context['subnote_gaps']``.
    Independently, :func:`detect_same_sheet_row_collisions` reports Sheet-12
    rows that received content from ≥2 distinct top-level notes (a non-catch-all
    pile-up), surfaced in ``context['row_collisions']``. Both are advisory
    candidates the validator agent investigates — gotcha-#14-safe.
    """
    deps = NotesValidatorAgentDeps(
        merged_workbook_path=merged_workbook_path,
        pdf_path=pdf_path,
        sidecar_paths=sidecar_paths,
        filing_level=filing_level,
        filing_standard=filing_standard,
        output_dir=output_dir,
        model=model,
    )

    # Prefer durable DB provenance (Step 2) so a manual re-review recomputes
    # findings from the database; fall back to the on-disk sidecars for legacy
    # runs (or when the provenance write was skipped).
    entries: list[dict] = []
    if run_id is not None and db_path:
        entries = load_provenance_entries(run_id, db_path)
    if not entries:
        entries = load_sidecar_entries(sidecar_paths)
    # Same DB-first preference for the scout inventory: when the caller didn't
    # pass it explicitly, hydrate from run_notes_inventory.
    if run_id is not None and db_path and (
        inventory_note_nums is None or inventory_subnotes is None
    ):
        _db_nums, _db_subs = load_inventory_from_db(run_id, db_path)
        if inventory_note_nums is None and _db_nums:
            inventory_note_nums = _db_nums
        if inventory_subnotes is None and _db_subs:
            inventory_subnotes = _db_subs
    duplicates = detect_cross_sheet_duplicates_by_ref(entries)
    overlap = detect_cross_sheet_overlap_candidates(entries)
    # N3 Stage 1: which inventory notes never got written anywhere.
    coverage_gaps = inventory_coverage_gaps(inventory_note_nums or [], entries)
    # Same-sheet pile-ups + partial sub-note coverage (advisory candidates).
    row_collisions = detect_same_sheet_row_collisions(entries)
    subnote_gaps = detect_subnote_coverage_gaps(inventory_subnotes or {}, entries)

    base_prompt = _PROMPT_PATH.read_text(encoding="utf-8").strip()
    dynamic_prompt = build_validator_prompt_body(
        duplicates, overlap, filing_level, filing_standard,
    )
    if coverage_gaps:
        dynamic_prompt += (
            "\n\n=== COVERAGE GAPS (scout saw these notes; no content was "
            "written for them on ANY sheet) ===\n"
            f"Inventory note numbers with no content: {coverage_gaps}.\n"
            "For each, view the cited PDF pages and judge: is this note "
            "genuinely absent from the filing (fine), or was it missed? Report "
            "missed disclosures — do NOT fabricate content."
        )
    dynamic_prompt += build_structural_findings_block(row_collisions, subnote_gaps)
    system_prompt = f"{base_prompt}\n\n{dynamic_prompt}"

    agent = Agent(
        model,
        deps_type=NotesValidatorAgentDeps,
        system_prompt=system_prompt,
        # Phase 2: provider-correct prompt caching of the static system prompt.
        model_settings=build_model_settings(
            model, cache_key="xbrl-notes-validator"
        ),
    )

    @agent.tool
    def view_pdf_pages(
        ctx: RunContext[NotesValidatorAgentDeps], pages: List[int],
    ) -> List[Union[str, BinaryContent]]:
        """View specific PDF pages as images."""
        ctx.deps.pdf_page_count = count_pdf_pages(ctx.deps.pdf_path)
        total_pages = ctx.deps.pdf_page_count
        requested = [p for p in pages if isinstance(p, int)]
        invalid = sorted({p for p in requested if p < 1 or p > total_pages})
        render_pages = sorted(set(p for p in requested if p not in invalid))

        results: List[Union[str, BinaryContent]] = []
        if invalid:
            results.append(
                f"Skipped invalid page(s) {invalid}. Valid range is 1-{total_pages}."
            )
        if not render_pages:
            results.append("No pages were rendered from this request.")
            return results

        rendered: dict[int, bytes] = {}
        with ThreadPoolExecutor(max_workers=min(len(render_pages), 8)) as pool:
            futures = {
                pool.submit(_render_single_page, ctx.deps.pdf_path, p): p
                for p in render_pages
            }
            for future in futures:
                page_num, png_bytes = future.result()
                rendered[page_num] = png_bytes

        for p in sorted(rendered):
            results.append(f"=== Page {p} ===")
            results.append(BinaryContent(data=rendered[p], media_type="image/png"))
        # Record every successfully-rendered page so the write guard can verify
        # an agent fix is grounded in a page it actually viewed (Step 4/6).
        ctx.deps.viewed_pages.update(rendered.keys())
        return results

    @agent.tool
    def read_cell(
        ctx: RunContext[NotesValidatorAgentDeps], sheet: str, row: int, col: int,
    ) -> str:
        """Read the current value of a merged-workbook cell."""
        # Serialise against concurrent rewrite_cell saves (see io_lock note).
        with ctx.deps.io_lock:
            wb = openpyxl.load_workbook(ctx.deps.merged_workbook_path)
        try:
            if sheet not in wb.sheetnames:
                return f"Sheet {sheet!r} not found. Available: {wb.sheetnames}"
            ws = wb[sheet]
            value = ws.cell(row=row, column=col).value
            return json.dumps({"sheet": sheet, "row": row, "col": col,
                               "value": value})
        finally:
            wb.close()

    @agent.tool
    def rewrite_cell(
        ctx: RunContext[NotesValidatorAgentDeps],
        sheet: str,
        row: int,
        col: int,
        content: str,
        evidence: Optional[str] = None,
    ) -> str:
        """Overwrite a data-entry cell in the merged workbook.

        Pass `content=""` to delete. Refuses to write to formula cells
        — those are pure template scaffolding and should never be
        touched by the validator.
        """
        return _rewrite_cell_impl(
            merged_workbook_path=ctx.deps.merged_workbook_path,
            filing_level=ctx.deps.filing_level,
            sheet=sheet,
            row=row,
            col=col,
            content=content,
            evidence=evidence,
            deps=ctx.deps,
        )

    @agent.tool
    def flag_duplication(
        ctx: RunContext[NotesValidatorAgentDeps],
        note_ref: str,
        decision: str,
        rationale: str,
    ) -> str:
        """Record that a cross-sheet duplicate has been resolved.

        ``decision`` should be one of: "kept_on_sheet_11", "kept_on_sheet_12",
        "deleted_from_both", "no_action". ``rationale`` is the agent's
        free-text explanation of its reasoning — stored for audit.
        """
        ctx.deps.correction_log.append({
            "note_ref": note_ref,
            "decision": decision,
            "rationale": rationale,
        })
        return f"Recorded decision for note {note_ref}: {decision}"

    context = {
        "duplicates": duplicates,
        "overlap_candidates": overlap,
        "entry_count": len(entries),
        # N3 Stage 1: inventory notes with no content anywhere (advisory).
        "coverage_gaps": coverage_gaps,
        # Sheet-12 rows holding ≥2 distinct top-level notes (advisory).
        "row_collisions": row_collisions,
        # Notes only partly covered at sub-reference granularity (advisory).
        "subnote_gaps": subnote_gaps,
    }
    return agent, deps, context




def _rewrite_cell_impl(
    merged_workbook_path: str,
    filing_level: str,
    sheet: str,
    row: int,
    col: int,
    content: str,
    evidence: Optional[str],
    deps: NotesValidatorAgentDeps,
) -> str:
    """Inner implementation so tests can exercise the tool without standing
    up a full RunContext. Also clears the evidence column when content
    becomes empty, so a deletion leaves no orphan citation behind."""
    from notes.writer import evidence_col_for

    # Defense-in-depth: refuse any sheet outside the validator's charter.
    # The prompt restricts the agent to Sheet 11 + Sheet 12, but a
    # hallucinated sheet name should fail loudly rather than corrupt
    # a face statement or a numeric notes sheet.
    if sheet not in _REWRITE_ALLOWED_SHEETS:
        return (
            f"Refusing to rewrite out-of-scope sheet {sheet!r} — "
            f"validator is restricted to "
            f"{sorted(_REWRITE_ALLOWED_SHEETS)}"
        )

    # Hold the run-wide io_lock across the entire read-modify-write so a
    # concurrent read_cell / rewrite_cell on another worker thread can't
    # observe the workbook mid-save (truncated zip → EOFError). See the
    # io_lock note on NotesValidatorAgentDeps.
    with deps.io_lock:
        wb = openpyxl.load_workbook(merged_workbook_path)
        try:
            if sheet not in wb.sheetnames:
                return f"Sheet {sheet!r} not found"
            ws = wb[sheet]
            target = ws.cell(row=row, column=col)
            if isinstance(target.value, str) and target.value.startswith("="):
                return (
                    f"Refusing to overwrite formula cell "
                    f"{target.coordinate} (formula: {target.value!r})"
                )
            # Empty string => explicit deletion (plan Step 5.2).
            target.value = content if content else None

            # Sync the evidence column. On deletion we always clear so the
            # audit trail can't keep a citation pointing at nothing. On an
            # update we apply the supplied evidence (if any) or leave the
            # existing value untouched.
            ev_col = evidence_col_for(filing_level)
            if col == 2:  # only sync when we're rewriting the primary value col
                ev_cell = ws.cell(row=row, column=ev_col)
                if not content:
                    ev_cell.value = None
                elif evidence is not None:
                    ev_cell.value = evidence
            _atomic_save_workbook(wb, merged_workbook_path)
            deps.writes_performed += 1
        finally:
            wb.close()

    deps.correction_log.append({
        "operation": "rewrite_cell",
        "sheet": sheet,
        "row": row,
        "col": col,
        "deleted": not content,
        "evidence": evidence,
    })
    action = "cleared" if not content else "wrote"
    return f"OK — {action} {sheet}!R{row}C{col}"
