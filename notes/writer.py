"""Writes a list of NotesPayload entries into an MBRS notes workbook.

Column layout (PLAN Section 2 #6):

  Company filing (4-col template):
    A=label, B=value, C=prior-year-value, D=source/evidence
    - Prose rows  → content to B, evidence to D (C left empty).
    - Numeric rows → values to B (CY) and C (PY), evidence to D.

  Group filing (6-col template):
    A=label, B=Group-CY, C=Group-PY, D=Company-CY, E=Company-PY, F=source
    - Prose rows  → content to B only (C/D/E empty), evidence to F.
    - Numeric rows → 4 values to B/C/D/E per role, evidence to F.

Char-limit guard: Excel caps cells at 32,767 chars. We truncate well below
that and append a footer pointing at the source pages.

Row resolution is fuzzy, label-based (same pattern as tools/fill_workbook.py):
normalise both sides (strip leading '*', lowercase) and exact-match first,
then SequenceMatcher fallback at ~0.7 similarity.
"""
from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

from notes.labels import normalize_label
from notes.html_to_text import (
    html_to_excel_text,
    rendered_length,
    truncate_html_to_rendered_length,
)
from notes.html_sanitize import sanitize_notes_html

import openpyxl

from notes.payload import NotesPayload

logger = logging.getLogger(__name__)


# Excel's hard limit is 32,767 chars; we keep ~2K of headroom for a footer.
CELL_CHAR_LIMIT = 30_000

# Fuzzy-match threshold for label resolution.
#
# Raised from 0.70 to 0.85 after a production run silently force-inserted
# payloads into wrong rows (e.g. "Disclosure of taxation" scored 0.78
# against "Disclosure of bonds" and landed income-tax prose in the bonds
# row). Legitimate near-matches from real traces all score >=0.90
# (leading `*`, case/whitespace, minor typos); the 0.70–0.85 band was
# almost entirely wrong-label cases. Kept equal to `BORDERLINE_FUZZY_SCORE`
# below so anything the writer accepts is, by definition, not borderline
# — the coordinator's "borderline warning" logic is dormant but harmless
# and the two constants cannot drift.
#
# Regression cases that protect this value live in
# tests/test_notes_writer_fuzzy_threshold.py — lowering the floor re-
# admits them.
_FUZZY_THRESHOLD = 0.85


@dataclass
class NotesWriteResult:
    success: bool
    rows_written: int = 0
    output_path: str = ""
    errors: list[str] = field(default_factory=list)
    # Labels resolved via fuzzy fallback rather than exact match. Populated
    # only for non-exact hits; each entry is (requested, chosen, score).
    # Surfaced so operators can review borderline matches that would
    # otherwise silently misroute payloads.
    fuzzy_matches: list[tuple[str, str, float]] = field(default_factory=list)
    # Human-readable notes about HTML fragments the sanitiser removed
    # before persisting (script tags, event handlers, disallowed tags).
    # Surfaced into the coordinator's warnings pipeline so dropped
    # content does not go unseen.
    sanitizer_warnings: list[str] = field(default_factory=list)
    # Per-cell HTML manifest (Step 6 of the notes rich-editor plan).
    # One entry per prose cell the writer landed, in template-row
    # order. Each dict carries the keys the persistence helper +
    # download overlay both expect: sheet, row, label, html,
    # evidence, source_pages. Numeric-only rows are omitted — the
    # editor focuses on prose.
    cells_written: list[dict] = field(default_factory=list)


# Scores below this threshold are logged at WARNING as "borderline". The
# writer still accepts them (the row was resolved), but the operator gets
# a visible nudge to review the match in the output. Public so the
# coordinator's warning-builder imports rather than duplicates this value.
BORDERLINE_FUZZY_SCORE = 0.85


def write_notes_workbook(
    template_path: str,
    payloads: list[NotesPayload],
    output_path: str,
    filing_level: str,
    sheet_name: str,
) -> NotesWriteResult:
    """Write NotesPayload entries to the given sheet of a notes template.

    Success is defined strictly as ``rows_written > 0``. Callers must
    pre-check for empty payloads and skip the write themselves if they
    want a no-op success — a zero-row write is treated as a failure so
    Sheet-12's "no payloads = all sub-agents lost coverage" case can't
    ship a silent green tick on an untouched template.
    """
    tpl = Path(template_path)
    if not tpl.exists():
        return NotesWriteResult(
            success=False,
            output_path="",
            errors=[f"Template not found: {template_path}"],
        )

    wb = openpyxl.load_workbook(template_path)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return NotesWriteResult(
            success=False,
            output_path="",
            errors=[
                f"Sheet '{sheet_name}' not found in template "
                f"(have: {wb.sheetnames})"
            ],
        )

    ws = wb[sheet_name]
    label_index = _build_label_index(ws)

    # Concatenate duplicate labels so Sheet-12 "Disclosure of other notes"
    # can collect multiple unmatched notes into a single cell.
    rows_consumed: dict[int, list[NotesPayload]] = {}
    errors: list[str] = []
    fuzzy_matches: list[tuple[str, str, float]] = []
    sanitizer_warnings: list[str] = []

    # Sanitise every prose payload before routing. Runs once per payload
    # (before combining) so warnings attribute to the right source and
    # `_write_row`'s truncation operates on clean HTML. Numeric-only
    # payloads skip sanitisation — they hold no HTML.
    payloads = [
        _sanitize_payload(p, sanitizer_warnings) for p in payloads
    ]

    for payload in payloads:
        resolution = _resolve_row(label_index, payload.chosen_row_label)
        if resolution is None:
            errors.append(
                f"No matching row for label '{payload.chosen_row_label}' in sheet '{sheet_name}'"
            )
            continue
        row, chosen_label, score = resolution
        if score < 1.0:
            fuzzy_matches.append((payload.chosen_row_label, chosen_label, score))
            level = logging.WARNING if score < BORDERLINE_FUZZY_SCORE else logging.DEBUG
            logger.log(
                level,
                "Fuzzy row match in %s: %r -> %r (score %.2f)",
                sheet_name, payload.chosen_row_label, chosen_label, score,
            )
        rows_consumed.setdefault(row, []).append(payload)

    evidence_col = _evidence_col(filing_level)

    rows_written = 0
    sidecar_entries: list[dict] = []
    cells_written: list[dict] = []
    # Reverse map row→template label so cells_written can carry the
    # template's verbatim col-A text (what the editor shows as the row
    # header) rather than the agent's possibly-fuzzy request.
    row_to_label: dict[int, str] = {e.row: e.original for e in label_index}
    for row, row_payloads in rows_consumed.items():
        combined = _combine_payloads(row_payloads)
        if _write_row(ws, row, combined, filing_level, evidence_col, errors):
            rows_written += 1
            # Peer-review I-6: the source_pages aggregation was
            # previously implemented twice in this block (once for
            # cells_written, once for sidecar_entries) with subtly
            # different variable names. Single helper keeps them
            # always-in-sync.
            aggregated_pages = _collect_unique_pages(row_payloads)
            # Record per-cell HTML for Step 6 persistence. Only prose
            # rows contribute — numeric rows have no HTML payload.
            if not combined.numeric_values:
                # Re-run the same HTML-aware truncation the writer used
                # on the cell so the DB row matches what the xlsx
                # rendered-from (pre-flatten). Keeps the editor's
                # payload identical in form to what shipped to Excel.
                html_for_db = truncate_with_footer(
                    combined.content, combined.source_pages,
                )
                cells_written.append({
                    "sheet": sheet_name,
                    "row": row,
                    "label": row_to_label.get(row, combined.chosen_row_label),
                    "html": html_for_db,
                    "evidence": combined.evidence or None,
                    "source_pages": aggregated_pages,
                })
            # Phase 4.3: collect per-cell provenance for the post-validator.
            # One entry per written row — combining all note-refs from the
            # contributing payloads so row-112 catch-alls retain the full
            # list. Preview is a short snippet of content (or the numeric
            # key names) to help operators cross-reference the sidecar
            # against the filled xlsx without loading the workbook.
            refs: list[str] = []
            seen: set[str] = set()
            for p in row_payloads:
                for r in p.source_note_refs:
                    if r not in seen:
                        seen.add(r)
                        refs.append(r)
            if combined.numeric_values:
                preview = f"numeric[{','.join(sorted(combined.numeric_values))}]"
            else:
                preview = combined.content[:120]
            sidecar_entries.append({
                "sheet": sheet_name,
                "row": row,
                "col": 2,  # all prose + all numeric rows start writing at col B
                "source_note_refs": refs,
                "source_pages": aggregated_pages,
                "content_preview": preview,
            })

    try:
        wb.save(output_path)
    finally:
        wb.close()

    # Phase 4.3: sidecar JSON for the Phase 5 post-validator. One file
    # per template output, same basename + "_payloads.json" — no shared
    # global file so the notes coordinator's `asyncio.gather` fan-out
    # can't race. Writing synchronously here is safe: each notes agent
    # owns exactly one template path.
    try:
        sidecar_path = payload_sidecar_path(output_path)
        sidecar_path.write_text(
            json.dumps(sidecar_entries, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    except OSError:
        # Best-effort — a filesystem failure here must not fail the
        # primary workbook write. The post-validator falls back to the
        # content-overlap heuristic when the sidecar is missing.
        logger.warning("Failed to write sidecar next to %s", output_path, exc_info=True)

    # Zero-row writes are failures — see docstring. Callers who want a
    # no-op success must short-circuit before calling this function.
    success = rows_written > 0
    return NotesWriteResult(
        success=success,
        rows_written=rows_written,
        output_path=output_path,
        errors=errors,
        fuzzy_matches=fuzzy_matches,
        sanitizer_warnings=sanitizer_warnings,
        cells_written=cells_written,
    )


def _inject_headings(payload: NotesPayload) -> NotesPayload:
    """Prepend `<h3>` lines to `payload.content` from parent_note / sub_note.

    Owns the "every note cell starts with its heading" contract. Running
    here (before sanitise + truncate) makes the headings impossible to drop:

      1. Heading injection happens first — the LLM cannot forget to prepend.
      2. The sanitiser treats `<h3>` as whitelisted, so the markup survives.
      3. Truncation operates on the combined text; if the body is huge the
         truncation footer will still sit AFTER the headings.

    Numeric-only payloads (no prose content) return unchanged — those cells
    hold a number, not prose, so there's no body to prepend to.
    """
    if not payload.content:
        return payload
    parent = payload.parent_note
    if parent is None:
        # Shouldn't happen on agent-parsed payloads (dataclass validator
        # rejects non-empty payloads without parent_note), but legacy test
        # constructions may still reach here — return unchanged in that
        # case rather than crash.
        return payload
    headings_html = f"<h3>{parent['number']} {parent['title']}</h3>"
    sub = payload.sub_note
    if sub is not None:
        headings_html += f"<h3>{sub['number']} {sub['title']}</h3>"
    # Headings always prepend. If the agent accidentally duplicated a
    # heading inside `content` (in defiance of the prompt's "do not
    # prepend <h3> manually" rule), the cell will carry both. The
    # prompt-contract test in Phase 3 guards the prompt side; at the
    # writer we keep the contract simple (always prepend, no dedup).
    return NotesPayload(
        chosen_row_label=payload.chosen_row_label,
        content=headings_html + payload.content,
        evidence=payload.evidence,
        source_pages=list(payload.source_pages),
        sub_agent_id=payload.sub_agent_id,
        numeric_values=payload.numeric_values,
        note_num=payload.note_num,
        source_note_refs=list(payload.source_note_refs),
        parent_note=payload.parent_note,
        sub_note=payload.sub_note,
    )


def _sanitize_payload(
    payload: NotesPayload, warnings: list[str],
) -> NotesPayload:
    """Return a payload with its `content` sanitised.

    Numeric-only payloads (no prose content) are returned unchanged —
    the sanitiser only applies to the HTML prose branch. Sanitiser
    warnings are prefixed with the target row label so the caller can
    attribute them when multiple payloads share the same batch.

    The heading prepend (`_inject_headings`) is applied first so the
    sanitiser sees `<h3>`s as part of the content and doesn't mistake
    them for untrusted markup injected mid-run.
    """
    payload = _inject_headings(payload)
    if not payload.content:
        return payload
    cleaned, sub_warnings = sanitize_notes_html(payload.content)
    if sub_warnings:
        for w in sub_warnings:
            warnings.append(f"{payload.chosen_row_label}: {w}")
    if cleaned == payload.content:
        return payload
    # Construct a replacement dataclass. Re-creating the NotesPayload
    # rather than mutating keeps the input list unaffected for any
    # caller that retained a reference (e.g. the sub-coordinator).
    return NotesPayload(
        chosen_row_label=payload.chosen_row_label,
        content=cleaned,
        evidence=payload.evidence,
        source_pages=list(payload.source_pages),
        sub_agent_id=payload.sub_agent_id,
        numeric_values=payload.numeric_values,
        note_num=payload.note_num,
        source_note_refs=list(payload.source_note_refs),
        # Heading hierarchy must survive the sanitise clone — without
        # this a caller that inspects parent_note on the returned object
        # would see None.
        parent_note=payload.parent_note,
        sub_note=payload.sub_note,
    )


def payload_sidecar_path(xlsx_output_path: str) -> Path:
    """Return the sidecar JSON path for a given filled xlsx output.

    Convention: ``{basename}_payloads.json`` in the same directory as the
    xlsx. Lifting into a named helper so both the writer and the Phase 5
    post-validator can import it and agree on the location. Public (no
    leading underscore) because it is a cross-module contract — the
    validator and the server orchestrator both resolve sidecar paths
    through this function.
    """
    p = Path(xlsx_output_path)
    return p.with_suffix("").with_name(p.stem + "_payloads.json")


# Backwards-compatibility alias — previous callers used the private form.
# New call sites should import ``payload_sidecar_path``.
_payload_sidecar_path = payload_sidecar_path


# ---------------------------------------------------------------------------
# Row resolution
# ---------------------------------------------------------------------------

@dataclass
class _LabelEntry:
    normalized: str
    row: int
    original: str


def _build_label_index(ws) -> list[_LabelEntry]:
    entries: list[_LabelEntry] = []
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if v is None:
            continue
        text = str(v)
        entries.append(_LabelEntry(
            normalized=_normalize(text),
            row=row,
            original=text,
        ))
    return entries


def _normalize(s: str) -> str:
    # Delegates to notes.labels.normalize_label so the writer and the
    # coverage validator stay in lock-step on what counts as "the same
    # label". Strips whitespace, a leading `*` marker, lowercases, and
    # removes a trailing SSM type suffix like `[text block]` — the
    # latter is what keeps MPERS labels matching bare-form payloads.
    return normalize_label(s)


def _resolve_row(
    entries: list[_LabelEntry], label: str,
) -> Optional[tuple[int, str, float]]:
    """Resolve a requested label to a template row.

    Returns `(row, chosen_label, score)` where `score == 1.0` for exact
    hits and in `[_FUZZY_THRESHOLD, 1.0)` for fuzzy fallbacks. Returns
    None when no label scored at or above the threshold.

    Peer-review S4: exact lookup is O(N) by design — the entries list is
    typically <200 rows and re-walking it is cheaper than building a
    dict per call. The fuzzy pass uses SequenceMatcher's quick_ratio /
    real_quick_ratio prefilters, both cheap upper bounds, to skip the
    full ratio() computation on obviously-distant labels.
    """
    target = _normalize(label)
    for e in entries:
        if e.normalized == target:
            return e.row, e.original, 1.0
    # Fuzzy fallback with cheap prefilter — see docstring.
    best_score = 0.0
    best: Optional[_LabelEntry] = None
    for e in entries:
        sm = SequenceMatcher(None, target, e.normalized)
        # real_quick_ratio is an O(1) length-based upper bound; skip
        # entries that can't possibly beat the threshold.
        if sm.real_quick_ratio() < _FUZZY_THRESHOLD:
            continue
        if sm.quick_ratio() < _FUZZY_THRESHOLD:
            continue
        score = sm.ratio()
        if score > best_score:
            best_score = score
            best = e
    if best is not None and best_score >= _FUZZY_THRESHOLD:
        return best.row, best.original, best_score
    return None


def top_candidates(
    entries: list[_LabelEntry], label: str, n: int = 3,
) -> list[tuple[str, float]]:
    """Return the top-`n` closest labels for a rejected match.

    When `_resolve_row` returns None the writer still knows which labels
    came closest — this helper surfaces them so the sub-agent branch of
    `write_notes` can ship them back as a retry hint. Without the hint
    the agent tends to fabricate another bad label on the next turn;
    with a short "did you mean" list it tends to pick one of the real
    options or legitimately skip the note.

    Scores are the same SequenceMatcher ratios used in `_resolve_row`
    so the hint is consistent with what the writer would have accepted.

    Peer-review S4: this path runs ONLY on the rejection path, so the
    full ratio() pass over every entry is cheap (rare). No prefilter
    here because we want a meaningful score for every entry to rank
    candidates — `quick_ratio` would lose ordering information.
    """
    target = _normalize(label)
    scored = [
        (e.original, SequenceMatcher(None, target, e.normalized).ratio())
        for e in entries
    ]
    scored.sort(key=lambda x: x[1], reverse=True)
    return scored[:n]


def resolve_payload_labels(
    entries: list[_LabelEntry],
    payloads: list[NotesPayload],
) -> tuple[list[NotesPayload], list[tuple[str, list[tuple[str, float]]]]]:
    """Partition a payload list into accepted vs rejected.

    - Accepted: every payload whose label resolves via `_resolve_row`
      (exact or fuzzy at/above `_FUZZY_THRESHOLD`). Returned in original
      input order — sub-coordinator row-112 concatenation and audit logs
      depend on a stable order.
    - Rejected: `(requested_label, top_3_candidates)` so the caller can
      ship a retry hint back to the agent. The actual row write is not
      attempted — this is pure triage.

    Used by the `write_notes` sub-agent branch in `notes/agent.py` so
    bad labels don't silently land in `payload_sink` (where they would
    only be discovered at the final write pass, long after the sub-
    agent has exited and can no longer retry).
    """
    accepted: list[NotesPayload] = []
    rejections: list[tuple[str, list[tuple[str, float]]]] = []
    for p in payloads:
        if _resolve_row(entries, p.chosen_row_label) is not None:
            accepted.append(p)
        else:
            rejections.append(
                (p.chosen_row_label, top_candidates(entries, p.chosen_row_label, n=3))
            )
    return accepted, rejections


# ---------------------------------------------------------------------------
# Concatenation + writing
# ---------------------------------------------------------------------------


def _collect_unique_pages(payloads: list[NotesPayload]) -> list[int]:
    """Flatten + deduplicate `source_pages` across a list of payloads.

    Used by the cells_written and sidecar_entries builders to aggregate
    the PDF pages that back a row (peer-review I-6 — previously both
    builders reimplemented the same seen-set/preserve-order dance with
    subtly different local names). Order is first-seen across the
    input list; callers rely on that for stable downstream output.
    """
    seen: set[int] = set()
    out: list[int] = []
    for p in payloads:
        for pg in p.source_pages:
            if pg not in seen:
                seen.add(pg)
                out.append(pg)
    return out


def _combine_payloads(payloads: list[NotesPayload]) -> NotesPayload:
    """Merge multiple payloads targeting the same row.

    Prose: concatenate content with blank line separators, ordered by the
    earliest PDF page each payload cited. Evidence is a semicolon-joined
    list in the same page order. Numeric: last-write-wins (multiple numeric
    payloads for one row is a bug upstream — a warning is logged).

    Ordering by ``min(source_pages)`` keeps row-112's concatenation
    stable across re-runs — without it, input order is
    ``asyncio.wait(ALL_COMPLETED)`` batch-completion order, which is
    non-deterministic and would churn the output on every re-run.

    Evidence fragments use ``;`` as a RESERVED separator. The dedup
    logic splits on ``;``, strips, and case-insensitively de-duplicates.
    Legitimate prose should not contain ``;`` inside a single citation
    (e.g. prefer "Page 14 Note 2(a)" over "Note 5; sub-note (b)");
    doing so would cause a fragment to be silently shredded.

    Single-payload inputs (peer-review #5): we still dedup even with
    one payload, because a sub-agent may emit an already-joined
    "Page 18; Page 18" within a single payload. Skipping dedup on len
    == 1 preserved exactly that duplicate in the filled workbook on
    the FINCO 2021 run. Short-circuit reserved for trivially-clean
    cases only.
    """
    if len(payloads) == 1:
        p = payloads[0]
        # Fast path only when there's nothing to dedup. `;` in evidence
        # means we need the full dedup loop below.
        if p.evidence is None or ";" not in (p.evidence or ""):
            return p

    # Sort by the earliest PDF page each payload cited. Payloads with no
    # source_pages sort to the front (key = 0) so they remain deterministic
    # rather than getting a ``min([])`` crash.
    payloads = sorted(
        payloads,
        key=lambda p: min(p.source_pages) if p.source_pages else 0,
    )

    # Numeric: warn and take first set of values.
    numeric_values = None
    numeric_payloads = [p for p in payloads if p.numeric_values]
    if numeric_payloads:
        numeric_values = numeric_payloads[0].numeric_values
        if len(numeric_payloads) > 1:
            logger.warning(
                "Multiple numeric payloads for row '%s' -- using first",
                payloads[0].chosen_row_label,
            )

    contents = [p.content.strip() for p in payloads if p.content.strip()]
    content = "\n\n".join(contents)

    # Phase 2.1: flatten + dedup evidence fragments across payloads.
    # Sub-agents frequently emit the same citation twice (e.g. two halves
    # of Note 13 Credit risk both citing "Pages 34-36, Note 13 (Credit
    # risk)"), and `_combine_payloads` used to concatenate them verbatim
    # which showed up in col D as "X; X". We split each payload's
    # evidence on ";" so ALREADY-joined strings get split back apart,
    # trim, and dedup case-insensitively while preserving first-seen
    # order (stable across re-runs because the outer payload list is
    # sorted above by min(source_pages)).
    seen_evidence: set[str] = set()
    evidence_parts: list[str] = []
    for p in payloads:
        for frag in p.evidence.split(";"):
            s = frag.strip()
            if not s:
                continue
            key = s.lower()
            if key in seen_evidence:
                continue
            seen_evidence.add(key)
            evidence_parts.append(s)
    evidence = "; ".join(evidence_parts)

    all_pages: list[int] = []
    seen: set[int] = set()
    for p in payloads:
        for pg in p.source_pages:
            if pg not in seen:
                seen.add(pg)
                all_pages.append(pg)

    # Preserve every contributing sub_agent_id so audit tooling can tell
    # which sub-agents wrote each chunk of a row-112 catch-all. Unique
    # entries, in first-seen order; stringified (comma-joined) because
    # the dataclass carries a single Optional[str].
    sub_ids: list[str] = []
    for p in payloads:
        if p.sub_agent_id and p.sub_agent_id not in sub_ids:
            sub_ids.append(p.sub_agent_id)
    combined_sub_id = ",".join(sub_ids) if sub_ids else None

    return NotesPayload(
        chosen_row_label=payloads[0].chosen_row_label,
        content=content,
        evidence=evidence,
        source_pages=all_pages,
        numeric_values=numeric_values,
        sub_agent_id=combined_sub_id,
        # Merged payloads inherit the first payload's heading hierarchy.
        # All payloads in a merge target the same row (same chosen_row_label),
        # which means they're all covering the same note — parent_note and
        # sub_note should match across them, so taking [0] is safe.
        parent_note=payloads[0].parent_note,
        sub_note=payloads[0].sub_note,
    )


# Canonical evidence-column map. Importers (e.g. notes/agent.py's prompt
# renderer) use `evidence_col_letter` to keep the model-facing instructions
# aligned with the writer's cell placement — no more silent drift between
# the docstring and the actual write target.
_EVIDENCE_COL = {"company": 4, "group": 6}


def evidence_col_for(filing_level: str) -> int:
    """Return the 1-indexed column for the evidence/source cell."""
    return _EVIDENCE_COL.get(filing_level, _EVIDENCE_COL["company"])


def evidence_col_letter(filing_level: str) -> str:
    """Return the Excel letter for the evidence column (e.g. 'D', 'F')."""
    return chr(ord("A") + evidence_col_for(filing_level) - 1)


def _evidence_col(filing_level: str) -> int:
    return evidence_col_for(filing_level)


def _write_row(
    ws,
    row: int,
    payload: NotesPayload,
    filing_level: str,
    evidence_col: int,
    errors: list[str],
) -> bool:
    # Refuse to overwrite formula cells in any write target.
    write_cols: list[tuple[int, object]] = []
    if payload.numeric_values:
        # Structured numeric — fill all four value cols for group, B+C for company.
        nv = payload.numeric_values
        if filing_level == "group":
            write_cols.extend([
                (2, nv.get("group_cy")),
                (3, nv.get("group_py")),
                (4, nv.get("company_cy")),
                (5, nv.get("company_py")),
            ])
        else:
            write_cols.extend([
                (2, nv.get("company_cy", nv.get("cy"))),
                (3, nv.get("company_py", nv.get("py"))),
            ])
    else:
        # Truncate first (HTML-aware when appropriate), then flatten to
        # Excel-plaintext before writing. The canonical HTML payload is
        # persisted to `notes_cells` by the coordinator (Step 6); Excel
        # cells always hold the flattened form because the workbook has
        # no HTML rendering layer — tables are pipe-separated, lists get
        # `- ` / `1. ` prefixes, and inline styling is dropped.
        text = truncate_with_footer(payload.content, payload.source_pages)
        if _looks_like_html(text):
            text = html_to_excel_text(text)
        # Prose — content goes to col B only (Group-CY for group, CY for company).
        # Group filings intentionally leave C/D/E empty for prose (PLAN Section 2 #6).
        write_cols.append((2, text))

    wrote_anything = False
    for col, value in write_cols:
        if value is None or value == "":
            continue
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            errors.append(
                f"Refusing to overwrite formula cell {cell.coordinate}: {cell.value}"
            )
            continue
        cell.value = value
        wrote_anything = True

    # Only write evidence when we actually wrote a value (or numeric block)
    # to the row. An evidence-only cell would be a "ghost row" — citation
    # text with nothing to cite against — and is almost always an upstream
    # bug in the payload (LLM produced evidence for a row it didn't fill).
    if payload.evidence and (wrote_anything or payload.numeric_values):
        ev_cell = ws.cell(row=row, column=evidence_col)
        if isinstance(ev_cell.value, str) and ev_cell.value.startswith("="):
            errors.append(
                f"Refusing to overwrite evidence formula cell {ev_cell.coordinate}"
            )
        else:
            ev_cell.value = payload.evidence

    return wrote_anything


def truncate_with_footer(text: str, source_pages: list[int]) -> str:
    """Cap a cell payload at ``CELL_CHAR_LIMIT`` rendered characters.

    Two code paths:

      * HTML payloads (post-rich-editor pipeline): the cap applies to
        the *rendered* text — tags don't count — so we measure with
        ``rendered_length`` and truncate at a tag boundary when needed.
        The footer is appended as HTML so the workbook round-trip (and
        the post-run editor) sees a well-formed document.

      * Plaintext payloads (legacy agents that haven't been migrated
        yet): the cap applies directly to ``len(text)``. Unchanged from
        pre-HTML behaviour so the existing plaintext tests still pin
        the exact pre-footer slice.

    Detection is "does this string contain any HTML tags at all?" —
    cheap heuristic that matches everything the notes prompts now emit
    without false-triggering on occasional angle brackets in ASCII
    schedules (those wouldn't parse as tags, but we check for ``<`` +
    a letter to avoid surprises on content like ``< 10``).
    """
    if not text:
        return text
    if _looks_like_html(text):
        if rendered_length(text) <= CELL_CHAR_LIMIT:
            return text
        pages = list(source_pages) if source_pages else []
        return truncate_html_to_rendered_length(
            text, max_rendered=CELL_CHAR_LIMIT, source_pages=pages,
        )
    if len(text) <= CELL_CHAR_LIMIT:
        return text
    pages_str = ", ".join(str(p) for p in source_pages) if source_pages else "n/a"
    footer = f"\n\n[truncated -- see PDF pages {pages_str}]"
    head_len = CELL_CHAR_LIMIT - len(footer)
    return text[:head_len] + footer


# Back-compat alias. `_truncate_with_footer` was the original name when the
# helper was considered module-private. Tests and an older call-site pin
# the old name; keeping the alias avoids a noisy mass-rename while the
# public entry point is now `truncate_with_footer`.
_truncate_with_footer = truncate_with_footer


def _looks_like_html(text: str) -> bool:
    """Cheap test: is ``text`` likely an HTML payload?

    Matches ``<`` followed by a known allowed tag name (from the
    sanitiser whitelist) or ``/`` for close tags. Tight enough to
    avoid false-positives on accounting placeholders like
    ``<CAPEX>`` / ``<Other Asset>`` (peer-review #8) and ASCII
    comparisons like ``< 10`` / ``<<``. Loose enough to catch every
    tag the notes prompts emit: the whitelist is kept in lock-step
    with ``notes.html_sanitize.ALLOWED_TAGS``.
    """
    # Import here to avoid a top-level cycle; the sanitiser module
    # imports nothing from writer but writer has multiple heavy
    # top-level imports the sanitiser doesn't need.
    from notes.html_sanitize import ALLOWED_TAGS
    # Lowercase tag list; include `/` for close-tag detection (e.g.
    # a bare ``</p>`` at the start of a cell still signals HTML).
    allowed_lower = {t.lower() for t in ALLOWED_TAGS}

    i = 0
    n = len(text)
    while i < n:
        if text[i] != "<":
            i += 1
            continue
        j = i + 1
        if j < n and text[j] == "/":
            j += 1
        # Scan the tag name. HTML5 tag names start with an ASCII
        # letter and may include digits after (h1, h3, …). Bail on
        # the first non-alphanumeric — space, `>`, `/` for
        # self-closing, attribute separator, or end-of-string.
        start = j
        if j < n and text[j].isalpha():
            j += 1
            while j < n and (text[j].isalpha() or text[j].isdigit()):
                j += 1
        name = text[start:j].lower()
        if name and name in allowed_lower:
            return True
        i = max(j, i + 1)
    return False
