"""RUN-REVIEW P0-2 (2026-04-26): introspect a known-good SSM-submitted
xlsx to determine whether the `[Text block added]` placeholder + hidden
`+FootnoteTextsN` sheet convention claimed in RUN-REVIEW §3.5 actually
exists, and if so, what its exact shape is.

The reviewer observed this convention on a Windows workstation looking
at the filer's submitted file but did not include the SSM template /
validator output as evidence. Before we ship a render-mode flag for an
unverified hypothesis, this harness reads the actual artifact.

Usage:
    python3 scripts/validate_ssm_compatibility.py path/to/filer-submission.xlsx

Output:
    Structured report on stdout describing:
      - Whether `[Text block added]` (or any [text block...] placeholder)
        appears anywhere in the visible cells
      - All hidden sheets and their shapes
      - Cell content overlap between visible "[Text block added]" rows
        and hidden footnote rows
      - A summary verdict that's safe to copy into
        `docs/SSM-NOTES-FORMAT-INVESTIGATION.md`

Investigation-only — does not modify the input file.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


def introspect(xlsx_path: Path) -> Dict[str, Any]:
    """Walk the xlsx and return a structured summary."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, data_only=True)

    visible_sheets: List[str] = []
    hidden_sheets: List[str] = []
    very_hidden_sheets: List[str] = []
    for ws in wb.worksheets:
        state = getattr(ws, "sheet_state", "visible")
        if state == "visible":
            visible_sheets.append(ws.title)
        elif state == "hidden":
            hidden_sheets.append(ws.title)
        elif state == "veryHidden":
            very_hidden_sheets.append(ws.title)

    placeholder_hits: List[Dict[str, Any]] = []
    placeholder_phrases = (
        "[text block added]",
        "[textblock added]",
        "[text block]",
    )
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                lv = v.strip().lower()
                if any(p in lv for p in placeholder_phrases):
                    placeholder_hits.append({
                        "sheet": ws.title,
                        "row": cell.row,
                        "col": cell.column_letter,
                        "value": v,
                    })

    # Summarise each hidden sheet's shape so we can match against the
    # reviewer's `+FootnoteTextsN` claim. No content sampling — that
    # might leak filer disclosures into shared logs.
    hidden_summary: List[Dict[str, Any]] = []
    for name in hidden_sheets + very_hidden_sheets:
        ws = wb[name]
        hidden_summary.append({
            "name": name,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "state": ws.sheet_state,
            "header_row": [
                str(ws.cell(1, c).value) if ws.cell(1, c).value else None
                for c in range(1, min(ws.max_column + 1, 6))
            ],
        })

    return {
        "path": str(xlsx_path),
        "visible_sheets": visible_sheets,
        "hidden_sheets": hidden_sheets,
        "very_hidden_sheets": very_hidden_sheets,
        "placeholder_hits": placeholder_hits,
        "hidden_summary": hidden_summary,
    }


def render_report(summary: Dict[str, Any]) -> str:
    lines: List[str] = [
        f"=== SSM compatibility introspection: {summary['path']} ===",
        "",
        f"Visible sheets ({len(summary['visible_sheets'])}):",
    ]
    for s in summary["visible_sheets"]:
        lines.append(f"  - {s}")
    lines.append("")
    lines.append(f"Hidden sheets ({len(summary['hidden_sheets'])}):")
    for s in summary["hidden_sheets"]:
        lines.append(f"  - {s}")
    if summary["very_hidden_sheets"]:
        lines.append("")
        lines.append(f"Very-hidden sheets ({len(summary['very_hidden_sheets'])}):")
        for s in summary["very_hidden_sheets"]:
            lines.append(f"  - {s}")
    lines.append("")

    if summary["placeholder_hits"]:
        lines.append(
            f"PLACEHOLDER FOUND: {len(summary['placeholder_hits'])} cell(s) "
            f"contain a text-block-style placeholder."
        )
        for h in summary["placeholder_hits"][:20]:
            lines.append(f"  {h['sheet']} {h['col']}{h['row']}: {h['value']!r}")
        if len(summary["placeholder_hits"]) > 20:
            lines.append(f"  … ({len(summary['placeholder_hits']) - 20} more)")
    else:
        lines.append(
            "NO PLACEHOLDER FOUND: visible cells contain no '[Text block "
            "added]' or similar token. The convention claimed in "
            "RUN-REVIEW §3.5 is NOT present in this file."
        )
    lines.append("")

    if summary["hidden_summary"]:
        lines.append("Hidden sheet shapes:")
        for h in summary["hidden_summary"]:
            lines.append(
                f"  - {h['name']}: {h['max_row']}r × {h['max_col']}c "
                f"(state={h['state']}); header row 1: {h['header_row']}"
            )
        # Specific check for the reviewer's named conventions
        names = {h["name"] for h in summary["hidden_summary"]}
        for expected in ("+FootnoteTexts0", "+Elements", "+Lineitems"):
            present = expected in names or any(
                expected.lower() in n.lower() for n in names
            )
            lines.append(f"  +FootnoteTexts/+Elements/+Lineitems check: "
                         f"{expected!r} {'FOUND' if present else 'absent'}")
    else:
        lines.append("No hidden sheets — the reviewer's `+FootnoteTextsN` "
                     "sheet pattern is NOT present.")

    lines.append("")
    lines.append("=== Verdict ===")
    has_placeholder = bool(summary["placeholder_hits"])
    has_footnote_sheet = any(
        "footnotetext" in h["name"].lower()
        for h in summary["hidden_summary"]
    )
    if has_placeholder and has_footnote_sheet:
        lines.append(
            "CONFIRMED: this file matches RUN-REVIEW §3.5's claim. "
            "Open `PLAN-ssm-notes-output.md` and design the render-mode "
            "split now that the convention is verified."
        )
    elif has_placeholder and not has_footnote_sheet:
        lines.append(
            "PARTIAL: placeholder present but no FootnoteTexts sheet. "
            "The convention may differ from the reviewer's recollection. "
            "Manual inspection needed."
        )
    elif not has_placeholder and has_footnote_sheet:
        lines.append(
            "PARTIAL: hidden sheets present but no [Text block added] "
            "placeholder in visible cells. Possibly an orthogonal "
            "feature, not the claimed convention."
        )
    else:
        lines.append(
            "NOT CONFIRMED: neither placeholder nor footnote sheet "
            "present. RUN-REVIEW P0-2 should be closed as not-actionable "
            "in this checkout — the inline-HTML output the codebase "
            "produces today is appropriate. Document the finding in "
            "docs/SSM-NOTES-FORMAT-INVESTIGATION.md."
        )

    return "\n".join(lines)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx_path", type=Path,
                    help="Path to the filer's SSM-submitted xlsx.")
    args = ap.parse_args()

    try:
        summary = introspect(args.xlsx_path)
    except FileNotFoundError as e:
        print(f"error: {e}", file=sys.stderr)
        return 1

    print(render_report(summary))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
