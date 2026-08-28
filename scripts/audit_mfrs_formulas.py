#!/usr/bin/env python3
"""Read-only MFRS face-formula audit against the SSM calculation linkbase.

Unlike the regeneration script, this command never writes a workbook. It
checks every same-sheet formula owned by the generated SOFP/SOPL/SOCI/
SOCF-Direct roles, the hand-curated SOCF-Indirect role, and the SOCIE movement
weights. Formula term order is ignored; child rows and coefficients must match.

Run: ``python -m scripts.audit_mfrs_formulas``
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from prompts._sign_conventions import _parse_total_formula
import scripts.regenerate_mfrs_sofp_sopl_formulas as mfrs


def _actual_parts(formula: str, column: str) -> set[tuple[int, int]]:
    return {
        (row, sign)
        for sign, referenced_column, row in _parse_total_formula(formula)
        if referenced_column == column
    }


def _audit_role_cells(
    path: Path,
    level: str,
    roles: list[tuple[str, str]],
) -> tuple[int, list[str]]:
    # These audits address formulas by coordinate. In openpyxl's read-only
    # mode every random cell lookup reparses worksheet XML from the start,
    # making this small audit quadratic in practice. The shipped workbooks
    # are small, so normal mode is both faster and still read-only at the
    # application level: this command never saves the workbook.
    wb = openpyxl.load_workbook(path, data_only=False)
    checked = 0
    issues: list[str] = []
    try:
        for sheet_name, role in roles:
            ws = wb[sheet_name]
            _rows, expected = mfrs.expected_same_sheet_formulas(role)
            for row, parts in expected.items():
                expected_set = set(parts)
                for column in mfrs.value_columns(level):
                    checked += 1
                    coordinate = f"{column}{row}"
                    formula = ws[coordinate].value
                    if not isinstance(formula, str) or not formula.startswith("="):
                        issues.append(f"{path}:{sheet_name}!{coordinate} missing formula")
                        continue
                    actual_set = _actual_parts(formula, column)
                    if actual_set != expected_set:
                        issues.append(
                            f"{path}:{sheet_name}!{coordinate} "
                            f"expected_only={sorted(expected_set - actual_set)} "
                            f"actual_only={sorted(actual_set - expected_set)}"
                        )
    finally:
        wb.close()
    return checked, issues


def _audit_socie(path: Path) -> tuple[int, list[str]]:
    wb = openpyxl.load_workbook(path, data_only=False)
    checked = 0
    issues: list[str] = []
    try:
        ws = wb["SOCIE"]
        dividend_rows = [
            row
            for row in range(1, ws.max_row + 1)
            if str(ws.cell(row, 1).value or "").strip().casefold() == "dividends paid"
        ]
        if not dividend_rows:
            return checked, [f"{path}: no Dividends paid rows"]
        for dividend_row in dividend_rows:
            for column_index in range(2, min(ws.max_column, 24) + 1):
                column = openpyxl.utils.get_column_letter(column_index)
                formulas = [
                    ws.cell(row, column_index).value
                    for row in range(dividend_row + 1, min(dividend_row + 10, ws.max_row) + 1)
                    if isinstance(ws.cell(row, column_index).value, str)
                    and ws.cell(row, column_index).value.startswith("=")
                ]
                # Some component columns are entirely formula/blank for a
                # block. Count only a formula that actually references the row.
                referencing = [f for f in formulas if f"{column}{dividend_row}" in f]
                if not referencing:
                    continue
                checked += 1
                if not any(
                    (dividend_row, -1) in _actual_parts(f, column)
                    for f in referencing
                ):
                    issues.append(
                        f"{path}:SOCIE!{column}{dividend_row} is not subtracted"
                    )
    finally:
        wb.close()
    return checked, issues


def audit() -> tuple[int, list[str]]:
    mfrs.configure_mfrs_taxonomy()
    checked = 0
    issues: list[str] = []
    root = mfrs.TEMPLATE_ROOT

    for level in ("Company", "Group"):
        for filename, roles in mfrs.TEMPLATE_ROLES.items():
            count, found = _audit_role_cells(root / level / filename, level, roles)
            checked += count
            issues.extend(found)

        # Hand-curated indirect cash flow still uses the same role-500100
        # calculation relationships. It differs only in formula term order.
        mfrs.taxonomy._PRE_TO_CALC_ROLE["520000"] = "500100"
        count, found = _audit_role_cells(
            root / level / "07-SOCF-Indirect.xlsx",
            level,
            [("SOCF-Indirect", "520000")],
        )
        checked += count
        issues.extend(found)

        count, found = _audit_socie(root / level / "09-SOCIE.xlsx")
        checked += count
        issues.extend(found)

    return checked, issues


def main() -> None:
    checked, issues = audit()
    print(f"MFRS formula/sign cells checked: {checked}")
    for issue in issues:
        print(f"ISSUE: {issue}")
    if issues:
        print("RESULT: ISSUES FOUND")
        raise SystemExit(1)
    print("RESULT: ALL MFRS FORMULAS MATCH THE LINKBASE SIGN CONTRACT")


if __name__ == "__main__":
    main()
