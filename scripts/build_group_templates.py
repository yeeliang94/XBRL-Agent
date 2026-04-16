"""
Builds Group-level templates from the current company-only templates.

Layout change (templates 01-08, 10-14):
  Before: A=label  B=Period(Cur)  C=Period(Prior)  D=Source
  After:  A=label  B=Group Cur    C=Group Prior    D=Company Cur    E=Company Prior    F=Source

  Row 1:  [Group]<merged B:C>   [Company]<merged D:E>   Source
  Row 2:  period                period                  period           period
  Row 3+: data (unchanged; all existing formulas preserved in B/C; D/E get duplicates of B/C formulas with B->D, C->E substitution)

Note: SOCIE (09) is handled separately in build_group_socie.py because it uses
columns B-X for equity components, not for period comparisons.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from copy import copy
import re
import shutil
from pathlib import Path

# Anchor to the repo root, not the caller's cwd. Previously running this
# from a different working directory would silently produce wrong output
# (peer-review I14).
_REPO_ROOT = Path(__file__).resolve().parent.parent
SRC_DIR = _REPO_ROOT / "XBRL-template-MFRS"
DST_DIR = _REPO_ROOT / "XBRL-template-MFRS" / "Group"

# Templates that follow the standard 2-column (period-pair) layout
STANDARD_TEMPLATES = [
    "01-SOFP-CuNonCu.xlsx",
    "02-SOFP-OrderOfLiquidity.xlsx",
    "03-SOPL-Function.xlsx",
    "04-SOPL-Nature.xlsx",
    "05-SOCI-BeforeTax.xlsx",
    "06-SOCI-NetOfTax.xlsx",
    "07-SOCF-Indirect.xlsx",
    "08-SOCF-Direct.xlsx",
    "10-Notes-CorporateInfo.xlsx",
    "11-Notes-AccountingPolicies.xlsx",
    "12-Notes-ListOfNotes.xlsx",
    "13-Notes-IssuedCapital.xlsx",
    "14-Notes-RelatedParty.xlsx",
]

# Regex to rewrite formulas: map B->D and C->E (for both same-sheet and cross-sheet refs)
# Use placeholders to avoid double-substitution.
_COL_PAT = re.compile(r"\$?([A-Z]+)(\$?\d+)")

def _shift_col_B_to_D(formula: str) -> str:
    """Rewrite a formula so that column B becomes D and C becomes E (absolute or relative).
    Leaves other columns untouched. Handles both 'B12' and '$B$12' forms, inside
    same-sheet and cross-sheet references.
    """
    def repl(m):
        col = m.group(1)
        rest = m.group(2)
        dollar = m.group(0).startswith('$')
        prefix = '$' if dollar else ''
        if col == 'B':
            return f"{prefix}D{rest}"
        elif col == 'C':
            return f"{prefix}E{rest}"
        return m.group(0)
    return _COL_PAT.sub(repl, formula)


def _copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def transform_sheet(ws):
    """In-place transform a sheet: shift Source D->F, duplicate B/C into D/E, add header row 1, set row 2 periods."""
    max_row = ws.max_row
    max_col = ws.max_column

    # Step 1: Move column D contents (Source) to column F for every row
    for r in range(1, max_row + 1):
        d_cell = ws.cell(row=r, column=4)  # D
        f_cell = ws.cell(row=r, column=6)  # F
        if d_cell.value is not None:
            f_cell.value = d_cell.value
            _copy_cell_style(d_cell, f_cell)
            d_cell.value = None
        # Clear style on D (will be overwritten)

    # Step 2: For every row starting at row 3, duplicate B -> D and C -> E,
    # rewriting formulas via column shift B->D, C->E.
    for r in range(3, max_row + 1):
        for src_col_letter, dst_col_letter in (('B', 'D'), ('C', 'E')):
            src = ws[f"{src_col_letter}{r}"]
            dst = ws[f"{dst_col_letter}{r}"]
            if src.value is None:
                # Copy style only so the cells look visually consistent for data entry
                _copy_cell_style(src, dst)
                continue
            if isinstance(src.value, str) and src.value.startswith('='):
                dst.value = _shift_col_B_to_D(src.value)
            else:
                # Static value (label, note) — copy as-is so Company column mirrors Group structure
                dst.value = src.value
            _copy_cell_style(src, dst)

    # Step 3: Overwrite row 1 with section labels.
    # Clear existing row 1 contents in B, C, D (period strings were there before).
    for col in (2, 3, 4, 5, 6):
        c = ws.cell(row=1, column=col)
        c.value = None
    ws.cell(row=1, column=2).value = "Group"   # B1
    ws.cell(row=1, column=4).value = "Company" # D1
    ws.cell(row=1, column=6).value = "Source"  # F1

    # Merge B1:C1 and D1:E1
    try:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
    except Exception:
        pass

    # Style the header row
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", start_color="E7E6E6")
    center = Alignment(horizontal="center", vertical="center")
    for col in (2, 4, 6):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    # Step 4: Put period strings in row 2 (previously empty in standard templates)
    period = "01/01/YYYY - 31/12/YYYY"
    for col in (2, 3, 4, 5):
        c = ws.cell(row=2, column=col)
        c.value = period
        c.font = Font(italic=True)
        c.alignment = Alignment(horizontal="center")

    # Step 5: Widen columns so the 4 period columns are readable
    for col_letter in ("B", "C", "D", "E"):
        ws.column_dimensions[col_letter].width = max(
            ws.column_dimensions[col_letter].width or 0, 18
        )
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width or 0, 20)


def _looks_already_group(src_path: Path) -> bool:
    """Idempotency guard: if the *source* file already has the 6-column
    Group layout (Company D/E columns present beyond row 2), running the
    transform again would silently double-shift columns. Detect by
    peeking at D3 / E3 — Company data cells that only exist post-build.
    """
    try:
        wb = openpyxl.load_workbook(src_path, read_only=True, data_only=False)
        ws = wb.active
        d3 = ws["D3"].value
        e3 = ws["E3"].value
        wb.close()
        return d3 is not None or e3 is not None
    except Exception:
        return False


def transform_workbook(src_path: Path, dst_path: Path):
    wb = openpyxl.load_workbook(src_path)
    for sn in wb.sheetnames:
        transform_sheet(wb[sn])
    dst_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dst_path)
    print(f"  -> {dst_path}")


def main():
    for name in STANDARD_TEMPLATES:
        src = SRC_DIR / name
        dst = DST_DIR / name
        if _looks_already_group(src):
            print(f"Skipping {name} — source already has Group D/E columns")
            continue
        print(f"Building Group/{name} from {src.name}")
        transform_workbook(src, dst)
    print("\nStandard templates done. SOCIE handled separately.")


if __name__ == "__main__":
    main()
