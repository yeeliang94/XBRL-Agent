#!/usr/bin/env python3
"""Regenerate MFRS SOFP/SOPL/SOCI/SOCF-Direct template formulas from the SSM linkbase.

This intentionally touches only the workbooks listed in ``TEMPLATE_ROLES``
under ``XBRL-template-MFRS/{Company,Group}``. It preserves existing workbook
layout, styles, labels, and data-entry cells; only value-column formulas are
cleared and re-emitted from the MFRS calculation/presentation linkbases.

SOCI and SOCF-Direct were added 2026-07-03 after an audit found the
hand-built originals deviated from the SSM calculation linkbase (missing
rollups into Total OCI, reclassification adjustments added instead of
subtracted, and a payments-entered-positive sign convention that
contradicted both the SSM weights and the SOCF prompt). SOCF-Indirect
already ties to the linkbase and is deliberately left out.
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import sys
from typing import Iterable

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

import scripts.generate_mpers_templates as taxonomy  # noqa: E402

MFRS_TAXONOMY_DIR = (
    REPO_ROOT / "SSMxT_2022v1.0" / "rep" / "ssm" / "ca-2016" / "fs" / "mfrs"
)
TEMPLATE_ROOT = REPO_ROOT / "XBRL-template-MFRS"

TEMPLATE_ROLES: dict[str, list[tuple[str, str]]] = {
    "01-SOFP-CuNonCu.xlsx": [
        ("SOFP-CuNonCu", "210000"),
        ("SOFP-Sub-CuNonCu", "210100"),
    ],
    "02-SOFP-OrderOfLiquidity.xlsx": [
        ("SOFP-OrdOfLiq", "220000"),
        ("SOFP-Sub-OrdOfLiq", "220100"),
    ],
    "03-SOPL-Function.xlsx": [
        ("SOPL-Function", "310000"),
        ("SOPL-Analysis-Function", "310100"),
    ],
    "04-SOPL-Nature.xlsx": [
        ("SOPL-Nature", "320000"),
        ("SOPL-Analysis-Nature", "320100"),
    ],
    "05-SOCI-BeforeTax.xlsx": [
        ("SOCI-BeforeOfTax", "420000"),
    ],
    "06-SOCI-NetOfTax.xlsx": [
        ("SOCI-NetOfTax", "410000"),
    ],
    "08-SOCF-Direct.xlsx": [
        ("SOCF-Direct", "510000"),
    ],
}

PRE_TO_CALC_ROLE = {
    "210000": "200100",
    "210100": "200200",
    "220000": "200100",
    "220100": "200200",
    "310000": "300100",
    "310100": "300200",
    "320000": "300100",
    "320100": "300200",
    "410000": "400100",
    "420000": "400100",
    "510000": "500100",
}

# Hand-built sheets whose body rows sit N rows below the presentation-walk
# default (``_FIRST_BODY_ROW + idx``). Verified by label-by-label alignment
# against the walk before these roles were added; every other sheet is 0.
ROLE_ROW_OFFSET = {
    "420000": 1,  # SOCI-BeforeOfTax body starts at row 4
}

# Concepts that must NOT be expanded through their calc children when they
# have no visible row. On the Direct-method SOCF, the calc linkbase lists the
# indirect-method subtotal ``CashFlowsFromUsedInOperations`` as a child of the
# operating total; expanding it walks the INDIRECT reconciliation tree, where
# the lease-payments concept carries weight -1 (an add-back) and cancels the
# direct-method +1 arc — silently dropping the row from the emitted formula.
ROLE_EXPANSION_EXCLUDE = {
    "510000": {"ifrs-full_CashFlowsFromUsedInOperations"},
}

# Accounting identities the calc linkbase cannot express (cross-period
# bridges), re-applied after linkbase emission. ``{c}`` is the value column.
SUPPLEMENTAL_FORMULAS: dict[str, dict[str, dict[int, str]]] = {
    # Cash and cash equivalents at end = beginning + net increase after FX.
    "08-SOCF-Direct.xlsx": {"SOCF-Direct": {77: "=1*{c}76+1*{c}75"}},
}


def configure_mfrs_taxonomy() -> None:
    """Point the existing taxonomy walker helpers at MFRS."""
    taxonomy._MPERS_TAXONOMY_DIR = MFRS_TAXONOMY_DIR
    taxonomy._ROLE_XSD = MFRS_TAXONOMY_DIR / "rol_ssmt-fs-mfrs_2022-12-31.xsd"
    taxonomy._LABEL_MAP_CACHE = None
    taxonomy._LABEL_ROLE_TABLE = {}
    taxonomy._PRE_TO_CALC_ROLE.update(PRE_TO_CALC_ROLE)
    taxonomy._pre_file_for_role = (
        lambda role: MFRS_TAXONOMY_DIR
        / f"pre_ssmt-fs-mfrs_2022-12-31_role-{role}.xml"
    )
    taxonomy._calc_file_for_role = (
        lambda role: MFRS_TAXONOMY_DIR
        / f"cal_ssmt-fs-mfrs_2022-12-31_role-{role}.xml"
    )


def value_columns(level: str) -> tuple[str, ...]:
    if level == "Company":
        return ("B", "C")
    if level == "Group":
        return ("B", "C", "D", "E")
    raise ValueError(f"Unsupported level: {level}")


def col_index(col_letter: str) -> int:
    return ord(col_letter) - ord("A") + 1


def formula_from_parts(col_letter: str, parts: Iterable[tuple[int, int]]) -> str:
    return "=" + "+".join(f"{weight}*{col_letter}{row}" for row, weight in parts)


def clear_value_column_formulas(ws, columns: tuple[str, ...]) -> int:
    cleared = 0
    for row in range(1, ws.max_row + 1):
        for col in columns:
            cell = ws.cell(row=row, column=col_index(col))
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None
                cleared += 1
    return cleared


def expected_same_sheet_formulas(
    role_number: str,
) -> tuple[
    list[tuple[int, str, str, bool]],
    dict[int, list[tuple[int, int]]],
]:
    """Return presentation rows and {xlsx_row: [(child_row, weight), ...]}."""
    rows = taxonomy.walk_role(taxonomy._pre_file_for_role(role_number))
    calc_blocks = taxonomy.parse_calc_linkbase_grouped_for_pre_role(role_number)
    row_offset = ROLE_ROW_OFFSET.get(role_number, 0)
    expansion_exclude = ROLE_EXPANSION_EXCLUDE.get(role_number, set())

    concept_to_rows: dict[str, list[int]] = defaultdict(list)
    first_occurrence: dict[str, int] = {}
    for idx, (_depth, concept_id, _label, _is_abstract) in enumerate(rows):
        xlsx_row = taxonomy._FIRST_BODY_ROW + idx + row_offset
        concept_to_rows[concept_id].append(xlsx_row)
        first_occurrence.setdefault(concept_id, xlsx_row)

    calc_children: dict[str, list[tuple[str, int]]] = {}
    parent_to_blocks: dict[str, list[list[tuple[str, int]]]] = defaultdict(list)
    for _link_role, calc_map in calc_blocks:
        for parent, children in calc_map.items():
            parent_to_blocks[parent].append(children)
            calc_children.setdefault(parent, children)

    def _resolve_child_rows(
        concept_id: str,
        weight: int,
        seen: set[str] | None = None,
    ) -> list[tuple[int, int]]:
        """Resolve calc children to presentation rows.

        Some MFRS face statements use split presentation rows where the calc
        role points to a generic subtotal concept. Example: ``Assets`` points
        to ``ifrs-full_BiologicalAssets``, while the CuNonCu presentation has
        ``NoncurrentBiologicalAssets`` and ``CurrentBiologicalAssets`` rows.
        Expand such missing generic concepts through their calc children.
        """
        if concept_id in first_occurrence:
            return [(first_occurrence[concept_id], weight)]
        if concept_id in expansion_exclude:
            return []

        seen = set() if seen is None else set(seen)
        if concept_id in seen:
            return []
        seen.add(concept_id)

        resolved: list[tuple[int, int]] = []
        for child_concept, child_weight in calc_children.get(concept_id, []):
            resolved.extend(
                _resolve_child_rows(
                    child_concept,
                    weight * child_weight,
                    seen,
                )
            )
        return resolved

    expected: dict[int, list[tuple[int, int]]] = {}
    for parent, blocks in parent_to_blocks.items():
        parent_rows = concept_to_rows.get(parent, [])
        if not parent_rows:
            continue

        n = min(len(parent_rows), len(blocks))
        row_start = len(parent_rows) - n
        block_start = len(blocks) - n
        for i in range(n):
            parent_row = parent_rows[row_start + i]
            children = blocks[block_start + i]
            by_row: dict[int, int] = defaultdict(int)
            for child_concept, weight in children:
                for child_row, resolved_weight in _resolve_child_rows(
                    child_concept, weight
                ):
                    by_row[child_row] += resolved_weight
            parts = [(row, by_row[row]) for row in sorted(by_row) if by_row[row]]
            if parts:
                expected[parent_row] = parts

    return rows, expected


def write_same_sheet_formulas(
    ws,
    expected: dict[int, list[tuple[int, int]]],
    columns: tuple[str, ...],
) -> int:
    written = 0
    for row, parts in sorted(expected.items()):
        for col in columns:
            ws.cell(row=row, column=col_index(col), value=formula_from_parts(col, parts))
            written += 1
    return written


def write_face_to_sub_formulas(
    face_ws,
    face_rows: list[tuple[int, str, str, bool]],
    face_same_formula_rows: set[int],
    sub_ws,
    sub_rows: list[tuple[int, str, str, bool]],
    columns: tuple[str, ...],
) -> int:
    """Wire face rows to the last matching concept on the sub-sheet."""
    # A few MFRS face concepts use different names from their sub-sheet total
    # equivalents. Keep these aliases explicit and narrow.
    concept_aliases = {
        "ifrs-full_CurrentPortionOfLongtermBorrowings": "ifrs-full_ShorttermBorrowings",
    }

    sub_last_row: dict[str, int] = {}
    for idx, (_depth, concept_id, _label, _is_abstract) in enumerate(sub_rows):
        sub_last_row[concept_id] = taxonomy._FIRST_BODY_ROW + idx

    written = 0
    for idx, (_depth, concept_id, _label, _is_abstract) in enumerate(face_rows):
        face_row = taxonomy._FIRST_BODY_ROW + idx
        if face_row <= 6:
            continue
        if face_row in face_same_formula_rows:
            continue
        sub_row = sub_last_row.get(concept_id)
        if sub_row is None:
            sub_row = sub_last_row.get(concept_aliases.get(concept_id, ""))
        if sub_row is None:
            continue
        for col in columns:
            formula = f"='{sub_ws.title}'!{col}{sub_row}"
            face_ws.cell(row=face_row, column=col_index(col), value=formula)
            written += 1
    return written


def regenerate_workbook(path: Path, level: str) -> dict[str, int | str]:
    roles = TEMPLATE_ROLES[path.name]
    columns = value_columns(level)
    wb = openpyxl.load_workbook(path)

    sheet_rows: dict[str, list[tuple[int, str, str, bool]]] = {}
    sheet_expected: dict[str, dict[int, list[tuple[int, int]]]] = {}

    for sheet_name, role_number in roles:
        rows, expected = expected_same_sheet_formulas(role_number)
        sheet_rows[sheet_name] = rows
        sheet_expected[sheet_name] = expected

    cleared = 0
    same_written = 0
    cross_written = 0

    for sheet_name, _role_number in roles:
        ws = wb[sheet_name]
        cleared += clear_value_column_formulas(ws, columns)
        same_written += write_same_sheet_formulas(
            ws, sheet_expected[sheet_name], columns
        )
        supplemental = SUPPLEMENTAL_FORMULAS.get(path.name, {}).get(sheet_name, {})
        for row, template in supplemental.items():
            for col in columns:
                ws.cell(
                    row=row,
                    column=col_index(col),
                    value=template.format(c=col),
                )
                same_written += 1

    if len(roles) == 2:
        face_sheet, _face_role = roles[0]
        sub_sheet, _sub_role = roles[1]
        cross_written += write_face_to_sub_formulas(
            wb[face_sheet],
            sheet_rows[face_sheet],
            set(sheet_expected[face_sheet]),
            wb[sub_sheet],
            sheet_rows[sub_sheet],
            columns,
        )

    wb.save(path)
    return {
        "path": str(path),
        "cleared": cleared,
        "same_written": same_written,
        "cross_written": cross_written,
    }


def main() -> None:
    configure_mfrs_taxonomy()
    summaries = []
    for level in ("Company", "Group"):
        for filename in TEMPLATE_ROLES:
            summaries.append(regenerate_workbook(TEMPLATE_ROOT / level / filename, level))

    for summary in summaries:
        print(
            "{path}: cleared={cleared} same_sheet={same_written} cross_sheet={cross_written}".format(
                **summary
            )
        )


if __name__ == "__main__":
    main()
