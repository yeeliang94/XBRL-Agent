"""
Builds the Group-level SOCIE template from the 2-block company-only SOCIE.

Current layout (09-SOCIE.xlsx):
  Row 1: B1 = period string
  Row 2: B2..X2 = equity component column headers (24 cols: Issued capital .. Total)
  Block 1 (rows 3-25): first SOCIE block  (Company Current Period in original)
  Block 2 (rows 27-49): second SOCIE block (Company Prior Period in original)

Group layout (4 blocks):
  Block 1 (Group, Current period): rows 3-25  (copied unchanged)
  Block 2 (Group, Prior period):  rows 27-49 (copied unchanged)
  Block 3 (Company, Current period): rows 51-73 (block 1 with +48 row shift)
  Block 4 (Company, Prior period):  rows 75-97 (block 2 with +48 row shift)

We also add a "section label" row just above each block so the user can see
what each block represents. The section label is placed in column A as text
and visually distinguished via bold + fill. The existing repeated
"Statement of changes in equity" banner rows are replaced with the section
labels so rows do not shift.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import re
from pathlib import Path

# Anchor to the repo root, not the caller's cwd (peer-review I14).
_REPO_ROOT = Path(__file__).resolve().parent.parent
SRC = _REPO_ROOT / "XBRL-template-MFRS" / "09-SOCIE.xlsx"
DST = _REPO_ROOT / "XBRL-template-MFRS" / "Group" / "09-SOCIE.xlsx"

# Row shift between the existing 2-block template and the new 4-block layout
BLOCK_STRIDE = 48  # Block 1 -> Block 3 (row 3 -> row 51), Block 2 -> Block 4

_CELL_REF_PAT = re.compile(r"(\$?[A-Z]+)(\$?)(\d+)")

def _shift_rows(formula: str, delta: int) -> str:
    """Shift every row reference in a formula by +delta. Handles $A$1 and A1."""
    def repl(m):
        col = m.group(1)
        dollar = m.group(2)
        row = int(m.group(3))
        return f"{col}{dollar}{row + delta}"
    return _CELL_REF_PAT.sub(repl, formula)


def _copy_cell(src_cell, dst_cell, row_delta: int = 0):
    """Copy a cell's value (shifting formula row refs) and style."""
    v = src_cell.value
    if isinstance(v, str) and v.startswith("=") and row_delta:
        dst_cell.value = _shift_rows(v, row_delta)
    else:
        dst_cell.value = v
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def build_socie_group():
    # Idempotency guard: the source is the 2-block company template. If a
    # prior run already converted the source in place (e.g. someone ran
    # this against an already-built artifact), rows 51+ will already
    # contain data and re-running would double-shift. Refuse.
    wb_probe = openpyxl.load_workbook(SRC, read_only=True, data_only=False)
    ws_probe = wb_probe["SOCIE"]
    max_row = ws_probe.max_row
    wb_probe.close()
    if max_row > 50:
        raise RuntimeError(
            f"Source SOCIE has {max_row} rows — looks like it was already "
            "transformed into a 4-block Group layout. Refusing to re-run."
        )

    wb = openpyxl.load_workbook(SRC)
    ws = wb["SOCIE"]
    max_col = ws.max_column  # should be 24 (X)

    # Row ranges for existing blocks
    BLOCK1_ROWS = range(3, 26)   # rows 3-25 inclusive
    BLOCK2_ROWS = range(26, 50)  # rows 26-49 inclusive (row 26 is the gap line in original)

    # Duplicate block 1 -> block 3 (rows 3-25 -> 51-73), shift formulas by +48
    # Duplicate block 2 -> block 4 (rows 26-49 -> 74-97), shift formulas by +48
    for src_rows, delta in ((BLOCK1_ROWS, 48), (BLOCK2_ROWS, 48)):
        for r in src_rows:
            for c in range(1, max_col + 1):
                src = ws.cell(row=r, column=c)
                if src.value is None and not src.has_style:
                    continue
                dst = ws.cell(row=r + delta, column=c)
                _copy_cell(src, dst, row_delta=delta)

    # Now update the section banner text in A3 (top of block 1), A27 (top of block 2),
    # A51 (top of block 3), A75 (top of block 4) so users can tell the blocks apart.
    # The original has "Statement of changes in equity" repeated in A3/A4/A5 and
    # A27/A28/A29 — we overwrite A3 and A27 with descriptive labels and leave the
    # other repeated rows untouched.
    block_labels = {
        3:  "Group - Current period",
        27: "Group - Prior period",
        51: "Company - Current period",
        75: "Company - Prior period",
    }

    banner_font = Font(bold=True, size=12, color="FFFFFF")
    banner_fill = PatternFill("solid", start_color="305496")
    banner_align = Alignment(horizontal="left", vertical="center", indent=1)

    for row, label in block_labels.items():
        cell = ws.cell(row=row, column=1)
        cell.value = label
        cell.font = banner_font
        cell.fill = banner_fill
        cell.alignment = banner_align
        # Blank out rows that were the repeated "Statement of changes in equity" banner
        # (rows row+1 and row+2) so they don't clutter the section. They were never
        # data rows so this is safe.
        for extra in (row + 1, row + 2):
            extra_cell = ws.cell(row=extra, column=1)
            extra_cell.value = None

    # Also ensure the column A widths are generous
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 56)

    # Widen column A and confirm other columns preserved
    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)
    print(f"Wrote Group SOCIE -> {DST}")


if __name__ == "__main__":
    build_socie_group()
