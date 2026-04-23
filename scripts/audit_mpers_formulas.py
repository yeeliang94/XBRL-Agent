#!/usr/bin/env python3
"""Complete MPERS formula audit report.

Compares every SUM formula emitted by ``scripts/generate_mpers_templates.py``
against the SSM MPERS calculation linkbase, honouring the per-link-role +
per-presentation-occurrence semantics introduced in commit c609d6c.

Run: ``python3 scripts/audit_mpers_formulas.py > audit_report.txt``

Run when a new SSM taxonomy version drops to catch any parent/child drift
in the generated bundle. The regression tests under
``tests/test_mpers_generator.py`` (``@pytest.mark.mpers_formulas``) are the
CI-level guarantee; this script is the broader sweep for humans.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl

from scripts.generate_mpers_templates import (
    _FIRST_BODY_ROW,
    _PRE_TO_CALC_ROLE,
    _pre_file_for_role,
    parse_calc_linkbase_grouped_for_pre_role,
    walk_role,
)

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATES_COMPANY = REPO_ROOT / "XBRL-template-MPERS/Company"
TEMPLATES_GROUP = REPO_ROOT / "XBRL-template-MPERS/Group"

# Mirrors scripts/generate_mpers_templates.py::_TEMPLATE_MAPPING (face/sub
# roles only — notes roles are kept out of the audit since they carry no
# calc linkbase).
_TEMPLATE_MAPPING: list[tuple[str, list[str]]] = [
    ("01-SOFP-CuNonCu.xlsx", ["210000", "210100"]),
    ("02-SOFP-OrderOfLiquidity.xlsx", ["220000", "220100"]),
    ("03-SOPL-Function.xlsx", ["310000", "310100"]),
    ("04-SOPL-Nature.xlsx", ["320000", "320100"]),
    ("05-SOCI-BeforeTax.xlsx", ["420000"]),
    ("06-SOCI-NetOfTax.xlsx", ["410000"]),
    ("07-SOCF-Indirect.xlsx", ["520000"]),
    ("08-SOCF-Direct.xlsx", ["510000"]),
    ("09-SOCIE.xlsx", ["610000"]),
    ("10-SoRE.xlsx", ["620000"]),
]

_SHEET_NAMES: dict[str, list[str]] = {
    "01-SOFP-CuNonCu.xlsx": ["SOFP-CuNonCu", "SOFP-Sub-CuNonCu"],
    "02-SOFP-OrderOfLiquidity.xlsx": ["SOFP-OrdOfLiq", "SOFP-Sub-OrdOfLiq"],
    "03-SOPL-Function.xlsx": ["SOPL-Function", "SOPL-Analysis-Function"],
    "04-SOPL-Nature.xlsx": ["SOPL-Nature", "SOPL-Analysis-Nature"],
    "05-SOCI-BeforeTax.xlsx": ["SOCI-BeforeOfTax"],
    "06-SOCI-NetOfTax.xlsx": ["SOCI-NetOfTax"],
    "07-SOCF-Indirect.xlsx": ["SOCF-Indirect"],
    "08-SOCF-Direct.xlsx": ["SOCF-Direct"],
    "09-SOCIE.xlsx": ["SOCIE"],
    "10-SoRE.xlsx": ["SoRE"],
}


def _parse_formula_refs(formula_str: str) -> list[tuple[int, int]]:
    """Parse an emitted SUM formula into ``[(row, weight), …]``.

    Formulas look like ``=1*B12+-1*B14+1*B16`` (the generator's style). Any
    reference not matching the ``<weight>*<col><row>`` pattern is ignored —
    cross-sheet refs use the form ``='SOFP-Sub-CuNonCu'!B50`` and aren't
    relevant to the same-sheet audit this script performs.
    """
    if not formula_str or not isinstance(formula_str, str) or not formula_str.startswith("="):
        return []
    # Normalise "-" to "+-" so a simple split on "+" works for signed terms.
    expr = formula_str[1:].replace("-", "+-")
    result: list[tuple[int, int]] = []
    for term in expr.split("+"):
        if not term:
            continue
        if "*" not in term:
            continue
        weight_s, ref = term.split("*", 1)
        try:
            weight = int(float(weight_s))
        except ValueError:
            continue
        match = re.match(r"^([A-Z]+)(\d+)$", ref)
        if not match:
            continue
        _col, row_s = match.groups()
        result.append((int(row_s), weight))
    return result


def _compute_expected_formulas(
    role_number: str,
) -> list[tuple[int, list[tuple[int, int]]]]:
    """Return ``[(xlsx_row, expected_children), …]`` for one role.

    Mirrors the per-presentation-occurrence + right-aligned assignment
    algorithm in ``scripts/generate_mpers_templates._inject_sum_formulas``.
    Children are resolved to their first-occurrence row (same as the
    emitter). Parents whose blocks have no resolvable children are skipped.
    """
    rows = walk_role(_pre_file_for_role(role_number))
    calc_blocks = parse_calc_linkbase_grouped_for_pre_role(role_number)

    # concept_id -> list of xlsx rows (presentation order).
    concept_to_rows: dict[str, list[int]] = {}
    first_occurrence: dict[str, int] = {}
    for idx, (_depth, concept_id, _label, _abs) in enumerate(rows):
        xlsx_row = _FIRST_BODY_ROW + idx
        concept_to_rows.setdefault(concept_id, []).append(xlsx_row)
        first_occurrence.setdefault(concept_id, xlsx_row)

    # Per-parent: ordered list of (link_role, children) from the calc file.
    parent_to_blocks: dict[str, list[list[tuple[str, int]]]] = {}
    for _role, calc_map in calc_blocks:
        for parent, children in calc_map.items():
            parent_to_blocks.setdefault(parent, []).append(children)

    expected: list[tuple[int, list[tuple[int, int]]]] = []
    for parent, blocks in parent_to_blocks.items():
        parent_rows = concept_to_rows.get(parent, [])
        if not parent_rows:
            continue
        n = min(len(parent_rows), len(blocks))
        row_start = len(parent_rows) - n  # right-aligned
        block_start = len(blocks) - n
        for i in range(n):
            xlsx_row = parent_rows[row_start + i]
            children = blocks[block_start + i]
            resolved: list[tuple[int, int]] = []
            for child_concept, weight in children:
                child_row = first_occurrence.get(child_concept)
                if child_row is not None:
                    resolved.append((child_row, weight))
            if resolved:
                expected.append((xlsx_row, resolved))
    return expected


def _value_columns_for(filename: str, level: str) -> Optional[list[str]]:
    """Columns the generator writes formulas into, or ``None`` to skip.

    Returns ``None`` for sheets that use a non-standard layout where the
    calc-based audit doesn't apply — specifically the Group SOCIE 4-block
    layout, which replaces the calc-driven single-row formulas entirely.
    """
    if filename == "09-SOCIE.xlsx" and level == "group":
        return None
    if level == "company":
        return ["B", "C"]
    return ["B", "C", "D", "E"]


def audit_template(template_path: Path, level: str) -> dict:
    """Audit one xlsx by comparing each expected formula against the cells."""
    filename = template_path.name
    role_numbers: Optional[list[str]] = None
    for fname, rns in _TEMPLATE_MAPPING:
        if fname == filename:
            role_numbers = rns
            break
    if role_numbers is None:
        return {"error": f"Unknown template {filename}"}

    sheet_names = _SHEET_NAMES.get(filename, [])
    wb = openpyxl.load_workbook(template_path)
    result: dict = {"filename": filename, "level": level, "sheets": {}}

    value_cols = _value_columns_for(filename, level)
    if value_cols is None:
        # e.g. Group SOCIE — skip the calc-based audit, record the reason so
        # the summary counts it as a deliberate skip, not a missing audit.
        for sheet_name in sheet_names:
            result["sheets"][sheet_name] = {"skipped": "group_socie_4block_layout"}
        return result

    for sheet_name, role_number in zip(sheet_names, role_numbers):
        if sheet_name not in wb.sheetnames:
            result["sheets"][sheet_name] = {"error": "Sheet not found"}
            continue

        ws = wb[sheet_name]
        expected = _compute_expected_formulas(role_number)

        sheet_result: dict = {
            "expected_formulas": len(expected),
            "formulas_correct": 0,
            "formulas_wrong": 0,
            "formulas_missing": 0,
            "issues": [],
        }

        for parent_row, expected_children in expected:
            expected_set = set(expected_children)
            for col_letter in value_cols:
                col_idx = ord(col_letter) - ord("A") + 1
                formula = ws.cell(row=parent_row, column=col_idx).value
                if not (isinstance(formula, str) and formula.startswith("=")):
                    sheet_result["formulas_missing"] += 1
                    sheet_result["issues"].append({
                        "type": "missing",
                        "cell": f"{col_letter}{parent_row}",
                        "expected": sorted(expected_set),
                    })
                    continue

                # Only compare refs to THIS column — cross-column refs are
                # invalid for the same-axis audit.
                actual_refs = _parse_formula_refs(formula)
                wrong_col = [f"{col_letter}{parent_row}"] if col_letter not in formula else []
                if wrong_col:
                    # A formula exists but references a different column — rare,
                    # but flag it rather than silently ignoring.
                    sheet_result["formulas_wrong"] += 1
                    sheet_result["issues"].append({
                        "type": "wrong_column",
                        "cell": f"{col_letter}{parent_row}",
                        "formula": formula[:120],
                    })
                    continue

                actual_set = set(actual_refs)
                if actual_set == expected_set:
                    sheet_result["formulas_correct"] += 1
                else:
                    sheet_result["formulas_wrong"] += 1
                    sheet_result["issues"].append({
                        "type": "mismatch",
                        "cell": f"{col_letter}{parent_row}",
                        "formula": formula[:120],
                        "expected_only": sorted(expected_set - actual_set),
                        "actual_only": sorted(actual_set - expected_set),
                    })

        result["sheets"][sheet_name] = sheet_result

    return result


def _accumulate(totals: dict, sheet_result: dict) -> None:
    totals["sheets"] += 1
    totals["expected_formulas"] += sheet_result.get("expected_formulas", 0)
    totals["formulas_correct"] += sheet_result.get("formulas_correct", 0)
    totals["formulas_wrong"] += sheet_result.get("formulas_wrong", 0)
    totals["formulas_missing"] += sheet_result.get("formulas_missing", 0)


def _print_sheet(filename: str, sheet_name: str, sheet_result: dict) -> None:
    if "skipped" in sheet_result:
        print(f"{filename} [{sheet_name}]: SKIPPED ({sheet_result['skipped']})")
        return
    if "error" in sheet_result:
        print(f"{filename} [{sheet_name}]: ERROR ({sheet_result['error']})")
        return
    correct = sheet_result["formulas_correct"]
    wrong = sheet_result["formulas_wrong"]
    missing = sheet_result["formulas_missing"]
    expected = sheet_result["expected_formulas"]
    status = "OK" if (wrong == 0 and missing == 0) else "ISSUES"
    print(
        f"{filename} [{sheet_name}]: {status} "
        f"(expected={expected}, correct={correct}, wrong={wrong}, missing={missing})"
    )
    for issue in sheet_result.get("issues", [])[:3]:
        kind = issue["type"].upper()
        cell = issue["cell"]
        extra = ""
        if issue["type"] == "mismatch":
            extra = f"  expected_only={issue['expected_only']} actual_only={issue['actual_only']}"
        print(f"    - {kind}: {cell}{extra}")


def main() -> None:
    print("=" * 100)
    print("MPERS FORMULA AUDIT — DETAILED REPORT")
    print("=" * 100)

    def _new_totals() -> dict:
        return {
            "templates": 0,
            "sheets": 0,
            "expected_formulas": 0,
            "formulas_correct": 0,
            "formulas_wrong": 0,
            "formulas_missing": 0,
        }

    company_totals = _new_totals()
    group_totals = _new_totals()

    print("\nCOMPANY LEVEL AUDIT")
    print("-" * 100)
    for filename, _ in _TEMPLATE_MAPPING:
        template_path = TEMPLATES_COMPANY / filename
        if not template_path.exists():
            print(f"{filename}: TEMPLATE NOT FOUND")
            continue
        company_totals["templates"] += 1
        result = audit_template(template_path, level="company")
        for sheet_name, sheet_result in result.get("sheets", {}).items():
            _print_sheet(filename, sheet_name, sheet_result)
            _accumulate(company_totals, sheet_result)

    print("\nGROUP LEVEL AUDIT")
    print("-" * 100)
    for filename, _ in _TEMPLATE_MAPPING:
        template_path = TEMPLATES_GROUP / filename
        if not template_path.exists():
            print(f"{filename}: TEMPLATE NOT FOUND")
            continue
        group_totals["templates"] += 1
        result = audit_template(template_path, level="group")
        for sheet_name, sheet_result in result.get("sheets", {}).items():
            _print_sheet(filename, sheet_name, sheet_result)
            _accumulate(group_totals, sheet_result)

    print()
    print("=" * 100)
    print("SUMMARY")
    print("=" * 100)
    for label, totals in (("COMPANY", company_totals), ("GROUP", group_totals)):
        print(
            f"\n{label}: templates={totals['templates']} sheets={totals['sheets']} "
            f"expected={totals['expected_formulas']} "
            f"correct={totals['formulas_correct']} "
            f"wrong={totals['formulas_wrong']} "
            f"missing={totals['formulas_missing']}"
        )
    clean = (
        company_totals["formulas_wrong"] == 0
        and company_totals["formulas_missing"] == 0
        and group_totals["formulas_wrong"] == 0
        and group_totals["formulas_missing"] == 0
    )
    print("\nRESULT: ALL FORMULAS CORRECT" if clean else "\nRESULT: ISSUES FOUND")


if __name__ == "__main__":
    main()
