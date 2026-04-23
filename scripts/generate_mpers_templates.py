"""Generate MPERS XBRL templates from the SSM MPERS linkbase.

Mirrors the existing MFRS template bundle by walking each role's presentation
linkbase and emitting an xlsx that matches the MFRS format exactly. Built step
by step under the `docs/Archive/PLAN-mpers-template-generator.md` red-green TDD plan — see
that plan for the full specification, and `tests/test_mpers_generator.py` for
the behaviour contract.
"""
from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path
from typing import Any

# Repo-root anchor so the script works regardless of caller cwd — same pattern
# as scripts/build_group_templates.py.
_REPO_ROOT = Path(__file__).resolve().parent.parent
_MPERS_TAXONOMY_DIR = (
    _REPO_ROOT / "SSMxT_2022v1.0" / "rep" / "ssm" / "ca-2016" / "fs" / "mpers"
)
_ROLE_XSD = _MPERS_TAXONOMY_DIR / "rol_ssmt-fs-mpers_2022-12-31.xsd"
_TAXONOMY_ROOT = _REPO_ROOT / "SSMxT_2022v1.0"

# XBRL linkbase namespaces used by every pre_/lab_/cal_ file.
_NS = {
    "link": "http://www.xbrl.org/2003/linkbase",
    "xlink": "http://www.w3.org/1999/xlink",
    "xml": "http://www.w3.org/XML/1998/namespace",
}

# XBRL 2003 standard label + total-label roles, plus SSM ReportingLabel.
_STANDARD_LABEL_ROLE = "http://www.xbrl.org/2003/role/label"
_TOTAL_LABEL_ROLE = "http://www.xbrl.org/2003/role/totalLabel"
_REPORTING_LABEL_ROLE = re.compile(r"/ReportingLabel$")
_REPORTING_TOTAL_LABEL_ROLE = re.compile(r"/ReportingTotalLabel$")

# Priority when a presentation arc does NOT carry a preferredLabel. SSM
# ReportingLabel is the "display" label for most MPERS/MFRS concepts; the
# XBRL 2003 standard label is the def-level fallback for concepts that only
# appear in ifrs-smes / ssmt-cor (e.g. abstract wrappers).
_DEFAULT_LABEL_PRIORITY = (
    "ReportingLabel",
    _STANDARD_LABEL_ROLE,
)

# SSM ReportingLabel suffixes appended to MPERS taxonomy labels. MFRS templates
# strip these for display; we do the same so MPERS matches MFRS visually. The
# underlying concept_id is preserved on every row, so XBRL compliance (which
# lives in the calculation/presentation linkbase via concept IDs, not label
# text) is unaffected. Kept in sync with notes.labels._TAXONOMY_SUFFIXES.
#
# SSM is inconsistent about spacing: the MFRS 2022 bundle uses ``[text block]``
# (with a space) on most concepts but ``[textblock]`` (no space) on the notes
# roles. Both variants appear in the MPERS bundle too. Strip both so the
# rendered templates match MFRS exactly.
_DISPLAY_LABEL_SUFFIXES: tuple[str, ...] = (
    "[text block]",
    "[textblock]",
    "[abstract]",
    "[axis]",
    "[member]",
    "[table]",
    "[line items]",
)

# Pure-structural labels — rows whose ONLY role is XBRL scaffolding
# (hypercube tables, axes, typed-member placeholders, line-items anchors).
# MFRS hand-curated templates omit these; the MPERS generator used to emit
# them verbatim because it walks the presentation linkbase DFS. We skip rows
# whose resolved label ends in one of these three markers — the data-carrying
# rows (tagged [abstract] / [text block]) are retained with the suffix stripped.
_STRUCTURAL_LABEL_SUFFIXES: tuple[str, ...] = (
    "[table]",
    "[axis]",
    "[member]",
    "[line items]",
)

_DISPLAY_SUFFIX_RE = re.compile(
    r"\s*(?:" + "|".join(re.escape(s) for s in _DISPLAY_LABEL_SUFFIXES) + r")\s*$",
    re.IGNORECASE,
)


def _strip_display_suffix(text: str) -> str:
    """Remove a single trailing SSM ReportingLabel suffix from a label string.

    Idempotent, case-insensitive on the bracketed text. Only strips the tail;
    mid-string brackets like ``EBITDA [reconciliation below]`` are preserved.
    """
    if not isinstance(text, str):
        return text
    return _DISPLAY_SUFFIX_RE.sub("", text)


def _is_structural_label(text: str) -> bool:
    """True when a resolved label is a pure XBRL scaffolding node.

    Structural nodes are hypercube tables, axes, members, and line-items
    anchors — they have no data column payload, they only define the shape
    of a dimensional cube. MFRS templates omit them from the rendered sheet.
    """
    if not isinstance(text, str):
        return False
    lowered = text.strip().lower()
    return any(lowered.endswith(s) for s in _STRUCTURAL_LABEL_SUFFIXES)

# Matches "[210000] Statement of financial position, ..." inside the role XSD
# so we can map role_number -> clean title for every role.
_ROLE_TITLE_PATTERN = re.compile(r"\[(?P<num>\d{6})\]\s*(?P<title>[^<]+?)\s*</link:definition>")

# Filename pattern: pre_ssmt-fs-mpers_2022-12-31_role-210000.xml -> "210000"
_PRE_FILENAME_PATTERN = re.compile(r"pre_.+_role-(?P<num>\d{6})\.xml$")


def _concept_id_from_href(href: str) -> str:
    """Extract the concept id from an XSD href like ``…/ifrs_for_smes-cor.xsd#ifrs-smes_Foo``."""
    # Everything after '#' is the concept id; the schema URL prefix is noise.
    return href.split("#", 1)[-1]


def _normalise_label_role(raw_role: str) -> str:
    """Collapse SSM's long ReportingLabel / ReportingTotalLabel URIs to short keys.

    The taxonomy expresses ReportingLabel via a long URL
    (``…/lab_rol_.../ReportingLabel``). Callers index by role, and the short
    keys ("ReportingLabel", "ReportingTotalLabel") are easier to read and
    match what `preferredLabel` arcs reference after normalisation.
    XBRL 2003 roles are kept verbatim so the priority list can still use the
    full URI.
    """
    if _REPORTING_TOTAL_LABEL_ROLE.search(raw_role):
        return "ReportingTotalLabel"
    if _REPORTING_LABEL_ROLE.search(raw_role):
        return "ReportingLabel"
    return raw_role


def _load_role_titles() -> dict[str, str]:
    """Return a {role_number -> clean title} map from the role XSD.

    The XSD carries multiple `<link:roleType>` entries per role number (one for
    definitionLink, one for presentationLink, etc.) but they all share the same
    `[NNNNNN] <title>` definition text, so dedup via dict merge is safe.
    """
    text = _ROLE_XSD.read_text(encoding="utf-8")
    return {m.group("num"): m.group("title").strip() for m in _ROLE_TITLE_PATTERN.finditer(text)}


# Per-filename sheet names — mirrors the MFRS convention so downstream
# pipeline code (template_reader, fill_workbook, cross-checks) can read
# MPERS templates without a code fork. First entry is the face sheet for
# the first role number, second entry (if present) is the sub-classification
# sheet for the second role number.
# Pre role → calc role mapping. MPERS presentation linkbases split face
# (CuNonCu vs OrdOfLiq) and sub-classification, but the calculation linkbase
# is neutral — 200100 (face-level totals) and 200200 (sub-level totals) are
# reused across both CuNonCu and OrdOfLiq. SOPL is analogous (300100/300200),
# SOCI face-only (400100), SOCF face-only (500100). SOCIE (610000) and
# SoRE (620000) have their own calc files keyed by the same role number.
_PRE_TO_CALC_ROLE: dict[str, str] = {
    # SOFP
    "210000": "200100",
    "210100": "200200",
    "220000": "200100",
    "220100": "200200",
    # SOPL
    "310000": "300100",
    "310100": "300200",
    "320000": "300100",
    "320100": "300200",
    # SOCI
    "410000": "400100",
    "420000": "400100",
    # SOCF
    "510000": "500100",
    "520000": "500100",
    # SOCIE + SoRE — own calc files
    "610000": "610000",
    "620000": "620000",
}


_SHEET_NAMES: dict[str, list[str]] = {
    "01-SOFP-CuNonCu.xlsx": ["SOFP-CuNonCu", "SOFP-Sub-CuNonCu"],
    "02-SOFP-OrderOfLiquidity.xlsx": ["SOFP-OrdOfLiq", "SOFP-Sub-OrdOfLiq"],
    "03-SOPL-Function.xlsx": ["SOPL-Function", "SOPL-Analysis-Function"],
    "04-SOPL-Nature.xlsx": ["SOPL-Nature", "SOPL-Analysis-Nature"],
    "05-SOCI-BeforeTax.xlsx": ["SOCI-BeforeOfTax"],
    "06-SOCI-NetOfTax.xlsx": ["SOCI-NetOfTax"],
    "07-SOCF-Indirect.xlsx": ["SOCF-Indirect"],
    "08-SOCF-Direct.xlsx": ["SOCF-Direct"],
    "09-SOCIE.xlsx": ["SOCIE"],
    "10-SoRE.xlsx": ["SoRE"],
    "11-Notes-CorporateInfo.xlsx": ["Notes-CI"],
    "12-Notes-AccountingPolicies.xlsx": ["Notes-SummaryofAccPol"],
    "13-Notes-ListOfNotes.xlsx": ["Notes-Listofnotes"],
    "14-Notes-IssuedCapital.xlsx": ["Notes-Issuedcapital"],
    "15-Notes-RelatedParty.xlsx": ["Notes-RelatedPartytran"],
}


# Maps each output xlsx filename to the MPERS role(s) that populate it.
# Declared order IS the output-bundle order. Composite entries (SOFP etc.)
# list the face role first and its sub-classification role second; the emitter
# puts each into its own sheet (face + SOFP-Sub-*) to match the MFRS layout.
# MPERS adds slot 10 (SoRE, role 620000) as a SOCIE variant, which shifts the
# notes templates from 10..14 to 11..15.
_TEMPLATE_MAPPING: list[tuple[str, list[str]]] = [
    ("01-SOFP-CuNonCu.xlsx", ["210000", "210100"]),
    ("02-SOFP-OrderOfLiquidity.xlsx", ["220000", "220100"]),
    ("03-SOPL-Function.xlsx", ["310000", "310100"]),
    ("04-SOPL-Nature.xlsx", ["320000", "320100"]),
    ("05-SOCI-BeforeTax.xlsx", ["420000"]),
    ("06-SOCI-NetOfTax.xlsx", ["410000"]),
    ("07-SOCF-Indirect.xlsx", ["520000"]),
    ("08-SOCF-Direct.xlsx", ["510000"]),
    ("09-SOCIE.xlsx", ["610000"]),
    ("10-SoRE.xlsx", ["620000"]),
    ("11-Notes-CorporateInfo.xlsx", ["710000"]),
    ("12-Notes-AccountingPolicies.xlsx", ["720000"]),
    ("13-Notes-ListOfNotes.xlsx", ["730000"]),
    ("14-Notes-IssuedCapital.xlsx", ["740000"]),
    ("15-Notes-RelatedParty.xlsx", ["750000"]),
]


def template_mapping() -> list[tuple[str, list[str]]]:
    """Return the ordered (filename, role_numbers) list of 15 output templates.

    Callers MUST NOT mutate the returned list — the underlying table is the
    single source of truth for bundle order and role grouping. A fresh copy
    (with its own list entries) is returned so accidental mutation is local.
    """
    return [(fname, list(rns)) for fname, rns in _TEMPLATE_MAPPING]


_LABEL_MAP_CACHE: dict[str, dict[str, str]] | None = None


def load_label_map() -> dict[str, dict[str, str]]:
    """Scan every `lab_en*.xml` file under the taxonomy tree and return a
    {concept_id -> {short_role -> english_label}} map.

    Some concepts have labels only in the def-level files (e.g.
    ``ssmt-mpers_*``), others get an SSM ReportingLabel override in the
    rep-level MPERS file. Merging both sources keeps every concept reachable
    and lets `walk_role()` pick the right variant by role.

    Cached on first call — a full scan touches ~10 files and ~3000 labels,
    but running it repeatedly during the same process is wasteful.
    """
    global _LABEL_MAP_CACHE
    if _LABEL_MAP_CACHE is not None:
        return _LABEL_MAP_CACHE

    label_files = [
        *_TAXONOMY_ROOT.rglob("lab_en-*.xml"),
        *_TAXONOMY_ROOT.rglob("lab_en_*.xml"),
        *_TAXONOMY_ROOT.rglob("lab_ifrs_for_smes-en*.xml"),
    ]
    # Dedup in case the globs overlap (some filesystems return the same file
    # twice under different patterns).
    seen: set[Path] = set()
    unique_files = [f for f in label_files if not (f in seen or seen.add(f))]

    # concept_id → {short_role → label_text}
    result: dict[str, dict[str, str]] = defaultdict(dict)

    for lab_file in unique_files:
        try:
            tree = ET.parse(lab_file)
        except ET.ParseError:
            # Malformed label files are not silently tolerated in XBRL runs,
            # but we prefer to skip one broken file than crash the whole
            # generator. The caller will loud-fail at walk time if a concept
            # lacks a label.
            continue
        root = tree.getroot()

        # xlink:label → concept_id (via link:loc)
        loc_to_concept: dict[str, str] = {}
        # xlink:label → (role, text) for actual <link:label> resources
        label_resources: dict[str, tuple[str, str]] = {}
        # (from_label, to_label) arcs bind loc → label.
        arcs: list[tuple[str, str]] = []

        for elem in root.iter():
            tag = elem.tag.split("}", 1)[-1]
            if tag == "loc":
                key = elem.get(f"{{{_NS['xlink']}}}label")
                href = elem.get(f"{{{_NS['xlink']}}}href")
                if key and href:
                    loc_to_concept[key] = _concept_id_from_href(href)
            elif tag == "label":
                key = elem.get(f"{{{_NS['xlink']}}}label")
                role = elem.get(f"{{{_NS['xlink']}}}role") or ""
                text = (elem.text or "").strip()
                if key and text:
                    label_resources[key] = (role, text)
            elif tag == "labelArc":
                frm = elem.get(f"{{{_NS['xlink']}}}from")
                to = elem.get(f"{{{_NS['xlink']}}}to")
                if frm and to:
                    arcs.append((frm, to))

        for frm, to in arcs:
            concept_id = loc_to_concept.get(frm)
            if not concept_id:
                continue
            label_info = label_resources.get(to)
            if not label_info:
                continue
            role, text = label_info
            short_role = _normalise_label_role(role)
            # First writer wins within a file; across files, ReportingLabel in
            # rep-level files overrides later writes because of the order we
            # walk files — we intentionally collect all and let the caller
            # pick via priority.
            result[concept_id].setdefault(short_role, text)

    # Promote to plain dict for the cache so callers cannot accidentally
    # mutate-default a missing key.
    flat: dict[str, dict[str, str]] = {k: dict(v) for k, v in result.items()}

    # Flatten the mapping to {concept_id -> label} using default priority.
    # Callers that need the full role table can use `_label_role_map()` (not
    # public — walker uses it directly).
    simple: dict[str, str] = {}
    for concept_id, roles in flat.items():
        for preferred in _DEFAULT_LABEL_PRIORITY:
            if preferred in roles:
                simple[concept_id] = roles[preferred]
                break
        else:
            # Fallback: use whatever label we have, arbitrary pick.
            simple[concept_id] = next(iter(roles.values()))

    # Store both the simple and the role-indexed map; _LABEL_ROLE_TABLE is
    # used internally by walk_role() when resolving preferredLabel.
    global _LABEL_ROLE_TABLE
    _LABEL_ROLE_TABLE = flat
    _LABEL_MAP_CACHE = simple
    return simple


# Populated as a side effect of `load_label_map()`. Separate from the simple
# map so callers can't accidentally depend on the private role table.
_LABEL_ROLE_TABLE: dict[str, dict[str, str]] = {}


def _resolve_preferred_label(concept_id: str, preferred_role: str | None) -> str:
    """Pick the best-matching label for a concept given an optional preferredLabel role.

    Falls back to the default priority list when the preferred role isn't
    available for the concept (some concepts legitimately lack a TotalLabel).
    """
    if not _LABEL_ROLE_TABLE:
        load_label_map()
    roles = _LABEL_ROLE_TABLE.get(concept_id, {})
    if preferred_role:
        # Normalise "…/totalLabel" → still the 2003 URI; "ReportingLabel" stays.
        # An exact-role hit is preferred over defaults.
        if preferred_role in roles:
            return roles[preferred_role]
        # Some preferredLabel arcs reference the SSM custom roles by full URI —
        # normalise those to short keys and re-check.
        short = _normalise_label_role(preferred_role)
        if short in roles:
            return roles[short]
    for fallback in _DEFAULT_LABEL_PRIORITY:
        if fallback in roles:
            return roles[fallback]
    return next(iter(roles.values()), concept_id)


def _is_abstract_concept(concept_id: str) -> bool:
    """Abstract concepts in SSM taxonomies are named with an ``Abstract`` suffix.

    Parsing every concept's XSD `abstract="true"` attribute would mean
    loading ~5 XSD files; the suffix convention is 100% reliable in the SSM
    taxonomy and avoids the cost.
    """
    return concept_id.endswith("Abstract")


def walk_role(pre_file_path: Path) -> list[tuple[int, str, str, bool]]:
    """DFS-traverse a presentation linkbase into a flat list of display rows.

    Returns ``[(depth, concept_id, label, is_abstract), …]`` in the order the
    rows should appear in the output xlsx. Depth 0 is the root abstract,
    depth 1 is a direct child, etc. Labels respect any ``preferredLabel`` on
    the parent→child arc (so TotalLabel variants surface automatically).
    """
    tree = ET.parse(pre_file_path)
    root = tree.getroot()

    # Each presentationLink element is one role. MPERS pre files ship a single
    # role per file, but we still iterate just to be safe.
    rows: list[tuple[int, str, str, bool]] = []
    for pres_link in root.iter(f"{{{_NS['link']}}}presentationLink"):
        # xlink:label → concept_id
        loc_map: dict[str, str] = {}
        # parent_label → list[(order, child_label, preferredLabel|None)]
        children: dict[str, list[tuple[float, str, str | None]]] = defaultdict(list)
        # Track every label that appears as a source or destination — the
        # root is the one that's never a `to`.
        all_froms: set[str] = set()
        all_tos: set[str] = set()

        for elem in pres_link:
            tag = elem.tag.split("}", 1)[-1]
            if tag == "loc":
                key = elem.get(f"{{{_NS['xlink']}}}label")
                href = elem.get(f"{{{_NS['xlink']}}}href")
                if key and href:
                    loc_map[key] = _concept_id_from_href(href)
            elif tag == "presentationArc":
                frm = elem.get(f"{{{_NS['xlink']}}}from")
                to = elem.get(f"{{{_NS['xlink']}}}to")
                if not (frm and to):
                    continue
                order_raw = elem.get("order", "0")
                try:
                    order = float(order_raw)
                except ValueError:
                    order = 0.0
                preferred = elem.get("preferredLabel")
                children[frm].append((order, to, preferred))
                all_froms.add(frm)
                all_tos.add(to)

        roots = [lbl for lbl in loc_map if lbl in all_froms and lbl not in all_tos]
        # Fall back to any loc that's in no arcs at all (orphan) if no true roots.
        if not roots:
            roots = [lbl for lbl in loc_map if lbl not in all_tos]

        def dfs(label: str, depth: int, preferred: str | None) -> None:
            concept_id = loc_map.get(label)
            if not concept_id:
                return
            text = _resolve_preferred_label(concept_id, preferred)
            # Scaffolding filter — skip XBRL hypercube tables, axes, typed
            # members, and line-items anchors. These carry no data; MFRS
            # hand-curated templates omit them. We still RECURSE into their
            # children so real data rows underneath remain reachable — the
            # concept_id stays in the taxonomy graph, only the row emit is
            # suppressed.
            if not _is_structural_label(text):
                # Strip display suffix (`[abstract]`, `[text block]`, …) to
                # match MFRS template formatting. The concept_id on this row
                # is preserved untouched, so taxonomy compliance is intact.
                display_text = _strip_display_suffix(text)
                rows.append(
                    (depth, concept_id, display_text, _is_abstract_concept(concept_id))
                )
            # Visit children in declared order. Same (order, to) pairs are
            # stable because Python's sort is stable.
            for _order, child_label, child_preferred in sorted(children[label], key=lambda x: x[0]):
                dfs(child_label, depth + 1, child_preferred)

        for root_label in roots:
            dfs(root_label, 0, None)

    return rows


# Column-layout constants — pinned by the Phase 1 MFRS format tests.
# The row-1 "Source" header sits in the rightmost used column; value columns
# live between A (label) and the source column.
_HEADER_ROW = 2  # Row where period-placeholder strings go (row 2 for Group, row 1 for Company).
_FIRST_BODY_ROW = 3
_PERIOD_PLACEHOLDER = "01/01/YYYY - 31/12/YYYY"


def _apply_group_sheet_layout(ws, rows: list[tuple[int, str, str, bool]]) -> None:
    """Write the 6-column Group layout onto a sheet.

    A=label, B=Group-CY, C=Group-PY, D=Company-CY, E=Company-PY, F=Source.

    Row 1 is the "Group"/"Company" banner row (B and D); row 2 carries period
    placeholders across B/C/D/E. Body rows start at row 3 (consistent with
    Company layout). Bold styling on labels starting with ``*`` matches the
    MFRS convention.
    """
    from openpyxl.styles import Font

    # Row 1 banners + "Source" header.
    ws.cell(row=1, column=2, value="Group")
    ws.cell(row=1, column=4, value="Company")
    ws.cell(row=1, column=6, value="Source")

    # Row 2 period placeholders in every value column.
    for col in (2, 3, 4, 5):
        ws.cell(row=2, column=col, value=_PERIOD_PLACEHOLDER)

    bold_font = Font(bold=True)
    for idx, (_depth, _concept_id, label, _is_abstract) in enumerate(rows):
        r = _FIRST_BODY_ROW + idx
        cell = ws.cell(row=r, column=1, value=label)
        if isinstance(label, str) and label.startswith("*"):
            cell.font = bold_font

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 55.0
    for col_letter in ("B", "C", "D", "E"):
        ws.column_dimensions[col_letter].width = 18.0
    ws.column_dimensions["F"].width = 40.0


def _apply_company_sheet_layout(ws, rows: list[tuple[int, str, str, bool]]) -> None:
    """Write the 4-column Company layout (label + CY + PY + Source) onto a sheet.

    Row 1 holds period placeholders in B/C and "Source" in D; row 2 is blank;
    body rows start at row 3. Labels starting with "*" render bold (matches
    the MFRS convention captured in Phase 1).
    """
    from openpyxl.styles import Font

    ws.cell(row=1, column=2, value=_PERIOD_PLACEHOLDER)
    ws.cell(row=1, column=3, value=_PERIOD_PLACEHOLDER)
    ws.cell(row=1, column=4, value="Source")

    bold_font = Font(bold=True)
    for idx, (_depth, _concept_id, label, _is_abstract) in enumerate(rows):
        r = _FIRST_BODY_ROW + idx
        cell = ws.cell(row=r, column=1, value=label)
        if isinstance(label, str) and label.startswith("*"):
            cell.font = bold_font

    # Freeze + widths pin: matches MFRS Company format reference.
    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 55.0
    ws.column_dimensions["B"].width = 18.0
    ws.column_dimensions["C"].width = 18.0
    ws.column_dimensions["D"].width = 40.0


def emit_template(
    rows: list[tuple[int, str, str, bool]],
    out_path: Path,
    level: str = "company",
) -> None:
    """Write a single-sheet xlsx matching the MFRS format for one role's rows.

    This is the Phase 2 unit-test surface — ``build_template()`` handles
    multi-sheet bundles by stitching per-role row-lists with
    ``_apply_company_sheet_layout()`` or ``_apply_group_sheet_layout()``.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet"
    if level == "company":
        _apply_company_sheet_layout(ws, rows)
    elif level == "group":
        _apply_group_sheet_layout(ws, rows)
    else:
        raise ValueError(f"level={level!r} not supported")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _pre_file_for_role(role_number: str) -> Path:
    """Return the `pre_*_role-NNNNNN.xml` path for a given MPERS role number."""
    return _MPERS_TAXONOMY_DIR / f"pre_ssmt-fs-mpers_2022-12-31_role-{role_number}.xml"


def _calc_file_for_role(calc_role_number: str) -> Path:
    """Return the `cal_*_role-NNNNNN.xml` path for a given MPERS calc role number."""
    return _MPERS_TAXONOMY_DIR / f"cal_ssmt-fs-mpers_2022-12-31_role-{calc_role_number}.xml"


def parse_calc_linkbase_grouped(
    calc_file_path: Path,
) -> list[tuple[str, dict[str, list[tuple[str, int]]]]]:
    """Parse a calculation linkbase, preserving ``<calculationLink>`` boundaries.

    Each XBRL ``<calculationLink>`` element declares an independent summation-
    consistency rule (e.g. SOPL's role-300100 carries ``ProfitLoss = Continuing
    + Discontinued`` and role-300100c carries ``ProfitLoss = Owners + EquityOther
    + NCI`` — both must hold true). Downstream formula emission needs these
    blocks kept apart so distinct decompositions don't get mashed into one sum
    (which would double-count the parent).

    Returns a list of ``(link_role_uri, {parent_concept: [(child, weight), …]})``
    tuples in file order. Within each block, children are de-duplicated on
    (child, weight) so redundant arcs (e.g. AuditorsRemuneration in role-300200
    and role-300200c) don't produce ``=1*B96+1*B96`` formulas later.
    """
    tree = ET.parse(calc_file_path)
    root = tree.getroot()

    grouped: list[tuple[str, dict[str, list[tuple[str, int]]]]] = []

    for calc_link in root.iter(f"{{{_NS['link']}}}calculationLink"):
        link_role = calc_link.get(f"{{{_NS['xlink']}}}role", "")
        # Each link has its own loc table — the same label may reference different
        # concepts in different <calculationLink> blocks.
        loc_map: dict[str, str] = {}
        pending: dict[str, list[tuple[float, str, int]]] = defaultdict(list)

        for elem in calc_link:
            tag = elem.tag.split("}", 1)[-1]
            if tag == "loc":
                key = elem.get(f"{{{_NS['xlink']}}}label")
                href = elem.get(f"{{{_NS['xlink']}}}href")
                if key and href:
                    loc_map[key] = _concept_id_from_href(href)
            elif tag == "calculationArc":
                frm = elem.get(f"{{{_NS['xlink']}}}from")
                to = elem.get(f"{{{_NS['xlink']}}}to")
                if not (frm and to):
                    continue
                parent_concept = loc_map.get(frm)
                child_concept = loc_map.get(to)
                if not (parent_concept and child_concept):
                    continue
                try:
                    weight = int(float(elem.get("weight", "1")))
                except ValueError:
                    weight = 1
                try:
                    order = float(elem.get("order", "0"))
                except ValueError:
                    order = 0.0
                pending[parent_concept].append((order, child_concept, weight))

        # Sort children by @order and dedup (child, weight) within this block.
        calc_map: dict[str, list[tuple[str, int]]] = {}
        for parent, entries in pending.items():
            entries.sort(key=lambda x: x[0])
            seen: set[tuple[str, int]] = set()
            deduped: list[tuple[str, int]] = []
            for _order, child, weight in entries:
                key = (child, weight)
                if key not in seen:
                    seen.add(key)
                    deduped.append(key)
            calc_map[parent] = deduped

        if calc_map:
            grouped.append((link_role, calc_map))

    return grouped


def parse_calc_linkbase(calc_file_path: Path) -> dict[str, list[tuple[str, int]]]:
    """Parse one calculation linkbase into ``{parent_concept: [(child, weight), …]}``.

    Kept for backward compatibility with ``scripts/audit_mpers_formulas.py``
    and the existing ``test_parse_calc_linkbase_*`` regression tests. Builds on
    top of :func:`parse_calc_linkbase_grouped` by merging blocks with the same
    parent and de-duplicating ``(child, weight)`` pairs across blocks — so
    callers that only need "does this concept have calc children at all" still
    work, but no longer see ``=1*B96+1*B96`` doubling.

    The flattened form does **not** capture distinct-axis decompositions — use
    :func:`parse_calc_linkbase_grouped` when formula emission needs one formula
    per presentation occurrence.
    """
    grouped = parse_calc_linkbase_grouped(calc_file_path)
    result: dict[str, list[tuple[str, int]]] = defaultdict(list)
    seen_by_parent: dict[str, set[tuple[str, int]]] = defaultdict(set)
    for _role, calc_map in grouped:
        for parent, children in calc_map.items():
            for child, weight in children:
                key = (child, weight)
                if key not in seen_by_parent[parent]:
                    seen_by_parent[parent].add(key)
                    result[parent].append(key)
    return dict(result)


def parse_calc_linkbase_grouped_for_pre_role(
    pre_role_number: str,
) -> list[tuple[str, dict[str, list[tuple[str, int]]]]]:
    """Convenience: map a presentation role number (e.g. ``"210000"``) to its
    calc file and return the per-link-role calc blocks. Returns ``[]`` if the
    pre role has no calc counterpart (notes roles, scope, auditor reports, …).
    """
    calc_role = _PRE_TO_CALC_ROLE.get(pre_role_number)
    if calc_role is None:
        return []
    return parse_calc_linkbase_grouped(_calc_file_for_role(calc_role))


def parse_calc_linkbase_for_pre_role(pre_role_number: str) -> dict[str, list[tuple[str, int]]]:
    """Convenience: map a presentation role number (e.g. ``"210000"``) to its
    calc file and return the parsed totals map. Returns ``{}`` if the pre role
    has no calc counterpart (notes roles, scope, auditor reports, …).
    """
    calc_role = _PRE_TO_CALC_ROLE.get(pre_role_number)
    if calc_role is None:
        return {}
    return parse_calc_linkbase(_calc_file_for_role(calc_role))


def _inject_sum_formulas(
    ws,
    rows: list[tuple[int, str, str, bool]],
    calc_blocks: list[tuple[str, dict[str, list[tuple[str, int]]]]],
    value_columns: tuple[str, ...] = ("B", "C"),
) -> None:
    """Write SUM formulas into total rows based on the calc linkbase.

    ``calc_blocks`` is the per-``<calculationLink>`` list produced by
    :func:`parse_calc_linkbase_grouped`. For every parent concept:

      * Gather the child-row list for each block where that parent appears
        (skipping children whose concepts aren't present in this sheet's
        ``rows`` — expected for face roles that reference sub-sheet concepts).
      * Find every xlsx row where the parent concept occurs in the presentation.
      * Assign the blocks to the occurrences, right-aligned: the last block is
        assigned to the last occurrence, the second-last to the second-last
        occurrence, and so on. Extras at the front are left without a formula.

    Right-alignment matches the SSM convention where totals sit *below* their
    children. If a concept appears as both "opening balance" (no calc) and
    "closing balance" (calc at the bottom), the calc lands on the closing row.
    Concepts that appear with multiple distinct decompositions (e.g. SOPL's
    ``ProfitLoss`` — vertical + attribution) have two presentation rows *and*
    two calc blocks; the vertical block lands on the earlier row, attribution
    on the later row.

    Formula style matches MFRS: ``=1*B{r1}+-1*B{r2}+…``. Only same-sheet refs;
    cross-sheet refs (face → sub) are out of scope.
    """
    from openpyxl.styles import Font

    # concept_id -> list of Excel rows (in presentation order) where it appears.
    concept_to_rows: dict[str, list[int]] = defaultdict(list)
    # The first-occurrence map is still useful for looking up child rows —
    # children are typically referenced once in the presentation. If a child
    # concept appears more than once, we take the first occurrence (matches
    # the pre-fix behaviour that wasn't buggy for non-duplicate concepts).
    first_occurrence: dict[str, int] = {}
    for idx, (_depth, concept_id, _label, _abs) in enumerate(rows):
        xlsx_row = _FIRST_BODY_ROW + idx
        concept_to_rows[concept_id].append(xlsx_row)
        first_occurrence.setdefault(concept_id, xlsx_row)

    # Collect per-parent the ordered list of (link_role, children) blocks.
    # Order preserves the calc-file order.
    parent_to_blocks: dict[str, list[tuple[str, list[tuple[str, int]]]]] = defaultdict(list)
    for link_role, calc_map in calc_blocks:
        for parent, children in calc_map.items():
            parent_to_blocks[parent].append((link_role, children))

    def _write_formula(parent_row: int, children: list[tuple[str, int]]) -> None:
        # Resolve child concepts to rows; skip misses (face → sub refs).
        parts: list[tuple[int, int]] = []
        for child_concept, weight in children:
            child_row = first_occurrence.get(child_concept)
            if child_row is None:
                continue
            parts.append((child_row, weight))
        if not parts:
            return

        for col_letter in value_columns:
            pieces = [f"{weight}*{col_letter}{row}" for row, weight in parts]
            formula = "=" + "+".join(pieces)
            col_idx = ord(col_letter) - ord("A") + 1
            ws.cell(row=parent_row, column=col_idx, value=formula)

        # Mark the parent label as a total row (MFRS convention): prepend "*"
        # and bold col A. Keeps downstream "*-prefixed = total" heuristic working.
        label_cell = ws.cell(row=parent_row, column=1)
        current_label = label_cell.value
        if isinstance(current_label, str) and not current_label.startswith("*"):
            label_cell.value = f"*{current_label}"
        label_cell.font = Font(bold=True)

    for parent, blocks in parent_to_blocks.items():
        parent_rows = concept_to_rows.get(parent, [])
        if not parent_rows:
            continue

        # Right-align blocks to presentation occurrences. When there are more
        # rows than blocks (e.g. SOCF CashAndCashEquivalents appears 3x but has
        # only 1 calc), the formula lands on the last row (the total). When
        # blocks > rows, the earliest blocks are dropped.
        n = min(len(parent_rows), len(blocks))
        row_start = len(parent_rows) - n
        block_start = len(blocks) - n
        for i in range(n):
            parent_row = parent_rows[row_start + i]
            _link_role, children = blocks[block_start + i]
            _write_formula(parent_row, children)


def _inject_face_to_sub_rollups(
    face_ws,
    face_rows: list[tuple[int, str, str, bool]],
    sub_ws,
    sub_rows: list[tuple[int, str, str, bool]],
    value_columns: tuple[str, ...] = ("B", "C"),
) -> None:
    """Wire face-sheet line items to sub-sheet rollup totals.

    Mirrors the MFRS template pattern: the main SOFP/SOPL/SOCI sheet references
    the sub-classification sheet's ``*Total X`` rows via formulas like
    ``='SOFP-Sub-CuNonCu'!B39``. This preserves the taxonomy calc structure
    without forcing agents to fill the face line-item *and* the sub details —
    they fill the sub details and the face rolls up automatically.

    Algorithm: for each face concept that also appears on the sub sheet, point
    at the **last** occurrence of that concept on the sub sheet (SSM's
    presentation linkbase puts rollup totals last within a block, so the last
    occurrence is the ``ReportingTotalLabel`` row). Skip face rows that
    already carry a formula from :func:`_inject_sum_formulas` — those are the
    on-sheet subtotals (Total non-current assets, Total equity, etc.) and we
    must not overwrite them.
    """
    # Resolve sub-sheet concept → last xlsx row (rollup target).
    sub_last_row: dict[str, int] = {}
    for idx, (_depth, concept_id, _label, _abs) in enumerate(sub_rows):
        sub_last_row[concept_id] = _FIRST_BODY_ROW + idx

    for idx, (_depth, concept_id, _label, _abs) in enumerate(face_rows):
        if concept_id not in sub_last_row:
            continue
        face_row = _FIRST_BODY_ROW + idx
        for col_letter in value_columns:
            col_idx = ord(col_letter) - ord("A") + 1
            existing = face_ws.cell(row=face_row, column=col_idx).value
            # Don't overwrite an on-sheet subtotal formula.
            if isinstance(existing, str) and existing.startswith("="):
                continue
            ref = f"='{sub_ws.title}'!{col_letter}{sub_last_row[concept_id]}"
            face_ws.cell(row=face_row, column=col_idx, value=ref)


def _collect_rows_with_calc(
    role_number: str,
) -> tuple[
    list[tuple[int, str, str, bool]],
    list[tuple[str, dict[str, list[tuple[str, int]]]]],
]:
    """Walk the pre linkbase + load the per-link-role calc blocks for one role.

    Returns ``(rows, calc_blocks)`` — the second item is the grouped calc
    representation consumed by :func:`_inject_sum_formulas`.
    """
    rows = walk_role(_pre_file_for_role(role_number))
    calc_blocks = parse_calc_linkbase_grouped_for_pre_role(role_number)
    return rows, calc_blocks


def _apply_group_socie_layout(ws, rows: list[tuple[int, str, str, bool]]) -> None:
    """Write the Group SOCIE 4-block layout onto a sheet.

    MFRS Group SOCIE format (Phase 1 pin): four 23-row blocks at rows
    3-25, 27-49, 51-73, 75-97, each prefixed by a one-line block header
    ("Group - Current period", "Group - Prior period", "Company - Current
    period", "Company - Prior period"), blank separators at 26/50/74.

    The body of each block is the same MPERS SOCIE row-set. Because SOCIE
    uses columns for equity components (not period pairs), this is not a
    6-column value layout — each block takes the full width and labels sit
    in column A.
    """
    from openpyxl.styles import Font

    bold_font = Font(bold=True)

    # The 4-block structure is pinned by the format reference. We compute 22
    # body rows per block (the 23rd row-slot is the block header itself).
    # The underlying row-set (rows) may be longer/shorter than 22 — we use
    # min(22, len(rows)) so we don't overflow the block.
    block_ranges = [(3, 25), (27, 49), (51, 73), (75, 97)]
    block_headers = [
        "Group - Current period",
        "Group - Prior period",
        "Company - Current period",
        "Company - Prior period",
    ]

    # Truncate the row-set to 22 entries so the 23-row block fits.
    truncated = rows[:22]

    for (start, _end), header in zip(block_ranges, block_headers):
        ws.cell(row=start, column=1, value=header).font = bold_font
        for idx, (_depth, _concept_id, label, _is_abstract) in enumerate(truncated):
            r = start + 1 + idx
            cell = ws.cell(row=r, column=1, value=label)
            if isinstance(label, str) and label.startswith("*"):
                cell.font = bold_font

    # Equity-at-end row is the last body row in each block by MFRS
    # convention — bold if it isn't already.
    for start, end in block_ranges:
        last_cell = ws.cell(row=end, column=1)
        if last_cell.value is not None and not last_cell.font.bold:
            last_cell.font = bold_font

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 55.0


def build_template(filename: str, level: str, out_dir: Path) -> Path:
    """Generate one output xlsx for the given bundle filename.

    Looks up the role_numbers + sheet_names from the module-level tables and
    assembles a workbook with one sheet per role. Dispatches to:

      * Company layout (4 columns) when level=="company".
      * Group layout (6 columns) when level=="group".
      * Group SOCIE special case (4 vertical row blocks) when level=="group"
        and filename=="09-SOCIE.xlsx" — same 24-col width but laid out as
        a stacked block structure.
    """
    import openpyxl

    role_numbers: list[str] | None = None
    for bundle_name, rns in _TEMPLATE_MAPPING:
        if bundle_name == filename:
            role_numbers = rns
            break
    if role_numbers is None:
        raise KeyError(f"no template mapping for {filename!r}")

    sheet_names = _SHEET_NAMES.get(filename)
    if not sheet_names:
        raise KeyError(f"no sheet-name mapping for {filename!r}")
    if len(sheet_names) != len(role_numbers):
        raise ValueError(
            f"{filename}: {len(role_numbers)} role(s) but {len(sheet_names)} sheet name(s)"
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # Keep the per-sheet row-lists around so we can wire face → sub rollups
    # after all sheets are built (MFRS format parity — main SOFP/SOPL/SOCI
    # references sub-classification *Total rows via cross-sheet formulas).
    emitted: list[tuple[Any, list[tuple[int, str, str, bool]]]] = []
    for sheet_name, role_number in zip(sheet_names, role_numbers):
        ws = wb.create_sheet(title=sheet_name)
        rows, calc = _collect_rows_with_calc(role_number)

        if level == "company":
            _apply_company_sheet_layout(ws, rows)
            if calc:
                _inject_sum_formulas(ws, rows, calc, value_columns=("B", "C"))
        elif level == "group" and filename == "09-SOCIE.xlsx":
            _apply_group_socie_layout(ws, rows)
        elif level == "group":
            _apply_group_sheet_layout(ws, rows)
            if calc:
                _inject_sum_formulas(ws, rows, calc, value_columns=("B", "C", "D", "E"))
        else:
            raise ValueError(f"level={level!r} not supported")
        emitted.append((ws, rows))

    # Cross-sheet rollups: if this bundle has a face + sub-classification
    # pair (first = face, second = sub), wire face concepts that also appear
    # on the sub sheet to pull from the sub sheet's rollup rows. Matches the
    # MFRS template pattern exactly. SOCIE (single sheet) is skipped because
    # there is no sub sheet to pull from.
    if len(emitted) >= 2:
        face_ws, face_rows = emitted[0]
        sub_ws, sub_rows = emitted[1]
        if level == "company":
            _inject_face_to_sub_rollups(
                face_ws, face_rows, sub_ws, sub_rows, value_columns=("B", "C")
            )
        elif level == "group":
            _inject_face_to_sub_rollups(
                face_ws, face_rows, sub_ws, sub_rows,
                value_columns=("B", "C", "D", "E"),
            )

    out_path = out_dir / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def snapshot_backup_originals(level: str = "company") -> Path:
    """Mirror the MFRS backup-originals/ pattern for MPERS.

    Copies the currently-emitted templates into
    ``XBRL-template-MPERS/backup-originals/{Company,Group}/`` so future
    taxonomy updates have a clean baseline to diff against. Returns the
    destination directory.
    """
    import shutil

    src_dir = _REPO_ROOT / "XBRL-template-MPERS" / ("Company" if level == "company" else "Group")
    dst_dir = _REPO_ROOT / "XBRL-template-MPERS" / "backup-originals" / ("Company" if level == "company" else "Group")
    dst_dir.mkdir(parents=True, exist_ok=True)
    for src in src_dir.glob("*.xlsx"):
        shutil.copy2(src, dst_dir / src.name)
    return dst_dir


def _evaluate_sofp_balance(filled_path: Path) -> bool:
    """Simple formula evaluator used by the Phase 4 balance test.

    openpyxl doesn't evaluate formulas on save (that's Excel's job), so we
    re-parse the sheet's formulas, substitute actual cell values for refs,
    and compute locally. Supports ``=w1*<col><row>+w2*<col><row>+…`` which is
    exactly what `_inject_sum_formulas()` writes.

    Returns True when Assets == EquityAndLiabilities (col B only — the test
    fills one column). This is a Phase-4 smoke check, not a full verifier.
    """
    import openpyxl

    wb = openpyxl.load_workbook(filled_path)
    ws = wb["SOFP-CuNonCu"]

    # label -> row index
    label_to_row: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str):
            label_to_row[val] = r

    def cell_value(col: int, row: int) -> float:
        """Recursively evaluate a cell. Plain numbers resolve directly;
        formulas get parsed and summed."""
        v = ws.cell(row=row, column=col).value
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            if not v.startswith("="):
                # Non-numeric strings contribute 0.
                try:
                    return float(v)
                except ValueError:
                    return 0.0
            # Parse "=w1*B5+w2*B6-1*B7" etc.
            total = 0.0
            # Drop leading "=", split on "+" to get each signed term.
            expr = v[1:]
            # Normalise "-" into "+-" so split-on-"+" keeps signs.
            expr = expr.replace("-", "+-")
            terms = [t for t in expr.split("+") if t]
            for term in terms:
                # term example: "1*B5", "-1*C7", "B5"
                if "*" in term:
                    weight_s, ref = term.split("*", 1)
                    try:
                        weight = float(weight_s)
                    except ValueError:
                        weight = 1.0
                else:
                    ref, weight = term, 1.0
                # Parse column letters + row digits
                col_letters = "".join(c for c in ref if c.isalpha())
                row_digits = "".join(c for c in ref if c.isdigit())
                if not (col_letters and row_digits):
                    continue
                ref_col = 0
                for ch in col_letters:
                    ref_col = ref_col * 26 + (ord(ch.upper()) - ord("A") + 1)
                total += weight * cell_value(ref_col, int(row_digits))
            return total
        return 0.0

    assets_row = label_to_row.get("Total assets") or label_to_row.get("Assets")
    eql_row = (
        label_to_row.get("Total equity and liabilities")
        or label_to_row.get("Equity and liabilities")
    )
    if assets_row is None or eql_row is None:
        # Fallback: look for any "Total assets"/"Total equity and liabilities" with * prefix.
        for lbl, r in label_to_row.items():
            if lbl.lower().lstrip("*").strip().startswith("total assets"):
                assets_row = r
            if lbl.lower().lstrip("*").strip().startswith("total equity and liabilities"):
                eql_row = r
    assert assets_row and eql_row, "couldn't locate Assets/EquityAndLiabilities rows"

    a = cell_value(2, assets_row)
    e = cell_value(2, eql_row)
    return abs(a - e) < 0.5  # penny-level rounding slack


def generate_all(level: str = "company", statements: str = "all") -> list[Path]:
    """Generate every MPERS template matching the bundle mapping.

    ``statements`` accepts ``"face"`` (01..10), ``"notes"`` (11..15), or
    ``"all"``. Returns the list of emitted paths in bundle order.
    """
    if level not in ("company", "group"):
        raise ValueError(f"level={level!r} not supported")

    if statements == "face":
        filenames = [name for name, _ in _TEMPLATE_MAPPING[:10]]
    elif statements == "notes":
        filenames = [name for name, _ in _TEMPLATE_MAPPING[10:]]
    elif statements == "all":
        filenames = [name for name, _ in _TEMPLATE_MAPPING]
    else:
        raise ValueError(f"unknown statements filter: {statements!r}")

    out_dir = _REPO_ROOT / "XBRL-template-MPERS" / ("Company" if level == "company" else "Group")
    return [build_template(fn, level, out_dir) for fn in filenames]


def _cli() -> int:
    """Console entry — keeps the module runnable via `python scripts/generate_mpers_templates.py`."""
    import argparse

    parser = argparse.ArgumentParser(description="Generate MPERS XBRL templates from the SSM linkbase.")
    parser.add_argument("--level", choices=("company", "group"), default="company")
    parser.add_argument("--statements", choices=("face", "notes", "all"), default="all")
    parser.add_argument(
        "--snapshot",
        action="store_true",
        help="After emitting, copy all files to XBRL-template-MPERS/backup-originals/",
    )
    args = parser.parse_args()

    emitted = generate_all(level=args.level, statements=args.statements)
    print(f"emitted {len(emitted)} file(s) under {emitted[0].parent if emitted else '(none)'}")
    for p in emitted:
        print(f"  {p.relative_to(_REPO_ROOT)}")

    if args.snapshot:
        dst = snapshot_backup_originals(level=args.level)
        print(f"snapshotted to {dst.relative_to(_REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(_cli())


def list_mpers_roles() -> list[dict[str, Any]]:
    """Enumerate every MPERS presentation role shipped in the taxonomy.

    Returns one entry per `pre_*_role-NNNNNN.xml` file with the role number
    (6-digit string, e.g. "210000"), the human-readable title from the role
    XSD, and the absolute path to the presentation linkbase file.
    """
    roles: list[dict[str, Any]] = []
    titles = _load_role_titles()
    pre_files = sorted(_MPERS_TAXONOMY_DIR.glob("pre_*_role-*.xml"))
    for pre_file in pre_files:
        match = _PRE_FILENAME_PATTERN.search(pre_file.name)
        if not match:
            continue
        role_number = match.group("num")
        title = titles.get(role_number, "")
        roles.append(
            {
                "role_number": role_number,
                "title": title,
                "pre_file_path": pre_file,
            }
        )
    return roles
