"""Workbook merger — combines per-statement filled workbooks into one file.

Each extraction sub-agent produces its own workbook (e.g. SOFP_filled.xlsx,
SOPL_filled.xlsx). This module copies all sheets from each per-statement
workbook into a single merged output file, preserving values, formulas,
and cell styles.

Cross-sheet formula references within the same statement (e.g. SOFP-CuNonCu
referencing SOFP-Sub-CuNonCu) work automatically because both sheets are
copied together. No cross-*statement* formula references exist in MBRS
templates, so no remapping is needed.
"""

from __future__ import annotations

import logging
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict

import openpyxl
from openpyxl.utils import get_column_letter

from statement_types import StatementType

logger = logging.getLogger(__name__)


@dataclass
class MergeResult:
    success: bool
    output_path: str = ""
    sheets_copied: int = 0
    errors: list[str] = field(default_factory=list)


def merge(
    workbook_paths: Dict[StatementType, str],
    output_path: str,
) -> MergeResult:
    """Merge per-statement workbooks into a single output file.

    Args:
        workbook_paths: mapping of statement type → path to filled workbook.
        output_path: where to write the merged workbook.

    Returns:
        MergeResult with success flag and error details.
    """
    if not workbook_paths:
        return MergeResult(success=False, errors=["No workbook paths provided"])

    errors: list[str] = []
    merged = openpyxl.Workbook()
    # Remove the default empty sheet — we'll add real ones
    merged.remove(merged.active)
    # Formula caches are not evaluated by openpyxl. Ask Excel to fully recalculate
    # the merged workbook when a user opens the downloaded file.
    if merged.calculation is not None:
        merged.calculation.fullCalcOnLoad = True
        merged.calculation.forceFullCalc = True
        merged.calculation.calcOnSave = True

    sheets_copied = 0

    # Process statements in a stable order (SOFP → SOPL → SOCI → SOCF → SOCIE)
    for stmt_type in StatementType:
        path = workbook_paths.get(stmt_type)
        if path is None:
            continue

        if not Path(path).exists():
            errors.append(f"{stmt_type.value}: file not found: {path}")
            continue

        try:
            src_wb = openpyxl.load_workbook(path, data_only=False)
        except Exception as e:
            errors.append(f"{stmt_type.value}: failed to open: {e}")
            continue

        for src_ws in src_wb.worksheets:
            _copy_sheet(src_ws, merged)
            sheets_copied += 1

        src_wb.close()

    if sheets_copied == 0:
        return MergeResult(
            success=False,
            output_path=output_path,
            errors=errors or ["No sheets copied — all workbooks empty or unreadable"],
        )

    try:
        merged.save(output_path)
    except Exception as e:
        return MergeResult(success=False, output_path=output_path, errors=[f"Failed to save: {e}"])
    finally:
        merged.close()

    if errors:
        logger.warning("Merge completed with errors: %s", errors)

    return MergeResult(
        success=True,
        output_path=output_path,
        sheets_copied=sheets_copied,
        errors=errors,
    )


def _copy_sheet(src_ws, dest_wb: openpyxl.Workbook) -> None:
    """Copy a worksheet into the destination workbook, preserving values, formulas, and styles."""
    dest_ws = dest_wb.create_sheet(title=src_ws.title)

    # Copy column dimensions (widths)
    for col_letter, dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col_letter].width = dim.width
        dest_ws.column_dimensions[col_letter].hidden = dim.hidden

    # Copy row dimensions (heights)
    for row_num, dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[row_num].height = dim.height
        dest_ws.row_dimensions[row_num].hidden = dim.hidden

    # Copy cell values, formulas, and styles
    for row in src_ws.iter_rows():
        for cell in row:
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)

            # Copy style attributes
            if cell.has_style:
                dest_cell.font = copy(cell.font)
                dest_cell.fill = copy(cell.fill)
                dest_cell.border = copy(cell.border)
                dest_cell.alignment = copy(cell.alignment)
                dest_cell.number_format = cell.number_format
                dest_cell.protection = copy(cell.protection)

    # Copy merged cell ranges
    for merged_range in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged_range))
