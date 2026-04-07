#!/usr/bin/env python3
"""
Fix +20 offset formula bugs in SOFP-Sub-CuNonCu sheet.

For each formula cell in columns B and C, this script:
1. Identifies cross-section subtotal references (B{row}/C{row} pointing to different sections)
2. Checks if row-20 contains a label in the correct section
3. If yes, replaces the reference (B{row} → B{row-20}, C{row} → C{row-20})
4. Only changes cross-section references, leaving within-section and cross-sheet refs alone
"""

import re
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# Configuration
TEMPLATE_PATH = Path(__file__).parent / "XBRL-template-MFRS" / "01-SOFP-CuNonCu.xlsx"
BACKUP_DIR = Path(__file__).parent / "backup-originals"
SHEET_NAME = "SOFP-Sub-CuNonCu"

# Section boundaries and order (from accounting structure)
# Each section is identified by its label pattern and row range
SECTIONS = {
    "PPE": {"start": None, "end": None, "labels": ["Land and buildings", "Leasehold improvements", "Plant and machinery", "Motor vehicles", "Fixtures and fittings", "Computer and office equipment", "Construction in progress", "Total PPE"]},
    "Intangible": {"start": None, "end": None, "labels": ["Goodwill", "Patents and trademarks", "Software", "Licenses", "Other intangible assets", "Accumulated amortization", "Total intangible"]},
    "NC Receivables": {"start": None, "end": None, "labels": ["Long-term receivables", "Accrued income", "Total NC receivables"]},
    "NC Derivatives": {"start": None, "end": None, "labels": ["Derivative assets", "Total NC derivative assets"]},
    "Prepayments": {"start": None, "end": None, "labels": ["Prepaid expenses", "Accrued income", "Total prepayments"]},
    "C Receivables": {"start": None, "end": None, "labels": ["Trade receivables", "Other receivables", "Total current receivables"]},
    "C Derivatives": {"start": None, "end": None, "labels": ["Derivative assets - current", "Total current derivative assets"]},
    "Cash": {"start": None, "end": None, "labels": ["Cash and cash equivalents", "Bank accounts", "Total cash"]},
    "Reserves": {"start": None, "end": None, "labels": ["Share capital", "Reserves", "Total reserves"]},
    "NC Borrowings": {"start": None, "end": None, "labels": ["Long-term debt", "Term loans", "Bonds", "Total NC borrowings"]},
    "NC Payables": {"start": None, "end": None, "labels": ["Trade payables", "Accrued expenses", "Other NC payables"]},
    "NC Derivatives": {"start": None, "end": None, "labels": ["Derivative liabilities", "Total NC derivative liabilities"]},
    "C Payables": {"start": None, "end": None, "labels": ["Trade payables", "Accrued expenses", "Other current payables"]},
    "C Derivatives": {"start": None, "end": None, "labels": ["Derivative liabilities - current", "Total current derivative liabilities"]},
}


def get_section_for_row(ws, row: int) -> str:
    """Determine which section a row belongs to by looking at column A."""
    cell_value = ws[f"A{row}"].value
    if not cell_value:
        return None

    label = str(cell_value).strip().lower()

    # Simple heuristic: look for subtotal/section markers
    if "total" in label or "*" in label:
        # This is a subtotal row - map to section
        if "ppe" in label or "land" in label or "plant" in label or "machinery" in label or "building" in label:
            return "PPE"
        elif "intangible" in label or "goodwill" in label or "patent" in label:
            return "Intangible"
        elif "nc receivable" in label or "long-term receivable" in label:
            return "NC Receivables"
        elif "nc derivative" in label and "asset" in label:
            return "NC Derivatives"
        elif "prepayment" in label or "accrued" in label and row < 170:
            return "Prepayments"
        elif "current receivable" in label or "trade receivable" in label:
            return "C Receivables"
        elif "current derivative" in label and "asset" in label:
            return "C Derivatives"
        elif "cash" in label:
            return "Cash"
        elif "reserve" in label or "share capital" in label or "equity" in label:
            return "Reserves"
        elif "nc borrowing" in label or "long-term debt" in label:
            return "NC Borrowings"
        elif "nc payable" in label or "nc" in label and "payable" in label:
            return "NC Payables"
        elif "nc derivative" in label and "liab" in label:
            return "NC Derivatives"
        elif "current payable" in label or ("payable" in label and row > 400):
            return "C Payables"
        elif "current derivative" in label and "liab" in label:
            return "C Derivatives"

    return None


def extract_cell_references(formula: str) -> list:
    """Extract all B{row}/C{row} references from a formula.

    Returns list of tuples: (column_letter, row_number, full_ref_str)
    """
    # Pattern: B123, C456, etc. (simple single-cell refs, not ranges or sheet refs)
    # Exclude cross-sheet refs like 'Sheet'!B123
    pattern = r"(?<![!'])([BC])(\d+)(?![\w'])"
    matches = re.findall(pattern, formula)

    refs = []
    for col, row in matches:
        refs.append((col, int(row), f"{col}{row}"))

    return refs


def get_label_at_row(ws, row: int) -> str:
    """Get the label text in column A for a given row."""
    cell = ws[f"A{row}"]
    value = cell.value
    return str(value).strip() if value else ""


def is_subtotal_row(label: str) -> bool:
    """Check if a label is a subtotal row."""
    return "total" in label.lower() or label.startswith("*")


def section_from_label(label: str, row: int) -> str:
    """Map a label to its section."""
    label_lower = label.lower()

    if "ppe" in label_lower or "land" in label_lower or "plant" in label_lower or "building" in label_lower or "machinery" in label_lower or "vehicle" in label_lower or "fixture" in label_lower or "computer" in label_lower:
        return "PPE"
    elif "intangible" in label_lower or "goodwill" in label_lower or "patent" in label_lower or "software" in label_lower or "license" in label_lower:
        return "Intangible"
    elif "nc receivable" in label_lower or "long-term receivable" in label_lower:
        return "NC Receivables"
    elif "nc derivative" in label_lower and "asset" in label_lower:
        return "NC Derivatives"
    elif "prepayment" in label_lower or ("accrued" in label_lower and row < 170):
        return "Prepayments"
    elif "current receivable" in label_lower or "trade receivable" in label_lower or "other receivable" in label_lower:
        return "C Receivables"
    elif "current derivative" in label_lower and "asset" in label_lower:
        return "C Derivatives"
    elif "cash" in label_lower:
        return "Cash"
    elif "reserve" in label_lower or "share capital" in label_lower or ("equity" in label_lower and row < 250):
        return "Reserves"
    elif "nc borrowing" in label_lower or "long-term debt" in label_lower or "term loan" in label_lower or "bond" in label_lower:
        return "NC Borrowings"
    elif "nc payable" in label_lower or ("nc" in label_lower and "payable" in label_lower):
        return "NC Payables"
    elif "nc derivative" in label_lower and "liab" in label_lower:
        return "NC Derivatives"
    elif "current payable" in label_lower or ("payable" in label_lower and row > 400):
        return "C Payables"
    elif "current derivative" in label_lower and "liab" in label_lower:
        return "C Derivatives"

    return None


def fix_formula_offsets():
    """Main fix function."""

    # Create backup directory
    BACKUP_DIR.mkdir(exist_ok=True)

    # Backup original file
    backup_path = BACKUP_DIR / TEMPLATE_PATH.name
    if not backup_path.exists():
        shutil.copy2(TEMPLATE_PATH, backup_path)
        print(f"Created backup: {backup_path}")
    else:
        print(f"Backup already exists: {backup_path}")

    # Load workbook
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    print(f"\nProcessing sheet: {SHEET_NAME}")
    print(f"Sheet dimensions: {ws.dimensions}")

    # Track changes
    changes = []

    # Scan all cells in columns B and C
    for row in range(1, ws.max_row + 1):
        for col in ["B", "C"]:
            cell = ws[f"{col}{row}"]

            # Only process formula cells
            if not cell.value or not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue

            formula = cell.value
            refs = extract_cell_references(formula)

            if not refs:
                continue

            # Determine section of the formula row (where the subtotal is)
            formula_row_label = get_label_at_row(ws, row)
            formula_section = section_from_label(formula_row_label, row)

            # For each reference in the formula
            changes_in_formula = []
            new_formula = formula

            for ref_col, ref_row, ref_str in refs:
                ref_label = get_label_at_row(ws, ref_row)
                ref_section = section_from_label(ref_label, ref_row)

                # Check if this is a cross-section reference (broken)
                if ref_section and formula_section and ref_section != formula_section:
                    # Check if row-20 has the correct section
                    alt_row = ref_row - 20
                    if alt_row > 0:
                        alt_label = get_label_at_row(ws, alt_row)
                        alt_section = section_from_label(alt_label, alt_row)

                        if alt_section == formula_section:
                            # This is a broken +20 offset! Fix it.
                            old_ref = ref_str
                            new_ref = f"{ref_col}{alt_row}"

                            new_formula = new_formula.replace(old_ref, new_ref, 1)

                            changes_in_formula.append({
                                "old_ref": old_ref,
                                "new_ref": new_ref,
                                "old_row": ref_row,
                                "new_row": alt_row,
                                "old_label": ref_label,
                                "new_label": alt_label,
                                "old_section": ref_section,
                                "new_section": alt_section,
                            })

            # Apply changes if any were made
            if changes_in_formula:
                change_record = {
                    "row": row,
                    "column": col,
                    "formula_label": formula_row_label,
                    "formula_section": formula_section,
                    "old_formula": formula,
                    "new_formula": new_formula,
                    "reference_changes": changes_in_formula,
                }
                changes.append(change_record)

                cell.value = new_formula
                print(f"\n[Row {row}, Col {col}] {formula_row_label}")
                print(f"  Old: {formula}")
                print(f"  New: {new_formula}")
                for change in changes_in_formula:
                    print(f"    {change['old_ref']} ({change['old_label'][:50]}) → {change['new_ref']} ({change['new_label'][:50]})")

    # Save workbook
    wb.save(TEMPLATE_PATH)
    print(f"\n{'='*80}")
    print(f"Total formulas fixed: {len(changes)}")
    print(f"File saved: {TEMPLATE_PATH}")

    # Summary
    print(f"\n{'='*80}")
    print("DETAILED CHANGE LOG:")
    print(f"{'='*80}")

    for i, change in enumerate(changes, 1):
        print(f"\n[Change {i}] Row {change['row']}, Column {change['column']}")
        print(f"  Label: {change['formula_label']}")
        print(f"  Section: {change['formula_section']}")
        print(f"  Old formula: {change['old_formula']}")
        print(f"  New formula: {change['new_formula']}")
        print(f"  References changed:")
        for ref_change in change['reference_changes']:
            print(f"    {ref_change['old_ref']} → {ref_change['new_ref']}")
            print(f"      Old: Row {ref_change['old_row']} ({ref_change['old_section']}) - {ref_change['old_label']}")
            print(f"      New: Row {ref_change['new_row']} ({ref_change['new_section']}) - {ref_change['new_label']}")

    return changes


if __name__ == "__main__":
    changes = fix_formula_offsets()
    print(f"\nScript completed. {len(changes)} formulas were fixed.")
