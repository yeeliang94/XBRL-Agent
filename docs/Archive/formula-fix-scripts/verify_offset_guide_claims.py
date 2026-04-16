#!/usr/bin/env python3
"""
Verify the TEMPLATE-FORMULA-FIX-GUIDE.md claims about +20 offset bugs.

According to the guide, these templates have +20 offset bugs in cross-section references:
1. 07-SOCF-Indirect.xlsx - SOCF-Indirect
2. 08-SOCF-Direct.xlsx - SOCF-Direct
3. 09-SOCIE.xlsx - SOCIE
4. 04-SOPL-Nature.xlsx - SOPL-Analysis-Nature

These should be CLEAN (no +20 bugs):
1. 03-SOPL-Function.xlsx
2. 05-SOCI-BeforeTax.xlsx
3. 06-SOCI-NetOfTax.xlsx

The guide specifically says:
- SOFP-Sub-CuNonCu: Has +20 offset bugs (HIGH severity)
- SOCF-Indirect: "operating/investing/financing roll-ups broken" with +20
- SOPL-Analysis-Nature: "LOW — 2 cells" affected by +20

The key pattern: A formula at row N references row N+20, but should reference row N-20.
This means referenced row should belong to a DIFFERENT accounting section.
"""

import openpyxl
from pathlib import Path
import re

class GuideClaimVerifier:
    def __init__(self, template_path: str):
        self.path = Path(template_path)
        self.wb = openpyxl.load_workbook(self.path, data_only=False)
        self.template_name = self.path.name

    def get_label(self, sheet_name: str, row: int) -> str:
        try:
            ws = self.wb[sheet_name]
            val = ws[f'A{row}'].value
            return str(val).strip() if val else ""
        except:
            return ""

    def extract_refs(self, formula: str):
        """Extract B{row} and C{row} references."""
        refs = []
        pattern = r'[\$]?([B-D])[\$]?(\d+)'
        for match in re.finditer(pattern, formula):
            col, row = match.groups()
            if col in ['B', 'C']:
                refs.append((col, int(row)))
        return refs

    def find_cross_section_refs(self, sheet_name: str):
        """
        Find formulas that reference rows in different sections.
        Per guide: "A formula is broken if it references a row in a DIFFERENT accounting section"

        Sections are determined by rows starting with '*' (subtotals/totals).
        """
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            return None

        # Build section map: row -> section_name
        section_map = {}
        current_section = None
        section_start = None

        for row in range(1, ws.max_row + 1):
            label = self.get_label(sheet_name, row)

            # Section header: starts with * or contains "Total"
            if label.startswith('*'):
                current_section = label
                section_start = row

            if current_section:
                section_map[row] = current_section

        # Find formulas with cross-section references
        issues = []

        for row in range(1, ws.max_row + 1):
            label = self.get_label(sheet_name, row)
            if not label:
                continue

            for col in ['B', 'C']:
                cell_addr = f'{col}{row}'
                cell_val = ws[cell_addr].value

                if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                    formula = cell_val
                    refs = self.extract_refs(formula)

                    # Check if any ref is in different section
                    formula_section = section_map.get(row)
                    broken_refs = []

                    for ref_col, ref_row in refs:
                        ref_label = self.get_label(sheet_name, ref_row)
                        ref_section = section_map.get(ref_row)

                        # Different section = potential cross-section ref
                        if (formula_section and ref_section and
                            formula_section != ref_section and
                            ref_label):  # Must have a label

                            # Check if ref_row +/- 20 has same section as formula
                            corrected_row = ref_row - 20
                            corrected_label = self.get_label(sheet_name, corrected_row)
                            corrected_section = section_map.get(corrected_row)

                            if corrected_section == formula_section:
                                # This is a +20 offset bug!
                                broken_refs.append({
                                    'ref': f'{ref_col}{ref_row}',
                                    'ref_label': ref_label[:50],
                                    'ref_section': ref_section[:40],
                                    'corrected_row': corrected_row,
                                    'corrected_label': corrected_label[:50] if corrected_label else '',
                                    'corrected_section': corrected_section[:40] if corrected_section else ''
                                })

                    if broken_refs:
                        issues.append({
                            'cell': cell_addr,
                            'row': row,
                            'label': label[:60],
                            'formula': formula,
                            'formula_section': formula_section[:40],
                            'broken_refs': broken_refs
                        })

        return issues, section_map

    def report(self, sheet_name: str):
        result = self.find_cross_section_refs(sheet_name)
        if not result:
            print(f"  Sheet '{sheet_name}' not found")
            return False

        issues, section_map = result

        print(f"  Sheet: {sheet_name}")
        print(f"  Total labeled rows: {len(section_map)}")
        print(f"  Cross-section reference bugs found: {len(issues)}")

        if issues:
            print(f"\n  BROKEN FORMULAS (referencing different section):\n")
            for i, issue in enumerate(issues[:10], 1):
                print(f"  {i}. Cell {issue['cell']} - Row {issue['row']}")
                print(f"     Label: {issue['label']}")
                print(f"     Section: {issue['formula_section']}")
                print(f"     Formula: {issue['formula'][:70]}...")
                for br in issue['broken_refs']:
                    print(f"       References {br['ref']}: {br['ref_label']}")
                    print(f"         Currently points to: {br['ref_section']}")
                    print(f"         Fix: {br['ref']} -> {br['ref'][0]}{br['corrected_row']} ({br['corrected_label']})")
                print()

            if len(issues) > 10:
                print(f"  ... and {len(issues) - 10} more\n")

            return True
        else:
            print(f"  No cross-section formula issues detected")
            return False

def main():
    specs = {
        "CLAIMED BUGGY (per guide)": [
            ("XBRL-template-MFRS/07-SOCF-Indirect.xlsx", "SOCF-Indirect"),
            ("XBRL-template-MFRS/08-SOCF-Direct.xlsx", "SOCF-Direct"),
            ("XBRL-template-MFRS/09-SOCIE.xlsx", "SOCIE"),
            ("XBRL-template-MFRS/04-SOPL-Nature.xlsx", "SOPL-Analysis-Nature"),
        ],
        "CLAIMED CLEAN (per guide)": [
            ("XBRL-template-MFRS/03-SOPL-Function.xlsx", "SOPL-Analysis-Function"),
            ("XBRL-template-MFRS/05-SOCI-BeforeTax.xlsx", "SOCI-BeforeOfTax"),
            ("XBRL-template-MFRS/06-SOCI-NetOfTax.xlsx", "SOCI-NetOfTax"),
        ]
    }

    base_path = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent")

    print("\n" + "=" * 100)
    print("VERIFY TEMPLATE-FORMULA-FIX-GUIDE.md CLAIMS")
    print("=" * 100)

    all_results = {}

    for category, templates in specs.items():
        print(f"\n### {category} ###\n")

        category_results = {}

        for template, sheet_name in templates:
            full_path = base_path / template
            if not full_path.exists():
                print(f"File not found: {template}")
                continue

            print(f"{Path(template).name}")
            verifier = GuideClaimVerifier(str(full_path))
            has_bugs = verifier.report(sheet_name)
            category_results[template] = has_bugs
            print()

        all_results[category] = category_results

    # Verdict
    print("=" * 100)
    print("VERDICT")
    print("=" * 100)

    print("\nBUGGY TEMPLATES (guide claims +20 offset bugs):")
    buggy_results = all_results.get("CLAIMED BUGGY (per guide)", {})
    for template, has_bugs in buggy_results.items():
        status = "✓ CONFIRMED" if has_bugs else "✗ CLAIM INCORRECT"
        print(f"  {Path(template).name}: {status}")

    print("\nCLEAN TEMPLATES (guide claims no bugs):")
    clean_results = all_results.get("CLAIMED CLEAN (per guide)", {})
    for template, has_bugs in clean_results.items():
        status = "✓ CONFIRMED" if not has_bugs else "✗ CLAIM INCORRECT"
        print(f"  {Path(template).name}: {status}")

    # Summary
    buggy_correct = sum(1 for v in buggy_results.values() if v)
    buggy_total = len(buggy_results)
    clean_correct = sum(1 for v in clean_results.values() if not v)
    clean_total = len(clean_results)

    print(f"\nSUMMARY:")
    print(f"  Buggy templates: {buggy_correct}/{buggy_total} claims verified")
    print(f"  Clean templates: {clean_correct}/{clean_total} claims verified")
    print(f"  Overall accuracy: {(buggy_correct + clean_correct) / (buggy_total + clean_total) * 100:.0f}%")

if __name__ == "__main__":
    main()
