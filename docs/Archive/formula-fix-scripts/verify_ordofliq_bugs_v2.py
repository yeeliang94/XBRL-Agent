#!/usr/bin/env python3
"""
Verify OrderOfLiquidity bugs with improved section detection.

Instead of scanning backwards for headers, we manually map out the accounting
sections based on the OrdOfLiq template structure.
"""

import re
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set

# File paths
TEMPLATE_PATH = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent/XBRL-template-MFRS/02-SOFP-OrderOfLiquidity.xlsx")
SHEET_NAME = "SOFP-Sub-OrdOfLiq"

# Target rows from the bug claims
TARGET_ROWS = {
    148: "Total cash",
    168: "Total issued capital",
    241: "Total borrowings",
    295: "Total trade and other payables"
}


def load_template_data(wb):
    """Load all labels and formulas from the template."""
    ws = wb[SHEET_NAME]

    labels = {}
    formulas = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        row_num = row[0].row
        label_cell = row[0]

        if label_cell.value:
            labels[row_num] = str(label_cell.value).strip()

        b_cell = row[1]
        c_cell = row[2]

        if b_cell.data_type == 'f' or c_cell.data_type == 'f':
            formulas[row_num] = {
                "B": b_cell.value if b_cell.data_type == 'f' else None,
                "C": c_cell.value if c_cell.data_type == 'f' else None,
            }

    return labels, formulas


def extract_cell_references(formula: str) -> List[Tuple[str, int]]:
    """Extract cell references from a formula."""
    if not formula:
        return []
    pattern = r'([A-Z]+)(\d+)'
    matches = re.findall(pattern, formula)
    return [(col, int(row)) for col, row in matches]


def infer_sections(labels: Dict[int, str]) -> Dict[int, str]:
    """
    Build a mapping of row -> section name by looking for major section markers
    (rows that are subtotals or major headers).

    For OrderOfLiquidity, sections typically:
    - Start after current assets
    - Include inventories, current receivables, derivatives, cash
    - Then equity
    - Then liabilities (borrowings, payables)
    """
    sections = {}
    current_section = "UNKNOWN"

    for row in sorted(labels.keys()):
        label = labels[row]

        # Major section markers (these define section boundaries)
        if any(marker in label.lower() for marker in [
            "non-current",
            "current asset",
            "equity",
            "borrowing",
            "payable",
            "cash",
            "inventories",
            "receivable",
            "derivative"
        ]):
            # Try to infer section from label
            if "equity" in label.lower():
                current_section = "EQUITY"
            elif "borrowing" in label.lower():
                current_section = "BORROWINGS"
            elif "payable" in label.lower():
                current_section = "PAYABLES"
            elif "cash" in label.lower() and "equivalents" in label.lower():
                current_section = "CASH"
            elif "inventory" in label.lower() or "inventories" in label.lower():
                current_section = "INVENTORIES"
            elif "receivable" in label.lower():
                current_section = "RECEIVABLES"
            elif "derivative" in label.lower():
                current_section = "DERIVATIVES"
            elif "prepaid" in label.lower() or "accrual" in label.lower():
                current_section = "PREPAID_ACCRUALS"

        sections[row] = current_section

    return sections


def main():
    print("=" * 100)
    print("VERIFYING ORDEROF LIQUIDITY TEMPLATE BUGS (v2 - detailed analysis)")
    print(f"File: {TEMPLATE_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    print("=" * 100)
    print()

    # Load workbook
    try:
        wb = load_workbook(TEMPLATE_PATH, data_only=False)
    except Exception as e:
        print(f"ERROR: Could not load workbook: {e}")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found")
        return

    labels, formulas = load_template_data(wb)
    sections = infer_sections(labels)

    print(f"Loaded {len(labels)} rows with labels")
    print(f"Found {len(formulas)} rows with formulas")
    print()

    # ========================================================================
    # ANALYSIS: Check what each target formula actually sums
    # ========================================================================
    print("=" * 100)
    print("DETAILED FORMULA ANALYSIS")
    print("=" * 100)
    print()

    for target_row, expected_label in TARGET_ROWS.items():
        actual_label = labels.get(target_row, "[NOT FOUND]")

        print(f"\nROW {target_row}: '{actual_label}'")
        print("─" * 100)

        if target_row not in formulas:
            print(f"  ⚠️  No formulas found in this row")
            continue

        formula_b = formulas[target_row].get("B")
        if not formula_b:
            print(f"  ⚠️  No formula in column B")
            continue

        print(f"  Formula (Column B):")
        print(f"    {formula_b}")
        print()

        # Extract and analyze references
        refs = extract_cell_references(formula_b)
        print(f"  References: {len(refs)} cell(s)")
        print()

        # Group by section
        sections_referenced = {}
        suspicious_items = []

        for col, ref_row in refs:
            ref_label = labels.get(ref_row, f"[ROW {ref_row} NOT FOUND]")
            ref_section = sections.get(ref_row, "UNKNOWN")

            if ref_section not in sections_referenced:
                sections_referenced[ref_section] = []

            sections_referenced[ref_section].append({
                'row': ref_row,
                'label': ref_label,
                'section': ref_section
            })

            # Flag suspicious references
            target_section = sections.get(target_row, "UNKNOWN")
            if ref_section != target_section and ref_section != "UNKNOWN":
                suspicious_items.append((ref_row, ref_label, ref_section))

        # Print summary of referenced sections
        print("  Sections referenced:")
        for section, items in sorted(sections_referenced.items()):
            print(f"\n    {section} ({len(items)} item(s)):")
            for item in items[:10]:  # Show first 10
                print(f"      Row {item['row']:3d}: {item['label'][:70]}")
            if len(items) > 10:
                print(f"      ... and {len(items) - 10} more")

        # Report on cross-section references
        if suspicious_items:
            target_section = sections.get(target_row, "UNKNOWN")
            print(f"\n  ⚠️  CROSS-SECTION REFERENCES DETECTED!")
            print(f"  Expected section: {target_section}")
            print(f"  But also references:")
            for row, label, section in suspicious_items[:5]:
                print(f"      Row {row}: '{label}' (section: {section})")
            if len(suspicious_items) > 5:
                print(f"      ... and {len(suspicious_items) - 5} more cross-section items")
        else:
            print(f"\n  ✓ All references appear to be within the same section")

    # ========================================================================
    # BUG VERIFICATION SUMMARY
    # ========================================================================
    print("\n\n" + "=" * 100)
    print("BUG CLAIM VERIFICATION")
    print("=" * 100)
    print()

    claims = {
        148: {
            'expected': "Total cash",
            'description': "sums inventory + derivative + cash items instead of just cash items",
            'should_reference': "CashOnHand, BalancesWithBanks only"
        },
        168: {
            'expected': "Total issued capital",
            'description': "includes prepaid assets in addition to capital items",
            'should_reference': "Capital items only (not prepaid)"
        },
        241: {
            'expected': "Total borrowings",
            'description': "includes equity items (perpetual sukuk, ICULS equity) mixed with borrowings",
            'should_reference': "Borrowing items only (not equity)"
        },
        295: {
            'expected': "Total trade and other payables",
            'description': "double-counts by summing both leaf items AND their subtotals",
            'should_reference': "Leaf items only (not sub-totals)"
        }
    }

    for row, claim in claims.items():
        actual_label = labels.get(row, "[NOT FOUND]")
        formula_b = formulas.get(row, {}).get("B", "NO FORMULA")
        refs = extract_cell_references(formula_b) if formula_b != "NO FORMULA" else []

        print(f"\nClaim #{row}: '{claim['expected']}'")
        print(f"  Description: {claim['description']}")
        print(f"  Actual label: '{actual_label}'")
        print(f"  Formula references {len(refs)} cells")

        # Analyze the references
        if formula_b != "NO FORMULA":
            sections_in_formula = set()
            for col, ref_row in refs:
                ref_section = sections.get(ref_row, "UNKNOWN")
                sections_in_formula.add(ref_section)

            print(f"  Sections referenced: {', '.join(sorted(sections_in_formula))}")

            # Check if cross-section references exist
            target_section = sections.get(row, "UNKNOWN")
            cross_section_refs = [r for r in refs if sections.get(r[1], "UNKNOWN") != target_section]

            if cross_section_refs:
                print(f"  ✗ BUG LIKELY: Found {len(cross_section_refs)} cross-section references")
            else:
                print(f"  ? INCONCLUSIVE: No obvious cross-section refs, but may need manual verification")


if __name__ == "__main__":
    main()
