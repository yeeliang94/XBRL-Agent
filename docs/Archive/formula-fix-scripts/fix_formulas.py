#!/usr/bin/env python3
"""
Fix mixed formula bugs in SOFP-OrderOfLiquidity template.

Strategy:
1. Parse XBRL calculation linkbase (role-200200 for Order of Liquidity)
2. Parse label linkbase to map concept names to human-readable labels
3. Read template and extract all labels from column A
4. For each formula cell, match its label to an XBRL concept
5. Build correct formula from the calculation tree
6. Compare and replace if different
"""

import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Tuple, Optional
from difflib import SequenceMatcher

# Paths
XBRL_BASE = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent/SSMxT_2022v1.0")
CAL_LINKBASE = XBRL_BASE / "rep/ssm/ca-2016/fs/mfrs/cal_ssmt-fs-mfrs_2022-12-31_role-200200.xml"
LAB_LINKBASE = XBRL_BASE / "rep/ssm/ca-2016/fs/mfrs/lab_en-ssmt-fs-mfrs_2022-12-31.xml"
TEMPLATE_PATH = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent/XBRL-template-MFRS/02-SOFP-OrderOfLiquidity.xlsx")
OUTPUT_PATH = TEMPLATE_PATH

# Backup path
BACKUP_DIR = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent/backup-originals")
BACKUP_PATH = BACKUP_DIR / "02-SOFP-OrderOfLiquidity.xlsx"

def load_labels_from_linkbase(lab_file: Path) -> Tuple[Dict[str, str], Dict[str, Tuple[Optional[str], str]]]:
    """
    Parse label linkbase and extract:
    - label_text: label_id -> text (e.g., label_43 -> "Total cash and cash equivalents")
    - label_to_concept: label_id -> (concept_name, text) (e.g., label_43 -> ("ssmt_CashAndBankBalances", "..."))
    """
    label_text = {}  # label_id -> text
    label_to_concept = {}  # label_id -> (concept, text)
    tree = ET.parse(lab_file)
    root = tree.getroot()

    # Fully-qualified namespace URIs
    ns_lb = '{http://www.xbrl.org/2003/linkbase}'
    ns_xlink = '{http://www.w3.org/1999/xlink}'

    # Step 1: Extract all label elements (label_N -> text)
    for label_elem in root.iter(f'{ns_lb}label'):
        label_id = label_elem.get(f'{ns_xlink}label')
        text = label_elem.text or ""
        if label_id and text:
            label_text[label_id] = text.strip()

    # Step 2: Extract loc elements (loc_X -> concept_name)
    loc_to_concept = {}
    for loc_elem in root.iter(f'{ns_lb}loc'):
        loc_id = loc_elem.get(f'{ns_xlink}label')
        href = loc_elem.get(f'{ns_xlink}href')
        if loc_id and href and '#' in href:
            concept = href.split('#')[1]
            loc_to_concept[loc_id] = concept

    # Step 3: Extract labelArc connections (loc_X -> label_N)
    for arc_elem in root.iter(f'{ns_lb}labelArc'):
        from_id = arc_elem.get(f'{ns_xlink}from')  # loc_ssmt_CashAndBankBalances
        to_id = arc_elem.get(f'{ns_xlink}to')      # label_43
        if from_id and to_id and to_id in label_text:
            label_to_concept[to_id] = (loc_to_concept.get(from_id), label_text[to_id])

    return label_text, label_to_concept

def parse_calculation_linkbase(cal_file: Path) -> Tuple[Dict[str, List[Tuple[str, float]]], Dict[str, str]]:
    """
    Parse calculation linkbase to build:
    1. parent_loc -> [(child_loc, weight), ...] mapping
    2. loc_id -> concept mapping

    Returns: (calc_tree, loc_to_concept)
    """
    calc_tree = {}  # loc_id -> [(child_loc_id, weight), ...]
    loc_to_concept = {}  # loc_id -> concept
    tree = ET.parse(cal_file)
    root = tree.getroot()

    # Fully-qualified namespace URIs
    ns_lb = '{http://www.xbrl.org/2003/linkbase}'
    ns_xlink = '{http://www.w3.org/1999/xlink}'

    # Step 1: Extract loc elements (loc_X -> concept_name)
    for loc in root.iter(f'{ns_lb}loc'):
        loc_id = loc.get(f'{ns_xlink}label')
        href = loc.get(f'{ns_xlink}href')
        if loc_id and href and '#' in href:
            concept = href.split('#')[1]
            loc_to_concept[loc_id] = concept

    # Step 2: Extract calculation arcs
    for arc in root.iter(f'{ns_lb}calculationArc'):
        from_loc = arc.get(f'{ns_xlink}from')
        to_loc = arc.get(f'{ns_xlink}to')
        weight = float(arc.get('weight', 1.0))

        if from_loc and to_loc:
            if from_loc not in calc_tree:
                calc_tree[from_loc] = []
            calc_tree[from_loc].append((to_loc, weight))

    return calc_tree, loc_to_concept

def fuzzy_match(excel_label: str, concept_label: str, threshold: float = 0.6) -> bool:
    """Fuzzy match between Excel label and XBRL concept label."""
    # Normalize both strings
    excel_norm = excel_label.lower().strip().replace('*', '').strip()
    concept_norm = concept_label.lower().strip().replace('*', '').strip()

    # Exact match
    if excel_norm == concept_norm:
        return True

    # Check if one contains the other
    if excel_norm in concept_norm or concept_norm in excel_norm:
        return True

    # Fuzzy similarity
    ratio = SequenceMatcher(None, excel_norm, concept_norm).ratio()
    return ratio >= threshold

def find_concept_for_label(excel_label: str, concept_to_label_map: Dict[str, str]) -> Optional[str]:
    """Find XBRL concept that matches Excel label."""
    if not excel_label:
        return None

    excel_label = excel_label.strip()

    # Try exact match first
    for concept, concept_label in concept_to_label_map.items():
        if excel_label.lower() == concept_label.lower():
            return concept

    # Try fuzzy match
    for concept, concept_label in concept_to_label_map.items():
        if fuzzy_match(excel_label, concept_label, threshold=0.75):
            return concept

    return None

def build_formula_from_tree(
    parent_loc: str,
    calc_tree: Dict[str, List[Tuple[str, float]]],
    loc_to_concept: Dict[str, str],
    concept_to_label_map: Dict[str, str],
    label_to_row: Dict[str, int],
) -> Optional[str]:
    """
    Build a formula for a parent loc by looking up its children in the calculation tree.
    Returns: "=weight1*B{row1}+weight2*B{row2}+..." or None if can't resolve
    """
    if parent_loc not in calc_tree:
        return None

    children = calc_tree[parent_loc]
    formula_parts = []

    for child_loc, weight in children:
        # Get concept for child loc
        child_concept = loc_to_concept.get(child_loc)
        if not child_concept:
            print(f"    WARNING: No concept found for locator {child_loc}")
            continue

        # Find the label for this child concept
        child_label = concept_to_label_map.get(child_concept)
        if not child_label:
            print(f"    WARNING: No label found for concept {child_concept}")
            continue

        # Find the row for this label in the Excel
        row = label_to_row.get(child_label)
        if row is None:
            print(f"    WARNING: No row found for label '{child_label}' (concept {child_concept})")
            continue

        # Build formula part: weight*B{row} or -1*B{row}
        if weight == 1.0 or weight == 1:
            formula_parts.append(f"1*B{row}")
        elif weight == -1.0 or weight == -1:
            formula_parts.append(f"-1*B{row}")
        else:
            formula_parts.append(f"{weight}*B{row}")

    if not formula_parts:
        return None

    return "=" + "+".join(formula_parts).replace("+-", "-")

def main():
    print("=" * 80)
    print("XBRL Formula Fixer for SOFP-OrderOfLiquidity Template")
    print("=" * 80)

    # Create backup directory
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    # Backup original file
    if not BACKUP_PATH.exists():
        import shutil
        shutil.copy2(TEMPLATE_PATH, BACKUP_PATH)
        print(f"Backed up original to: {BACKUP_PATH}\n")
    else:
        print(f"Backup already exists: {BACKUP_PATH}\n")

    # Load workbook
    print(f"Loading template: {TEMPLATE_PATH}")
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["SOFP-Sub-OrdOfLiq"]
    print(f"Sheet: {ws.title}\n")

    # Parse XBRL
    print("Parsing XBRL calculation linkbase...")
    calc_tree, loc_to_concept = parse_calculation_linkbase(CAL_LINKBASE)
    print(f"  Found {len(calc_tree)} parent locators with children")
    print(f"  Found {len(loc_to_concept)} locator -> concept mappings")

    print("Parsing XBRL label linkbase...")
    label_text, label_to_concept = load_labels_from_linkbase(LAB_LINKBASE)
    print(f"  Found {len(label_text)} label texts")
    print(f"  Found {len(label_to_concept)} label -> concept mappings\n")

    # Build concept -> label mapping (from label linkbase)
    concept_to_label_map = {}
    for label_id, (concept, label_str) in label_to_concept.items():
        if concept and label_str:
            concept_to_label_map[concept] = label_str

    print(f"  Built {len(concept_to_label_map)} concept -> label mappings")

    # Build label -> row mapping from Excel
    label_to_row = {}
    for row in range(1, ws.max_row + 1):
        cell_value = ws[f"A{row}"].value
        if cell_value:
            label_to_row[cell_value] = row

    print(f"  Found {len(label_to_row)} labels in Excel template\n")

    # Process all formula cells
    print("Scanning formula cells and building corrections...\n")
    changes = []

    for row in range(1, ws.max_row + 1):
        cell_label = ws[f"A{row}"].value
        cell_b = ws[f"B{row}"]
        cell_c = ws[f"C{row}"]

        # Only process formula cells
        if not (isinstance(cell_b.value, str) and cell_b.value.startswith("=")):
            continue

        old_formula_b = cell_b.value
        old_formula_c = cell_c.value

        # Find matching XBRL concept
        matching_concept = find_concept_for_label(cell_label, concept_to_label_map)

        if not matching_concept:
            print(f"Row {row}: '{cell_label}' - SKIPPED (no XBRL concept match)")
            continue

        # Find locator for this concept
        matching_loc = None
        for loc_id, concept in loc_to_concept.items():
            if concept == matching_concept:
                matching_loc = loc_id
                break

        if not matching_loc:
            print(f"Row {row}: '{cell_label}' - SKIPPED (no locator for concept {matching_concept})")
            continue

        # Build correct formula
        new_formula_b = build_formula_from_tree(
            matching_loc,
            calc_tree,
            loc_to_concept,
            concept_to_label_map,
            label_to_row,
        )

        if not new_formula_b:
            print(f"Row {row}: '{cell_label}' - SKIPPED (no children in calc tree)")
            continue

        # Build C formula (same but with C column)
        new_formula_c = new_formula_b.replace("B", "C")

        # Compare and record changes
        if old_formula_b != new_formula_b:
            changes.append({
                'row': row,
                'label': cell_label,
                'concept': matching_concept,
                'old_b': old_formula_b,
                'new_b': new_formula_b,
                'old_c': old_formula_c,
                'new_c': new_formula_c,
            })

            print(f"Row {row}: '{cell_label}'")
            print(f"  Concept: {matching_concept}")
            print(f"  OLD B: {old_formula_b[:100]}")
            print(f"  NEW B: {new_formula_b[:100]}")
            if old_formula_c != new_formula_c:
                print(f"  OLD C: {old_formula_c[:100]}")
                print(f"  NEW C: {new_formula_c[:100]}")
            print()
        else:
            print(f"Row {row}: '{cell_label}' - OK (no changes needed)")

    # Apply changes
    print("\n" + "=" * 80)
    print(f"Applying {len(changes)} formula fixes...")
    print("=" * 80 + "\n")

    for change in changes:
        row = change['row']
        ws[f"B{row}"].value = change['new_b']
        ws[f"C{row}"].value = change['new_c']
        print(f"Updated Row {row}: {change['label']}")

    # Save
    print(f"\nSaving fixed template to: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)

    print("\n" + "=" * 80)
    print(f"COMPLETE: Fixed {len(changes)} formulas")
    print("=" * 80)

    # Summary
    if changes:
        print("\nSummary of changes:")
        for change in changes:
            print(f"  Row {change['row']}: {change['label']}")

if __name__ == "__main__":
    main()
