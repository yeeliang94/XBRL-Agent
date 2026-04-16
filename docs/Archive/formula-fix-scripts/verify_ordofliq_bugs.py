#!/usr/bin/env python3
"""
Verify the mixed bug claims in TEMPLATE-FORMULA-FIX-GUIDE.md for OrderOfLiquidity.

Checks specifically:
- Row 148 "Total cash": sums inventory + derivative + cash items instead of just cash
- Row 168 "Total issued capital": includes prepaid assets
- Row 241 "Total borrowings": includes equity items
- Row 295 "Total trade and other payables": double-counts by summing both leaf + subtotals

Also scans all formulas to detect any cross-section references.
"""

import re
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, List, Tuple, Optional

# File paths
TEMPLATE_PATH = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent/XBRL-template-MFRS/02-SOFP-OrderOfLiquidity.xlsx")
SHEET_NAME = "SOFP-Sub-OrdOfLiq"

# Target rows from the bug claims
TARGET_ROWS = {
    148: "Total cash",
    168: "Total issued capital",
    241: "Total borrowings",
    295: "Total trade and other payables"
}


def load_labels_and_formulas(wb) -> Tuple[Dict[int, str], Dict[int, Dict[str, str]]]:
    """
    Load all labels from column A and all formulas from columns B/C.

    Returns:
        (labels_by_row, formulas_by_row)
        where formulas_by_row[row] = {"B": formula_str, "C": formula_str}
    """
    ws = wb[SHEET_NAME]

    labels = {}
    formulas = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        row_num = row[0].row

        # Column A: label
        label_cell = row[0]
        if label_cell.value:
            labels[row_num] = str(label_cell.value)

        # Columns B and C: formulas
        b_cell = row[1]
        c_cell = row[2]

        if b_cell.data_type == 'f' or c_cell.data_type == 'f':
            formulas[row_num] = {
                "B": b_cell.value if b_cell.data_type == 'f' else None,
                "C": c_cell.value if c_cell.data_type == 'f' else None,
            }

    return labels, formulas


def extract_cell_references(formula: str) -> List[Tuple[str, int]]:
    """
    Extract all cell references like B123, C456, etc. from a formula.

    Returns:
        List of (column, row_num) tuples
    """
    if not formula:
        return []

    # Pattern: letter(s) followed by digits
    pattern = r'([A-Z]+)(\d+)'
    matches = re.findall(pattern, formula)

    return [(col, int(row)) for col, row in matches]


def get_section_for_row(row_num: int, labels: Dict[int, str]) -> Optional[str]:
    """
    Determine which accounting section a row belongs to by scanning backwards
    for the last major section header (typically marked with asterisk or all caps).
    """
    for search_row in range(row_num, 0, -1):
        label = labels.get(search_row, "").strip()
        if label and label.startswith("*"):
            return f"{search_row}: {label}"
    return "UNKNOWN"


def is_likely_subtotal(label: str) -> bool:
    """Check if a label looks like it's a subtotal (starts with *)."""
    return label.strip().startswith("*")


def analyze_formula(
    row_num: int,
    column: str,
    formula: str,
    labels: Dict[int, str],
) -> Tuple[bool, List[str]]:
    """
    Analyze a formula for potential bugs.

    Returns:
        (has_cross_section_refs, list of issue descriptions)
    """
    if not formula:
        return False, []

    issues = []
    source_section = get_section_for_row(row_num, labels)
    references = extract_cell_references(formula)

    cross_section_refs = []

    for ref_col, ref_row in references:
        if ref_col != column:  # Skip cross-column references for now
            continue

        ref_label = labels.get(ref_row, f"[ROW {ref_row} NOT FOUND]")
        ref_section = get_section_for_row(ref_row, labels)

        # Check if this reference is to a different section
        if source_section != ref_section:
            cross_section_refs.append({
                'ref': f"{ref_col}{ref_row}",
                'label': ref_label,
                'section': ref_section,
                'is_subtotal': is_likely_subtotal(ref_label)
            })

    if cross_section_refs:
        issues.append(f"Cross-section reference(s) detected in {column}{row_num}:")
        for ref in cross_section_refs:
            issues.append(
                f"  → {ref['ref']}: '{ref['label']}' (section: {ref['section']})"
            )

    return len(cross_section_refs) > 0, issues


def main():
    print("=" * 80)
    print(f"VERIFYING ORDEROF LIQUIDITY TEMPLATE BUGS")
    print(f"File: {TEMPLATE_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    print("=" * 80)
    print()

    # Load workbook
    try:
        wb = load_workbook(TEMPLATE_PATH, data_only=False)
    except Exception as e:
        print(f"ERROR: Could not load workbook: {e}")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found in workbook")
        print(f"Available sheets: {wb.sheetnames}")
        return

    labels, formulas = load_labels_and_formulas(wb)

    print(f"Loaded {len(labels)} rows with labels")
    print(f"Found {len(formulas)} rows with formulas")
    print()

    # ========================================================================
    # CHECK SPECIFIC TARGET ROWS FROM BUG CLAIMS
    # ========================================================================
    print("=" * 80)
    print("CHECKING CLAIMED BUG ROWS")
    print("=" * 80)
    print()

    for target_row, expected_label in TARGET_ROWS.items():
        print(f"\n>>> ROW {target_row}: {expected_label}")
        print("-" * 80)

        actual_label = labels.get(target_row, "[NOT FOUND]")
        print(f"  Label in template: '{actual_label}'")

        if actual_label and expected_label.lower() not in actual_label.lower():
            print(f"  ⚠️  WARNING: Expected '{expected_label}' but found '{actual_label}'")

        if target_row not in formulas:
            print(f"  ⚠️  WARNING: No formulas found in row {target_row}")
            continue

        # Check column B
        formula_b = formulas[target_row].get("B")
        if formula_b:
            print(f"\n  Column B formula:")
            print(f"    {formula_b}")

            # Extract references
            refs_b = extract_cell_references(formula_b)
            print(f"    Cell references: {refs_b}")

            # Show what each reference points to
            for col, row in refs_b:
                ref_label = labels.get(row, f"[ROW {row} NOT FOUND]")
                ref_section = get_section_for_row(row, labels)
                print(f"      {col}{row} → '{ref_label}' (section: {ref_section})")

            # Check for cross-section issues
            has_cross, issues = analyze_formula(target_row, "B", formula_b, labels)
            if issues:
                print(f"\n    ⚠️  ISSUES FOUND:")
                for issue in issues:
                    print(f"    {issue}")

        # Check column C
        formula_c = formulas[target_row].get("C")
        if formula_c:
            print(f"\n  Column C formula:")
            print(f"    {formula_c}")

            # Extract references
            refs_c = extract_cell_references(formula_c)
            print(f"    Cell references: {refs_c}")

            # Show what each reference points to
            for col, row in refs_c:
                ref_label = labels.get(row, f"[ROW {row} NOT FOUND]")
                ref_section = get_section_for_row(row, labels)
                print(f"      {col}{row} → '{ref_label}' (section: {ref_section})")

            # Check for cross-section issues
            has_cross, issues = analyze_formula(target_row, "C", formula_c, labels)
            if issues:
                print(f"\n    ⚠️  ISSUES FOUND:")
                for issue in issues:
                    print(f"    {issue}")

    # ========================================================================
    # SCAN ALL FORMULAS FOR CROSS-SECTION REFERENCES
    # ========================================================================
    print("\n\n" + "=" * 80)
    print("FULL SCAN: ALL FORMULAS FOR CROSS-SECTION REFERENCES")
    print("=" * 80)
    print()

    all_issues = []

    for row_num in sorted(formulas.keys()):
        row_label = labels.get(row_num, "[NO LABEL]")

        for column in ["B", "C"]:
            formula = formulas[row_num].get(column)
            if not formula:
                continue

            has_cross, issues = analyze_formula(row_num, column, formula, labels)

            if has_cross:
                all_issues.append({
                    'row': row_num,
                    'label': row_label,
                    'column': column,
                    'formula': formula,
                    'issues': issues
                })

    if all_issues:
        print(f"Found {len(all_issues)} formula(s) with cross-section references:")
        print()

        for item in all_issues:
            print(f"Row {item['row']} ({item['column']}) - {item['label']}:")
            print(f"  Formula: {item['formula']}")
            for issue in item['issues']:
                print(f"  {issue}")
            print()
    else:
        print("No cross-section references detected in full scan.")

    # ========================================================================
    # SUMMARY
    # ========================================================================
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print()

    target_issues = []
    for target_row in TARGET_ROWS.keys():
        if target_row in formulas:
            formula_b = formulas[target_row].get("B")
            if formula_b:
                has_cross, issues = analyze_formula(target_row, "B", formula_b, labels)
                if has_cross:
                    target_issues.append(target_row)

    if target_issues:
        print(f"Claimed bug rows with detected issues: {target_issues}")
    else:
        print("No cross-section issues detected in the claimed bug rows.")

    if all_issues:
        print(f"Total formulas with potential cross-section bugs: {len(all_issues)}")
    else:
        print("No formulas with cross-section bugs found in the entire sheet.")


if __name__ == "__main__":
    main()
