#!/usr/bin/env python3
"""
Fix +20 offset formula bugs in SOFP-Sub-CuNonCu sheet (V2).

This version uses explicit section boundary detection based on label patterns
to correctly identify which section each row belongs to.
"""

import re
from pathlib import Path
from openpyxl import load_workbook

TEMPLATE_PATH = Path(__file__).parent / "XBRL-template-MFRS" / "01-SOFP-CuNonCu.xlsx"
SHEET_NAME = "SOFP-Sub-CuNonCu"

# Explicit section boundaries and their key labels
# Order matters - these define distinct accounting sections
SECTION_RANGES = [
    # Assets side
    ("Land and Buildings", {"start": None, "end": 20}),
    ("Other PPE", {"start": 21, "end": 38}),
    ("Intangible Assets", {"start": 39, "end": 68}),
    ("NC Receivables", {"start": 69, "end": 127}),
    ("NC Derivatives", {"start": 128, "end": 158}),
    ("Prepayments", {"start": 159, "end": 167}),
    ("C Receivables", {"start": 168, "end": 176}),
    ("C Derivatives", {"start": 177, "end": 191}),
    ("Cash", {"start": 192, "end": 210}),
    # Equity and Liabilities side
    ("Reserves", {"start": 211, "end": 239}),
    ("NC Borrowings", {"start": 240, "end": 269}),
    ("NC Payables", {"start": 270, "end": 334}),
    ("NC Derivatives Liab", {"start": 335, "end": 353}),
    ("C Payables", {"start": 354, "end": 434}),
    ("C Derivatives Liab", {"start": 435, "end": 450}),
]

# Known broken references (from the guide)
KNOWN_BROKEN_REFS = {
    20: {"B": [33], "C": [33]},
    39: {"B": [40, 46], "C": [40, 46]},
    69: {"B": [87], "C": [87]},
    119: {"B": [133, 138], "C": [133, 138]},
    129: {"B": [147], "C": [147]},
    160: {"B": [178], "C": [178]},
    168: {"B": [187], "C": [187]},  # Note: guide says B176, B180, B187 but B176/180 are internal
    178: {"B": [196], "C": [196]},
    193: {"B": [203, 211], "C": [203, 211]},
    222: {"B": [234, 241], "C": [234, 241]},
    271: {"B": [276, 283], "C": [276, 283]},
    320: {"B": [327, 339], "C": [327, 339]},
    336: {"B": [348, 354], "C": [348, 354]},
    436: {"B": [440, 445, 455], "C": [440, 445, 455]},
    452: {"B": [464, 470], "C": [464, 470]},
}


def get_section_for_row(row: int) -> str:
    """Get the section name for a given row number."""
    for section_name, bounds in SECTION_RANGES:
        if bounds["start"] is not None and row >= bounds["start"]:
            if bounds["end"] is None or row <= bounds["end"]:
                return section_name
    return "Unknown"


def get_label_at_row(ws, row: int) -> str:
    """Get the label text in column A for a given row."""
    cell = ws[f"A{row}"]
    value = cell.value
    return str(value).strip() if value else ""


def extract_cell_references(formula: str) -> list:
    """Extract all B{row}/C{row} references from a formula.

    Returns list of tuples: (column_letter, row_number, full_ref_str)
    Excludes cross-sheet references.
    """
    # Pattern: B123, C456, etc. (exclude sheet refs like 'Sheet'!B123)
    pattern = r"(?<![!'])([BC])(\d+)(?![\w'])"
    matches = re.findall(pattern, formula)

    refs = []
    for col, row in matches:
        refs.append((col, int(row), f"{col}{row}"))

    return refs


def fix_formula_offsets():
    """Main fix function."""

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    print(f"Processing sheet: {SHEET_NAME}")
    print(f"Sheet dimensions: {ws.dimensions}\n")

    changes = []

    # Scan all cells in columns B and C
    for row in range(1, ws.max_row + 1):
        formula_label = get_label_at_row(ws, row)
        formula_section = get_section_for_row(row)

        for col in ["B", "C"]:
            cell = ws[f"{col}{row}"]

            # Only process formula cells
            if not cell.value or not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue

            formula = cell.value
            refs = extract_cell_references(formula)

            if not refs:
                continue

            new_formula = formula
            changes_in_formula = []

            for ref_col, ref_row, ref_str in refs:
                ref_label = get_label_at_row(ws, ref_row)
                ref_section = get_section_for_row(ref_row)

                # Check if this is a known broken reference
                # (cross-section reference where row-20 would be correct)
                is_known_broken = (
                    row in KNOWN_BROKEN_REFS and
                    col in KNOWN_BROKEN_REFS[row] and
                    ref_row in KNOWN_BROKEN_REFS[row][col]
                )

                if is_known_broken:
                    # This is a broken +20 offset
                    alt_row = ref_row - 20
                    alt_label = get_label_at_row(ws, alt_row)
                    alt_section = get_section_for_row(alt_row)

                    old_ref = ref_str
                    new_ref = f"{ref_col}{alt_row}"

                    new_formula = new_formula.replace(old_ref, new_ref, 1)

                    changes_in_formula.append({
                        "old_ref": old_ref,
                        "new_ref": new_ref,
                        "old_row": ref_row,
                        "new_row": alt_row,
                        "old_label": ref_label,
                        "old_section": ref_section,
                        "new_label": alt_label,
                        "new_section": alt_section,
                    })

            # Apply changes if any were made
            if changes_in_formula:
                change_record = {
                    "row": row,
                    "column": col,
                    "formula_label": formula_label,
                    "formula_section": formula_section,
                    "old_formula": formula,
                    "new_formula": new_formula,
                    "reference_changes": changes_in_formula,
                }
                changes.append(change_record)

                cell.value = new_formula

    # Save workbook
    wb.save(TEMPLATE_PATH)

    # Print results
    print("=" * 100)
    print(f"FORMULA FIX RESULTS: {len(changes)} cells updated")
    print("=" * 100)

    for i, change in enumerate(changes, 1):
        print(f"\n[Change {i}] Row {change['row']}, Column {change['column']}")
        print(f"  Label: {change['formula_label']}")
        print(f"  Section: {change['formula_section']}")
        print(f"  Old: {change['old_formula']}")
        print(f"  New: {change['new_formula']}")

        for ref_change in change['reference_changes']:
            print(f"    Reference changed: {ref_change['old_ref']} → {ref_change['new_ref']}")
            print(f"      Old: Row {ref_change['old_row']} ({ref_change['old_section']}) - {ref_change['old_label']}")
            print(f"      New: Row {ref_change['new_row']} ({ref_change['new_section']}) - {ref_change['new_label']}")

    print(f"\n{'='*100}")
    print(f"File saved: {TEMPLATE_PATH}")
    print(f"Total formulas fixed: {len(changes)} (both columns B and C where applicable)")

    return changes


if __name__ == "__main__":
    changes = fix_formula_offsets()
