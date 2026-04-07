#!/usr/bin/env python3
"""
Verify +20 offset bug claims for XBRL templates.

Focuses on detecting when formulas reference rows that are ~20 rows off
from what they should reference.
"""

import openpyxl
from pathlib import Path
import re
from typing import Dict, List, Tuple, Optional, Set

class OffsetBugDetector:
    def __init__(self, template_path: str):
        self.path = Path(template_path)
        self.wb = openpyxl.load_workbook(self.path, data_only=False)
        self.template_name = self.path.name

    def get_label(self, sheet_name: str, row: int) -> str:
        """Get label from column A at given row."""
        try:
            ws = self.wb[sheet_name]
            cell = ws[f'A{row}']
            val = cell.value
            return str(val).strip() if val else ""
        except:
            return ""

    def extract_cell_refs(self, formula: str) -> List[Tuple[str, int]]:
        """Extract all cell references as (column, row) tuples."""
        refs = []
        # Match patterns: B10, $B$10, etc. (same-sheet refs only)
        pattern = r'[\$]?([A-Z]+)[\$]?(\d+)'
        matches = re.findall(pattern, formula)
        for col, row in matches:
            if col in ['B', 'C', 'D']:  # Only data columns
                refs.append((col, int(row)))
        return refs

    def analyze_sheet(self, sheet_name: str) -> List[Dict]:
        """
        Find formulas that might have +20 offset bugs.

        Returns list of findings with:
        - cell: formula location (B42)
        - label: row label
        - formula: actual formula
        - refs: list of (col, row) references
        - suspicious_refs: refs that seem offset from actual data
        """
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            return []

        findings = []

        # Build a map of "data rows" - rows with values in B or C
        data_rows = set()
        for row in ws.iter_rows(min_row=1, max_row=500, min_col=2, max_col=3):
            for cell in row:
                if cell.value and not isinstance(cell.value, str):
                    # Numeric value = data row
                    data_rows.add(cell.row)
                elif cell.value and isinstance(cell.value, str) and not cell.value.startswith('='):
                    # Text value (not formula) = might be data
                    if cell.value.strip() and not cell.value.strip().startswith('*'):
                        data_rows.add(cell.row)

        # Find all formulas
        for row in ws.iter_rows(min_row=1, max_row=500, min_col=2, max_col=3):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    col = cell.column_letter
                    row_num = cell.row
                    label = self.get_label(sheet_name, row_num)

                    refs = self.extract_cell_refs(formula)
                    if not refs:
                        continue

                    # Check if any ref looks suspiciously offset
                    suspicious = []
                    for ref_col, ref_row in refs:
                        # If reference points to a blank row but row-20 has data,
                        # that's suspicious
                        ref_label = self.get_label(sheet_name, ref_row)
                        if not ref_label or ref_label.strip() == "":
                            # Check if offset versions have data
                            for offset in [20, -20]:
                                offset_row = ref_row + offset
                                if offset_row > 0:
                                    offset_label = self.get_label(sheet_name, offset_row)
                                    if offset_label and offset_label.strip():
                                        suspicious.append({
                                            'ref': f"{ref_col}{ref_row}",
                                            'ref_label': ref_label,
                                            'offset': offset,
                                            'corrected_row': offset_row,
                                            'corrected_label': offset_label
                                        })

                    if suspicious:
                        findings.append({
                            'sheet': sheet_name,
                            'cell': f"{col}{row_num}",
                            'label': label,
                            'formula': formula,
                            'refs': refs,
                            'suspicious_refs': suspicious
                        })

        return findings

    def detailed_analysis(self, sheet_name: str) -> str:
        """Print detailed breakdown of all formulas in a sheet."""
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            return f"Sheet '{sheet_name}' not found"

        output = []
        output.append(f"\nDETAILED ANALYSIS: {sheet_name}")
        output.append("=" * 80)

        # Get row count
        max_row = ws.max_row
        output.append(f"Total rows in sheet: {max_row}")

        # Sample data rows
        output.append("\nData rows (rows with values in columns B or C):")
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=min(100, max_row), min_col=1, max_col=3):
            label = str(row[0].value or "").strip()
            b_val = row[1].value
            c_val = row[2].value

            if (b_val or c_val) and label:
                output.append(f"  Row {row[0].row}: {label[:60]}")
                if b_val:
                    output.append(f"    B{row[0].row}: {b_val}")
                if c_val:
                    output.append(f"    C{row[0].row}: {c_val}")
                count += 1
                if count >= 15:
                    output.append("  ...")
                    break

        # Sample formulas
        output.append("\nFormulas in columns B & C (first 15):")
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=2, max_col=3):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    label = self.get_label(sheet_name, cell.row)
                    output.append(f"  {cell.column_letter}{cell.row}: {label[:50]}")
                    output.append(f"    Formula: {cell.value}")
                    count += 1
                    if count >= 15:
                        output.append("  ...")
                        break
            if count >= 15:
                break

        return "\n".join(output)

def main():
    templates_to_check = {
        "BUGGY (claim: +20 offset)": [
            "XBRL-template-MFRS/07-SOCF-Indirect.xlsx",
            "XBRL-template-MFRS/08-SOCF-Direct.xlsx",
            "XBRL-template-MFRS/09-SOCIE.xlsx",
            "XBRL-template-MFRS/04-SOPL-Nature.xlsx",
        ],
        "CLEAN (claim: no bugs)": [
            "XBRL-template-MFRS/03-SOPL-Function.xlsx",
            "XBRL-template-MFRS/05-SOCI-BeforeTax.xlsx",
            "XBRL-template-MFRS/06-SOCI-NetOfTax.xlsx",
        ]
    }

    sheet_map = {
        "07-SOCF-Indirect.xlsx": "SOCF-Indirect",
        "08-SOCF-Direct.xlsx": "SOCF-Direct",
        "09-SOCIE.xlsx": "SOCIE",
        "04-SOPL-Nature.xlsx": "SOPL-Analysis-Nature",
        "03-SOPL-Function.xlsx": ["SOPL-Analysis-Function", "SOPL-by-Function"],
        "05-SOCI-BeforeTax.xlsx": "SOCI-Before-Tax",
        "06-SOCI-NetOfTax.xlsx": "SOCI-Net-of-Tax",
    }

    base_path = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent")

    print("\n" + "=" * 100)
    print("XBRL TEMPLATE +20 OFFSET BUG VERIFICATION")
    print("=" * 100)

    for category, templates in templates_to_check.items():
        print(f"\n### {category.upper()} ###\n")

        for template in templates:
            full_path = base_path / template
            if not full_path.exists():
                print(f"ERROR: {template} not found")
                continue

            print(f"\n{Path(template).name}")
            print("-" * 80)

            detector = OffsetBugDetector(str(full_path))

            # Print available sheets
            print(f"Available sheets: {detector.wb.sheetnames}")

            # Get sheet name to analyze
            filename = Path(template).name
            sheets_to_check = sheet_map.get(filename, detector.wb.sheetnames)
            if isinstance(sheets_to_check, str):
                sheets_to_check = [sheets_to_check]

            # Filter to sheets that exist
            sheets_to_check = [s for s in sheets_to_check if s in detector.wb.sheetnames]

            for sheet_name in sheets_to_check:
                print(f"\nAnalyzing sheet: {sheet_name}")

                # First print detailed structure
                print(detector.detailed_analysis(sheet_name))

                # Then look for suspicious patterns
                findings = detector.analyze_sheet(sheet_name)
                if findings:
                    print(f"\n  SUSPICIOUS OFFSET PATTERNS FOUND ({len(findings)} cells):")
                    for f in findings[:5]:  # Show first 5
                        print(f"    {f['cell']}: {f['label']}")
                        print(f"      Formula: {f['formula']}")
                        for s in f['suspicious_refs']:
                            print(f"      - {s['ref']} is empty, but {s['ref']} + {s['offset']} = Row {s['corrected_row']} ({s['corrected_label']})")
                else:
                    print(f"\n  No obvious offset patterns detected")

    print("\n" + "=" * 100)
    print("END OF REPORT")
    print("=" * 100)

if __name__ == "__main__":
    main()
