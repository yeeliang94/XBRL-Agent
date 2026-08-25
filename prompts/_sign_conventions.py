"""Build input-sign guidance from the live face-statement templates.

RUN-REVIEW P2-2 (2026-04-26): the Amway run had `(Gain) loss on disposal
of PPE` AI=-70 vs filer=70, and `Cash payments for the principal portion
of lease liabilities` AI=3,732 vs filer=-3,732. Both are valid signs in
isolation; which one is "right" depends on whether the *Total formula
adds or subtracts that cell. Mirroring the ADR-002 pattern for SOCIE
dividends, this module walks the live template's formula bar at
prompt-build time and surfaces a per-row signed-convention block to the
agent.  The other face statements receive concise guidance for their known
directional rows (for example treasury shares and dividends).

Use from the prompt-build path::

    from prompts._sign_conventions import face_sign_convention_block
    extra = face_sign_convention_block(template_path, statement_type)
    if extra:
        prompt += "\n\n" + extra
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Match a single signed-coefficient term in a *Total formula:
#   "+1*B11"        →  sign +1, ref B11
#   "-1*B13"        →  sign -1, ref B13
#   "1*B11"         →  sign +1, ref B11 (leading + omitted)
# The pre-pended sign captures the term-separator from a SUM expression.
_TERM_RE = re.compile(r"([+-]?\s*1)\s*\*\s*([A-Z]+)(\d+)")


@dataclass(frozen=True)
class _SheetSignMetadata:
    """Closed-workbook sign metadata safe to share across agent threads."""

    name: str
    labels_by_row: tuple[tuple[int, str], ...]
    occurrences: tuple[tuple[int, tuple[tuple[int, str], ...]], ...]


@dataclass(frozen=True)
class _TemplateSignMetadata:
    sheets: tuple[_SheetSignMetadata, ...]


def _template_cache_key(template_path: Path) -> tuple[str, int, int]:
    resolved = template_path.resolve()
    stat = resolved.stat()
    return str(resolved), stat.st_mtime_ns, stat.st_size


@lru_cache(maxsize=64)
def _load_template_sign_metadata(
    resolved_path: str,
    _mtime_ns: int,
    _size: int,
) -> _TemplateSignMetadata:
    """Read immutable metadata once per template version, then close it."""
    wb = load_workbook(resolved_path, data_only=False, read_only=True)
    sheets: list[_SheetSignMetadata] = []
    try:
        for ws in wb.worksheets:
            labels_by_row: list[tuple[int, str]] = []
            label_map: dict[int, str] = {}
            for row, (raw_label,) in enumerate(
                ws.iter_rows(min_col=1, max_col=1, values_only=True),
                start=1,
            ):
                label = str(raw_label).strip() if raw_label else ""
                if label:
                    labels_by_row.append((row, label))
                    label_map[row] = label

            occurrences: dict[int, list[tuple[int, str]]] = defaultdict(list)
            if "socf" in ws.title.casefold() or "sore" in ws.title.casefold():
                for row, (formula,) in enumerate(
                    ws.iter_rows(min_col=2, max_col=2, values_only=True),
                    start=1,
                ):
                    parent_label = label_map.get(row, "")
                    lowered = parent_label.casefold()
                    if not parent_label or not (
                        "total" in lowered
                        or parent_label.startswith("*")
                        or lowered.startswith("net ")
                    ):
                        continue
                    if not isinstance(formula, str) or not formula.startswith("="):
                        continue
                    for sign, _column, leaf_row in _parse_total_formula(formula):
                        if leaf_row not in label_map:
                            continue
                        occurrence = (sign, parent_label)
                        if occurrence not in occurrences[leaf_row]:
                            occurrences[leaf_row].append(occurrence)

            sheets.append(
                _SheetSignMetadata(
                    name=ws.title,
                    labels_by_row=tuple(labels_by_row),
                    occurrences=tuple(
                        (row, tuple(uses)) for row, uses in sorted(occurrences.items())
                    ),
                )
            )
    finally:
        wb.close()
    return _TemplateSignMetadata(sheets=tuple(sheets))


def _template_sign_metadata(template_path: Path) -> _TemplateSignMetadata:
    return _load_template_sign_metadata(*_template_cache_key(template_path))


def _clear_template_sign_metadata_cache() -> None:
    """Test/support hook; cached objects contain no workbook handles."""
    _load_template_sign_metadata.cache_clear()


def _parse_total_formula(formula: str) -> list[tuple[int, str, int]]:
    """Return [(sign, col_letter, row), ...] for each ±1*<cell> term.

    Returns empty list for formulas we don't recognise (SUM(), unusual
    forms, multi-row ranges) — the agent falls back to the generic
    sign rules in the prompt for those.
    """
    if not formula or not formula.startswith("="):
        return []
    out: list[tuple[int, str, int]] = []
    for sign_part, col, row in _TERM_RE.findall(formula):
        sign = -1 if "-" in sign_part else 1
        out.append((sign, col, int(row)))
    return out


def _label_at(ws, row: int) -> str:
    val = ws.cell(row, 1).value
    return str(val).strip() if val else ""


def _cash_direction(label: str) -> Optional[str]:
    """Return ``inflow``/``outflow`` only for unambiguous cash-flow labels.

    Directional labels such as ``increase (decrease)``, ``(gain) loss`` and
    ``refund (paid)`` deliberately return ``None``.  Their sign depends on the
    source amount, so a fixed hint would create the same error this helper is
    intended to prevent.
    """
    lowered = label.casefold()
    # Despite starting with "Proceeds", this taxonomy concept is an amount of
    # cash disposed with a discontinued operation.  The SSM calculation
    # linkbase assigns it weight -1 in investing cash flow, so treating it as
    # an ordinary receipts row would invert the agent's input.
    if "cash and cash equivalents disposed" in lowered:
        return None
    if any(
        phrase in lowered
        for phrase in (
            "cash receipts",
            "receipts from",
            "proceeds from",
            "proceeds on",
            "dividends received",
            "interest received",
            "withdrawal",
        )
    ):
        return "inflow"
    if any(
        phrase in lowered
        for phrase in (
            "cash payments",
            "other cash payment",
            "payments to",
            "payments for",
            "payments made",
            "payments of",
            "purchase of",
            "acquisition of",
            "acquisition and subscription",
            "repayments of borrowings",
            "repayment of loan",
            "dividends paid",
            "interest paid",
            "tax paid",
            "deposit placed",
            "development expenditure incurred",
            "issuance expenses",
            "repurchase of treasury shares",
            "cash flows used in obtaining control",
        )
    ):
        return "outflow"
    return None


def _expected_cash_input(
    label: str,
    coefficient: int,
) -> tuple[Optional[bool], Optional[str]]:
    """Return expected positivity and its reason for a clear SOCF row."""
    lowered = label.casefold()
    if "cash and cash equivalents disposed" in lowered and coefficient < 0:
        return (
            True,
            "the SSM linkbase subtracts this discontinued-operation disposed-cash "
            "concept despite its proceeds wording",
        )
    if "bank overdraft" in lowered and coefficient < 0:
        return True, "the closing-cash formula subtracts the overdraft"
    direction = _cash_direction(label)
    if direction is None:
        return None, None
    # V = C / coefficient. With coefficients restricted to ±1, the input is
    # positive for (inflow,+1) and (outflow,-1), negative for the mirror cases.
    positive = (direction == "inflow") == (coefficient > 0)
    action = "adds" if coefficient > 0 else "subtracts"
    return positive, f"the live subtotal {action} this cash {direction}"


def _direct_entry_hint(label: str, coefficient: int) -> str:
    """Translate an unambiguous cash direction into the template input sign."""
    expected_positive, reason = _expected_cash_input(label, coefficient)
    if expected_positive is None:
        return ""
    sign_word = "POSITIVE magnitude" if expected_positive else "NEGATIVE value"
    return f" Enter a {sign_word}: {reason}."


def _formula_occurrences(ws) -> tuple[
    dict[int, list[tuple[int, str]]],
    dict[int, str],
]:
    """Collect each row's signed uses in relevant subtotal formulas."""
    occurrences: dict[int, list[tuple[int, str]]] = defaultdict(list)
    row_labels: dict[int, str] = {}
    for r in range(1, ws.max_row + 1):
        label = _label_at(ws, r)
        if not label:
            continue
        lowered = label.lower()
        if not (
            "total" in lowered
            or label.startswith("*")
            or lowered.startswith("net ")
        ):
            continue
        formula = ws.cell(r, 2).value
        if not isinstance(formula, str) or not formula.startswith("="):
            continue
        terms = _parse_total_formula(formula)
        if not terms:
            continue
        for sign, _col, leaf_row in terms:
            leaf_label = _label_at(ws, leaf_row)
            if not leaf_label:
                continue
            row_labels[leaf_row] = leaf_label
            occurrence = (sign, label)
            if occurrence not in occurrences[leaf_row]:
                occurrences[leaf_row].append(occurrence)
    return occurrences, row_labels


def socf_sign_convention_block(template_path: str | Path) -> Optional[str]:
    """Build a prompt-injectable block listing each row that flows into
    a `*Total …` formula, alongside its add/subtract sign. Serves SOCF
    and the SoRE (SOCIE-family) statement — the title and wording are
    statement-neutral so SoRE no longer receives SOCF-branded prose.

    Returns None if the template can't be read or carries no `*Total`
    formulas — the agent falls back to the static generic rules. The
    matrix SOCIE sheet (named "SOCIE") is filtered out below, so only
    SoRE among the SOCIE family produces a block.

    The block is intentionally compact so the prompt cache stays warm. Each
    row has one output line, but that line preserves every distinct parent
    formula use. Conflicting coefficients are marked ``DUAL-USE``.
    """
    p = Path(template_path)
    if not p.exists():
        return None

    try:
        metadata = _template_sign_metadata(p)
    except Exception:  # noqa: BLE001 - static prompt remains the fallback
        return None

    # SOCF templates have one sheet but the helper is defensive. One row may
    # deliberately feed multiple subtotals, so cached metadata preserves every
    # distinct occurrence rather than a first-wins result.
    target = next(
        (
            sheet
            for sheet in metadata.sheets
            if "socf" in sheet.name.casefold() or "sore" in sheet.name.casefold()
        ),
        None,
    )
    if target is None:
        return None
    occurrences = {row: list(uses) for row, uses in target.occurrences}
    row_labels = dict(target.labels_by_row)

    if not occurrences:
        return None

    lines = [
        "=== PER-ROW SIGN CONVENTIONS — AUTHORITATIVE (from live template formulas) ===",
        "",
        "These signs are read directly from THIS template's live `*Total …`",
        "formulas, so for the rows listed below they OVERRIDE any general",
        "sign rule stated earlier in this prompt. (Rows not listed here fall",
        "back to the general sign rules above — this block is the single",
        "source of truth wherever the two disagree.)",
        "",
        "Each row below appears in a `*Total …` formula with the indicated",
        "coefficient. Enter values to MATCH the formula's intent:",
        "",
    ]
    for leaf_row, row_occurrences in sorted(occurrences.items()):
        full_leaf_label = row_labels.get(leaf_row, "")
        leaf_label = full_leaf_label
        # Truncate very long labels; the agent doesn't need the full
        # SSM URI suffix, just enough to recognise the row.
        if len(leaf_label) > 80:
            leaf_label = leaf_label[:77] + "..."
        coefficients = {sign for sign, _parent in row_occurrences}
        if len(coefficients) > 1:
            roles = "; ".join(
                f"{'ADDED' if sign > 0 else 'SUBTRACTED'} by `{parent[:55]}`"
                for sign, parent in row_occurrences
            )
            lowered = full_leaf_label.casefold()
            if "short-term lease payments" in lowered:
                lines.append(
                    f"- Row {leaf_row} `{leaf_label}` is DUAL-USE: {roles}. "
                    "Enter NEGATIVE for this cash payment; subtraction in the "
                    "adjustment subtotal adds the expense back, while addition "
                    "in operating cash records the outflow."
                )
            else:
                lines.append(
                    f"- Row {leaf_row} `{leaf_label}` is DUAL-USE: {roles}. "
                    "Do not infer its input sign from only one parent formula; "
                    "reconcile every listed subtotal."
                )
            continue

        sign = next(iter(coefficients))
        sign_word = "ADDED" if sign > 0 else "SUBTRACTED"
        lines.append(
            f"- Row {leaf_row} `{leaf_label}` is {sign_word} by its total."
            + _direct_entry_hint(full_leaf_label, sign)
        )
    lines.append("")
    lines.append(
        "THE ONE RULE THAT ALWAYS WORKS (apply it to every row, especially "
        "when a row's name is ambiguous): first decide the line's actual "
        "contribution C to the target subtotal — POSITIVE for an inflow, a "
        "non-cash add-back, or a loss added back; NEGATIVE for an outflow, a "
        "deduction from profit, or a gain reversed out. Then enter V = C / "
        "coefficient. So for an ADDED row "
        "(coefficient +1) enter V = C as-is; for a SUBTRACTED row (coefficient "
        "-1) enter V = -C — flip the sign, because the formula flips it back. "
        "This rule needs no judgement about the row's name and works for every "
        "row, statement and standard. The name-based hints below are just this "
        "rule spelled out for the common cases."
    )
    lines.append("")
    lines.append(
        "Worked examples of V = C / coefficient on SUBTRACTED rows (the case "
        "most often entered backwards):"
    )
    lines.append(
        "  - A SUBTRACTED gain on disposal of PPE has cash contribution "
        "C = -31,276 (a gain is deducted from profit) → enter V = -C = "
        "+31,276. Do NOT pre-negate the gain: the formula already subtracts "
        "the row, so entering -31,276 would double-negate and wrongly ADD the "
        "gain back to operating cash."
    )
    lines.append(
        "  - A SUBTRACTED non-cash add-back — e.g. 'Adjustments for accrued "
        "expenses (income) not yet paid (received)' — has C = +62,264 (a "
        "non-cash accrual added back to profit) → enter V = -C = -62,264. The "
        "blanket 'enter a positive magnitude' instinct is WRONG here: on a "
        "SUBTRACTED row a positive entry produces a NEGATIVE contribution."
    )
    lines.append("")
    lines.append(
        "If the formula ADDS a row, the total uses the cell's value AS-IS "
        "(no sign flip), so YOU must supply the correct sign:"
    )
    lines.append(
        "  - An ADDED row that is a cash OUTFLOW — its name is a 'payment', "
        "'repayment', 'purchase', 'repurchase', 'acquisition', 'deposit "
        "placed', a '…paid' line (dividends/interest/tax paid), or issuance "
        "'expenses' — takes a NEGATIVE value. Do NOT enter the bare positive "
        "magnitude: because the total ADDS (not subtracts) the cell, a "
        "positive number would wrongly INCREASE the section subtotal. "
        "(Worked example: 'Cash payments for the principal portion of the "
        "lease liability' is ADDED, so enter -3,732, NOT 3,732.)"
    )
    lines.append(
        "  - An ADDED row that is a cash INFLOW — 'Proceeds', 'Receipts', "
        "'Withdrawal', 'Dividends received', 'Interest received' — takes a "
        "POSITIVE value."
    )
    lines.append(
        "  - An ADDED gain/loss adjustment row follows its directional name: "
        "a 'Loss on disposal' takes a POSITIVE loss magnitude; a gain takes "
        "NEGATIVE."
    )
    lines.append(
        "If the formula SUBTRACTS a row, the total flips the cell's sign, so "
        "enter V = -C (the negative of the line's cash contribution):"
    )
    lines.append(
        "  - A SUBTRACTED cash OUTFLOW — 'Dividends paid', 'Cash payments', "
        "tax/interest paid — has C negative, so V = -C is a POSITIVE magnitude "
        "(do NOT pre-negate it)."
    )
    lines.append(
        "  - A SUBTRACTED gain/loss adjustment is the MIRROR of the ADDED "
        "gain/loss rule: a GAIN (C negative, deducted from profit) → enter a "
        "POSITIVE magnitude; a LOSS (C positive, added back) → enter NEGATIVE."
    )
    lines.append(
        "  - A SUBTRACTED non-cash add-back (accruals/provisions not yet paid, "
        "C positive) → enter NEGATIVE."
    )
    lines.append(
        "NOTE the contrast between the two branches: the SAME 'Cash payments' "
        "or 'gain on disposal' wording flips entry sign depending on whether "
        "THIS template's formula adds or subtracts that specific row — always "
        "obey the per-row ADDED/SUBTRACTED label listed above, not the row's "
        "name alone."
    )
    return "\n".join(lines)


def _live_labels(template_path: Path) -> tuple[list[str], list[str]]:
    """Return unique live labels and sheet names without retaining handles."""
    metadata = _template_sign_metadata(template_path)
    labels: list[str] = []
    seen: set[str] = set()
    for sheet in metadata.sheets:
        for _row, label in sheet.labels_by_row:
            if label.casefold() not in seen:
                seen.add(label.casefold())
                labels.append(label)
    return labels, [sheet.name for sheet in metadata.sheets]


def _find_live_label(labels: list[str], needle: str) -> Optional[str]:
    lowered = needle.casefold()
    return next((label for label in labels if lowered in label.casefold()), None)


def _non_socf_sign_block(
    template_path: Path,
    statement_key: str,
    labels: Optional[list[str]] = None,
) -> Optional[str]:
    """Build concise, audited sign guidance for the other face statements."""
    try:
        if labels is None:
            labels, _sheet_names = _live_labels(template_path)
    except Exception:  # noqa: BLE001 - caller retains static prompt fallback
        return None

    lines = [
        "=== FACE INPUT SIGN CONVENTIONS — AUTHORITATIVE (live template) ===",
        "",
        "Store the final signed number expected by this template. mTool exports",
        "that stored number unchanged; never apply a second sign conversion.",
        "",
    ]

    if statement_key == "SOFP":
        treasury = _find_live_label(labels, "Treasury shares")
        if treasury:
            lines.append(
                f"- `{treasury}`: Enter a POSITIVE magnitude. The live total-equity "
                "formula subtracts this balance. Do not pre-negate it."
            )
        lines.append(
            "- Retained earnings/accumulated losses are different: because the "
            "equity formula adds that row, an accumulated-loss balance is NEGATIVE."
        )
    elif statement_key == "SOPL":
        lines.extend([
            "- Ordinary expense, cost, tax-charge and loss rows take POSITIVE "
            "magnitudes; the profit calculation subtracts their expense totals.",
            "- Revenue, ordinary income and ordinary gain rows take POSITIVE "
            "magnitudes when their profit total adds them. Do not negate all "
            "income merely because expenses use positive inputs.",
            "- Parenthetical directional rows are not ordinary rows: resolve the "
            "stated direction and apply the relevant live subtotal. The audited "
            "inventory and tax conventions below are fixed examples.",
        ])
        inventory = _find_live_label(labels, "inventories of finished goods")
        if inventory:
            lines.append(
                f"- `{inventory}`: inventory decrease = POSITIVE; increase = NEGATIVE."
            )
        impairment = _find_live_label(labels, "(Reversal of)/Impairment loss")
        if impairment:
            lines.append(
                f"- `{impairment}`: reversal = POSITIVE; impairment loss = NEGATIVE."
            )
        tax = _find_live_label(labels, "Tax expense (income)") or _find_live_label(
            labels, "Total tax expense (income)"
        )
        if tax:
            lines.append(f"- `{tax}`: tax expense = POSITIVE; tax income = NEGATIVE.")
    elif statement_key == "SOCI":
        lines.extend([
            "- OCI gain (loss) rows are ADDED: gain = POSITIVE; loss = NEGATIVE.",
            "- Reclassification adjustments and amounts removed from equity are "
            "SUBTRACTED: a positive amount reclassified out is a POSITIVE magnitude; "
            "a reversal is NEGATIVE.",
            "- OCI tax rows are SUBTRACTED: tax charge = POSITIVE; tax benefit = NEGATIVE.",
        ])
        reclass = _find_live_label(labels, "Reclassification adjustments")
        if reclass:
            lines.append(
                f"- Live example `{reclass}` follows the reclassification rule above."
            )
    elif statement_key == "SOCIE":
        dividends = _find_live_label(labels, "Dividends paid")
        treasury = _find_live_label(labels, "Treasury shares transactions")
        if dividends:
            lines.append(
                f"- `{dividends}`: Enter a POSITIVE magnitude; the live movement "
                "formula subtracts it."
            )
        if treasury:
            lines.append(
                f"- `{treasury}`: The movement formula ADDS this row, so a treasury-"
                "share reduction is NEGATIVE and an increase is POSITIVE."
            )
        lines.append(
            "- Other increase (decrease) movement rows are ADDED and therefore keep "
            "their economic sign: increase POSITIVE, decrease NEGATIVE."
        )
    else:
        return None

    return "\n".join(lines)


def face_sign_convention_block(
    template_path: str | Path,
    statement_type: Any,
) -> Optional[str]:
    """Return authoritative live-template sign guidance for any face statement."""
    p = Path(template_path)
    if not p.exists():
        return None
    key = str(getattr(statement_type, "value", statement_type)).upper()
    if key == "SOCF":
        return socf_sign_convention_block(p)
    try:
        labels, _sheet_names = _live_labels(p)
    except Exception:  # noqa: BLE001 - prompt construction must degrade safely
        return None
    # SoRE already routes as StatementType.SOCIE. The equity-movement block is
    # intentionally shared; it avoids the cash-flow examples in the legacy
    # lower-level helper while retaining the dividend convention.
    return _non_socf_sign_block(p, key, labels)


def sign_warnings_for_resolved_writes(
    template_path: str | Path,
    resolved_writes: list[dict],
) -> list[str]:
    """Return advisory sign warnings without changing any written value.

    Only fixed conventions and formula-determinable cash directions are
    checked. Directional/reversal concepts remain model judgement and are not
    warned, avoiding false corrections of legitimate negative facts.
    """
    if not resolved_writes:
        return []
    p = Path(template_path)
    if not p.exists():
        return []
    try:
        metadata = _template_sign_metadata(p)
    except Exception:  # noqa: BLE001 - sign checks are advisory
        return []

    warnings: list[str] = []
    sheets = {sheet.name: sheet for sheet in metadata.sheets}
    label_maps = {
        sheet.name: dict(sheet.labels_by_row) for sheet in metadata.sheets
    }
    occurrence_maps = {
        sheet.name: dict(sheet.occurrences) for sheet in metadata.sheets
    }
    for write in resolved_writes:
        value = write.get("value")
        if isinstance(value, bool) or not isinstance(value, (int, float)) or value == 0:
            continue
        sheet_name = str(write.get("sheet") or "")
        row = write.get("row")
        col = write.get("col")
        sheet = sheets.get(sheet_name)
        if sheet is None or not isinstance(row, int):
            continue
        label = label_maps[sheet_name].get(row, "")
        lowered = label.casefold()
        sheet_key = sheet_name.casefold()
        expected_positive: Optional[bool] = None
        reason = ""

        if "sofp" in sheet_key and lowered == "treasury shares":
            expected_positive = True
            reason = "the live total-equity formula subtracts this balance"
        elif ("socie" in sheet_key or "sore" in sheet_key) and lowered == "dividends paid":
            expected_positive = True
            reason = "the live equity-movement formula subtracts dividends"
        elif "socf" in sheet_key:
            row_occurrences = occurrence_maps[sheet_name].get(row, ())
            coefficients = {sign for sign, _parent in row_occurrences}
            if len(coefficients) > 1 and "short-term lease payments" in lowered:
                expected_positive = False
                reason = (
                    "this dual-use row is added back in adjustments and "
                    "added as the actual operating cash payment"
                )
            elif len(coefficients) == 1:
                coefficient = next(iter(coefficients))
                expected_positive, expected_reason = _expected_cash_input(
                    label,
                    coefficient,
                )
                if expected_reason is not None:
                    reason = expected_reason

        if expected_positive is None:
            continue
        has_positive_sign = value > 0
        if has_positive_sign == expected_positive:
            continue
        expected = "POSITIVE magnitude" if expected_positive else "NEGATIVE value"
        coord = f"{get_column_letter(int(col))}{row}" if col else f"row {row}"
        warnings.append(
            f"Sign check for {sheet_name}!{coord} ('{label}'): expected a "
            f"{expected} because {reason}; got {value}. The value was "
            "written unchanged and mTool will also export it unchanged. "
            "Correct it unless the source documents a genuine exception."
        )
    return warnings
