#!/usr/bin/env python3
"""
Manually fix the 4 known broken formulas in SOFP-OrderOfLiquidity template.

Based on visual inspection of the template structure and accounting logic,
these formulas are mixing sections or double-counting.
"""

from openpyxl import load_workbook
from pathlib import Path

TEMPLATE_PATH = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent/XBRL-template-MFRS/02-SOFP-OrderOfLiquidity.xlsx")
BACKUP_PATH = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent/backup-originals/02-SOFP-OrderOfLiquidity.xlsx")

def get_formula_parts_from_rows(rows: list) -> str:
    """Build formula: =1*B{row1}+1*B{row2}+..."""
    parts = [f"1*B{row}" for row in rows]
    return "=" + "+".join(parts)

def main():
    print("=" * 80)
    print("Manual Fix for Known Formula Bugs in SOFP-OrderOfLiquidity")
    print("=" * 80)
    print()

    # Load workbook
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["SOFP-Sub-OrdOfLiq"]

    fixes = []

    # FIX 1: Row 148 - Total cash
    # Should sum ONLY direct cash items (rows 146-147), not inventory or derivatives
    old_b_148 = ws["B148"].value
    new_b_148 = get_formula_parts_from_rows([146, 147])
    new_c_148 = new_b_148.replace("B", "C")
    fixes.append({
        'row': 148,
        'label': ws["A148"].value,
        'reason': 'Included inventory and derivative items - should only sum cash items (146-147)',
        'old_b': old_b_148,
        'new_b': new_b_148,
        'old_c': ws["C148"].value,
        'new_c': new_c_148,
    })

    # FIX 2: Row 168 - Total issued capital
    # Should sum ONLY capital items (rows 165-167), not prepaid assets (rows 160-162)
    old_b_168 = ws["B168"].value
    new_b_168 = get_formula_parts_from_rows([165, 166, 167])
    new_c_168 = new_b_168.replace("B", "C")
    fixes.append({
        'row': 168,
        'label': ws["A168"].value,
        'reason': 'Included prepaid assets - should only sum capital items (165-167)',
        'old_b': old_b_168,
        'new_b': new_b_168,
        'old_c': ws["C168"].value,
        'new_c': new_c_168,
    })

    # FIX 3: Row 241 - Total borrowings
    # Should sum ONLY borrowing subtotals (rows 207, 218, 225, 232, 240)
    # NOT equity components (rows 189-194)
    old_b_241 = ws["B241"].value
    new_b_241 = get_formula_parts_from_rows([207, 218, 225, 232, 240])
    new_c_241 = new_b_241.replace("B", "C")
    fixes.append({
        'row': 241,
        'label': ws["A241"].value,
        'reason': 'Included equity items and subtotals - should only sum borrowing subtotals',
        'old_b': old_b_241,
        'new_b': new_b_241,
        'old_c': ws["C241"].value,
        'new_c': new_c_241,
    })

    # FIX 4: Row 295 - Total trade and other payables
    # Should sum ONLY the three payables subtotals (rows 270, 294)
    # NOT the individual items or intermediate subtotals
    # Row 270: Trade payables (subtotal)
    # Row 294: Other payables (subtotal - which includes rows 278, 283, 293)
    old_b_295 = ws["B295"].value
    new_b_295 = get_formula_parts_from_rows([270, 294])
    new_c_295 = new_b_295.replace("B", "C")
    fixes.append({
        'row': 295,
        'label': ws["A295"].value,
        'reason': 'Double-counted leaf items and subtotals - should sum only main subtotals (270, 294)',
        'old_b': old_b_295,
        'new_b': new_b_295,
        'old_c': ws["C295"].value,
        'new_c': new_c_295,
    })

    # Apply fixes
    print("Formula Corrections:")
    print("-" * 80)
    print()

    for fix in fixes:
        row = fix['row']
        label = fix['label']
        reason = fix['reason']
        old_b = fix['old_b']
        new_b = fix['new_b']

        print(f"Row {row}: {label}")
        print(f"  Reason: {reason}")
        print(f"  OLD: {old_b[:80]}{'...' if len(old_b) > 80 else ''}")
        print(f"  NEW: {new_b}")
        print()

        # Apply the fix
        ws[f"B{row}"].value = new_b
        ws[f"C{row}"].value = fix['new_c']

    # Save
    print("=" * 80)
    print(f"Saving {len(fixes)} fixes to: {TEMPLATE_PATH}")
    print("=" * 80)
    wb.save(TEMPLATE_PATH)

    print("\nDone!")

if __name__ == "__main__":
    main()
