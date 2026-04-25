"""Cross-check correction agent factory (Phase 3).

Runs exactly once per merged workbook after cross-checks fail. The agent
receives the list of failed `CrossCheckResult` objects plus the merged
workbook + PDF path, and is empowered to:

    - view PDF pages (same tool the face agents use),
    - rewrite cells in the merged workbook via `fill_workbook`,
    - re-run the intra-statement verifier on an edited sheet,
    - re-run cross-statement checks to confirm the fix.

Bounded to 1 iteration by the coordinator; remaining failures after that
surface in the Validator tab for human review.
"""
from __future__ import annotations

import json
import logging
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Iterable, List, Optional, Set, Union

import openpyxl
from pydantic_ai import Agent, RunContext
from pydantic_ai.messages import BinaryContent
from pydantic_ai.models import Model
from pydantic_ai.settings import ModelSettings

from cross_checks.framework import CrossCheckResult
from statement_types import StatementType
from tools.fill_workbook import fill_workbook as _fill_workbook_impl
from tools.pdf_viewer import count_pdf_pages, render_pages_to_png_bytes
from tools.verifier import verify_statement as _verify_statement_impl

logger = logging.getLogger(__name__)

_PROMPT_PATH = Path(__file__).resolve().parent.parent / "prompts" / "correction.md"


class CorrectionAgentDeps:
    """Dependencies carried through the correction agent's tool calls."""

    def __init__(
        self,
        merged_workbook_path: str,
        pdf_path: str,
        failed_checks: List[CrossCheckResult],
        infopack: Any,  # scout.infopack.Infopack or None
        filing_level: str,
        filing_standard: str,
        output_dir: str,
        model: Any,
        statements_to_run: Iterable[StatementType],
        variants: Optional[dict] = None,
    ):
        self.merged_workbook_path = merged_workbook_path
        self.pdf_path = pdf_path
        self.failed_checks = list(failed_checks)
        self.infopack = infopack
        self.filing_level = filing_level
        self.filing_standard = filing_standard
        self.output_dir = output_dir
        self.model = model
        # Peer-review S2: thread the outer run's variant map so the agent's
        # own `run_cross_checks` tool sees the same variant-gating as the
        # server's post-run re-check. Without this, variant-gated checks
        # (e.g. SoRE retained-earnings for MPERS) would mark themselves
        # not_applicable during the correction agent's re-verify and the
        # agent would never learn it's fixed them.
        self.variants = dict(variants) if variants else {}
        # Scope the agent's run_cross_checks tool to the same statements the
        # outer run actually extracted. Using set(StatementType) here would
        # invent missing-sheet failures for statements that were never part
        # of the run (e.g. a SOFP-only run would see fake SOPL/SOCI/SOCIE
        # failures after fixing its real one).
        self.statements_to_run: Set[StatementType] = {
            StatementType(s) if not isinstance(s, StatementType) else s
            for s in statements_to_run
        }
        # Every write bumps this so we can tell whether the agent actually
        # touched the workbook. The coordinator inspects it to decide
        # whether to re-run cross-checks.
        self.writes_performed = 0
        self.pdf_page_count = 0


def _render_single_page(pdf_path: str, page_num: int, dpi: int = 200):
    images = render_pages_to_png_bytes(pdf_path, start=page_num, end=page_num, dpi=dpi)
    return page_num, images[0]


def _format_failed_checks(failed: List[CrossCheckResult]) -> str:
    """Render the failed-checks block for the correction agent prompt."""
    lines = ["=== FAILED CROSS-CHECKS ==="]
    if not failed:
        lines.append("(none — coordinator should not have launched you)")
        return "\n".join(lines)
    for cr in failed:
        bits = [f"• {cr.name}"]
        if cr.expected is not None or cr.actual is not None:
            bits.append(
                f"  expected={cr.expected} actual={cr.actual} diff={cr.diff}"
            )
        if cr.message:
            bits.append(f"  {cr.message}")
        lines.append("\n".join(bits))
    return "\n".join(lines)


def _format_page_hints(infopack: Any) -> str:
    """Render infopack-derived page hints for the prompt. Empty on None."""
    if infopack is None:
        return ""
    lines = ["=== SCOUT PAGE HINTS ==="]
    # scout.infopack.Infopack carries per-statement face_page entries and
    # a notes_page_hints() helper. We avoid importing the class to keep
    # this module agnostic of scout optional imports.
    face_refs = getattr(infopack, "face_refs", None) or {}
    for stmt, ref in face_refs.items():
        page = getattr(ref, "face_page", None)
        if page:
            lines.append(f"- {stmt}: face page {page}")
    note_hints = []
    try:
        if hasattr(infopack, "notes_page_hints"):
            note_hints = list(infopack.notes_page_hints()) or []
    except Exception:  # noqa: BLE001 — advisory block
        note_hints = []
    if note_hints:
        lines.append(f"- note pages: {sorted(set(note_hints))[:20]}")
    if len(lines) == 1:
        return ""  # no hints to show
    return "\n".join(lines)


def _norm_label(value: Any) -> str:
    return str(value or "").strip().lstrip("*").strip().lower()


def _coerce_int(value: Any, default: int, lo: int, hi: int) -> int:
    """Coerce a JSON scalar to an int and clamp to [lo, hi].

    Falls back to ``default`` when the value is None, missing, or
    cannot be parsed (e.g. the model emits ``"context_rows": "nearby"``).
    Avoids leaking a ValueError out of the tool body — pydantic-ai
    aborts the agent run on unhandled tool exceptions.
    """
    if value is None:
        coerced = default
    else:
        try:
            coerced = int(value)
        except (TypeError, ValueError):
            coerced = default
    return max(lo, min(coerced, hi))


def _cell_summary(ws, row: int, max_col: int = 8) -> str:
    parts: list[str] = []
    for col in range(1, min(ws.max_column, max_col) + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            parts.append(f"{cell.coordinate}={cell.value}")
    return "; ".join(parts)


def _nearby_formula_refs(ws, target_row: int, window: int = 10, max_col: int = 8) -> list[str]:
    refs: list[str] = []
    start = max(1, target_row - window)
    end = min(ws.max_row, target_row + window)
    for row in range(start, end + 1):
        for col in range(2, min(ws.max_column, max_col) + 1):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            # Column-specific refs catch the useful case (e.g. B24 subtracts B17)
            # without flooding the model with unrelated row-number hits.
            target_ref = f"{ws.cell(row=target_row, column=col).coordinate}"
            if target_ref in value:
                refs.append(f"{cell.coordinate}: {value}")
    return refs[:12]


def create_correction_agent(
    merged_workbook_path: str,
    pdf_path: str,
    failed_checks: List[CrossCheckResult],
    infopack: Any,
    filing_level: str,
    filing_standard: str,
    model: Union[str, Model],
    output_dir: str,
    statements_to_run: Iterable[StatementType],
    variants: Optional[dict] = None,
) -> tuple[Agent[CorrectionAgentDeps, str], CorrectionAgentDeps]:
    """Build a correction agent wired to the four required tools.

    Exactly one instance is created per run; the coordinator invokes it at
    most once (PLAN D4). Failures that remain after this single pass are
    surfaced to the Validator tab, not retried.
    """
    deps = CorrectionAgentDeps(
        merged_workbook_path=merged_workbook_path,
        pdf_path=pdf_path,
        failed_checks=failed_checks,
        infopack=infopack,
        filing_level=filing_level,
        filing_standard=filing_standard,
        output_dir=output_dir,
        model=model,
        statements_to_run=statements_to_run,
        variants=variants,
    )

    system_prompt_parts: list[str] = [_PROMPT_PATH.read_text(encoding="utf-8").strip()]
    system_prompt_parts.append(_format_failed_checks(failed_checks))
    hints = _format_page_hints(infopack)
    if hints:
        system_prompt_parts.append(hints)
    system_prompt_parts.append(
        f"=== RUN CONTEXT ===\n"
        f"filing_level: {filing_level}\n"
        f"filing_standard: {filing_standard}\n"
        f"merged_workbook: {merged_workbook_path}\n"
    )
    system_prompt = "\n\n".join(system_prompt_parts)

    # Pin temperature=1.0 for the same reason as extraction/agent.py — Gemini 3
    # through the enterprise proxy requires it.
    agent = Agent(
        model,
        deps_type=CorrectionAgentDeps,
        system_prompt=system_prompt,
        model_settings=ModelSettings(temperature=1.0),
    )

    @agent.tool
    def inspect_workbook(
        ctx: RunContext[CorrectionAgentDeps],
        query_json: str,
    ) -> str:
        """Inspect labels, values, and nearby formulas in the merged workbook.

        Args:
            query_json: JSON object:
              {
                "sheet": "SOPL-Nature",              // optional; omit to search all sheets
                "labels": ["Finance costs"],         // optional label substrings
                "rows": [17, 24],                    // optional exact row numbers
                "context_rows": 1,                   // optional surrounding rows (default 1)
                "max_col": 8                         // optional display cap (default 8)
              }

        Use this before sign-sensitive corrections. Formula rows near a
        target data row reveal whether the template adds or subtracts that
        row, which is safer than guessing from words like "loss" or "paid".
        """
        try:
            query = json.loads(query_json or "{}")
        except json.JSONDecodeError as exc:
            return f"Invalid query_json: {exc}"

        sheet_filter = query.get("sheet")
        labels = [
            _norm_label(label)
            for label in query.get("labels", []) or []
            if str(label).strip()
        ]
        rows = {
            int(row)
            for row in query.get("rows", []) or []
            if isinstance(row, int) or str(row).isdigit()
        }
        context_rows = _coerce_int(
            query.get("context_rows"), default=1, lo=0, hi=5,
        )
        max_col = _coerce_int(
            query.get("max_col"), default=8, lo=2, hi=24,
        )

        try:
            wb = openpyxl.load_workbook(ctx.deps.merged_workbook_path, data_only=False)
        except Exception as exc:  # noqa: BLE001
            return f"Could not open merged workbook: {type(exc).__name__}: {exc}"

        try:
            sheets = [sheet_filter] if sheet_filter else list(wb.sheetnames)
            missing = [s for s in sheets if s not in wb.sheetnames]
            if missing:
                return f"Sheet(s) not found: {missing}. Available: {wb.sheetnames}"

            lines: list[str] = ["=== Workbook inspection ==="]
            match_count = 0
            for sheet_name in sheets:
                ws = wb[sheet_name]
                target_rows: set[int] = set(r for r in rows if 1 <= r <= ws.max_row)
                if labels:
                    for row in range(1, ws.max_row + 1):
                        label = _norm_label(ws.cell(row=row, column=1).value)
                        if label and any(term in label for term in labels):
                            target_rows.add(row)

                for target_row in sorted(target_rows):
                    match_count += 1
                    start = max(1, target_row - context_rows)
                    end = min(ws.max_row, target_row + context_rows)
                    lines.append(f"\n-- {sheet_name} row {target_row} --")
                    for row in range(start, end + 1):
                        prefix = ">" if row == target_row else " "
                        lines.append(f"{prefix} row {row}: {_cell_summary(ws, row, max_col=max_col)}")
                    refs = _nearby_formula_refs(ws, target_row, max_col=max_col)
                    if refs:
                        lines.append("Nearby formulas referencing target row:")
                        lines.extend(f"  {ref}" for ref in refs)

            if match_count == 0:
                lines.append(
                    "No matching rows found. Try a broader label substring "
                    "or provide exact sheet/row from the failed check."
                )
            return "\n".join(lines)
        finally:
            wb.close()

    @agent.tool
    def view_pdf_pages(
        ctx: RunContext[CorrectionAgentDeps], pages: List[int],
    ) -> List[Union[str, BinaryContent]]:
        """View specific PDF pages as PNG images."""
        ctx.deps.pdf_page_count = count_pdf_pages(ctx.deps.pdf_path)
        total_pages = ctx.deps.pdf_page_count
        requested = [p for p in pages if isinstance(p, int)]
        invalid = sorted({p for p in requested if p < 1 or p > total_pages})
        render_pages = sorted(set(p for p in requested if p not in invalid))

        results: List[Union[str, BinaryContent]] = []
        if invalid:
            results.append(
                f"Skipped invalid page(s) {invalid}. Valid range is 1-{total_pages}."
            )
        if not render_pages:
            results.append("No pages were rendered from this request.")
            return results

        rendered: dict[int, bytes] = {}
        with ThreadPoolExecutor(max_workers=min(len(render_pages), 8)) as pool:
            futures = {
                pool.submit(_render_single_page, ctx.deps.pdf_path, p): p
                for p in render_pages
            }
            for future in futures:
                page_num, png_bytes = future.result()
                rendered[page_num] = png_bytes

        for p in sorted(rendered):
            results.append(f"=== Page {p} ===")
            results.append(BinaryContent(data=rendered[p], media_type="image/png"))
        return results

    @agent.tool
    def fill_workbook(ctx: RunContext[CorrectionAgentDeps], fields_json: str) -> str:
        """Rewrite cells in the merged workbook.

        Uses the standard fill_workbook writer with the merged workbook
        as both source and target — edits overwrite existing values in
        place. Accepts the same JSON shape as the extraction agent tool.
        """
        path = ctx.deps.merged_workbook_path
        result = _fill_workbook_impl(
            template_path=path,
            output_path=path,
            fields_json=fields_json,
            filing_level=ctx.deps.filing_level,
        )
        if result.success:
            ctx.deps.writes_performed += result.fields_written
            msg = (
                f"Successfully wrote {result.fields_written} fields to the "
                f"merged workbook."
            )
            if result.errors:
                msg += f"\nErrors: {result.errors}"
            return msg
        return f"Failed to fill workbook. Errors: {result.errors}"

    @agent.tool
    def verify_totals(ctx: RunContext[CorrectionAgentDeps], statement: str) -> str:
        """Re-verify a single statement sheet after a correction.

        ``statement`` is the StatementType value (e.g. "SOFP"). Variants
        are resolved by the verifier from the sheet labels themselves —
        the correction agent does not need to know which variant was
        originally selected.
        """
        try:
            stmt = StatementType(statement.upper())
        except ValueError:
            return (
                f"Unknown statement type {statement!r}. "
                f"Valid values: {[s.value for s in StatementType]}"
            )
        result = _verify_statement_impl(
            ctx.deps.merged_workbook_path,
            stmt,
            variant="",  # verifier auto-detects from sheet names
            filing_level=ctx.deps.filing_level,
        )
        lines = [
            f"Statement: {statement}",
            f"Balanced: {result.is_balanced}",
        ]
        if result.mismatches:
            lines.append("Mismatches: " + json.dumps(result.mismatches, indent=2))
        if result.mandatory_unfilled:
            lines.append(
                "Mandatory unfilled: " + json.dumps(result.mandatory_unfilled)
            )
        if result.feedback:
            lines.append(f"Feedback: {result.feedback}")
        return "\n".join(lines)

    @agent.tool
    def run_cross_checks(ctx: RunContext[CorrectionAgentDeps]) -> str:
        """Re-run cross-statement checks against the merged workbook.

        Returns a summary so the agent can decide whether further edits
        are needed. The coordinator still has the final word — it bounds
        the correction agent to one iteration and inspects the result
        itself.
        """
        # Import the registry from its home in cross_checks.framework —
        # peer-review I2 moved it here from server.py to break a latent
        # circular-import hazard.
        from cross_checks.framework import (
            run_all,
            build_default_cross_checks,
            DEFAULT_TOLERANCE_RM,
        )

        checks = build_default_cross_checks()
        # Scope wb_paths + statements_to_run to the statements the outer
        # run actually extracted. Widening to set(StatementType) here
        # invents fake missing-sheet failures on partial runs and misleads
        # the agent into thinking it hasn't fixed the real issue.
        statements = set(ctx.deps.statements_to_run)
        wb_paths = {s: ctx.deps.merged_workbook_path for s in statements}
        run_config = {
            "statements_to_run": statements,
            "variants": getattr(ctx.deps, "variants", None) or {},
            "filing_level": ctx.deps.filing_level,
            "filing_standard": ctx.deps.filing_standard,
        }
        results = run_all(
            checks, wb_paths, run_config, tolerance=DEFAULT_TOLERANCE_RM,
        )
        lines = ["=== Cross-check results ==="]
        for cr in results:
            lines.append(
                f"- {cr.name}: {cr.status}"
                + (f" (diff={cr.diff})" if cr.diff is not None else "")
            )
            if cr.message:
                lines.append(f"  {cr.message}")
        return "\n".join(lines)

    return agent, deps
