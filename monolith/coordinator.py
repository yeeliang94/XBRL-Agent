"""Monolith coordinator — single PydanticAI agent fills all five face
statements onto one workbook.

Parallel to `coordinator.py` (the split-pipeline coordinator) but with a
materially different shape:

  * One workbook on disk (`monolith_filled.xlsx`), snapshotted after every
    successful `write_cells` batch so a crash mid-run preserves whatever
    landed (PRD §6a).
  * One `run_agents` row with `statement_type = "MONOLITH"` (gotcha #6
    explicitly carves this out — single-agent path, full per-turn rows).
  * One `monolith_conversation_trace.json` via `save_agent_trace`; failure
    path uses `save_messages_trace` (gotcha #6).
  * Structured outcomes for `iteration_exhausted` and `wallclock_exhausted`
    — never silent (gotcha #18-style invariant for this path).
  * Reuses the existing pipeline-stage SSE event family (gotcha #19) so the
    frontend's silent-gap labelling still works.
"""
from __future__ import annotations

import asyncio
import json
import logging
import shutil
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import openpyxl
# Hoisted to module scope so the `@agent.tool` decorators below can
# resolve `RunContext[_MonolithDeps]` annotations via module globals at
# tool-registration time. `from __future__ import annotations` turns
# type hints into strings; pydantic-ai's tool plumbing then re-evaluates
# them against the function's module globals — local imports inside
# `_build_agent` are invisible at that point.
from pydantic_ai import RunContext
from pydantic_ai.messages import (
    BinaryContent,
    FunctionToolCallEvent,
    FunctionToolResultEvent,
    PartDeltaEvent,
    TextPartDelta,
    ThinkingPartDelta,
)

from agent_tracing import save_agent_trace, save_messages_trace
from monolith.config import (
    MAX_AGENT_ITERATIONS_MONOLITH,
    MONOLITH_REQUEST_LIMIT,
    MONOLITH_TURN_TIMEOUT,
    MONOLITH_WALLCLOCK_SECONDS,
    MONOLITH_WALLCLOCK_WARNING_SECONDS,
)
from monolith.prompt_renderer import render as render_monolith_prompt
from monolith.tools import (
    MonolithToolContext,
    done as tool_done,
    get_state as tool_get_state,
    write_cells as tool_write_cells,
)
from statement_types import (
    StatementType,
    template_path as get_template_path,
)
from tools.pdf_viewer import count_pdf_pages, render_pages_to_png_bytes
from workbook_merger import _copy_sheet

logger = logging.getLogger(__name__)


# Monolith tool → pipeline phase. Mirror of coordinator.py:PHASE_MAP for
# the monolith's smaller tool surface (no separate read_template /
# fill_workbook / verify_totals — the monolith tools collapse those
# responsibilities). Emitted as a `status` SSE event so the run page's
# phase indicator tracks monolith activity the same way it tracks
# split activity.
_MONOLITH_PHASE_MAP = {
    "get_state": "reading_state",
    "view_pdf_pages": "viewing_pdf",
    "write_cells": "writing_cells",
    "done": "completing",
}


@dataclass
class MonolithRunConfig:
    pdf_path: str
    output_dir: str
    model: Any                              # str | provider model object
    statements: Set[StatementType] = field(
        default_factory=lambda: set(StatementType),
    )
    variants: Dict[StatementType, str] = field(default_factory=dict)
    filing_level: str = "company"
    filing_standard: str = "mfrs"
    page_hints: Optional[dict] = None       # scout output, optional
    run_id: Optional[int] = None
    db_path: Optional[str] = None


@dataclass
class MonolithResult:
    status: str                             # succeeded | failed | aborted
    workbook_path: Optional[str] = None
    error: Optional[str] = None
    failing_checks: List[str] = field(default_factory=list)
    accepted_residuals: List[dict] = field(default_factory=list)
    total_tokens: int = 0
    total_cost: float = 0.0
    turns: list = field(default_factory=list)
    prompt_tokens: int = 0
    completion_tokens: int = 0
    turn_count: int = 0
    tool_call_count: int = 0
    # The set of sheets that actually received writes — used for the
    # partial_merge SSE event when the run is cancelled mid-flight.
    statements_included: List[str] = field(default_factory=list)


class _IterationLimitReached(RuntimeError):
    """Mirror of coordinator._IterationLimitReached for the monolith cap."""


class _WallClockReached(RuntimeError):
    """The 15-min wall-clock cap fired."""


# ---------------------------------------------------------------------------
# Public entry
# ---------------------------------------------------------------------------


async def run_monolith(
    config: MonolithRunConfig,
    *,
    event_queue: Optional[asyncio.Queue] = None,
) -> MonolithResult:
    """Run the monolith agent end-to-end.

    Emits SSE events into `event_queue` when supplied. The caller is
    responsible for draining the queue and persisting `run_agents` /
    `run_agent_turns` rows (server-side wiring in
    `server.run_multi_agent_stream`).
    """
    output_dir = Path(config.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook_path = output_dir / "monolith_filled.xlsx"
    _materialise_workbook(workbook_path, config)

    pdf_page_count = _safe_pdf_page_count(config.pdf_path)
    # Canonical-mode plumbing: derive {sheet_name → template_id} once so
    # write_cells can route each landed cell into `run_concept_facts`
    # under the right template. Skipped cleanly when canonical mode is
    # off (run_id / db_path None) — projection then short-circuits.
    template_id_by_sheet = _build_template_id_by_sheet(config)
    ctx = MonolithToolContext(
        workbook_path=str(workbook_path),
        pdf_page_count=pdf_page_count,
        filing_standard=config.filing_standard,
        filing_level=config.filing_level,
        statements=list(config.statements),
        variants=config.variants,
        run_id=config.run_id,
        db_path=config.db_path,
        template_id_by_sheet=template_id_by_sheet,
    )

    async def _emit(event_type: str, data: dict) -> None:
        if event_queue is not None:
            await event_queue.put({
                "event": event_type,
                "data": {**data, "agent_id": "monolith", "agent_role": "MONOLITH"},
            })

    rendered = render_monolith_prompt(
        config.pdf_path,
        filing_standard=config.filing_standard,
        filing_level=config.filing_level,
        statements=list(config.statements),
        variants=config.variants,
        page_hints=config.page_hints or {},
    )

    agent, deps = _build_agent(
        model=config.model,
        rendered_prompt=rendered.full,
        ctx=ctx,
        pdf_path=config.pdf_path,
        pdf_page_count=pdf_page_count,
    )

    await _emit("status", {
        "phase": "starting",
        "message": "Starting monolith extraction (5 face statements).",
    })

    turn_records: list[dict] = []
    started = time.monotonic()
    warned_wallclock = False
    final_done_result: Optional[dict] = None
    # Per-run state for the live SSE emission path (peer-review MEDIUM
    # #3). _tool_start_times pairs a tool_call event's call_id with the
    # wall-clock at call so the matching tool_result event can report
    # duration_ms. _thinking_counter is the monotonically-increasing
    # id used to suffix thinking_id so the frontend can group
    # contiguous thinking deltas as one collapsible block.
    _tool_start_times: dict[str, float] = {}
    _thinking_counter = 0
    # Bound up-front so the except branches' best-effort trace helper
    # can reference it by name regardless of whether ``agent.iter(...)``
    # raised before the ``as`` binding.
    agent_run = None

    try:
        # Explicit UsageLimits so pydantic-ai's silent default of
        # request_limit=50 doesn't race the monolith iteration cap
        # (gotcha #18 — same incident the split path was hardened against).
        # MONOLITH_REQUEST_LIMIT is pinned strictly > MAX_AGENT_ITERATIONS_MONOLITH
        # in monolith/config.py and `tests/test_monolith_iteration_cap.py`.
        from pydantic_ai.usage import UsageLimits

        async with agent.iter(
            "Begin. Call get_state() and follow the workflow contract.",
            deps=deps,
            usage_limits=UsageLimits(request_limit=MONOLITH_REQUEST_LIMIT),
        ) as agent_run:
            iteration = 0
            prev_prompt = 0
            prev_completion = 0
            prev_total = 0
            async for node in _iter_with_turn_timeout(
                agent_run, MONOLITH_TURN_TIMEOUT,
            ):
                iteration += 1
                if iteration > MAX_AGENT_ITERATIONS_MONOLITH:
                    raise _IterationLimitReached(
                        f"Hit monolith iteration limit "
                        f"({MAX_AGENT_ITERATIONS_MONOLITH}). Agent stuck."
                    )

                # Wall-clock soft + hard cap.
                elapsed = time.monotonic() - started
                if (
                    not warned_wallclock
                    and elapsed >= MONOLITH_WALLCLOCK_WARNING_SECONDS
                ):
                    warned_wallclock = True
                    await _emit("pipeline_stage", {
                        "stage": "extracting",
                        "message": (
                            f"Monolith run has used "
                            f"{int(elapsed)}s of "
                            f"{int(MONOLITH_WALLCLOCK_SECONDS)}s wall-clock."
                        ),
                    })
                if elapsed >= MONOLITH_WALLCLOCK_SECONDS:
                    raise _WallClockReached(
                        f"Hit monolith wall-clock cap "
                        f"({int(MONOLITH_WALLCLOCK_SECONDS)}s)."
                    )

                # Streaming token telemetry per node (mirror of coordinator.py
                # — keeps Telemetry tab populated with per-turn rows).
                node_start = time.monotonic()
                node_kind = _node_kind(node)
                tool_names: list[str] = []
                if _is_call_tools_node(node):
                    async with node.stream(agent_run.ctx) as tool_stream:
                        # Bound the INNER event iteration too: the outer
                        # per-node timeout only covers node acquisition;
                        # a provider that stalls mid-stream on a single
                        # tool's output would hang the whole monolith
                        # run until pydantic-ai's UsageLimit fires.
                        # Mirror of coordinator.py:601-604 (split path).
                        async for event in _iter_with_turn_timeout(
                            tool_stream, MONOLITH_TURN_TIMEOUT,
                        ):
                            # Peer-review MEDIUM #3 (2026-05-28): emit
                            # live tool_call / tool_result SSE events
                            # so the Agents tab populates the same way
                            # it does for split agents. Mirror of
                            # coordinator.py:605-647.
                            if isinstance(event, FunctionToolCallEvent):
                                tool_name = event.part.tool_name
                                tool_names.append(tool_name)
                                phase = _MONOLITH_PHASE_MAP.get(tool_name)
                                if phase:
                                    await _emit("status", {
                                        "phase": phase,
                                        "message": (
                                            f"Monolith: "
                                            f"{phase.replace('_', ' ').title()}"
                                        ),
                                    })
                                raw_args = event.part.args
                                if isinstance(raw_args, str):
                                    try:
                                        parsed_args = json.loads(raw_args)
                                    except (json.JSONDecodeError, TypeError):
                                        parsed_args = {}
                                elif isinstance(raw_args, dict):
                                    parsed_args = raw_args
                                else:
                                    parsed_args = {}
                                await _emit("tool_call", {
                                    "tool_name": tool_name,
                                    "tool_call_id": event.part.tool_call_id,
                                    "args": parsed_args,
                                })
                                _tool_start_times[event.part.tool_call_id] = (
                                    time.monotonic()
                                )
                            elif isinstance(event, FunctionToolResultEvent):
                                content = event.result.content
                                summary = str(content)[:800] if content else ""
                                call_id = event.result.tool_call_id
                                start_t = _tool_start_times.pop(call_id, None)
                                duration_ms = (
                                    int((time.monotonic() - start_t) * 1000)
                                    if start_t else 0
                                )
                                await _emit("tool_result", {
                                    "tool_name": event.result.tool_name,
                                    "tool_call_id": call_id,
                                    "result_summary": summary,
                                    "duration_ms": duration_ms,
                                })
                elif _is_model_request_node(node):
                    # Stream text + thinking deltas (mirror of
                    # coordinator.py:649-687). Without this the
                    # Agents tab "Thinking" / "Response" panes stay
                    # empty for monolith runs, which is exactly when
                    # observability matters — a single long-running
                    # agent doing 30+ turns is harder to supervise
                    # than 5 short specialists in parallel.
                    _thinking_id = (
                        f"monolith_think_{_thinking_counter}"
                    )
                    _thinking_active = False
                    async with node.stream(agent_run.ctx) as model_stream:
                        async for event in _iter_with_turn_timeout(
                            model_stream, MONOLITH_TURN_TIMEOUT,
                        ):
                            if isinstance(event, PartDeltaEvent):
                                delta = event.delta
                                if isinstance(delta, TextPartDelta):
                                    if _thinking_active:
                                        await _emit("thinking_end", {
                                            "thinking_id": _thinking_id,
                                            "summary": "",
                                            "full_length": 0,
                                        })
                                        _thinking_active = False
                                        _thinking_counter += 1
                                        _thinking_id = (
                                            f"monolith_think_"
                                            f"{_thinking_counter}"
                                        )
                                    await _emit("text_delta", {
                                        "content": delta.content_delta,
                                    })
                                elif isinstance(delta, ThinkingPartDelta):
                                    _thinking_active = True
                                    await _emit("thinking_delta", {
                                        "content": delta.content_delta or "",
                                        "thinking_id": _thinking_id,
                                    })
                    if _thinking_active:
                        await _emit("thinking_end", {
                            "thinking_id": _thinking_id,
                            "summary": "",
                            "full_length": 0,
                        })
                        _thinking_counter += 1
                # Sweep the workbook to disk after every node — cheap and
                # the snapshot-after-write contract is what the cancel path
                # depends on.
                _snapshot_workbook(workbook_path)

                # Per-turn token deltas (best-effort).
                try:
                    u = agent_run.usage()
                    prompt_t = int(u.request_tokens or 0)
                    completion_t = int(u.response_tokens or 0)
                    total_t = int(u.total_tokens or 0)
                    turn_records.append({
                        "turn_index": iteration,
                        "node_kind": node_kind,
                        "tool_names": ",".join(tool_names) or None,
                        "_n_tool_calls": len(tool_names),
                        "prompt_tokens": max(prompt_t - prev_prompt, 0),
                        "completion_tokens": max(completion_t - prev_completion, 0),
                        "total_tokens": max(total_t - prev_total, 0),
                        "cumulative_tokens": total_t,
                        "cost_estimate": 0.0,
                        "duration_ms": int(
                            (time.monotonic() - node_start) * 1000,
                        ),
                    })
                    prev_prompt, prev_completion, prev_total = (
                        prompt_t, completion_t, total_t,
                    )
                except Exception:  # noqa: BLE001 — telemetry is advisory
                    logger.debug("per-turn telemetry skipped for monolith")

        result = agent_run.result
        save_agent_trace(
            result, str(output_dir), "monolith", turns=turn_records,
        )
        final_done_result = deps.final_done_result

        return _finalize_success(
            ctx=ctx,
            workbook_path=workbook_path,
            agent_run=agent_run,
            turn_records=turn_records,
            done_result=final_done_result,
        )

    except asyncio.CancelledError:
        # Partial-merge equivalent: the workbook on disk already holds
        # the last snapshot. Emit a `partial_merge` SSE event with the
        # full PartialMergeData shape (web/src/lib/types.ts:227) so the
        # frontend reducer can render the "Saved partial workbook"
        # banner consistently with the split path.
        #
        # Peer-review MEDIUM #4 (2026-05-28): the live xlsx may be
        # mid-`wb.save()` when CancelledError fires — openpyxl doesn't
        # write atomically, so the file on disk could be a truncated
        # zip. `_resolve_workbook_for_recovery` validates the live file
        # and falls back to the last-known-good `.snap` snapshot if
        # needed, so the download link the user gets always points at
        # a file Excel can open.
        recovered = _resolve_workbook_for_recovery(workbook_path)
        if recovered is not None:
            statements_included = _statements_with_writes(
                recovered, config.statements,
            )
        else:
            statements_included = []
        statements_missing = sorted(
            s.value for s in config.statements
            if s.value not in statements_included
        )
        await _emit("partial_merge", {
            "merged": recovered is not None,
            "merged_path": str(recovered) if recovered else str(workbook_path),
            "statements_included": statements_included,
            "statements_missing": statements_missing,
            # Monolith is face-only by design (PRD §3) — no notes path.
            "notes_included": [],
            "notes_missing": [],
            "error": (
                None
                if recovered is not None
                else "live workbook unreadable and no snapshot available"
            ),
        })
        # Save whatever messages we accumulated for debuggability.
        _best_effort_partial_trace(agent_run, output_dir, turn_records)
        raise

    except _IterationLimitReached as exc:
        await _emit("error", {
            "message": str(exc),
            "type": "iteration_exhausted",
        })
        _best_effort_partial_trace(agent_run, output_dir, turn_records)
        return MonolithResult(
            status="failed",
            workbook_path=str(workbook_path),
            error=str(exc),
            turns=turn_records,
            turn_count=len(turn_records),
            tool_call_count=sum(
                int(t.get("_n_tool_calls") or 0) for t in turn_records
            ),
            statements_included=_statements_with_writes(
                workbook_path, config.statements,
            ),
        )

    except _WallClockReached as exc:
        await _emit("error", {
            "message": str(exc),
            "type": "wallclock_exhausted",
        })
        _best_effort_partial_trace(agent_run, output_dir, turn_records)
        return MonolithResult(
            status="failed",
            workbook_path=str(workbook_path),
            error=str(exc),
            turns=turn_records,
            turn_count=len(turn_records),
            tool_call_count=sum(
                int(t.get("_n_tool_calls") or 0) for t in turn_records
            ),
            statements_included=_statements_with_writes(
                workbook_path, config.statements,
            ),
        )

    except asyncio.TimeoutError as exc:
        await _emit("error", {
            "message": f"Monolith stalled past {MONOLITH_TURN_TIMEOUT}s on a single turn.",
            "type": "turn_timeout",
        })
        _best_effort_partial_trace(agent_run, output_dir, turn_records)
        return MonolithResult(
            status="failed",
            workbook_path=str(workbook_path),
            error=str(exc),
            turns=turn_records,
            turn_count=len(turn_records),
            tool_call_count=sum(
                int(t.get("_n_tool_calls") or 0) for t in turn_records
            ),
            statements_included=_statements_with_writes(
                workbook_path, config.statements,
            ),
        )

    except Exception as exc:  # noqa: BLE001
        logger.exception("Monolith coordinator raised")
        await _emit("error", {
            "message": str(exc),
            "type": "monolith_exception",
        })
        _best_effort_partial_trace(agent_run, output_dir, turn_records)
        return MonolithResult(
            status="failed",
            workbook_path=str(workbook_path) if workbook_path.exists() else None,
            error=str(exc),
            turns=turn_records,
            turn_count=len(turn_records),
            tool_call_count=sum(
                int(t.get("_n_tool_calls") or 0) for t in turn_records
            ),
            statements_included=_statements_with_writes(
                workbook_path, config.statements,
            ),
        )


# ---------------------------------------------------------------------------
# Agent construction
# ---------------------------------------------------------------------------


class _MonolithDeps:
    """Carry the tool context + post-run capture slots through to tools."""

    def __init__(self, ctx: MonolithToolContext):
        self.ctx = ctx
        self.final_done_result: Optional[dict] = None


def _build_agent(
    *,
    model,
    rendered_prompt: str,
    ctx: MonolithToolContext,
    pdf_path: str,
    pdf_page_count: int,
):
    """Construct the PydanticAI agent + deps.

    Tools are bound as `tool_plain` (no RunContext arg) — we use a closure
    over `deps` instead. This keeps the agent surface tight and the tool
    signatures match what the model sees in the cached prompt.
    """
    from pydantic_ai import Agent
    from pydantic_ai.settings import ModelSettings

    deps = _MonolithDeps(ctx)
    # Pin temperature=1.0 — Gemini 3 through the LiteLLM proxy hard-requires
    # this (CLAUDE.md gotcha #5); other models tolerate it. Same value used
    # by extraction/agent.py and notes/agent.py so the monolith stays
    # apples-to-apples with the split path it's being compared against.
    agent: Agent = Agent(
        model=model,
        system_prompt=rendered_prompt,
        deps_type=_MonolithDeps,
        model_settings=ModelSettings(temperature=1.0),
    )

    @agent.tool
    async def get_state(rc: RunContext[_MonolithDeps]) -> dict:
        return tool_get_state(rc.deps.ctx)

    @agent.tool
    async def write_cells(
        rc: RunContext[_MonolithDeps], writes: list[dict],
    ) -> dict:
        return tool_write_cells(rc.deps.ctx, writes)

    @agent.tool
    async def done(
        rc: RunContext[_MonolithDeps],
        accept_imbalance: Optional[list[dict]] = None,
    ) -> dict:
        result = tool_done(rc.deps.ctx, accept_imbalance=accept_imbalance)
        # Capture the *latest* call; the agent may call done() multiple
        # times if the first attempt returned not_done.
        rc.deps.final_done_result = result
        return result

    @agent.tool
    async def view_pdf_pages(
        rc: RunContext[_MonolithDeps], start_page: int, end_page: int,
    ) -> list:
        """Render PDF pages as PNG vision payloads.

        Returns a list of (page header string + BinaryContent) entries,
        matching the existing extraction agent's `view_pdf_pages` shape
        (extraction/agent.py:459). The PydanticAI runtime binds
        `BinaryContent` into the next assistant turn's message envelope —
        a dict-only return loses the bytes and the agent gets no vision.
        """
        # Pre-validate the range — out-of-range pages would otherwise raise
        # inside the renderer and surface as an opaque tool error.
        start = max(1, int(start_page))
        end = min(pdf_page_count, int(end_page))
        if start > end or start < 1:
            return [
                f"page range [{start_page}, {end_page}] invalid for "
                f"PDF with {pdf_page_count} pages.",
            ]
        png_bytes_list = render_pages_to_png_bytes(
            pdf_path, start=start, end=end,
        )
        results: list = []
        for offset, png in enumerate(png_bytes_list):
            page_num = start + offset
            results.append(f"=== page {page_num} ===")
            results.append(BinaryContent(data=png, media_type="image/png"))
        return results

    return agent, deps


# ---------------------------------------------------------------------------
# Workbook handling
# ---------------------------------------------------------------------------


def _materialise_workbook(target: Path, config: MonolithRunConfig) -> None:
    """Concatenate per-statement templates into the live workbook.

    Builds a fresh workbook and copies every sheet from every face
    template into it via `workbook_merger._copy_sheet`, which preserves
    cell values, formulas, styles (fills, fonts, borders, alignment,
    number formats, protection), merged-cell ranges, and row/column
    dimensions.

    Critical gotcha #17 invariant: the abstract-row writer guard in
    `tools/section_headers.discover_section_headers` detects header rows
    by **dark-navy fill colour**. A value-only copy drops those fills
    and the guard silently no-ops on every non-first sheet — exactly
    the 2026-04-26 SOPL-Analysis-Function failure mode. So we route
    through `_copy_sheet`, not a hand-rolled value copy.

    If the target already exists (resumed run), leave it alone.
    """
    if target.exists():
        return
    statements = sorted(
        config.statements, key=lambda s: list(StatementType).index(s),
    )

    merged = openpyxl.Workbook()
    # Drop the default empty sheet so the output carries only real templates.
    if merged.sheetnames:
        merged.remove(merged.active)
    # Force Excel to recalc all formulas on open — same flag as
    # workbook_merger.merge() so monolith downloads behave identically
    # to split-pipeline ones.
    if merged.calculation is not None:
        merged.calculation.fullCalcOnLoad = True
        merged.calculation.forceFullCalc = True
        merged.calculation.calcOnSave = True

    sheets_copied = 0
    for stmt in statements:
        variant = config.variants.get(stmt) or _default_variant(
            stmt, config.filing_standard,
        )
        try:
            tpl_path = get_template_path(
                stmt, variant,
                level=config.filing_level,
                standard=config.filing_standard,
            )
        except ValueError:
            logger.warning(
                "Skipping %s/%s — no template for standard %s level %s",
                stmt.value, variant,
                config.filing_standard, config.filing_level,
            )
            continue
        wb_src = openpyxl.load_workbook(str(tpl_path), data_only=False)
        try:
            for src_ws in wb_src.worksheets:
                if src_ws.title in merged.sheetnames:
                    continue
                _copy_sheet(src_ws, merged)
                sheets_copied += 1
        finally:
            wb_src.close()

    if sheets_copied == 0:
        raise RuntimeError(
            f"_materialise_workbook copied no sheets — check template "
            f"paths for {config.filing_standard}/{config.filing_level}."
        )
    merged.save(str(target))
    merged.close()


def _build_template_id_by_sheet(config: MonolithRunConfig) -> dict[str, str]:
    """Compute `{sheet_name → template_id}` across all 5 face templates.

    Used by canonical projection in `MonolithToolContext` so each cell
    write resolves under the right template. Cheap — opens each template
    once at coordinator startup just to read sheet names + filename slug.

    Returns an empty dict when canonical mode is off (no run_id/db_path
    on config) — projection then short-circuits with no DB hit.
    """
    if config.run_id is None or not config.db_path:
        return {}
    try:
        from concept_model.parser import _derive_template_id
    except Exception:  # noqa: BLE001 — concept_model is optional
        logger.debug("concept_model not importable; canonical projection off")
        return {}

    out: dict[str, str] = {}
    for stmt in config.statements:
        variant = config.variants.get(stmt) or _default_variant(
            stmt, config.filing_standard,
        )
        try:
            tpl_path = get_template_path(
                stmt, variant,
                level=config.filing_level,
                standard=config.filing_standard,
            )
        except ValueError:
            continue
        try:
            tid = _derive_template_id(tpl_path)
        except Exception:  # noqa: BLE001
            logger.debug("template_id derivation failed for %s", tpl_path)
            continue
        # Read sheet names from the template file directly — cheap because
        # openpyxl loads metadata first.
        try:
            wb_src = openpyxl.load_workbook(str(tpl_path), data_only=False)
        except Exception:  # noqa: BLE001
            continue
        try:
            for sheet_name in wb_src.sheetnames:
                out[sheet_name] = tid
        finally:
            wb_src.close()
    return out


def _default_variant(stmt: StatementType, filing_standard: str) -> str:
    from statement_types import variants_for_standard

    for v in variants_for_standard(stmt, filing_standard):
        if v.template_filename and v.name != "NotPrepared":
            return v.name
    return "Default"


def _is_xlsx_openable(path: Path) -> bool:
    """Lightweight integrity probe for a workbook on disk.

    openpyxl write is NOT atomic — a crash mid-`wb.save()` can leave a
    truncated xlsx (incomplete zip central directory). This helper
    confirms the file is at least readable as a workbook, so callers
    can distinguish "complete" from "in-progress / corrupt".

    openpyxl gates ``load_workbook`` on the path's suffix (rejects
    anything that isn't .xlsx/.xlsm/.xltx/.xltm). Our `.xlsx.snap`
    sidecar trips that check even when the bytes are a perfectly
    valid xlsx, so we read the bytes and load through BytesIO — the
    same file is then valid input regardless of the on-disk name.
    """
    if not path.exists():
        return False
    try:
        import io

        data = path.read_bytes()
        wb = openpyxl.load_workbook(
            io.BytesIO(data), data_only=False, read_only=True,
        )
        wb.close()
        return True
    except Exception:  # noqa: BLE001 — any failure means "not openable"
        return False


def _snapshot_workbook(path: Path) -> None:
    """Capture a known-good copy of the live workbook to ``.xlsx.snap``.

    Peer-review MEDIUM #4 (2026-05-28): this previously dropped a
    ``.snap`` sidecar that nobody read, so the documented "if a later
    crash zaps the file mid-save, the previous snapshot stays intact"
    contract was a lie.

    The current implementation:
    - Validates the live workbook is openable BEFORE overwriting the
      snapshot. If the live file is mid-save / corrupt, we keep the
      last known-good ``.snap`` instead of propagating corruption.
    - Stages via a ``.snapshot`` temp path and atomically renames into
      ``.snap``. On POSIX `Path.replace` is atomic within a filesystem;
      on Windows it overwrites in one syscall.

    The recovery consumer is ``_resolve_workbook_for_recovery`` below
    (used by the cancel / partial-merge SSE path), which validates
    ``monolith_filled.xlsx`` and falls back to the snapshot if needed.
    """
    if not _is_xlsx_openable(path):
        # Live workbook can't be opened — most likely we caught it
        # mid-`wb.save()`. Don't overwrite the snapshot with corruption.
        return
    try:
        snap = path.with_suffix(".xlsx.snapshot")
        shutil.copy(str(path), str(snap))
        # POSIX rename is atomic on the same fs; on Windows it overwrites.
        snap.replace(path.with_suffix(".xlsx.snap"))
    except Exception:  # noqa: BLE001
        logger.debug("Workbook snapshot skipped for %s", path)


def _resolve_workbook_for_recovery(path: Path) -> Optional[Path]:
    """Pick the best available copy of ``path`` for crash recovery.

    Returns the live workbook if it's openable; otherwise the latest
    ``.snap`` if that's openable; otherwise None. Used by the cancel
    / partial-merge path so the download link the user is offered
    points at a file Excel can actually open.
    """
    if _is_xlsx_openable(path):
        return path
    snap_path = path.with_suffix(".xlsx.snap")
    if _is_xlsx_openable(snap_path):
        return snap_path
    return None


def _statements_with_writes(
    workbook_path: Path, statements: Set[StatementType],
) -> list[str]:
    """Best-effort: list which face statements have any non-formula values."""
    if not workbook_path.exists():
        return []
    out: list[str] = []
    try:
        wb = openpyxl.load_workbook(str(workbook_path), data_only=False)
    except Exception:  # noqa: BLE001
        return []
    try:
        from monolith.state import _SHEET_PREFIX  # type: ignore

        for stmt in statements:
            sheet = None
            for name in wb.sheetnames:
                for prefix in _SHEET_PREFIX.get(stmt, ()):
                    if name == prefix or name.startswith(prefix):
                        sheet = name
                        break
                if sheet:
                    break
            if not sheet:
                continue
            ws = wb[sheet]
            has_value = False
            for row in ws.iter_rows(min_row=3, max_row=min(ws.max_row, 200)):
                for cell in row:
                    if cell.column == 1:
                        continue
                    v = cell.value
                    if v is None:
                        continue
                    if isinstance(v, str) and v.startswith("="):
                        continue
                    if isinstance(v, str) and not v.strip():
                        continue
                    has_value = True
                    break
                if has_value:
                    break
            if has_value:
                out.append(stmt.value)
    finally:
        wb.close()
    return out


# ---------------------------------------------------------------------------
# Iter helpers (mirror coordinator.py)
# ---------------------------------------------------------------------------


async def _iter_with_turn_timeout(async_iterable, timeout: float):
    iterator = async_iterable.__aiter__()
    while True:
        try:
            node = await asyncio.wait_for(iterator.__anext__(), timeout=timeout)
        except StopAsyncIteration:
            return
        yield node


def _is_call_tools_node(node) -> bool:
    try:
        from pydantic_ai import Agent

        return Agent.is_call_tools_node(node)
    except Exception:  # noqa: BLE001
        return False


def _is_model_request_node(node) -> bool:
    try:
        from pydantic_ai import Agent

        return Agent.is_model_request_node(node)
    except Exception:  # noqa: BLE001
        return False


def _node_kind(node) -> Optional[str]:
    try:
        from pydantic_ai import Agent

        if Agent.is_call_tools_node(node):
            return "call_tools"
        if Agent.is_model_request_node(node):
            return "model_request"
    except Exception:  # noqa: BLE001
        return None
    return None


def _maybe_tool_name(event) -> Optional[str]:
    try:
        from pydantic_ai.messages import FunctionToolCallEvent

        if isinstance(event, FunctionToolCallEvent):
            return event.part.tool_name
    except Exception:  # noqa: BLE001
        return None
    return None


def _safe_pdf_page_count(path: str) -> int:
    if not path:
        return 0
    try:
        return count_pdf_pages(path)
    except Exception:  # noqa: BLE001
        return 0


def _best_effort_partial_trace(
    agent_run, output_dir: Path, turn_records: list[dict],
) -> None:
    if agent_run is None:
        return
    try:
        msgs = agent_run.ctx.state.message_history
    except Exception:  # noqa: BLE001
        msgs = None
    if not msgs:
        return
    save_messages_trace(msgs, str(output_dir), "monolith", turns=turn_records)


def _finalize_success(
    *,
    ctx: MonolithToolContext,
    workbook_path: Path,
    agent_run,
    turn_records: list[dict],
    done_result: Optional[dict],
) -> MonolithResult:
    """Compose the success-path MonolithResult."""
    failing: list[str] = []
    accepted: list[dict] = []
    error_msg: Optional[str] = None
    # Strict success contract: success ONLY when the agent called
    # `done({...})` AND the server returned status="done" (every failing
    # check either passed or was server-validated as accepted).
    # `not_done` means at least one accept_imbalance entry failed
    # validation OR a failing check wasn't named; either way the run
    # didn't finish honestly. A missing done_result means the agent
    # ended conversationally without calling `done` at all — also failed.
    # Mirror of extraction/agent.py's `result_saved` gate (gotcha that
    # carried over from the peer-review on the split path).
    if done_result is None:
        status = "failed"
        error_msg = (
            "Monolith agent finished without calling done() — extraction "
            "did not finalise. Whatever landed on disk is the snapshot at "
            "the last write."
        )
    elif done_result.get("status") == "done":
        status = "succeeded"
        failing = list(done_result.get("failing_checks") or [])
        accepted = list(done_result.get("accepted_residuals") or [])
    else:
        status = "failed"
        failing = list(done_result.get("failing_checks") or [])
        accepted = list(done_result.get("accepted_residuals") or [])
        msg = done_result.get("message") or "done() returned not_done."
        invalid = done_result.get("invalid_accepts")
        if invalid:
            msg += (
                f" {len(invalid)} accept_imbalance entries failed "
                "server-side validation."
            )
        error_msg = msg

    # Token rollup: derive prompt/completion from per-turn deltas and
    # set total = prompt + completion so the three rollup fields are
    # always internally consistent. Pydantic-ai's cumulative
    # `agent_run.usage().total_tokens` can include billing-only buckets
    # (cache reads, etc.) that don't appear in request/response splits;
    # using sum-of-deltas keeps Telemetry tab math addable.
    prompt_total = sum(int(t.get("prompt_tokens") or 0) for t in turn_records)
    completion_total = sum(
        int(t.get("completion_tokens") or 0) for t in turn_records
    )
    return MonolithResult(
        status=status,
        workbook_path=str(workbook_path),
        error=error_msg,
        failing_checks=failing,
        accepted_residuals=accepted,
        total_tokens=prompt_total + completion_total,
        turns=turn_records,
        prompt_tokens=prompt_total,
        completion_tokens=completion_total,
        turn_count=len(turn_records),
        tool_call_count=sum(int(t.get("_n_tool_calls") or 0) for t in turn_records),
        statements_included=_statements_with_writes(
            workbook_path, set(ctx.statements),
        ),
    )
