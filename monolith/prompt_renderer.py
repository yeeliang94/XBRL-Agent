"""Render the monolith agent's cached system prompt.

The output is one block: role + rules (from `prompts/monolith_face.md`)
+ template structure for each of the 5 sheets + page-marked PDF text.

The renderer enforces a hard byte ceiling (slice-0a derived) — if the
prompt would exceed it, PDF text is trimmed first, then the rendered
prompt is returned with a warning header so the operator can see the
cache pressure.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

import openpyxl

from monolith.config import MONOLITH_PROMPT_BYTE_CEILING
from statement_types import (
    StatementType,
    template_path as get_template_path,
)
from tools.template_reader import read_template

logger = logging.getLogger(__name__)


# Default ordering — matches statement_types canonical order.
_FACE_STATEMENTS: list[StatementType] = [
    StatementType.SOFP,
    StatementType.SOPL,
    StatementType.SOCI,
    StatementType.SOCF,
    StatementType.SOCIE,
]


# Cross-reference annotations — one per identity. The agent reads these
# alongside the template structure so it doesn't have to re-derive the
# relationships every turn (PRD §7).
_CROSS_REFERENCES: list[str] = [
    "SOFP!`*Total assets` ↔ SOFP!`*Total equity and liabilities`  (sofp_balance)",
    "SOPL!`Profit (loss)` ↔ SOCIE current-year profit row (Total col)  (sopl_to_socie_profit)",
    "SOCI!`Total comprehensive income` ↔ SOCIE TCI row (Total col)  (soci_to_socie_tci)",
    "SOCIE!`Equity at end of period` (Total col X) ↔ SOFP!`*Total equity`  (socie_to_sofp_equity)",
    "SOCF!`Cash and cash equivalents at end of period` ↔ SOFP!`Cash and cash equivalents`  (socf_to_sofp_cash)",
]


@dataclass
class RenderedPrompt:
    """Output of `render()` — pre-split so callers can introspect."""
    rules_block: str               # role + rules (from monolith_face.md)
    template_structure: str        # the per-sheet indexed row list
    pdf_text: str                  # page-marked PDF text (trimmed to fit)
    full: str                      # concatenated cached chunk
    byte_size: int
    trimmed: bool


# ---------------------------------------------------------------------------


def render(
    pdf_path: str,
    filing_standard: str = "mfrs",
    filing_level: str = "company",
    *,
    statements: Optional[list[StatementType]] = None,
    variants: Optional[dict[StatementType, str]] = None,
    byte_ceiling: int = MONOLITH_PROMPT_BYTE_CEILING,
    page_hints: Optional[dict[StatementType, dict]] = None,
) -> RenderedPrompt:
    """Render the full cached prompt chunk.

    `variants` picks which variant to load per statement (SOFP CuNonCu vs
    OrderOfLiquidity, etc.). Default uses the first registry variant per
    statement.

    `page_hints` is the scout's `{statement: {face_page, note_pages}}`
    output — used to bias which PDF pages get inlined into the cached
    PDF text block. If absent, the renderer falls back to the full PDF
    text and trims uniformly.
    """
    statements = statements or list(_FACE_STATEMENTS)
    variants = variants or {}

    rules_block = _load_rules_block()
    template_structure = _render_template_structure(
        statements,
        variants=variants,
        filing_standard=filing_standard,
        filing_level=filing_level,
    )
    pdf_text_full = _extract_pdf_text(pdf_path)

    # First-pass assemble; if too large, trim PDF text.
    overhead = _byte_length(rules_block + template_structure) + _byte_length(
        "\n\n=== PDF TEXT (cached) ===\n"
    )
    pdf_budget = max(byte_ceiling - overhead, 4096)
    pdf_trimmed, trimmed = _maybe_trim_pdf_text(
        pdf_text_full, pdf_budget, page_hints=page_hints,
    )

    full = _assemble(rules_block, template_structure, pdf_trimmed, trimmed=trimmed)
    byte_size = _byte_length(full)
    # If the prompt is still over the ceiling — usually because the
    # template-structure block alone exceeds it — flag `trimmed=True` so
    # the operator knows the cache budget is exceeded. The PDF text was
    # already squeezed; further trimming would mean cutting structure,
    # which we won't do silently.
    over_budget = byte_size > byte_ceiling
    return RenderedPrompt(
        rules_block=rules_block,
        template_structure=template_structure,
        pdf_text=pdf_trimmed,
        full=full,
        byte_size=byte_size,
        trimmed=trimmed or over_budget,
    )


# ---------------------------------------------------------------------------
# Rules block (monolith_face.md)
# ---------------------------------------------------------------------------


_PROMPT_PATH = Path(__file__).resolve().parent.parent / "prompts" / "monolith_face.md"


def _load_rules_block() -> str:
    """Read prompts/monolith_face.md verbatim.

    The test `tests/test_monolith_prompt.py` grep-asserts the load-bearing
    invariants are present in this string — keep them in the .md file
    rather than templating them here so a single source of truth survives.
    """
    return _PROMPT_PATH.read_text(encoding="utf-8")


# ---------------------------------------------------------------------------
# Template structure
# ---------------------------------------------------------------------------


def _render_template_structure(
    statements: Iterable[StatementType],
    *,
    variants: dict[StatementType, str],
    filing_standard: str,
    filing_level: str,
) -> str:
    """Per-sheet indexed row list + cross-reference annotations."""
    lines: list[str] = [
        "## Template structure (cached)",
        "",
        "Per-sheet writable-row index. `kind` mirrors what `get_state()`",
        "returns. Abstract rows are read-only; formula rows compute their",
        "values from the leaves around them.",
        "",
        "### Cross-statement identities",
        "",
    ]
    for ref in _CROSS_REFERENCES:
        lines.append(f"- {ref}")
    lines.append("")

    for stmt in statements:
        variant = variants.get(stmt) or _default_variant(stmt, filing_standard)
        try:
            tpl_path = get_template_path(
                stmt, variant, level=filing_level, standard=filing_standard,
            )
        except ValueError as exc:
            lines.append(f"### {stmt.value} ({variant}) — UNAVAILABLE: {exc}")
            lines.append("")
            continue
        lines.append(f"### {stmt.value} — variant {variant}")
        lines.append("")
        lines.append("```")
        lines.extend(_describe_sheet_rows(tpl_path))
        lines.append("```")
        lines.append("")
    return "\n".join(lines)


def _default_variant(stmt: StatementType, filing_standard: str) -> str:
    """Pick a default variant for prompt rendering when none is supplied.

    The runtime coordinator does its own (scout-driven) variant resolution;
    this is only for the cached structure block. Returns the first
    registered variant that applies to the requested standard and carries
    a template file.
    """
    from statement_types import variants_for_standard

    for v in variants_for_standard(stmt, filing_standard):
        if v.template_filename and v.name != "NotPrepared":
            return v.name
    # Should not happen on MFRS Company, but stay graceful.
    return "Default"


def _describe_sheet_rows(template_path: Path) -> list[str]:
    """One line per labelled row: `row | kind | label`.

    Reuses `tools.template_reader.read_template` so the kind classification
    (abstract / data-entry / formula) matches what the writer accepts.
    """
    rows_out: list[str] = []
    fields = read_template(str(template_path))
    # Group by sheet so output reads naturally on multi-sheet templates.
    by_sheet: dict[str, dict[int, dict]] = {}
    for f in fields:
        if f.col != 1:
            continue
        if f.value is None:
            continue
        label = str(f.value).strip()
        if not label:
            continue
        by_sheet.setdefault(f.sheet, {})[f.row] = {
            "label": label,
            "is_abstract": f.is_abstract,
        }
    # Detect formula rows by looking at col B/C (or matrix cols for SOCIE).
    wb = openpyxl.load_workbook(str(template_path), data_only=False)
    try:
        for sheet_name, rows in by_sheet.items():
            ws = wb[sheet_name]
            rows_out.append(f"[sheet: {sheet_name}]")
            for row_idx in sorted(rows):
                meta = rows[row_idx]
                kind = _row_kind(ws, row_idx, is_abstract=meta["is_abstract"])
                rows_out.append(f"  {row_idx:>4} | {kind:<14} | {meta['label']}")
    finally:
        wb.close()
    return rows_out


def _row_kind(ws, row: int, *, is_abstract: bool) -> str:
    if is_abstract:
        return "abstract"
    # Scan a handful of columns; if any holds a formula, this is a formula row.
    for col in range(2, min(ws.max_column + 1, 25)):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, str) and v.startswith("="):
            return "formula"
    return "leaf"


# ---------------------------------------------------------------------------
# PDF text
# ---------------------------------------------------------------------------


def _extract_pdf_text(pdf_path: str) -> str:
    """Page-marked PyMuPDF extract of the PDF.

    Format:
        === page 1 ===
        <text of page 1>

        === page 2 ===
        ...
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        logger.warning("PyMuPDF not available — PDF text will be empty.")
        return ""

    out: list[str] = []
    doc = fitz.open(pdf_path)
    try:
        for i in range(len(doc)):
            page = doc[i]
            text = page.get_text("text") or ""
            out.append(f"=== page {i + 1} ===\n{text.rstrip()}")
    finally:
        doc.close()
    return "\n\n".join(out)


def _maybe_trim_pdf_text(
    text: str,
    budget_bytes: int,
    *,
    page_hints: Optional[dict[StatementType, dict]] = None,
) -> tuple[str, bool]:
    """Trim PDF text to fit within `budget_bytes`.

    Strategy:
      - If scout hints exist, keep only the union of `face_page` +
        `note_pages` ranges (extended by ±2 pages for context) and drop
        the rest.
      - Else, keep pages front-to-back until we hit the budget.
    The trim is byte-based on the UTF-8 encoding so a 200 KB ceiling
    really means 200 KB on the wire.
    """
    if _byte_length(text) <= budget_bytes:
        return text, False

    pages = _split_pages(text)
    if page_hints:
        keep = _pages_from_hints(page_hints, total_pages=len(pages))
        kept_text = _join_pages(
            [p for i, p in enumerate(pages, start=1) if i in keep]
        )
        if _byte_length(kept_text) <= budget_bytes:
            return kept_text, True
        # Hints still over budget — fall through to the front-pack path
        # using the hinted subset.
        pages = [p for i, p in enumerate(pages, start=1) if i in keep]

    # Front-pack: greedily pack pages until budget exhausted.
    out: list[str] = []
    used = 0
    for p in pages:
        chunk_size = _byte_length(p) + 2  # +2 for the `\n\n` separator
        if used + chunk_size > budget_bytes:
            break
        out.append(p)
        used += chunk_size
    return _join_pages(out), True


def _split_pages(text: str) -> list[str]:
    """Round-trip safe split: returns one entry per `=== page N ===` block."""
    if not text:
        return []
    parts = text.split("\n\n=== page ")
    if not parts[0].startswith("=== page "):
        # `text` started with the marker; the split swallowed it.
        if len(parts) > 1:
            parts = [parts[0]] + ["=== page " + p for p in parts[1:]]
    else:
        parts = ["=== page " + p for p in parts]
    return [p for p in parts if p.strip()]


def _join_pages(pages: list[str]) -> str:
    return "\n\n".join(pages)


def _pages_from_hints(
    page_hints: dict, total_pages: int, expand: int = 2,
) -> set[int]:
    keep: set[int] = set()
    for entry in page_hints.values():
        if not isinstance(entry, dict):
            continue
        face = entry.get("face_page")
        if isinstance(face, int) and 1 <= face <= total_pages:
            for p in range(
                max(1, face - expand), min(total_pages, face + expand) + 1,
            ):
                keep.add(p)
        for note in entry.get("note_pages") or []:
            if isinstance(note, int) and 1 <= note <= total_pages:
                for p in range(
                    max(1, note - expand), min(total_pages, note + expand) + 1,
                ):
                    keep.add(p)
    return keep


# ---------------------------------------------------------------------------
# Assembly helpers
# ---------------------------------------------------------------------------


def _byte_length(s: str) -> int:
    return len(s.encode("utf-8"))


def _assemble(
    rules: str, structure: str, pdf_text: str, *, trimmed: bool,
) -> str:
    parts: list[str] = [rules.strip(), "", structure.strip(), ""]
    parts.append("## PDF TEXT (cached)")
    if trimmed:
        parts.append("")
        parts.append(
            "_NOTE: PDF text trimmed to fit the cached-prefix byte ceiling._"
            " Use `view_pdf_pages` for any page not below."
        )
    parts.append("")
    parts.append(pdf_text.strip())
    return "\n".join(parts).strip() + "\n"
