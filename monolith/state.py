"""StateSnapshot — the agent's per-turn dashboard.

Composes existing primitives:

  - `tools.template_reader.read_template` for row shape + abstract-row flags
    (gotcha #17 — those rows must never be writable).
  - `tools.verifier.verify_statement` for per-sheet balance identities.
  - `tools.verifier._resolve_cell_value` to evaluate formula cells so the
    agent sees both the formula text and its computed value (PRD §5).
  - `cross_checks.framework.run_all` for cross-statement identities, with
    direction strings added on top of the framework's raw `expected/actual`.

The snapshot is recomputed from disk every call — no cached state — so a
crash mid-write leaves no stale view behind.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl

from statement_types import StatementType, template_path as get_template_path
from tools.template_reader import read_template
from tools.verifier import _resolve_cell_value, verify_statement

logger = logging.getLogger(__name__)

# Five face-statement sheet roots. Each is a hint: the actual sheet name
# in a filled workbook depends on the chosen variant (e.g. SOPL-Function
# vs SOPL-Nature). We inspect the workbook's sheet list to pick the
# right one per sheet root.
_FACE_STATEMENTS: list[StatementType] = [
    StatementType.SOFP,
    StatementType.SOPL,
    StatementType.SOCI,
    StatementType.SOCF,
    StatementType.SOCIE,
]

# Sheet-name prefix per statement type, for resolving the live sheet
# inside a multi-sheet workbook. SOCIE has just the bare prefix.
_SHEET_PREFIX: dict[StatementType, tuple[str, ...]] = {
    StatementType.SOFP: ("SOFP-",),
    StatementType.SOPL: ("SOPL-",),
    StatementType.SOCI: ("SOCI-",),
    StatementType.SOCF: ("SOCF-",),
    StatementType.SOCIE: ("SOCIE",),  # exact match
}

# Column mapping for non-matrix sheets. Mirrors `FieldMapping.col` in
# tools/fill_workbook.py (gotcha #17 leans on these).
_COL_CY = 2
_COL_PY = 3
_COL_EVIDENCE_COMPANY = 4
_COL_EVIDENCE_GROUP = 6

# SOCIE matrix layout — components live in cols B..W (2..23), Total at
# X (24). PY blocks live in separate row ranges, not a PY column.
_SOCIE_TOTAL_COL_MFRS = 24


@dataclass
class CellWriteResult:
    """Carried inside `RowSnapshot.cy`/`py` when the cell is a formula.

    Mirrors PRD §5: `{formula, computed, warnings}`. Plain numeric cells
    surface as bare floats (not this struct).
    """
    formula: str
    computed: Optional[float]
    warnings: list[str] = field(default_factory=list)


@dataclass
class RowSnapshot:
    row: int
    concept: Optional[str]          # XBRL concept id, when known (placeholder for v1)
    label: str
    kind: str                       # 'leaf' | 'formula' | 'abstract' | 'matrix_leaf'
    cy: Any = None                  # float | CellWriteResult | None
    py: Any = None
    evidence: Optional[str] = None
    # SOCIE matrix only. `matrix_cols` keyed by equity-component label
    # (resolved from `concept_nodes.matrix_col_label` if available, else
    # the row-2 header text). Values are floats or None.
    matrix_cols: Optional[dict[str, Any]] = None


@dataclass
class SheetSnapshot:
    sheet: str                      # the actual openpyxl sheet name
    statement: str                  # StatementType.value
    filled: int = 0
    writable: int = 0
    rows: list[RowSnapshot] = field(default_factory=list)


@dataclass
class CheckSnapshot:
    """Common shape for both verifier rows and cross-check results."""
    id: str
    pass_: bool                     # `pass` is a reserved word; serialised as 'pass'
    lhs: Optional[float] = None
    rhs: Optional[float] = None
    diff: Optional[float] = None
    direction: Optional[str] = None
    lhs_ref: Optional[str] = None
    rhs_ref: Optional[str] = None
    message: str = ""
    sheet: Optional[str] = None
    target_row: Optional[int] = None


@dataclass
class StateSnapshot:
    """The thing `get_state()` returns.

    Designed to serialise to JSON cleanly via `to_dict()` — pydantic-ai
    tools render dict results directly into the assistant message.
    """
    filing_standard: str
    filing_level: str
    turn: int
    sheets: dict[str, SheetSnapshot] = field(default_factory=dict)
    verifier: list[CheckSnapshot] = field(default_factory=list)
    cross_checks: list[CheckSnapshot] = field(default_factory=list)
    history_hints: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "filing": {
                "standard": self.filing_standard,
                "level": self.filing_level,
            },
            "turn": self.turn,
            "sheets": {
                name: {
                    "sheet": s.sheet,
                    "statement": s.statement,
                    "filled": s.filled,
                    "writable": s.writable,
                    "rows": [_row_to_dict(r) for r in s.rows],
                }
                for name, s in self.sheets.items()
            },
            "verifier": [_check_to_dict(c) for c in self.verifier],
            "cross_checks": [_check_to_dict(c) for c in self.cross_checks],
            "history_hints": list(self.history_hints),
        }


def _row_to_dict(r: RowSnapshot) -> dict:
    out: dict = {
        "row": r.row,
        "label": r.label,
        "kind": r.kind,
    }
    if r.concept is not None:
        out["concept"] = r.concept
    if r.kind == "matrix_leaf":
        out["matrix_cols"] = r.matrix_cols or {}
        if r.evidence is not None:
            out["evidence"] = r.evidence
    elif r.kind == "abstract":
        pass  # abstract rows carry no values by contract
    else:
        out["cy"] = _value_to_dict(r.cy)
        out["py"] = _value_to_dict(r.py)
        if r.evidence is not None:
            out["evidence"] = r.evidence
    return out


def _value_to_dict(v: Any) -> Any:
    if isinstance(v, CellWriteResult):
        return {
            "formula": v.formula,
            "computed": v.computed,
            "warnings": list(v.warnings),
        }
    return v


def _check_to_dict(c: CheckSnapshot) -> dict:
    out: dict = {"id": c.id, "pass": c.pass_}
    if c.lhs is not None:
        out["lhs"] = c.lhs
    if c.rhs is not None:
        out["rhs"] = c.rhs
    if c.diff is not None:
        out["diff"] = c.diff
    if c.direction:
        out["direction"] = c.direction
    if c.lhs_ref:
        out["lhs_ref"] = c.lhs_ref
    if c.rhs_ref:
        out["rhs_ref"] = c.rhs_ref
    if c.message:
        out["message"] = c.message
    if c.sheet:
        out["sheet"] = c.sheet
    if c.target_row:
        out["target_row"] = c.target_row
    return out


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


def build_state_snapshot(
    workbook_path: str,
    *,
    filing_standard: str = "mfrs",
    filing_level: str = "company",
    statements: Optional[list[StatementType]] = None,
    variants: Optional[dict[StatementType, str]] = None,
    turn: int = 0,
    history_hints: Optional[list[dict]] = None,
) -> StateSnapshot:
    """Compose a StateSnapshot from the live workbook at `workbook_path`.

    `variants` maps each statement type to its variant name (e.g.
    {SOFP: "CuNonCu"}). It only affects cross-check dispatch + the
    verifier; sheet detection works off the workbook's actual sheet names.

    `history_hints` is injected by the coordinator (see coordinator step 4
    in the plan). State.py is dumb about history.
    """
    statements = statements or list(_FACE_STATEMENTS)
    variants = variants or {}

    wb_path = Path(workbook_path)
    if not wb_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(str(wb_path), data_only=False)

    # Read the template structure ONCE for every sheet (one load), then
    # bucket by sheet so per-statement snapshots reuse the result instead
    # of reloading the whole workbook five times via `read_template`.
    all_fields = read_template(str(wb_path))
    fields_by_sheet: dict[str, list] = {}
    for f in all_fields:
        fields_by_sheet.setdefault(f.sheet, []).append(f)

    sheets: dict[str, SheetSnapshot] = {}
    for stmt in statements:
        sheet_name = _resolve_sheet_name(wb, stmt)
        if sheet_name is None:
            # The variant for this statement isn't in the workbook (e.g.
            # SOCI/NotPrepared on an MFRS run that omitted it). Skip
            # silently — the agent's prompt will note the missing sheet.
            continue
        sheets[stmt.value] = _build_sheet_snapshot(
            wb,
            sheet_name=sheet_name,
            statement=stmt,
            filing_standard=filing_standard,
            filing_level=filing_level,
            sheet_fields=fields_by_sheet.get(sheet_name, []),
        )

    verifier_results: list[CheckSnapshot] = []
    for stmt in statements:
        if stmt.value not in sheets:
            continue
        variant = variants.get(stmt) or ""
        verifier_results.extend(
            _run_verifier_snapshot(
                workbook_path=str(wb_path),
                statement=stmt,
                variant=variant,
                filing_level=filing_level,
                filing_standard=filing_standard,
            )
        )

    cross_check_results = _run_cross_checks_snapshot(
        workbook_path=str(wb_path),
        statements=statements,
        variants=variants,
        filing_level=filing_level,
        filing_standard=filing_standard,
    )

    wb.close()

    return StateSnapshot(
        filing_standard=filing_standard,
        filing_level=filing_level,
        turn=turn,
        sheets=sheets,
        verifier=verifier_results,
        cross_checks=cross_check_results,
        history_hints=list(history_hints or []),
    )


# ---------------------------------------------------------------------------
# Sheet-snapshot helpers
# ---------------------------------------------------------------------------


def _resolve_sheet_name(wb, stmt: StatementType) -> Optional[str]:
    """Find the live sheet name in the workbook for a given statement.

    SOFP/SOPL/SOCI/SOCF have variant-suffixed names ("SOFP-CuNonCu"); SOCIE
    is just "SOCIE". Returns None when the workbook doesn't carry this
    statement (e.g. SOCI/NotPrepared).
    """
    prefixes = _SHEET_PREFIX.get(stmt, ())
    for name in wb.sheetnames:
        for p in prefixes:
            if name == p or name.startswith(p):
                return name
    return None


def _build_sheet_snapshot(
    wb,
    *,
    sheet_name: str,
    statement: StatementType,
    filing_standard: str,
    filing_level: str,
    sheet_fields: list,
) -> SheetSnapshot:
    # `sheet_fields` is the per-sheet slice of the single workbook-wide
    # `read_template` walk done by the caller. Avoids reloading the
    # full xlsx once per statement (peer-review I4).
    rows_by_index: dict[int, RowSnapshot] = {}
    abstract_rows: set[int] = set()

    for f in sheet_fields:
        if f.col == 1:
            label = (f.value or "").strip()
            if not label:
                continue
            kind = "abstract" if f.is_abstract else "leaf"
            if f.is_abstract:
                abstract_rows.add(f.row)
            rows_by_index[f.row] = RowSnapshot(
                row=f.row,
                concept=None,
                label=label,
                kind=kind,
            )

    ws = wb[sheet_name]
    is_socie = statement is StatementType.SOCIE and filing_standard == "mfrs"

    socie_headers = _read_socie_matrix_headers(ws) if is_socie else {}

    filled = 0
    writable = 0
    for row in sorted(rows_by_index):
        snap = rows_by_index[row]
        if snap.kind == "abstract":
            continue
        if is_socie:
            snap.kind = "matrix_leaf"
            snap.matrix_cols = _read_socie_matrix_cells(
                wb, ws, row=row, headers=socie_headers,
            )
            snap.evidence = _read_evidence_cell(ws, row=row, socie=True)
            # Writable column count: every matrix component column counts
            # as a separate writable cell.
            writable += len(snap.matrix_cols)
            filled += sum(
                1 for v in snap.matrix_cols.values() if _is_filled(v)
            )
        else:
            snap.cy = _read_value_or_formula(
                wb, ws, row=row, col=_COL_CY,
            )
            snap.py = _read_value_or_formula(
                wb, ws, row=row, col=_COL_PY,
            )
            snap.evidence = _read_evidence_cell(
                ws, row=row, socie=False, filing_level=filing_level,
            )
            # Mark formula rows. A formula on either CY or PY is enough —
            # the writer refuses to overwrite formulas (gotcha #3).
            if (
                isinstance(snap.cy, CellWriteResult)
                or isinstance(snap.py, CellWriteResult)
            ):
                snap.kind = "formula"
            writable += 2  # CY + PY
            filled += int(_is_filled(snap.cy)) + int(_is_filled(snap.py))

    return SheetSnapshot(
        sheet=sheet_name,
        statement=statement.value,
        filled=filled,
        writable=writable,
        rows=sorted(rows_by_index.values(), key=lambda r: r.row),
    )


def _read_value_or_formula(wb, ws, *, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    raw = cell.value
    if raw is None:
        return None
    if isinstance(raw, str) and raw.startswith("="):
        warnings: list[str] = []
        computed = _resolve_cell_value(
            wb, ws.title, cell.coordinate, warnings=warnings,
        )
        return CellWriteResult(
            formula=raw, computed=computed, warnings=warnings,
        )
    if isinstance(raw, (int, float)):
        return float(raw)
    # Strings (notes, e.g. "Note 14") — surface verbatim. The agent uses
    # these as evidence pointers; the writer overwrites them via `col`
    # if the agent re-targets the cell.
    return raw


def _read_evidence_cell(
    ws, *, row: int, socie: bool, filing_level: str = "company",
) -> Optional[str]:
    if socie:
        col = _resolve_socie_evidence_col(ws)
    else:
        col = _COL_EVIDENCE_GROUP if filing_level == "group" else _COL_EVIDENCE_COMPANY
    raw = ws.cell(row=row, column=col).value
    if raw is None:
        return None
    return str(raw)


def _resolve_socie_evidence_col(ws) -> int:
    # Mirror of tools.fill_workbook._resolve_socie_evidence_col. Centralised
    # inline because importing the writer here would create a cycle for
    # tests that import state.py without the writer's heavier dependencies.
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if isinstance(v, str) and v.strip().lower() == "source":
            return col
    return 25  # MFRS matrix fall-back (col Y)


def _is_filled(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, CellWriteResult):
        return v.computed is not None
    if isinstance(v, str) and not v.strip():
        return False
    return True


def _read_socie_matrix_headers(ws) -> dict[int, str]:
    """Read row-2 component labels off an MFRS SOCIE sheet.

    Returns {col_index: label}. Column 1 (col A) is the row label
    column and is excluded. Empty headers are skipped.
    """
    headers: dict[int, str] = {}
    for col in range(2, _SOCIE_TOTAL_COL_MFRS + 1):
        v = ws.cell(row=2, column=col).value
        if v is None:
            continue
        text = str(v).strip()
        if not text:
            continue
        headers[col] = text
    return headers


def _read_socie_matrix_cells(
    wb, ws, *, row: int, headers: dict[int, str],
) -> dict[str, Any]:
    """Read every component-column value on a SOCIE row.

    Returns a label→value map (None for empty cells). Formula cells get
    the same `{formula, computed, warnings}` treatment as linear sheets.
    """
    out: dict[str, Any] = {}
    for col, label in headers.items():
        out[label] = _read_value_or_formula(wb, ws, row=row, col=col)
    return out


# ---------------------------------------------------------------------------
# Verifier snapshot
# ---------------------------------------------------------------------------


def _run_verifier_snapshot(
    *,
    workbook_path: str,
    statement: StatementType,
    variant: str,
    filing_level: str,
    filing_standard: str,
) -> list[CheckSnapshot]:
    """Turn one statement's verifier result into per-check snapshots."""
    try:
        result = verify_statement(
            workbook_path,
            statement_type=statement,
            variant=variant,
            filing_level=filing_level,
            filing_standard=filing_standard,
        )
    except Exception as exc:  # noqa: BLE001
        logger.debug(
            "verify_statement(%s/%s) raised in snapshot: %s",
            statement.value, variant, exc,
        )
        return [CheckSnapshot(
            id=f"verifier:{statement.value.lower()}",
            pass_=False,
            message=f"Verifier raised: {exc}",
        )]

    # `is_balanced` is None when the check is N/A (e.g. SOPL without
    # attribution). Surface those as `pass: true` so they don't show as
    # failures the agent has to act on.
    if result.is_balanced is None and not result.mismatches:
        return []

    pass_ = bool(result.is_balanced) and not result.mismatches
    direction: Optional[str] = None
    diff: Optional[float] = None
    lhs: Optional[float] = None
    rhs: Optional[float] = None

    # Try to extract the headline diff direction for SOFP (where we have
    # signed lhs/rhs available). Other statements get the mismatches text
    # without a quantitative direction.
    if statement is StatementType.SOFP:
        ass = result.computed_totals.get("total_assets_cy")
        eql = result.computed_totals.get("total_equity_liabilities_cy")
        if ass is not None and eql is not None:
            lhs = ass
            rhs = eql
            diff = round(ass - eql, 4)
            if abs(diff) > 0.01:
                direction = (
                    "assets > equity+liab" if diff > 0
                    else "assets < equity+liab"
                )

    return [CheckSnapshot(
        id=f"verifier:{statement.value.lower()}",
        pass_=pass_,
        lhs=lhs,
        rhs=rhs,
        diff=diff,
        direction=direction,
        message=("\n".join(result.mismatches) if result.mismatches else result.feedback),
        sheet=_SHEET_PREFIX[statement][0].rstrip("-") if statement in _SHEET_PREFIX else None,
    )]


# ---------------------------------------------------------------------------
# Cross-check snapshot
# ---------------------------------------------------------------------------


def _run_cross_checks_snapshot(
    *,
    workbook_path: str,
    statements: list[StatementType],
    variants: dict[StatementType, str],
    filing_level: str,
    filing_standard: str,
) -> list[CheckSnapshot]:
    """Run the existing cross-check framework against the single workbook.

    The split pipeline normally calls cross_checks against per-statement
    files (one path per StatementType). The monolith has one workbook for
    all five sheets, so we point every required path at the same file.
    Cross-check implementations resolve the sheet by name internally, so
    pointing five paths at the same workbook works as-is.
    """
    from cross_checks.framework import (
        build_default_cross_checks,
        run_all,
    )

    checks = build_default_cross_checks()
    workbook_paths = {stmt: workbook_path for stmt in statements}
    run_config = {
        "statements_to_run": set(statements),
        "variants": dict(variants),
        "filing_level": filing_level,
        "filing_standard": filing_standard,
    }

    results = run_all(checks, workbook_paths, run_config)
    out: list[CheckSnapshot] = []
    for r in results:
        # Only surface checks the snapshot can act on. Pending /
        # not_applicable are no-ops for the agent.
        if r.status in ("pending", "not_applicable"):
            continue
        pass_ = r.status == "passed"
        diff: Optional[float] = None
        direction: Optional[str] = None
        if r.expected is not None and r.actual is not None:
            diff = round(r.expected - r.actual, 4)
            if not pass_ and abs(diff) > 0.01:
                # Direction string per PRD §5 — "SOFP higher by 45".
                lhs_label, rhs_label = _check_id_endpoints(r.name)
                if diff > 0:
                    direction = f"{lhs_label} higher by {abs(diff):.2f}"
                else:
                    direction = f"{rhs_label} higher by {abs(diff):.2f}"
        out.append(CheckSnapshot(
            id=r.name,
            pass_=pass_,
            lhs=r.expected,
            rhs=r.actual,
            diff=diff,
            direction=direction,
            message=r.message,
            sheet=r.target_sheet,
            target_row=r.target_row,
        ))
    return out


# Minimal map from cross-check `name` to (lhs_label, rhs_label) for
# direction strings. Falls back to "LHS"/"RHS" if a name isn't known —
# the diff magnitude still tells the agent which side is high.
_CHECK_ENDPOINTS: dict[str, tuple[str, str]] = {
    "sofp_balance": ("SOFP assets", "SOFP equity+liab"),
    "sopl_to_socie_profit": ("SOPL profit", "SOCIE profit"),
    "soci_to_socie_tci": ("SOCI TCI", "SOCIE TCI"),
    "socie_to_sofp_equity": ("SOCIE equity", "SOFP equity"),
    "socf_to_sofp_cash": ("SOCF cash", "SOFP cash"),
    "sore_to_sofp_retained_earnings": ("SoRE retained", "SOFP retained"),
}


def _check_id_endpoints(name: str) -> tuple[str, str]:
    return _CHECK_ENDPOINTS.get(name, ("LHS", "RHS"))
