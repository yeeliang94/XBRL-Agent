"""Monolith agent tools — `get_state`, `write_cells`, `done`.

The agent's three handles to the world (PRD §6). All three are pure
Python — no PydanticAI bindings here so they're trivially unit-testable.
The coordinator wraps them as `Agent.tool_plain` instances at agent build
time.
"""
from __future__ import annotations

import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import openpyxl

from statement_types import StatementType
from tools.fill_workbook import FieldMapping, fill_workbook as _fill_workbook_impl
from monolith.state import (
    StateSnapshot,
    build_state_snapshot,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------


@dataclass
class CellWrite:
    """One cell write request (PRD §6 — mirrors `FieldMapping`)."""
    sheet: str
    col: str = ""                       # "cy" | "py" | "evidence" (non-matrix)
    row: Optional[int] = None
    label: Optional[str] = None
    section: Optional[str] = None
    matrix_col: Optional[str] = None    # SOCIE only
    value: Any = None
    evidence: Optional[str] = None      # always written to evidence col


@dataclass
class WriteResult:
    """Outcome of one `write_cells` batch (PRD §6)."""
    written: list[dict] = field(default_factory=list)
    rejected: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "written": list(self.written),
            "rejected": list(self.rejected),
        }


@dataclass
class Accept:
    """One server-validated `accept_imbalance` entry (PRD §6 `done`)."""
    check_id: str
    reason: str
    pdf_page: int
    evidence_excerpt: str


@dataclass
class CompletionResult:
    status: str                          # "done" | "not_done"
    failing_checks: list[str] = field(default_factory=list)
    accepted_residuals: list[dict] = field(default_factory=list)
    message: str = ""

    def to_dict(self) -> dict:
        return {
            "status": self.status,
            "failing_checks": list(self.failing_checks),
            "accepted_residuals": list(self.accepted_residuals),
            "message": self.message,
        }


# Column index mapping. `evidence` is NOT a `col` value — it rides
# alongside value writes via the `evidence` field per PRD §6 worked
# examples. The writer (tools.fill_workbook) takes evidence as metadata
# on the value write and routes it to col D (Company) / F (Group) /
# resolved-source-col (SOCIE) automatically.
_COL_NAME_TO_INDEX = {"cy": 2, "py": 3}
_MATRIX_SHEETS = {"SOCIE"}  # MFRS SOCIE only — MPERS SOCIE/SoRE is out-of-scope (PRD §3)


# ---------------------------------------------------------------------------
# Tool factories
# ---------------------------------------------------------------------------


class MonolithToolContext:
    """Carries the per-run state every tool call needs.

    Held by the coordinator as the agent's `deps`. The agent itself only
    sees the tool functions; this object stays server-side.
    """

    def __init__(
        self,
        *,
        workbook_path: str,
        pdf_page_count: int,
        filing_standard: str = "mfrs",
        filing_level: str = "company",
        statements: Optional[list[StatementType]] = None,
        variants: Optional[dict[StatementType, str]] = None,
        history_repeat_threshold: int = 3,
        history_max_entries: int = 5,
        # Canonical mode (Phase B): when all three are set, every
        # successful fill_workbook write is projected into
        # `run_concept_facts` so the Values / Concepts page populates.
        # template_id_by_sheet is built by the coordinator at startup —
        # one entry per worksheet across all 5 face templates. Without it
        # the monolith path skips projection and the Values page stays
        # empty (the bug surfaced by the first real run).
        run_id: Optional[int] = None,
        db_path: Optional[str] = None,
        template_id_by_sheet: Optional[dict[str, str]] = None,
    ):
        self.workbook_path = workbook_path
        self.pdf_page_count = max(int(pdf_page_count), 0)
        self.filing_standard = filing_standard
        self.filing_level = filing_level
        self.statements = list(statements or list(StatementType))
        self.variants = dict(variants or {})
        self.run_id = run_id
        self.db_path = db_path
        self.template_id_by_sheet = dict(template_id_by_sheet or {})
        # Per-cell write history. Key: (sheet, row, col_or_matrix_col).
        # Value: list of values written in order. Surfaces a hint when the
        # same value repeats ≥ threshold times — see PRD §5 `history_hints`.
        self._writes_by_cell: dict[tuple[str, int, str], list[Any]] = {}
        self.turn = 0
        self._history_repeat_threshold = history_repeat_threshold
        self._history_max_entries = history_max_entries
        # Last computed snapshot, exposed so `done()` can check the
        # currently-failing checks without recomputing from disk.
        self.last_snapshot: Optional[StateSnapshot] = None

    # ---------------------------------------------------------------
    # Helpers
    # ---------------------------------------------------------------

    def record_write(self, sheet: str, row: int, col_key: str, value: Any) -> None:
        key = (sheet, int(row), col_key)
        history = self._writes_by_cell.setdefault(key, [])
        history.append(value)
        if len(history) > self._history_max_entries:
            del history[0 : len(history) - self._history_max_entries]

    def history_hints(self) -> list[dict]:
        out: list[dict] = []
        for (sheet, row, col_key), values in self._writes_by_cell.items():
            if len(values) < self._history_repeat_threshold:
                continue
            tail = values[-self._history_repeat_threshold :]
            if all(_eq_lenient(v, tail[0]) for v in tail):
                out.append({
                    "sheet": sheet,
                    "row": row,
                    "col": col_key,
                    "value": tail[0],
                    "note": (
                        f"you've written this same value {len(tail)} times "
                        "— try a different approach"
                    ),
                })
        return out


def _eq_lenient(a: Any, b: Any) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return a == b


def _workbook_has_explicit_numerics(workbook_path: str) -> bool:
    """True if any data column on any sheet carries a non-formula numeric.

    Templates ship with formulas + labels only — zero numeric values in
    columns B+. After a successful `write_cells` batch the writer
    deposits real numbers. This is the persisted-state half of the
    empty-workbook guard (peer-review HIGH #1): on a resumed run, the
    in-memory `_writes_by_cell` tracker is empty, but the workbook
    already has data — the guard must not refuse `done` in that case.

    Reads with `data_only=False` so formula cells appear as their
    formula string (skipped) rather than their cached value.
    """
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
    except Exception as exc:
        logger.debug(
            "Could not open %s to check pre-existing writes: %s",
            workbook_path, exc,
        )
        return False
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.column < 2:
                        # Col A is labels; never a data write.
                        continue
                    v = cell.value
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        return True
        return False
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# get_state
# ---------------------------------------------------------------------------


def get_state(ctx: MonolithToolContext) -> dict:
    """Return the per-turn dashboard. PRD §5."""
    ctx.turn += 1
    snap = build_state_snapshot(
        ctx.workbook_path,
        filing_standard=ctx.filing_standard,
        filing_level=ctx.filing_level,
        statements=ctx.statements,
        variants=ctx.variants,
        turn=ctx.turn,
        history_hints=ctx.history_hints(),
    )
    ctx.last_snapshot = snap
    return snap.to_dict()


# ---------------------------------------------------------------------------
# write_cells
# ---------------------------------------------------------------------------


def write_cells(ctx: MonolithToolContext, writes: list[dict]) -> dict:
    """Persist a batch of cell writes onto the live workbook.

    Validation per PRD §6:
      - matrix sheet writes use matrix_col (never col=cy/py)
      - non-matrix sheets reject matrix_col
      - abstract / formula cells refused (the writer carries the actual
        guard — gotchas #3, #17 — we surface its rejection here)
      - duplicate writes in one batch refused
      - unknown matrix_col label refused
      - type mismatch (string value on a numeric col) refused
    """
    parsed, batch_rejections = _parse_and_validate_batch(ctx, writes)

    written: list[dict] = []
    rejected: list[dict] = list(batch_rejections)

    if not parsed:
        return WriteResult(written=written, rejected=rejected).to_dict()

    # SOCIE matrix writes route through col-by-col `FieldMapping`s with
    # the matrix column index resolved from row-2 headers. Linear-sheet
    # writes use the standard col mapping. Both flow through the existing
    # `fill_workbook` so the gotcha #17 abstract-row guard fires.
    mappings, mapping_rejections = _to_field_mappings(ctx, parsed)
    rejected.extend(mapping_rejections)

    if not mappings:
        return WriteResult(written=written, rejected=rejected).to_dict()

    fields_json = json.dumps([_mapping_to_jsonable(m) for m in mappings])
    fill_result = _fill_workbook_impl(
        template_path=ctx.workbook_path,
        output_path=ctx.workbook_path,  # in-place overwrite — the live workbook
        fields_json=fields_json,
        filing_level=ctx.filing_level,
    )

    # The writer reports errors as a flat list of strings. Surface every
    # one as a rejection keyed back to the originating write — best-effort
    # match by `(sheet, label-or-row)`.
    if fill_result.errors:
        for err in fill_result.errors:
            rejected.append({
                "write": None,
                "reason": err,
            })

    # Mark the resolved writes as written. `fill_result.resolved_writes`
    # carries (sheet, row, col, value, evidence) per accepted write.
    for rw in fill_result.resolved_writes:
        written.append({
            "sheet": rw["sheet"],
            "row": rw["row"],
            "col": _index_to_col_name(rw["col"]),
            "value": rw["value"],
        })
        # Track for history_hints. Use the col name (cy/py) so the same
        # cell-key is used regardless of which side wrote (matrix writes
        # use the matrix_col label instead — recorded below).
        ctx.record_write(
            rw["sheet"], rw["row"], _index_to_col_name(rw["col"]), rw["value"],
        )

    # Canonical-mode projection: route each resolved write into
    # `run_concept_facts` so the Values / Concepts page populates.
    # Mirrors extraction/agent.py::_project_facts_if_canonical, but the
    # monolith spans 5 templates in one workbook — so we group writes by
    # sheet → template_id and call project_writes once per template.
    # Never raises; a projection hiccup never breaks the run.
    _project_monolith_writes_if_canonical(ctx, fill_result.resolved_writes)

    # Surface writer-side warnings (double-booking) as advisory entries
    # under `rejected[*].reason` prefixed with "warning:" — the agent
    # should treat them as hints, not blockers. PRD §6 leaves the format
    # open; this keeps a single channel.
    for w in fill_result.warnings:
        rejected.append({"write": None, "reason": f"warning: {w}"})

    return WriteResult(written=written, rejected=rejected).to_dict()


def _parse_and_validate_batch(
    ctx: MonolithToolContext, writes: list[dict],
) -> tuple[list[CellWrite], list[dict]]:
    """Initial pass — schema validation + duplicate detection + matrix gate.

    Returns (accepted_writes, rejections). Each rejection carries the
    offending write dict + a structured reason.
    """
    accepted: list[CellWrite] = []
    rejections: list[dict] = []
    seen_keys: set[tuple[str, Any, str]] = set()

    for raw in writes:
        try:
            cw = _coerce_write(raw)
        except ValueError as exc:
            rejections.append({"write": raw, "reason": str(exc)})
            continue

        is_matrix = cw.sheet.upper() in _MATRIX_SHEETS
        # Matrix gating per PRD §6.
        if is_matrix:
            if cw.col:
                rejections.append({
                    "write": raw,
                    "reason": (
                        f"col {cw.col!r} not valid on matrix sheet "
                        f"{cw.sheet} — use `matrix_col` instead "
                        "(equity-component label, e.g. 'RetainedEarnings')."
                    ),
                })
                continue
            if not cw.matrix_col:
                rejections.append({
                    "write": raw,
                    "reason": (
                        f"matrix sheet {cw.sheet} requires `matrix_col` "
                        "(equity-component label)."
                    ),
                })
                continue
        else:
            if cw.matrix_col:
                rejections.append({
                    "write": raw,
                    "reason": (
                        f"matrix_col is not valid on non-matrix sheet "
                        f"{cw.sheet}."
                    ),
                })
                continue
            if not cw.col:
                rejections.append({
                    "write": raw,
                    "reason": "`col` is required on non-matrix sheet writes.",
                })
                continue
            if cw.col.lower() not in _COL_NAME_TO_INDEX:
                rejections.append({
                    "write": raw,
                    "reason": (
                        f"col {cw.col!r} not recognised; valid: "
                        f"{sorted(_COL_NAME_TO_INDEX)} (evidence rides via "
                        "the `evidence` field, not as a col)."
                    ),
                })
                continue

        # Row OR (label + section) required.
        if cw.row is None and not cw.label:
            rejections.append({
                "write": raw,
                "reason": "either `row` or `label` is required.",
            })
            continue

        # Type sanity for the value field. cy/py and matrix_col carry a
        # number.
        col_key = cw.matrix_col or cw.col.lower()
        if cw.value is not None and not isinstance(cw.value, (int, float)):
            rejections.append({
                "write": raw,
                "reason": (
                    f"value for {col_key!r} must be a number; "
                    f"got {type(cw.value).__name__}."
                ),
            })
            continue

        # Duplicate gating within one batch. Same (sheet, row-or-label, col-key).
        row_part = cw.row if cw.row is not None else cw.label
        dup_key = (cw.sheet, row_part, col_key)
        if dup_key in seen_keys:
            rejections.append({
                "write": raw,
                "reason": (
                    "duplicate write to the same cell in this batch — "
                    "order your writes; second occurrence dropped."
                ),
            })
            continue
        seen_keys.add(dup_key)

        accepted.append(cw)

    return accepted, rejections


def _coerce_write(raw: dict) -> CellWrite:
    if not isinstance(raw, dict):
        raise ValueError(f"write must be an object, got {type(raw).__name__}")
    sheet = raw.get("sheet")
    if not sheet or not isinstance(sheet, str):
        raise ValueError("write missing required `sheet` field")
    row = raw.get("row")
    if row is not None:
        try:
            row = int(row)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"`row` must be an integer: {exc}") from exc
    return CellWrite(
        sheet=sheet,
        col=str(raw.get("col") or ""),
        row=row,
        label=raw.get("label"),
        section=raw.get("section"),
        matrix_col=raw.get("matrix_col"),
        value=raw.get("value"),
        evidence=raw.get("evidence"),
    )


def _to_field_mappings(
    ctx: MonolithToolContext, writes: list[CellWrite],
) -> tuple[list[FieldMapping], list[dict]]:
    """Translate accepted CellWrites into `FieldMapping`s for fill_workbook.

    SOCIE matrix writes need the matrix col index resolved from row-2
    headers (the writer doesn't carry matrix awareness — the monolith
    path resolves it here). Unknown matrix_col labels become rejections.
    """
    if not writes:
        return [], []

    mappings: list[FieldMapping] = []
    rejections: list[dict] = []

    # Only load the workbook when the batch actually contains a SOCIE
    # write — non-matrix batches don't need the row-2 header map and the
    # load is the most expensive part of write_cells.
    has_matrix_write = any(cw.sheet.upper() in _MATRIX_SHEETS for cw in writes)
    socie_header_index: dict[str, int] = {}
    wb = None
    if has_matrix_write:
        wb = openpyxl.load_workbook(ctx.workbook_path, data_only=False)
        if "SOCIE" in wb.sheetnames:
            ws = wb["SOCIE"]
            for col in range(2, ws.max_column + 1):
                v = ws.cell(row=2, column=col).value
                if v is None:
                    continue
                text = str(v).strip()
                if text:
                    socie_header_index[text] = col
                    # Also index a lowercase / dehyphenated alias so the
                    # agent can pass labels in either form.
                    socie_header_index.setdefault(text.lower(), col)

    try:
        for cw in writes:
            is_matrix = cw.sheet.upper() in _MATRIX_SHEETS
            if is_matrix:
                # Resolve matrix column from header text.
                key = cw.matrix_col or ""
                col_idx = socie_header_index.get(key) or socie_header_index.get(
                    key.lower(),
                )
                if col_idx is None:
                    # Build the human-friendly valid set without the
                    # lowercase aliases we tucked into the index for
                    # tolerant lookup.
                    valid = sorted({
                        v for v in socie_header_index.keys()
                        if any(c.isupper() for c in v)
                    })
                    rejections.append({
                        "write": _cw_to_dict(cw),
                        "reason": (
                            f"matrix_col {cw.matrix_col!r} not present on "
                            f"SOCIE; valid headers: {valid}"
                        ),
                    })
                    continue
                col = col_idx
            else:
                col = _COL_NAME_TO_INDEX[cw.col.lower()]

            mappings.append(FieldMapping(
                sheet=cw.sheet,
                field_label=cw.label or "",
                col=col,
                value=cw.value,
                section=cw.section or "",
                row=cw.row,
                evidence=cw.evidence or "",
            ))
    finally:
        if wb is not None:
            wb.close()
    return mappings, rejections


def _cw_to_dict(cw: CellWrite) -> dict:
    out = {"sheet": cw.sheet}
    if cw.row is not None:
        out["row"] = cw.row
    if cw.label:
        out["label"] = cw.label
    if cw.section:
        out["section"] = cw.section
    if cw.col:
        out["col"] = cw.col
    if cw.matrix_col:
        out["matrix_col"] = cw.matrix_col
    if cw.value is not None:
        out["value"] = cw.value
    if cw.evidence:
        out["evidence"] = cw.evidence
    return out


def _mapping_to_jsonable(m: FieldMapping) -> dict:
    out: dict = {
        "sheet": m.sheet,
        "col": int(m.col),
        "value": m.value,
    }
    if m.field_label:
        out["field_label"] = m.field_label
    if m.section:
        out["section"] = m.section
    if m.row is not None:
        out["row"] = m.row
    if m.evidence:
        out["evidence"] = m.evidence
    return out


def _project_monolith_writes_if_canonical(
    ctx: "MonolithToolContext", resolved_writes: list[dict],
) -> None:
    """Project the just-written cells into `run_concept_facts`.

    Active only when canonical mode is on (ctx.run_id + db_path +
    template_id_by_sheet all populated). Groups writes by sheet → matches
    the per-template template_id → calls `concept_model.cell_resolver
    .project_writes` per template so each call's writes belong to ONE
    template (the API contract).

    Best-effort: a failure here logs and returns. The xlsx is the
    authoritative artefact regardless; the facts table is a derivative
    view for the Concepts UI.
    """
    if not (
        ctx.run_id is not None
        and ctx.db_path
        and ctx.template_id_by_sheet
        and resolved_writes
    ):
        return
    try:
        from collections import defaultdict
        from concept_model.cell_resolver import project_writes

        by_template: dict[str, list[dict]] = defaultdict(list)
        unmapped = 0
        for w in resolved_writes:
            tid = ctx.template_id_by_sheet.get(w.get("sheet"))
            if tid is None:
                unmapped += 1
                continue
            by_template[tid].append(w)
        if unmapped:
            logger.debug(
                "monolith projection: %d writes had no template_id mapping "
                "(sheets %s)",
                unmapped,
                sorted({
                    w.get("sheet") for w in resolved_writes
                    if w.get("sheet") not in ctx.template_id_by_sheet
                }),
            )
        total_projected = 0
        for tid, writes in by_template.items():
            proj = project_writes(
                ctx.db_path,
                ctx.run_id,
                tid,
                writes,
                filing_level=ctx.filing_level,
            )
            total_projected += proj.projected
            if proj.has_gaps:
                logger.info(
                    "monolith projection (template %s): %d saved, "
                    "%d skipped, %d rejected",
                    tid, proj.projected,
                    len(proj.skipped), len(proj.rejected),
                )
        if total_projected:
            logger.debug(
                "monolith projection: %d fact(s) saved across %d template(s)",
                total_projected, len(by_template),
            )
    except Exception:
        # Canonical projection is advisory. The xlsx write already
        # succeeded; a failure here must never break the run.
        logger.warning(
            "monolith canonical projection failed (writes=%d)",
            len(resolved_writes), exc_info=True,
        )


def _index_to_col_name(col: int) -> str:
    # Linear sheets: B=cy / C=py. SOCIE matrix columns (B..W) get rendered
    # as `col_<n>` so the agent can still trace what landed; cleaner names
    # would require the matrix-header map plumbed through here.
    return {2: "cy", 3: "py"}.get(int(col), f"col_{col}")


# ---------------------------------------------------------------------------
# done
# ---------------------------------------------------------------------------


def done(
    ctx: MonolithToolContext,
    accept_imbalance: Optional[list[dict]] = None,
) -> dict:
    """Agent signals completion. PRD §6.

    Validation contract:
      1. `check_id` must correspond to a currently-failing check.
      2. `pdf_page` must satisfy 1 ≤ page ≤ N.
      3. `evidence_excerpt` non-empty, ≤ 200 chars.

    Any failure returns `status="not_done"` with the offending entries
    surfaced.
    """
    accept_imbalance = list(accept_imbalance or [])

    # Always recompute the snapshot — even if `last_snapshot` is set, the
    # workbook may have moved underneath it since the agent's last
    # `get_state` call.
    snap = build_state_snapshot(
        ctx.workbook_path,
        filing_standard=ctx.filing_standard,
        filing_level=ctx.filing_level,
        statements=ctx.statements,
        variants=ctx.variants,
        turn=ctx.turn,
        history_hints=ctx.history_hints(),
    )
    ctx.last_snapshot = snap

    # Empty-workbook guard. An effectively-blank workbook makes most
    # cross-checks pass vacuously (0 == 0), which means a single weak
    # accept_imbalance could rubber-stamp a completely empty filling.
    # Refuse `done` when neither the current agent has written nor the
    # workbook on disk contains any explicit numeric data — regardless
    # of cross-check state. The 2026-05-28 scanned-PDF incident (run
    # 82dd3ac8) triggered exactly this exit path. Reading persisted
    # data (not just in-memory) covers the resumed-run path documented
    # at `_materialise_workbook` and flagged by peer-review HIGH #1.
    if not ctx._writes_by_cell and not _workbook_has_explicit_numerics(
        ctx.workbook_path,
    ):
        return CompletionResult(
            status="not_done",
            failing_checks=sorted(
                {c.id for c in snap.cross_checks if not c.pass_}
                | {c.id for c in snap.verifier if not c.pass_}
            ),
            accepted_residuals=[],
            message=(
                "Cannot finalise — the workbook has no cell writes yet. "
                "Extract values from the PDF and call write_cells before "
                "calling done."
            ),
        ).to_dict()

    failing_now = {
        c.id for c in snap.cross_checks if not c.pass_
    } | {
        c.id for c in snap.verifier if not c.pass_
    }

    # Validate every accept entry against the contract.
    accepted: list[dict] = []
    invalid: list[dict] = []
    accepted_check_ids: set[str] = set()
    for entry in accept_imbalance:
        if not isinstance(entry, dict):
            invalid.append({
                "entry": entry,
                "reason": "accept entry must be an object",
            })
            continue
        check_id = entry.get("check_id")
        reason = entry.get("reason") or ""
        pdf_page = entry.get("pdf_page")
        excerpt = entry.get("evidence_excerpt") or ""

        if not check_id or check_id not in failing_now:
            invalid.append({
                "entry": entry,
                "reason": (
                    f"check_id {check_id!r} is not currently failing — "
                    "you can only accept a failing check."
                ),
            })
            continue
        try:
            page = int(pdf_page)
        except (TypeError, ValueError):
            invalid.append({
                "entry": entry,
                "reason": "pdf_page must be an integer.",
            })
            continue
        if page < 1 or page > ctx.pdf_page_count:
            invalid.append({
                "entry": entry,
                "reason": (
                    f"pdf_page {page} out of range [1, {ctx.pdf_page_count}]."
                ),
            })
            continue
        if not isinstance(excerpt, str) or not excerpt.strip():
            invalid.append({
                "entry": entry,
                "reason": "evidence_excerpt is required and must be non-empty.",
            })
            continue
        if len(excerpt) > 200:
            invalid.append({
                "entry": entry,
                "reason": "evidence_excerpt is too long (max 200 chars).",
            })
            continue

        accepted.append({
            "check_id": check_id,
            "reason": reason,
            "pdf_page": page,
            "evidence_excerpt": excerpt,
        })
        accepted_check_ids.add(check_id)

    if invalid:
        return CompletionResult(
            status="not_done",
            failing_checks=sorted(failing_now),
            accepted_residuals=accepted,
            message=(
                "Some accept_imbalance entries failed server-side validation."
            ),
        ).to_dict() | {"invalid_accepts": invalid}

    # Any failing checks not in the accept list block completion.
    unhandled = sorted(failing_now - accepted_check_ids)
    if unhandled:
        return CompletionResult(
            status="not_done",
            failing_checks=unhandled,
            accepted_residuals=accepted,
            message=(
                "Cannot finalise — these checks are still failing and "
                "not in accept_imbalance: " + ", ".join(unhandled)
            ),
        ).to_dict()

    return CompletionResult(
        status="done",
        failing_checks=[],
        accepted_residuals=accepted,
        message="All checks pass or are accepted residuals.",
    ).to_dict()
