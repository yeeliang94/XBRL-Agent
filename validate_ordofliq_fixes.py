#!/usr/bin/env python3
"""
Validation script for OrderOfLiquidity template fixes.

After applying fixes to the template, run this script to verify that:
1. Row 148 (Total cash) only references cash items
2. Row 168 (Total issued capital) only references capital items
3. Row 241 (Total borrowings) doesn't reference equity items
4. Row 295 (Total payables) only sums subtotals, not leaf items

Usage:
  python validate_ordofliq_fixes.py [path_to_fixed_template.xlsx]

Defaults to: /sessions/happy-zealous-dirac/mnt/xbrl-agent/XBRL-template-MFRS/02-SOFP-OrderOfLiquidity.xlsx
"""

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TEMPLATE_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
    "/sessions/happy-zealous-dirac/mnt/xbrl-agent/XBRL-template-MFRS/02-SOFP-OrderOfLiquidity.xlsx"
)
SHEET_NAME = "SOFP-Sub-OrdOfLiq"

# Expected correct formulas after fixes
EXPECTED_FORMULAS = {
    148: {
        'description': 'Total cash',
        'refs': {146, 147},  # Cash on hand, Balances with banks
        'categories': {'CASH_ITEM'},
    },
    168: {
        'description': 'Total issued capital',
        'refs': {165, 166, 167},  # Capital items only
        'categories': {'CAPITAL'},
    },
    241: {
        'description': 'Total borrowings',
        'refs': None,  # Too many to list, just check no equity items
        'categories': {'BORROWING'},  # Should only have borrowing subtotals
        'excluded_refs': {189, 190, 191, 192, 193, 194},  # Equity items to exclude
    },
    295: {
        'description': 'Total trade and other payables',
        'refs': {270, 294},  # Trade payables + rolled-up other payables
        'categories': {'SUBTOTAL'},
        'excluded_refs': {
            262, 263, 264, 265, 266, 267, 268, 269,  # Trade payable leaf items
            273, 274, 275, 276, 277,  # Other payables due to related leaf items
            280, 281, 282,  # NCI leaf items
            285, 286, 287, 288, 289, 290, 291, 292,  # Other non-trade leaf items
        }
    }
}


def load_data(wb):
    ws = wb[SHEET_NAME]
    labels = {}
    formulas = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        row_num = row[0].row
        if row[0].value:
            labels[row_num] = str(row[0].value).strip()

        b_cell = row[1]
        if b_cell.data_type == 'f':
            formulas[row_num] = b_cell.value

    return labels, formulas


def extract_refs(formula: str) -> set:
    if not formula or not formula.startswith('='):
        return set()
    pattern = r'B(\d+)'
    matches = re.findall(pattern, formula)
    return set(int(m) for m in matches)


def categorize_reference(label: str, has_formula: bool = False) -> str:
    label_lower = label.lower()

    if has_formula and (
        label.startswith("*")
        or label_lower.startswith("total ")
        or label_lower in {"trade payables", "other payables"}
        or " due to " not in label_lower and "payables" in label_lower
    ):
        return "SUBTOTAL"
    elif ("cash" in label_lower and ("hand" in label_lower or "bank" in label_lower)) or (
        "balances with bank" in label_lower
    ):
        return "CASH_ITEM"
    elif "inventory" in label_lower or "raw material" in label_lower or "work in progress" in label_lower or "finished good" in label_lower or "spare part" in label_lower:
        return "INVENTORY"
    elif "derivative" in label_lower:
        return "DERIVATIVE"
    elif "prepaid" in label_lower or "accrual" in label_lower:
        return "PREPAID_ACCRUAL"
    elif "capital" in label_lower or ("share" in label_lower and "equity" not in label_lower):
        return "CAPITAL"
    elif "perpetual sukuk" in label_lower or "equity component" in label_lower or "iculs" in label_lower:
        return "EQUITY"
    elif any(x in label_lower for x in ["loan", "bond", "sukuk", "financing", "hire purchase", "mtn", "bankers acceptance", "bank overdraft", "trade financing", "revolving credit", "borrowing"]):
        return "BORROWING"
    elif "subtotal" in label_lower or label.startswith('*') or (any(x in label_lower for x in ["total", "payable"]) and "due to" not in label_lower and "deferred" not in label_lower):
        # Heuristic: if it looks like a total/subtotal and not a leaf item
        return "SUBTOTAL"
    elif "payable" in label_lower:
        return "PAYABLE"
    else:
        return "OTHER"


def validate_template(template_path: Path) -> dict[str, Any]:
    """Validate the fixed OrderOfLiquidity template and return structured results."""
    if not template_path.exists():
        raise FileNotFoundError(f"File not found: {template_path}")

    wb = load_workbook(template_path, data_only=False)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise KeyError(f"Sheet '{SHEET_NAME}' not found")

        labels, formulas = load_data(wb)
        results: list[dict[str, Any]] = []

        for row, expected in EXPECTED_FORMULAS.items():
            outcome: dict[str, Any] = {
                "row": row,
                "description": expected["description"],
                "passed": False,
                "messages": [],
            }

            if row not in formulas:
                outcome["messages"].append(f"No formula found in row {row}")
                results.append(outcome)
                continue

            formula = formulas[row]
            refs = extract_refs(formula)
            outcome["formula"] = formula
            outcome["refs"] = sorted(refs)

            excluded_refs = expected.get("excluded_refs") or set()
            excluded_found = refs & excluded_refs
            if excluded_found:
                outcome["messages"].append(
                    f"Found excluded references: {sorted(excluded_found)}"
                )
                results.append(outcome)
                continue

            if expected.get("refs") is not None:
                expected_refs = expected["refs"]
                missing = expected_refs - refs
                extra = refs - expected_refs

                if missing:
                    outcome["messages"].append(
                        f"Missing expected references: {sorted(missing)}"
                    )
                    results.append(outcome)
                    continue

                if extra:
                    outcome["messages"].append(
                        f"Found unexpected references: {sorted(extra)}"
                    )
                    results.append(outcome)
                    continue

            allowed_categories = expected.get("categories")
            if allowed_categories:
                bad_categories: dict[str, list[tuple[int, str]]] = {}
                for ref in refs:
                    ref_label = labels.get(ref, "[NOT FOUND]")
                    category = categorize_reference(ref_label, has_formula=ref in formulas)
                    if category not in allowed_categories:
                        bad_categories.setdefault(category, []).append((ref, ref_label))

                if bad_categories:
                    rendered = []
                    for category, items in sorted(bad_categories.items()):
                        rendered.append(
                            f"{category}: "
                            + ", ".join(f"row {ref} ({label})" for ref, label in items[:3])
                        )
                    outcome["messages"].append(
                        "Found unexpected categories: " + "; ".join(rendered)
                    )
                    results.append(outcome)
                    continue

            outcome["passed"] = True
            results.append(outcome)

        passed = sum(1 for item in results if item["passed"])
        failed = len(results) - passed
        return {
            "results": results,
            "passed": passed,
            "failed": failed,
        }
    finally:
        wb.close()


def main():
    print("=" * 100)
    print("ORDEROF LIQUIDITY TEMPLATE FIX VALIDATION")
    print(f"File: {TEMPLATE_PATH}")
    print("=" * 100)
    print()

    try:
        validation = validate_template(TEMPLATE_PATH)
    except Exception as e:
        print(f"ERROR: {e}")
        return 1

    # Validate each expected formula
    for result in validation["results"]:
        row = result["row"]
        print(f"\nValidating Row {row}: {result['description']}")
        print("─" * 100)
        if "formula" in result:
            print(f"  Formula: {result['formula'][:80]}...")
            print(f"  References: {result['refs']}")
        if result["passed"]:
            print("  ✓ PASS")
        else:
            for message in result["messages"]:
                print(f"  ✗ FAIL: {message}")

    # Summary
    print("\n" + "=" * 100)
    print(f"RESULTS: {validation['passed']} passed, {validation['failed']} failed")
    print("=" * 100)

    if validation["failed"] == 0:
        print("\n✓ All validations passed! The template has been correctly fixed.")
        return 0
    else:
        print(f"\n✗ {validation['failed']} validation(s) failed. Review the errors above.")
        return 1


if __name__ == "__main__":
    sys.exit(main())
