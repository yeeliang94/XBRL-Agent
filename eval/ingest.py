"""Reverse-ingestion — a human-filled workbook becomes gold facts.

A reviewer fills the SSM MBRS template ``.xlsx`` by hand (the same template the
agents target) and uploads it. This module parses that workbook into
``gold_concept_facts`` for a benchmark, reusing the existing
``cell_resolver.resolve_cell`` so there is NO new mapping logic — the same
``(sheet, row, col) → (concept_uuid, period, entity_scope)`` inverse the
extraction write-path already trusts.

Key rules (docs/PRD-eval-benchmark.md §"Clarifications"):

* **Multi-statement.** A filled workbook spans many sheets/templates.
  ``ingest_workbook`` takes the benchmark's template *set* and resolves each
  worksheet to the matching ``template_id`` by sheet name. Sheets outside the
  set are skipped; a workbook whose sheets match NO benchmark template is
  rejected loudly.
* **Leaves only.** Only ``LEAF`` / ``MATRIX_CELL`` concepts are ingested.
  COMPUTED totals are skipped (they're Excel-formula-derived; grading excludes
  them anyway).
* **Deterministic read.** ``openpyxl(data_only=True)`` reads the cached
  computed value, not the formula string. Only value columns resolve (col A
  labels + source/evidence columns resolve to ``None`` and are skipped).
  Accountant text a human may have typed (``(95)`` → -95, thousands commas) is
  parsed.
"""
from __future__ import annotations

import logging
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from concept_model.cell_resolver import resolve_cell

logger = logging.getLogger(__name__)

# Widest value-column span we scan per row. Company uses B/C (2-3), Group
# B/C/D/E (2-5), and the SOCIE matrix spans the equity-component grid out to
# col X (24) on MFRS. resolve_cell returns None for any non-value column, so
# over-scanning is harmless — it just costs a few extra resolver lookups.
_MAX_VALUE_COL = 30


@dataclass
class IngestResult:
    """Outcome of ingesting one workbook into a benchmark's gold facts."""
    ingested: int = 0
    skipped: int = 0
    # Gradeable cells (LEAF/MATRIX_CELL) SILENTLY dropped because the workbook
    # cell holds a live formula with no cached value — openpyxl's
    # ``data_only=True`` reads those as ``None``. A machine-exported workbook
    # stores the SOCIE matrix + cross-sheet face rollups this way, so an upload
    # of an un-recalculated export loses them (the 2026-06-05 sub-sheet-loss
    # incident). Surfaced as a warning; seeding from a run avoids it entirely.
    skipped_formula_cells: int = 0
    matched_sheets: list[str] = field(default_factory=list)
    unmatched_sheets: list[str] = field(default_factory=list)


def _now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace(
        "+00:00", "Z"
    )


# Matches accountant-style numbers a human might type as text:
# "1,595", "(95)", "-95", "1234.5", "1,234.56". A bare dash / "N/A" / blank
# returns None (ambiguous nil — we don't guess a value for gold). Parentheses
# must be BALANCED — "(95)" is a negative, but an unbalanced "(95" / "95)" is
# malformed text and rejected (the old `\(?...\)?` form silently coerced it to
# a positive value and stored it as gold).
_NUMBER_RE = re.compile(
    r"^(?:\([\d,]+(?:\.\d+)?\)|-?[\d,]+(?:\.\d+)?)$"
)


def parse_accounting_number(raw) -> float | None:
    """Parse a cell value into a float, or ``None`` if it isn't a number.

    Handles native numerics directly and accountant text conventions
    (parentheses = negative, thousands commas). Returns ``None`` for blanks,
    dashes, and free text so non-value cells never become spurious gold.
    """
    if raw is None:
        return None
    if isinstance(raw, bool):
        # openpyxl can surface TRUE/FALSE; never a financial value.
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    if not isinstance(raw, str):
        return None
    text = raw.strip()
    if not text:
        return None
    if not _NUMBER_RE.match(text):
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "")
    try:
        value = float(text)
    except ValueError:
        return None
    return -value if negative else value


def _sheet_to_template(
    conn: sqlite3.Connection, template_ids: list[str]
) -> dict[str, str]:
    """Map every render/alias sheet name in the template set to its owning
    ``template_id`` so a worksheet title resolves to one template.

    Each statement variant has uniquely-named sheets (e.g. ``SOFP-CuNonCu``,
    ``SOFP-Sub-CuNonCu``), so within one benchmark's set the mapping is
    unambiguous.
    """
    mapping: dict[str, str] = {}
    if not template_ids:
        return mapping
    placeholders = ",".join("?" for _ in template_ids)
    for r in conn.execute(
        f"SELECT DISTINCT render_sheet, template_id FROM concept_nodes "
        f"WHERE template_id IN ({placeholders})",
        tuple(template_ids),
    ).fetchall():
        mapping[r[0]] = r[1]
    # Cross-sheet face aliases (schema v11) — the face sheet that mirrors a
    # sub-sheet total. Map it too so a value typed on the face row resolves.
    for r in conn.execute(
        f"SELECT DISTINCT a.alias_sheet, n.template_id "
        f"FROM concept_render_aliases a "
        f"JOIN concept_nodes n ON n.concept_uuid = a.concept_uuid "
        f"WHERE n.template_id IN ({placeholders})",
        tuple(template_ids),
    ).fetchall():
        mapping.setdefault(r[0], r[1])
    return mapping


def _gradeable_kinds(
    conn: sqlite3.Connection, template_ids: list[str]
) -> dict[str, str]:
    """``{concept_uuid: kind}`` for the template set, so the ingester can skip
    COMPUTED concepts without a per-cell query."""
    if not template_ids:
        return {}
    placeholders = ",".join("?" for _ in template_ids)
    return {
        r[0]: r[1]
        for r in conn.execute(
            f"SELECT concept_uuid, kind FROM concept_nodes "
            f"WHERE template_id IN ({placeholders})",
            tuple(template_ids),
        ).fetchall()
    }


def ingest_workbook(
    conn: sqlite3.Connection,
    benchmark_id: int,
    xlsx_path: str | Path,
    template_ids: list[str],
) -> IngestResult:
    """Parse a human-filled workbook into ``gold_concept_facts``.

    ``template_ids`` is the benchmark's explicit template SET. For each
    worksheet whose title belongs to the set, every populated value cell is
    resolved to a ``(concept_uuid, period, entity_scope)`` and upserted as a
    gold fact. Only LEAF / MATRIX_CELL concepts are stored; COMPUTED and
    non-resolving cells are counted in ``skipped``.

    Raises ``ValueError`` if the workbook's sheets match NO benchmark template
    (a loud signal that the wrong file or wrong standard/level was uploaded —
    silently producing zero gold facts would be worse).

    The caller owns the connection and the commit; this batches all writes and
    leaves the transaction open so an enclosing handler can roll back on error.
    """
    xlsx_path = str(xlsx_path)
    if not template_ids:
        raise ValueError(
            "Benchmark has no template set — add templates before ingesting."
        )

    sheet_map = _sheet_to_template(conn, template_ids)
    kinds = _gradeable_kinds(conn, template_ids)
    now = _now()
    source = f"ingested from {Path(xlsx_path).name}"

    # data_only=True returns the cached computed value, not the formula text —
    # exactly what we want for human-typed leaf cells.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # A second view that keeps formula STRINGS (data_only=False) so we can tell
    # an empty cell apart from an un-recalculated formula cell. Only consulted
    # on the None path, and only the cheap "starts with '='" check runs per
    # blank cell — resolve_cell fires solely for actual formula cells. Loaded
    # inside the try so a failure here still closes ``wb`` in the finally.
    wb_formulas = None
    result = IngestResult()
    try:
        wb_formulas = openpyxl.load_workbook(xlsx_path, data_only=False)
        for ws in wb.worksheets:
            template_id = sheet_map.get(ws.title)
            if template_id is None:
                result.unmatched_sheets.append(ws.title)
                continue
            result.matched_sheets.append(ws.title)
            ws_f = wb_formulas[ws.title]
            max_col = min(ws.max_column or 0, _MAX_VALUE_COL)
            for row in range(1, (ws.max_row or 0) + 1):
                for col in range(2, max_col + 1):
                    value = parse_accounting_number(
                        ws.cell(row=row, column=col).value
                    )
                    if value is None:
                        # Distinguish a genuinely blank cell from a gradeable
                        # cell lost to a missing formula cache (so the caller
                        # can warn instead of silently shipping sparse gold).
                        raw_f = ws_f.cell(row=row, column=col).value
                        if isinstance(raw_f, str) and raw_f.startswith("="):
                            resolved_f = resolve_cell(
                                conn, template_id, ws.title, row, col
                            )
                            if resolved_f is not None and kinds.get(
                                resolved_f[0]
                            ) in ("LEAF", "MATRIX_CELL"):
                                result.skipped_formula_cells += 1
                        continue
                    resolved = resolve_cell(conn, template_id, ws.title, row, col)
                    if resolved is None:
                        # Non-value column or a row with no concept — skip.
                        result.skipped += 1
                        continue
                    concept_uuid, period, entity_scope = resolved
                    if kinds.get(concept_uuid) not in ("LEAF", "MATRIX_CELL"):
                        # COMPUTED total (or unknown) — never gold.
                        result.skipped += 1
                        continue
                    conn.execute(
                        "INSERT INTO gold_concept_facts(benchmark_id, "
                        "concept_uuid, period, entity_scope, value, "
                        "value_status, source, updated_at) "
                        "VALUES (?, ?, ?, ?, ?, 'observed', ?, ?) "
                        "ON CONFLICT(benchmark_id, concept_uuid, period, "
                        "entity_scope) DO UPDATE SET value = excluded.value, "
                        "value_status = excluded.value_status, "
                        "source = excluded.source, "
                        "updated_at = excluded.updated_at",
                        (benchmark_id, concept_uuid, period, entity_scope,
                         value, source, now),
                    )
                    result.ingested += 1
    finally:
        wb.close()
        if wb_formulas is not None:
            wb_formulas.close()

    if not result.matched_sheets:
        raise ValueError(
            "No worksheet in the uploaded workbook matched any of the "
            f"benchmark's templates (sheets seen: "
            f"{result.unmatched_sheets[:8]}). Check the filing standard / "
            "level and that you uploaded an MBRS template workbook."
        )

    logger.info(
        "ingest_workbook: benchmark %s — %d gold facts from %d sheets "
        "(%d cells skipped, %d gradeable cells lost to un-cached formulas)",
        benchmark_id, result.ingested, len(result.matched_sheets),
        result.skipped, result.skipped_formula_cells,
    )
    return result
