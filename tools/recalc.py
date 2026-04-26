"""Force formula recalc on a merged workbook.

RUN-REVIEW P0-3 (2026-04-26): the merged `filled.xlsx` ships with
`*Total …` formulas that openpyxl cannot evaluate. With
``data_only=True`` they return ``None``, which silently breaks
downstream programmatic readers — `compare_results.py`, the agent's
own ``verify_totals`` when it reads sub-sheet rollups, and any future
verifier that wants to cross-check sub-sheet leaf sums.

This helper uses the `formulas` package (pure-Python, MIT-licensed,
already in the venv) to compute every formula's value, then writes
the literal results back into the original workbook. The original
sheet names, formatting, and column structure are preserved; only
formula cells are rewritten with their cached values.

Trade-off: this REPLACES formulas with literals. Excel users no longer
see ``=B8+B9+...`` in the formula bar — they see the computed number.
For an XBRL submission workbook this is fine (SSM consumes values,
not formulas). The fallback path (recalc fails gracefully) leaves the
formulas intact and relies on Excel's ``fullCalcOnLoad=True`` flag
the merger sets — that path is the failsafe.

If `formulas` evaluation fails for any reason the caller's workbook
is left untouched and a warning is logged; the run does NOT fail on
advisory recalc telemetry.
"""
from __future__ import annotations

import logging
import os
import re
import tempfile
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Match the formulas package's cell-key shape, e.g.
#   '[sofp_company_mfrs.xlsx]SOFP-SUB-CUNONCU'!B39
# Captures sheet name (uppercase) + cell ref so we can map back to the
# original workbook's case-preserved sheet names.
_CELL_KEY_RE = re.compile(r"'\[[^\]]+\]([^']+)'!([A-Z]+\d+)")


def _extract_value(cell_obj) -> Optional[float]:
    """Pull a scalar number out of a `formulas` ``Ranges`` value object.

    ``formulas`` returns Solution[key] = Ranges instance. The numeric
    representation lives in ``.value`` as a 1×1 numpy array. Empty
    cells render as ``[[empty]]`` — those return None.
    """
    try:
        arr = getattr(cell_obj, "value", None)
        if arr is None:
            return None
        # Handle 2D numpy array (most common) and scalar fallbacks
        try:
            scalar = arr.item(0, 0) if arr.ndim == 2 else arr.item(0)
        except (AttributeError, ValueError, IndexError):
            scalar = arr
        # The "empty" sentinel object stringifies to "empty"
        if str(scalar).lower() == "empty":
            return None
        # Bool / numeric types both pass through float()
        return float(scalar)
    except Exception:  # noqa: BLE001 — best-effort scalar extraction
        return None


def recalc_workbook(path: str | Path) -> Path:
    """Recompute every formula in ``path`` and write cached values back.

    Returns the same ``path`` on success or on any failure (caller can
    treat the result as "either recalculated or unchanged"). Logs a
    WARNING on failure so operators know to fall back to opening the
    workbook in Excel for programmatic-reader use cases.

    Idempotent: running this twice produces the same output as running
    it once (the second pass has no formulas to compute).
    """
    workbook_path = Path(path)
    if not workbook_path.exists():
        logger.warning("recalc: workbook not found: %s", workbook_path)
        return workbook_path

    try:
        # Lazy import — `formulas` is a heavy dep with scipy/numpy in tow,
        # we don't want to pay the import cost on every server boot when
        # most callers won't recalc.
        import formulas  # type: ignore[import-not-found]
    except ImportError:
        logger.warning(
            "recalc: `formulas` package not installed; skipping recalc — "
            "downstream openpyxl readers will see None for formula cells"
        )
        return workbook_path

    try:
        # Compute all formulas. The formulas package builds an Excel-style
        # dependency graph and evaluates in order, handling cross-sheet
        # refs (which our face→sub rollups rely on).
        xl_model = formulas.ExcelModel().loads(str(workbook_path)).finish()
        sol = xl_model.calculate()
    except Exception:  # noqa: BLE001
        logger.warning(
            "recalc: formulas evaluation failed for %s — workbook left "
            "untouched; fullCalcOnLoad will trigger Excel-side recalc",
            workbook_path, exc_info=True,
        )
        return workbook_path

    # Build {(uppercase_sheet, cell_ref): computed_value} dict so we can
    # cheaply look up every cell when iterating openpyxl's formula cells.
    computed: dict[tuple[str, str], float] = {}
    for key, range_obj in sol.items():
        m = _CELL_KEY_RE.match(str(key))
        if not m:
            continue
        sheet_upper = m.group(1).upper()
        cell_ref = m.group(2)
        val = _extract_value(range_obj)
        if val is None:
            continue
        computed[(sheet_upper, cell_ref)] = val

    if not computed:
        logger.warning(
            "recalc: formulas evaluation returned no scalar values for %s",
            workbook_path,
        )
        return workbook_path

    # Open the ORIGINAL workbook (preserving sheet-name casing, fills,
    # widths, etc.) and overwrite formula cells with their cached values.
    # We do this in-place — the merged workbook is run-output, not a
    # template, so destructive replacement of formula strings is safe.
    try:
        wb = load_workbook(workbook_path, data_only=False)
    except Exception:  # noqa: BLE001
        logger.warning(
            "recalc: openpyxl reload failed for %s — recalc abandoned",
            workbook_path, exc_info=True,
        )
        return workbook_path

    cells_replaced = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_upper = sheet_name.upper()
        for row in ws.iter_rows():
            for cell in row:
                # Only touch formula cells. data_type "f" is openpyxl's
                # marker for formulas; skipping non-formula cells keeps
                # the rest of the workbook bit-identical to the input.
                if cell.data_type != "f":
                    continue
                key = (sheet_upper, cell.coordinate)
                if key in computed:
                    cell.value = computed[key]
                    cells_replaced += 1

    if cells_replaced == 0:
        # Nothing to write — workbook had formulas but none mapped to
        # computed values (e.g. all-empty leaves). Leave the file alone.
        return workbook_path

    # Atomic save: write to a sibling temp file, then replace. Avoids
    # leaving a half-written xlsx on disk if the process is killed
    # mid-save (which would otherwise corrupt the merged output).
    # Close the fd `mkstemp` opens — leaving it open leaks a descriptor
    # and on Windows holds an exclusive lock that can break `wb.save`.
    fd, tmp_str = tempfile.mkstemp(
        prefix=workbook_path.stem + ".",
        suffix=".recalc.xlsx",
        dir=str(workbook_path.parent),
    )
    os.close(fd)
    tmp_path = Path(tmp_str)
    try:
        wb.save(tmp_path)
        tmp_path.replace(workbook_path)
    except Exception:  # noqa: BLE001
        logger.warning(
            "recalc: atomic save failed for %s — original workbook intact",
            workbook_path, exc_info=True,
        )
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        return workbook_path

    logger.info(
        "recalc: replaced %d formula cell(s) with cached values in %s",
        cells_replaced, workbook_path,
    )
    return workbook_path
