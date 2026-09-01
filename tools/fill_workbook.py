import logging
import re
import unicodedata
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Annotated, Optional, Sequence, Union

import openpyxl
from pydantic import BaseModel, StringConstraints

from utils.workbook_io import atomic_save_workbook
from tools.guard_result import GuardResult
from tools.section_headers import (
    discover_section_headers,
    header_set,
    keyword_fallback_for_sheet,
)

logger = logging.getLogger(__name__)


class FactWrite(BaseModel):
    """One typed cell write proposed by the extraction agent.

    Phase 3 of the first-principles rewrite replaced the stringly-typed
    ``fields_json`` blob with this model so pydantic-ai validates each
    proposal (and injects the schema into the tool signature) BEFORE the
    tool body runs — malformed proposals never reach `fill_workbook`.

    Mirrors the real cell contract rather than the report's store-first
    ``(concept, period, scope)`` shape, because today's fill is still
    cell-based (Phase 4 is what flips render-last). Evidence is REQUIRED:
    the evidence column is the audit trail (gotcha #16) and making it a
    routed typed field kills the old silent evidence-column override.
    """

    sheet: str
    # 2 = CY (col B), 3 = PY (col C); group filings add D/E; the SOCIE
    # 24-col matrix uses the equity-component column directly.
    col: int = 2
    # PDF page + short quote. Required — never let a value land without
    # provenance (gotcha #16). `strip_whitespace=True` + `min_length=1` means a
    # whitespace-only string ("   ") is stripped to "" and then rejected, so
    # blank-after-trim evidence can't sneak past as "present".
    evidence: Annotated[str, StringConstraints(strip_whitespace=True, min_length=1)]
    # Label-matching mode (preferred): match against column-A text.
    field_label: str = ""
    # Section hint to disambiguate duplicate labels (e.g. "current" vs
    # "non-current").
    section: str = ""
    # Explicit-coordinate mode (SOCIE matrix and other complex layouts).
    row: Optional[int] = None
    # Numeric for data rows; str for the row-1 reporting-period date cells
    # (e.g. "01/01/2022 - 31/12/2022"). The legacy dataclass annotation was
    # `Optional[float]` but dataclasses don't validate, so date strings rode
    # through. pydantic DOES validate — keep the union so date cells still
    # work. int|float preserves integer-ness for the common numeric case.
    value: Optional[Union[int, float, str]] = None


@dataclass
class FieldMapping:
    sheet: str
    field_label: str
    col: int  # 2 = CY (column B), 3 = PY (column C)
    value: Optional[float]
    # Section hint for disambiguating duplicate labels (e.g. "current" vs "non-current")
    section: str = ""
    # Legacy row-based fallback
    row: Optional[int] = None
    evidence: str = ""


@dataclass
class FillResult:
    success: bool
    fields_written: int
    output_path: str
    errors: list[str]
    # RUN-REVIEW P1-1 (2026-04-26): non-fatal warnings surfaced to the
    # agent so it can self-correct double-booking before the next
    # verify_totals pass. Empty list when no concerns detected.
    warnings: list[str] = field(default_factory=list)
    # Canonical mode (Phase B): resolved cell coordinates that actually
    # landed, so the caller can project each into run_concept_facts. Each
    # entry is {sheet, row, col, value, evidence} with `row` already
    # label-resolved. Empty in legacy mode / on total failure.
    resolved_writes: list[dict] = field(default_factory=list)
    # Harness Item 2 follow-through: per-kind counts of GuardResult
    # refusals in this call (e.g. {"abstract_row": 2}) — the machine-
    # countable side of the verdicts whose messages land in `errors`.
    guard_rejections: dict[str, int] = field(default_factory=dict)
    # Stable request identities let the agent harness retain unresolved
    # partial-write errors across later unrelated calls, while clearing an
    # error when that same logical fact is successfully retried.
    successful_request_keys: list[dict[str, str]] = field(default_factory=list)
    failed_request_keys: list[dict[str, str]] = field(default_factory=list)


# Default SOCIE evidence column for the MFRS 24-col equity-component matrix.
# Lives just past Total (col X = 24); MFRS templates have no row-1 "Source"
# header so we fall back to this. MPERS templates declare a real Source
# header at col D / F and `_resolve_socie_evidence_col` honours that.
_DEFAULT_MATRIX_SOCIE_EVIDENCE_COL = 25


def _resolve_socie_evidence_col(ws) -> int:
    """Return the column where SOCIE evidence/source should be written.

    Looks for a row-1 cell whose text equals "Source" (case-insensitive).
    The MPERS Group/Company SOCIE templates publish this header at col D
    (4-col layout) and the MPERS Group SoRE template at col F. The MFRS
    matrix SOCIE templates carry no Source header — fall back to col Y
    (25) so existing MFRS behaviour is preserved.
    """
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if isinstance(value, str) and value.strip().lower() == "source":
            return col
    return _DEFAULT_MATRIX_SOCIE_EVIDENCE_COL


def _evidence_token_overlap(a: str, b: str) -> int:
    """Count distinct lowercase tokens (≥4 chars) shared by two evidence
    strings. Used by the double-booking guard (RUN-REVIEW P1-1) to decide
    whether two same-value writes really refer to the same disclosure.
    Short tokens are ignored because words like "of", "and", "the", "RM"
    overlap on every pair and would make the guard fire constantly.
    """
    if not a or not b:
        return 0
    norm = lambda s: {  # noqa: E731
        tok for tok in
        ''.join(c.lower() if c.isalnum() else ' ' for c in s).split()
        if len(tok) >= 4
    }
    return len(norm(a) & norm(b))


def _detect_double_bookings(
    label_index: dict[str, "list[_LabelEntry]"],
    written: list[FieldMapping],
    *,
    min_value: float = 1.0,
    overlap_threshold: int = 3,
) -> list[str]:
    """Return human-readable warnings about same-value/same-section writes.

    RUN-REVIEW §3.3-D: the Amway run wrote restoration provision PY 1,881
    onto BOTH row 287 (Provision for decommissioning…) and row 318 (Other
    non-current non-trade payables) in the same Non-current liabilities
    section. The face balance still passes because *Total non-current
    liabilities sums both — but the value is double-booked.

    The guard is intentionally narrow:

    * Same sheet, same column, same numeric value.
    * Evidence-string token overlap ≥ ``overlap_threshold`` distinct
      tokens of length ≥ 4. Disjoint evidence (two different
      disclosures that happen to round to the same number) does NOT
      trip the guard — that's a coincidence, not a double-book.
    * Tiny values (|val| < ``min_value``) are ignored — zeros and
      single-digit RM amounts coincide too often to be meaningful.

    Note we do NOT require both rows to share an exact section label.
    The Amway bug had row 287 (section "non-current provisions") and
    row 318 (section "non-current non-trade payables") — peer
    sub-sections under Non-current liabilities. Forcing exact-section
    match would silently miss this real failure mode. The
    evidence-overlap discriminator carries the load instead.

    The guard is column-scoped so legitimate consolidation pass-through
    on Group filings (same value in Group-CY col B AND Company-CY col
    D for the same row) does NOT trigger — that's by design, the
    discriminator is whether we're seeing two ROWS with the same value
    in ONE column.
    """
    warnings: list[str] = []
    if not written:
        return warnings

    # Group writes by (sheet, col) so we only compare within the same
    # column — Group consolidation pass-through (same value in Group +
    # Company columns) is legitimate and must not warn.
    by_col: dict[tuple[str, int], list[tuple[FieldMapping, "_LabelEntry"]]] = {}
    for m in written:
        if m.value is None:
            continue
        try:
            v = float(m.value)
        except (TypeError, ValueError):
            continue
        if abs(v) < min_value:
            continue
        sheet_entries = label_index.get(m.sheet, [])
        # Find the entry for the row we wrote to (already resolved during
        # the main loop; here we just look it up by mapping fields).
        entry = None
        for e in sheet_entries:
            if m.row is not None and e.row == m.row:
                entry = e
                break
            if m.field_label and e.normalized_label == _normalize_label(m.field_label):
                # Section-aware match when row coordinate isn't carried.
                if m.section and m.section.lower() not in e.section.lower():
                    continue
                entry = e
                break
        if entry is None or entry.is_header:
            continue
        by_col.setdefault((m.sheet, m.col), []).append((m, entry))

    for (sheet, col), items in by_col.items():
        # Find pairs sharing (value, overlapping-evidence) within this
        # (sheet, col). We do NOT gate on exact section match — the
        # canonical Amway bug straddled peer sub-sections, see the
        # docstring rationale above. Section labels are still surfaced
        # in the warning text so the agent has navigation context.
        for i, (m_a, e_a) in enumerate(items):
            for m_b, e_b in items[i + 1:]:
                if e_a.row == e_b.row:
                    continue
                # Cast through float so 1881 == 1881.0 etc.
                if abs(float(m_a.value) - float(m_b.value)) > 0.5:
                    continue
                overlap = _evidence_token_overlap(m_a.evidence, m_b.evidence)
                if overlap < overlap_threshold:
                    continue
                col_letter = openpyxl.utils.get_column_letter(col)
                same_section = e_a.section == e_b.section
                section_note = (
                    f"section '{e_a.section}'"
                    if same_section
                    else f"sections '{e_a.section}' / '{e_b.section}'"
                )
                warnings.append(
                    f"Possible double-booking on {sheet} col {col_letter} "
                    f"{section_note}: value {m_a.value} appears "
                    f"on row {e_a.row} ('{m_a.field_label or ''}') AND "
                    f"row {e_b.row} ('{m_b.field_label or ''}') with "
                    f"overlapping evidence ({overlap} shared token(s)). "
                    f"If both are correct, leave them; if one is the "
                    f"wrong row, remove it before the next verify_totals."
                )
    return warnings


def fill_workbook(
    template_path: str,
    output_path: str,
    facts: Sequence[Union["FactWrite", dict]],
    filing_level: str = "company",
    canonical_template_path: str | None = None,
) -> FillResult:
    """Apply typed cell writes to an Excel template.

    Matches fields by label (column A text) with section-aware disambiguation.
    When a label appears multiple times (e.g. "Lease liabilities" under both
    non-current and current), the section hint ("current"/"non-current") picks
    the correct occurrence.

    ``facts`` is a sequence of `FactWrite` models (the agent-facing contract
    validated by pydantic-ai before this body runs) or plain dicts with the
    same keys (the internal/test contract). Phase 3 of the rewrite removed the
    old ``fields_json`` string + its JSON-decode error branch — proposals are
    now typed end-to-end.

    ``template_path`` is the mutable workbook to read.  On incremental writes
    that is normally the agent's scratch workbook.  ``canonical_template_path``
    identifies the immutable managed template whose filing-target manifest is
    authoritative for writability; when omitted it defaults to
    ``template_path`` for existing callers and synthetic workbooks.
    """
    template = Path(template_path)
    if not template.exists():
        return FillResult(
            success=False,
            fields_written=0,
            output_path="",
            errors=[f"Template not found: {template_path}"],
        )

    mappings = _coerce_facts(facts)

    # The taxonomy/slot manifest is authoritative for writability. Workbook
    # colour remains useful for section disambiguation, but it cannot decide
    # whether an XBRL concept is allowed to carry a fact.
    from concept_model.filing_targets import (
        list_writable_targets,
        writable_coordinates,
    )

    canonical_template = Path(canonical_template_path or template_path)
    if not canonical_template.exists():
        return FillResult(
            success=False,
            fields_written=0,
            output_path="",
            errors=[f"Canonical template not found: {canonical_template}"],
        )

    writable_targets = list_writable_targets(canonical_template)
    writable_rows_by_sheet = {
        sheet: frozenset(
            target.row for target in writable_targets if target.sheet == sheet
        )
        for sheet in {target.sheet for target in writable_targets}
    }
    writable_coordinates_by_sheet = {
        sheet: writable_coordinates(canonical_template, sheet)
        for sheet in {target.sheet for target in writable_targets}
    }
    linear_value_columns = (
        frozenset({2, 3, 4, 5})
        if filing_level == "group"
        else frozenset({2, 3})
    )

    wb = openpyxl.load_workbook(template_path)
    errors: list[str] = []
    failed_mapping_messages: list[tuple[str, str, str]] = []

    def reject(mapping: FieldMapping, message: str) -> None:
        """Attach one deterministic refusal to its own request identity."""
        errors.append(message)
        key, base_key = _mapping_request_keys(mapping)
        failed_mapping_messages.append((key, base_key, message))

    guard_rejections: dict[str, int] = {}
    fields_written = 0
    # RUN-REVIEW P1-1: track successful writes so the post-loop double-
    # booking guard only sees mappings that actually landed (skipped
    # writes shouldn't raise spurious warnings).
    successful_writes: list[FieldMapping] = []
    successful_mapping_keys: list[tuple[str, str]] = []
    resolved_cells: dict[tuple[str, int, int], object] = {}

    # Build section-aware label index per sheet
    label_index = _build_label_index(wb)

    for mapping in mappings:
        if mapping.sheet not in wb.sheetnames:
            reject(mapping, f"Sheet '{mapping.sheet}' not found in template")
            continue

        ws = wb[mapping.sheet]

        # Resolve the target row: match by label first, fall back to explicit row
        target_row = None
        if mapping.field_label:
            allowed_rows = None
            if writable_targets:
                exact_coordinates = writable_coordinates_by_sheet.get(mapping.sheet)
                if exact_coordinates is None:
                    allowed_rows = writable_rows_by_sheet.get(
                        mapping.sheet, frozenset()
                    )
                else:
                    col_letter = openpyxl.utils.get_column_letter(mapping.col)
                    allowed_rows = {
                        row for row, col in exact_coordinates if col == col_letter
                    }
            resolution = _resolve_row_by_label(
                label_index.get(mapping.sheet, []),
                mapping.field_label,
                section_hint=mapping.section,
                allowed_rows=allowed_rows,
            )
            target_row = resolution.row
            if target_row is None:
                if resolution.error_kind == "ambiguous_label":
                    rows = ", ".join(str(row) for row in resolution.candidate_rows)
                    reject(
                        mapping,
                        f"Ambiguous label '{mapping.field_label}' in sheet "
                        f"'{mapping.sheet}' matches writable rows {rows}. "
                        "Supply the exact period/scope section from "
                        "read_template(); no row was selected.",
                    )
                    guard_rejections["ambiguous_label"] = (
                        guard_rejections.get("ambiguous_label", 0) + 1
                    )
                    continue
                msg = (
                    f"No matching label for '{mapping.field_label}'"
                    f"{f' (section: {mapping.section})' if mapping.section else ''}"
                    f" in sheet '{mapping.sheet}'."
                )
                redirect = _writable_label_in_other_sheets(
                    wb, label_index, mapping.sheet,
                    mapping.field_label, mapping.col,
                )
                if redirect is not None:
                    other_sheet, other_row = redirect
                    msg += (
                        f" It IS a writable cell at '{other_sheet}'!row"
                        f" {other_row} — write it there instead. This row"
                        f" has no equivalent on '{mapping.sheet}' (some face"
                        f" rows are direct data-entry with no sub-sheet line)."
                    )
                else:
                    msg += " Check the exact label text from read_template()."
                reject(mapping, msg)
                continue
        elif mapping.row is not None:
            target_row = mapping.row
            # Bug 5b — if the agent supplied a row coordinate (no field_label),
            # check that col A at that row actually has a label. The MPERS
            # SOCIE bug was exactly this: socie.md's MFRS-matrix instructions
            # told the agent to write at rows 30/35/49, which on the MPERS
            # Company template have NO label in col A — the writes landed on
            # blank cells silently. Row 1 is the documented carve-out for
            # date cells (see `prompts/_base.md`). Any other labelless row
            # means the agent is targeting a row that does not exist in the
            # current template.
            if target_row != 1:
                col_a_value = ws.cell(row=target_row, column=1).value
                if col_a_value is None or not str(col_a_value).strip():
                    # S-5: earlier wording said "this row does not exist in
                    # the loaded template" — technically wrong (the row
                    # exists, the LABEL is absent). The new phrasing points
                    # at the real fix: field_label matching, and cross-
                    # check against read_template if the agent believed
                    # the row was intentional.
                    reject(
                        mapping,
                        f"Refusing to write to {mapping.sheet} row {target_row}: "
                        f"col A is empty — this row has no label. Use "
                        f"field_label matching, or call read_template() to "
                        f"confirm the row is the one you intended.",
                    )
                    continue
        else:
            reject(mapping, f"Field has neither label nor row: {mapping}")
            continue

        cell = ws.cell(row=target_row, column=mapping.col)

        # Row 1 is the existing period-metadata carve-out. Every other write
        # must resolve to a reportable primary item on an INPUT slot.
        exact_coordinates = writable_coordinates_by_sheet.get(mapping.sheet)
        if exact_coordinates is None:
            coordinate_is_writable = (
                target_row
                in writable_rows_by_sheet.get(mapping.sheet, frozenset())
                and mapping.col in linear_value_columns
            )
        else:
            coordinate_is_writable = (
                target_row,
                openpyxl.utils.get_column_letter(mapping.col),
            ) in exact_coordinates
        if writable_targets and target_row != 1 and not coordinate_is_writable:
            label_text = ws.cell(row=target_row, column=1).value
            row_is_writable = target_row in writable_rows_by_sheet.get(
                mapping.sheet, frozenset()
            )
            if row_is_writable:
                column_letter = openpyxl.utils.get_column_letter(mapping.col)
                if exact_coordinates is None:
                    allowed_columns = [
                        openpyxl.utils.get_column_letter(col)
                        for col in sorted(linear_value_columns)
                    ]
                else:
                    allowed_columns = sorted(
                        col for row, col in exact_coordinates if row == target_row
                    )
                allowed_text = ", ".join(allowed_columns) or "none"
                message = (
                    f"Refusing to write to {mapping.sheet}!{cell.coordinate}: "
                    f"row {target_row} ('{label_text}') is writable, but column "
                    f"{column_letter} is a non-entry template column for this "
                    f"row. Use a value column shown by read_template() "
                    f"({allowed_text})."
                )
            else:
                message = (
                    f"Refusing to write to {mapping.sheet}!{cell.coordinate}: "
                    f"row {target_row} ('{label_text}') is a heading, formula, "
                    "or other non-entry template row. Choose a writable "
                    "canonical field from read_template()."
                )
            verdict = GuardResult.retry(
                message,
                kind="non_writable_template_slot",
            )
            reject(mapping, verdict.message)
            guard_rejections[verdict.kind] = guard_rejections.get(verdict.kind, 0) + 1
            continue

        # Never overwrite formula cells
        if cell.value is not None and str(cell.value).startswith("="):
            verdict = GuardResult.retry(
                f"Refusing to overwrite formula cell {mapping.sheet}!{cell.coordinate}: {cell.value}",
                kind="formula_cell",
            )
            reject(mapping, verdict.message)
            guard_rejections[verdict.kind] = guard_rejections.get(verdict.kind, 0) + 1
            logger.warning(
                "fill_workbook guard rejected (%s): %s!%s",
                verdict.kind, mapping.sheet, cell.coordinate,
            )
            continue

        # Bug A (2026-04-26): refuse writes to abstract section-header rows.
        # The screenshot bug on SOPL-Analysis-Function had the agent writing
        # 6,092 onto the dark-navy "Interest income" row instead of the
        # leaves below — the formula-driven "Total interest income" then
        # evaluated to 0 because the leaves were empty. The header is an
        # XBRL abstract concept, never a data target. We look the row up
        # in the same `label_index` we just built so the check stays cheap
        # and consistent with `_find_row_by_label`'s leaf-preference logic.
        sheet_entries = label_index.get(mapping.sheet, [])
        target_entry = next(
            (e for e in sheet_entries if e.row == target_row),
            None,
        )
        if target_entry is not None and target_entry.is_header:
            label_text = ws.cell(row=target_row, column=1).value
            verdict = GuardResult.retry(
                f"Refusing to write to {mapping.sheet}!{cell.coordinate}: "
                f"row {target_row} ('{label_text}') is an XBRL abstract "
                f"section header, not a data-entry cell. Write to a leaf "
                f"row under it (call read_template() and look for non-"
                f"[ABSTRACT] rows in this section), or roll the value up "
                f"into the nearest matching leaf. Never plug a residual "
                f"into a catch-all to make totals reconcile.",
                kind="abstract_row",
            )
            reject(mapping, verdict.message)
            guard_rejections[verdict.kind] = guard_rejections.get(verdict.kind, 0) + 1
            logger.warning(
                "fill_workbook guard rejected (%s): %s!%s",
                verdict.kind, mapping.sheet, cell.coordinate,
            )
            continue

        target_key = (mapping.sheet, target_row, mapping.col)
        if target_key in resolved_cells:
            prior_value = resolved_cells[target_key]
            if prior_value != mapping.value:
                verdict = GuardResult.retry(
                    f"Conflicting writes in one request for "
                    f"{mapping.sheet}!{cell.coordinate}: {prior_value!r} and "
                    f"{mapping.value!r}. Submit one grounded value for the "
                    "physical slot; the later value was not written.",
                    kind="conflicting_write",
                )
                reject(mapping, verdict.message)
                guard_rejections[verdict.kind] = (
                    guard_rejections.get(verdict.kind, 0) + 1
                )
            # An identical duplicate is an idempotent no-op, not a second
            # field write or a second canonical projection.
            if prior_value == mapping.value:
                successful_mapping_keys.append(_mapping_request_keys(mapping))
            continue

        cell.value = mapping.value
        resolved_cells[target_key] = mapping.value
        fields_written += 1
        # Stash a copy with the resolved row coordinate so the double-
        # booking guard doesn't have to redo label/section matching.
        successful_writes.append(FieldMapping(
            sheet=mapping.sheet,
            field_label=mapping.field_label,
            col=mapping.col,
            value=mapping.value,
            section=mapping.section,
            row=target_row,
            evidence=mapping.evidence,
        ))
        successful_mapping_keys.append(_mapping_request_keys(mapping))

        # Write evidence/source to a single column per sheet so notes don't repeat.
        #
        # SOCIE sheets historically used col Y (25) because the MFRS template is
        # a 24-col equity-component matrix with no Source header. MPERS SOCIE
        # templates publish a real Source header at col D (Company) or F (Group),
        # so writing to col 25 there hides the audit trail off-screen and leaves
        # the visible Source column empty (peer-review H2). For SOCIE sheets we
        # now look up the Source header by name and only fall back to 25 when
        # no header is found (MFRS matrix layouts). Other sheets keep the
        # filing-level branch as before.
        if mapping.evidence:
            if "socie" in mapping.sheet.lower():
                evidence_col = _resolve_socie_evidence_col(ws)
            elif filing_level == "group":
                evidence_col = 6  # F — after Company PY (E=5)
            else:
                evidence_col = 4  # D — after PY (C=3)
            evidence_cell = ws.cell(row=target_row, column=evidence_col)
            # Always overwrite evidence so correction passes don't accumulate
            # stale provenance from values that were later replaced.
            evidence_cell.value = mapping.evidence

    # Item 8 / gotcha #22: atomic save — pydantic-ai runs batched tool calls
    # concurrently, so a reader hitting this path mid-save must see
    # old-or-new, never a truncated zip.
    atomic_save_workbook(wb, output_path)
    wb.close()

    # RUN-REVIEW P1-1: scan successful writes for double-bookings now
    # that the workbook is closed. Warnings are advisory — they don't
    # flip success to False — but they bubble up to the agent so it
    # can decide whether to keep both rows or remove one before the
    # next verify_totals pass.
    warnings = _detect_double_bookings(label_index, successful_writes)

    resolved_writes = [
        {
            "sheet": w.sheet,
            "row": w.row,
            "col": w.col,
            "value": w.value,
            "evidence": w.evidence,
        }
        for w in successful_writes
        if w.row is not None
    ]

    successful_request_keys = [
        {"key": key, "base_key": base_key}
        for key, base_key in successful_mapping_keys
    ]
    failed_request_keys = [
        {"key": key, "base_key": base_key, "message": message}
        for key, base_key, message in failed_mapping_messages
    ]

    if errors and fields_written == 0:
        return FillResult(
            success=False,
            fields_written=0,
            output_path=output_path,
            errors=errors,
            warnings=warnings,
            guard_rejections=guard_rejections,
            successful_request_keys=successful_request_keys,
            failed_request_keys=failed_request_keys,
        )

    return FillResult(
        success=True,
        fields_written=fields_written,
        output_path=output_path,
        errors=errors,
        warnings=warnings,
        resolved_writes=resolved_writes,
        guard_rejections=guard_rejections,
        successful_request_keys=successful_request_keys,
        failed_request_keys=failed_request_keys,
    )


@dataclass
class _LabelEntry:
    """A label in the template with its row and the section it belongs to."""
    normalized_label: str
    row: int
    section: str  # e.g. "non-current assets", "current liabilities"
    # Ordered enclosing headers. Group SOCIE needs both the outer
    # period/scope block and the nested subsection (for example
    # "group - prior period" -> "comprehensive income").
    section_path: tuple[str, ...] = ()
    # Bug A (2026-04-26): True when this label is itself a section-header
    # (XBRL-abstract) row. Used by `_find_row_by_label` to prefer leaves
    # over headers on duplicate labels, and by the writer to refuse writes
    # whose target lands on an abstract row.
    is_header: bool = False


# Keyword fallback registry now lives in `tools.section_headers` so the
# reader (template_reader) and the writer share one source of truth — see
# peer-review #1 (2026-04-26) and `keyword_fallback_for_sheet`.


def _build_label_index(wb: openpyxl.Workbook) -> dict[str, list[_LabelEntry]]:
    """Build a section-aware label index per sheet.

    Walks column A top-to-bottom, tracking which section we're in based on
    header rows. Each label is stored with its section context.
    """
    index: dict[str, list[_LabelEntry]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        entries: list[_LabelEntry] = []
        current_section = ""
        current_block = ""

        # Detect header rows by row index (not label string). The legacy
        # form returned a set of normalised labels, which mis-marked any
        # leaf with the same text as a header — that was itself part of
        # the SOPL-Analysis duplicate-label bug. Keyword fallback selection
        # is shared with template_reader via section_headers — see
        # peer-review #1 (2026-04-26).
        fallback = keyword_fallback_for_sheet(name)
        header_rows = {
            h.row for h in discover_section_headers(ws, extra_keywords=fallback)
        }

        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val is None:
                continue

            normalized = _normalize_label(str(cell_val))
            is_header = row in header_rows
            # Section transitions: every header switches the running
            # subsection. MPERS Group SOCIE block headers are outer context,
            # so a nested "Comprehensive income" header must not erase the
            # period/scope identity needed to route repeated labels.
            if is_header:
                if _is_socie_block_section(normalized):
                    current_block = normalized
                    current_section = ""
                else:
                    current_section = normalized

            section_path = tuple(
                part for part in (current_block, current_section) if part
            )

            entries.append(_LabelEntry(
                normalized_label=normalized,
                row=row,
                section=current_section or current_block,
                section_path=section_path,
                is_header=is_header,
            ))

        index[name] = entries
    return index


def _normalize_label(label: str) -> str:
    """Normalize harmless Unicode/spacing variants without weakening identity."""
    normalized = unicodedata.normalize("NFKC", label)
    normalized = re.sub(r"[\u2010-\u2015\u2212]", "-", normalized)
    normalized = " ".join(normalized.strip().lstrip("*").strip().split())
    return normalized.casefold()


def _mapping_request_keys(mapping: FieldMapping) -> tuple[str, str]:
    """Return ``(scoped, base)`` identities for unresolved-write tracking."""
    sheet = _normalize_label(mapping.sheet)
    if mapping.field_label:
        locator = f"label={_normalize_label(mapping.field_label)}"
    else:
        locator = f"row={mapping.row}"
    base = f"{sheet}|col={mapping.col}|{locator}"
    section = _normalize_label(mapping.section) if mapping.section else ""
    return (f"{base}|section={section}", base)


_SOCIE_BLOCK_SECTION_RE = re.compile(
    r"^(group|company)\s*-\s*(current|prior)\s+period$",
)


def _is_socie_block_section(normalized: str) -> bool:
    return bool(_SOCIE_BLOCK_SECTION_RE.fullmatch(normalized))


def _entry_matches_section(entry: _LabelEntry, section_hint: str) -> bool:
    hint = _normalize_label(section_hint)
    if not hint:
        return False
    sections = entry.section_path or ((entry.section,) if entry.section else ())
    if any(section == hint for section in sections):
        return True
    if any(section.startswith(hint) or hint.startswith(section) for section in sections):
        return True

    hint_has_current = "current" in hint
    hint_has_noncurrent = "non-current" in hint
    if hint_has_noncurrent:
        return any("non-current" in section for section in sections)
    if hint_has_current:
        return any(
            "current" in section and "non-current" not in section
            for section in sections
        )
    return False


@dataclass(frozen=True)
class _LabelResolution:
    row: Optional[int]
    error_kind: Optional[str] = None
    candidate_rows: tuple[int, ...] = ()


def _resolve_row_by_label(
    entries: list[_LabelEntry],
    field_label: str,
    section_hint: str = "",
    threshold: float = 0.7,
    allowed_rows: Optional[set[int]] = None,
) -> _LabelResolution:
    """Resolve one label, failing closed when context cannot make it unique."""
    normalized = _normalize_label(field_label)

    exact_matches = [e for e in entries if e.normalized_label == normalized]
    if exact_matches and any(not e.is_header for e in exact_matches):
        exact_matches = [e for e in exact_matches if not e.is_header]
    if allowed_rows is not None:
        writable_exact = [e for e in exact_matches if e.row in allowed_rows]
        if writable_exact:
            exact_matches = writable_exact

    if exact_matches:
        if len(exact_matches) == 1:
            # A section hint is advisory when the label itself is unique.
            return _LabelResolution(row=exact_matches[0].row)
        if section_hint:
            filtered = [
                entry for entry in exact_matches
                if _entry_matches_section(entry, section_hint)
            ]
            if len(filtered) == 1:
                return _LabelResolution(row=filtered[0].row)
            if filtered:
                exact_matches = filtered
        return _LabelResolution(
            row=None,
            error_kind="ambiguous_label",
            candidate_rows=tuple(entry.row for entry in exact_matches),
        )

    candidates = [entry for entry in entries if not entry.is_header]
    if allowed_rows is not None:
        writable_candidates = [entry for entry in candidates if entry.row in allowed_rows]
        if writable_candidates:
            candidates = writable_candidates
    if section_hint:
        section_candidates = [
            entry for entry in candidates
            if _entry_matches_section(entry, section_hint)
        ]
        if section_candidates:
            candidates = section_candidates

    scored = sorted(
        [
            (
                SequenceMatcher(None, normalized, entry.normalized_label).ratio(),
                entry,
            )
            for entry in candidates
        ],
        key=lambda item: item[0],
        reverse=True,
    )
    if not scored or scored[0][0] < threshold:
        return _LabelResolution(row=None, error_kind="unknown_label")
    best_score = scored[0][0]
    near_best = [entry for score, entry in scored if best_score - score < 0.05]
    if len(near_best) != 1:
        return _LabelResolution(
            row=None,
            error_kind="ambiguous_label",
            candidate_rows=tuple(entry.row for entry in near_best),
        )
    return _LabelResolution(row=near_best[0].row)


def _writable_label_in_other_sheets(
    wb: openpyxl.Workbook,
    label_index: dict[str, list[_LabelEntry]],
    current_sheet: str,
    field_label: str,
    col: int,
) -> Optional[tuple[str, int]]:
    """Find `field_label` as a writable leaf on a sheet other than `current_sheet`.

    Returns ``(sheet, row)`` for the first non-header occurrence whose cell at
    `col` is empty or non-formula (i.e. a real data-entry target), else None.

    Motivated by the 2026-06-15 SOFP OrderOfLiquidity incident: the agent
    tried to write a non-derivative FVTPL unit-trust value to the SUB-sheet,
    where "Investments other than investments accounted for using equity
    method" does not exist — the row is a direct DATA_ENTRY leaf on the FACE
    sheet (row 16, no sub-sheet equivalent). The bare "check the label text"
    error sent the agent in circles until it gave up with a flagged imbalance
    equal to exactly the FVTPL amount. Exact-match only — a fuzzy redirect
    could send the agent to the wrong sheet, which is worse than no hint.
    """
    normalized = _normalize_label(field_label)
    for sheet_name, entries in label_index.items():
        if sheet_name == current_sheet or sheet_name not in wb.sheetnames:
            continue
        for entry in entries:
            if entry.normalized_label != normalized or entry.is_header:
                continue
            cell = wb[sheet_name].cell(row=entry.row, column=col)
            if cell.value is None or not str(cell.value).startswith("="):
                return (sheet_name, entry.row)
    return None


def _find_row_by_label(
    entries: list[_LabelEntry],
    field_label: str,
    section_hint: str = "",
    threshold: float = 0.7,
) -> Optional[int]:
    """Find the best matching row for a field label.

    When section_hint is provided, it disambiguates duplicate labels against
    their full enclosing section path. Unresolved duplicates return ``None``.
    """
    return _resolve_row_by_label(
        entries,
        field_label,
        section_hint=section_hint,
        threshold=threshold,
    ).row


def _coerce_facts(facts: Sequence[Union["FactWrite", dict]]) -> list[FieldMapping]:
    """Normalise the typed `facts` argument into internal `FieldMapping`s.

    Accepts `FactWrite` models (the agent path, already validated by
    pydantic-ai) or plain dicts (internal callers + tests). The internal
    loop, the double-booking guard, and the canonical projection all work
    in terms of `FieldMapping`, so we convert once here.
    """
    mappings: list[FieldMapping] = []
    for f in facts:
        if isinstance(f, FactWrite):
            mappings.append(
                FieldMapping(
                    sheet=f.sheet,
                    field_label=f.field_label,
                    col=f.col,
                    value=f.value,
                    section=f.section,
                    row=f.row,
                    evidence=f.evidence,
                )
            )
        elif isinstance(f, dict):
            mappings.append(
                FieldMapping(
                    sheet=f["sheet"],
                    field_label=f.get("field_label", ""),
                    col=int(f.get("col", 2)),
                    value=f.get("value"),
                    section=f.get("section") or "",
                    row=int(f["row"]) if f.get("row") is not None else None,
                    evidence=f.get("evidence", ""),
                )
            )
        else:
            raise TypeError(
                f"fill_workbook: unsupported fact entry {type(f).__name__}; "
                "expected FactWrite or dict"
            )
    return mappings
