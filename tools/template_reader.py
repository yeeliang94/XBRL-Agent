from dataclasses import dataclass
from typing import Optional

import openpyxl

# Reader symmetry with the writer: both call discover_section_headers with
# the same keyword fallback, so a row classified as abstract here is also
# treated as one by fill_workbook (peer-review #1, 2026-04-26). Without
# matching keyword fallbacks, MPERS Group SOCIE block dividers and SOFP
# sub-sheet sub-section dividers would surface as [DATA_ENTRY] to the agent
# but be refused at write time — confusing.
from tools.section_headers import (
    discover_section_headers,
    keyword_fallback_for_sheet,
)


@dataclass
class TemplateField:
    sheet: str
    coordinate: str
    row: int
    col: int
    value: Optional[str]
    formula: Optional[str]
    has_formula: bool
    is_data_entry: bool
    # Bug A: True when this cell sits on an XBRL abstract section-header row.
    # Set only on the col-A label cell (cols B/C/D on header rows are None
    # and therefore never appear in the field list). Agents must not write
    # to these rows — fill_workbook refuses such writes.
    is_abstract: bool = False

    @property
    def label(self) -> str:
        return self.value or self.formula or ""


def read_template(path: str, sheet: Optional[str] = None) -> list[TemplateField]:
    wb = openpyxl.load_workbook(path, data_only=False)
    sheet_names = [sheet] if sheet else wb.sheetnames
    fields: list[TemplateField] = []

    for name in sheet_names:
        ws = wb[name]
        # Same detector + same fallback keywords as fill_workbook, so the
        # agent's read_template summary marks exactly the rows the writer
        # will refuse — no spurious [DATA_ENTRY] surprises.
        fallback = keyword_fallback_for_sheet(name)
        abstract_rows = {
            h.row for h in discover_section_headers(ws, extra_keywords=fallback)
        }

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                formula = cell.value if is_formula else None
                value = cell.value if not is_formula else None

                # Mark the col-A label cell on every abstract row. Cells
                # outside col A are not flagged — partly because header rows
                # don't carry B/C/D values in the templates, and partly to
                # keep the flag's semantics row-level rather than cell-level.
                is_abstract = (cell.column == 1 and cell.row in abstract_rows)

                fields.append(
                    TemplateField(
                        sheet=name,
                        coordinate=cell.coordinate,
                        row=cell.row,
                        col=cell.column,
                        value=str(value) if value is not None else None,
                        formula=formula,
                        has_formula=is_formula,
                        is_data_entry=not is_formula,
                        is_abstract=is_abstract,
                    )
                )

    return fields
