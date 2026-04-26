import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.formula.tokenizer import Tokenizer

# Forward reference: statement_types imports are kept local to the dispatch
# function to avoid a circular import if verifier ever grows into a package.


# Peer-review #2 (2026-04-26): factor the imbalance feedback so every
# branch (CY, PY, Group Company-CY, Group Company-PY) gets the diagnostic
# direction marker AND the no-plug guidance — the original Bug B fix
# only updated the CY branch, leaving the other three with the legacy
# bare line and no anti-plug guard.
def _sofp_imbalance_feedback(period_label: str, diff: float) -> list[str]:
    """Return the IMBALANCE + diagnostic + no-plug lines for a SOFP period.

    `period_label` is what appears in the IMBALANCE prefix — "CY", "PY",
    "Company CY", or "Company PY". `diff = assets - (equity+liabilities)`,
    so a positive sign means assets exceeds equity+liabilities (something
    missing on the equity/liabilities side, or assets carries an extra
    value); a negative sign is the mirror image. The legacy phrasing
    inverted the negative-diff direction (peer-review 2026-04-26) — the
    correct reading is that assets is LOWER, not higher.
    """
    direction = (
        "equity+liabilities section is lower than assets, or assets "
        "carries an extra value"
        if diff > 0
        else "assets section is lower than equity+liabilities, or "
             "equity+liabilities carries an extra value"
    )
    return [
        f"IMBALANCE ({period_label}): assets - (equity+liabilities) = {diff}",
        # Diagnostic + no-plug guidance. Wording deliberately avoids the
        # legacy "Action: re-examine X side" framing because, paired with
        # the save-gate's hard block, that pushed the agent toward
        # plugging a residual into a catch-all row to satisfy the gate.
        f"Diagnostic: {direction}. Re-examine the relevant notes for any "
        f"sub-items you may have missed. Do NOT plug a catch-all row "
        f"('Other …', 'Other miscellaneous …', similar) to balance — if "
        f"the discrepancy persists, leave the leaves untouched and "
        f"finish honestly with the gap flagged."
    ]


# Footer appended to every non-SOFP verifier's feedback when imbalanced,
# so SOPL/SOCI/SOCF/SOCIE attribution and balance failures also reach the
# agent with the no-plug guard. SOFP uses `_sofp_imbalance_feedback` per
# branch instead because it has the assets-vs-equity directional cue.
_NO_PLUG_FOOTER = (
    "Reminder: do NOT plug a catch-all row ('Other …', 'Other "
    "miscellaneous …', similar) to absorb the discrepancy. If you cannot "
    "locate the missing component in the notes, leave the leaves "
    "untouched and finish honestly with the gap flagged."
)


def _compose_feedback(
    mismatches: list[str],
    is_balanced: Optional[bool],
    passed_message: str,
) -> str:
    """Build the per-statement feedback string with the no-plug footer
    appended whenever the statement is unbalanced. Used by SOCIE / SOCF /
    SOPL / SOCI verifiers (SOFP injects no-plug guidance per branch via
    `_sofp_imbalance_feedback`)."""
    if not mismatches:
        return passed_message
    lines = list(mismatches)
    if is_balanced is False:
        lines.append(_NO_PLUG_FOOTER)
    return "\n".join(lines)


@dataclass
class VerificationResult:
    # `is_balanced` is True/False for statements where the concept applies
    # (currently SOFP) and None when the check is not applicable to the
    # statement type (e.g. SOPL has no balance identity to verify).
    is_balanced: Optional[bool]
    # True/False when PDF reference values were compared; None when no
    # comparison was performed (e.g. non-SOFP statements without per-statement
    # comparison maps). Consumers must not treat None as "passed".
    matches_pdf: Optional[bool]
    computed_totals: dict[str, float] = field(default_factory=dict)
    pdf_values: dict[str, float] = field(default_factory=dict)
    mismatches: list[str] = field(default_factory=list)
    feedback: str = ""
    # Phase 1.1: labels whose mandatory ('*') rows were not filled. Populated
    # by `_collect_unfilled_mandatory` from each statement verifier so
    # `verify_totals` feedback can route the agent back to the gap instead
    # of silently shipping a blank cell. Empty list = no gaps.
    mandatory_unfilled: list[str] = field(default_factory=list)


# The exact labels we look for in the SOFP main sheet to verify balance
_TOTAL_ASSETS_LABEL = "*total assets"
_TOTAL_EQ_LIAB_LABEL = "*total equity and liabilities"
# Target the main SOFP sheet — falls back to active sheet if not found
_SOFP_SHEET_NAME = "SOFP-CuNonCu"


def _resolve_cell_value(
    wb: openpyxl.Workbook,
    sheet_name: str,
    cell_ref: str,
    visited: Optional[set[str]] = None,
    warnings: Optional[list[str]] = None,
) -> float:
    """Resolve a cell's value, recursing through formulas with cycle detection."""
    if visited is None:
        visited = set()

    key = f"{sheet_name}!{cell_ref}"
    if key in visited:
        return 0.0  # cycle — break it
    visited.add(key)

    try:
        raw = wb[sheet_name][cell_ref].value
    except KeyError:
        if warnings is not None:
            warnings.append(f"Missing reference: sheet '{sheet_name}' or cell {cell_ref} not found")
        return 0.0

    if raw is None:
        return 0.0

    if isinstance(raw, str) and raw.startswith("="):
        return _evaluate_formula(wb, sheet_name, raw, visited, warnings)

    try:
        return float(raw)
    except (ValueError, TypeError):
        if warnings is not None:
            warnings.append(f"Unparseable value in {sheet_name}!{cell_ref}: {raw!r}")
        return 0.0


def _expand_range(range_ref: str) -> list[str]:
    """Expand a cell range like E6:L6 into individual cell references.

    Handles single-row ranges (E6:L6), single-column ranges (B3:B10),
    and rectangular ranges (A1:C3).
    """
    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)$", range_ref)
    if not match:
        return [range_ref]

    col1_str, row1_str, col2_str, row2_str = match.groups()
    row1, row2 = int(row1_str), int(row2_str)

    # Convert column letters to numbers (A=1, B=2, ..., Z=26, AA=27, ...)
    def col_to_num(s: str) -> int:
        n = 0
        for ch in s:
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n

    def num_to_col(n: int) -> str:
        result = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            result = chr(65 + remainder) + result
        return result

    c1, c2 = col_to_num(col1_str), col_to_num(col2_str)
    cells = []
    for r in range(min(row1, row2), max(row1, row2) + 1):
        for c in range(min(c1, c2), max(c1, c2) + 1):
            cells.append(f"{num_to_col(c)}{r}")
    return cells


def _parse_range_operand(value: str) -> tuple[Optional[str], str]:
    """Split a Tokenizer RANGE operand into (sheet_name, cell_or_range).

    Examples:
        "'SOFP-Sub'!B39"  -> ("SOFP-Sub", "B39")
        "Sheet!B39:B40"   -> ("Sheet", "B39:B40")
        "B39"             -> (None, "B39")
    """
    if "!" not in value:
        return None, value
    sheet_part, cell_part = value.rsplit("!", 1)
    sheet_part = sheet_part.strip()
    # Excel wraps sheet names with spaces / punctuation in single quotes.
    if sheet_part.startswith("'") and sheet_part.endswith("'"):
        sheet_part = sheet_part[1:-1].replace("''", "'")
    return sheet_part, cell_part


def _sum_range_operand(
    wb: openpyxl.Workbook,
    default_sheet: str,
    operand: str,
    visited: set[str],
    warnings: Optional[list[str]],
) -> float:
    """Resolve a RANGE operand (single cell, A1:C3 range, or sheet-qualified
    form) and return the sum of the referenced cell values."""
    sheet, cell_part = _parse_range_operand(operand)
    sheet_name = sheet if sheet is not None else default_sheet
    total = 0.0
    if ":" in cell_part:
        for c in _expand_range(cell_part):
            total += _resolve_cell_value(wb, sheet_name, c, visited, warnings)
    else:
        total += _resolve_cell_value(wb, sheet_name, cell_part, visited, warnings)
    return total


# Functions the evaluator supports at the top level. Anything else produces
# a formula warning and a sentinel value — we refuse to guess.
_SUPPORTED_FUNCTIONS = {"SUM("}


def _evaluate_formula(
    wb: openpyxl.Workbook,
    sheet_name: str,
    formula: str,
    visited: Optional[set[str]] = None,
    warnings: Optional[list[str]] = None,
) -> float:
    """Evaluate a cell formula, recursing into referenced cells.

    Uses openpyxl's Tokenizer for a real tokenizer-driven walk (peer-review
    fix for C3). Supported constructs:

      * Cross-sheet cell/range references: ='SOFP-Sub'!B39, ='S'!A1:B2
      * Signed sums of references, including with coefficient form:
          =B139+B140-B141         → +B139 +B140 -B141
          =1*B139+1*B140-1*B141   → same, accepted for template compatibility
      * Top-level SUM(): =SUM(E6:L6) or =SUM(A1,B2,C3:C5)

    Anything outside that grammar — other functions, `#REF!`, division,
    array formulas — emits a warning and returns 0.0. Do NOT extend the
    grammar by guessing; if a real template needs more, add the case
    explicitly with a test.
    """
    if visited is None:
        visited = set()

    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return 0.0

    try:
        tokens = list(Tokenizer(formula).items)
    except Exception as exc:
        if warnings is not None:
            warnings.append(f"Could not tokenize formula {formula!r}: {exc}")
        return 0.0

    if not tokens:
        return 0.0

    # Special case: a whole formula that is just SUM(args) at the top level.
    # The body between SUM( and the matching ) is a comma-separated arg list.
    first = tokens[0]
    last = tokens[-1]
    if (
        first.type == "FUNC" and first.subtype == "OPEN"
        and first.value.upper() == "SUM("
        and last.type == "FUNC" and last.subtype == "CLOSE"
    ):
        total = 0.0
        inner = tokens[1:-1]
        # Walk args, split on SEP/ARG tokens at depth 0. We only accept pure
        # range operands inside SUM; any nested call or operator is refused.
        current_arg: list = []
        depth = 0

        def flush_arg(arg_tokens: list) -> Optional[float]:
            if not arg_tokens:
                return 0.0
            if len(arg_tokens) != 1 or arg_tokens[0].type != "OPERAND" or arg_tokens[0].subtype != "RANGE":
                if warnings is not None:
                    warnings.append(
                        f"SUM argument not a plain range: {[t.value for t in arg_tokens]!r}"
                    )
                return None
            return _sum_range_operand(
                wb, sheet_name, arg_tokens[0].value, visited, warnings
            )

        for tok in inner:
            # Skip Tokenizer whitespace tokens — Excel allows spaces after
            # commas in multi-arg SUM (e.g. `=SUM(A1, B2)`), and they
            # would otherwise poison the arg shape check in flush_arg.
            if tok.type == "WHITE-SPACE":
                continue
            if tok.type == "FUNC" and tok.subtype == "OPEN":
                depth += 1
                current_arg.append(tok)
            elif tok.type == "FUNC" and tok.subtype == "CLOSE":
                depth -= 1
                current_arg.append(tok)
            elif tok.type == "SEP" and tok.subtype == "ARG" and depth == 0:
                val = flush_arg(current_arg)
                if val is None:
                    return 0.0
                total += val
                current_arg = []
            else:
                current_arg.append(tok)
        val = flush_arg(current_arg)
        if val is None:
            return 0.0
        total += val
        return total

    # General case: signed sum of (coefficient × range) terms. Walk tokens
    # with a small state machine. Any unsupported token produces a warning
    # and a sentinel; we never silently ignore.
    total = 0.0
    sign = 1
    pending_coeff: Optional[float] = None

    for tok in tokens:
        # Excel tokenizes whitespace as its own token; skip without warning.
        # Formulas like `=B1 + B2` are otherwise valid.
        if tok.type == "WHITE-SPACE":
            continue
        if tok.type == "OPERATOR-INFIX":
            if tok.value == "+":
                if pending_coeff is not None:
                    # Standalone numeric term, e.g. =5+B1 — add it.
                    total += sign * pending_coeff
                    pending_coeff = None
                sign = 1
            elif tok.value == "-":
                if pending_coeff is not None:
                    total += sign * pending_coeff
                    pending_coeff = None
                sign = -1
            elif tok.value == "*":
                # Expect the next OPERAND to multiply pending_coeff by.
                # pending_coeff must already be set from the previous NUMBER.
                continue
            else:
                if warnings is not None:
                    warnings.append(
                        f"Unsupported operator {tok.value!r} in {formula!r}"
                    )
                return 0.0
        elif tok.type == "OPERATOR-PREFIX":
            # Unary prefix operators appear in two situations we care about:
            # (1) at the start of a formula like `=-1*B1+B2`,
            # (2) right after a binary `+`/`-`, as in `=1*B7+-1*B11` —
            # Excel's tokenizer splits `+-1` into INFIX `+` followed by
            # PREFIX `-` rather than a single negative-coefficient token.
            # The XBRL cashflow linkbase emits exactly this pattern, so
            # without this branch every statement with a negative-signed
            # coefficient row (SOCF, SOCIE, SOPL, SOFP equity) silently
            # evaluated to 0 and masked real imbalances.
            if tok.value == "-":
                sign = -sign
            elif tok.value == "+":
                # No-op: `+-1*B3` == `-1*B3`, `++B3` == `+B3`.
                pass
            else:
                if warnings is not None:
                    warnings.append(
                        f"Unsupported prefix operator {tok.value!r} in {formula!r}"
                    )
                return 0.0
        elif tok.type == "OPERAND" and tok.subtype == "NUMBER":
            try:
                pending_coeff = float(tok.value)
            except ValueError:
                if warnings is not None:
                    warnings.append(f"Unparseable number {tok.value!r} in {formula!r}")
                return 0.0
        elif tok.type == "OPERAND" and tok.subtype == "RANGE":
            val = _sum_range_operand(wb, sheet_name, tok.value, visited, warnings)
            coeff = pending_coeff if pending_coeff is not None else 1.0
            total += sign * coeff * val
            pending_coeff = None
        elif tok.type == "OPERAND" and tok.subtype == "ERROR":
            # Tokens like #REF!, #DIV/0! — refuse to evaluate.
            if warnings is not None:
                warnings.append(f"Error token {tok.value!r} in {formula!r}")
            return 0.0
        elif tok.type == "FUNC" and tok.subtype == "OPEN":
            # A top-level SUM was handled above; any other function (or SUM
            # embedded mid-expression) is out of scope.
            if warnings is not None:
                warnings.append(
                    f"Unsupported function {tok.value!r} in {formula!r}"
                )
            return 0.0
        else:
            # Whitespace / parens / other — treat as unsupported to stay safe.
            if warnings is not None:
                warnings.append(
                    f"Unsupported token {tok.type}/{tok.subtype} {tok.value!r} in {formula!r}"
                )
            return 0.0

    # Flush any trailing standalone numeric term.
    if pending_coeff is not None:
        total += sign * pending_coeff

    return total


def verify_totals(
    path: str,
    pdf_values: Optional[dict[str, float]] = None,
    filing_level: str = "company",
) -> VerificationResult:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Template not found: {path}")

    # Load with data_only=False to access formulas for evaluation
    wb = openpyxl.load_workbook(path, data_only=False)

    # Target the SOFP main sheet explicitly; fall back to active sheet
    if _SOFP_SHEET_NAME in wb.sheetnames:
        ws = wb[_SOFP_SHEET_NAME]
    else:
        ws = wb.active

    computed_totals: dict[str, float] = {}
    is_balanced = True
    mismatches: list[str] = []
    feedback_lines: list[str] = []
    formula_warnings: list[str] = []

    # Scan for the exact total labels we care about
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.column != 1:
                continue
            label = _normalize_label(str(cell.value))

            if label == _TOTAL_ASSETS_LABEL or label == "total assets":
                val_b = _get_cell_value(wb, ws, cell.row, 2, warnings=formula_warnings)
                val_c = _get_cell_value(wb, ws, cell.row, 3, warnings=formula_warnings)
                if val_b is not None:
                    computed_totals["total_assets_cy"] = val_b
                if val_c is not None:
                    computed_totals["total_assets_py"] = val_c
                if filing_level == "group":
                    val_d = _get_cell_value(wb, ws, cell.row, 4, warnings=formula_warnings)
                    val_e = _get_cell_value(wb, ws, cell.row, 5, warnings=formula_warnings)
                    if val_d is not None:
                        computed_totals["company_total_assets_cy"] = val_d
                    if val_e is not None:
                        computed_totals["company_total_assets_py"] = val_e

            elif label == _TOTAL_EQ_LIAB_LABEL or label == "total equity and liabilities":
                val_b = _get_cell_value(wb, ws, cell.row, 2, warnings=formula_warnings)
                val_c = _get_cell_value(wb, ws, cell.row, 3, warnings=formula_warnings)
                if val_b is not None:
                    computed_totals["total_equity_liabilities_cy"] = val_b
                if val_c is not None:
                    computed_totals["total_equity_liabilities_py"] = val_c
                if filing_level == "group":
                    val_d = _get_cell_value(wb, ws, cell.row, 4, warnings=formula_warnings)
                    val_e = _get_cell_value(wb, ws, cell.row, 5, warnings=formula_warnings)
                    if val_d is not None:
                        computed_totals["company_total_equity_liabilities_cy"] = val_d
                    if val_e is not None:
                        computed_totals["company_total_equity_liabilities_py"] = val_e

    # Surface formula resolution warnings as mismatches
    for w in dict.fromkeys(formula_warnings):  # deduplicate, preserve order
        mismatches.append(f"Formula warning: {w}")

    # Check CY balance
    if (
        "total_assets_cy" in computed_totals
        and "total_equity_liabilities_cy" in computed_totals
    ):
        diff = computed_totals["total_assets_cy"] - computed_totals["total_equity_liabilities_cy"]
        if abs(diff) > 0.01:
            is_balanced = False
            mismatches.append(
                f"CY: assets={computed_totals['total_assets_cy']} "
                f"!= equity+liabilities={computed_totals['total_equity_liabilities_cy']}"
            )
            feedback_lines.extend(_sofp_imbalance_feedback("CY", diff))

    # Check PY balance
    if (
        "total_assets_py" in computed_totals
        and "total_equity_liabilities_py" in computed_totals
    ):
        diff = computed_totals["total_assets_py"] - computed_totals["total_equity_liabilities_py"]
        if abs(diff) > 0.01:
            is_balanced = False
            mismatches.append(
                f"PY: assets={computed_totals['total_assets_py']} "
                f"!= equity+liabilities={computed_totals['total_equity_liabilities_py']}"
            )
            feedback_lines.extend(_sofp_imbalance_feedback("PY", diff))

    # Group filing: also check Company columns (D/E)
    if filing_level == "group":
        if (
            "company_total_assets_cy" in computed_totals
            and "company_total_equity_liabilities_cy" in computed_totals
        ):
            diff = computed_totals["company_total_assets_cy"] - computed_totals["company_total_equity_liabilities_cy"]
            if abs(diff) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"Company CY: assets={computed_totals['company_total_assets_cy']} "
                    f"!= equity+liabilities={computed_totals['company_total_equity_liabilities_cy']}"
                )
                feedback_lines.extend(_sofp_imbalance_feedback("Company CY", diff))
        if (
            "company_total_assets_py" in computed_totals
            and "company_total_equity_liabilities_py" in computed_totals
        ):
            diff = computed_totals["company_total_assets_py"] - computed_totals["company_total_equity_liabilities_py"]
            if abs(diff) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"Company PY: assets={computed_totals['company_total_assets_py']} "
                    f"!= equity+liabilities={computed_totals['company_total_equity_liabilities_py']}"
                )
                feedback_lines.extend(_sofp_imbalance_feedback("Company PY", diff))

    if not computed_totals:
        is_balanced = False
        mismatches.append("No totals found in workbook — cannot verify balance")
        feedback_lines.append("Action: No total rows detected. Check that the template has 'Total assets' and 'Total equity and liabilities' labels.")

    # Compare against PDF reference values if provided
    matches_pdf = True
    if pdf_values:
        for key, expected in pdf_values.items():
            actual = computed_totals.get(key)
            if actual is None:
                mismatches.append(f"Computed total '{key}' not found")
                matches_pdf = False
            elif abs(actual - expected) > 0.01:
                mismatches.append(f"{key}: computed={actual}, expected={expected}")
                matches_pdf = False

    # Phase 2.2: SOFP Group equity attribution check — owners + NCI must
    # sum to Total equity. Group-only because standalone Company SOFPs
    # don't carry NCI. Label matching works on both MFRS and MPERS
    # (identical SSM labels).
    if filing_level == "group":
        total_equity_row = None
        owners_equity_row = None
        nci_equity_row = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.column != 1 or cell.value is None:
                    continue
                norm = _normalize_label(str(cell.value))
                if norm == "total equity":
                    total_equity_row = cell.row
                elif "equity" in norm and "owners of parent" in norm and "attribut" in norm:
                    owners_equity_row = cell.row
                elif norm == "non-controlling interests":
                    nci_equity_row = cell.row

        if total_equity_row and owners_equity_row and nci_equity_row:
            for cy_col, prefix in _cy_columns(filing_level):
                pfx = f"{prefix} " if prefix else ""
                sfx = f"_{prefix.lower()}" if prefix else ""
                te = _get_cell_value(
                    wb, ws, total_equity_row, cy_col, warnings=formula_warnings,
                ) or 0.0
                owners = _get_cell_value(
                    wb, ws, owners_equity_row, cy_col, warnings=formula_warnings,
                ) or 0.0
                nci = _get_cell_value(
                    wb, ws, nci_equity_row, cy_col, warnings=formula_warnings,
                ) or 0.0
                computed_totals[f"total_equity_cy{sfx}"] = te
                computed_totals[f"equity_owners_cy{sfx}"] = owners
                computed_totals[f"equity_nci_cy{sfx}"] = nci
                expected = owners + nci
                if abs(te - expected) > 0.01:
                    is_balanced = False
                    mismatches.append(
                        f"{pfx}Total equity ({te}) != "
                        f"owners ({owners}) + non-controlling interests ({nci}) = {expected}"
                    )

    # Phase 1.1: mandatory-field scan. Runs on the SOFP sheet (ws above)
    # — it is the only sheet verify_totals inspects.
    mandatory_unfilled = _collect_unfilled_mandatory(wb, ws, filing_level)

    wb.close()

    return VerificationResult(
        is_balanced=is_balanced,
        matches_pdf=matches_pdf,
        computed_totals=computed_totals,
        pdf_values=pdf_values or {},
        mismatches=mismatches,
        feedback="\n".join(feedback_lines),
        mandatory_unfilled=mandatory_unfilled,
    )


def _normalize_label(label: str) -> str:
    return label.strip().lstrip("*").strip().lower()


def _get_cell_value(
    wb: openpyxl.Workbook, ws, row: int, col: int,
    warnings: Optional[list[str]] = None,
) -> Optional[float]:
    """Get a cell's effective value — evaluate its formula if it has one,
    otherwise return the literal value. Recurses through formula chains."""
    cell = ws.cell(row=row, column=col)
    raw = cell.value

    if raw is None:
        return None

    if isinstance(raw, str) and raw.startswith("="):
        return _evaluate_formula(wb, ws.title, raw, warnings=warnings)

    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Parametric verifier entry point (Phase 1, Step 1.3)
#
# `verify_totals` above keeps doing SOFP-specific balance checks so the
# legacy single-agent SOFP pipeline keeps passing byte-identically.
# `verify_statement` is the new front door that dispatches on statement type.
# Only SOFP has a meaningful balance identity right now; the other four
# statements return is_balanced=None ("not applicable") while still letting
# the caller supply PDF reference values for cross-checking.
# ---------------------------------------------------------------------------


def _cy_columns(filing_level: str) -> list[tuple[int, str]]:
    """Return (column_index, label_prefix) pairs for the CY data columns."""
    if filing_level == "group":
        return [(2, "Group"), (4, "Company")]
    return [(2, "")]


# Phase 1.1: mandatory-field helper. Scans col A of the given worksheet for
# labels prefixed with `*` and reports any whose CY value column(s) are
# None or empty string. Face sheets only (notes have legitimate blank `*`
# rows). The caller decides which worksheet to pass so this helper stays
# dumb about sheet names / variant quirks.
def _collect_unfilled_mandatory(
    wb: openpyxl.Workbook,
    ws,
    filing_level: str,
) -> list[str]:
    """Return mandatory (`*`-prefixed) row labels whose CY cells are blank.

    A row counts as "unfilled" when every CY column (col B for company
    filings; cols B and D for group filings) is either None, an empty
    string, or a formula that evaluates to 0 AND the underlying cell is
    literally blank (i.e. no data was entered and no upstream cell either).
    Formulas that resolve to a real number — including 0 from a genuine
    sum — count as filled, so agents aren't nagged into fabricating a
    non-zero value for a legitimately zero line item.
    """
    unfilled: list[str] = []
    cy_cols = [c for c, _ in _cy_columns(filing_level)]
    for row in range(1, ws.max_row + 1):
        raw = ws.cell(row=row, column=1).value
        if raw is None:
            continue
        label = str(raw).strip()
        if not label.startswith("*"):
            continue
        # Treat a row as filled if ANY CY column carries a non-empty value
        # (literal or formula-resolved). Note: for group, we require BOTH
        # Group CY and Company CY to be populated — a group filing with a
        # blank Company column is exactly the gap we want to surface.
        row_unfilled = False
        for col in cy_cols:
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is None or (isinstance(val, str) and not val.strip()):
                # Literal blank — unfilled in this column.
                row_unfilled = True
                break
            if isinstance(val, str) and val.startswith("="):
                # Formula present ⇒ treat as filled. We can't distinguish
                # a real zero from an unresolvable formula (both return
                # 0.0), so we prefer quiet false negatives over noisy
                # false positives. Skipping the `_evaluate_formula` walk
                # keeps this hot path O(rows) — full evaluation is
                # expensive (tokenizer + recursive cell resolution) and
                # its return value is unused.
                continue
        if row_unfilled:
            unfilled.append(label)
    return unfilled


def verify_statement(
    path: str,
    statement_type: "object",  # statement_types.StatementType, duck-typed
    variant: str = "",
    pdf_values: Optional[dict[str, float]] = None,
    filing_level: str = "company",
) -> VerificationResult:
    """Verify a filled workbook for a given statement type.

    Each statement type has its own balance identity:
    - SOFP: Total assets == Total equity and liabilities
    - SOCIE: Equity at end == Equity at beginning (restated) + Total increase
    - SOCF: Cash at end == Cash at beginning + Net increase after FX
    - SOPL: Profit == Revenue - Costs (attribution check)
    - SOCI: Total comprehensive income == P&L + Total OCI (attribution check)
    """
    from statement_types import StatementType

    name = statement_type.value if hasattr(statement_type, "value") else str(statement_type)

    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Template not found: {path}")

    if name == StatementType.SOFP.value:
        return verify_totals(path, pdf_values=pdf_values, filing_level=filing_level)
    elif name == StatementType.SOCIE.value:
        return _verify_socie(path, pdf_values=pdf_values, filing_level=filing_level)
    elif name == StatementType.SOCF.value:
        return _verify_socf(path, variant=variant, pdf_values=pdf_values, filing_level=filing_level)
    elif name == StatementType.SOPL.value:
        return _verify_sopl(path, variant=variant, pdf_values=pdf_values, filing_level=filing_level)
    elif name == StatementType.SOCI.value:
        return _verify_soci(path, variant=variant, pdf_values=pdf_values, filing_level=filing_level)

    return VerificationResult(
        is_balanced=None,
        matches_pdf=None,
        computed_totals={},
        pdf_values=pdf_values or {},
        mismatches=[],
        feedback=f"No intra-statement balance check defined for {name}/{variant or 'Default'}.",
    )


# ---------------------------------------------------------------------------
# SOCIE verifier: closing equity == restated opening + total increase
# ---------------------------------------------------------------------------

def _verify_socie(
    path: str,
    pdf_values: Optional[dict[str, float]] = None,
    filing_level: str = "company",
) -> VerificationResult:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb["SOCIE"] if "SOCIE" in wb.sheetnames else wb.active

    computed_totals: dict[str, float] = {}
    mismatches: list[str] = []
    is_balanced = True
    formula_warnings: list[str] = []

    # Find key rows by label (column A)
    label_rows: dict[str, list[int]] = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            norm = _normalize_label(str(val))
            label_rows.setdefault(norm, []).append(row)

    restated_rows = label_rows.get("equity at beginning of period, restated", [])
    total_inc_rows = label_rows.get("total increase (decrease) in equity", [])
    closing_rows = label_rows.get("equity at end of period", [])

    # Fail closed: if we can't find the rows we need, report it
    if not restated_rows or not total_inc_rows or not closing_rows:
        missing = []
        if not restated_rows:
            missing.append("'Equity at beginning of period, restated'")
        if not total_inc_rows:
            missing.append("'Total increase (decrease) in equity'")
        if not closing_rows:
            missing.append("'Equity at end of period'")
        wb.close()
        return VerificationResult(
            is_balanced=False,
            matches_pdf=_check_pdf_values({}, pdf_values),
            computed_totals={},
            pdf_values=pdf_values or {},
            mismatches=[f"Required label not found: {', '.join(missing)}"],
            feedback=f"SOCIE verification failed: missing labels {', '.join(missing)}",
        )

    # Check each period block — the "Total" column is X (col 24)
    total_col = 24  # Column X = grand total

    if filing_level == "group":
        block_labels = ["group_cy", "group_py", "company_cy", "company_py"]
    else:
        block_labels = ["cy", "py"]

    for i, (rest_r, inc_r, close_r) in enumerate(
        zip(restated_rows, total_inc_rows, closing_rows)
    ):
        label = block_labels[i] if i < len(block_labels) else f"block_{i}"

        restated = _get_cell_value(wb, ws, rest_r, total_col, warnings=formula_warnings) or 0.0
        increase = _get_cell_value(wb, ws, inc_r, total_col, warnings=formula_warnings) or 0.0
        closing = _get_cell_value(wb, ws, close_r, total_col, warnings=formula_warnings) or 0.0

        computed_totals[f"restated_equity_{label}"] = restated
        computed_totals[f"total_increase_{label}"] = increase
        computed_totals[f"closing_equity_{label}"] = closing

        expected = restated + increase
        diff = closing - expected
        if abs(diff) > 0.01:
            is_balanced = False
            mismatches.append(
                f"{label}: closing equity ({closing}) != "
                f"restated ({restated}) + total increase ({increase}) = {expected}"
            )

    for w in dict.fromkeys(formula_warnings):
        mismatches.append(f"Formula warning: {w}")

    mandatory_unfilled = _collect_unfilled_mandatory(wb, ws, filing_level)

    wb.close()

    matches_pdf = _check_pdf_values(computed_totals, pdf_values)

    return VerificationResult(
        is_balanced=is_balanced,
        matches_pdf=matches_pdf,
        computed_totals=computed_totals,
        pdf_values=pdf_values or {},
        mismatches=mismatches,
        feedback=_compose_feedback(
            mismatches, is_balanced, "SOCIE balance check passed.",
        ),
        mandatory_unfilled=mandatory_unfilled,
    )


# ---------------------------------------------------------------------------
# SOCF verifier: cash at end == cash at beginning + net increase after FX
# ---------------------------------------------------------------------------

def _verify_socf(
    path: str,
    variant: str = "",
    pdf_values: Optional[dict[str, float]] = None,
    filing_level: str = "company",
) -> VerificationResult:
    wb = openpyxl.load_workbook(path, data_only=False)

    # Try to find the right sheet
    sheet_names_to_try = ["SOCF-Indirect", "SOCF-Direct"]
    ws = None
    for sn in sheet_names_to_try:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        ws = wb.active

    computed_totals: dict[str, float] = {}
    mismatches: list[str] = []
    is_balanced = True
    formula_warnings: list[str] = []

    # Find key rows
    rows_by_label: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            norm = _normalize_label(str(val))
            if "net cash flows" in norm and "operating" in norm:
                rows_by_label["net_operating"] = row
            elif "net cash flows" in norm and "investing" in norm:
                rows_by_label["net_investing"] = row
            elif "net cash flows" in norm and "financing" in norm:
                rows_by_label["net_financing"] = row
            elif "net increase" in norm and "after" in norm:
                rows_by_label["net_increase_after_fx"] = row
            elif "net increase" in norm and "before" in norm:
                rows_by_label["net_increase_before_fx"] = row
            elif "cash and cash equivalents at end" in norm:
                rows_by_label["cash_end"] = row
            elif "cash and cash equivalents at beginning" in norm:
                rows_by_label["cash_beginning"] = row

    # Fail closed: require at least operating + net increase rows
    required = ["net_operating", "net_increase_before_fx"]
    missing = [k for k in required if k not in rows_by_label]
    if missing:
        wb.close()
        return VerificationResult(
            is_balanced=False,
            matches_pdf=_check_pdf_values({}, pdf_values),
            computed_totals={},
            pdf_values=pdf_values or {},
            mismatches=[f"Required SOCF rows not found: {', '.join(missing)}"],
            feedback=f"SOCF verification failed: missing rows {', '.join(missing)}",
        )

    for cy_col, prefix in _cy_columns(filing_level):
        pfx = f"{prefix} " if prefix else ""
        sfx = f"_{prefix.lower()}" if prefix else ""

        # Check: operating + investing + financing == net increase before FX
        col_totals: dict[str, float] = {}
        for key in ["net_operating", "net_investing", "net_financing", "net_increase_before_fx"]:
            if key in rows_by_label:
                col_totals[key] = _get_cell_value(wb, ws, rows_by_label[key], cy_col, warnings=formula_warnings) or 0.0
                computed_totals[f"{key}{sfx}"] = col_totals[key]

        if all(k in col_totals for k in ["net_operating", "net_investing", "net_financing", "net_increase_before_fx"]):
            expected = col_totals["net_operating"] + col_totals["net_investing"] + col_totals["net_financing"]
            actual = col_totals["net_increase_before_fx"]
            if abs(actual - expected) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"{pfx}Net increase before FX ({actual}) != "
                    f"Operating ({col_totals['net_operating']}) + "
                    f"Investing ({col_totals['net_investing']}) + "
                    f"Financing ({col_totals['net_financing']}) = {expected}"
                )

        # Check: cash at end == cash at beginning + net increase after FX
        for key in ["cash_beginning", "cash_end", "net_increase_after_fx"]:
            if key in rows_by_label:
                col_totals[key] = _get_cell_value(wb, ws, rows_by_label[key], cy_col, warnings=formula_warnings) or 0.0
                computed_totals[f"{key}{sfx}"] = col_totals[key]

        if all(k in col_totals for k in ["cash_beginning", "cash_end", "net_increase_after_fx"]):
            expected = col_totals["cash_beginning"] + col_totals["net_increase_after_fx"]
            actual = col_totals["cash_end"]
            if abs(actual - expected) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"{pfx}Cash at end ({actual}) != "
                    f"Cash at beginning ({col_totals['cash_beginning']}) + "
                    f"Net increase after FX ({col_totals['net_increase_after_fx']}) = {expected}"
                )

    for w in dict.fromkeys(formula_warnings):
        mismatches.append(f"Formula warning: {w}")

    mandatory_unfilled = _collect_unfilled_mandatory(wb, ws, filing_level)

    wb.close()

    matches_pdf = _check_pdf_values(computed_totals, pdf_values)

    return VerificationResult(
        is_balanced=is_balanced,
        matches_pdf=matches_pdf,
        computed_totals=computed_totals,
        pdf_values=pdf_values or {},
        mismatches=mismatches,
        feedback=_compose_feedback(
            mismatches, is_balanced, "SOCF balance check passed.",
        ),
        mandatory_unfilled=mandatory_unfilled,
    )


# ---------------------------------------------------------------------------
# SOPL verifier: attribution check (total P&L == owners + NCI)
# ---------------------------------------------------------------------------

def _verify_sopl(
    path: str,
    variant: str = "",
    pdf_values: Optional[dict[str, float]] = None,
    filing_level: str = "company",
) -> VerificationResult:
    wb = openpyxl.load_workbook(path, data_only=False)

    sheet_names_to_try = ["SOPL-Function", "SOPL-Nature"]
    ws = None
    for sn in sheet_names_to_try:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        ws = wb.active

    computed_totals: dict[str, float] = {}
    mismatches: list[str] = []
    is_balanced = True
    formula_warnings: list[str] = []

    # Find profit/loss and attribution total rows.
    # SOPL-Function uses "Total profit (loss)" for the attribution row.
    # SOPL-Nature uses a second "Profit (loss)" row instead.
    profit_loss_row = None
    total_profit_row = None
    last_profit_loss_row = None  # track the last "profit (loss)" for Nature variant
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            norm = _normalize_label(str(val))
            if norm == "profit (loss)":
                if profit_loss_row is None:
                    profit_loss_row = row
                last_profit_loss_row = row
            elif norm == "total profit (loss)":
                total_profit_row = row

    # For Nature variant: if no "Total profit (loss)" exists but there are
    # multiple "Profit (loss)" rows, the last one is the attribution total.
    if total_profit_row is None and last_profit_loss_row and last_profit_loss_row != profit_loss_row:
        total_profit_row = last_profit_loss_row

    # Fail closed: require at least the primary profit/loss row
    if not profit_loss_row:
        wb.close()
        return VerificationResult(
            is_balanced=False,
            matches_pdf=_check_pdf_values({}, pdf_values),
            computed_totals={},
            pdf_values=pdf_values or {},
            mismatches=["Required label not found: 'Profit (loss)'"],
            feedback="SOPL verification failed: missing 'Profit (loss)' label",
        )

    # Phase 2.1: SOPL Group attribution check — owners + NCI must sum to
    # the Profit (loss) row. Only fires on group filings because standalone
    # filings don't carry the attribution rows. We find the rows by label
    # match so MFRS and MPERS (which use the identical `attributable to,
    # owners of parent` / `non-controlling interests` wording) both work.
    owners_profit_row = None
    nci_profit_row = None
    if filing_level == "group":
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if not val:
                continue
            norm = _normalize_label(str(val))
            # Match the full SSM label "Profit (loss), attributable to,
            # owners of parent" and also tolerate shorter variants that
            # some agents may produce.
            if "profit" in norm and "owners of parent" in norm and "attribut" in norm:
                owners_profit_row = row
            elif (
                "profit" in norm
                and "non-controlling interest" in norm
                and "attribut" in norm
            ):
                nci_profit_row = row

    for cy_col, prefix in _cy_columns(filing_level):
        pfx = f"{prefix} " if prefix else ""
        sfx = f"_{prefix.lower()}" if prefix else ""

        pl_val = _get_cell_value(wb, ws, profit_loss_row, cy_col, warnings=formula_warnings) or 0.0
        computed_totals[f"profit_loss_cy{sfx}"] = pl_val

        if total_profit_row:
            attr_val = _get_cell_value(wb, ws, total_profit_row, cy_col, warnings=formula_warnings) or 0.0
            computed_totals[f"total_profit_attribution_cy{sfx}"] = attr_val

            if abs(pl_val - attr_val) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"{pfx}Profit/loss ({pl_val}) != attribution total ({attr_val})"
                )

        # Phase 2.1: owners + NCI attribution sum check. Only fires when
        # filing_level == "group" and both rows were found.
        if filing_level == "group" and owners_profit_row and nci_profit_row:
            owners = _get_cell_value(
                wb, ws, owners_profit_row, cy_col, warnings=formula_warnings,
            ) or 0.0
            nci = _get_cell_value(
                wb, ws, nci_profit_row, cy_col, warnings=formula_warnings,
            ) or 0.0
            computed_totals[f"profit_owners_cy{sfx}"] = owners
            computed_totals[f"profit_nci_cy{sfx}"] = nci
            expected = owners + nci
            if abs(pl_val - expected) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"{pfx}Profit/loss ({pl_val}) != "
                    f"owners ({owners}) + non-controlling interests ({nci}) = {expected}"
                )

    for w in dict.fromkeys(formula_warnings):
        mismatches.append(f"Formula warning: {w}")

    mandatory_unfilled = _collect_unfilled_mandatory(wb, ws, filing_level)

    wb.close()

    matches_pdf = _check_pdf_values(computed_totals, pdf_values)

    return VerificationResult(
        is_balanced=is_balanced,
        matches_pdf=matches_pdf,
        computed_totals=computed_totals,
        pdf_values=pdf_values or {},
        mismatches=mismatches,
        feedback=_compose_feedback(
            mismatches, is_balanced, "SOPL attribution check passed.",
        ),
        mandatory_unfilled=mandatory_unfilled,
    )


# ---------------------------------------------------------------------------
# SOCI verifier: total comprehensive income == P&L + OCI; attribution check
# ---------------------------------------------------------------------------

def _verify_soci(
    path: str,
    variant: str = "",
    pdf_values: Optional[dict[str, float]] = None,
    filing_level: str = "company",
) -> VerificationResult:
    wb = openpyxl.load_workbook(path, data_only=False)

    sheet_names_to_try = ["SOCI-BeforeOfTax", "SOCI-NetOfTax"]
    ws = None
    for sn in sheet_names_to_try:
        if sn in wb.sheetnames:
            ws = wb[sn]
            break
    if ws is None:
        ws = wb.active

    computed_totals: dict[str, float] = {}
    mismatches: list[str] = []
    is_balanced = True
    formula_warnings: list[str] = []

    # Find key rows
    pl_row = None
    total_oci_row = None
    total_ci_rows = []

    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            norm = _normalize_label(str(val))
            if norm == "profit (loss)" and pl_row is None:
                pl_row = row
            elif norm == "total other comprehensive income" and total_oci_row is None:
                total_oci_row = row
            elif norm == "total comprehensive income":
                total_ci_rows.append(row)

    # Fail closed: require at least P&L and one Total CI row
    if not pl_row or not total_ci_rows:
        missing = []
        if not pl_row:
            missing.append("'Profit (loss)'")
        if not total_ci_rows:
            missing.append("'Total comprehensive income'")
        wb.close()
        return VerificationResult(
            is_balanced=False,
            matches_pdf=_check_pdf_values({}, pdf_values),
            computed_totals={},
            pdf_values=pdf_values or {},
            mismatches=[f"Required label not found: {', '.join(missing)}"],
            feedback=f"SOCI verification failed: missing labels {', '.join(missing)}",
        )

    for cy_col, prefix in _cy_columns(filing_level):
        pfx = f"{prefix} " if prefix else ""
        sfx = f"_{prefix.lower()}" if prefix else ""

        pl_val = _get_cell_value(wb, ws, pl_row, cy_col, warnings=formula_warnings) or 0.0
        computed_totals[f"profit_loss_cy{sfx}"] = pl_val

        oci_val = None
        if total_oci_row:
            oci_val = _get_cell_value(wb, ws, total_oci_row, cy_col, warnings=formula_warnings) or 0.0
            computed_totals[f"total_oci_cy{sfx}"] = oci_val

        if total_ci_rows:
            ci_val = _get_cell_value(wb, ws, total_ci_rows[0], cy_col, warnings=formula_warnings) or 0.0
            computed_totals[f"total_comprehensive_income_cy{sfx}"] = ci_val

            if oci_val is not None:
                expected = pl_val + oci_val
                if abs(ci_val - expected) > 0.01:
                    is_balanced = False
                    mismatches.append(
                        f"{pfx}Total CI ({ci_val}) != P&L ({pl_val}) "
                        f"+ OCI ({oci_val}) = {expected}"
                    )

        if len(total_ci_rows) >= 2:
            ci_main = _get_cell_value(wb, ws, total_ci_rows[0], cy_col, warnings=formula_warnings) or 0.0
            ci_attr = _get_cell_value(wb, ws, total_ci_rows[1], cy_col, warnings=formula_warnings) or 0.0
            computed_totals[f"total_ci_attribution_cy{sfx}"] = ci_attr
            if abs(ci_main - ci_attr) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"{pfx}Total CI ({ci_main}) != attribution total ({ci_attr})"
                )

    for w in dict.fromkeys(formula_warnings):
        mismatches.append(f"Formula warning: {w}")

    mandatory_unfilled = _collect_unfilled_mandatory(wb, ws, filing_level)

    wb.close()

    matches_pdf = _check_pdf_values(computed_totals, pdf_values)

    return VerificationResult(
        is_balanced=is_balanced,
        matches_pdf=matches_pdf,
        computed_totals=computed_totals,
        pdf_values=pdf_values or {},
        mismatches=mismatches,
        feedback=_compose_feedback(
            mismatches, is_balanced, "SOCI balance check passed.",
        ),
        mandatory_unfilled=mandatory_unfilled,
    )


# ---------------------------------------------------------------------------
# Cross-sheet validation: run all inter-statement consistency checks
# ---------------------------------------------------------------------------

def verify_cross_sheet(
    workbook_paths: dict[str, str],
) -> VerificationResult:
    """Check inter-statement identities across filled workbooks.

    `workbook_paths` maps statement type names (e.g. "SOFP", "SOCIE") to
    file paths of their filled workbooks.

    Checks performed:
    - SOFP Total equity == SOCIE Closing equity (Total column)
    - SOPL Profit/loss == SOCI Profit/loss (top of SOCI)
    - SOCF Cash at end == SOFP Cash and cash equivalents
    """
    mismatches: list[str] = []
    computed_totals: dict[str, float] = {}
    is_balanced = True

    # Helper to get a labelled value from a workbook.
    # Uses exact-match-first to avoid "total equity" matching
    # "total equity attributable to owners" before "total equity".
    def _find_value(path: str, sheet_names: list[str], label_target: str, col: int = 2) -> Optional[float]:
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = None
        for sn in sheet_names:
            if sn in wb.sheetnames:
                ws = wb[sn]
                break
        if ws is None:
            ws = wb.active

        target = label_target.strip().lower()
        exact_row = None
        substr_row = None
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val is None:
                continue
            normalized = _normalize_label(str(val))
            if normalized == target:
                exact_row = row
                break
            if substr_row is None and target in normalized:
                substr_row = row

        match_row = exact_row or substr_row
        if match_row is None:
            wb.close()
            return None

        result = _get_cell_value(wb, ws, match_row, col)
        wb.close()
        return result

    # Check 1: SOFP Total equity == SOCIE Closing equity
    if "SOFP" in workbook_paths and "SOCIE" in workbook_paths:
        sofp_equity = _find_value(
            workbook_paths["SOFP"],
            ["SOFP-CuNonCu", "SOFP-OrdOfLiq"],
            "total equity",
            col=2,
        )
        # SOCIE closing equity is in column X (24), last "Equity at end" row
        socie_wb = openpyxl.load_workbook(workbook_paths["SOCIE"], data_only=False)
        socie_ws = socie_wb["SOCIE"] if "SOCIE" in socie_wb.sheetnames else socie_wb.active
        socie_closing = None
        for row in range(1, socie_ws.max_row + 1):
            val = socie_ws.cell(row=row, column=1).value
            if val and "equity at end of period" in _normalize_label(str(val)):
                socie_closing = _get_cell_value(socie_wb, socie_ws, row, 24)  # col X
                break  # First period = CY
        socie_wb.close()

        if sofp_equity is not None:
            computed_totals["sofp_total_equity"] = sofp_equity
        if socie_closing is not None:
            computed_totals["socie_closing_equity"] = socie_closing

        if sofp_equity is not None and socie_closing is not None:
            if abs(sofp_equity - socie_closing) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"SOFP Total equity ({sofp_equity}) != SOCIE Closing equity ({socie_closing})"
                )

    # Check 2: SOPL Profit/loss == SOCI Profit/loss
    if "SOPL" in workbook_paths and "SOCI" in workbook_paths:
        sopl_profit = _find_value(
            workbook_paths["SOPL"],
            ["SOPL-Function", "SOPL-Nature"],
            "profit (loss)",
            col=2,
        )
        soci_profit = _find_value(
            workbook_paths["SOCI"],
            ["SOCI-BeforeOfTax", "SOCI-NetOfTax"],
            "profit (loss)",
            col=2,
        )

        if sopl_profit is not None:
            computed_totals["sopl_profit_loss"] = sopl_profit
        if soci_profit is not None:
            computed_totals["soci_profit_loss"] = soci_profit

        if sopl_profit is not None and soci_profit is not None:
            if abs(sopl_profit - soci_profit) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"SOPL Profit/loss ({sopl_profit}) != SOCI Profit/loss ({soci_profit})"
                )

    # Check 3: SOCF Cash at end == SOFP Cash and cash equivalents
    if "SOCF" in workbook_paths and "SOFP" in workbook_paths:
        socf_cash = _find_value(
            workbook_paths["SOCF"],
            ["SOCF-Indirect", "SOCF-Direct"],
            "cash and cash equivalents at end",
            col=2,
        )
        sofp_cash = _find_value(
            workbook_paths["SOFP"],
            ["SOFP-CuNonCu", "SOFP-OrdOfLiq"],
            "cash and cash equivalents",
            col=2,
        )

        if socf_cash is not None:
            computed_totals["socf_cash_end"] = socf_cash
        if sofp_cash is not None:
            computed_totals["sofp_cash"] = sofp_cash

        if socf_cash is not None and sofp_cash is not None:
            if abs(socf_cash - sofp_cash) > 0.01:
                is_balanced = False
                mismatches.append(
                    f"SOCF Cash at end ({socf_cash}) != SOFP Cash ({sofp_cash})"
                )

    return VerificationResult(
        is_balanced=is_balanced,
        matches_pdf=None,
        computed_totals=computed_totals,
        pdf_values={},
        mismatches=mismatches,
        feedback="\n".join(mismatches) if mismatches else "All cross-sheet checks passed.",
    )


def _check_pdf_values(
    computed_totals: dict[str, float],
    pdf_values: Optional[dict[str, float]],
) -> Optional[bool]:
    """Compare computed totals against PDF reference values if provided."""
    if not pdf_values:
        return None
    matches = True
    for key, expected in pdf_values.items():
        actual = computed_totals.get(key)
        if actual is None:
            matches = False
        elif abs(actual - expected) > 0.01:
            matches = False
    return matches
