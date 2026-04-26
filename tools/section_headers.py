"""Discover section-header rows by walking a worksheet's column A.

Why: `fill_workbook` needs to know which rows are section boundaries so it can
disambiguate duplicate labels (e.g. "Lease liabilities" appears under both
non-current and current liabilities). Previously this was hard-coded against
one SOFP template. For the multi-statement rollout, the same detector has to
work across 9 MBRS templates — and it has to keep working on the legacy
`SOFP-Xbrl-template.xlsx` until Phase 7 retires it.

How: rows in MBRS templates use fill color to signal their role:

    Header rows     -> coloured fill, not bold-only; repeated per section
    Total rows      -> coloured fill with "Total" in the label
    Line items      -> default / white fill

We recognise header rows by (a) a whitelist of header fill colours observed
across both the legacy and new template families, and (b) a keyword set the
caller can pass in (scout populates this from the registry).
"""
from __future__ import annotations

from dataclasses import dataclass

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


# ARGB fill codes used to paint section-header rows. openpyxl reports the
# leading alpha byte as "00" when the colour is saved with a zero alpha and
# "FF" when saved opaque — we accept both variants of each hue.
#   1F3864 — dark navy, main section headers in Company + Group templates
#   305496 — mid blue, Group SOCIE block dividers ("Group - Current period", …)
_HEADER_FILL_RGB = frozenset({
    "001F3864", "FF1F3864",
    "00305496", "FF305496",
})

# ARGB fill codes used to paint Total rows. We exclude these so totals don't
# get mistaken for section headers.
#   EEF2F8 — pale blue, Total rows in Company + Group templates
_TOTAL_FILL_RGB = frozenset({
    "00EEF2F8", "FFEEF2F8",
})


@dataclass(frozen=True)
class SectionHeader:
    row: int
    label: str             # original cell text, trimmed
    normalized: str        # lowercased, leading '*' stripped


def _normalize(label: str) -> str:
    return label.strip().lstrip("*").strip().lower()


def _cell_fill_rgb(cell) -> str | None:
    """Return the fill's ARGB string if set, else None.

    openpyxl sometimes returns a `Color` object whose `rgb` is `None` when the
    fill is theme-indexed rather than RGB — we treat that as "no fill" since
    none of our templates use themed fills for headers.
    """
    if not cell.fill or not cell.fill.fgColor:
        return None
    rgb = cell.fill.fgColor.rgb
    if isinstance(rgb, str):
        return rgb
    return None


def discover_section_headers(
    ws: Worksheet,
    extra_keywords: frozenset[str] | None = None,
) -> list[SectionHeader]:
    """Return the section-header rows in worksheet order.

    Walks column A top-to-bottom. A row is classified as a header if either:
      1. its column-A cell has a header-coloured fill AND is not a total row, or
      2. its normalised label matches one of `extra_keywords`.
    """
    extra = extra_keywords or frozenset()
    headers: list[SectionHeader] = []

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value is None:
            continue
        label_raw = str(cell.value).strip()
        if not label_raw:
            continue
        norm = _normalize(label_raw)

        rgb = _cell_fill_rgb(cell)
        is_header_fill = rgb in _HEADER_FILL_RGB
        is_total_fill = rgb in _TOTAL_FILL_RGB

        # Total rows are coloured too but they're not section headers.
        if is_total_fill or norm.startswith("total "):
            continue

        if is_header_fill or norm in extra:
            headers.append(SectionHeader(row=row, label=label_raw, normalized=norm))

    return headers


def header_set(
    wb: openpyxl.Workbook,
    sheet_name: str,
    extra_keywords: frozenset[str] | None = None,
) -> frozenset[str]:
    """Convenience: normalized-label set for a sheet, used by fill_workbook."""
    if sheet_name not in wb.sheetnames:
        return frozenset()
    return frozenset(
        h.normalized for h in discover_section_headers(wb[sheet_name], extra_keywords)
    )


# ---------------------------------------------------------------------------
# Shared keyword fallback registry (peer-review #1, 2026-04-26)
#
# Header detection by fill colour covers ~95% of cases. The remaining 5% are
# uncoloured-but-load-bearing rows: SOFP sub-sheet sub-section dividers and
# MPERS Group SOCIE block dividers ("Group - Current period", etc.).
# fill_workbook's label-disambiguation pass needs them; template_reader's
# is_abstract marking needs them too. Owning the keyword fallback selection
# here keeps reader and writer symmetric — without it, an agent could see
# [DATA_ENTRY] in the read_template summary and then have its write refused.
# ---------------------------------------------------------------------------

_LEGACY_MAIN_HEADER_KEYWORDS: frozenset[str] = frozenset({
    "non-current assets",
    "current assets",
    "equity",
    "non-current liabilities",
    "current liabilities",
})

_LEGACY_SUB_HEADER_KEYWORDS: frozenset[str] = _LEGACY_MAIN_HEADER_KEYWORDS | frozenset({
    "property, plant and equipment",
    "investment property",
    "biological assets",
    "intangible assets",
    "investments in subsidiaries",
    "investments in associates",
    "investments in joint ventures",
    "non-current trade receivables",
    "current trade receivables",
    "non-current derivative financial assets",
    "current derivative financial assets",
    "inventories",
    "cash and cash equivalents",
    "non-current borrowings",
    "current borrowings",
    "non-current employee benefit liabilities",
    "current employee benefit liabilities",
    "non-current provisions",
    "current provisions",
    "non-current trade payables",
    "current trade payables",
    "non-current non-trade payables",
    "current non-trade payables",
    "non-current derivative financial liabilities",
    "current derivative financial liabilities",
})

_MPERS_GROUP_SOCIE_BLOCK_HEADERS: frozenset[str] = frozenset({
    "group - current period",
    "group - prior period",
    "company - current period",
    "company - prior period",
})


def keyword_fallback_for_sheet(sheet_name: str) -> frozenset[str]:
    """Return the keyword fallback set appropriate for a sheet.

    Used by both the writer (fill_workbook) and the reader (template_reader)
    so that abstract-row detection stays symmetric across the two layers.
    Sheet-name heuristics mirror the legacy logic from fill_workbook's
    `_build_label_index`: SOCIE sheets get the MPERS block-divider keywords;
    sub-sheets / analysis sheets get the broader sub-section keyword set;
    everything else gets the main-statement keyword set.
    """
    name = sheet_name.lower()
    if "socie" in name:
        return _LEGACY_MAIN_HEADER_KEYWORDS | _MPERS_GROUP_SOCIE_BLOCK_HEADERS
    if "sub" in name or "analysis" in name:
        return _LEGACY_SUB_HEADER_KEYWORDS
    return _LEGACY_MAIN_HEADER_KEYWORDS
