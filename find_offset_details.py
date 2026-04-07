#!/usr/bin/env python3
"""
Deep dive analysis: find actual +20 offset issues by examining:
1. Which rows have data labels
2. Which rows are referenced by formulas
3. Check if references point to blank rows when rows ±20 have data
"""

import openpyxl
from pathlib import Path
from collections import defaultdict

class OffsetDiveAnalyzer:
    def __init__(self, template_path: str):
        self.path = Path(template_path)
        self.wb = openpyxl.load_workbook(self.path, data_only=False)
        self.template_name = self.path.name

    def get_label(self, sheet_name: str, row: int) -> str:
        """Get label from column A."""
        try:
            ws = self.wb[sheet_name]
            val = ws[f'A{row}'].value
            return str(val).strip() if val else ""
        except:
            return ""

    def extract_refs_from_formula(self, formula: str):
        """Extract all B/C row references from formula."""
        import re
        refs = []
        pattern = r'[\$]?([B-D])[\$]?(\d+)'
        for match in re.finditer(pattern, formula):
            col, row = match.groups()
            refs.append((col, int(row)))
        return refs

    def analyze_sheet_deep(self, sheet_name: str):
        """
        Find all formulas and check:
        - What rows do they reference?
        - Are those rows blank with data at row±20?
        """
        try:
            ws = self.wb[sheet_name]
        except KeyError:
            return None

        # Map all rows with labels
        label_map = {}
        for row in range(1, ws.max_row + 1):
            label = self.get_label(sheet_name, row)
            if label:
                label_map[row] = label

        # Find all formulas and their references
        formula_info = {}  # formula_cell -> {formula, refs, label}
        for row in range(1, ws.max_row + 1):
            for col in ['B', 'C']:
                cell = ws[f'{col}{row}']
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    label = self.get_label(sheet_name, row)
                    refs = self.extract_refs_from_formula(formula)
                    if refs:
                        formula_info[f'{col}{row}'] = {
                            'formula': formula,
                            'refs': refs,
                            'label': label,
                            'row': row
                        }

        # Now check each reference: does it point to a blank row with data at ±20?
        suspicious_formulas = []

        for cell_addr, info in formula_info.items():
            for ref_col, ref_row in info['refs']:
                ref_label = label_map.get(ref_row, "")

                # If reference is blank, check ±20
                if not ref_label:
                    for offset in [20, -20]:
                        target_row = ref_row + offset
                        target_label = label_map.get(target_row, "")
                        if target_label:
                            # Found a suspicious pattern
                            suspicious_formulas.append({
                                'formula_cell': cell_addr,
                                'formula_label': info['label'],
                                'formula': info['formula'],
                                'formula_row': info['row'],
                                'ref': f'{ref_col}{ref_row}',
                                'ref_label': ref_label or '[BLANK]',
                                'offset': offset,
                                'corrected_row': target_row,
                                'corrected_label': target_label
                            })

        return suspicious_formulas, label_map, formula_info

def main():
    templates_to_check = {
        "BUGGY (claim: +20 offset)": [
            ("XBRL-template-MFRS/07-SOCF-Indirect.xlsx", "SOCF-Indirect"),
            ("XBRL-template-MFRS/08-SOCF-Direct.xlsx", "SOCF-Direct"),
            ("XBRL-template-MFRS/09-SOCIE.xlsx", "SOCIE"),
            ("XBRL-template-MFRS/04-SOPL-Nature.xlsx", "SOPL-Analysis-Nature"),
        ],
        "CLEAN (claim: no bugs)": [
            ("XBRL-template-MFRS/03-SOPL-Function.xlsx", "SOPL-Analysis-Function"),
            ("XBRL-template-MFRS/05-SOCI-BeforeTax.xlsx", "SOCI-BeforeOfTax"),
            ("XBRL-template-MFRS/06-SOCI-NetOfTax.xlsx", "SOCI-NetOfTax"),
        ]
    }

    base_path = Path("/sessions/happy-zealous-dirac/mnt/xbrl-agent")

    print("\n" + "=" * 100)
    print("DETAILED +20 OFFSET BUG ANALYSIS")
    print("=" * 100)

    for category, templates in templates_to_check.items():
        print(f"\n{'#' * 3} {category.upper()}")
        print()

        for template, sheet_name in templates:
            full_path = base_path / template
            if not full_path.exists():
                print(f"File not found: {template}")
                continue

            print(f"Template: {Path(template).name}")
            print(f"Sheet: {sheet_name}")
            print("-" * 80)

            analyzer = OffsetDiveAnalyzer(str(full_path))
            result = analyzer.analyze_sheet_deep(sheet_name)

            if not result:
                print(f"  ERROR: Could not analyze sheet")
                continue

            suspicious, label_map, formula_info = result

            if suspicious:
                print(f"  FOUND {len(suspicious)} SUSPICIOUS PATTERNS:\n")
                for i, finding in enumerate(suspicious[:10], 1):
                    print(f"  {i}. Formula at {finding['formula_cell']}")
                    print(f"     Label: {finding['formula_label']}")
                    print(f"     Formula: {finding['formula']}")
                    print(f"     References {finding['ref']} which is BLANK")
                    print(f"     But {finding['ref']} + {finding['offset']:+d} = Row {finding['corrected_row']}: {finding['corrected_label']}")
                    print()

                if len(suspicious) > 10:
                    print(f"  ... and {len(suspicious) - 10} more patterns\n")
            else:
                print(f"  No suspicious +20 offset patterns found")
                print(f"  Total formulas: {len(formula_info)}")
                print(f"  Total labeled rows: {len(label_map)}")
                # Show some sample formulas to verify we're reading them correctly
                print(f"\n  Sample formulas (first 3):")
                for cell_addr, info in list(formula_info.items())[:3]:
                    print(f"    {cell_addr}: {info['formula']}")

            print()

    print("=" * 100)
    print("INTERPRETATION GUIDE")
    print("=" * 100)
    print("""
If "BUGGY (claim: +20 offset)" templates show FOUND patterns:
  -> CLAIM IS CORRECT - The guides correctly identified offset bugs

If "BUGGY" templates show NO suspicious patterns:
  -> CLAIM IS INCORRECT OR MISLEADING - No obvious offset bugs detected

If "CLEAN (claim: no bugs)" templates show FOUND patterns:
  -> CLAIM IS WRONG - Unexpected bugs found in "clean" templates

If "CLEAN" templates show NO patterns:
  -> CLAIM IS CORRECT - These templates are indeed clean
""")

if __name__ == "__main__":
    main()
