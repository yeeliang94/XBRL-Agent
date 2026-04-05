"""Compare agent extraction output against the FINCO reference template.

The reference file has extra blank rows inserted at the top of each sheet
(1 extra in sub-sheet, 2 extra in main sheet), so we compare by LABEL,
not by row number. This gives an accurate picture of extraction quality.
"""
import sys
import openpyxl

ref_path = sys.argv[1] if len(sys.argv) > 1 else "SOFP-Xbrl-reference-FINCO-filled.xlsx"
agent_path = sys.argv[2] if len(sys.argv) > 2 else "output/run_001/filled.xlsx"

ref = openpyxl.load_workbook(ref_path, data_only=True)
agent = openpyxl.load_workbook(agent_path, data_only=False)

sheets = ["SOFP-CuNonCu", "SOFP-Sub-CuNonCu"]
total_correct = 0
total_wrong = 0
total_missing = 0


def normalize(label):
    if not label:
        return ""
    return str(label).strip().lstrip("*").strip().lower()


def is_formula(val):
    return isinstance(val, str) and val.startswith("=")


for sheet_name in sheets:
    ref_ws = ref[sheet_name]
    agent_ws = agent[sheet_name]
    # Also load reference without data_only to check which cells are formulas
    ref_raw_ws = openpyxl.load_workbook(ref_path, data_only=False)[sheet_name]

    # Build label -> {col -> value} maps for both files.
    # For duplicate labels, use section context (rows above) to disambiguate.
    def build_value_map(ws, raw_ws=None, use_data_only=True):
        """Build a map of (normalized_label, section, col) -> value for data-entry cells."""
        values = {}
        current_section = ""
        section_headers = {
            "non-current assets", "current assets", "equity",
            "non-current liabilities", "current liabilities",
            "property, plant and equipment", "investment property",
            "current trade receivables", "current non-trade receivables",
            "cash and cash equivalents", "cash",
            "current trade payables", "current non-trade payables",
        }
        for row in range(1, ws.max_row + 1):
            label_raw = ws.cell(row, 1).value
            if not label_raw:
                continue
            norm = normalize(label_raw)
            if norm in section_headers:
                current_section = norm

            for col in [2, 3]:
                val = ws.cell(row, col).value
                # Skip formula cells — we only compare data-entry values
                check_ws = raw_ws if raw_ws else ws
                raw_val = check_ws.cell(row, col).value
                if is_formula(raw_val):
                    continue
                if val is None:
                    continue
                # Skip header text values like "RM", date strings
                try:
                    float(val)
                except (ValueError, TypeError):
                    continue

                key = (norm, current_section, col)
                values[key] = float(val)
        return values

    ref_values = build_value_map(ref_ws, ref_raw_ws, use_data_only=True)
    agent_values = build_value_map(agent_ws, agent_ws, use_data_only=False)

    header = f"{'Label':<50} {'Section':<25} {'Col':<5} {'Reference':<12} {'Agent':<12} {'Status'}"
    print(f"\n=== {sheet_name} ===")
    print(header)
    print("-" * len(header))

    correct = 0
    wrong = 0
    missing = 0

    # Check all reference values
    all_keys = set(ref_values.keys()) | set(agent_values.keys())
    for key in sorted(all_keys, key=lambda k: (k[1], k[0], k[2])):
        norm_label, section, col = key
        col_label = "CY" if col == 2 else "PY"
        ref_val = ref_values.get(key)
        agent_val = agent_values.get(key)

        # Both zero or both absent
        if (ref_val is None or ref_val == 0) and (agent_val is None or agent_val == 0):
            if ref_val is not None or agent_val is not None:
                correct += 1
            continue

        if ref_val is not None and (agent_val is None or agent_val == 0) and ref_val != 0:
            missing += 1
            print(f"{norm_label[:50]:<50} {section[:25]:<25} {col_label:<5} {ref_val:<12.0f} {'---':<12} MISSING")
        elif ref_val is not None and agent_val is not None and abs(ref_val - agent_val) < 1:
            correct += 1
        elif ref_val is not None and agent_val is not None:
            wrong += 1
            print(f"{norm_label[:50]:<50} {section[:25]:<25} {col_label:<5} {ref_val:<12.0f} {agent_val:<12.0f} WRONG")
        elif agent_val is not None and ref_val is None:
            # Agent wrote a value not in reference — not an error, just extra
            pass
        else:
            correct += 1

    total_correct += correct
    total_wrong += wrong
    total_missing += missing
    total_ref = sum(1 for v in ref_values.values() if v and v != 0)
    matched = correct
    print(f"\n  {sheet_name}: {correct} matched, {wrong} wrong, {missing} missing (of {total_ref} reference values)")

print(f"\n{'='*60}")
total = total_correct + total_wrong + total_missing
accuracy = (total_correct / total * 100) if total > 0 else 0
print(f"RESULT: {total_correct}/{total} correct ({accuracy:.1f}%)")
print(f"  Wrong values: {total_wrong}")
print(f"  Missing values: {total_missing}")
print(f"\nNote: Compares by label+section (not row number).")
print(f"Formula cells excluded (they calculate when opened in Excel).")
